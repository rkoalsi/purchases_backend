import io
import logging
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from ..database import get_database

router = APIRouter()
logger = logging.getLogger(__name__)

SIZE_MAP = {
    "no size": "00",
    "xxxs": "01",
    "xxs": "02",
    "x-small": "03",
    "xsmall": "03",
    "xs": "03",
    "small": "04",
    "s": "04",
    "medium": "05",
    "m": "05",
    "large": "06",
    "l": "06",
    "x-large": "07",
    "xlarge": "07",
    "xl": "07",
    "xxl": "08",
    "xxxl": "09",
}

# Shoe/boot sizes 1-8 map directly to zero-padded strings
SHOE_SIZE_MAP = {str(i): f"{i:02d}" for i in range(1, 9)}

TEMPLATE_COLUMNS = [
    "Brand Name",
    "Series Name",
    "Is Wand Toy (Yes/No)",
    "Animal Type (Dog/Cat)",
    "Item Type (Regular/Shoes/Boots)",
    "Size",
    "Item Name",
    "Color",
]

OUTPUT_COLUMNS = TEMPLATE_COLUMNS + ["Generated BB Code", "Error"]

COL_WIDTHS_TEMPLATE = [20, 20, 22, 22, 28, 15, 20, 15]
COL_WIDTHS_OUTPUT   = [20, 20, 22, 22, 28, 15, 20, 15, 18, 35]


def _get_initials(name: str) -> str:
    """2-letter initials: first letter of each word if multi-word, else first 2 chars."""
    parts = name.strip().split()
    if len(parts) >= 2:
        return (parts[0][0] + parts[1][0]).upper()
    if len(parts) == 1 and len(parts[0]) >= 2:
        return parts[0][:2].upper()
    if len(parts) == 1:
        return (parts[0][0] * 2).upper()
    return "XX"


def _get_first_two(name: str) -> str:
    """First 2 letters of the name (ignoring spaces), padded with X if short."""
    letters = name.strip().replace(" ", "")
    if len(letters) >= 2:
        return letters[:2].upper()
    if len(letters) == 1:
        return (letters[0] * 2).upper()
    return "XX"


def _generate_code(
    brand: str,
    series: str,
    is_wand_toy: str,
    animal_type: str,
    item_type: str,
    size: str,
    item_name: str,
    color: str,
) -> str:
    brand_code = _get_initials(brand)

    is_wand = is_wand_toy.strip().lower() in ("yes", "y", "true", "1")
    series_code = "WA" if is_wand else _get_first_two(series)

    is_cat = animal_type.strip().lower() == "cat"
    is_footwear = item_type.strip().lower() in ("shoes", "boots")

    if is_cat:
        size_code = "00"
    elif is_footwear:
        size_key = size.strip().lower().lstrip("0") or "0"
        # accept both "1" and "01" etc.
        size_code = SHOE_SIZE_MAP.get(size_key)
        if size_code is None:
            raise ValueError(
                f"Invalid shoe/boot size '{size}'. Must be 1–8."
            )
    else:
        size_code = SIZE_MAP.get(size.strip().lower(), "00")

    item_code = _get_first_two(item_name)
    color_code = _get_initials(color)
    return brand_code + series_code + size_code + item_code + color_code


def _build_template_workbook(brands: list[str] | None = None) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "BB Code Template"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, (col_name, width) in enumerate(
        zip(TEMPLATE_COLUMNS, COL_WIDTHS_TEMPLATE), start=1
    ):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20

    # Brand Name  (col A) — populated from DB if brands provided
    if brands:
        brands_csv = ",".join(b.replace(",", "") for b in brands)
        dv_brand = DataValidation(type="list", formula1=f'"{brands_csv}"', allow_blank=False)
        dv_brand.sqref = "A2:A10000"
        ws.add_data_validation(dv_brand)

    # Is Wand Toy  (col C)
    dv_wand = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    dv_wand.sqref = "C2:C10000"
    ws.add_data_validation(dv_wand)

    # Animal Type  (col D)
    dv_animal = DataValidation(type="list", formula1='"Dog,Cat"', allow_blank=False)
    dv_animal.sqref = "D2:D10000"
    ws.add_data_validation(dv_animal)

    # Item Type  (col E)
    dv_item_type = DataValidation(
        type="list", formula1='"Regular,Shoes,Boots"', allow_blank=False
    )
    dv_item_type.sqref = "E2:E10000"
    ws.add_data_validation(dv_item_type)

    # Size  (col F) — regular sizes + shoe sizes 1-8
    sizes_csv = "No Size,XXS,X-Small,Small,Medium,Large,X-Large,XXL,XXXL,1,2,3,4,5,6,7,8"
    dv_size = DataValidation(
        type="list", formula1=f'"{sizes_csv}"', allow_blank=False
    )
    dv_size.sqref = "F2:F10000"
    ws.add_data_validation(dv_size)

    return wb


def _process_rows(ws) -> list[dict]:
    results = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue

        brand      = str(row[0] or "").strip()
        series     = str(row[1] or "").strip()
        is_wand    = str(row[2] or "No").strip()
        animal     = str(row[3] or "Dog").strip()
        item_type  = str(row[4] or "Regular").strip()
        size       = str(row[5] or "No Size").strip()
        item_name  = str(row[6] or "").strip()
        color      = str(row[7] or "").strip()

        base = {
            "row": row_idx,
            "brand": brand,
            "series": series,
            "is_wand_toy": is_wand,
            "animal_type": animal,
            "item_type": item_type,
            "size": size,
            "item_name": item_name,
            "color": color,
        }

        if not brand or not item_name or not color:
            results.append({**base, "bb_code": None, "error": "Brand Name, Item Name, and Color are required"})
            continue

        try:
            bb_code = _generate_code(brand, series, is_wand, animal, item_type, size, item_name, color)
            results.append({**base, "bb_code": bb_code, "error": None})
        except Exception as e:
            results.append({**base, "bb_code": None, "error": str(e)})

    return results


def _build_output_workbook(results: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "BB Codes"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    success_fill = PatternFill("solid", fgColor="E8F5E9")
    success_font = Font(bold=True, color="1B5E20")
    error_fill = PatternFill("solid", fgColor="FFEBEE")
    error_font = Font(color="B71C1C")

    for col_idx, (col_name, width) in enumerate(
        zip(OUTPUT_COLUMNS, COL_WIDTHS_OUTPUT), start=1
    ):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, r in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=r["brand"])
        ws.cell(row=row_idx, column=2, value=r["series"])
        ws.cell(row=row_idx, column=3, value=r["is_wand_toy"])
        ws.cell(row=row_idx, column=4, value=r["animal_type"])
        ws.cell(row=row_idx, column=5, value=r["item_type"])
        ws.cell(row=row_idx, column=6, value=r["size"])
        ws.cell(row=row_idx, column=7, value=r["item_name"])
        ws.cell(row=row_idx, column=8, value=r["color"])

        code_cell = ws.cell(row=row_idx, column=9, value=r["bb_code"])
        if r["bb_code"]:
            code_cell.fill = success_fill
            code_cell.font = success_font

        err_cell = ws.cell(row=row_idx, column=10, value=r["error"])
        if r["error"]:
            err_cell.fill = error_fill
            err_cell.font = error_font

    return wb


@router.get("/template")
def download_template(db=Depends(get_database)):
    """Download the BB Code Generator XLSX template."""
    try:
        brands = sorted(db.get_collection("products").find({"status": "active"}).distinct("brand"))
        brands = [b for b in brands if b]
    except Exception:
        brands = []
    wb = _build_template_workbook(brands)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bb_code_template.xlsx"},
    )


@router.post("/generate")
async def generate_bb_codes(file: UploadFile = File(...)):
    """Upload filled template and receive JSON results with generated BB codes."""
    contents = await file.read()
    try:
        wb = load_workbook(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")

    if "BB Code Template" not in wb.sheetnames:
        raise HTTPException(
            status_code=400,
            detail='Sheet "BB Code Template" not found. Please use the downloaded template.',
        )

    results = _process_rows(wb["BB Code Template"])
    if not results:
        raise HTTPException(status_code=400, detail="No data rows found in the template.")

    return JSONResponse(content={"results": results})


@router.post("/generate/download")
async def download_bb_codes(file: UploadFile = File(...)):
    """Upload filled template and download an XLSX with generated BB codes."""
    contents = await file.read()
    try:
        wb = load_workbook(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")

    if "BB Code Template" not in wb.sheetnames:
        raise HTTPException(
            status_code=400,
            detail='Sheet "BB Code Template" not found. Please use the downloaded template.',
        )

    results = _process_rows(wb["BB Code Template"])
    if not results:
        raise HTTPException(status_code=400, detail="No data rows found in the template.")

    out_wb = _build_output_workbook(results)
    buffer = io.BytesIO()
    out_wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bb_codes_output.xlsx"},
    )
