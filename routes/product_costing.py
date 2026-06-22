"""
routes/product_costing.py

Generate product costing Excel workbooks.

POST /product-costing/generate
    Accept brand tab configs + exchange rates → download .xlsx
    One tab per brand; columns mirror the vendor costing sheets (A–BG)
    plus optional live Zoho stock + 3-month sales columns appended at end.
"""

from __future__ import annotations

import asyncio
import io
import logging
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from fastapi import APIRouter, Depends, Form, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from ..database import get_database

router = APIRouter()
logger = logging.getLogger(__name__)

# ── Column positions (1-indexed, openpyxl convention) ─────────────────────────
# Matches the vendor costing sheet template exactly (A–BG = cols 1–59).
COL_SR          =  1  # A
COL_STATUS      =  2  # B
COL_CODE        =  3  # C  — Manufacturer Code (cf_item_code)
COL_BB          =  4  # D  — BB code (cf_sku_code)
COL_NAME        =  5  # E  — PRODUCT NAME
COL_PACKING     =  6  # F  — PACKING (e.g. "24 packs")
COL_CBM         =  7  # G  — CBM as per PL
COL_TOTAL_CBM   =  8  # H  — formula =G*K
COL_DUTY_PCT    =  9  # I  — Custom Duty %
COL_HSN         = 10  # J  — HSN
COL_CARTON      = 11  # K  — Cartons
COL_QTY         = 12  # L  — formula =M*K  (casepack × carton)
COL_CASEPACK    = 13  # M  — Casepack
COL_PER_PC_1    = 14  # N  — per pc USD (raw from supplier)
COL_PER_PC_2    = 15  # O  — formula =SUBSTITUTE(N,"US$","")
COL_FX_TOTAL    = 16  # P  — USD TOTAL = L*O
COL_CP_ACTUAL   = 17  # Q  — CP (Actual) = IFERROR(R/L,0)
COL_COST_ACT    = 18  # R  — Total Cost (Actual) = (O*bank_rate)*L
COL_CP_ASS      = 19  # S  — CP (Ass. Val.) = IFERROR(T/L,0)
COL_ASS_VAL     = 20  # T  — Ass. Val. = R+U+V
COL_ITEM_FREIGHT   = 21  # U  — Item Freight    [blue / optional]
COL_ITEM_INSURANCE = 22  # V  — Item Insurance  [blue / optional]
COL_CUS_DUTY    = 23  # W  — Cus duty = T*I
COL_SURCHARGE   = 24  # X  — Surcharge = W*10%
COL_DUTY_PC     = 25  # Y  — Cus Duty / PC = IFERROR(W/L,0)
COL_SURCH_PC    = 26  # Z  — Surcharge / PC = IFERROR(X/L,0)
# AA=27: intentionally blank
COL_SHIPPING    = 28  # AB — Shipping fee = G*150*customs_rate*K
COL_PACKAGING   = 29  # AC — Packaging & designing  [blue / optional]
COL_ZIPTIES     = 30  # AD — zip ties                [blue / optional]
COL_LABELS      = 31  # AE — labels                  [blue / optional]
COL_PLASTIC     = 32  # AF — plastic pack             [blue / optional]
COL_COURIER     = 33  # AG — courier box              [blue / optional]
# AH=34: intentionally blank
COL_GST_PCT     = 35  # AI — GST %
COL_LANDED      = 36  # AJ — Total cost to warehouse = IFERROR(((R+W+X+AB)/L),0)
COL_GST_PAID    = 37  # AK — Gst paid till now = (S+Y+Z)*AI
COL_COST_RETAIL = 38  # AL — Total cost to retailer = (AJ+AH+AG+AE+AF+AD+AC)
COL_GST_TOTAL   = 39  # AM — Total gst paid = AK*L
COL_INVESTMENT  = 40  # AN — Total investment = (AL*L)+AM
COL_OLD_MRP     = 41  # AO — OLD MRP (manual / from product rate)
COL_NEW_MRP     = 42  # AP — NEW MRP = AO (copies OLD MRP; user overrides manually)
COL_DISCOUNT    = 43  # AQ — Discount % = MAX(0, IFERROR((AO-AP)/AO,0))
COL_MARGIN      = 44  # AR — Margin (default 0.55)
COL_PRICE_RTL   = 45  # AS — price to retailer = AP-(AP*AR)
COL_GST_RATE_R  = 46  # AT — Gst % = 1+(AI/1)
COL_GST_AMT_R   = 47  # AU — Gst = AS-(AS/AT)
COL_GST_TO_PAY  = 48  # AV — Gst to be paid = AU-AK
COL_CASH_AFTER  = 49  # AW — Cash collected After gst = AS-AV
COL_TOTAL_COST  = 50  # AX — Total cost = IFERROR(AN/L,0)
COL_NET_TOTAL   = 51  # AY — Net Total Cost = AX*L+AZ
COL_LOGISTICS   = 52  # AZ — logistics total = BE*L     [computed from BE]
COL_COLLECTED   = 53  # BA — To be collected after gst = AW*L
COL_PROFIT      = 54  # BB — Total profit = BA-AY
COL_OP_PROFIT   = 55  # BC — Opreating profit % = IFERROR(BB/AY*100,0)
COL_GP_MARGIN   = 56  # BD — Gross profit Margin = IFERROR((BB/BA)*100,0)
COL_LOGISTICS_PC= 57  # BE — Logistics cost per Piece   [blue / optional, user input]
COL_ROI         = 58  # BF — ROI % = BB/(AL*L)*100
COL_PROFIT_PC   = 59  # BG — Profit Per Pcs = BB/L

# Columns appended after BG (always present, then optional live-data)
COL_IMAGE       = 60  # BH — Image or image URL (from upload template)
COL_SALES       = 61  # BI — Total sale of Last 3 months      [live data]
COL_DAYS        = 62  # BJ — Total No. of days in stock       [live data]
COL_ZOHO_WH     = 63  # BK — Zoho Stock (Pupscribe WH)        [live data]
COL_TRANSIT     = 64  # BL — Total Stock in Transit (1+2+3)   [live data]
COL_AVG_SALES   = 65  # BM — Average Sales per Day (DRR)      [live data]
COL_DAYS_COVER  = 66  # BN — Days Until Stock Lasts = IFERROR((BK+BL)/BM,0)

# Exchange rate metadata — placed in META_ROW (row 1), cols A–F.
# Using cols 1–6 avoids overlap with the data columns N/O/P that hold prices.
META_ROW        = 1
META_BANK_LABEL = 1   # A1 — "Conversion rate By Bank"
META_BANK_VAL   = 2   # B1 — bank rate value  ← $B$1 in formulas
META_CUST_LABEL = 3   # C1 — "Conversion rate Customs"
META_CUST_VAL   = 4   # D1 — customs rate value  ← $D$1 in formulas
META_FRET_LABEL = 5   # E1 — "Conversion rate (Freight)"
META_FRET_VAL   = 6   # F1 — freight rate value  ← $F$1 in formulas

HEADER_ROW = 6
DATA_START = 7

# Colours
_BLUE_FILL   = PatternFill("solid", fgColor="00B0F0")  # optional blue columns
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
_META_FONT   = Font(bold=True, size=9)
_DATA_FONT   = Font(size=9)

# Blue (optional) data columns that get blue header fill
_BLUE_COLS = {COL_ITEM_FREIGHT, COL_ITEM_INSURANCE, COL_PACKAGING, COL_ZIPTIES, COL_LABELS, COL_PLASTIC, COL_COURIER, COL_LOGISTICS_PC}


# ── Request models ─────────────────────────────────────────────────────────────

class ExchangeRates(BaseModel):
    bank:    float = 96.0
    customs: float = 92.0
    freight: float = 92.0


class BrandTabConfig(BaseModel):
    label:           str
    brand_names:     list[str]
    currency_label:  str  = "USD"
    currency_filter: str | None = None
    exchange_rates:  ExchangeRates = ExchangeRates()


class ProductCostingRequest(BaseModel):
    tabs:              list[BrandTabConfig]
    include_live_data: bool = True


# ── DB helpers ─────────────────────────────────────────────────────────────────

def _fetch_products(db, tab: BrandTabConfig) -> list[dict]:
    query: dict = {"brand": {"$in": tab.brand_names}}
    if tab.currency_filter:
        query["currency"] = tab.currency_filter
    fields = {
        "_id": 0,
        "cf_sku_code": 1, "cf_item_code": 1,
        "item_name": 1, "brand": 1,
        "purchase_rate": 1, "rate": 1,
        "hsn_or_sac": 1, "cbm": 1, "case_pack": 1,
        "status": 1, "purchase_status": 1,
        "costing_carton": 1, "custom_duty": 1, "purchase_price": 1,
        "item_tax_preferences": 1,
    }
    return list(db.products.find(query, fields).sort([("item_name", 1)]))


# ── Helpers ────────────────────────────────────────────────────────────────────

def _col(c: int) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(c)


def _set(ws, row: int, col: int, value, bold: bool = False, pct: bool = False, fmt: str | None = None):
    cell = ws.cell(row, col)
    cell.value = value
    cell.font = Font(bold=bold, size=9) if bold else _DATA_FONT
    if pct:
        cell.number_format = "0%"
    elif fmt:
        cell.number_format = fmt


def _gst_decimal(product: dict) -> float | None:
    if product.get("gst_pct") is not None:
        return float(product["gst_pct"]) / 100
    for pref in product.get("item_tax_preferences") or []:
        if pref.get("tax_specification") == "intra":
            pct = pref.get("tax_percentage")
            if pct is not None:
                return float(pct) / 100
    return None


# ── Excel builder ──────────────────────────────────────────────────────────────

def _build_headers(cur: str, include_live: bool, stock_date: str, sales_label: str) -> list[str]:
    headers = [
        "Sr",                    #  1 A
        "Status",                #  2 B
        "Code",                  #  3 C
        "BB code",               #  4 D
        "PRODUCT NAME",          #  5 E
        "PACKING",               #  6 F
        "CBM as per PL",         #  7 G
        "Total CBM",             #  8 H
        "Custom Duty",           #  9 I
        "HSN",                   # 10 J
        "Carton",                # 11 K
        "Quantity",              # 12 L
        "Casepack",              # 13 M
        f"per pc {cur}",         # 14 N
        f"per pc {cur}",         # 15 O
        f"{cur} TOTAL",          # 16 P
        "CP (Actual)",           # 17 Q
        "Total Cost (Actual)",   # 18 R
        "CP (Ass. Val.)",        # 19 S
        "Ass. Val.",             # 20 T
        "Item Freight",           # 21 U  [blue / optional]
        "Item Insurance",         # 22 V  [blue / optional]
        "Cus duty",              # 23 W
        "Surcharge",             # 24 X
        "Cus Duty / PC",         # 25 Y
        "Surcharge / PC",        # 26 Z
        "",                      # 27 AA
        "Shipping fee, stamp, trans to ware, Agency, Handling, exami & Gst",  # 28 AB
        "Packaging & designing", # 29 AC
        "zip ties",              # 30 AD
        "labels",                # 31 AE
        "plastic pack",          # 32 AF
        "courier box",           # 33 AG
        "",                      # 34 AH
        "GST %",                 # 35 AI
        "Total cost to warehouse",          # 36 AJ
        "Gst paid till now",                # 37 AK
        "Total cost to retailer",           # 38 AL
        "Total gst paid",                   # 39 AM
        "Total investment",                 # 40 AN
        "OLD MRP",                          # 41 AO
        "NEW MRP",                          # 42 AP
        "Discount %",                       # 43 AQ
        "Margin",                           # 44 AR
        "price to retailer @ Target price", # 45 AS
        "Gst %",                            # 46 AT
        "Gst",                              # 47 AU
        "Gst to be paid",                   # 48 AV
        "Cash collected After gst",         # 49 AW
        "Total cost",                       # 50 AX
        "Net Total Cost",                   # 51 AY
        "logistics",                        # 52 AZ
        "To be collected after gst",        # 53 BA
        "Total profit",                     # 54 BB
        "Opreating profit %",               # 55 BC
        "Gross profit Margin",              # 56 BD
        "Logistics cost per Piece",         # 57 BE
        "ROI %",                            # 58 BF
        "Profit Per Pcs",                   # 59 BG
        "Image / Link",                     # 60 BH  — always present
    ]
    if include_live:
        headers += [
            f"Total sale of Last 3 months ({sales_label})",  # 61 BI
            "Total No. of days in stock last 3 months",       # 62 BJ
            f"Zoho Stock - Pupscribe WH ({stock_date})",      # 63 BK
            "Total Stock in Transit (1+2+3)",                  # 64 BL
            "Average Sales per Day",                           # 65 BM
            "Days Until Stock Lasts",                          # 66 BN
        ]
    return headers


def _write_row(
    ws,
    r: int,
    sr: int,
    product: dict,
    include_live: bool,
    sku_sales_data: dict | None,
):
    sku       = product.get("cf_sku_code", "")
    cbm       = product.get("cbm")
    case_pack = product.get("case_pack")
    pr        = product.get("purchase_price")
    carton    = product.get("costing_carton", 1)
    duty_pct  = product.get("custom_duty")  # stored as %, e.g. 10.0 → 10%

    # ── Identity columns ──────────────────────────────────────────────────────
    _set(ws, r, COL_SR,      sr)
    _set(ws, r, COL_STATUS,  product.get("purchase_status", ""))
    _set(ws, r, COL_CODE,    product.get("cf_item_code", ""))
    _set(ws, r, COL_BB,      sku)
    _set(ws, r, COL_NAME,    product.get("item_name", ""))
    _set(ws, r, COL_PACKING, product.get("packing") or product.get("notes") or "")

    # ── Physical / quantity ───────────────────────────────────────────────────
    if cbm:
        _set(ws, r, COL_CBM, round(cbm, 5))
        ws.cell(r, COL_TOTAL_CBM).value = f"=G{r}*K{r}"

    if duty_pct is not None:
        _set(ws, r, COL_DUTY_PCT, duty_pct / 100, pct=True)

    _set(ws, r, COL_HSN,    product.get("hsn_or_sac", ""))
    _set(ws, r, COL_CARTON, carton)

    if case_pack:
        _set(ws, r, COL_CASEPACK, int(case_pack))

    # Quantity (col L): for order-wise costing, use the actual ordered qty from
    # the PO line item; otherwise fall back to the casepack × carton formula.
    po_qty = product.get("_po_qty")
    if po_qty is not None:
        _set(ws, r, COL_QTY, po_qty)
    elif case_pack:
        ws.cell(r, COL_QTY).value = f"=M{r}*K{r}"  # casepack × carton

    # ── Price columns ─────────────────────────────────────────────────────────
    if pr and pr > 0:
        _set(ws, r, COL_PER_PC_1, pr)
    ws.cell(r, COL_PER_PC_2).value = f'=SUBSTITUTE(N{r},"US$","")'
    ws.cell(r, COL_FX_TOTAL).value = f"=L{r}*O{r}"

    # ── Cost formulas ($B$1 = bank rate in META row 1) ────────────────────────
    ws.cell(r, COL_COST_ACT).value  = f"=(O{r}*$B$1)*L{r}"
    ws.cell(r, COL_CP_ACTUAL).value = f"=IFERROR(R{r}/L{r},0)"
    ws.cell(r, COL_ASS_VAL).value   = f"=R{r}+U{r}+V{r}"
    ws.cell(r, COL_CP_ASS).value    = f"=IFERROR(T{r}/L{r},0)"
    _set(ws, r, COL_ITEM_FREIGHT,   0)
    _set(ws, r, COL_ITEM_INSURANCE, 0)

    # ── Duty formulas ─────────────────────────────────────────────────────────
    ws.cell(r, COL_CUS_DUTY).value  = f"=T{r}*I{r}"
    ws.cell(r, COL_SURCHARGE).value = f"=W{r}*10%"
    ws.cell(r, COL_DUTY_PC).value   = f"=IFERROR(W{r}/L{r},0)"
    ws.cell(r, COL_SURCH_PC).value  = f"=IFERROR(X{r}/L{r},0)"

    # ── Shipping ($D$1 = customs/freight rate) ────────────────────────────────
    if cbm:
        ws.cell(r, COL_SHIPPING).value = f"=G{r}*150*$D$1*K{r}"

    # ── Blue optional columns (default 0, user fills) ─────────────────────────
    for col in (COL_PACKAGING, COL_ZIPTIES, COL_LABELS, COL_PLASTIC, COL_COURIER):
        _set(ws, r, col, 0)

    # ── GST + landed cost ─────────────────────────────────────────────────────
    gst = _gst_decimal(product)
    _set(ws, r, COL_GST_PCT, gst if gst is not None else 0, pct=True)

    # AJ: Total cost to warehouse (landed cost per pc)
    ws.cell(r, COL_LANDED).value     = f"=IFERROR(((R{r}+W{r}+X{r}+AB{r})/L{r}),0)"
    # AK: GST paid till now = (CP_Ass + Duty/PC + Surch/PC) × GST%
    ws.cell(r, COL_GST_PAID).value   = f"=(S{r}+Y{r}+Z{r})*AI{r}"
    # AL: Total cost to retailer = landed + optional packaging cols
    ws.cell(r, COL_COST_RETAIL).value= f"=(AJ{r}+AG{r}+AE{r}+AF{r}+AD{r}+AC{r})"
    # AM: Total GST paid = AK × qty
    ws.cell(r, COL_GST_TOTAL).value  = f"=AK{r}*L{r}"
    # AN: Total investment = (cost_retail × qty) + total_gst
    ws.cell(r, COL_INVESTMENT).value = f"=(AL{r}*L{r})+AM{r}"

    # ── MRP / retail pricing ──────────────────────────────────────────────────
    mrp = product.get("rate")
    if mrp:
        _set(ws, r, COL_OLD_MRP, mrp)

    ws.cell(r, COL_NEW_MRP).value   = f"=AO{r}"
    c_disc = ws.cell(r, COL_DISCOUNT)
    c_disc.value = f"=MAX(0,IFERROR((AO{r}-AP{r})/AO{r},0))"
    c_disc.number_format = "0%"

    _set(ws, r, COL_MARGIN, 0.55, pct=True)

    ws.cell(r, COL_PRICE_RTL).value  = f"=AP{r}-(AP{r}*AR{r})"
    ws.cell(r, COL_GST_RATE_R).value = f"=1+AI{r}"
    ws.cell(r, COL_GST_AMT_R).value  = f"=AS{r}-(AS{r}/AT{r})"
    ws.cell(r, COL_GST_TO_PAY).value = f"=AU{r}-AK{r}"
    ws.cell(r, COL_CASH_AFTER).value = f"=AS{r}-AV{r}"

    # ── Profit calculation ────────────────────────────────────────────────────
    ws.cell(r, COL_TOTAL_COST).value = f"=IFERROR(AN{r}/L{r},0)"
    # BE: logistics per piece (user fills, blue) — default 0
    _set(ws, r, COL_LOGISTICS_PC, 0)
    # AZ: logistics total = BE × qty
    ws.cell(r, COL_LOGISTICS).value  = f"=BE{r}*L{r}"
    # AY: net total cost = total_cost × qty + logistics
    ws.cell(r, COL_NET_TOTAL).value  = f"=AX{r}*L{r}+AZ{r}"
    # BA: cash collected after gst × qty
    ws.cell(r, COL_COLLECTED).value  = f"=AW{r}*L{r}"
    # BB: total profit
    ws.cell(r, COL_PROFIT).value     = f"=BA{r}-AY{r}"
    c_op = ws.cell(r, COL_OP_PROFIT)
    c_op.value = f"=IFERROR(BB{r}/AY{r}*100,0)"
    c_gp = ws.cell(r, COL_GP_MARGIN)
    c_gp.value = f"=IFERROR((BB{r}/BA{r})*100,0)"
    ws.cell(r, COL_ROI).value        = f"=BB{r}/(AL{r}*L{r})*100"
    ws.cell(r, COL_PROFIT_PC).value  = f"=BB{r}/L{r}"

    # ── Image / Link (always) ─────────────────────────────────────────────────
    img_bytes = product.get("image_bytes")
    img_url   = product.get("image_link", "")
    if img_bytes:
        try:
            from openpyxl.drawing.image import Image as XLImage
            xl_img = XLImage(io.BytesIO(img_bytes))
            xl_img.width  = 80
            xl_img.height = 60
            xl_img.anchor = f"{_col(COL_IMAGE)}{r}"
            ws.add_image(xl_img)
            ws.row_dimensions[r].height = 50
        except Exception:
            _set(ws, r, COL_IMAGE, "(image)")
    elif img_url:
        cell = ws.cell(r, COL_IMAGE)
        cell.value = img_url
        if str(img_url).startswith("http"):
            cell.hyperlink = img_url
            cell.font = Font(color="0563C1", underline="single", size=9)

    # ── Optional live data ────────────────────────────────────────────────────
    if include_live:
        if sku_sales_data and sku:
            sd = sku_sales_data.get(sku, {})
            total_sales = sd.get("total_units_sold", "")
            days_in_stk = sd.get("total_days_in_stock", "")
            zoho_wh     = sd.get("pupscribe_wh_stock", "")
            transit     = sd.get("total_transit", "")
            drr         = sd.get("avg_daily_run_rate", "")
            if total_sales != "":
                _set(ws, r, COL_SALES, total_sales)
            if days_in_stk != "":
                _set(ws, r, COL_DAYS, days_in_stk)
            if zoho_wh != "":
                _set(ws, r, COL_ZOHO_WH, zoho_wh)
            if transit != "":
                _set(ws, r, COL_TRANSIT, transit)
            if drr != "":
                _set(ws, r, COL_AVG_SALES, drr)

        ws.cell(r, COL_DAYS_COVER).value = f"=IFERROR((BK{r}+BL{r})/BM{r},0)"


def _build_sheet(
    wb,
    tab: BrandTabConfig,
    products: list[dict],
    sku_sales_data: dict | None,
    stock_date: str,
    sales_label: str,
):
    ws = wb.create_sheet(title=tab.label[:31])
    rates = tab.exchange_rates
    cur   = tab.currency_label

    # ── Row 1: exchange rate metadata (cols A–F) ──────────────────────────────
    _set(ws, META_ROW, META_BANK_LABEL, "Conversion rate By Bank",  bold=True)
    _set(ws, META_ROW, META_BANK_VAL,   rates.bank)
    _set(ws, META_ROW, META_CUST_LABEL, "Conversion rate Customs",  bold=True)
    _set(ws, META_ROW, META_CUST_VAL,   rates.customs)
    _set(ws, META_ROW, META_FRET_LABEL, "Conversion rate (Freight)", bold=True)
    _set(ws, META_ROW, META_FRET_VAL,   rates.freight)

    # ── Row 2: date + currency ────────────────────────────────────────────────
    _set(ws, 2, 1, "Date", bold=True)
    _set(ws, 2, 3, datetime.now().strftime("%d-%m-%Y"))
    _set(ws, 2, 5, cur)

    # ── Row 3: brand title ────────────────────────────────────────────────────
    _set(ws, 3, 1, f"{tab.label} Latest Price", bold=True)

    # ── Row 6: column headers ─────────────────────────────────────────────────
    include_live = sku_sales_data is not None
    headers = _build_headers(cur, include_live, stock_date, sales_label)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(HEADER_ROW, c)
        cell.value     = h
        cell.font      = _HEADER_FONT
        cell.fill      = _BLUE_FILL if c in _BLUE_COLS else _HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    ws.row_dimensions[HEADER_ROW].height = 40

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, product in enumerate(products, 1):
        r = DATA_START + i - 1
        _write_row(ws, r, i, product, include_live, sku_sales_data)

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions[_col(COL_NAME)].width    = 45
    ws.column_dimensions[_col(COL_PACKING)].width = 16
    ws.column_dimensions[_col(COL_BB)].width      = 16
    ws.column_dimensions[_col(COL_CODE)].width    = 12
    ws.column_dimensions[_col(COL_HSN)].width     = 12
    ws.column_dimensions[_col(COL_STATUS)].width  = 22
    ws.column_dimensions[_col(COL_SHIPPING)].width = 20
    ws.column_dimensions[_col(COL_IMAGE)].width    = 20

    ws.freeze_panes = f"A{DATA_START}"


def _build_workbook(
    tabs: list[BrandTabConfig],
    products_by_tab: list[list[dict]],
    sku_sales_data: dict | None,
    stock_date: str,
    sales_label: str,
) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for tab, products in zip(tabs, products_by_tab):
        _build_sheet(wb, tab, products, sku_sales_data, stock_date, sales_label)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Endpoint ───────────────────────────────────────────────────────────────────

@router.post("/generate")
async def generate_product_costing(
    body: ProductCostingRequest,
    db=Depends(get_database),
):
    if not body.tabs:
        raise HTTPException(status_code=400, detail="Select at least one brand tab.")

    products_by_tab: list[list[dict]] = await asyncio.to_thread(
        lambda: [_fetch_products(db, tab) for tab in body.tabs]
    )

    sku_sales_data: dict | None = None
    stock_date = datetime.now().strftime("%d %b %Y")
    sales_label = ""

    if body.include_live_data:
        from .sheets_updater import _last_3_months_range
        from ..services.master_report import _generate_master_report_data

        start_date, end_date = _last_3_months_range()
        sales_label = f"{start_date} to {end_date}"

        try:
            report = await _generate_master_report_data(
                start_date=start_date, end_date=end_date, db=db,
            )
            stock_date = report.get("dates", {}).get("zoho") or stock_date
            sku_sales_data = {}
            for item in report.get("combined_data", []):
                sku = item.get("sku_code", "")
                if sku:
                    m = item.get("combined_metrics", {})
                    sku_sales_data[sku] = {
                        "total_units_sold":    round(m.get("total_units_sold", 0) or 0, 2),
                        "total_days_in_stock": int(m.get("total_days_in_stock", 0) or 0),
                        "pupscribe_wh_stock":  round(m.get("pupscribe_wh_stock", 0) or 0, 2),
                        "avg_daily_run_rate":  round(m.get("avg_daily_run_rate", 0) or 0, 4),
                        "total_transit":       round(
                            (item.get("stock_in_transit_1") or 0) +
                            (item.get("stock_in_transit_2") or 0) +
                            (item.get("stock_in_transit_3") or 0), 2
                        ),
                    }
        except Exception as exc:
            logger.warning("Live data fetch failed (continuing without): %s", exc)
            sku_sales_data = None

    excel_bytes = await asyncio.to_thread(
        _build_workbook,
        body.tabs, products_by_tab,
        sku_sales_data,
        stock_date, sales_label,
    )

    filename = f"product_costing_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Template download ──────────────────────────────────────────────────────────

_TEMPLATE_COLUMNS = [
    ("Product Name",             True),
    ("Unit Price (USD)",         True),
    ("Units per Carton",         True),
    ("CBM per Carton",           True),
    ("Custom Duty %",            True),
    ("HSN Code",                 True),
    ("GST %",                    True),
    ("Packing",                  False),
    ("Image Link",               False),
    ("Vendor SKU / Article No.", False),
    ("Unit Price (RMB)",         False),
    ("Carton Weight (kg)",       False),
    ("Notes",                    False),
]

_TEMPLATE_WIDTHS = [40, 18, 18, 16, 15, 14, 10, 16, 40, 24, 16, 18, 20]

_REQ_FILL = PatternFill("solid", fgColor="1F3864")
_OPT_FILL = PatternFill("solid", fgColor="4472C4")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)


def _build_template_workbook() -> bytes:
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Price List"

    for col, ((name, required), width) in enumerate(zip(_TEMPLATE_COLUMNS, _TEMPLATE_WIDTHS), 1):
        cell = ws.cell(1, col, name)
        cell.font      = _HDR_FONT
        cell.fill      = _REQ_FILL if required else _OPT_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get("/brands")
async def list_brands(db=Depends(get_database)):
    """Return all brand names from the brands collection, sorted alphabetically."""
    brands = await asyncio.to_thread(
        lambda: sorted(db.get_collection("brands").distinct("name"))
    )
    return {"brands": brands}


@router.get("/template")
async def download_template():
    excel_bytes = await asyncio.to_thread(_build_template_workbook)
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="product_costing_template.xlsx"'},
    )


# ── Template upload → costing sheet ───────────────────────────────────────────

_COL_MAP = {
    "Product Name":             "item_name",
    "Unit Price (USD)":         "purchase_price",
    "Units per Carton":         "case_pack",
    "CBM per Carton":           "cbm",
    "Custom Duty %":            "custom_duty",
    "HSN Code":                 "hsn_or_sac",
    "GST %":                    "gst_pct",
    "Packing":                  "packing",
    "Image Link":               "image_link",
    "Vendor SKU / Article No.": "cf_item_code",
    "Unit Price (RMB)":         "rmb_price",
    "Carton Weight (kg)":       "ctn_weight",
    "Notes":                    "notes",
}


def _parse_template(file_bytes: bytes) -> list[dict]:
    # read_only=False so ws._images is populated (needed for embedded image extraction)
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Build row → image-bytes map from any images embedded in the sheet
    image_by_row: dict[int, bytes] = {}
    for img in getattr(ws, "_images", []):
        try:
            anchor = img.anchor
            if hasattr(anchor, "_from"):
                row_1idx = anchor._from.row + 1   # _from.row is 0-indexed
            elif isinstance(anchor, str):
                import re
                m = re.match(r"[A-Z]+(\d+)", anchor)
                row_1idx = int(m.group(1)) if m else None
            else:
                row_1idx = None
            if row_1idx and callable(getattr(img, "_data", None)):
                data = img._data()
                if data:
                    image_by_row[row_1idx] = data
        except Exception:
            pass

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    header_row_idx = None
    col_indices: dict[str, int] = {}
    for ri, row in enumerate(rows):
        mapping = {}
        for ci, cell in enumerate(row):
            if cell and str(cell).strip() in _COL_MAP:
                mapping[str(cell).strip()] = ci
        if mapping:
            header_row_idx = ri
            col_indices = mapping
            break

    if header_row_idx is None:
        raise HTTPException(
            status_code=400,
            detail="Could not find header row. Make sure the file uses the standard template."
        )

    if "Product Name" not in col_indices or "Unit Price (USD)" not in col_indices:
        raise HTTPException(
            status_code=400,
            detail="Template must have at least 'Product Name' and 'Unit Price (USD)' columns."
        )

    products = []
    for ws_row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
        if not any(row):
            continue

        def _get(col_name: str, _row=row):
            idx = col_indices.get(col_name)
            if idx is None or idx >= len(_row):
                return None
            v = _row[idx]
            return v if v not in ("", None) else None

        name      = _get("Product Name")
        price_usd = _get("Unit Price (USD)")

        if not name or not isinstance(price_usd, (int, float)) or float(price_usd) <= 0:
            continue

        case_pack  = _get("Units per Carton")
        cbm        = _get("CBM per Carton")
        duty       = _get("Custom Duty %")
        hsn        = _get("HSN Code")
        gst        = _get("GST %")
        packing    = _get("Packing")
        image_link = _get("Image Link")
        vendor_sku = _get("Vendor SKU / Article No.")

        product: dict = {
            "item_name":       str(name).strip(),
            "purchase_price":  float(price_usd),
            "cf_sku_code":     "",
            "cf_item_code":    str(vendor_sku) if vendor_sku is not None else "",
            "purchase_status": "",
            "packing":         str(packing).strip() if packing is not None else "",
            "image_link":      str(image_link).strip() if image_link is not None else "",
        }
        # Prefer embedded image over text link; fall back to text if no image found
        img_bytes = image_by_row.get(ws_row_idx)
        if img_bytes:
            product["image_bytes"] = img_bytes
        if case_pack is not None:
            try:
                product["case_pack"] = int(float(case_pack))
            except (ValueError, TypeError):
                pass
        if cbm is not None:
            try:
                product["cbm"] = float(cbm)
            except (ValueError, TypeError):
                pass
        if duty is not None:
            try:
                product["custom_duty"] = float(duty)
            except (ValueError, TypeError):
                pass
        if hsn is not None:
            product["hsn_or_sac"] = str(hsn)
        if gst is not None:
            try:
                product["gst_pct"] = float(gst)
            except (ValueError, TypeError):
                pass

        products.append(product)

    if not products:
        raise HTTPException(status_code=400, detail="No valid product rows found in the uploaded file.")

    return products


@router.post("/upload")
async def upload_price_list(
    file:           UploadFile = File(...),
    tab_label:      str        = Form("Uploaded"),
    currency_label: str        = Form("USD"),
    bank_rate:      float      = Form(96.0),
    customs_rate:   float      = Form(92.0),
    freight_rate:   float      = Form(92.0),
):
    content = await file.read()

    products = await asyncio.to_thread(_parse_template, content)

    tab = BrandTabConfig(
        label=tab_label,
        brand_names=[],
        currency_label=currency_label,
        exchange_rates=ExchangeRates(bank=bank_rate, customs=customs_rate, freight=freight_rate),
    )

    excel_bytes = await asyncio.to_thread(
        _build_workbook,
        [tab], [products],
        None, None,
        datetime.now().strftime("%d %b %Y"), "",
    )

    safe = tab_label.replace(" ", "_").replace("/", "-")
    filename = f"product_costing_{safe}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Order Wise Costing ─────────────────────────────────────────────────────────
# Costing sheets driven by the actual rates inside a vendor's purchase orders.
#
# Brand → vendor: the `brands` collection maps `name → vendor_id`; `vendor_id`
# matches `purchase_orders.vendor_id`. Merged brands (e.g. Petfest = Dogfest +
# Catfest) resolve to the union of their constituent vendor_ids (deduped — the
# constituents often share a single vendor).


def _vendor_ids_for_brands(db, brand_names: list[str]) -> list[str]:
    """Resolve a list of constituent brand names → deduped vendor_ids.

    A brand may map to multiple vendors (e.g. Truelove has a USD vendor and a
    CNY vendor) via the `vendor_ids` array; older docs only have the scalar
    `vendor_id`. Read both so every vendor's POs are reachable.
    """
    seen: list[str] = []
    for b in db.brands.find(
        {"name": {"$in": brand_names}}, {"_id": 0, "vendor_id": 1, "vendor_ids": 1}
    ):
        vids = list(b.get("vendor_ids") or [])
        if b.get("vendor_id"):
            vids.append(b["vendor_id"])
        for vid in vids:
            if vid and vid not in seen:
                seen.append(vid)
    return seen


def _li_has_sku(li: dict) -> bool:
    return bool(li.get("sku")) and (li.get("quantity") or 0) > 0


def _fetch_brand_pos(db, brand_names: list[str]) -> list[dict]:
    vendor_ids = _vendor_ids_for_brands(db, brand_names)
    if not vendor_ids:
        return []
    cursor = db.purchase_orders.find(
        {"vendor_id": {"$in": vendor_ids}},
        {
            "_id": 0, "purchaseorder_number": 1, "date": 1,
            "currency_code": 1, "exchange_rate": 1, "vendor_name": 1,
            "total": 1, "line_items": 1, "status": 1,
        },
    )
    out: list[dict] = []
    for po in cursor:
        items = [li for li in (po.get("line_items") or []) if _li_has_sku(li)]
        if not items:
            continue
        out.append({
            "po_number":     po.get("purchaseorder_number", ""),
            "date":          str(po.get("date") or "")[:10],
            "currency_code": po.get("currency_code", "USD"),
            "exchange_rate": round(po.get("exchange_rate") or 0, 4),
            "vendor_name":   po.get("vendor_name", ""),
            "status":        po.get("status", ""),
            "total":         round(po.get("total") or 0, 2),
            "num_items":     len(items),
        })

    # Attach the brand-order name (e.g. "Order #2") for each PO, matched on
    # purchaseorder_number — same field shown on the brand_orders page.
    po_nums = [p["po_number"] for p in out if p["po_number"]]
    name_by_po = {
        bo["purchaseorder_number"]: bo.get("name", "")
        for bo in db.brand_orders.find(
            {"purchaseorder_number": {"$in": po_nums}},
            {"_id": 0, "name": 1, "purchaseorder_number": 1},
        )
    }
    for p in out:
        p["order_name"] = name_by_po.get(p["po_number"], "")

    out.sort(key=lambda p: p["date"], reverse=True)
    return out


@router.get("/order-wise/pos")
async def list_order_wise_pos(brands: str, db=Depends(get_database)):
    """List purchase orders for the given brand(s).

    `brands` = comma-separated constituent brand names (e.g. "Dogfest,Catfest").
    """
    brand_names = [b.strip() for b in brands.split(",") if b.strip()]
    if not brand_names:
        raise HTTPException(status_code=400, detail="No brand provided.")
    pos = await asyncio.to_thread(_fetch_brand_pos, db, brand_names)
    return {"purchase_orders": pos}


def _assemble_po_line_items(db, po: dict) -> list[dict]:
    """Build costing-row dicts from a PO's line items, joined to products.

    PO line items reference products by Zoho `item_id` and barcode `sku` — NOT by
    `cf_sku_code`. We look up products on both and use the product's canonical
    `cf_sku_code` (needed for the BB-code column and the live-data join).
    """
    items = [li for li in (po.get("line_items") or []) if _li_has_sku(li)]
    item_ids = list({li["item_id"] for li in items if li.get("item_id")})
    barcodes = list({li["sku"] for li in items if li.get("sku")})
    fields = {
        "_id": 0, "cf_sku_code": 1, "cf_item_code": 1, "item_name": 1,
        "purchase_status": 1, "rate": 1, "case_pack": 1, "cbm": 1,
        "costing_carton": 1, "custom_duty": 1, "hsn_or_sac": 1,
        "item_tax_preferences": 1, "packing": 1, "image_link": 1,
        "item_id": 1, "sku": 1,
    }
    by_item_id: dict[str, dict] = {}
    by_sku:     dict[str, dict] = {}
    for p in db.products.find(
        {"$or": [{"item_id": {"$in": item_ids}}, {"sku": {"$in": barcodes}}]},
        fields,
    ):
        if p.get("item_id"):
            by_item_id[p["item_id"]] = p
        if p.get("sku"):
            by_sku[p["sku"]] = p

    rows: list[dict] = []
    for li in items:
        p = by_item_id.get(li.get("item_id")) or by_sku.get(li.get("sku")) or {}
        row = {
            "cf_sku_code":      p.get("cf_sku_code") or li.get("sku", ""),
            "cf_item_code":     p.get("cf_item_code", ""),
            "item_name":        li.get("name") or p.get("item_name", ""),
            "purchase_status":  p.get("purchase_status", ""),
            "purchase_price":   li.get("rate") or 0,          # ← PO line-item rate
            "_po_qty":          li.get("quantity") or 0,       # ← actual ordered qty
            "case_pack":        p.get("case_pack"),
            "cbm":              p.get("cbm"),
            "costing_carton":   p.get("costing_carton", 1),
            "custom_duty":      p.get("custom_duty"),
            "hsn_or_sac":       li.get("hsn_or_sac") or p.get("hsn_or_sac", ""),
            "rate":             p.get("rate"),                 # MRP
            "packing":          p.get("packing"),
            "image_link":       p.get("image_link", ""),
        }
        # GST sourced from products.item_tax_preferences (intra spec) via _gst_decimal
        if p.get("item_tax_preferences"):
            row["item_tax_preferences"] = p["item_tax_preferences"]
        rows.append(row)
    return rows


class OrderWisePO(BaseModel):
    po_number:      str
    currency_label: str = "USD"
    exchange_rates: ExchangeRates = ExchangeRates()


class OrderWiseRequest(BaseModel):
    brand_names:       list[str]
    pos:               list[OrderWisePO]
    include_live_data: bool = False


@router.post("/order-wise/generate")
async def generate_order_wise_costing(body: OrderWiseRequest, db=Depends(get_database)):
    if not body.brand_names:
        raise HTTPException(status_code=400, detail="No brand provided.")
    if not body.pos:
        raise HTTPException(status_code=400, detail="Select at least one purchase order.")

    vendor_ids = await asyncio.to_thread(_vendor_ids_for_brands, db, body.brand_names)
    if not vendor_ids:
        raise HTTPException(status_code=404, detail="Brand has no mapped vendor.")

    po_numbers = [p.po_number for p in body.pos]

    def _load() -> tuple[list[BrandTabConfig], list[list[dict]]]:
        docs = {
            po.get("purchaseorder_number"): po
            for po in db.purchase_orders.find(
                {"vendor_id": {"$in": vendor_ids},
                 "purchaseorder_number": {"$in": po_numbers}}
            )
        }
        tabs: list[BrandTabConfig] = []
        products_by_tab: list[list[dict]] = []
        for req_po in body.pos:
            po = docs.get(req_po.po_number)
            if not po:
                continue
            tabs.append(BrandTabConfig(
                label=req_po.po_number[:31],
                brand_names=body.brand_names,
                currency_label=req_po.currency_label,
                exchange_rates=req_po.exchange_rates,
            ))
            products_by_tab.append(_assemble_po_line_items(db, po))
        return tabs, products_by_tab

    tabs, products_by_tab = await asyncio.to_thread(_load)
    if not tabs:
        raise HTTPException(status_code=404, detail="None of the selected POs were found.")

    sku_sales_data: dict | None = None
    stock_date  = datetime.now().strftime("%d %b %Y")
    sales_label = ""

    if body.include_live_data:
        from .sheets_updater import _last_3_months_range
        from ..services.master_report import _generate_master_report_data

        start_date, end_date = _last_3_months_range()
        sales_label = f"{start_date} to {end_date}"
        try:
            report = await _generate_master_report_data(
                start_date=start_date, end_date=end_date, db=db,
            )
            stock_date = report.get("dates", {}).get("zoho") or stock_date
            sku_sales_data = {}
            for item in report.get("combined_data", []):
                sku = item.get("sku_code", "")
                if sku:
                    m = item.get("combined_metrics", {})
                    sku_sales_data[sku] = {
                        "total_units_sold":    round(m.get("total_units_sold", 0) or 0, 2),
                        "total_days_in_stock": int(m.get("total_days_in_stock", 0) or 0),
                        "pupscribe_wh_stock":  round(m.get("pupscribe_wh_stock", 0) or 0, 2),
                        "avg_daily_run_rate":  round(m.get("avg_daily_run_rate", 0) or 0, 4),
                        "total_transit":       round(
                            (item.get("stock_in_transit_1") or 0) +
                            (item.get("stock_in_transit_2") or 0) +
                            (item.get("stock_in_transit_3") or 0), 2
                        ),
                    }
        except Exception as exc:
            logger.warning("Live data fetch failed (continuing without): %s", exc)
            sku_sales_data = None

    excel_bytes = await asyncio.to_thread(
        _build_workbook,
        tabs, products_by_tab,
        sku_sales_data,
        stock_date, sales_label,
    )

    brand_slug = "_".join(body.brand_names).replace(" ", "-").replace("/", "-")
    filename = f"order_costing_{brand_slug}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
