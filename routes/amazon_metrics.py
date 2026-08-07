# amazon_metrics.py
"""
Amazon per-ASIN metrics workbook ("amz metrics").

Rebuilds, from live data, the sheet that was previously maintained by hand:
one row per ASIN, a static unit-economics block, then one block per calendar
month with units, sessions, ad spend and P&L.

Formulas are written into the cells rather than pre-computed values, so the
workbook stays live — editing ASP, COGS or transport cost recalculates the
whole P&L in Excel, exactly as the manual version did.

Data sources
------------
units / sales / sessions   amazon_sales_traffic   (SC sales & traffic, per ASIN per day)
ad spend by ASIN           amazon_ads_product_daily (spAdvertisedProduct / sdAdvertisedProduct)
ad spend, Brands           amazon_ads_daily, mapped campaign-name -> SKU (approximate)
ASP / margin               vendor_margins
COGS                       Zoho Books inventory valuation (same source as the master report)
"""
import io
import re
import logging
import asyncio
from datetime import datetime, date, timedelta, timezone

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from ..database import get_database
from .amazon_ads import DAILY_COLLECTION, PRODUCT_DAILY_COLLECTION

logger = logging.getLogger(__name__)

router = APIRouter()

SALES_TRAFFIC_COLLECTION = "amazon_sales_traffic"   # Seller Central: units, sales, sessions
VENDOR_SALES_COLLECTION = "amazon_vendor_sales"     # Vendor Central: ordered/shipped units + revenue
SKU_MAPPING_COLLECTION = "amazon_sku_mapping"
VENDOR_MARGINS_COLLECTION = "vendor_margins"

# Flat per-unit transport cost. Kept as a constant for now; when this needs to
# vary per SKU it moves to its own collection without touching the formulas.
DEFAULT_TRANSPORT_COST = 3.0
DEFAULT_GST = 0.18
DEFAULT_MARGIN = 0.33

# An ASIN first seen within this many days is treated as a new launch.
NEW_LAUNCH_DAYS = 180

DEFAULT_MONTHS = 6


# --------------------------------------------------------------------------
# Period helpers
# --------------------------------------------------------------------------
def _month_key(d: date) -> str:
    return d.strftime("%Y-%m")


def _month_windows(end: date, months: int) -> list[tuple[str, str, str, str]]:
    """
    Returns [(month_key, label, first_day, last_day)] oldest first.

    The most recent month is clipped at `end` so a partial month reports only
    the days that exist rather than implying a full month of data.
    """
    out = []
    cursor = date(end.year, end.month, 1)
    starts = []
    for _ in range(months):
        starts.append(cursor)
        cursor = date(cursor.year - 1, 12, 1) if cursor.month == 1 else date(cursor.year, cursor.month - 1, 1)
    for first in reversed(starts):
        nxt = date(first.year + 1, 1, 1) if first.month == 12 else date(first.year, first.month + 1, 1)
        last = min(nxt - timedelta(days=1), end)
        out.append((_month_key(first), first.strftime("%b %y"), first.isoformat(), last.isoformat()))
    return out


# --------------------------------------------------------------------------
# Data loading
# --------------------------------------------------------------------------
def _load_sales_traffic_sync(db, start: str, end: str) -> tuple[dict, dict, dict]:
    """
    Per (asin, month): units, sales, sessions. Also returns parent ASIN and
    the earliest date each ASIN was seen, used for the Old/New launch flag.

    Dates are stored as datetimes in this collection, so the match is a range
    over datetime rather than the string comparison used elsewhere.
    """
    start_dt = datetime.fromisoformat(start)
    end_dt = datetime.fromisoformat(end) + timedelta(days=1)

    per_month: dict[tuple[str, str], dict] = {}
    parents: dict[str, str] = {}
    first_seen: dict[str, str] = {}

    cursor = db[SALES_TRAFFIC_COLLECTION].find(
        {"date": {"$gte": start_dt, "$lt": end_dt}},
        {"date": 1, "parentAsin": 1, "salesByAsin": 1, "trafficByAsin": 1, "childAsin": 1, "asin": 1, "_id": 0},
    )
    for doc in cursor:
        d = doc.get("date")
        if not isinstance(d, datetime):
            continue
        asin = doc.get("childAsin") or doc.get("asin") or doc.get("parentAsin")
        if not asin:
            continue
        mk = _month_key(d.date())
        sales = doc.get("salesByAsin") or {}
        traffic = doc.get("trafficByAsin") or {}
        amount = (sales.get("orderedProductSales") or {}).get("amount") or 0

        rec = per_month.setdefault((asin, mk), {"units": 0.0, "sales": 0.0, "sessions": 0.0})
        rec["units"] += sales.get("unitsOrdered") or 0
        rec["sales"] += amount
        rec["sessions"] += traffic.get("sessions") or 0

        if doc.get("parentAsin"):
            parents[asin] = doc["parentAsin"]
        iso = d.date().isoformat()
        if asin not in first_seen or iso < first_seen[asin]:
            first_seen[asin] = iso

    return per_month, parents, first_seen


def _load_vendor_sales_sync(db, start: str, end: str) -> tuple[dict, dict]:
    """
    Per (asin, month) ordered units and revenue from Vendor Central.

    This is the correct denominator when the advertising profile is a vendor
    entity: Seller Central figures cover a different channel entirely and
    produce wildly overstated TACOS if used here.

    Vendor Central reports no session data, so conversion rate is unavailable.
    """
    start_dt = datetime.fromisoformat(start)
    end_dt = datetime.fromisoformat(end) + timedelta(days=1)

    per_month: dict[tuple[str, str], dict] = {}
    first_seen: dict[str, str] = {}

    pipeline = [
        {"$match": {"date": {"$gte": start_dt, "$lt": end_dt}}},
        {"$group": {
            "_id": {"asin": "$asin", "month": {"$dateToString": {"format": "%Y-%m", "date": "$date"}}},
            "units": {"$sum": "$orderedUnits"},
            "sales": {"$sum": "$orderedRevenue.amount"},
            "shipped_units": {"$sum": "$shippedUnits"},
            "shipped_cogs": {"$sum": "$shippedCogs.amount"},
            "returns": {"$sum": "$customerReturns"},
            "first": {"$min": "$date"},
        }},
    ]
    for r in db[VENDOR_SALES_COLLECTION].aggregate(pipeline):
        asin, mk = r["_id"]["asin"], r["_id"]["month"]
        if not asin:
            continue
        per_month[(asin, mk)] = {
            "units": r.get("units") or 0,
            "sales": r.get("sales") or 0,
            "sessions": 0,  # not reported by Vendor Central
            "shipped_units": r.get("shipped_units") or 0,
            "shipped_cogs": r.get("shipped_cogs") or 0,
            "returns": r.get("returns") or 0,
        }
        f = r.get("first")
        if isinstance(f, datetime):
            iso = f.date().isoformat()
            if asin not in first_seen or iso < first_seen[asin]:
                first_seen[asin] = iso

    return per_month, first_seen


def _load_first_seen_all_time_sync(db) -> dict:
    """
    Earliest sale date per ASIN across ALL history.

    Must not be derived from the report window: inside a 6-month window every
    ASIN looks like it first sold within 180 days, which would mark the entire
    catalogue "New".
    """
    first: dict[str, str] = {}
    for coll, asin_field in ((VENDOR_SALES_COLLECTION, "$asin"), (SALES_TRAFFIC_COLLECTION, "$childAsin")):
        try:
            for r in db[coll].aggregate([
                {"$group": {"_id": asin_field, "first": {"$min": "$date"}}},
            ]):
                asin, f = r.get("_id"), r.get("first")
                if not asin or not isinstance(f, datetime):
                    continue
                iso = f.date().isoformat()
                if asin not in first or iso < first[asin]:
                    first[asin] = iso
        except Exception as e:
            logger.warning("first-seen scan failed for %s: %s", coll, e)
    return first


def _load_parent_map_sync(db) -> dict:
    """
    child ASIN -> parent ASIN, from Seller Central.

    Vendor Central carries no parent/child relationship, so this is used for
    both sources; it is purely descriptive and never feeds a calculation.
    """
    parents: dict[str, str] = {}
    try:
        for r in db[SALES_TRAFFIC_COLLECTION].aggregate([
            {"$match": {"childAsin": {"$ne": None}, "parentAsin": {"$ne": None}}},
            {"$group": {"_id": "$childAsin", "parent": {"$last": "$parentAsin"}}},
        ]):
            if r.get("_id"):
                parents[r["_id"]] = r.get("parent")
    except Exception as e:
        logger.warning("parent map scan failed: %s", e)
    return parents


def _combine_channels(vendor: dict, seller: dict) -> dict:
    """
    Vendor + Seller Central, matching how the manual sheet counted units.

    Units use Vendor *shipped* rather than *ordered*: shipped + Seller
    reproduced the manual sheet's July total to 5 units in 7,637 (0.07%),
    where ordered + Seller was 42 out. Revenue still uses Vendor *ordered*,
    because Amazon reports shippedRevenue as 0 for this account — so units and
    revenue come from different Vendor measures by necessity, not by choice.

    Sessions come only from Seller Central while units span both channels, so
    conversion rate is not meaningful here and is left at zero.
    """
    out: dict[tuple[str, str], dict] = {}
    for key, v in vendor.items():
        out[key] = {
            "units": v.get("shipped_units") or 0,
            "sales": v.get("sales") or 0,
            "sessions": 0,
            "vendor_units": v.get("shipped_units") or 0,
            "seller_units": 0,
        }
    for key, s in seller.items():
        rec = out.setdefault(key, {"units": 0, "sales": 0, "sessions": 0,
                                   "vendor_units": 0, "seller_units": 0})
        rec["units"] += s.get("units") or 0
        rec["sales"] += s.get("sales") or 0
        rec["sessions"] += s.get("sessions") or 0
        rec["seller_units"] = s.get("units") or 0
    return out


def _load_product_ads_sync(db, start: str, end: str) -> dict:
    """Per (asin, month, ad_product): true ASIN-level spend and ad sales."""
    out: dict[tuple[str, str, str], dict] = {}
    pipeline = [
        {"$match": {"date": {"$gte": start, "$lte": end}}},
        {"$group": {
            "_id": {"asin": "$asin", "month": {"$substr": ["$date", 0, 7]}, "ad_product": "$ad_product"},
            "cost": {"$sum": "$cost"},
            "sales": {"$sum": "$sales"},
            "units": {"$sum": "$units"},
        }},
    ]
    for r in db[PRODUCT_DAILY_COLLECTION].aggregate(pipeline):
        k = (r["_id"]["asin"], r["_id"]["month"], r["_id"]["ad_product"])
        out[k] = {"cost": r["cost"], "sales": r["sales"], "units": r["units"]}
    return out


_TOKEN_RE = re.compile(r"[A-Za-z0-9]+")


def _load_brand_ads_by_asin_sync(db, start: str, end: str, sku_to_asin: dict) -> tuple[dict, dict]:
    """
    Sponsored Brands has no per-ASIN report — spend is brand-level by design.

    Campaign names embed SKU codes (e.g. "FOSE04OCBL/SB/KT/Gen"), so spend is
    attributed by matching a name token against a known SKU. This is an
    approximation and is deliberately kept in its own columns; whatever cannot
    be matched is returned separately rather than being silently spread.
    """
    mapped: dict[tuple[str, str], float] = {}
    unmapped: dict[str, float] = {}

    pipeline = [
        {"$match": {"date": {"$gte": start, "$lte": end}, "ad_product": "SPONSORED_BRANDS"}},
        {"$group": {
            "_id": {"name": "$campaign_name", "month": {"$substr": ["$date", 0, 7]}},
            "cost": {"$sum": "$cost"},
        }},
    ]
    for r in db[DAILY_COLLECTION].aggregate(pipeline):
        name = r["_id"].get("name") or ""
        month = r["_id"]["month"]
        cost = r["cost"] or 0
        asin = None
        for tok in _TOKEN_RE.findall(name.upper()):
            if tok in sku_to_asin:
                asin = sku_to_asin[tok]
                break
        if asin:
            mapped[(asin, month)] = mapped.get((asin, month), 0.0) + cost
        else:
            unmapped[month] = unmapped.get(month, 0.0) + cost
    return mapped, unmapped


def _load_statics_sync(db) -> dict:
    """ASIN -> identity and unit economics."""
    # etrade_* fields are mirrored from the Etrade Master sheet by
    # POST /amazon/sync-etrade-master and are the primary source for ASP,
    # margin and GST — GST in particular is per-SKU (5% or 18%), not flat.
    mapping = {}
    for m in db[SKU_MAPPING_COLLECTION].find(
        {},
        {"item_id": 1, "sku_code": 1, "item_name": 1, "amazon_status": 1,
         "etrade_asp": 1, "etrade_mrp": 1, "etrade_margin": 1, "etrade_gst": 1,
         "etrade_cost_price_wo_tax": 1, "etrade_status": 1, "etrade_brand": 1, "_id": 0},
    ):
        asin = m.get("item_id")
        if asin:
            mapping[asin] = {
                "sku_code": m.get("sku_code"),
                "item_name": m.get("item_name"),
                "amazon_status": m.get("amazon_status"),
                "etrade_asp": m.get("etrade_asp"),
                "etrade_mrp": m.get("etrade_mrp"),
                "etrade_margin": m.get("etrade_margin"),
                "etrade_gst": m.get("etrade_gst"),
                "etrade_cost_price_wo_tax": m.get("etrade_cost_price_wo_tax"),
                "etrade_status": m.get("etrade_status"),
                "etrade_brand": m.get("etrade_brand"),
            }

    margins = {}
    for v in db[VENDOR_MARGINS_COLLECTION].find(
        {}, {"asin": 1, "margin": 1, "etrade_asp": 1, "cost_price_wo_tax": 1, "_id": 0}
    ):
        if v.get("asin"):
            margins[v["asin"]] = v

    sku_codes = {v["sku_code"] for v in mapping.values() if v.get("sku_code")}
    products = {}
    for p in db["products"].find(
        {"cf_sku_code": {"$in": list(sku_codes)}},
        {"cf_sku_code": 1, "name": 1, "item_name": 1, "status": 1, "purchase_status": 1,
         "brand": 1, "rate": 1, "category": 1, "sub_category": 1, "_id": 0},
    ):
        cf = p.get("cf_sku_code")
        if not cf:
            continue
        # A SKU code can appear twice (e.g. a GST-rate change created a second
        # entry). Prefer active rows, and among equals prefer the one that
        # actually carries category/brand — the stale twin is often blank.
        prev = products.get(cf)
        if prev is None:
            products[cf] = p
            continue
        score = lambda d: ((d.get("status") == "active") * 2) + bool(d.get("category")) + bool(d.get("brand"))
        if score(p) > score(prev):
            products[cf] = p

    return {"mapping": mapping, "margins": margins, "products": products}


# --------------------------------------------------------------------------
# Workbook
# --------------------------------------------------------------------------
_STATIC_COLUMNS = [
    # "Is Parent" comes from amazon_sku_mapping.amazon_status == 'Parent' — a
    # flag on the listing, not a child->parent mapping (Amazon gives us no such
    # mapping). Category/Sub Category stand in for the manual "series" grouping.
    "ASIN", "SKU Code", "Item Name", "Brand", "Category", "Sub Category", "Is Parent",
    "Status", "SKU Launch",
    "MRP", "ASP", "Etrade Margin", "GST",
    "Basic Cost (ASP)", "Cash Collected with GST", "Cash Collected",
    "COGS", "Avg Transport Cost",
]
# Per-month block. Value columns are written as numbers; the rest are formulas
# so the sheet recalculates when the static economics are edited.
_MONTH_COLUMNS = [
    # "Amazon Sales" is Amazon's gross ordered revenue; "Gross Sales" keeps the
    # manual sheet's meaning of Units x Cash Collected (net of margin and GST).
    # TACOS divides by the latter, matching how the manual sheet computed it.
    "Units", "Amazon Sales (gross)", "Sessions", "Conv %",
    "SP Ad Spend", "SD Ad Spend", "SB Ad Spend (approx)", "Total Ad Spend", "Ad Sales",
    "TACOS %", "ACOS %", "Ad Spend / Unit", "Per Unit PNL", "PNL", "Gross Sales", "PNL %",
]
HEADER_ROW = 3
DATA_START_ROW = 4


def _build_workbook_sync(db, months_meta: list, rows: list[dict], unmapped_sb: dict,
                         cogs_date: str | None, sales_source: str = "combined") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter as gcl

    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"

    # --- headers -----------------------------------------------------------
    for i, name in enumerate(_STATIC_COLUMNS, start=1):
        ws.cell(HEADER_ROW, i, name)
    col = len(_STATIC_COLUMNS) + 1
    month_col_start = {}
    for mk, label, _s, _e in months_meta:
        month_col_start[mk] = col
        ws.cell(1, col, label)  # month banner above the block
        for name in _MONTH_COLUMNS:
            ws.cell(HEADER_ROW, col, f"{label} {name}")
            col += 1
    last_col = col - 1
    last_row = DATA_START_ROW + len(rows) - 1

    # --- data --------------------------------------------------------------
    S = {n: i + 1 for i, n in enumerate(_STATIC_COLUMNS)}
    for idx, r in enumerate(rows):
        rw = DATA_START_ROW + idx
        ws.cell(rw, S["ASIN"], r["asin"])
        ws.cell(rw, S["SKU Code"], r.get("sku_code"))
        ws.cell(rw, S["Item Name"], r.get("item_name"))
        ws.cell(rw, S["Brand"], r.get("brand"))
        ws.cell(rw, S["Category"], r.get("category"))
        ws.cell(rw, S["Sub Category"], r.get("sub_category"))
        ws.cell(rw, S["Is Parent"], "Yes" if r.get("is_parent") else "")
        ws.cell(rw, S["Status"], r.get("status"))
        ws.cell(rw, S["SKU Launch"], r.get("launch"))
        ws.cell(rw, S["MRP"], r.get("mrp"))
        ws.cell(rw, S["ASP"], r.get("asp"))
        ws.cell(rw, S["Etrade Margin"], r.get("margin"))
        ws.cell(rw, S["GST"], r.get("gst"))

        asp, margin, gst = gcl(S["ASP"]), gcl(S["Etrade Margin"]), gcl(S["GST"])
        cash_gst, cogs_c, tr_c = gcl(S["Cash Collected with GST"]), gcl(S["COGS"]), gcl(S["Avg Transport Cost"])
        cash_c = gcl(S["Cash Collected"])

        ws.cell(rw, S["Basic Cost (ASP)"], f"=IFERROR({asp}{rw}/(1+{gst}{rw}),0)")
        ws.cell(rw, S["Cash Collected with GST"], f"=IFERROR({asp}{rw}*(1-{margin}{rw}),0)")
        ws.cell(rw, S["Cash Collected"], f"=IFERROR({cash_gst}{rw}/(1+{gst}{rw}),0)")
        ws.cell(rw, S["COGS"], r.get("cogs"))
        ws.cell(rw, S["Avg Transport Cost"], r.get("transport"))

        for mk, *_ in months_meta:
            c0 = month_col_start[mk]
            m = r["months"].get(mk, {})
            C = {n: gcl(c0 + i) for i, n in enumerate(_MONTH_COLUMNS)}

            ws.cell(rw, c0 + 0, m.get("units", 0))
            ws.cell(rw, c0 + 1, round(m.get("sales", 0), 2))
            ws.cell(rw, c0 + 2, m.get("sessions", 0))
            ws.cell(rw, c0 + 3, f"=IFERROR({C['Units']}{rw}/{C['Sessions']}{rw},0)")
            ws.cell(rw, c0 + 4, round(m.get("sp_cost", 0), 2))
            ws.cell(rw, c0 + 5, round(m.get("sd_cost", 0), 2))
            ws.cell(rw, c0 + 6, round(m.get("sb_cost", 0), 2))
            ws.cell(rw, c0 + 7,
                    f"={C['SP Ad Spend']}{rw}+{C['SD Ad Spend']}{rw}+{C['SB Ad Spend (approx)']}{rw}")
            ws.cell(rw, c0 + 8, round(m.get("ad_sales", 0), 2))
            ws.cell(rw, c0 + 9, f"=IFERROR({C['Total Ad Spend']}{rw}/{C['Gross Sales']}{rw},0)")
            ws.cell(rw, c0 + 10, f"=IFERROR({C['Total Ad Spend']}{rw}/{C['Ad Sales']}{rw},0)")
            ws.cell(rw, c0 + 11, f"=IFERROR({C['Total Ad Spend']}{rw}/{C['Units']}{rw},0)")
            # Contribution per unit after cost, freight and advertising.
            ws.cell(rw, c0 + 12,
                    f"=IFERROR({cash_c}{rw}-{cogs_c}{rw}-{tr_c}{rw}-{C['Ad Spend / Unit']}{rw},0)")
            ws.cell(rw, c0 + 13, f"=IFERROR({C['Per Unit PNL']}{rw}*{C['Units']}{rw},0)")
            ws.cell(rw, c0 + 14, f"=IFERROR({C['Units']}{rw}*{cash_c}{rw},0)")
            ws.cell(rw, c0 + 15, f"=IFERROR({C['PNL']}{rw}/{C['Gross Sales']}{rw},0)")

    # --- totals row (row 2), SUBTOTAL so it follows filtering ---------------
    if rows:
        for mk, *_ in months_meta:
            c0 = month_col_start[mk]
            for i, name in enumerate(_MONTH_COLUMNS):
                c = c0 + i
                letter = gcl(c)
                if name in ("Units", "Amazon Sales (gross)", "Sessions", "SP Ad Spend", "SD Ad Spend",
                            "SB Ad Spend (approx)", "Total Ad Spend", "Ad Sales", "PNL", "Gross Sales"):
                    ws.cell(2, c, f"=SUBTOTAL(9,{letter}{DATA_START_ROW}:{letter}{last_row})")
            C = {n: gcl(c0 + i) for i, n in enumerate(_MONTH_COLUMNS)}
            # Ratios at the top must divide the totals, never average the rows.
            ws.cell(2, c0 + 9, f"=IFERROR({C['Total Ad Spend']}2/{C['Gross Sales']}2,0)")
            ws.cell(2, c0 + 10, f"=IFERROR({C['Total Ad Spend']}2/{C['Ad Sales']}2,0)")
            ws.cell(2, c0 + 15, f"=IFERROR({C['PNL']}2/{C['Gross Sales']}2,0)")
        ws.cell(2, 1, "TOTALS")

    # --- formatting --------------------------------------------------------
    head_fill = PatternFill("solid", fgColor="1F3864")
    for c in range(1, last_col + 1):
        cell = ws.cell(HEADER_ROW, c)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = head_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for c in range(1, last_col + 1):
        ws.cell(2, c).font = Font(bold=True)
    ws.row_dimensions[HEADER_ROW].height = 42
    # Freeze only ASIN / SKU / Item Name. Freezing the whole static block made
    # the locked region wider than the screen, leaving nothing to scroll.
    ws.freeze_panes = ws.cell(DATA_START_ROW, 4)
    ws.auto_filter.ref = f"A{HEADER_ROW}:{gcl(last_col)}{max(last_row, DATA_START_ROW)}"

    pct = {"Conv %", "TACOS %", "ACOS %", "PNL %"}
    for mk, *_ in months_meta:
        c0 = month_col_start[mk]
        for i, name in enumerate(_MONTH_COLUMNS):
            letter = gcl(c0 + i)
            ws.column_dimensions[letter].width = 15
            if name in pct:
                for rw in list(range(DATA_START_ROW, last_row + 1)) + [2]:
                    ws.cell(rw, c0 + i).number_format = "0.00%"
    for i, name in enumerate(_STATIC_COLUMNS, start=1):
        ws.column_dimensions[gcl(i)].width = 22 if name in ("Item Name", "SKU Code") else 13
    for rw in range(DATA_START_ROW, last_row + 1):
        ws.cell(rw, S["Etrade Margin"]).number_format = "0.00%"
        ws.cell(rw, S["GST"]).number_format = "0.00%"

    # --- notes sheet -------------------------------------------------------
    notes = wb.create_sheet("Notes")
    notes.column_dimensions["A"].width = 34
    notes.column_dimensions["B"].width = 96
    lines = [
        ("Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Rows", str(len(rows))),
        ("", ""),
        ("Sales source", sales_source),
        ("Units", {"combined": "Vendor SHIPPED units + Seller Central units (both channels)",
                   "vendor": "Vendor Central ordered units",
                   "seller": "Seller Central units"}.get(sales_source, sales_source)),
        ("Sales", {"combined": "Vendor ORDERED revenue + Seller Central sales. Vendor shippedRevenue "
                               "is reported as 0 by Amazon, so revenue uses the ordered measure while "
                               "units use shipped.",
                   "vendor": "Vendor Central ordered revenue",
                   "seller": "Seller Central ordered product sales"}.get(sales_source, sales_source)),
        ("Sessions / Conv %", "Sessions exist only for Seller Central. In combined mode units span both "
                              "channels, so conversion rate is not meaningful and is left blank."
                              if sales_source != "seller" else "amazon_sales_traffic"),
        ("SP / SD Ad Spend", "amazon_ads_product_daily — true ASIN-level spend (spAdvertisedProduct / sdAdvertisedProduct)"),
        ("SB Ad Spend (approx)", "Sponsored Brands has no per-ASIN report. Spend is attributed by matching a SKU "
                                 "code inside the campaign name. Treat as indicative, not exact."),
        ("ASP / Margin", "vendor_margins (etrade_asp, margin)"),
        ("COGS", f"Zoho Books inventory valuation, Pupscribe WH, as of {cogs_date or 'n/a'}"),
        ("Avg Transport Cost", f"Flat INR {DEFAULT_TRANSPORT_COST} per unit"),
        ("", ""),
        ("Cash Collected with GST", "= ASP x (1 - Etrade Margin)"),
        ("Cash Collected", "= Cash Collected with GST / (1 + GST)"),
        ("Ad Spend / Unit", "= Total Ad Spend / Units"),
        ("Per Unit PNL", "= Cash Collected - COGS - Avg Transport Cost - Ad Spend / Unit"),
        ("PNL", "= Per Unit PNL x Units"),
        ("Gross Sales", "= Units x Cash Collected — net of Amazon margin and GST. This is the "
                        "manual sheet's 'Gross Sales' definition and the TACOS/PNL denominator."),
        ("Amazon Sales (gross)", "Amazon's reported ordered revenue, before margin and GST. Shown for "
                                 "reference; it is NOT the P&L basis."),
        ("TACOS %", "= Total Ad Spend / Gross Sales (net realisation), matching the manual sheet. "
                    "Dividing by Amazon's gross revenue instead would read roughly a third lower."),
        ("ACOS %", "= Total Ad Spend / Ad Sales"),
        ("", ""),
        ("Note", "Row 2 uses SUBTOTAL(9,...) so totals follow the filter. Ratios in row 2 divide "
                 "the totals rather than averaging rows."),
    ]
    # Months with sales but no ad data (retention already dropped them) would
    # otherwise read as 0% TACOS and an inflated PNL. Call them out.
    no_ads = []
    for mk, label, *_ in months_meta:
        sales = sum(r["months"].get(mk, {}).get("sales", 0) or 0 for r in rows)
        spend = sum((r["months"].get(mk, {}).get("sp_cost", 0) or 0)
                    + (r["months"].get(mk, {}).get("sd_cost", 0) or 0)
                    + (r["months"].get(mk, {}).get("sb_cost", 0) or 0) for r in rows)
        if sales > 0 and spend <= 0:
            no_ads.append(label)
    if no_ads:
        lines.append(("", ""))
        lines.append(("WARNING — no ad data", ", ".join(no_ads)))
        lines.append(("", "These months have sales but zero recorded ad spend, because Amazon's "
                          "reporting retention had already dropped them when the data was first "
                          "pulled. TACOS reads 0% and PNL is overstated for those months — they "
                          "are not comparable with the rest."))

    if unmapped_sb:
        lines.append(("", ""))
        lines.append(("Unmapped SB spend", "Sponsored Brands spend whose campaign name held no known SKU code:"))
        for mk in sorted(unmapped_sb):
            lines.append((f"   {mk}", f"INR {unmapped_sb[mk]:,.2f} not attributed to any ASIN"))
    for i, (k, v) in enumerate(lines, start=1):
        notes.cell(i, 1, k).font = Font(bold=True)
        notes.cell(i, 2, v)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# --------------------------------------------------------------------------
# Assembly
# --------------------------------------------------------------------------
def _assemble_rows_sync(db, months_meta: list, start: str, end: str, cogs_by_sku: dict,
                        sales_source: str = "combined") -> tuple[list, dict]:
    statics = _load_statics_sync(db)
    mapping, margins, products = statics["mapping"], statics["margins"], statics["products"]

    sku_to_asin = {v["sku_code"].upper(): a for a, v in mapping.items() if v.get("sku_code")}

    # The ads profile is a vendor entity, so vendor sales are the default and
    # correct denominator. Seller Central remains available for the SC channel.
    if sales_source == "vendor":
        per_month, _ = _load_vendor_sales_sync(db, start, end)
    elif sales_source == "seller":
        per_month, _, _ = _load_sales_traffic_sync(db, start, end)
    elif sales_source == "combined":
        per_month = _combine_channels(*_load_vendor_sales_sync(db, start, end)[:1],
                                      _load_sales_traffic_sync(db, start, end)[0])
    else:
        raise HTTPException(
            status_code=400, detail="sales_source must be 'vendor', 'seller' or 'combined'"
        )

    # Parent ASIN has no source: Seller Central rows here are keyed by
    # parentAsin with no child, and Vendor Central carries no parent/child at
    # all. The column stays for layout parity with the manual sheet and is
    # filled once a mapping is supplied.
    parents: dict[str, str] = {}
    # Derived from full history, not the reporting window.
    first_seen = _load_first_seen_all_time_sync(db)
    product_ads = _load_product_ads_sync(db, start, end)
    sb_by_asin, sb_unmapped = _load_brand_ads_by_asin_sync(db, start, end, sku_to_asin)

    asins = {a for a, _ in per_month}
    asins |= {k[0] for k in product_ads}
    asins |= {k[0] for k in sb_by_asin}

    today = datetime.now(timezone.utc).date()
    rows = []
    for asin in sorted(asins):
        info = mapping.get(asin, {})
        sku = info.get("sku_code")
        prod = products.get(sku, {}) if sku else {}
        vm = margins.get(asin, {})

        fs = first_seen.get(asin)
        launch = ""
        if fs:
            launch = "New" if (today - date.fromisoformat(fs)).days <= NEW_LAUNCH_DAYS else "Old"

        months = {}
        has_activity = False
        for mk, *_ in months_meta:
            st = per_month.get((asin, mk), {})
            sp = product_ads.get((asin, mk, "SPONSORED_PRODUCTS"), {})
            sd = product_ads.get((asin, mk, "SPONSORED_DISPLAY"), {})
            entry = {
                "units": st.get("units", 0),
                "sales": st.get("sales", 0),
                "sessions": st.get("sessions", 0),
                "sp_cost": sp.get("cost", 0),
                "sd_cost": sd.get("cost", 0),
                "sb_cost": sb_by_asin.get((asin, mk), 0),
                "ad_sales": sp.get("sales", 0) + sd.get("sales", 0),
            }
            if any(entry.values()):
                has_activity = True
            months[mk] = entry

        # An ASIN with no units and no spend in the whole window is noise.
        if not has_activity:
            continue

        rows.append({
            "asin": asin,
            "sku_code": sku,
            "item_name": prod.get("name") or info.get("item_name"),
            "brand": prod.get("brand"),
            "category": prod.get("category"),
            "sub_category": prod.get("sub_category"),
            "is_parent": info.get("amazon_status") == "Parent",
            # Etrade Master (mirrored to the DB) wins; vendor_margins and the
            # Zoho product are fallbacks for ASINs the sheet does not cover.
            "status": (info.get("etrade_status") or prod.get("purchase_status")
                       or prod.get("status") or info.get("amazon_status")),
            "launch": launch,
            "mrp": info.get("etrade_mrp") or prod.get("rate"),
            "asp": info.get("etrade_asp") or vm.get("etrade_asp") or prod.get("rate"),
            "margin": info.get("etrade_margin") or vm.get("margin") or DEFAULT_MARGIN,
            "gst": info.get("etrade_gst") if info.get("etrade_gst") is not None else DEFAULT_GST,
            "cogs": (cogs_by_sku.get(sku) or {}).get("unit_cost") if sku else None,
            "transport": DEFAULT_TRANSPORT_COST,
            "months": months,
        })

    return rows, sb_unmapped


@router.get("/report/download")
async def download_metrics_report(
    end_date: str | None = Query(None, description="Last day to include (default: yesterday)"),
    months: int = Query(DEFAULT_MONTHS, ge=1, le=24),
    include_cogs: bool = Query(True, description="Fetch unit cost from Zoho Books (adds ~30s)"),
    sales_source: str = Query("combined", description="'combined' (vendor + seller, matches the manual sheet), 'vendor' or 'seller'"),
    db=Depends(get_database),
):
    """Generate the amz-metrics workbook with live formulas."""
    end = date.fromisoformat(end_date) if end_date else datetime.now(timezone.utc).date() - timedelta(days=1)
    months_meta = _month_windows(end, months)
    start = months_meta[0][2]
    end_str = months_meta[-1][3]

    cogs_by_sku, cogs_date = {}, None
    if include_cogs:
        try:
            from ..services.master_service import OptimizedMasterReportService

            svc = OptimizedMasterReportService(db)
            res = await svc.fetch_inventory_valuation_cogs(end_str)
            cogs_by_sku = (res or {}).get("by_sku", {}) or {}
            cogs_date = (res or {}).get("as_of_date")
        except Exception as e:
            # COGS is one column; losing it must not lose the workbook.
            logger.error("amz-metrics: COGS fetch failed, continuing without it: %s", e)

    rows, sb_unmapped = await asyncio.to_thread(
        _assemble_rows_sync, db, months_meta, start, end_str, cogs_by_sku, sales_source
    )
    if not rows:
        raise HTTPException(status_code=404, detail=f"No Amazon activity found between {start} and {end_str}")

    content = await asyncio.to_thread(
        _build_workbook_sync, db, months_meta, rows, sb_unmapped, cogs_date, sales_source
    )
    fname = f"amz_metrics_{start}_to_{end_str}.xlsx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/report/preview")
async def preview_metrics_report(
    end_date: str | None = Query(None),
    months: int = Query(DEFAULT_MONTHS, ge=1, le=24),
    limit: int = Query(20, le=200),
    sales_source: str = Query("combined"),
    db=Depends(get_database),
):
    """Same aggregation as the download, as JSON — for sanity-checking numbers."""
    end = date.fromisoformat(end_date) if end_date else datetime.now(timezone.utc).date() - timedelta(days=1)
    months_meta = _month_windows(end, months)
    start, end_str = months_meta[0][2], months_meta[-1][3]

    rows, sb_unmapped = await asyncio.to_thread(
        _assemble_rows_sync, db, months_meta, start, end_str, {}, sales_source
    )

    totals = {}
    for mk, label, *_ in months_meta:
        t = {"units": 0.0, "sales": 0.0, "sessions": 0.0, "sp_cost": 0.0,
             "sd_cost": 0.0, "sb_cost": 0.0, "ad_sales": 0.0}
        for r in rows:
            for k in t:
                t[k] += r["months"].get(mk, {}).get(k, 0) or 0
        t["total_ad_spend"] = round(t["sp_cost"] + t["sd_cost"] + t["sb_cost"], 2)
        t["tacos_pct"] = round(t["total_ad_spend"] / t["sales"] * 100, 2) if t["sales"] else None
        totals[label] = {k: (round(v, 2) if isinstance(v, float) else v) for k, v in t.items()}

    return {
        "period": {"start": start, "end": end_str},
        "sales_source": sales_source,
        "months": [m[1] for m in months_meta],
        "asin_count": len(rows),
        "monthly_totals": totals,
        "unmapped_sb_spend": {k: round(v, 2) for k, v in sb_unmapped.items()},
        "sample": [{"asin": r["asin"], "sku_code": r["sku_code"], "item_name": r["item_name"],
                    "asp": r["asp"], "margin": r["margin"], "months": r["months"]} for r in rows[:limit]],
    }
