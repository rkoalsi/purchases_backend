"""
Amazon Listing Queue.

Lets a manager select products (from the Zoho product list) that still need to be
listed on Amazon, and lets a lister mark each one "listed" once done. Both actions
are timestamped and attributed:

  requested_at / requested_by  — manager added the product to the queue
  listed_at    / listed_by     — lister completed the Amazon listing

The page also shows an "Already Listed on Amazon" reference built from
amazon_sku_mapping (see routes/amazon.py get_amazon_sku_mapping).

Auth model follows the rest of the app: endpoints are frontend-gated and the acting
user's identity is passed in the request body (like tasks.py created_by/created_by_name).
"""

import asyncio
import io
import logging
import math
import re
from datetime import datetime, timedelta, timezone
from typing import List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from bson import ObjectId
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel
from pymongo.errors import PyMongoError

from ..database import get_database, serialize_mongo_document
from ..helpers.datetime_utils import utcnow, to_ist_str

router = APIRouter()
logger = logging.getLogger(__name__)

QUEUE_COLLECTION = "amazon_listing_queue"
PRODUCTS_COLLECTION = "products"
SKU_MAPPING_COLLECTION = "amazon_sku_mapping"

_ACTIVE_STATUSES = ("pending", "listed")
IST = timezone(timedelta(hours=5, minutes=30))

_indexes_ready = False


def _ensure_indexes(database) -> None:
    """Create indexes needed by the products join / queue queries (idempotent, once)."""
    global _indexes_ready
    if _indexes_ready:
        return
    try:
        database.get_collection(SKU_MAPPING_COLLECTION).create_index("sku_code")
        # amazon_sku_mapping.item_id already has a unique index (used by ASIN search)
        database.get_collection(PRODUCTS_COLLECTION).create_index("cf_sku_code")
        database.get_collection(QUEUE_COLLECTION).create_index([("requested_at", -1), ("_id", -1)])
        database.get_collection(QUEUE_COLLECTION).create_index("product_id")
    except Exception as e:  # non-fatal — queries still work, just slower
        logger.warning(f"amazon_listing index ensure failed: {e}")
    _indexes_ready = True


# ─── Request models ───────────────────────────────────────────────────────────

class AddToQueueRequest(BaseModel):
    product_ids: List[str]
    requested_by: str
    requested_by_name: str
    notes: Optional[str] = None


class UpdateStatusRequest(BaseModel):
    status: str  # "pending" | "listed"
    listed_by: Optional[str] = None
    listed_by_name: Optional[str] = None


# ─── Shared query pipeline ────────────────────────────────────────────────────

def _ist_day_range(date_str: str) -> tuple[datetime, datetime]:
    """Convert an IST 'YYYY-MM-DD' to a [start, end) naive-UTC datetime range."""
    d = datetime.strptime(date_str, "%Y-%m-%d")
    start_ist = datetime(d.year, d.month, d.day, tzinfo=IST)
    start_utc = start_ist.astimezone(timezone.utc).replace(tzinfo=None)
    return start_utc, start_utc + timedelta(days=1)


def _base_pipeline(status: str, search, brand, purchase_status, date) -> list:
    """Stages that filter the queue: queue-level match → product lookup → live-field match.

    brand / purchase_status filter the *live* product fields (they can change after
    the item was queued); status/search/date filter the queue document itself.
    """
    queue_match: dict = {}
    if status in ("pending", "listed"):
        queue_match["status"] = status
    if search:
        rgx = {"$regex": re.escape(search), "$options": "i"}
        queue_match["$or"] = [{"product_name": rgx}, {"sku_code": rgx}]
    if date:
        start, end = _ist_day_range(date)
        queue_match["requested_at"] = {"$gte": start, "$lt": end}

    stages: list = [
        {"$match": queue_match},
        {"$addFields": {
            "_pid": {"$convert": {"input": "$product_id", "to": "objectId", "onError": None, "onNull": None}}
        }},
        {"$lookup": {
            "from": PRODUCTS_COLLECTION,
            "localField": "_pid",
            "foreignField": "_id",
            "as": "_product",
            "pipeline": [{"$project": {"name": 1, "brand": 1, "status": 1, "purchase_status": 1, "cf_sku_code": 1}}],
        }},
        {"$addFields": {"product": {"$arrayElemAt": ["$_product", 0]}}},
    ]

    if brand or purchase_status:
        conds = []
        if brand:
            conds.append({"$eq": [{"$ifNull": ["$product.brand", "$brand"]}, brand]})
        if purchase_status:
            conds.append({"$eq": [{"$ifNull": ["$product.purchase_status", "$purchase_status"]}, purchase_status]})
        stages.append({"$match": {"$expr": {"$and": conds}}})

    return stages


def _finalize_items(items: list) -> list:
    """Flatten the joined product onto each row (live value, snapshot fallback)."""
    out = []
    for it in items:
        prod = it.pop("product", None) or {}
        it.pop("_product", None)
        it.pop("_pid", None)
        it["product_name"] = prod.get("name") or it.get("product_name")
        it["brand"] = prod.get("brand") or it.get("brand")
        it["sku_code"] = prod.get("cf_sku_code") or it.get("sku_code")
        it["purchase_status"] = prod.get("purchase_status") or it.get("purchase_status")
        it["zoho_status"] = prod.get("status")
        out.append(it)
    return out


# ─── Queue listing ──────────────────────────────────────────────────────────--

@router.get("/queue")
async def list_queue(
    page: int = Query(1, ge=1),
    limit: int = Query(500, ge=1, le=2000),
    status: str = Query("all", description="pending | listed | all"),
    search: str = Query(None, description="Search by product name or SKU"),
    brand: str = Query(None, description="Filter by live product brand"),
    purchase_status: str = Query(None, description="Filter by live product purchase_status"),
    date: str = Query(None, description="IST day 'YYYY-MM-DD' (a requested-date group)"),
    database=Depends(get_database),
):
    """Queue entries joined to the live product for name/brand/purchase-status/zoho-status."""
    def _fetch():
        col = database.get_collection(QUEUE_COLLECTION)
        base = _base_pipeline(status, search, brand, purchase_status, date)

        count_res = list(col.aggregate(base + [{"$count": "n"}]))
        total_count = count_res[0]["n"] if count_res else 0
        total_pages = math.ceil(total_count / limit) if total_count > 0 else 1
        skip = (page - 1) * limit

        pipeline = base + [
            {"$sort": {"requested_at": -1, "_id": -1}},
            {"$skip": skip},
            {"$limit": limit},
            {"$project": {"_product": 0, "_pid": 0}},
        ]
        items = _finalize_items(serialize_mongo_document(list(col.aggregate(pipeline))))

        return {
            "items": items,
            "pagination": {
                "currentPage": page,
                "totalPages": total_pages,
                "totalItems": total_count,
                "limit": limit,
            },
        }

    try:
        result = await asyncio.to_thread(_fetch)
        return JSONResponse(content=result)
    except PyMongoError as e:
        logger.error(f"DB error in amazon_listing/queue: {e}")
        raise HTTPException(status_code=500, detail=f"Database error: {e}")


@router.get("/queue/groups")
async def queue_groups(
    status: str = Query("all"),
    search: str = Query(None),
    brand: str = Query(None),
    purchase_status: str = Query(None),
    database=Depends(get_database),
):
    """Requested-date groups (IST day) with counts, for the collapsed accordion view."""
    def _fetch():
        col = database.get_collection(QUEUE_COLLECTION)
        base = _base_pipeline(status, search, brand, purchase_status, None)
        pipeline = base + [
            {"$group": {
                "_id": {"$dateToString": {"format": "%Y-%m-%d", "date": "$requested_at", "timezone": "+05:30"}},
                "count": {"$sum": 1},
                "pending": {"$sum": {"$cond": [{"$eq": ["$status", "pending"]}, 1, 0]}},
                "listed": {"$sum": {"$cond": [{"$eq": ["$status", "listed"]}, 1, 0]}},
            }},
            {"$sort": {"_id": -1}},
        ]
        groups = [
            {"date": g["_id"], "count": g["count"], "pending": g["pending"], "listed": g["listed"]}
            for g in col.aggregate(pipeline)
            if g["_id"]
        ]
        return {"groups": groups}

    result = await asyncio.to_thread(_fetch)
    return JSONResponse(content=result)


@router.get("/queue/sku-status")
async def queue_sku_status(database=Depends(get_database)):
    """Lightweight {sku_code: status} map of all active queue entries, for badging
    the Zoho-products browser without an N+1 of per-row lookups."""
    def _fetch():
        col = database.get_collection(QUEUE_COLLECTION)
        docs = col.find(
            {"status": {"$in": list(_ACTIVE_STATUSES)}},
            {"sku_code": 1, "product_id": 1, "status": 1},
        )
        by_sku: dict = {}
        by_pid: dict = {}
        for d in docs:
            if d.get("sku_code"):
                by_sku[d["sku_code"]] = d["status"]
            if d.get("product_id"):
                by_pid[d["product_id"]] = d["status"]
        return {"by_sku": by_sku, "by_product_id": by_pid}

    result = await asyncio.to_thread(_fetch)
    return JSONResponse(content=result)


# ─── Products browser (joined with Amazon listing status) ─────────────────────

@router.get("/products")
async def list_products(
    page: int = Query(1, ge=1),
    limit: int = Query(25, ge=1, le=200),
    search: str = Query(None, description="Search by product name, SKU or sku"),
    brand: str = Query(None),
    purchase_status: str = Query(None),
    zoho_status: str = Query(None, description="active | inactive"),
    listed: str = Query("all", description="all | listed | not_listed (matched via amazon_sku_mapping)"),
    database=Depends(get_database),
):
    """Non-combo products annotated with whether they are already listed on Amazon.

    A product is "listed" when a matching amazon_sku_mapping doc exists
    (amazon_sku_mapping.sku_code == products.cf_sku_code). Exposes the matched
    amazon_status / active_platforms / ASIN so the row can be shown + selection disabled.
    """
    def _fetch():
        col = database.get_collection(PRODUCTS_COLLECTION)
        _ensure_indexes(database)

        # Include all non-combo products (any status) PLUS *active* combo products.
        # Combos are Zoho-status "inactive" by design (not sold standalone); their
        # active marker is purchase_status == "active - combo".
        # Built as an $and list so multiple $or blocks (this + search) don't collide.
        and_conds: list = [
            {"$or": [
                {"is_combo_product": {"$ne": True}},
                {"is_combo_product": True, "purchase_status": "active - combo"},
            ]}
        ]
        if brand:
            and_conds.append({"brand": brand})
        if purchase_status:
            and_conds.append({"purchase_status": purchase_status})
        if zoho_status:
            and_conds.append({"status": zoho_status.lower()})
        if search:
            rgx = {"$regex": re.escape(search), "$options": "i"}
            or_conds = [{"name": rgx}, {"cf_sku_code": rgx}, {"sku": rgx}]
            # also match by ASIN → resolve to the linked SKU(s) in amazon_sku_mapping
            asin_skus = [
                s for s in database.get_collection(SKU_MAPPING_COLLECTION).distinct(
                    "sku_code", {"item_id": rgx}
                ) if s
            ]
            if asin_skus:
                or_conds.append({"cf_sku_code": {"$in": asin_skus}})
            and_conds.append({"$or": or_conds})

        # Indexable lookup (localField/foreignField uses the index on sku_code).
        # is_listed is blank-guarded so products with an empty cf_sku_code never
        # false-match amazon_sku_mapping rows that also have a blank sku_code.
        lookup_stages: list = [
            {"$lookup": {
                "from": SKU_MAPPING_COLLECTION,
                "localField": "cf_sku_code",
                "foreignField": "sku_code",
                "as": "_amz",
                "pipeline": [{"$project": {"item_id": 1, "amazon_status": 1, "active_platforms": 1, "_id": 0}}],
            }},
            {"$addFields": {
                "is_listed": {"$and": [
                    {"$gt": [{"$size": "$_amz"}, 0]},
                    {"$ne": ["$cf_sku_code", None]},
                    {"$ne": ["$cf_sku_code", ""]},
                ]},
                "_amazon": {"$arrayElemAt": ["$_amz", 0]},
            }},
        ]
        project = {"$project": {
            "cf_sku_code": 1, "name": 1, "brand": 1, "status": 1, "purchase_status": 1,
            "is_combo_product": {"$eq": ["$is_combo_product", True]},
            "is_listed": 1,
            "amazon_status": {"$cond": ["$is_listed", "$_amazon.amazon_status", None]},
            "active_platforms": {"$cond": ["$is_listed", "$_amazon.active_platforms", None]},
            "amazon_asin": {"$cond": ["$is_listed", "$_amazon.item_id", None]},
        }}

        # Fold the listing filter into the indexed $match instead of joining the
        # whole collection: resolve the set of listed SKUs once, then filter
        # cf_sku_code by $in / $nin. Every path then joins only the page's docs.
        if listed in ("listed", "not_listed"):
            listed_skus = [
                s for s in database.get_collection(SKU_MAPPING_COLLECTION).distinct("sku_code")
                if s
            ]
            if listed == "listed":
                and_conds.append({"cf_sku_code": {"$in": listed_skus}})
            else:  # not_listed → blank/missing cf_sku_code also counts as not listed
                and_conds.append({"cf_sku_code": {"$nin": listed_skus}})

        match = {"$and": and_conds}

        skip = (page - 1) * limit
        total_count = col.count_documents(match)
        pipeline = [
            {"$match": match},
            {"$sort": {"created_at": -1, "_id": -1}},
            {"$skip": skip},
            {"$limit": limit},
            *lookup_stages,
            project,
        ]

        total_pages = math.ceil(total_count / limit) if total_count > 0 else 1
        products = serialize_mongo_document(list(col.aggregate(pipeline)))

        return {
            "products": products,
            "pagination": {
                "currentPage": page,
                "totalPages": total_pages,
                "totalProducts": total_count,
                "limit": limit,
            },
        }

    try:
        result = await asyncio.to_thread(_fetch)
        return JSONResponse(content=result)
    except PyMongoError as e:
        logger.error(f"DB error in amazon_listing/products: {e}")
        raise HTTPException(status_code=500, detail=f"Database error: {e}")


# ─── Add products to the queue ──────────────────────────────────────────────--

@router.post("/queue")
async def add_to_queue(request: AddToQueueRequest, database=Depends(get_database)):
    """Add products to the listing queue. Skips products already queued or listed."""
    if not request.product_ids:
        raise HTTPException(status_code=400, detail="product_ids is required")

    def _add():
        col = database.get_collection(QUEUE_COLLECTION)
        products_col = database.get_collection(PRODUCTS_COLLECTION)
        now = utcnow()

        added: list = []
        skipped: list = []

        existing = {
            d["product_id"]
            for d in col.find(
                {"product_id": {"$in": request.product_ids}, "status": {"$in": list(_ACTIVE_STATUSES)}},
                {"product_id": 1},
            )
        }

        for pid in request.product_ids:
            if pid in existing:
                skipped.append({"product_id": pid, "reason": "already in queue"})
                continue

            prod = None
            try:
                prod = products_col.find_one(
                    {"_id": ObjectId(pid)},
                    {"name": 1, "brand": 1, "cf_sku_code": 1, "purchase_status": 1},
                )
            except Exception:
                prod = None

            if not prod:
                skipped.append({"product_id": pid, "reason": "product not found"})
                continue

            doc = {
                "product_id": pid,
                "sku_code": prod.get("cf_sku_code"),
                "product_name": prod.get("name"),
                "brand": prod.get("brand"),
                "purchase_status": prod.get("purchase_status"),
                "status": "pending",
                "notes": request.notes,
                "requested_by": request.requested_by,
                "requested_by_name": request.requested_by_name,
                "requested_at": now,
                "listed_by": None,
                "listed_by_name": None,
                "listed_at": None,
                "created_at": now,
                "updated_at": now,
            }
            res = col.insert_one(doc)
            doc["_id"] = res.inserted_id
            added.append(serialize_mongo_document(doc))

        return {"added": added, "skipped": skipped}

    try:
        result = await asyncio.to_thread(_add)
        return JSONResponse(content=result)
    except PyMongoError as e:
        logger.error(f"DB error adding to amazon_listing queue: {e}")
        raise HTTPException(status_code=500, detail=f"Database error: {e}")


# ─── Mark listed / revert ─────────────────────────────────────────────────────

@router.patch("/queue/{entry_id}/status")
async def update_status(entry_id: str, request: UpdateStatusRequest, database=Depends(get_database)):
    """Mark an entry listed (records listed_at/by) or revert it to pending (clears them)."""
    if request.status not in _ACTIVE_STATUSES:
        raise HTTPException(status_code=400, detail=f"status must be one of {_ACTIVE_STATUSES}")

    def _update():
        col = database.get_collection(QUEUE_COLLECTION)
        now = utcnow()

        if request.status == "listed":
            update = {"$set": {
                "status": "listed",
                "listed_by": request.listed_by,
                "listed_by_name": request.listed_by_name,
                "listed_at": now,
                "updated_at": now,
            }}
        else:  # revert to pending
            update = {"$set": {
                "status": "pending",
                "listed_by": None,
                "listed_by_name": None,
                "listed_at": None,
                "updated_at": now,
            }}

        return col.update_one({"_id": ObjectId(entry_id)}, update)

    try:
        result = await asyncio.to_thread(_update)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid entry id")

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Queue entry not found")
    return {"message": "Status updated"}


@router.delete("/queue/{entry_id}")
async def delete_entry(entry_id: str, database=Depends(get_database)):
    """Remove an entry from the queue (manager-only, UI-gated)."""
    def _delete():
        return database.get_collection(QUEUE_COLLECTION).delete_one({"_id": ObjectId(entry_id)})

    try:
        result = await asyncio.to_thread(_delete)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid entry id")

    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Queue entry not found")
    return {"message": "Deleted"}


# ─── Bulk xlsx download / upload (mark-listed only) ────────────────────────────

_XLSX_HEADERS = [
    "Entry ID", "SKU", "Product", "Brand", "Purchase Status",
    "Zoho Status", "Status", "Mark Listed (Yes/No)",
    "Requested By", "Requested At (IST)", "Listed By", "Listed At (IST)",
]
_TRUE_TOKENS = {"yes", "y", "true", "1", "listed", "done"}
_FALSE_TOKENS = {"no", "n", "false", "0", "pending"}


@router.get("/queue/download")
async def download_queue(
    status: str = Query("all"),
    search: str = Query(None),
    brand: str = Query(None),
    purchase_status: str = Query(None),
    date: str = Query(None),
    database=Depends(get_database),
):
    """Download the (filtered) queue as an editable xlsx.

    The lister sets 'Mark Listed (Yes/No)' to Yes and re-uploads to bulk mark listed.
    Only the listed status is updated on upload; matching is by Entry ID.
    """
    def _fetch():
        col = database.get_collection(QUEUE_COLLECTION)
        base = _base_pipeline(status, search, brand, purchase_status, date)
        pipeline = base + [
            {"$sort": {"requested_at": -1, "_id": -1}},
            {"$project": {"_product": 0, "_pid": 0}},
        ]
        return _finalize_items(serialize_mongo_document(list(col.aggregate(pipeline))))

    items = await asyncio.to_thread(_fetch)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Amazon Listing Queue"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(_XLSX_HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _ist(v):
        if not v:
            return ""
        try:
            return to_ist_str(datetime.fromisoformat(v)) if isinstance(v, str) else to_ist_str(v)
        except Exception:
            return str(v)

    for r, it in enumerate(items, start=2):
        listed = it.get("status") == "listed"
        ws.cell(row=r, column=1, value=str(it.get("_id", "")))
        ws.cell(row=r, column=2, value=it.get("sku_code") or "")
        ws.cell(row=r, column=3, value=it.get("product_name") or "")
        ws.cell(row=r, column=4, value=it.get("brand") or "")
        ws.cell(row=r, column=5, value=it.get("purchase_status") or "")
        ws.cell(row=r, column=6, value=it.get("zoho_status") or "")
        ws.cell(row=r, column=7, value=it.get("status") or "")
        ws.cell(row=r, column=8, value="Yes" if listed else "No")
        ws.cell(row=r, column=9, value=it.get("requested_by_name") or "")
        ws.cell(row=r, column=10, value=_ist(it.get("requested_at")))
        ws.cell(row=r, column=11, value=it.get("listed_by_name") or "")
        ws.cell(row=r, column=12, value=_ist(it.get("listed_at")))

    widths = [26, 16, 40, 16, 20, 14, 12, 20, 18, 22, 18, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="amazon_listing_queue.xlsx"'},
    )


@router.post("/queue/upload")
async def upload_queue(
    file: UploadFile = File(...),
    listed_by: str = Form(...),
    listed_by_name: str = Form(...),
    database=Depends(get_database),
):
    """Bulk update listed status from the edited xlsx. Matches rows by 'Entry ID';
    sets each entry to listed / pending based on 'Mark Listed (Yes/No)'."""
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")

    # Locate header row + columns.
    header_row_idx = None
    headers: dict[str, int] = {}
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        row_lower = [str(c).strip().lower() if c is not None else "" for c in row]
        if "entry id" in row_lower:
            header_row_idx = r_idx
            for c_idx, val in enumerate(row_lower):
                headers[val] = c_idx
            break
    if header_row_idx is None:
        raise HTTPException(status_code=400, detail="Could not find 'Entry ID' header row")

    def _col(name: str) -> Optional[int]:
        for k, v in headers.items():
            if name in k:
                return v
        return None

    col_id = _col("entry id")
    col_listed = _col("mark listed")
    if col_id is None or col_listed is None:
        raise HTTPException(status_code=400, detail="Required columns 'Entry ID' and 'Mark Listed (Yes/No)' not found")

    now = utcnow()
    to_listed: list = []
    to_pending: list = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        raw_id = row[col_id] if col_id < len(row) else None
        if not raw_id or str(raw_id).strip().lower() in ("none", "entry id", ""):
            continue
        try:
            oid = ObjectId(str(raw_id).strip())
        except Exception:
            continue

        val = str(row[col_listed]).strip().lower() if col_listed < len(row) and row[col_listed] is not None else ""
        if val in _TRUE_TOKENS:
            to_listed.append(oid)
        elif val in _FALSE_TOKENS:
            to_pending.append(oid)

    def _apply():
        col = database.get_collection(QUEUE_COLLECTION)
        n_listed = 0
        n_pending = 0
        if to_listed:
            r = col.update_many(
                {"_id": {"$in": to_listed}, "status": {"$ne": "listed"}},
                {"$set": {
                    "status": "listed",
                    "listed_by": listed_by,
                    "listed_by_name": listed_by_name,
                    "listed_at": now,
                    "updated_at": now,
                }},
            )
            n_listed = r.modified_count
        if to_pending:
            r = col.update_many(
                {"_id": {"$in": to_pending}, "status": {"$ne": "pending"}},
                {"$set": {
                    "status": "pending",
                    "listed_by": None,
                    "listed_by_name": None,
                    "listed_at": None,
                    "updated_at": now,
                }},
            )
            n_pending = r.modified_count
        return {"marked_listed": n_listed, "reverted_pending": n_pending}

    try:
        result = await asyncio.to_thread(_apply)
    except PyMongoError as e:
        logger.error(f"DB error in amazon_listing upload: {e}")
        raise HTTPException(status_code=500, detail=f"Database error: {e}")

    return JSONResponse(result)
