import os
import json
import asyncio
import logging
import requests
import io
from datetime import datetime

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..database import get_database

logger = logging.getLogger(__name__)

router = APIRouter()

ORGANIZATION_ID = os.getenv("ORGANIZATION_ID", "776755316")
BOOKS_URL = os.getenv("BOOKS_URL")
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
BOOKS_REFRESH_TOKEN = os.getenv("BOOKS_REFRESH_TOKEN")
ZOHO_BOOKS_BASE = "https://books.zoho.com/api/v3"


def _get_zoho_token() -> str:
    url = BOOKS_URL.format(
        clientId=CLIENT_ID,
        clientSecret=CLIENT_SECRET,
        grantType="refresh_token",
        books_refresh_token=BOOKS_REFRESH_TOKEN,
    )
    r = requests.post(url, timeout=30)
    r.raise_for_status()
    return r.json()["access_token"]


def _fetch_all_aging_items(token: str, to_date: str) -> list:
    headers = {"Authorization": f"Zoho-oauthtoken {token}"}
    select_columns = json.dumps([
        {"field": "item_name", "group": "item"},
        {"field": "intervals", "group": "report"},
    ])
    base_params = {
        "organization_id": ORGANIZATION_ID,
        "per_page": 500,
        "interval_type": "days",
        "number_of_columns": 4,
        "interval_range": 90,
        "select_columns": select_columns,
        "to_date": to_date,
        "sort_column": "item_id",
        "sort_order": "A",
        "response_option": 1,
    }

    all_items = []
    page = 1
    while True:
        params = {**base_params, "page": page}
        resp = requests.get(
            f"{ZOHO_BOOKS_BASE}/reports/inventoryagingsummary",
            headers=headers,
            params=params,
            timeout=60,
        )
        resp.raise_for_status()
        data = resp.json()

        if data.get("code", 0) != 0:
            raise Exception(f"Zoho API error: {data.get('message')}")

        items = data.get("inventory_aging_summary", [])
        all_items.extend(items)
        logger.info(f"Fetched page {page}, {len(items)} items (total so far: {len(all_items)})")

        page_ctx = data.get("page_context", {})
        if not page_ctx.get("has_more_page", False):
            break
        page += 1

    return all_items


def _extract_interval(intervals: list, label: str) -> tuple[float, float]:
    """Return (stock_on_hand, asset_value) for the given interval label."""
    qty = 0.0
    val = 0.0
    for entry in intervals:
        iv = entry.get("interval", "").strip()
        if iv == label:
            qty = entry.get("stock_on_hand", 0.0) or 0.0
        elif iv == f"Asset Value ({label})":
            val = entry.get("asset_value", 0.0) or 0.0
    return qty, val


def _apply_header_style(ws, row: int, col_count: int, fill_hex: str):
    fill = PatternFill("solid", fgColor=fill_hex)
    bold_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = bold_font
        cell.alignment = center
        cell.border = border


def _apply_row_style(ws, row: int, col_count: int, alt: bool):
    fill = PatternFill("solid", fgColor="F5F5F5" if alt else "FFFFFF")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _build_item_sheet(ws, interval_label: str, rows: list, fill_hex: str):
    headers = [
        "Item ID", "Item Name", "Brand",
        interval_label, f"Asset Value ({interval_label})",
        "MRP", "Total MRP Value", "Collection Value",
    ]
    ws.append(headers)
    _apply_header_style(ws, 1, len(headers), fill_hex)
    ws.row_dimensions[1].height = 20

    for i, row in enumerate(rows, start=2):
        qty = row["qty"]
        asset_val = row["asset_value"]
        mrp = row["mrp"]

        ws.append([
            row["item_id"],
            row["item_name"],
            row["brand"],
            qty,
            asset_val,
            mrp if mrp else "",
            f"=D{i}*F{i}" if mrp else "",
            f"=G{i}/2" if mrp else "",
        ])
        _apply_row_style(ws, i, len(headers), i % 2 == 0)

    _auto_width(ws)


def _build_summary_sheet(ws, curr_rows: list, prev_rows: list, curr_label: str, prev_label: str, fill_hex: str):
    """Build a comparison summary sheet with all unique items from both periods."""
    curr_map = {r["item_name"]: r["qty"] for r in curr_rows}
    prev_map = {r["item_name"]: r["qty"] for r in prev_rows}

    all_names = sorted(set(curr_map) | set(prev_map))

    headers = ["Item Name", prev_label, curr_label, "Status", "Difference Qty", "Change %"]
    ws.append(headers)
    _apply_header_style(ws, 1, len(headers), fill_hex)
    ws.row_dimensions[1].height = 20

    pct_fmt = '0.00%'

    for i, name in enumerate(all_names, start=2):
        prev_qty = prev_map.get(name, 0)
        curr_qty = curr_map.get(name, 0)
        diff = curr_qty - prev_qty

        if prev_qty == 0:
            status = "New Added"
            change_pct = "New Added"
        elif curr_qty == 0:
            status = "Removed"
            change_pct = -1.0
        elif curr_qty > prev_qty:
            status = "Increased"
            change_pct = diff / prev_qty
        elif curr_qty < prev_qty:
            status = "Decreased"
            change_pct = diff / prev_qty
        else:
            status = "No Change"
            change_pct = 0.0

        ws.append([name, prev_qty, curr_qty, status, diff, change_pct if isinstance(change_pct, float) else change_pct])
        _apply_row_style(ws, i, len(headers), i % 2 == 0)

        # Format Change % column as percentage if it's a number
        pct_cell = ws.cell(row=i, column=6)
        if isinstance(change_pct, float):
            pct_cell.number_format = pct_fmt

        # Colour-code the Status cell
        status_cell = ws.cell(row=i, column=4)
        if status == "New Added":
            status_cell.fill = PatternFill("solid", fgColor="D5F5E3")
            status_cell.font = Font(color="1E8449")
        elif status == "Removed":
            status_cell.fill = PatternFill("solid", fgColor="FADBD8")
            status_cell.font = Font(color="C0392B")
        elif status == "Increased":
            status_cell.fill = PatternFill("solid", fgColor="D6EAF8")
            status_cell.font = Font(color="1A5276")
        elif status == "Decreased":
            status_cell.fill = PatternFill("solid", fgColor="FEF9E7")
            status_cell.font = Font(color="9A7D0A")

    _auto_width(ws)


def _aggregate_by_brand(rows: list) -> dict:
    """Sum qty and total_mrp (qty * mrp) per brand from item rows."""
    totals: dict[str, dict] = {}
    for r in rows:
        brand = r["brand"]
        if brand not in totals:
            totals[brand] = {"qty": 0.0, "total_mrp": 0.0}
        totals[brand]["qty"] += r["qty"]
        totals[brand]["total_mrp"] += r["qty"] * r["mrp"]
    return totals


def _write_brand_section(ws, title: str, fill_hex: str, stock_col_label: str,
                         brand_data: dict, start_row: int) -> int:
    """Write a titled brand table starting at start_row. Returns the next free row."""
    # Section title
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, color="FFFFFF", size=11)
    title_cell.fill = PatternFill("solid", fgColor=fill_hex)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[start_row].height = 18
    for col in range(2, 5):
        cell = ws.cell(row=start_row, column=col)
        cell.fill = PatternFill("solid", fgColor=fill_hex)
    start_row += 1

    # Headers
    headers = ["Brand", stock_col_label, "Total MRP Value", "Collection Value"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col, value=h)
    _apply_header_style(ws, start_row, len(headers), fill_hex)
    ws.row_dimensions[start_row].height = 18
    data_start = start_row + 1
    start_row += 1

    for brand, data in sorted(brand_data.items()):
        qty = round(data.get("stock", data.get("qty", 0)), 2)
        tmrp = round(data.get("total_mrp", 0), 2)
        ws.cell(row=start_row, column=1, value=brand)
        ws.cell(row=start_row, column=2, value=qty)
        ws.cell(row=start_row, column=3, value=tmrp)
        ws.cell(row=start_row, column=4, value=f"=C{start_row}/2")
        _apply_row_style(ws, start_row, len(headers), start_row % 2 == 0)
        start_row += 1

    total_row = start_row
    end_data_row = total_row - 1
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=2, value=f"=SUM(B{data_start}:B{end_data_row})")
    ws.cell(row=total_row, column=3, value=f"=SUM(C{data_start}:C{end_data_row})")
    ws.cell(row=total_row, column=4, value=f"=SUM(D{data_start}:D{end_data_row})")
    bold = Font(bold=True)
    for col in range(1, 5):
        ws.cell(row=total_row, column=col).font = bold
    start_row += 1

    return start_row


def _build_brand_sheet(ws, brand_totals: dict, stock_date: str,
                       slow_rows: list, dead_rows: list):
    # Stock date label in col F
    ws.cell(row=1, column=6, value="Zoho Stock Date")
    ws.cell(row=1, column=6).font = Font(bold=True)
    ws.cell(row=2, column=6, value=stock_date)
    ws.column_dimensions["F"].width = 20

    next_row = 1

    # Section 1 – Current Stock
    next_row = _write_brand_section(
        ws, "Current Stock", "1F5C99", "Zoho Stock",
        brand_totals, next_row,
    )
    next_row += 2  # gap

    # Section 2 – Slow Movers
    slow_by_brand = _aggregate_by_brand(slow_rows)
    next_row = _write_brand_section(
        ws, "Slow Movers (181–270 days)", "C0392B", "Qty (181–270 days)",
        slow_by_brand, next_row,
    )
    next_row += 2  # gap

    # Section 3 – Deadstock
    dead_by_brand = _aggregate_by_brand(dead_rows)
    _write_brand_section(
        ws, "Deadstock (>270 days)", "7D3C98", "Qty (>270 days)",
        dead_by_brand, next_row,
    )

    _auto_width(ws)


def _build_rows_from_items(items: list, product_map: dict) -> tuple[list, list]:
    """Extract slow_rows and dead_rows from Zoho aging items."""
    slow_rows = []
    dead_rows = []

    for item in items:
        item_id = item["item_id"]
        item_name = item["item_name"]
        intervals = item.get("intervals", [])
        prod = product_map.get(item_id, {})
        brand = (prod.get("brand") or "Unknown").strip() or "Unknown"
        mrp = prod.get("rate") or 0.0

        slow_qty, slow_val = _extract_interval(intervals, "181 - 270 days")
        dead_qty, dead_val = _extract_interval(intervals, "> 270 days")

        base = {"item_id": item_id, "item_name": item_name, "brand": brand, "mrp": mrp}

        if slow_qty > 0:
            slow_rows.append({**base, "qty": slow_qty, "asset_value": slow_val})

        if dead_qty > 0:
            dead_rows.append({**base, "qty": dead_qty, "asset_value": dead_val})

    slow_rows.sort(key=lambda r: (r["brand"], r["item_name"]))
    dead_rows.sort(key=lambda r: (r["brand"], r["item_name"]))
    return slow_rows, dead_rows


def _date_label(date_str: str) -> str:
    """Convert YYYY-MM-DD to a short label like '09 May 26'."""
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").strftime("%d %b %y")
    except Exception:
        return date_str


@router.get("/download")
async def download_inventory_aging(
    to_date: str = Query(..., description="Current report date (YYYY-MM-DD)"),
    prev_date: str = Query(..., description="Previous period date (YYYY-MM-DD)"),
    db=Depends(get_database),
):
    for label, val in [("to_date", to_date), ("prev_date", prev_date)]:
        try:
            datetime.strptime(val, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail=f"{label} must be YYYY-MM-DD")

    if prev_date >= to_date:
        raise HTTPException(status_code=400, detail="prev_date must be earlier than to_date")

    # Fetch Zoho token once, then fetch both periods in parallel
    token = await asyncio.to_thread(_get_zoho_token)
    curr_items, prev_items = await asyncio.gather(
        asyncio.to_thread(_fetch_all_aging_items, token, to_date),
        asyncio.to_thread(_fetch_all_aging_items, token, prev_date),
    )

    if not curr_items:
        raise HTTPException(status_code=404, detail="No inventory aging data returned for current date")

    # Batch-load products and brands in parallel
    all_item_ids = list({item["item_id"] for item in curr_items + prev_items})
    products_col = db["products"]
    brands_col = db["brands"]
    stock_col = db["zoho_warehouse_stock"]

    def _load_products():
        return list(products_col.find(
            {"item_id": {"$in": all_item_ids}},
            {"item_id": 1, "brand": 1, "rate": 1},
        ))

    def _load_brands():
        return list(brands_col.find({}, {"name": 1}))

    def _find_stock_date():
        rec = stock_col.find_one(
            {},
            sort=[("date", -1)],
            projection={"date": 1},
        )
        return rec["date"] if rec else None

    product_docs, brand_docs, stock_date = await asyncio.gather(
        asyncio.to_thread(_load_products),
        asyncio.to_thread(_load_brands),
        asyncio.to_thread(_find_stock_date),
    )

    product_map = {p["item_id"]: p for p in product_docs}

    # Build brand_totals from zoho_warehouse_stock
    brand_names = [b["name"] for b in brand_docs if b.get("name")]
    brand_totals: dict[str, dict] = {
        name: {"stock": 0.0, "total_mrp": 0.0, "collection_value": 0.0}
        for name in brand_names
    }
    stock_date_str = stock_date.strftime("%Y-%m-%d") if stock_date else to_date

    if stock_date:
        def _load_stock_records():
            return list(stock_col.find(
                {"date": stock_date},
                {"zoho_item_id": 1, "warehouses": 1},
            ))

        def _load_all_products_for_brands():
            return list(products_col.find(
                {},
                {"item_id": 1, "brand": 1, "rate": 1},
            ))

        stock_records, all_product_docs = await asyncio.gather(
            asyncio.to_thread(_load_stock_records),
            asyncio.to_thread(_load_all_products_for_brands),
        )

        all_product_map = {p["item_id"]: p for p in all_product_docs}

        for rec in stock_records:
            zoho_id = rec.get("zoho_item_id")
            warehouses = rec.get("warehouses") or {}
            total_stock = sum(v for v in warehouses.values() if isinstance(v, (int, float)))
            if total_stock <= 0:
                continue
            prod = all_product_map.get(zoho_id, {})
            brand = (prod.get("brand") or "").strip()
            mrp = prod.get("rate") or 0.0
            if brand not in brand_totals:
                continue
            total_mrp_val = total_stock * mrp
            brand_totals[brand]["stock"] += total_stock
            brand_totals[brand]["total_mrp"] += total_mrp_val
            brand_totals[brand]["collection_value"] += total_mrp_val / 2

    # Build slow/dead rows for both periods
    curr_slow, curr_dead = _build_rows_from_items(curr_items, product_map)
    prev_slow, prev_dead = _build_rows_from_items(prev_items, product_map)

    curr_label = _date_label(to_date)
    prev_label = _date_label(prev_date)

    # Build XLSX
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Slow Movers sheets
    ws_slow_summary = wb.create_sheet("Summary - Slow Movers")
    _build_summary_sheet(ws_slow_summary, curr_slow, prev_slow, curr_label, prev_label, "C0392B")

    ws_slow_curr = wb.create_sheet(f"Slow Movers ({curr_label})")
    _build_item_sheet(ws_slow_curr, "181 - 270 days", curr_slow, "C0392B")

    ws_slow_prev = wb.create_sheet(f"Slow Movers ({prev_label})")
    _build_item_sheet(ws_slow_prev, "181 - 270 days", prev_slow, "E74C3C")

    # Deadstock sheets
    ws_dead_summary = wb.create_sheet("Summary - Deadstock")
    _build_summary_sheet(ws_dead_summary, curr_dead, prev_dead, curr_label, prev_label, "7D3C98")

    ws_dead_curr = wb.create_sheet(f"Deadstock ({curr_label})")
    _build_item_sheet(ws_dead_curr, "> 270 days", curr_dead, "7D3C98")

    ws_dead_prev = wb.create_sheet(f"Deadstock ({prev_label})")
    _build_item_sheet(ws_dead_prev, "> 270 days", prev_dead, "9B59B6")

    # Brand sheet — three sections: current stock, slow movers, deadstock
    ws_brand = wb.create_sheet("Brand wise collection value")
    _build_brand_sheet(ws_brand, brand_totals, stock_date_str, curr_slow, curr_dead)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Inventory_Aging_{to_date}_vs_{prev_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
