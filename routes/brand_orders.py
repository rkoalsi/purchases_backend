from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form, Request, Query
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel
from datetime import datetime, timedelta, timezone
from bson import ObjectId
from ..database import get_database, serialize_mongo_document
import asyncio
import io
import os
import re
import zipfile
import boto3
from botocore.config import Config as BotocoreConfig
from typing import Optional
import logging
from jose import jwt, JWTError
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests

logger = logging.getLogger(__name__)
router = APIRouter()

COLLECTION = "brand_orders"
PO_COLLECTION = "purchase_orders"
VENDORS_COLLECTION = "vendors"
CATEGORIES_COLLECTION = "brand_order_upload_categories"
NOTIFICATIONS_COLLECTION = "notifications"
S3_BUCKET = os.getenv("S3_BUCKET", "pupscribe-purchases")
AWS_REGION = os.getenv("AWS_REGION", "ap-south-1")
DEFAULT_CATEGORIES = ["PI", "CL", "Bill of Lading", "Bill of Entry", "Insurance"]


class AddCategoryRequest(BaseModel):
    name: str


class CreateFolderRequest(BaseModel):
    name: str
    parent_path: Optional[str] = ""


class RenameFolderRequest(BaseModel):
    name: str


def _sanitize_folder_name(name: str) -> str:
    """Convert folder name to a safe path segment (preserves readability)."""
    name = name.strip()
    safe = re.sub(r"[^\w\s.\-]", "", name)
    safe = re.sub(r"\s+", "_", safe).strip("_")
    return safe or "folder"


def _slugify(text: str) -> str:
    text = text.lower().strip()
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[\s_-]+", "_", text)
    return text


def _s3_client():
    return boto3.client(
        "s3",
        region_name=AWS_REGION,
        endpoint_url=f"https://s3.{AWS_REGION}.amazonaws.com",
        config=BotocoreConfig(signature_version="s3v4"),
    )


def _upload_to_s3(file_bytes: bytes, s3_key: str, content_type: str) -> None:
    boto3.client("s3", region_name=AWS_REGION).put_object(
        Bucket=S3_BUCKET,
        Key=s3_key,
        Body=file_bytes,
        ContentType=content_type,
    )


def _presign_s3(s3_key: str, expires: int = 3600) -> str:
    return _s3_client().generate_presigned_url(
        "get_object",
        Params={"Bucket": S3_BUCKET, "Key": s3_key},
        ExpiresIn=expires,
    )


def _delete_from_s3(s3_key: str) -> None:
    boto3.client("s3", region_name=AWS_REGION).delete_object(
        Bucket=S3_BUCKET, Key=s3_key
    )


# ─── orders CRUD ──────────────────────────────────────────────────────────────

def _validate_optional_dates(fields_map: dict) -> dict:
    out = {}
    for field, val in fields_map.items():
        if val:
            try:
                datetime.strptime(val, "%Y-%m-%d")
                out[field] = val
            except ValueError:
                raise HTTPException(status_code=400, detail=f"{field} must be YYYY-MM-DD")
        else:
            out[field] = None
    return out


@router.post("/")
async def create_order(
    brand: str = Form(...),
    vendor_id: Optional[str] = Form(None),
    order_date: Optional[str] = Form(None),
    shipment_eta: Optional[str] = Form(None),
    purchaseorder_number: Optional[str] = Form(None),
    initiation_date: Optional[str] = Form(None),
    proforma_date: Optional[str] = Form(None),
    ready_date: Optional[str] = Form(None),
    etd_date: Optional[str] = Form(None),
    eta_port_date: Optional[str] = Form(None),
    duty_payment_date: Optional[str] = Form(None),
    inward_date: Optional[str] = Form(None),
    po_due_date: Optional[str] = Form(None),
    custom_duty: Optional[float] = Form(None),
    custom_duty_due_date: Optional[str] = Form(None),
    shipping_charges: Optional[float] = Form(None),
    shipping_charges_due_date: Optional[str] = Form(None),
    db=Depends(get_database),
):
    if order_date:
        try:
            datetime.strptime(order_date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="order_date must be YYYY-MM-DD")
    if shipment_eta:
        try:
            datetime.strptime(shipment_eta, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="shipment_eta must be YYYY-MM-DD")

    extra_dates = _validate_optional_dates({
        "initiation_date": initiation_date,
        "proforma_date": proforma_date,
        "ready_date": ready_date,
        "etd_date": etd_date,
        "eta_port_date": eta_port_date,
        "duty_payment_date": duty_payment_date,
        "inward_date": inward_date,
        "po_due_date": po_due_date,
        "custom_duty_due_date": custom_duty_due_date,
        "shipping_charges_due_date": shipping_charges_due_date,
    })

    def _insert():
        brand_name = brand.strip()
        vid = (vendor_id or "").strip() or None
        po_num = (purchaseorder_number or "").strip() or None

        # Block duplicate: same brand + PO number
        if po_num and db[COLLECTION].find_one({"brand": brand_name, "purchaseorder_number": po_num}):
            raise ValueError(f"A brand order for {brand_name} with PO {po_num} already exists")

        # Derive next order number from max existing name for this brand+vendor
        existing = list(db[COLLECTION].find(
            {"brand": brand_name, "vendor_id": vid},
            {"name": 1},
        ))
        max_num = 0
        import re as _re
        for ex in existing:
            m = _re.search(r"#(\d+)", ex.get("name", ""))
            if m:
                max_num = max(max_num, int(m.group(1)))
        name = f"Order #{max_num + 1}"

        now = datetime.now()
        doc = {
            "brand": brand_name,
            "vendor_id": vid,
            "name": name,
            "order_date": order_date or None,
            "shipment_eta": shipment_eta or None,
            "purchaseorder_number": po_num,
            "documents": [],
            "created_at": now,
            "updated_at": now,
            **extra_dates,
            "custom_duty": custom_duty,
            "shipping_charges": shipping_charges,
        }
        result = db[COLLECTION].insert_one(doc)
        return str(result.inserted_id), doc

    try:
        order_id, doc = await asyncio.to_thread(_insert)
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))

    async def _notify_brand_order():
        def _fan_out():
            recipients = list(db["purchase_users"].find(
                {"department": {"$in": ["purchases", "design"]}, "status": "active"},
                {"_id": 1},
            ))
            if not recipients:
                return
            now = datetime.utcnow()
            brand_name = doc["brand"]
            order_name = doc["name"]
            po_num = doc.get("purchaseorder_number")
            snippet = f"created {brand_name} — {order_name}"
            if po_num:
                snippet += f" (PO {po_num})"
            notif_docs = [
                {
                    "user_id": str(r["_id"]),
                    "source": "brand_order",
                    "order_id": order_id,
                    "brand": brand_name,
                    "order_name": order_name,
                    "type": "brand_order_created",
                    "actor_name": "System",
                    "snippet": snippet,
                    "read": False,
                    "created_at": now,
                }
                for r in recipients
            ]
            db[NOTIFICATIONS_COLLECTION].insert_many(notif_docs)
        try:
            await asyncio.to_thread(_fan_out)
        except Exception as _e:
            logger.warning(f"Brand order notification fan-out failed: {_e}")

    asyncio.create_task(_notify_brand_order())

    return JSONResponse(status_code=201, content={"_id": order_id, **serialize_mongo_document(doc)})


@router.get("/")
async def list_orders(brand: Optional[str] = None, db=Depends(get_database)):
    def _fetch():
        query = {"brand": brand} if brand else {}
        pipeline = [
            {"$match": query},
            {"$addFields": {"doc_count": {"$size": {"$ifNull": ["$documents", []]}}}},
            {"$project": {"documents": 0}},
            {"$lookup": {
                "from": PO_COLLECTION,
                "localField": "purchaseorder_number",
                "foreignField": "purchaseorder_number",
                "as": "_po",
            }},
            {"$lookup": {
                "from": VENDORS_COLLECTION,
                "localField": "vendor_id",
                "foreignField": "contact_id",
                "as": "_vendor",
            }},
            {"$addFields": {
                "po_status": {
                    "$let": {
                        "vars": {
                            "fmt": {"$arrayElemAt": ["$_po.order_status_formatted", 0]},
                            "raw": {"$arrayElemAt": ["$_po.order_status", 0]},
                        },
                        "in": {
                            "$cond": {
                                "if": {"$gt": [{"$strLenCP": {"$ifNull": ["$$fmt", ""]}}, 0]},
                                "then": "$$fmt",
                                "else": {
                                    "$cond": {
                                        "if": {"$gt": [{"$strLenCP": {"$ifNull": ["$$raw", ""]}}, 0]},
                                        "then": {"$concat": [
                                            {"$toUpper": {"$substrCP": ["$$raw", 0, 1]}},
                                            {"$substrCP": ["$$raw", 1, {"$subtract": [{"$strLenCP": "$$raw"}, 1]}]},
                                        ]},
                                        "else": None,
                                    }
                                },
                            }
                        },
                    }
                },
                "po_sub_total": {"$arrayElemAt": ["$_po.total", 0]},
                "po_currency_code": {"$arrayElemAt": ["$_po.currency_code", 0]},
                "vendor_name": {"$arrayElemAt": ["$_vendor.contact_name", 0]},
            }},
            {"$project": {"_po": 0, "_vendor": 0}},
            {"$sort": {"created_at": -1, "_id": -1}},
        ]
        return list(db[COLLECTION].aggregate(pipeline))

    orders = await asyncio.to_thread(_fetch)
    return serialize_mongo_document(orders)


@router.get("/lead-time-report")
async def download_lead_time_report(
    request: Request,
    brand: Optional[str] = None,
    db=Depends(get_database),
):
    """Generate Lead Time report Excel for all orders (admin only)."""
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = auth_header.split(" ", 1)[1]
    try:
        secret = os.getenv("SECRET_KEY") or ""
        algo = os.getenv("ALGORITHM") or "HS256"
        payload = jwt.decode(token, secret, algorithms=[algo], options={"verify_aud": False, "verify_exp": False})
        email = payload.get("sub")
    except Exception as exc:
        logger.warning("lead-time-report token decode failed: %s", exc)
        raise HTTPException(status_code=401, detail="Invalid token")

    def _fetch(email, brand_filter):
        user = db["purchase_users"].find_one({"email": email}, {"role": 1})
        if not user or user.get("role") != "admin":
            return None
        query = {"brand": brand_filter} if brand_filter else {}
        return list(db[COLLECTION].find(query, {
            "brand": 1, "name": 1, "purchaseorder_number": 1,
            "initiation_date": 1, "proforma_date": 1, "ready_date": 1,
            "etd_date": 1, "eta_port_date": 1, "duty_payment_date": 1,
            "inward_date": 1,
        }).sort([("brand", 1), ("name", 1)]))

    orders = await asyncio.to_thread(_fetch, email, brand)
    if orders is None:
        raise HTTPException(status_code=403, detail="Admin access required")

    buf = await asyncio.to_thread(_build_lead_time_excel, orders)
    filename = f"Lead_Time_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/payment-report")
async def download_payment_report(
    request: Request,
    brand: Optional[str] = None,
    db=Depends(get_database),
):
    """Generate Payment Dates & Details report (admin only)."""
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = auth_header.split(" ", 1)[1]
    try:
        secret = os.getenv("SECRET_KEY") or ""
        algo = os.getenv("ALGORITHM") or "HS256"
        payload = jwt.decode(token, secret, algorithms=[algo], options={"verify_aud": False, "verify_exp": False})
        email = payload.get("sub")
    except Exception as exc:
        logger.warning("payment-report token decode failed: %s", exc)
        raise HTTPException(status_code=401, detail="Invalid token")

    def _fetch_payment(email, brand_filter):
        user = db["purchase_users"].find_one({"email": email}, {"role": 1})
        if not user or user.get("role") != "admin":
            return None
        query = {"brand": brand_filter} if brand_filter else {}
        pipeline = [
            {"$match": query},
            {"$lookup": {
                "from": PO_COLLECTION,
                "localField": "purchaseorder_number",
                "foreignField": "purchaseorder_number",
                "as": "_po",
            }},
            {"$addFields": {
                "po_sub_total": {"$arrayElemAt": ["$_po.total", 0]},
                "po_currency_code": {"$arrayElemAt": ["$_po.currency_code", 0]},
                "po_status": {
                    "$let": {
                        "vars": {
                            "fmt": {"$arrayElemAt": ["$_po.order_status_formatted", 0]},
                            "raw": {"$arrayElemAt": ["$_po.order_status", 0]},
                        },
                        "in": {
                            "$cond": {
                                "if": {"$gt": [{"$strLenCP": {"$ifNull": ["$$fmt", ""]}}, 0]},
                                "then": "$$fmt",
                                "else": {
                                    "$cond": {
                                        "if": {"$gt": [{"$strLenCP": {"$ifNull": ["$$raw", ""]}}, 0]},
                                        "then": {
                                            "$concat": [
                                                {"$toUpper": {"$substrCP": ["$$raw", 0, 1]}},
                                                {"$substrCP": ["$$raw", 1, {"$subtract": [{"$strLenCP": "$$raw"}, 1]}]},
                                            ]
                                        },
                                        "else": "",
                                    }
                                },
                            }
                        },
                    }
                },
            }},
            {"$project": {
                "brand": 1, "name": 1, "purchaseorder_number": 1,
                "po_sub_total": 1, "po_currency_code": 1, "po_status": 1,
                "po_due_date": 1, "custom_duty": 1, "custom_duty_due_date": 1,
                "shipping_charges": 1, "shipping_charges_due_date": 1,
            }},
            {"$sort": {"brand": 1, "name": 1}},
        ]
        return list(db[COLLECTION].aggregate(pipeline))

    orders = await asyncio.to_thread(_fetch_payment, email, brand)
    if orders is None:
        raise HTTPException(status_code=403, detail="Admin access required")

    currencies = {o.get("po_currency_code") for o in orders if o.get("po_currency_code")}
    fx_rates = await asyncio.to_thread(_fetch_fx_rates, currencies)

    buf = await asyncio.to_thread(_build_payment_report_excel, orders, fx_rates)
    filename = f"Payment_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── upload categories ────────────────────────────────────────────────────────

@router.get("/categories")
async def get_categories(db=Depends(get_database)):
    def _fetch():
        cats = list(db[CATEGORIES_COLLECTION].find({}, {"name": 1, "_id": 0}))
        if not cats:
            db[CATEGORIES_COLLECTION].insert_many([{"name": n} for n in DEFAULT_CATEGORIES])
            return DEFAULT_CATEGORIES
        return [c["name"] for c in cats]
    return await asyncio.to_thread(_fetch)


@router.post("/categories")
async def add_category(request: AddCategoryRequest, db=Depends(get_database)):
    name = request.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Category name required")
    def _add():
        if not db[CATEGORIES_COLLECTION].find_one({"name": name}):
            db[CATEGORIES_COLLECTION].insert_one({"name": name})
    await asyncio.to_thread(_add)
    return {"name": name}


@router.delete("/categories/{name}")
async def delete_category(name: str, db=Depends(get_database)):
    def _delete():
        result = db[CATEGORIES_COLLECTION].delete_one({"name": name})
        return result.deleted_count
    deleted = await asyncio.to_thread(_delete)
    if not deleted:
        raise HTTPException(status_code=404, detail="Category not found")
    return {"name": name, "deleted": True}


@router.patch("/categories/{name}")
async def rename_category(name: str, request: Request, db=Depends(get_database)):
    body = await request.json()
    new_name = (body.get("new_name") or "").strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="new_name required")
    def _rename():
        if db[CATEGORIES_COLLECTION].find_one({"name": new_name}):
            raise HTTPException(status_code=409, detail="Category already exists")
        result = db[CATEGORIES_COLLECTION].update_one({"name": name}, {"$set": {"name": new_name}})
        if not result.matched_count:
            raise HTTPException(status_code=404, detail="Category not found")
        # Update all documents in orders that used the old category name
        db[COLLECTION].update_many(
            {"documents.category": name},
            {"$set": {"documents.$[elem].category": new_name}},
            array_filters=[{"elem.category": name}]
        )
    await asyncio.to_thread(_rename)
    return {"old_name": name, "new_name": new_name}


# ─── folders ──────────────────────────────────────────────────────────────────

@router.get("/{order_id}/folders")
async def list_folders(order_id: str, db=Depends(get_database)):
    def _fetch():
        doc = db[COLLECTION].find_one({"_id": ObjectId(order_id)}, {"folders": 1})
        if not doc:
            raise ValueError("Order not found")
        return doc.get("folders", [])
    try:
        folders = await asyncio.to_thread(_fetch)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return serialize_mongo_document(folders)


@router.post("/{order_id}/folders")
async def create_folder(order_id: str, body: CreateFolderRequest, db=Depends(get_database)):
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Folder name required")
    parent = (body.parent_path or "").strip().strip("/")
    safe_name = _sanitize_folder_name(name)
    path = f"{parent}/{safe_name}" if parent else safe_name

    def _create():
        doc = db[COLLECTION].find_one({"_id": ObjectId(order_id)}, {"folders": 1})
        if not doc:
            raise ValueError("Order not found")
        if any(f.get("path") == path for f in doc.get("folders", [])):
            raise ValueError(f"A folder named '{name}' already exists here")
        folder_entry = {
            "folder_id": str(ObjectId()),
            "name": name,
            "path": path,
            "parent_path": parent,
            "created_at": datetime.now(),
        }
        db[COLLECTION].update_one(
            {"_id": ObjectId(order_id)},
            {"$push": {"folders": folder_entry}, "$set": {"updated_at": datetime.now()}},
        )
        return folder_entry

    try:
        folder = await asyncio.to_thread(_create)
    except ValueError as e:
        code = 409 if "already exists" in str(e) else 404
        raise HTTPException(status_code=code, detail=str(e))
    return serialize_mongo_document(folder)


@router.patch("/{order_id}/folders/{folder_id}")
async def rename_folder(order_id: str, folder_id: str, body: RenameFolderRequest, db=Depends(get_database)):
    new_name = body.name.strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="Folder name required")

    def _rename():
        doc = db[COLLECTION].find_one({"_id": ObjectId(order_id)}, {"folders": 1, "documents": 1})
        if not doc:
            raise ValueError("Order not found")
        folders = doc.get("folders", [])
        folder = next((f for f in folders if f.get("folder_id") == folder_id), None)
        if not folder:
            raise ValueError("Folder not found")

        old_path = folder["path"]
        parent = folder.get("parent_path", "")
        safe_name = _sanitize_folder_name(new_name)
        new_path = f"{parent}/{safe_name}" if parent else safe_name

        if new_path == old_path:
            return {**folder, "name": new_name, "path": new_path}

        # Rename this folder record
        db[COLLECTION].update_one(
            {"_id": ObjectId(order_id), "folders.folder_id": folder_id},
            {"$set": {"folders.$.name": new_name, "folders.$.path": new_path, "updated_at": datetime.now()}},
        )
        # Cascade rename into child folder paths
        for f in folders:
            f_path = f.get("path", "")
            if f_path.startswith(old_path + "/"):
                new_child_path = new_path + f_path[len(old_path):]
                new_parent = f.get("parent_path", "")
                if new_parent.startswith(old_path):
                    new_parent = new_path + new_parent[len(old_path):]
                db[COLLECTION].update_one(
                    {"_id": ObjectId(order_id), "folders.folder_id": f["folder_id"]},
                    {"$set": {"folders.$.path": new_child_path, "folders.$.parent_path": new_parent}},
                )
        # Cascade rename into document folder fields
        for d in doc.get("documents", []):
            d_folder = d.get("folder", "")
            if d_folder == old_path:
                db[COLLECTION].update_one(
                    {"_id": ObjectId(order_id), "documents.doc_id": d["doc_id"]},
                    {"$set": {"documents.$.folder": new_path}},
                )
            elif d_folder.startswith(old_path + "/"):
                db[COLLECTION].update_one(
                    {"_id": ObjectId(order_id), "documents.doc_id": d["doc_id"]},
                    {"$set": {"documents.$.folder": new_path + d_folder[len(old_path):]}},
                )
        return {**folder, "name": new_name, "path": new_path}

    try:
        result = await asyncio.to_thread(_rename)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return serialize_mongo_document(result)


@router.delete("/{order_id}/folders/{folder_id}")
async def delete_folder(order_id: str, folder_id: str, db=Depends(get_database)):
    """Delete a folder (and all subfolders); documents inside are moved to the parent folder."""
    def _delete():
        doc = db[COLLECTION].find_one({"_id": ObjectId(order_id)}, {"folders": 1, "documents": 1})
        if not doc:
            raise ValueError("Order not found")
        folders = doc.get("folders", [])
        folder = next((f for f in folders if f.get("folder_id") == folder_id), None)
        if not folder:
            raise ValueError("Folder not found")

        path = folder["path"]
        parent_path = folder.get("parent_path", "")

        # Move documents in this folder (and sub-folders) to parent
        for d in doc.get("documents", []):
            d_folder = d.get("folder", "")
            if d_folder == path or d_folder.startswith(path + "/"):
                db[COLLECTION].update_one(
                    {"_id": ObjectId(order_id), "documents.doc_id": d["doc_id"]},
                    {"$set": {"documents.$.folder": parent_path}},
                )

        # Collect folder_ids to remove (this folder + all descendants)
        ids_to_remove = [folder_id] + [
            f["folder_id"] for f in folders if f.get("path", "").startswith(path + "/")
        ]
        db[COLLECTION].update_one(
            {"_id": ObjectId(order_id)},
            {
                "$pull": {"folders": {"folder_id": {"$in": ids_to_remove}}},
                "$set": {"updated_at": datetime.now()},
            },
        )
        return True

    try:
        await asyncio.to_thread(_delete)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return {"folder_id": folder_id, "deleted": True}


@router.get("/{order_id}")
async def get_order(order_id: str, db=Depends(get_database)):
    def _fetch():
        pipeline = [
            {"$match": {"_id": ObjectId(order_id)}},
            {"$lookup": {
                "from": PO_COLLECTION,
                "localField": "purchaseorder_number",
                "foreignField": "purchaseorder_number",
                "as": "_po",
            }},
            {"$lookup": {
                "from": VENDORS_COLLECTION,
                "localField": "vendor_id",
                "foreignField": "contact_id",
                "as": "_vendor",
            }},
            {"$addFields": {
                "po_status": {
                    "$let": {
                        "vars": {
                            "fmt": {"$arrayElemAt": ["$_po.order_status_formatted", 0]},
                            "raw": {"$arrayElemAt": ["$_po.order_status", 0]},
                        },
                        "in": {
                            "$cond": {
                                "if": {"$gt": [{"$strLenCP": {"$ifNull": ["$$fmt", ""]}}, 0]},
                                "then": "$$fmt",
                                "else": {
                                    "$cond": {
                                        "if": {"$gt": [{"$strLenCP": {"$ifNull": ["$$raw", ""]}}, 0]},
                                        "then": {"$concat": [
                                            {"$toUpper": {"$substrCP": ["$$raw", 0, 1]}},
                                            {"$substrCP": ["$$raw", 1, {"$subtract": [{"$strLenCP": "$$raw"}, 1]}]},
                                        ]},
                                        "else": None,
                                    }
                                },
                            }
                        },
                    }
                },
                "po_sub_total": {"$arrayElemAt": ["$_po.total", 0]},
                "po_currency_code": {"$arrayElemAt": ["$_po.currency_code", 0]},
                "vendor_name": {"$arrayElemAt": ["$_vendor.contact_name", 0]},
            }},
            {"$project": {"_po": 0, "_vendor": 0}},
        ]
        docs = list(db[COLLECTION].aggregate(pipeline))
        return docs[0] if docs else None

    doc = await asyncio.to_thread(_fetch)
    if not doc:
        raise HTTPException(status_code=404, detail="Order not found")
    return serialize_mongo_document(doc)


@router.patch("/{order_id}")
async def update_order(
    order_id: str,
    brand: Optional[str] = Form(None),
    name: Optional[str] = Form(None),
    vendor_id: Optional[str] = Form(None),
    order_date: Optional[str] = Form(None),
    shipment_eta: Optional[str] = Form(None),
    purchaseorder_number: Optional[str] = Form(None),
    initiation_date: Optional[str] = Form(None),
    proforma_date: Optional[str] = Form(None),
    ready_date: Optional[str] = Form(None),
    etd_date: Optional[str] = Form(None),
    eta_port_date: Optional[str] = Form(None),
    duty_payment_date: Optional[str] = Form(None),
    inward_date: Optional[str] = Form(None),
    po_due_date: Optional[str] = Form(None),
    custom_duty: Optional[float] = Form(None),
    custom_duty_due_date: Optional[str] = Form(None),
    shipping_charges: Optional[float] = Form(None),
    shipping_charges_due_date: Optional[str] = Form(None),
    db=Depends(get_database),
):
    fields: dict = {"updated_at": datetime.now()}
    if brand is not None:
        fields["brand"] = brand.strip()
    if name is not None:
        fields["name"] = name.strip()
    if order_date is not None:
        if order_date:
            try:
                datetime.strptime(order_date, "%Y-%m-%d")
            except ValueError:
                raise HTTPException(status_code=400, detail="order_date must be YYYY-MM-DD")
        fields["order_date"] = order_date or None
    if shipment_eta is not None:
        if shipment_eta:
            try:
                datetime.strptime(shipment_eta, "%Y-%m-%d")
            except ValueError:
                raise HTTPException(status_code=400, detail="shipment_eta must be YYYY-MM-DD")
        fields["shipment_eta"] = shipment_eta or None
    if purchaseorder_number is not None:
        po_num = purchaseorder_number.strip()
        fields["purchaseorder_number"] = po_num or None
    if vendor_id is not None:
        fields["vendor_id"] = vendor_id.strip() or None
    for field, val in [
        ("initiation_date", initiation_date),
        ("proforma_date", proforma_date),
        ("ready_date", ready_date),
        ("etd_date", etd_date),
        ("eta_port_date", eta_port_date),
        ("duty_payment_date", duty_payment_date),
        ("inward_date", inward_date),
        ("po_due_date", po_due_date),
        ("custom_duty_due_date", custom_duty_due_date),
        ("shipping_charges_due_date", shipping_charges_due_date),
    ]:
        if val is not None:
            if val:
                try:
                    datetime.strptime(val, "%Y-%m-%d")
                except ValueError:
                    raise HTTPException(status_code=400, detail=f"{field} must be YYYY-MM-DD")
            fields[field] = val or None

    if custom_duty is not None:
        fields["custom_duty"] = custom_duty
    if shipping_charges is not None:
        fields["shipping_charges"] = shipping_charges

    def _update():
        order_doc = db[COLLECTION].find_one(
            {"_id": ObjectId(order_id)},
            {"brand": 1, "name": 1, "ready_date": 1, "etd_date": 1},
        )
        if not order_doc:
            return None, None, {}
        prev_dates = {
            "ready_date": order_doc.get("ready_date"),
            "etd_date": order_doc.get("etd_date"),
        }
        result = db[COLLECTION].update_one(
            {"_id": ObjectId(order_id)}, {"$set": fields}
        )
        return result.matched_count, order_doc, prev_dates

    matched, order_doc, prev_dates = await asyncio.to_thread(_update)
    if not matched:
        raise HTTPException(status_code=404, detail="Order not found")

    # Notify design Slack channel only when a date transitions from unset → set
    _DATE_NOTIFICATIONS = {
        "ready_date": (":package: Order Ready Date Set", "The order is ready for shipment."),
        "etd_date": (":ship: ETD Date Set", "The estimated departure date has been confirmed."),
    }
    if order_doc:
        def _notify_dates(brand: str, order_name: str, notifications: list):
            urls = [u for u in [os.getenv("SLACK_URL_DESIGN"), os.getenv("SLACK_URL_PURCHASE")] if u]
            if not urls:
                return
            for title, subtitle, date_val in notifications:
                blocks = [
                    {
                        "type": "header",
                        "text": {"type": "plain_text", "text": title, "emoji": True},
                    },
                    {
                        "type": "section",
                        "fields": [
                            {"type": "mrkdwn", "text": f"*Brand:*\n{brand}"},
                            {"type": "mrkdwn", "text": f"*Order:*\n{order_name}"},
                            {"type": "mrkdwn", "text": f"*Date:*\n{date_val}"},
                        ],
                    },
                    {
                        "type": "context",
                        "elements": [{"type": "mrkdwn", "text": f"{subtitle}  ·  {datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=5, minutes=30))).strftime('%d %b %Y, %H:%M')} IST"}],
                    },
                ]
                for url in urls:
                    try:
                        requests.post(url, json={"blocks": blocks}, timeout=10)
                    except Exception as exc:
                        logger.warning("Brand order Slack notification failed: %s", exc)

        notifications_to_send = []
        for field, (title, subtitle) in _DATE_NOTIFICATIONS.items():
            new_val = fields.get(field)
            old_val = prev_dates.get(field)
            if new_val and not old_val:
                notifications_to_send.append((title, subtitle, new_val))

        if notifications_to_send:
            await asyncio.to_thread(
                _notify_dates,
                order_doc.get("brand", ""),
                order_doc.get("name", ""),
                notifications_to_send,
            )

    return {"_id": order_id, **{k: v for k, v in fields.items() if k != "updated_at"}}


@router.delete("/{order_id}")
async def delete_order(order_id: str, db=Depends(get_database)):
    def _delete():
        doc = db[COLLECTION].find_one({"_id": ObjectId(order_id)}, {"documents": 1})
        if not doc:
            return None
        s3 = boto3.client("s3", region_name=AWS_REGION)
        for d in doc.get("documents", []):
            if d.get("s3_key"):
                try:
                    s3.delete_object(Bucket=S3_BUCKET, Key=d["s3_key"])
                except Exception:
                    pass
        db[COLLECTION].delete_one({"_id": ObjectId(order_id)})
        return True

    result = await asyncio.to_thread(_delete)
    if result is None:
        raise HTTPException(status_code=404, detail="Order not found")
    return {"_id": order_id, "deleted": True}


# ─── backfill vendor IDs ──────────────────────────────────────────────────────

@router.post("/backfill-vendor-ids")
async def backfill_vendor_ids(db=Depends(get_database)):
    """Assign vendor_id to brand orders missing one, by matching PO currency to the brand's vendors."""
    def _run():
        updated = 0
        orders = list(db[COLLECTION].find(
            {"vendor_id": {"$in": [None, ""]}, "purchaseorder_number": {"$ne": None}},
            {"_id": 1, "brand": 1, "purchaseorder_number": 1},
        ))

        for order in orders:
            po_num = order.get("purchaseorder_number")
            if not po_num:
                continue

            po = db[PO_COLLECTION].find_one({"purchaseorder_number": po_num}, {"currency_code": 1})
            if not po:
                continue
            currency = po.get("currency_code")

            brand_doc = db["brands"].find_one({"name": order["brand"]}, {"vendor_id": 1, "vendor_ids": 1})
            if not brand_doc:
                continue

            vendor_ids = brand_doc.get("vendor_ids") or []
            if not vendor_ids and brand_doc.get("vendor_id"):
                vendor_ids = [brand_doc["vendor_id"]]
            if not vendor_ids:
                continue

            matched = None
            if currency:
                for vid in vendor_ids:
                    v = db[VENDORS_COLLECTION].find_one({"contact_id": vid}, {"contact_id": 1, "currency_code": 1})
                    if v and v.get("currency_code") == currency:
                        matched = v["contact_id"]
                        break

            if not matched and len(vendor_ids) == 1:
                matched = vendor_ids[0]

            if matched:
                db[COLLECTION].update_one(
                    {"_id": order["_id"]},
                    {"$set": {"vendor_id": matched, "updated_at": datetime.now()}},
                )
                updated += 1

        return updated

    count = await asyncio.to_thread(_run)
    return {"backfilled": count}


# ─── PO search ────────────────────────────────────────────────────────────────

@router.get("/po/search")
async def search_po(q: str, db=Depends(get_database)):
    """Search purchase_orders by purchaseorder_number prefix, excluding POs already attached to an order."""
    def _search():
        used_pos = {
            o["purchaseorder_number"]
            for o in db[COLLECTION].find(
                {"purchaseorder_number": {"$nin": [None, ""]}},
                {"purchaseorder_number": 1, "_id": 0},
            )
            if o.get("purchaseorder_number")
        }
        match_filter: dict = {
            "purchaseorder_number": {"$regex": f"^{re.escape(q.strip())}", "$options": "i"},
        }
        if used_pos:
            match_filter["purchaseorder_number"]["$nin"] = list(used_pos)
        pipeline = [
            {"$match": match_filter},
            {"$addFields": {
                "order_status_formatted": {
                    "$cond": {
                        "if": {"$gt": [{"$strLenCP": {"$ifNull": ["$order_status_formatted", ""]}}, 0]},
                        "then": "$order_status_formatted",
                        "else": {
                            "$cond": {
                                "if": {"$gt": [{"$strLenCP": {"$ifNull": ["$order_status", ""]}}, 0]},
                                "then": {
                                    "$concat": [
                                        {"$toUpper": {"$substrCP": ["$order_status", 0, 1]}},
                                        {"$substrCP": ["$order_status", 1, {"$subtract": [{"$strLenCP": "$order_status"}, 1]}]},
                                    ]
                                },
                                "else": "",
                            }
                        },
                    }
                }
            }},
            {"$project": {
                "purchaseorder_number": 1,
                "order_status_formatted": 1,
                "vendor_id": 1,
                "bill_date": {"$arrayElemAt": ["$bills.date", 0]},
                "po_date": {
                    "$cond": {
                        "if": {"$eq": [{"$type": "$date"}, "date"]},
                        "then": {"$dateToString": {"format": "%Y-%m-%d", "date": "$date"}},
                        "else": "$date",
                    }
                },
                "_id": 0,
            }},
            {"$limit": 10},
        ]
        return list(db[PO_COLLECTION].aggregate(pipeline))

    results = await asyncio.to_thread(_search)
    return serialize_mongo_document(results)


# ─── document search ──────────────────────────────────────────────────────────

@router.get("/documents/search")
async def search_documents(q: str, db=Depends(get_database)):
    """Search document filenames across all brand orders."""
    def _search():
        pipeline = [
            {"$unwind": "$documents"},
            {"$match": {"documents.filename": {"$regex": re.escape(q.strip()), "$options": "i"}}},
            {"$project": {
                "order_id": {"$toString": "$_id"},
                "order_name": "$name",
                "brand": "$brand",
                "doc": "$documents",
            }},
            {"$limit": 25},
        ]
        return list(db[COLLECTION].aggregate(pipeline))

    results = await asyncio.to_thread(_search)
    return serialize_mongo_document(results)


# ─── line items ────────────────────────────────────────────────────────────────

_LINE_ITEM_EXPORT_FIELDS = [
    ("name", "Product Name"),
    ("sku", "SKU"),
    ("quantity", "Quantity"),
    ("quantity_received", "Qty Received"),
    ("quantity_billed", "Qty Billed"),
    ("quantity_cancelled", "Qty Cancelled"),
    ("quantity_intransit", "Qty In-Transit"),
    ("quantity_marked_as_received", "Qty Marked Received"),
    ("rate", "Rate"),
    ("_currency_code", "Currency Code"),
    ("item_total", "Item Total"),
    ("tax_name", "Tax Name"),
    ("tax_percentage", "Tax %"),
    ("unit", "Unit"),
    ("hsn_or_sac", "HSN/SAC"),
    ("warehouse_name", "Warehouse"),
    ("line_item_category", "Category"),
    ("header_name", "Header"),
    ("description", "Description"),
]


def _build_line_items_excel(line_items: list, currency_code: str) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Line Items"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    data_font = Font(size=9)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    ws.append([label for _, label in _LINE_ITEM_EXPORT_FIELDS])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 20

    for item in line_items:
        row = []
        for key, _ in _LINE_ITEM_EXPORT_FIELDS:
            if key == "_currency_code":
                row.append(currency_code or "")
            else:
                row.append(item.get(key, ""))
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = data_font
            cell.border = border
            cell.alignment = left

    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/{order_id}/line-items/download")
async def download_order_line_items(order_id: str, db=Depends(get_database)):
    """Download line items for the order's linked PO as XLSX, including the PO's currency code."""
    def _fetch():
        order = db[COLLECTION].find_one(
            {"_id": ObjectId(order_id)},
            {"purchaseorder_number": 1, "brand": 1, "name": 1},
        )
        if not order or not order.get("purchaseorder_number"):
            return None, None, None
        po = db[PO_COLLECTION].find_one(
            {"purchaseorder_number": order["purchaseorder_number"]},
            {"line_items": 1, "currency_code": 1},
        )
        if not po:
            return order, None, None
        return order, po.get("line_items", []), po.get("currency_code", "")

    order, line_items, currency_code = await asyncio.to_thread(_fetch)
    if not order:
        raise HTTPException(status_code=404, detail="Order not found or no PO attached")
    if line_items is None:
        raise HTTPException(status_code=404, detail="Linked PO not found")

    buf = await asyncio.to_thread(_build_line_items_excel, line_items, currency_code or "")
    po_number = order.get("purchaseorder_number", order_id)
    filename = f"{po_number}_line_items.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{order_id}/line-items")
async def get_order_line_items(order_id: str, db=Depends(get_database)):
    def _fetch():
        order = db[COLLECTION].find_one(
            {"_id": ObjectId(order_id)}, {"purchaseorder_number": 1}
        )
        if not order or not order.get("purchaseorder_number"):
            return []
        po = db[PO_COLLECTION].find_one(
            {"purchaseorder_number": order["purchaseorder_number"]},
            {"line_items": 1, "date": 1}
        )
        if not po:
            return []
        line_items = po.get("line_items", [])
        po_date = po.get("date")
        if isinstance(po_date, str):
            try:
                po_date = datetime.strptime(po_date, "%Y-%m-%d")
            except ValueError:
                po_date = None
        if po_date and line_items:
            item_ids = [li.get("item_id") for li in line_items if li.get("item_id")]
            products = {
                str(p["item_id"]): p.get("created_at")
                for p in db["products"].find(
                    {"item_id": {"$in": item_ids}},
                    {"item_id": 1, "created_at": 1, "_id": 0},
                )
                if p.get("item_id")
            }
            for li in line_items:
                prod_created = products.get(str(li.get("item_id")))
                if isinstance(prod_created, datetime) and isinstance(po_date, datetime):
                    li["is_new"] = abs((prod_created - po_date).days) <= 30
                else:
                    li["is_new"] = False
        return line_items
    items = await asyncio.to_thread(_fetch)
    return serialize_mongo_document(items)


# ─── documents ────────────────────────────────────────────────────────────────

@router.post("/{order_id}/documents")
async def upload_document(
    order_id: str,
    file: UploadFile = File(...),
    relative_path: Optional[str] = Form(None),
    category: Optional[str] = Form(None),
    folder: Optional[str] = Form(None),
    item_id: Optional[str] = Form(None),
    item_name: Optional[str] = Form(None),
    db=Depends(get_database),
):
    file_bytes = await file.read()
    filename = file.filename or "document"
    content_type = file.content_type or "application/octet-stream"
    cat = (category or "general").strip()
    folder_path = (folder or "").strip().strip("/")

    def _process():
        doc = db[COLLECTION].find_one({"_id": ObjectId(order_id)}, {"brand": 1, "name": 1})
        if not doc:
            raise ValueError("Order not found")

        brand_slug = _slugify(doc["brand"])
        name_slug = _slugify(doc["name"])
        cat_slug = _slugify(cat)

        # Build S3 key: folder path takes precedence over category slug for new uploads
        if relative_path:
            parts = relative_path.replace("\\", "/").strip("/").split("/")
            safe_parts = [re.sub(r"[^\w.\-]", "_", p) for p in parts if p]
            if folder_path:
                safe_folder = "/".join(re.sub(r"[^\w.\-]", "_", p) for p in folder_path.split("/") if p)
                s3_key = f"brand_orders/{brand_slug}/{name_slug}/{safe_folder}/{'/'.join(safe_parts)}"
            else:
                s3_key = f"brand_orders/{brand_slug}/{name_slug}/{cat_slug}/{'/'.join(safe_parts)}"
        else:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_filename = re.sub(r"[^\w.\-]", "_", filename)
            if folder_path:
                safe_folder = "/".join(re.sub(r"[^\w.\-]", "_", p) for p in folder_path.split("/") if p)
                s3_key = f"brand_orders/{brand_slug}/{name_slug}/{safe_folder}/{ts}_{safe_filename}"
            else:
                s3_key = f"brand_orders/{brand_slug}/{name_slug}/{cat_slug}/{ts}_{safe_filename}"

        _upload_to_s3(file_bytes, s3_key, content_type)

        doc_entry: dict = {
            "doc_id": str(ObjectId()),
            "filename": filename,
            "s3_key": s3_key,
            "content_type": content_type,
            "size": len(file_bytes),
            "uploaded_at": datetime.now(),
            "category": cat,
            "folder": folder_path,
        }
        if item_id:
            doc_entry["item_id"] = item_id
        if item_name:
            doc_entry["item_name"] = item_name

        db[COLLECTION].update_one(
            {"_id": ObjectId(order_id)},
            {
                "$push": {"documents": doc_entry},
                "$set": {"updated_at": datetime.now()},
            },
        )
        return doc_entry

    try:
        doc_entry = await asyncio.to_thread(_process)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.exception("upload_document failed")
        raise HTTPException(status_code=500, detail=str(e))

    return serialize_mongo_document(doc_entry)


@router.get("/{order_id}/documents/zip")
async def download_documents_zip(
    order_id: str,
    item_id: Optional[str] = Query(None),
    db=Depends(get_database),
):
    """Download documents for an order as a ZIP archive. Pass item_id to filter to one line item."""
    def _fetch_docs():
        doc = db[COLLECTION].find_one({"_id": ObjectId(order_id)}, {"documents": 1, "name": 1, "brand": 1})
        if not doc:
            raise ValueError("Order not found")
        return doc

    try:
        order_doc = await asyncio.to_thread(_fetch_docs)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    documents = order_doc.get("documents", [])
    if item_id:
        documents = [d for d in documents if d.get("item_id") == item_id]
    if not documents:
        raise HTTPException(status_code=404, detail="No documents to download")

    def _build_zip():
        s3 = boto3.client("s3", region_name=AWS_REGION)
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for d in documents:
                s3_key = d.get("s3_key")
                if not s3_key:
                    continue
                try:
                    obj = s3.get_object(Bucket=S3_BUCKET, Key=s3_key)
                    data = obj["Body"].read()
                    # Preserve relative path inside zip if key contains subfolders beyond order prefix
                    arcname = d.get("filename") or s3_key.split("/")[-1]
                    zf.writestr(arcname, data)
                except Exception:
                    logger.warning("Skipping %s in zip — download failed", s3_key)
        buf.seek(0)
        return buf

    zip_buf = await asyncio.to_thread(_build_zip)

    safe_name = re.sub(r"[^\w.\-]", "_", order_doc.get("name", order_id))
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}.zip"'},
    )


@router.patch("/{order_id}/documents/{doc_id}")
async def update_document(order_id: str, doc_id: str, body: dict, db=Depends(get_database)):
    updates: dict = {}
    category = (body.get("category") or "").strip()
    if category:
        updates["documents.$.category"] = category
    if "folder" in body:
        updates["documents.$.folder"] = (body["folder"] or "").strip().strip("/")
    if not updates:
        raise HTTPException(status_code=400, detail="Nothing to update — provide category or folder")
    updates["updated_at"] = datetime.now()

    def _update():
        result = db[COLLECTION].update_one(
            {"_id": ObjectId(order_id), "documents.doc_id": doc_id},
            {"$set": updates},
        )
        return result.matched_count

    matched = await asyncio.to_thread(_update)
    if not matched:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"doc_id": doc_id, **{k.replace("documents.$.", ""): v for k, v in updates.items() if k != "updated_at"}}


@router.get("/{order_id}/documents/{doc_id}/url")
async def get_document_url(order_id: str, doc_id: str, db=Depends(get_database)):
    def _fetch():
        doc = db[COLLECTION].find_one(
            {"_id": ObjectId(order_id), "documents.doc_id": doc_id},
            {"documents.$": 1},
        )
        if not doc or not doc.get("documents"):
            return None
        return doc["documents"][0].get("s3_key")

    s3_key = await asyncio.to_thread(_fetch)
    if not s3_key:
        raise HTTPException(status_code=404, detail="Document not found")

    try:
        url = await asyncio.to_thread(_presign_s3, s3_key)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    return {"url": url}


def _days_between(a: Optional[str], b: Optional[str]) -> Optional[int]:
    if not a or not b:
        return None
    try:
        da = datetime.strptime(a, "%Y-%m-%d")
        db_ = datetime.strptime(b, "%Y-%m-%d")
        return (db_ - da).days
    except ValueError:
        return None


def _build_lead_time_excel(orders: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lead Time Report"

    # ── Styles ──────────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    brand_fill = PatternFill("solid", fgColor="D9E1F2")
    brand_font = Font(bold=True, size=9)
    data_font = Font(size=9)
    avg_fill = PatternFill("solid", fgColor="E2EFDA")
    avg_font = Font(bold=True, size=9)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=False)

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    orange_fill = PatternFill("solid", fgColor="FFCC99")
    red_fill = PatternFill("solid", fgColor="FFC7CE")

    def improvement_fill(days_over: Optional[int]) -> Optional[PatternFill]:
        if days_over is None:
            return None
        if days_over <= 0:
            return green_fill
        if days_over <= 3:
            return yellow_fill
        if days_over <= 6:
            return orange_fill
        return red_fill

    # ── Headers ──────────────────────────────────────────────────────────────────
    headers = [
        "Sr. No", "Supplier Name", "Order No",
        "Date of Initiation", "Proforma Invoice Date",
        "Order Processing Days\n(G = F-E)",
        "Order Ready Date",
        "Order Preparing Days\n(I = H-F)",
        "Manufacturer Lead Time\n(J = G+I)",
        "ETD / Sailing Date",
        "Actual Mfg. Lead Time\n(N = G+I)",
        "Ready→ETD Days\n(O = M-H, Target 7-10)",
        "Ready→ETD %\n(P = O/N)",
        "Ready→ETD Over Target\n(Q = O-7)",
        "Total Days\n(R = O+N)",
        "Port / ETA Date",
        "Sail Days\n(T = S-M)",
        "Inward Date",
        "Port→WH Days\n(V = U-S, Target 7)",
        "Port→WH %\n(W = V/N)",
        "Port→WH Over Target\n(X = V-7)",
        "Lead Time\n(AA = Z-F)",
        "Brand",
        "Avg. Lead Time",
    ]
    ws.append(headers)
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    ws.row_dimensions[1].height = 36

    # ── Group by brand ────────────────────────────────────────────────────────────
    from itertools import groupby
    grouped = {}
    for o in orders:
        grouped.setdefault(o.get("brand", "Unknown"), []).append(o)

    sr = 1
    for brand_name, brand_orders in sorted(grouped.items()):
        lead_times = []
        data_rows = []

        for order in brand_orders:
            init = order.get("initiation_date")
            pf = order.get("proforma_date")
            ready = order.get("ready_date")
            etd = order.get("etd_date")
            port = order.get("eta_port_date")
            inward = order.get("inward_date")

            G = _days_between(init, pf)
            I_ = _days_between(pf, ready)
            J = (G + I_) if G is not None and I_ is not None else None
            O_ = _days_between(ready, etd)
            P = round(O_ / J, 4) if O_ is not None and J else None
            Q = (O_ - 7) if O_ is not None else None
            R = (O_ + J) if O_ is not None and J is not None else None
            T_ = _days_between(etd, port)
            V = _days_between(port, inward)
            W = round(V / J, 4) if V is not None and J else None
            X = (V - 7) if V is not None else None
            AA = _days_between(pf, inward)

            if AA is not None:
                lead_times.append(AA)

            po_num = order.get("purchaseorder_number") or order.get("name", "")
            data_rows.append({
                "sr": sr, "brand": brand_name, "po_num": po_num,
                "init": init, "pf": pf, "G": G,
                "ready": ready, "I": I_, "J": J,
                "etd": etd, "N": J, "O": O_, "P": P, "Q": Q, "R": R,
                "port": port, "T": T_,
                "inward": inward, "V": V, "W": W, "X": X,
                "AA": AA,
            })
            sr += 1

        avg_lt = round(sum(lead_times) / len(lead_times), 1) if lead_times else None

        for i, row_data in enumerate(data_rows):
            is_last = (i == len(data_rows) - 1)

            def fmt_date(d):
                if not d:
                    return None
                try:
                    return datetime.strptime(d, "%Y-%m-%d")
                except Exception:
                    return None

            row = [
                row_data["sr"],
                row_data["brand"],
                row_data["po_num"],
                fmt_date(row_data["init"]),
                fmt_date(row_data["pf"]),
                row_data["G"],
                fmt_date(row_data["ready"]),
                row_data["I"],
                row_data["J"],
                fmt_date(row_data["etd"]),
                row_data["N"],
                row_data["O"],
                row_data["P"],
                row_data["Q"],
                row_data["R"],
                fmt_date(row_data["port"]),
                row_data["T"],
                fmt_date(row_data["inward"]),
                row_data["V"],
                row_data["W"],
                row_data["X"],
                row_data["AA"],
                brand_name if is_last else None,
                avg_lt if is_last else None,
            ]
            ws.append(row)
            excel_row = ws.max_row

            for col_idx, cell in enumerate(ws[excel_row], 1):
                cell.font = brand_font if col_idx <= 2 else data_font
                cell.border = border

                # Date columns: format as date
                if col_idx in (4, 5, 7, 10, 16, 18):
                    if cell.value:
                        cell.number_format = "DD-MMM-YYYY"
                    cell.alignment = center
                # Percentage columns
                elif col_idx in (13, 20):
                    if cell.value is not None:
                        cell.number_format = "0.0%"
                    cell.alignment = center
                # Numeric columns
                elif col_idx in (6, 8, 9, 11, 12, 14, 15, 17, 19, 21, 22):
                    cell.alignment = center
                # Brand / avg
                elif col_idx in (23, 24):
                    cell.font = avg_font
                    cell.fill = avg_fill
                    cell.alignment = center
                else:
                    cell.alignment = left

                # Color coding for Q (col 14) and X (col 21)
                if col_idx == 14:
                    f = improvement_fill(row_data["Q"])
                    if f:
                        cell.fill = f
                elif col_idx == 21:
                    f = improvement_fill(row_data["X"])
                    if f:
                        cell.fill = f

        # Blank separator row between brand groups
        ws.append([None] * len(headers))

    # ── Column widths ─────────────────────────────────────────────────────────────
    col_widths = [6, 22, 16, 14, 14, 12, 14, 12, 14, 14, 12, 12, 10, 12, 10, 14, 10, 14, 12, 10, 12, 14, 20, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _fetch_fx_rates(currencies: set) -> dict:
    """Return {currency_code: inr_rate} using Frankfurter (free, no key needed)."""
    rates: dict = {}
    for cur in currencies:
        if not cur or cur == "INR":
            rates[cur] = 1.0
            continue
        try:
            resp = requests.get(
                "https://api.frankfurter.app/latest",
                params={"from": cur, "to": "INR"},
                timeout=10,
            )
            if resp.status_code == 200:
                rates[cur] = resp.json().get("rates", {}).get("INR")
            else:
                rates[cur] = None
        except Exception:
            rates[cur] = None
    return rates


def _build_payment_report_excel(orders: list, fx_rates: dict) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Report"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    data_font = Font(size=9)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    num_fmt = '#,##0.00'

    headers = [
        "Brand — Order Number",
        "PO Number",
        "PO Status",
        "PO Total Value (Foreign Currency)",
        "Live INR Value",
        "Exchange Rate Used (INR +₹0.20)",
        "PO Due Date",
        "Custom Duty (INR)",
        "Custom Duty Due Date",
        "Shipping Charges (INR)",
        "Shipping Charges Due Date",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 20

    for order in orders:
        brand_order_label = f"{order.get('brand', '')} — {order.get('name', '')}"
        po_total = order.get("po_sub_total")
        po_currency = order.get("po_currency_code") or ""

        po_total_str = f"{po_currency} {po_total:,.2f}" if po_total is not None else ""

        inr_value = None
        rate_used = None
        if po_total is not None and po_currency:
            raw_rate = fx_rates.get(po_currency)
            if raw_rate is not None:
                # INR stays at 1.0; all other currencies get +₹0.20 spread
                rate_used = 1.0 if po_currency == "INR" else round(raw_rate + 0.20, 4)
                inr_value = round(po_total * rate_used, 2)

        def fmt_date_cell(d):
            if not d:
                return None
            try:
                return datetime.strptime(d, "%Y-%m-%d")
            except Exception:
                return None

        row = [
            brand_order_label,
            order.get("purchaseorder_number", ""),
            order.get("po_status", ""),
            po_total_str,
            inr_value,
            rate_used,
            fmt_date_cell(order.get("po_due_date")),
            order.get("custom_duty"),
            fmt_date_cell(order.get("custom_duty_due_date")),
            order.get("shipping_charges"),
            fmt_date_cell(order.get("shipping_charges_due_date")),
        ]
        ws.append(row)
        excel_row = ws.max_row
        for col_idx, cell in enumerate(ws[excel_row], 1):
            cell.font = data_font
            cell.border = border
            if col_idx in (7, 9, 11):  # date columns
                if cell.value:
                    cell.number_format = "DD-MMM-YYYY"
                cell.alignment = center
            elif col_idx in (5, 6, 8, 10):  # numeric columns (INR value, rate, custom duty, shipping)
                cell.number_format = num_fmt
                cell.alignment = center
            else:
                cell.alignment = left

    col_widths = [40, 20, 16, 28, 20, 26, 16, 20, 20, 22, 24]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    # ── FX Rates sheet ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("FX Rates")
    fx_headers = ["Currency", "Live Rate (Frankfurter, INR per 1 unit)", "Rate Used (+₹0.20 spread)", "Note"]
    ws2.append(fx_headers)
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    ws2.row_dimensions[1].height = 20

    # Collect unique currencies from orders (preserving encounter order)
    seen: dict = {}
    for order in orders:
        cur = order.get("po_currency_code") or ""
        if cur and cur not in seen:
            seen[cur] = fx_rates.get(cur)

    rate4_fmt = '#,##0.0000'
    for cur, raw_rate in seen.items():
        if cur == "INR":
            adj_rate = 1.0
            note = "Base currency — no spread applied"
        elif raw_rate is not None:
            adj_rate = round(raw_rate + 0.20, 4)
            note = "Frankfurter live rate + ₹0.20 spread"
        else:
            adj_rate = None
            note = "Rate unavailable"
        row2 = [cur, raw_rate, adj_rate, note]
        ws2.append(row2)
        r = ws2.max_row
        for col_idx, cell in enumerate(ws2[r], 1):
            cell.font = data_font
            cell.border = border
            if col_idx in (2, 3):
                cell.number_format = rate4_fmt
                cell.alignment = center
            elif col_idx == 1:
                cell.alignment = center
                cell.font = Font(bold=True, size=9)
            else:
                cell.alignment = left

    for i, w in enumerate([14, 38, 30, 38], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.delete("/{order_id}/documents/{doc_id}")
async def delete_document(order_id: str, doc_id: str, db=Depends(get_database)):
    def _delete():
        doc = db[COLLECTION].find_one(
            {"_id": ObjectId(order_id), "documents.doc_id": doc_id},
            {"documents.$": 1},
        )
        if not doc or not doc.get("documents"):
            return False
        s3_key = doc["documents"][0].get("s3_key")
        if s3_key:
            try:
                _delete_from_s3(s3_key)
            except Exception:
                pass
        db[COLLECTION].update_one(
            {"_id": ObjectId(order_id)},
            {
                "$pull": {"documents": {"doc_id": doc_id}},
                "$set": {"updated_at": datetime.now()},
            },
        )
        return True

    found = await asyncio.to_thread(_delete)
    if not found:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"order_id": order_id, "doc_id": doc_id, "deleted": True}
