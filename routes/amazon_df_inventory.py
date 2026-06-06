from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from datetime import datetime
from ..helpers.datetime_utils import utcnow
from ..database import get_database
import asyncio
import csv
import io
import uuid
import re
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)
router = APIRouter()

COLLECTION = "amazon_df_inventory"

COLUMNS = ["Date", "SKU", "ASIN", "Title", "Warehouse", "Warehouse name", "Available units", "Status"]


def _clean_sku(raw: str) -> str:
    """Strip Excel ="..." formula wrapper from SKU values."""
    s = raw.strip()
    m = re.match(r'^="?([^"]*)"?$', s)
    return m.group(1) if m else s


def _parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        rows.append({
            "date": row.get("Date", "").strip(),
            "sku": _clean_sku(row.get("SKU", "")),
            "asin": row.get("ASIN", "").strip(),
            "title": row.get("Title", "").strip(),
            "warehouse": row.get("Warehouse", "").strip(),
            "warehouse_name": row.get("Warehouse name", "").strip(),
            "available_units": _safe_int(row.get("Available units", "")),
            "status": row.get("Status", "").strip(),
        })
    return rows


def _parse_xlsx(content: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {h: i for i, h in enumerate(headers)}

    def _get(row_vals, name):
        i = col_idx.get(name)
        return str(row_vals[i]).strip() if i is not None and row_vals[i] is not None else ""

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rows.append({
            "date": _get(row, "Date"),
            "sku": _clean_sku(_get(row, "SKU")),
            "asin": _get(row, "ASIN"),
            "title": _get(row, "Title"),
            "warehouse": _get(row, "Warehouse"),
            "warehouse_name": _get(row, "Warehouse name"),
            "available_units": _safe_int(_get(row, "Available units")),
            "status": _get(row, "Status"),
        })
    return rows


def _safe_int(val: str) -> int | None:
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _build_template_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Amazon DF Inventory"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    widths = [14, 18, 14, 60, 12, 40, 16, 12]

    for i, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ─── Endpoints ────────────────────────────────────────────────────────────────

@router.get("/template")
async def download_template():
    xlsx_bytes = _build_template_xlsx()
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="amazon_df_inventory_template.xlsx"'},
    )


@router.post("/upload")
async def upload_inventory(
    file: UploadFile = File(...),
    start_date: str = Form(...),
    end_date: str = Form(...),
    db=Depends(get_database),
):
    content = await file.read()
    filename = file.filename or ""

    if filename.lower().endswith(".csv"):
        rows = _parse_csv(content)
    elif filename.lower().endswith((".xlsx", ".xls")):
        rows = _parse_xlsx(content)
    else:
        raise HTTPException(status_code=400, detail="Only CSV or XLSX files are supported")

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in file")

    upload_id = str(uuid.uuid4())
    uploaded_at = utcnow()

    docs = [
        {
            **row,
            "upload_id": upload_id,
            "start_date": start_date,
            "end_date": end_date,
            "uploaded_at": uploaded_at,
            "filename": filename,
        }
        for row in rows
    ]

    def _insert():
        db[COLLECTION].insert_many(docs)

    await asyncio.to_thread(_insert)

    return {
        "upload_id": upload_id,
        "rows_inserted": len(docs),
        "start_date": start_date,
        "end_date": end_date,
        "filename": filename,
        "uploaded_at": uploaded_at.isoformat(),
    }


@router.get("/uploads")
async def list_uploads(db=Depends(get_database)):
    def _fetch():
        pipeline = [
            {"$sort": {"uploaded_at": -1}},
            {
                "$group": {
                    "_id": "$upload_id",
                    "upload_id": {"$first": "$upload_id"},
                    "start_date": {"$first": "$start_date"},
                    "end_date": {"$first": "$end_date"},
                    "filename": {"$first": "$filename"},
                    "uploaded_at": {"$first": "$uploaded_at"},
                    "row_count": {"$sum": 1},
                }
            },
            {"$sort": {"uploaded_at": -1}},
        ]
        return list(db[COLLECTION].aggregate(pipeline))

    uploads = await asyncio.to_thread(_fetch)
    result = []
    for u in uploads:
        result.append({
            "upload_id": u["upload_id"],
            "start_date": u["start_date"],
            "end_date": u["end_date"],
            "filename": u["filename"],
            "uploaded_at": u["uploaded_at"].isoformat() if isinstance(u["uploaded_at"], datetime) else u["uploaded_at"],
            "row_count": u["row_count"],
        })
    return result


@router.get("/{upload_id}/rows")
async def get_upload_rows(upload_id: str, db=Depends(get_database)):
    def _fetch():
        return list(
            db[COLLECTION].find(
                {"upload_id": upload_id},
                {"_id": 0, "upload_id": 0, "uploaded_at": 0, "filename": 0},
            )
        )

    rows = await asyncio.to_thread(_fetch)
    if not rows:
        raise HTTPException(status_code=404, detail="Upload not found")
    return rows


@router.delete("/{upload_id}")
async def delete_upload(upload_id: str, db=Depends(get_database)):
    def _delete():
        return db[COLLECTION].delete_many({"upload_id": upload_id})

    result = await asyncio.to_thread(_delete)
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Upload not found")
    return {"deleted": result.deleted_count}
