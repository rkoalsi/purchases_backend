from fastapi import APIRouter, HTTPException, Depends, Query, Body, UploadFile, File
from fastapi.responses import StreamingResponse, JSONResponse
import io
import os
import re
import asyncio
import logging
import requests
import pandas as pd
from functools import lru_cache
from openpyxl import Workbook, load_workbook
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import List, Dict, Optional
from pydantic import BaseModel
from ..database import get_database, serialize_mongo_document
from .task_triggers import fire_trigger

router = APIRouter()
logger = logging.getLogger(__name__)

ORGANIZATION_ID = os.getenv("ORGANIZATION_ID", "776755316")
BOOKS_URL = os.getenv("BOOKS_URL")
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
BOOKS_REFRESH_TOKEN = os.getenv("BOOKS_REFRESH_TOKEN")
ZOHO_BOOKS_BASE = "https://books.zoho.com/api/v3"

PRODUCTS_COLLECTION = "products"
PURCHASE_ORDERS_COLLECTION = "purchase_orders"
DRAFT_PURCHASE_ORDERS_COLLECTION = "draft_purchase_orders"
NOTIFICATIONS_COLLECTION = "notifications"


def _get_zoho_token() -> str:
    url = BOOKS_URL.format(
        clientId=CLIENT_ID,
        clientSecret=CLIENT_SECRET,
        grantType="refresh_token",
        books_refresh_token=BOOKS_REFRESH_TOKEN,
    )
    r = requests.post(url, timeout=30)
    r.raise_for_status()
    return r.json()["access_token"]


def _parse_draft_order_excel(file_bytes: bytes) -> list[dict]:
    """Parse draft order Excel. Detects column positions from headers so files with an
    extra SKU Codes column (or other variations) are handled correctly.
    Expected headers (case-insensitive, partial match): Manufacturer Code, BBCode,
    Item Name, Qty, Unit Price. Reads number_format on the Unit Price cell to detect currency.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Prefer the "Draft Order …" sheet; fall back to the active (first) sheet.
    # Master report files have the main sales sheet first, so wb.active would
    # land on the wrong sheet without this lookup.
    draft_sheet = next(
        (wb[name] for name in wb.sheetnames if name.lower().startswith("draft order")),
        wb.active,
    )
    ws = draft_sheet

    header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
    if header_row is None:
        raise ValueError("Order file has no data rows")

    # Map column name keywords → 0-based index
    col_map: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        h = str(cell.value or "").strip().lower()
        if "manufacturer" in h:
            col_map.setdefault("mfr", idx)
        elif "bbcode" in h or "bb code" in h or "bb_code" in h:
            col_map.setdefault("bb", idx)
        elif "item name" in h or "item_name" in h:
            col_map.setdefault("name", idx)
        elif "qty" in h or "quantity" in h:
            col_map.setdefault("qty", idx)
        elif "unit price" in h or "unit_price" in h:
            col_map.setdefault("price", idx)
        elif "hsn" in h:
            col_map.setdefault("hsn", idx)
        elif "sub category" in h or "sub_category" in h:
            col_map.setdefault("sub_category", idx)
        elif "category" in h:
            col_map.setdefault("category", idx)
        elif "series" in h:
            col_map.setdefault("series", idx)
        elif "mrp" in h:
            col_map.setdefault("mrp", idx)
        elif "sku" in h:
            col_map.setdefault("sku_barcode", idx)

    # Fall back to positional defaults if headers not recognised
    mfr_idx = col_map.get("mfr", 0)
    bb_idx = col_map.get("bb", 1)
    name_idx = col_map.get("name", 2)
    qty_idx = col_map.get("qty", 3)
    price_idx = col_map.get("price", 4)
    hsn_idx = col_map.get("hsn")
    sku_barcode_idx = col_map.get("sku_barcode")
    category_idx = col_map.get("category")
    sub_category_idx = col_map.get("sub_category")
    series_idx = col_map.get("series")
    mrp_idx = col_map.get("mrp")

    def _cell_str(row, idx):
        if idx is None or idx >= len(row):
            return ""
        v = row[idx].value
        return str(v).strip() if v is not None else ""

    items = []
    for row in ws.iter_rows(min_row=2):
        if len(row) <= max(mfr_idx, bb_idx, name_idx, qty_idx, price_idx):
            continue
        if not row[mfr_idx].value and not row[bb_idx].value:
            continue
        qty = row[qty_idx].value
        if qty is not None and isinstance(qty, str):
            try:
                qty = float(qty)
            except ValueError:
                qty = None
        if not isinstance(qty, (int, float)) or qty <= 0:
            continue

        price_cell = row[price_idx]
        fmt = price_cell.number_format or ""
        currency = "CNY" if ("¥" in fmt or "\\¥" in fmt) else "USD"

        mrp_val = None
        if mrp_idx is not None and mrp_idx < len(row) and row[mrp_idx].value is not None:
            try:
                mrp_val = float(row[mrp_idx].value)
            except (ValueError, TypeError):
                pass

        items.append(
            {
                "manufacturer_code": str(row[mfr_idx].value).strip() if row[mfr_idx].value else "",
                "bb_code": str(row[bb_idx].value).strip() if row[bb_idx].value else "",
                "item_name": str(row[name_idx].value).strip() if row[name_idx].value else "",
                "qty": int(qty),
                "unit_price": float(price_cell.value) if price_cell.value is not None else 0.0,
                "currency": currency,
                "hsn_or_sac": _cell_str(row, hsn_idx),
                "sku_code": _cell_str(row, sku_barcode_idx),
                "category": _cell_str(row, category_idx),
                "sub_category": _cell_str(row, sub_category_idx),
                "series": _cell_str(row, series_idx),
                "mrp": mrp_val,
            }
        )
    return items


TAX_RATE_MAP: dict[int, dict] = {
    5: {
        "intra": {"tax_id": "3220178000000935227", "tax_specification": "intra"},
        "inter": {"tax_id": "3220178000000085045", "tax_specification": "inter"},
    },
    12: {
        "intra": {"tax_id": "3220178000000085117", "tax_specification": "intra"},
        "inter": {"tax_id": "3220178000000085049", "tax_specification": "inter"},
    },
    18: {
        "intra": {"tax_id": "3220178000000085129", "tax_specification": "intra"},
        "inter": {"tax_id": "3220178000000085053", "tax_specification": "inter"},
    },
}

# HSN code → GST rate mapping. Sourced from HSN master sheet.
# Conditional entries include a price_threshold: rate is below_rate when price <= threshold,
# otherwise the standard rate applies.
HSN_GST_MAP: dict[str, dict] = {
    "33059030":  {"rate": 5},
    "63079099":  {"rate": 5, "price_threshold": 2500, "above_threshold_rate": 18},
    "96190090":  {"rate": 5},
    "1517":      {"rate": 5},
    "62171030":  {"rate": 5},
    "23091000":  {"rate": 18},
    "15159099":  {"rate": 5},
    "33073090":  {"rate": 18},
    "39269099":  {"rate": 18},
    "40169990":  {"rate": 18},
    "40169100":  {"rate": 18},
    "3926":      {"rate": 18},
    "96161010":  {"rate": 5},
    "21069099":  {"rate": 18},
    "48239090":  {"rate": 18},
    "48189000":  {"rate": 18},
    "25081090":  {"rate": 5},
    "63079090":  {"rate": 5},
    "42010000":  {"rate": 18},
    "23099090":  {"rate": 18},
    "38249900":  {"rate": 18},
    "82142010":  {"rate": 18},
    "40170090":  {"rate": 18},
    "33051090":  {"rate": 18},
    "14049090":  {"rate": 5},
    "58110090":  {"rate": 5},
}


def resolve_gst_rate(hsn_or_sac: str, price: Optional[float]) -> int:
    """Return GST rate % for the given HSN code and price. Defaults to 18 if HSN not in map."""
    entry = HSN_GST_MAP.get(str(hsn_or_sac).strip())
    if not entry:
        return 18
    if "price_threshold" in entry and price is not None:
        if price > entry["price_threshold"]:
            return entry["above_threshold_rate"]
    return entry["rate"]


class ZohoItemToCreate(BaseModel):
    item_name: str
    manufacturer_code: str = ""
    bb_code: str = ""
    sku_code: str = ""
    hsn_or_sac: str = ""
    category: str = ""
    sub_category: str = ""
    series: str = ""
    mrp: Optional[float] = None
    brand: str = ""
    tax_rate: int = 18
    upc_code: str = ""
    ean_code: str = ""


class CreateZohoItemsRequest(BaseModel):
    items: List[ZohoItemToCreate]


class SavePendingItemRequest(BaseModel):
    bb_code: str
    manufacturer_code: str = ""
    tax_rate: int = 18
    upc_code: str = ""
    ean_code: str = ""


class CreateDraftPORequest(BaseModel):
    contact_id: str
    date: str
    items: List[dict]
    notes: Optional[str] = ""
    reference_number: Optional[str] = ""
    purchaseorder_number: Optional[str] = ""
    description: Optional[str] = ""
    draft_id: Optional[str] = None


class SaveDraftOrderRequest(BaseModel):
    description: Optional[str] = ""
    items: List[dict]
    detected_vendor: Optional[dict] = None
    date: str
    notes: Optional[str] = ""
    reference_number: Optional[str] = ""
    purchaseorder_number: Optional[str] = ""


@router.get("/brands")
def get_available_brands(db=Depends(get_database)):
    """Get list of available brands with their assigned vendors (supports multi-vendor)."""
    try:
        brands_collection = db.get_collection("brands")

        pipeline = [
            {
                # Normalize: prefer vendor_ids array; fall back to wrapping legacy vendor_id
                "$addFields": {
                    "normalized_ids": {
                        "$cond": {
                            "if": {"$isArray": "$vendor_ids"},
                            "then": "$vendor_ids",
                            "else": {
                                "$cond": {
                                    "if": {"$gt": ["$vendor_id", None]},
                                    "then": ["$vendor_id"],
                                    "else": [],
                                }
                            },
                        }
                    }
                }
            },
            {
                "$lookup": {
                    "from": "vendors",
                    "localField": "normalized_ids",
                    "foreignField": "contact_id",
                    "as": "vendor_data",
                }
            },
            {
                "$addFields": {
                    "vendors": {
                        "$map": {
                            "input": "$vendor_data",
                            "as": "v",
                            "in": {
                                "contact_id": "$$v.contact_id",
                                "contact_name": "$$v.contact_name",
                                "currency_code": {"$ifNull": ["$$v.currency_code", ""]},
                            },
                        }
                    }
                }
            },
            {"$project": {"vendor_data": 0, "normalized_ids": 0}},
            {"$sort": {"name": 1}},
        ]

        brands = list(brands_collection.aggregate(pipeline))
        return {"brands": serialize_mongo_document(brands)}

    except Exception as e:
        logger.error(f"Error getting brands: {e}")
        raise HTTPException(status_code=500, detail="Error retrieving brands")


class CreateBrandRequest(BaseModel):
    name: str


class RenameBrandRequest(BaseModel):
    name: str


@router.post("/brands")
def create_brand(body: CreateBrandRequest, db=Depends(get_database)):
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Brand name is required")
    brands_collection = db.get_collection("brands")
    if brands_collection.find_one({"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}}):
        raise HTTPException(status_code=409, detail="Brand already exists")
    result = brands_collection.insert_one({"name": name})
    return {"_id": str(result.inserted_id), "name": name}


@router.patch("/brands/{brand_id}")
def rename_brand(brand_id: str, body: RenameBrandRequest, db=Depends(get_database)):
    from bson import ObjectId
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Brand name is required")
    try:
        oid = ObjectId(brand_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid brand_id")
    brands_collection = db.get_collection("brands")
    conflict = brands_collection.find_one(
        {"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}, "_id": {"$ne": oid}}
    )
    if conflict:
        raise HTTPException(status_code=409, detail="A brand with that name already exists")
    result = brands_collection.update_one({"_id": oid}, {"$set": {"name": name}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Brand not found")
    return {"_id": brand_id, "name": name}


@router.delete("/brands/{brand_id}")
def delete_brand(brand_id: str, db=Depends(get_database)):
    from bson import ObjectId
    try:
        oid = ObjectId(brand_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid brand_id")
    brands_collection = db.get_collection("brands")
    brand = brands_collection.find_one({"_id": oid}, {"vendor_ids": 1, "vendor_id": 1})
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    ids = brand.get("vendor_ids") or ([brand["vendor_id"]] if brand.get("vendor_id") else [])
    if ids:
        raise HTTPException(status_code=400, detail="Remove all vendors before deleting this brand")
    brands_collection.delete_one({"_id": oid})
    return {"deleted": True}


@router.get("")
def get_vendors(
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    search: str | None = Query(None, description="Search by contact_name"),
    db=Depends(get_database),
):
    """Get paginated list of vendors with optional search on contact_name"""
    try:
        vendors_collection = db.get_collection("vendors")

        # Build filter
        query = {}
        if search:
            query["contact_name"] = {
                "$regex": re.escape(search),
                "$options": "i",  # case-insensitive
            }

        skip = (page - 1) * page_size

        vendors_cursor = vendors_collection.find(query).skip(skip).limit(page_size)
        vendors = list(vendors_cursor)

        total_count = vendors_collection.count_documents(query)

        return {
            "page": page,
            "page_size": page_size,
            "total": total_count,
            "total_pages": (total_count + page_size - 1) // page_size,
            "search": search,
            "vendors": serialize_mongo_document(vendors),
        }

    except Exception as e:
        logger.error(f"Error getting vendors: {e}")
        raise HTTPException(status_code=500, detail="Error retrieving vendors")


@router.put("/brands/vendor")
def update_brand_vendor(
    name: str = Body(..., embed=True),
    vendor_ids: List[str] = Body(..., embed=True),
    db=Depends(get_database),
):
    """Update vendor_ids list for a brand (max 5 vendors)."""
    if len(vendor_ids) > 5:
        raise HTTPException(status_code=400, detail="Maximum 5 vendors allowed per brand")

    try:
        brands_collection = db.get_collection("brands")

        query = {"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}}
        update = {"$set": {"vendor_ids": vendor_ids}}

        result = brands_collection.update_many(query, update)

        if result.matched_count == 0:
            raise HTTPException(
                status_code=404, detail="No brand found with given name"
            )

        return {
            "message": "Vendors updated successfully",
            "matched_count": result.matched_count,
            "modified_count": result.modified_count,
        }

    except Exception as e:
        logger.error(f"Error updating brand vendor_id: {e}")
        raise HTTPException(status_code=500, detail="Error updating brand")


@router.post("/draft_orders/validate")
async def validate_draft_order(
    file: UploadFile = File(...),
    db=Depends(get_database),
):
    """Parse draft order Excel and check all items exist in the product database."""
    file_bytes = await file.read()

    def _process():
        items = _parse_draft_order_excel(file_bytes)
        if not items:
            raise ValueError("No valid data rows found in file")

        bb_codes = [it["bb_code"] for it in items if it["bb_code"]]
        mfr_codes = [it["manufacturer_code"] for it in items if it["manufacturer_code"]]

        products_col = db.get_collection(PRODUCTS_COLLECTION)

        PRODUCT_PROJECTION = {
            "cf_sku_code": 1, "cf_item_code": 1, "item_id": 1,
            "name": 1, "item_name": 1, "brand": 1,
            "hsn_or_sac": 1, "purchase_account_id": 1, "status": 1,
        }

        products_by_bb: dict[str, dict] = {}
        for p in products_col.find({"cf_sku_code": {"$in": bb_codes}}, PRODUCT_PROJECTION):
            key = p["cf_sku_code"]
            existing = products_by_bb.get(key)
            # Prefer active product over inactive when multiple share the same SKU code
            if existing is None or (p.get("status") == "active" and existing.get("status") != "active"):
                products_by_bb[key] = p

        remaining_mfr = [
            it["manufacturer_code"]
            for it in items
            if it["manufacturer_code"]
            and (not it["bb_code"] or it["bb_code"] not in products_by_bb)
        ]
        products_by_mfr: dict[str, dict] = {}
        if remaining_mfr:
            for p in products_col.find({"cf_item_code": {"$in": remaining_mfr}}, PRODUCT_PROJECTION):
                key = p.get("cf_item_code", "")
                existing = products_by_mfr.get(key)
                if existing is None or (p.get("status") == "active" and existing.get("status") != "active"):
                    products_by_mfr[key] = p

        validated = []
        missing = []
        inactive_items = []
        first_brand: str | None = None
        for it in items:
            product = None
            if it["bb_code"] and it["bb_code"] in products_by_bb:
                product = products_by_bb[it["bb_code"]]
            elif it["manufacturer_code"] and it["manufacturer_code"] in products_by_mfr:
                product = products_by_mfr[it["manufacturer_code"]]

            if product:
                brand = product.get("brand") or ""
                if brand and first_brand is None:
                    first_brand = brand
                item_status = product.get("status", "active")
                enriched = {
                    **it,
                    "item_id": product.get("item_id", ""),
                    "product_name": product.get("name") or product.get("item_name", ""),
                    "hsn_or_sac": product.get("hsn_or_sac", ""),
                    "purchase_account_id": product.get("purchase_account_id", ""),
                    "status": item_status,
                }
                if item_status != "active":
                    inactive_items.append({
                        "manufacturer_code": it["manufacturer_code"],
                        "bb_code": it["bb_code"],
                        "item_name": enriched["product_name"] or it["item_name"],
                        "status": item_status,
                    })
                else:
                    validated.append(enriched)
            else:
                barcode = it.get("sku_code", "") or ""
                missing.append(
                    {
                        "manufacturer_code": it["manufacturer_code"],
                        "bb_code": it["bb_code"],
                        "item_name": it["item_name"],
                        "hsn_or_sac": it.get("hsn_or_sac", ""),
                        "sku_code": barcode,
                        "category": it.get("category", ""),
                        "sub_category": it.get("sub_category", ""),
                        "series": it.get("series", ""),
                        "mrp": it.get("mrp"),
                        "tax_rate": resolve_gst_rate(it.get("hsn_or_sac", ""), it.get("mrp")),
                        "upc_code": barcode,
                        "ean_code": barcode,
                    }
                )

        # Detect vendors from the first product's brand (supports multi-vendor)
        detected_vendors = []
        if first_brand:
            brand_doc = db.get_collection("brands").find_one(
                {"name": {"$regex": f"^{re.escape(first_brand)}$", "$options": "i"}},
                {"vendor_id": 1, "vendor_ids": 1},
            )
            if brand_doc:
                ids = brand_doc.get("vendor_ids") or []
                if not ids and brand_doc.get("vendor_id"):
                    ids = [brand_doc["vendor_id"]]
                for vid in ids:
                    vendor_doc = db.get_collection("vendors").find_one(
                        {"contact_id": vid},
                        {"contact_id": 1, "contact_name": 1, "currency_code": 1},
                    )
                    if vendor_doc:
                        detected_vendors.append({
                            "contact_id": vendor_doc.get("contact_id", ""),
                            "contact_name": vendor_doc.get("contact_name", ""),
                            "currency_code": vendor_doc.get("currency_code", "USD"),
                        })

        return validated, missing, inactive_items, detected_vendors

    try:
        validated, missing, inactive_items, detected_vendors = await asyncio.to_thread(_process)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    if missing:
        # Merge any previously saved edits (tax_rate, upc_code, ean_code) from DB
        keys = [m["bb_code"] or m["manufacturer_code"] for m in missing if m["bb_code"] or m["manufacturer_code"]]
        if keys:
            saved = {
                doc["_key"]: doc
                for doc in db.get_collection("draft_order_pending_items").find({"_key": {"$in": keys}})
            }
            for m in missing:
                key = m["bb_code"] or m["manufacturer_code"]
                if key in saved:
                    s = saved[key]
                    if "tax_rate" in s:
                        m["tax_rate"] = s["tax_rate"]
                    if "upc_code" in s:
                        m["upc_code"] = s["upc_code"]
                    if "ean_code" in s:
                        m["ean_code"] = s["ean_code"]
        return JSONResponse(
            status_code=200,
            content={
                "valid": False,
                "missing_items": missing,
                "found_count": len(validated),
                "missing_count": len(missing),
            },
        )

    return JSONResponse(
        status_code=200,
        content={
            "valid": True,
            "items": serialize_mongo_document(validated),
            "total_items": len(validated),
            "inactive_items": inactive_items,
            "detected_vendors": detected_vendors,
        },
    )


@router.put("/draft_orders/pending_item")
async def save_pending_item(body: SavePendingItemRequest, db=Depends(get_database)):
    """Upsert a missing item's editable fields into draft_order_pending_items collection."""
    key = body.bb_code or body.manufacturer_code
    if not key:
        raise HTTPException(status_code=400, detail="bb_code or manufacturer_code required")
    col = db.get_collection("draft_order_pending_items")
    col.update_one(
        {"_key": key},
        {"$set": {
            "_key": key,
            "bb_code": body.bb_code,
            "manufacturer_code": body.manufacturer_code,
            "tax_rate": body.tax_rate,
            "upc_code": body.upc_code,
            "ean_code": body.ean_code,
            "updated_at": datetime.utcnow(),
        }},
        upsert=True,
    )
    return {"ok": True}


@router.post("/draft_orders/create_zoho_items")
async def create_zoho_items(body: CreateZohoItemsRequest):
    """Create missing items on Zoho Books. Returns per-item created/failed results."""

    FIXED_ACCOUNT_ID         = "3220178000000000388"
    FIXED_PURCHASE_ACCOUNT   = "3220178000000034003"
    FIXED_INVENTORY_ACCOUNT  = "3220178000000034001"
    FIXED_UNIT               = "pcs"
    FIXED_UNIT_ID            = "3220178000000075076"
    CUSTOM_FIELD_MFR_CODE    = "3220178000000075178"
    CUSTOM_FIELD_SKU_CODE    = "3220178000000075190"
    CUSTOM_FIELD_CATEGORY    = "3220178000505344016"
    CUSTOM_FIELD_SUB_CAT     = "3220178000505344022"
    CUSTOM_FIELD_SERIES      = "3220178000505344028"
    CUSTOM_FIELD_EXTRA       = "3220178000517626009"

    def _do_create():
        token = _get_zoho_token()
        headers = {"Authorization": f"Zoho-oauthtoken {token}"}
        created = []
        failed = []

        for item in body.items:
            barcode = item.sku_code or ""
            upc = item.upc_code or barcode
            ean = item.ean_code or barcode
            tax_prefs = TAX_RATE_MAP.get(item.tax_rate, TAX_RATE_MAP[18])
            item_tax_preferences = [tax_prefs["intra"], tax_prefs["inter"]]
            payload = {
                "name": item.item_name,
                "rate": item.mrp or 0,
                "account_id": FIXED_ACCOUNT_ID,
                "sku": barcode,
                "upc": upc,
                "ean": ean,
                "purchase_rate": 0,
                "purchase_account_id": FIXED_PURCHASE_ACCOUNT,
                "purchase_description": "",
                "inventory_account_id": FIXED_INVENTORY_ACCOUNT,
                "hsn_or_sac": item.hsn_or_sac or "",
                "brand": item.brand or "",
                "product_type": "goods",
                "is_taxable": True,
                "taxability_type": "none",
                "item_tax_preferences": item_tax_preferences,
                "track_serial_number": False,
                "track_batch_number": False,
                "is_returnable": True,
                "can_be_sold": True,
                "can_be_purchased": True,
                "track_inventory": True,
                "inventory_valuation_method": "fifo",
                "unit": FIXED_UNIT,
                "unit_id": FIXED_UNIT_ID,
                "unit_group_id": "",
                "default_sales_unit_conversion_id": "",
                "default_purchase_unit_conversion_id": "",
                "package_details": {
                    "length": "", "width": "", "height": "",
                    "weight": "", "weight_unit": "kg", "dimension_unit": "cm",
                },
                "custom_fields": [
                    {"customfield_id": CUSTOM_FIELD_MFR_CODE, "value": item.manufacturer_code or ""},
                    {"customfield_id": CUSTOM_FIELD_SKU_CODE, "value": item.bb_code or ""},
                    {"customfield_id": CUSTOM_FIELD_CATEGORY, "value": item.category or ""},
                    {"customfield_id": CUSTOM_FIELD_SUB_CAT,  "value": item.sub_category or ""},
                    {"customfield_id": CUSTOM_FIELD_SERIES,   "value": item.series or ""},
                    {"customfield_id": CUSTOM_FIELD_EXTRA,    "value": ""},
                ],
            }
            try:
                r = requests.post(
                    f"{ZOHO_BOOKS_BASE}/items",
                    headers=headers,
                    json=payload,
                    params={"organization_id": ORGANIZATION_ID},
                    timeout=30,
                )
                r.raise_for_status()
                data = r.json()
                if data.get("code") != 0:
                    raise ValueError(data.get("message", "Zoho error"))
                zoho_item = data.get("item", {})
                created.append({
                    "item_name": item.item_name,
                    "bb_code": item.bb_code,
                    "manufacturer_code": item.manufacturer_code,
                    "item_id": zoho_item.get("item_id", ""),
                })
            except Exception as exc:
                logger.error("Failed to create Zoho item '%s': %s", item.item_name, exc)
                failed.append({
                    "item_name": item.item_name,
                    "bb_code": item.bb_code,
                    "manufacturer_code": item.manufacturer_code,
                    "error": str(exc),
                })

        return created, failed

    try:
        created, failed = await asyncio.to_thread(_do_create)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    if created:
        def _notify():
            import os as _os
            purchase_url = _os.getenv("SLACK_URL_PURCHASE")
            design_url   = _os.getenv("SLACK_URL_DESIGN")
            n = len(created)
            item_lines = "\n".join(
                f"• *{c['item_name']}*  `{c['bb_code'] or c['manufacturer_code']}`"
                for c in created
            )
            blocks = [
                {
                    "type": "header",
                    "text": {"type": "plain_text", "text": f":sparkles: {n} New Product{'s' if n != 1 else ''} Created on Zoho", "emoji": True},
                },
                {
                    "type": "section",
                    "text": {"type": "mrkdwn", "text": item_lines},
                },
                {
                    "type": "context",
                    "elements": [{"type": "mrkdwn", "text": f"Created via Draft Order Upload · {datetime.utcnow().strftime('%d %b %Y, %H:%M UTC')}"}],
                },
            ]
            payload = {"blocks": blocks}
            for url in {u for u in (purchase_url, design_url) if u}:
                try:
                    requests.post(url, json=payload, timeout=10)
                except Exception:
                    pass
        await asyncio.to_thread(_notify)

    return {"created": created, "failed": failed}


@router.get("/draft_orders")
def list_draft_orders(db=Depends(get_database)):
    """List all saved draft purchase orders (excluding items for brevity)."""
    col = db.get_collection(DRAFT_PURCHASE_ORDERS_COLLECTION)
    pipeline = [
        {"$sort": {"created_at": -1}},
        {"$limit": 100},
        {"$lookup": {
            "from": PURCHASE_ORDERS_COLLECTION,
            "localField": "purchaseorder_number",
            "foreignField": "purchaseorder_number",
            "as": "_po",
        }},
        {"$addFields": {
            "po_status": {
                "$let": {
                    "vars": {
                        "fmt": {"$arrayElemAt": ["$_po.order_status_formatted", 0]},
                        "raw": {"$arrayElemAt": ["$_po.order_status", 0]},
                    },
                    "in": {
                        "$cond": {
                            "if": {"$gt": [{"$strLenCP": {"$ifNull": ["$$fmt", ""]}}, 0]},
                            "then": "$$fmt",
                            "else": {
                                "$cond": {
                                    "if": {"$gt": [{"$strLenCP": {"$ifNull": ["$$raw", ""]}}, 0]},
                                    "then": {"$concat": [
                                        {"$toUpper": {"$substrCP": ["$$raw", 0, 1]}},
                                        {"$substrCP": ["$$raw", 1, {"$subtract": [{"$strLenCP": "$$raw"}, 1]}]},
                                    ]},
                                    "else": None,
                                }
                            },
                        }
                    },
                }
            },
        }},
        {"$project": {"_po": 0}},
    ]
    docs = list(col.aggregate(pipeline))
    return serialize_mongo_document(docs)


@router.post("/draft_orders/save")
async def save_draft_order(body: SaveDraftOrderRequest, db=Depends(get_database)):
    """Persist a validated draft order to the draft_purchase_orders collection."""
    def _insert():
        col = db.get_collection(DRAFT_PURCHASE_ORDERS_COLLECTION)
        doc = {
            "description": body.description or "",
            "items": body.items,
            "detected_vendor": body.detected_vendor,
            "date": body.date,
            "item_count": len(body.items),
            "po_created": False,
            "notes": body.notes or "",
            "reference_number": body.reference_number or "",
            "purchaseorder_number": body.purchaseorder_number or "",
            "created_at": datetime.now(),
            "updated_at": datetime.now(),
        }
        result = col.insert_one(doc)
        doc["_id"] = result.inserted_id
        return doc

    doc = await asyncio.to_thread(_insert)
    return JSONResponse(status_code=201, content=serialize_mongo_document(doc))


def _increment_po_number(po_number: str) -> str:
    match = re.search(r'(\d+)$', po_number)
    if match:
        num_str = match.group(1)
        incremented = str(int(num_str) + 1).zfill(len(num_str))
        return po_number[:match.start()] + incremented
    return po_number


@router.get("/draft_orders/last_po_number")
def get_last_po_number(vendor_id: str = Query(...), db=Depends(get_database)):
    """Get the last PO number for a vendor and return the suggested next one."""
    col = db.get_collection(PURCHASE_ORDERS_COLLECTION)
    last_po = col.find_one(
        {"vendor_id": vendor_id},
        {"purchaseorder_number": 1},
        sort=[("created_at", -1)],
    )
    if not last_po or not last_po.get("purchaseorder_number"):
        return {"last_po_number": None, "next_po_number": None}
    last_num = last_po["purchaseorder_number"]
    return {"last_po_number": last_num, "next_po_number": _increment_po_number(last_num)}


@router.delete("/draft_orders/{draft_id}")
async def delete_draft_order(draft_id: str, db=Depends(get_database)):
    """Delete a saved draft order by its MongoDB _id."""
    from bson import ObjectId
    try:
        oid = ObjectId(draft_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid draft_id")

    def _delete():
        col = db.get_collection(DRAFT_PURCHASE_ORDERS_COLLECTION)
        return col.delete_one({"_id": oid})

    result = await asyncio.to_thread(_delete)
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Draft order not found")
    return {"deleted": True}


@router.post("/draft_orders/create_po")
async def create_draft_order_po(
    body: CreateDraftPORequest,
    db=Depends(get_database),
):
    """Create a Zoho Books purchase order from a validated draft order and save it to the purchase_orders collection."""
    try:
        datetime.strptime(body.date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")

    if not body.contact_id:
        raise HTTPException(status_code=400, detail="contact_id must not be empty")
    if not body.items:
        raise HTTPException(status_code=400, detail="items must not be empty")

    def _create():
        token = _get_zoho_token()
        headers = {"Authorization": f"Zoho-oauthtoken {token}"}

        vendor_doc = db.get_collection("vendors").find_one(
            {"contact_id": body.contact_id},
            {
                "currency_id": 1, "gst_treatment": 1,
                "billing_address": 1, "shipping_address": 1,
                "destination_of_supply": 1,
            },
        )
        vd = vendor_doc or {}

        DEFAULT_ACCOUNT_ID = "3220178000000034001"
        DEFAULT_WAREHOUSE_ID = "3220178000000403010"

        line_items = [
            {
                "item_id": it["item_id"],
                "name": it.get("product_name", ""),
                "quantity": it["qty"],
                "rate": it["unit_price"],
                "account_id": it.get("purchase_account_id") or DEFAULT_ACCOUNT_ID,
                "hsn_or_sac": it.get("hsn_or_sac", ""),
                "unit": "pcs",
                "tags": [],
                "item_custom_fields": [],
            }
            for it in body.items
            if it.get("item_id")
        ]

        billing_addr = vd.get("billing_address") or {}
        shipping_addr = vd.get("shipping_address") or {}

        payload: dict = {
            "vendor_id": body.contact_id,
            "date": body.date,
            "location_id": "3220178000143298047",
            "delivery_org_address_id": DEFAULT_WAREHOUSE_ID,
            "template_id": "3220178000000017007",
            "destination_of_supply": vd.get("destination_of_supply") or "MH",
            "gst_treatment": vd.get("gst_treatment", "overseas"),
            "gst_no": "",
            "is_inclusive_tax": False,
            "discount": 0,
            "discount_type": "entity_level",
            "is_discount_before_tax": True,
            "payment_terms": 0,
            "payment_terms_label": "Due on Receipt",
            "contact_persons": [],
            "custom_fields": [],
            "documents": [],
            "line_items": line_items,
        }
        if vd.get("currency_id"):
            payload["currency_id"] = vd["currency_id"]
        if billing_addr.get("address_id"):
            payload["billing_address_id"] = billing_addr["address_id"]
        if shipping_addr.get("address_id"):
            payload["shipping_address_id"] = shipping_addr["address_id"]
        if body.notes:
            payload["notes"] = body.notes
        if body.reference_number:
            payload["reference_number"] = body.reference_number
        if body.purchaseorder_number:
            payload["purchaseorder_number"] = body.purchaseorder_number

        logger.info("Zoho PO payload: %s", payload)
        r = requests.post(
            f"{ZOHO_BOOKS_BASE}/purchaseorders",
            headers=headers,
            json=payload,
            params={"organization_id": ORGANIZATION_ID},
            timeout=30,
        )
        logger.error("Zoho PO response %s: %s", r.status_code, r.text)
        r.raise_for_status()
        data = r.json()

        if data.get("code") != 0:
            raise ValueError(f"Zoho error: {data.get('message', 'Unknown error')}")

        po_doc = data["purchaseorder"]
        now = datetime.now()
        db[PURCHASE_ORDERS_COLLECTION].update_one(
            {"purchaseorder_id": po_doc.get("purchaseorder_id")},
            {
                "$set": {**po_doc, "updated_at": now},
                "$setOnInsert": {"created_at": now},
            },
            upsert=True,
        )

        if body.draft_id:
            try:
                from bson import ObjectId
                db.get_collection(DRAFT_PURCHASE_ORDERS_COLLECTION).update_one(
                    {"_id": ObjectId(body.draft_id)},
                    {"$set": {
                        "po_created": True,
                        "po_number": po_doc.get("purchaseorder_number"),
                        "po_status": po_doc.get("status"),
                        "updated_at": now,
                    }},
                )
            except Exception:
                pass

        return po_doc

    try:
        po_doc = await asyncio.to_thread(_create)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except requests.HTTPError as e:
        body = e.response.text if e.response is not None else ""
        raise HTTPException(status_code=502, detail=f"Zoho API error: {e} — {body}")

    async def _notify_po_created():
        def _fan_out():
            recipients = list(db["purchase_users"].find(
                {"status": "active"},
                {"_id": 1},
            ))
            if not recipients:
                return
            now = datetime.utcnow()
            po_number = po_doc.get("purchaseorder_number", "")
            vendor_name = po_doc.get("vendor_name", "")
            snippet = f"Draft order converted to PO {po_number} — {vendor_name}"
            notif_docs = [
                {
                    "user_id": str(r["_id"]),
                    "source": "purchase_order",
                    "purchaseorder_id": po_doc.get("purchaseorder_id"),
                    "purchaseorder_number": po_number,
                    "vendor_name": vendor_name,
                    "type": "draft_order_po_created",
                    "actor_name": "System",
                    "snippet": snippet,
                    "read": False,
                    "created_at": now,
                }
                for r in recipients
            ]
            db[NOTIFICATIONS_COLLECTION].insert_many(notif_docs)
        try:
            await asyncio.to_thread(_fan_out)
        except Exception as _e:
            logger.warning(f"Draft order PO notification fan-out failed: {_e}")

    asyncio.create_task(_notify_po_created())

    asyncio.create_task(fire_trigger("po_created", {
        "po_number": po_doc.get("purchaseorder_number", ""),
        "vendor_name": po_doc.get("vendor_name", ""),
        "brand": po_doc.get("brand_name", ""),
    }, db))

    return JSONResponse(
        status_code=201,
        content=serialize_mongo_document(
            {
                "purchaseorder_id": po_doc.get("purchaseorder_id"),
                "purchaseorder_number": po_doc.get("purchaseorder_number"),
                "vendor_name": po_doc.get("vendor_name"),
                "date": po_doc.get("date"),
                "status": po_doc.get("status"),
                "total": po_doc.get("total"),
                "currency_code": po_doc.get("currency_code"),
            }
        ),
    )
