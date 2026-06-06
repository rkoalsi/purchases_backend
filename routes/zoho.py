import os
import requests
import pandas as pd
from datetime import datetime, timedelta
import io, logging, math, json, re
import openpyxl
from fastapi import (
    APIRouter,
    File,
    Form,
    Header,
    UploadFile,
    HTTPException,
    status,
    Depends,
    Query,
    BackgroundTasks,
    Body,
)
import threading
from fastapi.responses import JSONResponse, StreamingResponse
import concurrent.futures
from pymongo.errors import PyMongoError
from bson import ObjectId
from ..database import get_database, serialize_mongo_document
from .task_triggers import fire_trigger
from pydantic import BaseModel, validator
from typing import List, Optional

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# Request model
class ReportRequest(BaseModel):
    start_date: str
    end_date: str
    brand: str = ""  # empty = all brands
    exclude_customers: bool = False
    min_quantity: float = None
    report_type: str = "customer"  # "customer" or "transfer_order"


    @validator("start_date", "end_date")
    def validate_date_format(cls, v):
        try:
            datetime.strptime(v, "%Y-%m-%d")
            return v
        except ValueError:
            raise ValueError("Date must be in YYYY-MM-DD format")

    @validator("end_date")
    def validate_date_range(cls, v, values):
        if "start_date" in values:
            start = datetime.strptime(values["start_date"], "%Y-%m-%d")
            end = datetime.strptime(v, "%Y-%m-%d")
            if start > end:
                raise ValueError("End date must be after start date")
        return v


# --- Configuration ---
# Use environment variables for production

PRODUCTS_COLLECTION = "products"
COMPOSITE_COLLECTION = "composite_products"
INVOICES_COLLECTION = "invoices"
PURCHASE_ORDER_COLLECTION = "purchase_orders"

# Standardized excluded customers list - used across multiple endpoints
EXCLUDED_CUSTOMERS_LIST = [
    "(amzb2b) Pupscribe Enterprises Pvt Ltd",
    "Pupscribe Enterprises Private Limited",
    "(OSAMP) Office samples",
    "(PUPEV) PUPSCRIBE EVENTS",
    "(SSAM) Sales samples",
    "(RS) Retail samples",
    "(MKT) Marketing Campaign",
    "(DON) Donations",
    "KIRANAKART (Hyderabad)",
    "KIRANAKART (Gurugram)",
    "KIRANAKART (Chennai)",
    "KIRANAKART (Bhiwandi)",
    "KIRANAKART (Bangalore)",
    "Pupscribe Enterprises Private Limited (Blinkit)",
    "(SPUR) Staff purchase",
]

router = APIRouter()


@router.get("/excluded-customers")
async def get_excluded_customers():
    """
    Get the standardized list of excluded customers.
    This list is used across multiple endpoints for filtering out internal/sample customers.
    """
    return {
        "excluded_customers": EXCLUDED_CUSTOMERS_LIST,
        "count": len(EXCLUDED_CUSTOMERS_LIST),
    }


@router.get("/products")
def get_products(
    page: int = Query(1, ge=1, description="Page number (starts from 1)"),
    limit: int = Query(
        10, ge=1, le=100, description="Number of items per page (max 100)"
    ),
    search: str = Query(None, description="Search term for product name or SKU"),
    category: str = Query(None, description="Filter by category"),
    status: str = Query(None, description="Filter by status (active/inactive)"),
    purchase_status: str = Query(None, description="Filter by purchase_status"),
    brand: str = Query(None, description="Filter by brand"),
    sort_by: str = Query(
        "name", description="Sort by field (name, price, stock, created_date)"
    ),
    sort_order: str = Query(
        "asc", regex="^(asc|desc)$", description="Sort order (asc/desc)"
    ),
):
    """
    Get products with pagination, search, filtering, and sorting capabilities.

    Args:
        page: Page number (1-based)
        limit: Items per page (1-100)
        search: Search in product name, SKU, or description
        category: Filter by category name
        status: Filter by active/inactive status
        sort_by: Field to sort by
        sort_order: Sort order (asc/desc)

    Returns:
        JSON response with products, pagination info, and metadata
    """
    try:
        db = get_database()
        collection = db[PRODUCTS_COLLECTION]

        # Build query filter using $and so each condition is independent
        conditions = []

        if search:
            search_regex = {"$regex": re.escape(search), "$options": "i"}
            conditions.append({"$or": [
                {"name": search_regex},
                {"sku": search_regex},
                {"cf_sku_code": search_regex},
                {"item_id": search_regex},
                {"description": search_regex},
            ]})

        if brand:
            conditions.append({"brand": brand})

        if category:
            conditions.append({"$or": [
                {"category_name": {"$regex": re.escape(category), "$options": "i"}},
                {"item_type": {"$regex": re.escape(category), "$options": "i"}},
            ]})

        if status:
            if status.lower() == "active":
                conditions.append({"$or": [{"status": "active"}, {"is_active": True}]})
            elif status.lower() == "inactive":
                conditions.append({"$or": [{"status": {"$ne": "active"}}, {"is_active": {"$ne": True}}]})

        if purchase_status:
            conditions.append({"purchase_status": purchase_status})

        query_filter = {"$and": conditions} if conditions else {}

        # Sort configuration
        sort_field_mapping = {
            "name": "name",
            "price": "rate",
            "stock": "stock_on_hand",
            "created_date": "created_at",
        }

        actual_sort_field = sort_field_mapping.get(sort_by, "name")
        sort_direction = 1 if sort_order.lower() == "asc" else -1

        # Calculate pagination
        skip = (page - 1) * limit

        # Get total count for pagination
        total_count = collection.count_documents(query_filter)

        # Calculate total pages
        total_pages = math.ceil(total_count / limit) if total_count > 0 else 1

        # Get products with pagination
        cursor = (
            collection.find(query_filter)
            .sort([(actual_sort_field, sort_direction), ("_id", sort_direction)])
            .skip(skip)
            .limit(limit)
        )
        products = list(cursor)

        # Serialize MongoDB documents
        serialized_products = serialize_mongo_document(products)

        # Prepare response
        response_data = {
            "products": serialized_products,
            "pagination": {
                "currentPage": page,
                "totalPages": total_pages,
                "totalProducts": total_count,
                "limit": limit,
                "hasNextPage": page < total_pages,
                "hasPrevPage": page > 1,
                "nextPage": page + 1 if page < total_pages else None,
                "prevPage": page - 1 if page > 1 else None,
            },
            "filters": {
                "search": search,
                "category": category,
                "status": status,
                "sort_by": sort_by,
                "sort_order": sort_order,
            },
            "meta": {
                "timestamp": datetime.now().isoformat(),
                "resultsOnPage": len(serialized_products),
            },
        }

        return JSONResponse(content=response_data)

    except PyMongoError as e:
        logger.info(f"MongoDB Error Getting Products: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database error occurred while fetching products: {str(e)}",
        )
    except HTTPException as e:
        raise e
    except Exception as e:
        logger.info(f"Error Getting Products: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An unexpected error occurred while getting the products: {str(e)}",
        )


PURCHASE_STATUS_OPTIONS = {"active", "inactive", "discontinued until stock lasts"}


class PurchaseStatusUpdate(BaseModel):
    purchase_status: str


@router.patch("/products/{item_id}/purchase-status")
def update_purchase_status(item_id: str, body: PurchaseStatusUpdate):
    """Update the purchase_status field of a product."""
    if body.purchase_status not in PURCHASE_STATUS_OPTIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid purchase_status. Must be one of: {', '.join(PURCHASE_STATUS_OPTIONS)}",
        )
    try:
        db = get_database()
        collection = db[PRODUCTS_COLLECTION]
        result = collection.update_one(
            {"_id": ObjectId(item_id)},
            {"$set": {"purchase_status": body.purchase_status}},
        )
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Product not found")
        return JSONResponse(
            content={"success": True, "purchase_status": body.purchase_status}
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.info(f"Error updating purchase_status: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sku-brand-map")
def get_sku_brand_map():
    """
    Returns a flat {sku: brand} mapping for all products that have a brand set.
    Used by platform item pages (Amazon, Blinkit) for client-side brand filtering.
    """
    try:
        db = get_database()
        collection = db[PRODUCTS_COLLECTION]
        cursor = collection.find(
            {"brand": {"$exists": True, "$ne": ""}},
            {"cf_sku_code": 1, "brand": 1, "_id": 0},
        )
        result = {}
        for doc in cursor:
            sku = doc.get("cf_sku_code")
            brand = doc.get("brand")
            if sku and brand:
                result[str(sku)] = brand
        return JSONResponse(content=result)
    except PyMongoError as e:
        logger.info(f"MongoDB Error Getting SKU Brand Map: {e}")
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        logger.info(f"Error Getting SKU Brand Map: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/composite-products")
def get_composite_products(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=100),
    search: str = Query(None),
):
    """Paginated composite products from the composite_products collection."""
    try:
        db = get_database()
        collection = db[COMPOSITE_COLLECTION]

        query_filter = {}
        if search:
            rx = {"$regex": search, "$options": "i"}
            query_filter["$or"] = [
                {"name": rx},
                {"sku_code": rx},
                {"composite_item_id": rx},
            ]

        total_count = collection.count_documents(query_filter)
        total_pages = math.ceil(total_count / limit) if total_count > 0 else 1
        skip = (page - 1) * limit

        items = serialize_mongo_document(
            list(collection.find(query_filter).sort("name", 1).skip(skip).limit(limit))
        )

        return JSONResponse(
            content={
                "items": items,
                "pagination": {
                    "currentPage": page,
                    "totalPages": total_pages,
                    "totalItems": total_count,
                    "limit": limit,
                    "hasNextPage": page < total_pages,
                    "hasPrevPage": page > 1,
                },
            }
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/products/summary")
def get_products_summary():
    """
    Get a summary of products including total count, active/inactive counts, and categories.

    Returns:
        JSON response with product summary statistics
    """
    try:
        db = get_database()
        collection = db[PRODUCTS_COLLECTION]

        # Get total count
        total_count = collection.count_documents({})

        # Get active/inactive counts
        active_count = collection.count_documents({"status": "active"})

        inactive_count = total_count - active_count

        # Get categories
        categories_pipeline = [
            {
                "$group": {
                    "_id": {
                        "$ifNull": [
                            "$category",
                            {"$ifNull": ["$item_type", "Uncategorized"]},
                        ]
                    },
                    "count": {"$sum": 1},
                }
            },
            {"$sort": {"count": -1}},
        ]

        # Get stock status
        low_stock_count = collection.count_documents({"stock": {"$lt": 10}})

        out_of_stock_count = collection.count_documents({"stock": 0})

        summary_data = {
            "totalProducts": total_count,
            "activeProducts": active_count,
            "inactiveProducts": inactive_count,
            "lowStockProducts": low_stock_count,
            "outOfStockProducts": out_of_stock_count,
            "timestamp": datetime.now().isoformat(),
        }

        return JSONResponse(content=summary_data)

    except Exception as e:
        logger.info(f"Error Getting Products Summary: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An error occurred getting the products summary: {str(e)}",
        )


@router.get("/sales")
def get_sales(
    page: int = Query(1, ge=1, description="Page number (starts from 1)"),
    limit: int = Query(
        10, ge=1, le=100, description="Number of items per page (max 100)"
    ),
    search: str = Query(
        None, description="Search term for invoice number or customer name"
    ),
    status: str = Query(
        None, description="Filter by status (paid, pending, overdue, etc.)"
    ),
    date_from: str = Query(None, description="Start date filter (YYYY-MM-DD)"),
    date_to: str = Query(None, description="End date filter (YYYY-MM-DD)"),
    sort_by: str = Query(
        "invoice_date",
        description="Sort by field (invoice_date, total, customer_name, status)",
    ),
    sort_order: str = Query(
        "desc", regex="^(asc|desc)$", description="Sort order (asc/desc)"
    ),
):
    """
    Get sales/invoices with pagination, search, filtering, and sorting capabilities.

    Args:
        page: Page number (1-based)
        limit: Items per page (1-100)
        search: Search in invoice number or customer name
        status: Filter by invoice status
        date_from: Filter invoices from this date
        date_to: Filter invoices up to this date
        sort_by: Field to sort by
        sort_order: Sort order (asc/desc)

    Returns:
        JSON response with sales data, pagination info, and metadata
    """
    try:
        db = get_database()
        collection = db[INVOICES_COLLECTION]

        # Build query filter
        query_filter = {}

        # Search functionality - search in invoice number and customer name
        if search:
            search_regex = {"$regex": search, "$options": "i"}
            query_filter["$or"] = [
                {"invoice_number": search_regex},
                {"number": search_regex},
                {"customer_name": search_regex},
                {"customer.display_name": search_regex},
                {"customer.name": search_regex},
            ]

        # Status filter
        if status:
            query_filter["status"] = {"$regex": status, "$options": "i"}

        # Date range filter
        if date_from or date_to:
            date_filter = {}
            if date_from:
                try:
                    from_date = datetime.strptime(date_from, "%Y-%m-%d")
                    date_filter["$gte"] = from_date
                except ValueError:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Invalid date_from format. Use YYYY-MM-DD",
                    )

            if date_to:
                try:
                    to_date = datetime.strptime(date_to, "%Y-%m-%d")
                    # Add 1 day and subtract 1 second to include the entire end date
                    to_date = to_date + timedelta(days=1) - timedelta(seconds=1)
                    date_filter["$lt"] = to_date
                except ValueError:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Invalid date_to format. Use YYYY-MM-DD",
                    )

            # Apply date filter to possible date fields
            if date_filter:
                query_filter["$or"] = query_filter.get("$or", []) + [
                    {"invoice_date": date_filter},
                    {"date": date_filter},
                    {"created_time": date_filter},
                ]

        # Sort configuration
        sort_field_mapping = {
            "invoice_date": "invoice_date",
            "date": "date",
            "total": "total",
            "amount": "total",
            "customer_name": "customer_name",
            "status": "status",
            "created_date": "created_time",
        }

        actual_sort_field = sort_field_mapping.get(sort_by, "invoice_date")
        sort_direction = 1 if sort_order.lower() == "asc" else -1

        # Calculate pagination
        skip = (page - 1) * limit

        # Get total count for pagination
        total_count = collection.count_documents(query_filter)

        # Calculate total pages
        total_pages = math.ceil(total_count / limit) if total_count > 0 else 1

        # Get sales with pagination
        cursor = (
            collection.find(query_filter)
            .sort(actual_sort_field, sort_direction)
            .skip(skip)
            .limit(limit)
        )
        sales = list(cursor)

        # Process sales data to ensure consistent field names
        processed_sales = []
        for sale in sales:
            processed_sale = sale.copy()

            # Normalize customer name
            if not processed_sale.get("customer_name"):
                if processed_sale.get("customer", {}).get("display_name"):
                    processed_sale["customer_name"] = processed_sale["customer"][
                        "display_name"
                    ]
                elif processed_sale.get("customer", {}).get("name"):
                    processed_sale["customer_name"] = processed_sale["customer"]["name"]

            # Normalize customer email
            if not processed_sale.get("customer_email"):
                if processed_sale.get("customer", {}).get("email"):
                    processed_sale["customer_email"] = processed_sale["customer"][
                        "email"
                    ]

            # Normalize date fields
            if not processed_sale.get("invoice_date") and processed_sale.get("date"):
                processed_sale["invoice_date"] = processed_sale["date"]

            # Normalize amount/total
            if not processed_sale.get("total") and processed_sale.get("amount"):
                processed_sale["total"] = processed_sale["amount"]

            # Calculate days overdue if applicable
            if processed_sale.get("due_date") and processed_sale.get(
                "status", ""
            ).lower() not in ["paid", "sent"]:
                try:
                    due_date = processed_sale["due_date"]
                    if isinstance(due_date, str):
                        due_date = datetime.strptime(due_date.split("T")[0], "%Y-%m-%d")

                    days_diff = (datetime.now() - due_date).days
                    if days_diff > 0:
                        processed_sale["days_overdue"] = days_diff
                    else:
                        processed_sale["days_overdue"] = 0
                except:
                    processed_sale["days_overdue"] = 0
            else:
                processed_sale["days_overdue"] = 0

            processed_sales.append(processed_sale)

        # Serialize MongoDB documents
        serialized_sales = serialize_mongo_document(processed_sales)

        # Calculate summary statistics
        total_amount = sum(sale.get("total", 0) for sale in processed_sales)
        paid_count = len(
            [
                sale
                for sale in processed_sales
                if sale.get("status", "").lower() == "paid"
            ]
        )
        pending_count = len(
            [
                sale
                for sale in processed_sales
                if sale.get("status", "").lower() == "pending"
            ]
        )
        overdue_count = len(
            [sale for sale in processed_sales if sale.get("days_overdue", 0) > 0]
        )

        # Prepare response
        response_data = {
            "sales": serialized_sales,
            "pagination": {
                "currentPage": page,
                "totalPages": total_pages,
                "totalSales": total_count,
                "limit": limit,
                "hasNextPage": page < total_pages,
                "hasPrevPage": page > 1,
                "nextPage": page + 1 if page < total_pages else None,
                "prevPage": page - 1 if page > 1 else None,
            },
            "filters": {
                "search": search,
                "status": status,
                "date_from": date_from,
                "date_to": date_to,
                "sort_by": sort_by,
                "sort_order": sort_order,
            },
            "summary": {
                "totalAmount": total_amount,
                "paidCount": paid_count,
                "pendingCount": pending_count,
                "overdueCount": overdue_count,
                "currentPageAmount": sum(
                    sale.get("total", 0) for sale in serialized_sales
                ),
            },
            "meta": {
                "timestamp": datetime.now().isoformat(),
                "resultsOnPage": len(serialized_sales),
            },
        }

        return JSONResponse(content=response_data)

    except PyMongoError as e:
        logger.info(f"MongoDB Error Getting Sales: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database error occurred while fetching sales: {str(e)}",
        )
    except HTTPException as e:
        raise e
    except Exception as e:
        logger.info(f"Error Getting Sales: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An unexpected error occurred while getting the sales: {str(e)}",
        )


_item_name_cache = {}
_cache_ttl = 300  # 5 minutes
_cache_lock = threading.Lock()


def get_item_names_by_brand(db, brand: str) -> tuple:
    """
    Cached version of item name lookup with TTL for better performance.
    Pass empty string to get all item names across all brands.
    Returns (item_names, sku_to_name) where sku_to_name maps EAN sku → current product name.
    Used to match invoice line items by either name or sku (handles renamed products).
    """
    current_time = datetime.now().timestamp()

    cache_key = brand if brand else "__all__"
    with _cache_lock:
        if cache_key in _item_name_cache:
            cache_entry = _item_name_cache[cache_key]
            if current_time - cache_entry["timestamp"] < _cache_ttl:
                logger.info(f"Using cached item names for brand '{brand or 'all'}'")
                return cache_entry["item_names"], cache_entry["sku_to_name"]

    # Fetch from database
    try:
        products_collection = db.get_collection(PRODUCTS_COLLECTION)

        if brand and brand.lower() == "petfest":
            match_stage = {"$match": {"brand": {"$in": ["Dogfest", "Catfest"]}}}
        elif brand:
            match_stage = {"$match": {"brand": brand}}
        else:
            match_stage = {"$match": {}}
        pipeline = [
            match_stage,
            {"$project": {"name": 1, "sku": 1, "_id": 0}},
        ]

        result = list(products_collection.aggregate(pipeline))

        item_names = [p["name"] for p in result if p.get("name")]
        sku_to_name = {
            str(p["sku"]): p["name"]
            for p in result
            if p.get("sku") and p.get("name")
        }

        with _cache_lock:
            _item_name_cache[cache_key] = {
                "item_names": item_names,
                "sku_to_name": sku_to_name,
                "timestamp": current_time,
            }

        logger.info(
            f"Found {len(item_names)} item names, {len(sku_to_name)} skus for brand '{brand or 'all'}'"
        )
        return item_names, sku_to_name

    except Exception as e:
        logger.error(f"Error getting item names for brand '{brand}': {e}")
        raise HTTPException(
            status_code=500, detail=f"Error retrieving item names for brand '{brand}'"
        )


def query_invoices_for_item_names(
    db,
    item_names: List[str],
    sku_to_name: dict,
    start_date: datetime,
    end_date: datetime,
    exclude_customers: bool,
    report_type: str = "customer",
    include_brand: bool = False,
) -> List[dict]:
    """
    Query invoices by matching item names or EAN sku.
    Matching by sku handles products that were renamed after invoices were created.
    report_type="transfer_order" filters to only internal/excluded customers.
    report_type="customer" with exclude_customers=True filters them out.
    """
    try:
        invoices_collection = db.get_collection(INVOICES_COLLECTION)
        products_collection = db.get_collection(PRODUCTS_COLLECTION)

        sku_list = list(sku_to_name.keys())

        base_match: dict = {
            "$expr": {
                "$and": [
                    {"$gte": [{"$toDate": "$date"}, start_date]},
                    {"$lte": [{"$toDate": "$date"}, end_date]},
                ]
            },
            "status": {"$nin": ["draft", "void"]},
            "line_items": {
                "$elemMatch": {
                    "$or": [
                        {"name": {"$in": item_names}},
                        {"sku": {"$in": sku_list}},
                    ]
                }
            },
        }

        if report_type == "transfer_order":
            base_match["$or"] = [
                {"customer_name": {"$in": []}},
                {"customer_name": {"$regex": "pupscribe", "$options": "i"}},
            ]
        elif exclude_customers:
            base_match["$nor"] = [
                {"customer_name": {"$in": EXCLUDED_CUSTOMERS_LIST}},
                {"customer_name": {"$regex": "pupscribe", "$options": "i"}},
            ]

        pipeline = [
            {"$match": base_match},
            {
                "$addFields": {
                    "_matchingItems": {
                        "$filter": {
                            "input": "$line_items",
                            "as": "item",
                            "cond": {
                                "$or": [
                                    {"$in": ["$$item.name", item_names]},
                                    {"$in": [{"$toString": "$$item.sku"}, sku_list]},
                                ]
                            },
                        }
                    }
                }
            },
            {"$unwind": "$_matchingItems"},
            {
                "$project": {
                    "_id": 1,
                    "invoice_number": 1,
                    "customer_name": 1,
                    "invoice_status": "$status",
                    "created_date_str": "$date",
                    "created_at": 1,
                    "quantity": "$_matchingItems.quantity",
                    "item_total": "$_matchingItems.item_total",
                    "item_name": "$_matchingItems.name",
                    "item_sku": {"$toString": "$_matchingItems.sku"},
                    "sku_code": {
                        "$ifNull": [
                            "$_matchingItems.sku",
                            {
                                "$arrayElemAt": [
                                    "$_matchingItems.item_custom_fields.value",
                                    0,
                                ]
                            },
                        ]
                    },
                }
            },
        ]

        logger.info(
            f"Executing item name aggregation for {len(item_names)} items, {len(sku_list)} skus (report_type={report_type})"
        )
        results = list(invoices_collection.aggregate(pipeline, allowDiskUse=True))
        logger.info(f"Found {len(results)} total matching line items")

        # Batch-load product info (brand, status, purchase_status) by item name.
        # Also build sku→product map for items whose names changed since invoice was created.
        result_item_names = list(
            {doc.get("item_name") for doc in results if doc.get("item_name")}
        )
        product_docs = list(
            products_collection.find(
                {"name": {"$in": result_item_names}},
                {"name": 1, "brand": 1, "status": 1, "purchase_status": 1, "_id": 0},
            )
        )
        product_info = {p["name"]: p for p in product_docs}

        all_results = []
        for doc in results:
            raw_item_name = doc.get("item_name", "N/A")
            item_sku = doc.get("item_sku", "")
            # Normalize to current product name if invoice used an old name
            item_name = sku_to_name.get(item_sku, raw_item_name)
            product = product_info.get(item_name) or product_info.get(raw_item_name, {})
            row = {
                "SKU Code": doc.get("sku_code", "N/A"),
                "Item": item_name,
                "Customer": doc.get("customer_name", "N/A"),
                "Quantity": doc.get("quantity", 0),
                "Total Amount": doc.get("item_total", 0),
                "Status": product.get("status", "N/A"),
                "Purchase Status": product.get("purchase_status", "N/A"),
                "Invoice Number": doc.get("invoice_number", "N/A"),
                "Invoice ID": str(doc["_id"]),
                "Created Date": doc.get("created_date_str", "N/A"),
            }
            if include_brand:
                row["Brand"] = product.get("brand", "N/A")
            all_results.append(row)

        return all_results

    except Exception as e:
        logger.error(f"Error in item name invoice query: {e}")
        raise HTTPException(status_code=500, detail="Error querying invoice data")


def build_summary_data(
    invoice_data: List[dict],
) -> List[dict]:
    """
    Build the summary: one row per (SKU, Customer), all combinations included.
    """
    # Aggregate by (SKU, Customer)
    agg: dict = {}
    for row in invoice_data:
        sku = row.get("SKU Code") or ""
        customer = row.get("Customer") or ""
        key = (sku, customer)
        if key not in agg:
            agg[key] = {
                "SKU Code": sku,
                "Item": row.get("Item", ""),
                "Customer": customer,
                "Total Quantity": 0.0,
                "Total Amount": 0.0,
                "_invoices": set(),
                "Status": row.get("Status", ""),
                "Purchase Status": row.get("Purchase Status", ""),
            }
        agg[key]["Total Quantity"] += float(row.get("Quantity") or 0)
        agg[key]["Total Amount"] += float(row.get("Total Amount") or 0)
        agg[key]["_invoices"].add(row.get("Invoice Number"))

    summary_rows = []
    for key, entry in agg.items():
        summary_rows.append(
            {
                "SKU Code": entry["SKU Code"],
                "Item": entry["Item"],
                "Customer": entry["Customer"],
                "Total Quantity": round(entry["Total Quantity"], 2),
                "Total Amount": round(entry["Total Amount"], 2),
                "Total no. of orders in current period": len(entry["_invoices"]),
                "Status": entry["Status"],
                "Purchase Status": entry["Purchase Status"],
                "Remark": "",
                "Sales Confirmation - Reorder Expected?": "",
            }
        )

    # Sort by Total Quantity descending
    summary_rows.sort(key=lambda r: r["Total Quantity"], reverse=True)
    return summary_rows


def create_excel_file(data: List[dict], summary_data: List[dict] = None) -> io.BytesIO:
    """
    Create Excel file with parallel processing for large datasets.
    """
    try:
        if not data:
            raise HTTPException(
                status_code=404, detail="No data found for the specified criteria"
            )

        logger.info(f"Creating Excel file with {len(data)} records")

        # For large datasets, process in parallel
        if len(data) > 10000:
            chunk_size = len(data) // 4
            chunks = [data[i : i + chunk_size] for i in range(0, len(data), chunk_size)]
            with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
                df_chunks = list(executor.map(pd.DataFrame, chunks))
            df = pd.concat(df_chunks, ignore_index=True)
        else:
            df = pd.DataFrame(data)

        # Enforce column order: SKU Code first, then rest
        desired_order = [
            "SKU Code",
            "Item",
            "Customer",
            "Brand",
            "Quantity",
            "Total Amount",
            "Status",
            "Purchase Status",
            "Invoice Number",
            "Invoice ID",
            "Created Date",
        ]
        ordered_cols = [c for c in desired_order if c in df.columns]
        extra_cols = [c for c in df.columns if c not in desired_order]
        df = df[ordered_cols + extra_cols]

        # Create Excel file in memory
        excel_buffer = io.BytesIO()

        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Invoice Report", index=False)

            workbook = writer.book
            worksheet = writer.sheets["Invoice Report"]

            for column in df:
                column_length = max(df[column].astype(str).map(len).max(), len(column))
                col_idx = df.columns.get_loc(column)
                worksheet.column_dimensions[chr(65 + col_idx)].width = min(
                    column_length + 2, 50
                )

            # Summary / reorder-trigger sheet
            if summary_data:
                from openpyxl.utils import get_column_letter
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name="Summary Report", index=False)
                ws_summary = writer.sheets["Summary Report"]
                for col_idx, column in enumerate(summary_df.columns, start=1):
                    col_letter = get_column_letter(col_idx)
                    col_width = max(
                        summary_df[column].astype(str).map(len).max(),
                        len(column),
                    )
                    ws_summary.column_dimensions[col_letter].width = min(col_width + 2, 60)

        excel_buffer.seek(0)
        logger.info("Excel file created successfully")
        return excel_buffer

    except Exception as e:
        logger.error(f"Error creating Excel file: {e}")
        raise HTTPException(status_code=500, detail="Error generating Excel file")


@router.post("/generate-invoice-report")
def generate_invoice_report(
    request: ReportRequest, background_tasks: BackgroundTasks, db=Depends(get_database)
):
    try:
        start_time = datetime.now()
        logger.info(
            f"Starting report generation for brand: {request.brand}, date range: {request.start_date} to {request.end_date}, min_quantity: {request.min_quantity}"
        )

        # Parse dates
        start_date = datetime.strptime(request.start_date, "%Y-%m-%d")
        end_date = datetime.strptime(request.end_date, "%Y-%m-%d")

        # STEP 1: Get item names (empty brand = all brands)
        item_name_start = datetime.now()
        item_names, sku_to_name = get_item_names_by_brand(db, request.brand)
        item_name_duration = (datetime.now() - item_name_start).total_seconds()
        logger.info(f"Item name lookup took {item_name_duration:.2f} seconds")

        if not item_names:
            label = request.brand or "all brands"
            raise HTTPException(status_code=404, detail=f"No items found for '{label}'")

        include_brand = not request.brand  # show Brand column when all brands selected

        # STEP 2: Query invoices by item names or sku
        query_start = datetime.now()
        invoice_data = query_invoices_for_item_names(
            db,
            item_names,
            sku_to_name,
            start_date,
            end_date,
            request.exclude_customers,
            request.report_type,
            include_brand,
        )
        query_duration = (datetime.now() - query_start).total_seconds()
        logger.info(f"Invoice query took {query_duration:.2f} seconds")

        if not invoice_data:
            raise HTTPException(
                status_code=404, detail="No invoices found for the specified criteria"
            )

        summary_data = build_summary_data(invoice_data)
        logger.info(f"Summary sheet: {len(summary_data)} rows")

        # Filter by min_quantity: keep only rows where quantity >= threshold
        if request.min_quantity is not None and request.min_quantity > 0:
            before_count = len(invoice_data)
            invoice_data = [
                row
                for row in invoice_data
                if float(row.get("Quantity", 0)) >= request.min_quantity
            ]
            logger.info(
                f"Min quantity filter ({request.min_quantity}): {before_count} -> {len(invoice_data)} rows"
            )

            if not invoice_data:
                raise HTTPException(
                    status_code=404,
                    detail=f"No line items found with quantity >= {request.min_quantity}",
                )

        # STEP 3: Create Excel file (parallel processing for large datasets)
        excel_start = datetime.now()
        excel_file = create_excel_file(invoice_data, summary_data)
        excel_duration = (datetime.now() - excel_start).total_seconds()
        logger.info(f"Excel generation took {excel_duration:.2f} seconds")

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        brand_label = request.brand or "all_brands"
        report_label = (
            "transfer_order" if request.report_type == "transfer_order" else "invoice"
        )
        filename = f"{report_label}_report_{brand_label}_{request.start_date}_to_{request.end_date}_{timestamp}.xlsx"

        total_duration = (datetime.now() - start_time).total_seconds()
        logger.info(
            f"Report generated successfully in {total_duration:.2f} seconds with {len(invoice_data)} records"
        )

        # Return Excel file as streaming response
        return StreamingResponse(
            io.BytesIO(excel_file.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Unexpected error generating report: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")


_BOOKS_URL = os.getenv("BOOKS_URL")
_BOOKS_CLIENT_ID = os.getenv("CLIENT_ID")
_BOOKS_CLIENT_SECRET = os.getenv("CLIENT_SECRET")
_BOOKS_REFRESH_TOKEN = os.getenv("BOOKS_REFRESH_TOKEN")
_BOOKS_ORGANIZATION_ID = os.getenv("ORGANIZATION_ID", "776755316")
_ZOHO_BOOKS_BASE = "https://books.zoho.com/api/v3"


def _get_zoho_books_token() -> str:
    url = _BOOKS_URL.format(
        clientId=_BOOKS_CLIENT_ID,
        clientSecret=_BOOKS_CLIENT_SECRET,
        grantType="refresh_token",
        books_refresh_token=_BOOKS_REFRESH_TOKEN,
    )
    r = requests.post(url, timeout=30)
    r.raise_for_status()
    return r.json()["access_token"]


def _fetch_pl_data_from_zoho(from_date: str, to_date: str) -> dict:
    """
    Fetch horizontal P&L from Zoho Books and extract specific account totals.
    Returns dict with: other_charges, shipping_charges, post_supply_discount
    """
    token = _get_zoho_books_token()
    headers = {"Authorization": f"Zoho-oauthtoken {token}"}
    params = {
        "organization_id": _BOOKS_ORGANIZATION_ID,
        "from_date": from_date,
        "to_date": to_date,
        "filter_by": "TransactionDate.CustomDate",
        "response_option": "1",
    }
    r = requests.get(
        f"{_ZOHO_BOOKS_BASE}/reports/horizontalprofitandloss",
        headers=headers,
        params=params,
        timeout=30,
    )
    r.raise_for_status()
    data = r.json()

    if data.get("code") != 0:
        raise ValueError(f"Zoho Books P&L API error: {data.get('message')}")

    # Map exact Zoho account names (case-insensitive) to result keys
    target_names = {
        "other charges (sales)": "other_charges",
        "shipping charges (sales)": "shipping_charges",
        "post supply discount": "post_supply_discount",
    }
    result = {k: 0.0 for k in target_names.values()}

    def _walk(obj):
        if isinstance(obj, dict):
            name = (obj.get("name") or obj.get("total_label") or "").lower().strip()
            if name in target_names:
                val = obj.get("total") or 0
                try:
                    result[target_names[name]] = float(str(val).replace(",", ""))
                except (TypeError, ValueError):
                    pass
            for v in obj.values():
                _walk(v)
        elif isinstance(obj, list):
            for item in obj:
                _walk(item)

    _walk(data)
    logger.info("Zoho Books P&L charges fetched: %s", result)
    return result


class BrandSalesReportRequest(BaseModel):
    start_date: str
    end_date: str

    @validator("start_date", "end_date")
    def validate_date_format(cls, v):
        try:
            datetime.strptime(v, "%Y-%m-%d")
            return v
        except ValueError:
            raise ValueError("Date must be in YYYY-MM-DD format")


def _query_line_items_for_brand_report(
    invoices_col, start_date: datetime, end_date: datetime, filter_type: str
) -> list:
    """
    filter_type: 'all' | 'transfer_order' | 'customer_only'
    Returns list of {item_name, item_sku, item_total}.
    """
    match: dict = {
        "$expr": {
            "$and": [
                {"$gte": [{"$toDate": "$date"}, start_date]},
                {"$lte": [{"$toDate": "$date"}, end_date]},
            ]
        },
        "status": {"$nin": ["draft", "void"]},
    }
    if filter_type == "transfer_order":
        match["customer_name"] = {"$regex": "pupscribe enterprises", "$options": "i"}
    elif filter_type == "customer_only":
        match["$nor"] = [
            {"customer_name": {"$in": EXCLUDED_CUSTOMERS_LIST}},
            {"customer_name": {"$regex": "pupscribe", "$options": "i"}},
        ]

    pipeline = [
        {"$match": match},
        {"$unwind": "$line_items"},
        {"$match": {"line_items.name": {"$exists": True, "$nin": [None, ""]}}},
        {
            "$project": {
                "_id": 0,
                "item_name": "$line_items.name",
                "item_sku": {"$toString": "$line_items.sku"},
                "item_total": "$line_items.item_total",
            }
        },
    ]
    return list(invoices_col.aggregate(pipeline, allowDiskUse=True))


def _query_credit_notes_for_brand_report(
    credit_notes_col, start_date: datetime, end_date: datetime, filter_type: str
) -> list:
    """
    filter_type: 'all' | 'customer_only' | 'transfer_order'
    'all'            — all credit notes excluding transfer order customers (matches Zoho P&L behaviour)
    'customer_only'  — credit notes excluding transfer orders AND excluded customers
    'transfer_order' — credit notes only for pupscribe-regex customers (mirrors invoice transfer_order filter)
    Returns list of {item_name, item_sku, item_total} for credit notes in the period.
    credit_notes stores date as ISODate or string — both are handled.
    """
    cn_end = end_date.replace(hour=23, minute=59, second=59)
    start_str = start_date.strftime("%Y-%m-%d")
    end_str = end_date.strftime("%Y-%m-%d")
    match: dict = {
        "$or": [
            {"date": {"$gte": start_str, "$lte": end_str}},
            {"date": {"$gte": start_date, "$lte": cn_end}},
        ],
        "status": {"$nin": ["draft", "void"]},
    }
    if filter_type == "transfer_order":
        match["customer_name"] = {"$regex": "pupscribe enterprises", "$options": "i"}
    elif filter_type == "customer_only":
        match["$nor"] = [
            {"customer_name": {"$in": EXCLUDED_CUSTOMERS_LIST}},
            {"customer_name": {"$regex": "pupscribe", "$options": "i"}},
        ]
    # 'all': no customer filter — include all CNs to match Zoho P&L Total Operating Income

    pipeline = [
        {"$match": match},
        {"$unwind": "$line_items"},
        {"$match": {"line_items.name": {"$exists": True, "$nin": [None, ""]}}},
        {
            "$project": {
                "_id": 0,
                "item_name": "$line_items.name",
                "item_sku": {"$toString": "$line_items.sku"},
                "item_total": "$line_items.item_total",
            }
        },
    ]
    return list(credit_notes_col.aggregate(pipeline, allowDiskUse=True))


@router.post("/generate-brand-sales-report")
def generate_brand_sales_report(
    request: BrandSalesReportRequest, db=Depends(get_database)
):
    try:
        import calendar as _cal
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from collections import defaultdict

        start_date = datetime.strptime(request.start_date, "%Y-%m-%d")
        end_date = datetime.strptime(request.end_date, "%Y-%m-%d")
        month_label = start_date.strftime("%B %Y")

        # ── Auto-compute previous calendar month ─────────────────────────
        if start_date.month == 1:
            prev_month_num, prev_year = 12, start_date.year - 1
        else:
            prev_month_num, prev_year = start_date.month - 1, start_date.year

        prev_start = start_date.replace(year=prev_year, month=prev_month_num, day=1)
        last_day_prev = _cal.monthrange(prev_year, prev_month_num)[1]
        prev_end = prev_start.replace(day=last_day_prev)
        prev_month_label = prev_start.strftime("%B %Y")
        prev_start_str = prev_start.strftime("%Y-%m-%d")
        prev_end_str = prev_end.strftime("%Y-%m-%d")

        # Fetch P&L charges from Zoho Books for both months
        pl_curr = _fetch_pl_data_from_zoho(request.start_date, request.end_date)
        pl_prev = _fetch_pl_data_from_zoho(prev_start_str, prev_end_str)

        invoices_col = db.get_collection(INVOICES_COLLECTION)
        credit_notes_col = db.get_collection("credit_notes")
        products_col = db.get_collection(PRODUCTS_COLLECTION)

        # Load product catalog: name/sku → {brand, category, sub_category}
        products = list(
            products_col.find(
                {},
                {
                    "name": 1,
                    "sku": 1,
                    "brand": 1,
                    "category": 1,
                    "sub_category": 1,
                    "_id": 0,
                },
            )
        )
        name_to_prod = {p["name"]: p for p in products if p.get("name")}
        sku_to_prod = {str(p["sku"]): p for p in products if p.get("sku")}

        def get_prod(item_name, item_sku):
            return sku_to_prod.get(item_sku) or name_to_prod.get(item_name) or {}

        def _aggregate_month(s: datetime, e: datetime):
            """Return (all_by_brand, transfer_by_brand, customer_by_brand, bwc, bwsc)."""
            _all = _query_line_items_for_brand_report(invoices_col, s, e, "all")
            _tr  = _query_line_items_for_brand_report(invoices_col, s, e, "transfer_order")
            _cu  = _query_line_items_for_brand_report(invoices_col, s, e, "customer_only")

            # Credit notes reduce revenue — subtract them to match Zoho's Total Operating Income.
            _cn_all = _query_credit_notes_for_brand_report(credit_notes_col, s, e, "all")
            _cn_tr  = _query_credit_notes_for_brand_report(credit_notes_col, s, e, "transfer_order")
            _cn_cu  = _query_credit_notes_for_brand_report(credit_notes_col, s, e, "customer_only")

            by_all = defaultdict(float)
            by_tr  = defaultdict(float)
            by_cu  = defaultdict(float)
            _bwc   = defaultdict(lambda: defaultdict(float))
            _bwsc  = defaultdict(lambda: defaultdict(float))

            for item in _all:
                by_all[get_prod(item["item_name"], item["item_sku"]).get("brand", "Unknown")] += item.get("item_total") or 0
            for item in _cn_all:
                by_all[get_prod(item["item_name"], item["item_sku"]).get("brand", "Unknown")] -= item.get("item_total") or 0

            for item in _tr:
                by_tr[get_prod(item["item_name"], item["item_sku"]).get("brand", "Unknown")] += item.get("item_total") or 0
            for item in _cn_tr:
                by_tr[get_prod(item["item_name"], item["item_sku"]).get("brand", "Unknown")] -= item.get("item_total") or 0

            for item in _cu:
                prod    = get_prod(item["item_name"], item["item_sku"])
                brand   = prod.get("brand", "Unknown")
                amt     = item.get("item_total") or 0
                by_cu[brand] += amt
            for item in _cn_cu:
                prod    = get_prod(item["item_name"], item["item_sku"])
                brand   = prod.get("brand", "Unknown")
                amt     = item.get("item_total") or 0
                by_cu[brand] -= amt

            # BWC/BWSC use by_all - by_tr logic so Grand Total matches BWS Pure Sales (=B-C)
            for item in _all:
                prod    = get_prod(item["item_name"], item["item_sku"])
                brand   = prod.get("brand", "Unknown")
                cat     = prod.get("category") or "Uncategorized"
                sub_cat = prod.get("sub_category") or "Uncategorized"
                amt     = item.get("item_total") or 0
                _bwc[brand][cat]      += amt
                _bwsc[brand][sub_cat] += amt
            for item in _tr:
                prod    = get_prod(item["item_name"], item["item_sku"])
                brand   = prod.get("brand", "Unknown")
                cat     = prod.get("category") or "Uncategorized"
                sub_cat = prod.get("sub_category") or "Uncategorized"
                amt     = item.get("item_total") or 0
                _bwc[brand][cat]      -= amt
                _bwsc[brand][sub_cat] -= amt
            for item in _cn_all:
                prod    = get_prod(item["item_name"], item["item_sku"])
                brand   = prod.get("brand", "Unknown")
                cat     = prod.get("category") or "Uncategorized"
                sub_cat = prod.get("sub_category") or "Uncategorized"
                amt     = item.get("item_total") or 0
                _bwc[brand][cat]      -= amt
                _bwsc[brand][sub_cat] -= amt
            for item in _cn_tr:
                prod    = get_prod(item["item_name"], item["item_sku"])
                brand   = prod.get("brand", "Unknown")
                cat     = prod.get("category") or "Uncategorized"
                sub_cat = prod.get("sub_category") or "Uncategorized"
                amt     = item.get("item_total") or 0
                _bwc[brand][cat]      += amt
                _bwsc[brand][sub_cat] += amt

            return by_all, by_tr, by_cu, _bwc, _bwsc

        # Aggregate current and previous months
        curr_all, curr_tr, curr_cu, curr_bwc, curr_bwsc = _aggregate_month(start_date, end_date)
        prev_all, prev_tr, prev_cu, prev_bwc, prev_bwsc = _aggregate_month(prev_start, prev_end)

        # Union of all brands across both months
        all_brands = sorted(
            set(curr_all) | set(curr_tr) | set(curr_bwc) | set(curr_bwsc) |
            set(prev_all) | set(prev_tr) | set(prev_bwc) | set(prev_bwsc)
        )

        # ── Styles ──────────────────────────────────────────────────────────
        curr_header_fill = PatternFill("solid", fgColor="4472C4")   # blue  – current month
        prev_header_fill = PatternFill("solid", fgColor="70AD47")   # green – previous month
        header_font  = Font(bold=True, color="FFFFFF")
        total_fill   = PatternFill("solid", fgColor="D9E1F2")
        total_font   = Font(bold=True)
        title_font   = Font(bold=True, size=12)
        green_fill   = PatternFill("solid", fgColor="C6EFCE")
        red_fill     = PatternFill("solid", fgColor="FFC7CE")
        thin         = Side(style="thin")
        thin_border  = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_header(cell, fill=None):
            cell.font      = header_font
            cell.fill      = fill or curr_header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border    = thin_border

        def style_total(cell):
            cell.font   = total_font
            cell.fill   = total_fill
            cell.border = thin_border

        wb = Workbook()

        # ════════════════════════════════════════════════════════════════════
        # Helper: write one BWS block starting at `row_offset`.
        # Returns the row number of the first blank row after the block.
        # ════════════════════════════════════════════════════════════════════
        def _write_bws_block(ws, row_offset: int, by_all, by_tr, by_cu, pl, label: str, hdr_fill):
            # Title
            ws.cell(row=row_offset, column=1,
                    value=f"{label} (Sales Data With Transfer Order and without Transfer order)").font = title_font
            ws.merge_cells(f"A{row_offset}:D{row_offset}")

            # Headers (row_offset + 2)
            hrow = row_offset + 2
            for col, h in enumerate(["Row Labels", "Sum of Amount", "Transfer Order Value", "Pure Sales"], 1):
                style_header(ws.cell(row=hrow, column=col, value=h), hdr_fill)

            data_start = hrow + 1
            for r, brand in enumerate(all_brands, data_start):
                ws.cell(row=r, column=1, value=brand).border = thin_border
                total_val    = by_all.get(brand, 0)
                transfer_val = by_tr.get(brand, 0)
                pure_val     = by_cu.get(brand, 0)
                ws.cell(row=r, column=2, value=round(total_val, 2)).number_format = "#,##0.00"
                ws.cell(row=r, column=2).border = thin_border
                if transfer_val:
                    ws.cell(row=r, column=3, value=round(transfer_val, 2)).number_format = "#,##0.00"
                ws.cell(row=r, column=3).border = thin_border
                pure_cell = ws.cell(row=r, column=4, value=f"=B{r}-C{r}")
                pure_cell.number_format = "#,##0.00"
                pure_cell.border = thin_border

            grand_row = data_start + len(all_brands)
            ws.cell(row=grand_row, column=1, value="Grand Total")
            style_total(ws.cell(row=grand_row, column=1))
            for col, letter in [(2, "B"), (3, "C"), (4, "D")]:
                cell = ws.cell(row=grand_row, column=col,
                               value=f"=SUM({letter}{data_start}:{letter}{grand_row - 1})")
                cell.number_format = "#,##0.00"
                style_total(cell)

            charges = [
                ("Other Charges (Sales)",   pl["other_charges"]),
                ("Shipping Charges (Sales)", pl["shipping_charges"]),
                ("Post Supply Discount",     pl["post_supply_discount"]),
            ]
            for i, (lbl, val) in enumerate(charges):
                r = grand_row + 1 + i
                ws.cell(row=r, column=1, value=lbl)
                ws.cell(row=r, column=2, value=round(val, 2)).number_format = "#,##0.00"

            pl_row = grand_row + 1 + len(charges)
            ws.cell(row=pl_row, column=1, value="Total Sales as per P & l account").font = total_font
            charge_sum = "+".join(f"B{grand_row + 1 + i}" for i in range(len(charges)))
            ws.cell(row=pl_row, column=2,
                    value=f"=B{grand_row}+{charge_sum}").number_format = "#,##0.00"

            return pl_row + 3   # first blank row after this block

        # ════════════════════════════════════════════════════════════════════
        # SHEET 1: BWS
        # ════════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = f"(BWS) {month_label}"

        next_row = _write_bws_block(ws1, 1, curr_all, curr_tr, curr_cu, pl_curr, month_label, curr_header_fill)
        _write_bws_block(ws1, next_row, prev_all, prev_tr, prev_cu, pl_prev, prev_month_label, prev_header_fill)

        ws1.column_dimensions["A"].width = 35
        for col in ["B", "C", "D"]:
            ws1.column_dimensions[col].width = 20

        # ════════════════════════════════════════════════════════════════════
        # Helper: build one pivot block (BWC or BWSC) starting at `row_offset`.
        # Returns the row number of the first blank row after the block.
        # ════════════════════════════════════════════════════════════════════
        def _write_pivot_block(ws, row_offset: int, pivot_data: dict, all_cols: list,
                               sheet_title: str, hdr_fill, brands_in_pivot: list):
            n_data_cols = len(all_cols)
            gt_col      = n_data_cols + 2   # col index of "Grand Total"

            ws.cell(row=row_offset, column=1, value=sheet_title).font = title_font
            ws.merge_cells(f"A{row_offset}:{get_column_letter(gt_col)}{row_offset}")

            hrow = row_offset + 2
            style_header(ws.cell(row=hrow, column=1, value="Row Labels"), hdr_fill)
            for ci, col_name in enumerate(all_cols, 2):
                style_header(ws.cell(row=hrow, column=ci, value=col_name), hdr_fill)
            style_header(ws.cell(row=hrow, column=gt_col, value="Grand Total"), hdr_fill)

            data_start = hrow + 1
            for ri, brand in enumerate(brands_in_pivot, data_start):
                ws.cell(row=ri, column=1, value=brand).border = thin_border
                for ci, col_name in enumerate(all_cols, 2):
                    val = pivot_data[brand].get(col_name, 0)
                    c = ws.cell(row=ri, column=ci, value=round(val, 2) if val else None)
                    c.number_format = "#,##0.00"
                    c.border = thin_border
                gt_cell = ws.cell(
                    row=ri, column=gt_col,
                    value=f"=SUM({get_column_letter(2)}{ri}:{get_column_letter(n_data_cols + 1)}{ri})"
                )
                gt_cell.number_format = "#,##0.00"
                gt_cell.border = thin_border

            grand_row = data_start + len(brands_in_pivot)
            ws.cell(row=grand_row, column=1, value="Grand Total")
            style_total(ws.cell(row=grand_row, column=1))
            for ci in range(2, gt_col + 1):
                letter = get_column_letter(ci)
                cell = ws.cell(row=grand_row, column=ci,
                               value=f"=SUM({letter}{data_start}:{letter}{grand_row - 1})")
                cell.number_format = "#,##0.00"
                style_total(cell)

            return grand_row + 3    # first blank row after this block

        cmp_header_fill = PatternFill("solid", fgColor="7030A0")   # purple – comparison block

        def _write_pivot_comparison_block(ws, row_offset: int, curr_pivot: dict, prev_pivot: dict,
                                          all_cols: list, comp_title: str, brands_list: list):
            """
            Writes two stacked comparison sub-blocks (Growth ₹ then Growth %) for a
            brand × category pivot.  Returns the first blank row after the block.
            """
            n_data_cols = len(all_cols)
            gt_col      = n_data_cols + 2

            def _pct(curr_val, prev_val):
                if prev_val:
                    return (curr_val - prev_val) / prev_val   # as fraction for "0.00%" format
                return None

            def _write_sub_block(ws, r0: int, title: str, mode: str):
                """mode = 'abs' (₹) or 'pct' (%)"""
                ws.cell(row=r0, column=1, value=title).font = title_font
                ws.merge_cells(f"A{r0}:{get_column_letter(gt_col)}{r0}")

                hrow = r0 + 2
                style_header(ws.cell(row=hrow, column=1, value="Row Labels"), cmp_header_fill)
                for ci, col_name in enumerate(all_cols, 2):
                    style_header(ws.cell(row=hrow, column=ci, value=col_name), cmp_header_fill)
                style_header(ws.cell(row=hrow, column=gt_col, value="Grand Total"), cmp_header_fill)

                data_start = hrow + 1
                for ri, brand in enumerate(brands_list, data_start):
                    ws.cell(row=ri, column=1, value=brand).border = thin_border
                    row_curr_total = 0.0
                    row_prev_total = 0.0
                    for ci, col_name in enumerate(all_cols, 2):
                        cv = curr_pivot.get(brand, {}).get(col_name, 0) or 0
                        pv = prev_pivot.get(brand, {}).get(col_name, 0) or 0
                        row_curr_total += cv
                        row_prev_total += pv
                        if mode == "abs":
                            diff = round(cv - pv, 2)
                            c = ws.cell(row=ri, column=ci, value=diff if (cv or pv) else None)
                            c.number_format = "#,##0.00"
                        else:
                            pct = _pct(cv, pv)
                            if pct is not None:
                                c = ws.cell(row=ri, column=ci, value=pct)
                                c.number_format = "0.00%"
                            else:
                                c = ws.cell(row=ri, column=ci,
                                            value="NEW" if cv else (None if not pv else None))
                                c.alignment = Alignment(horizontal="center")
                        c.border = thin_border
                        if mode == "abs":
                            val_for_color = cv - pv
                        else:
                            val_for_color = (cv - pv) if pv else (cv if cv else 0)
                        if val_for_color > 0:
                            c.fill = green_fill
                        elif val_for_color < 0:
                            c.fill = red_fill

                    # Grand Total cell for this row
                    if mode == "abs":
                        gt_diff = round(row_curr_total - row_prev_total, 2)
                        gt_c = ws.cell(row=ri, column=gt_col, value=gt_diff if (row_curr_total or row_prev_total) else None)
                        gt_c.number_format = "#,##0.00"
                    else:
                        gt_pct = _pct(row_curr_total, row_prev_total)
                        if gt_pct is not None:
                            gt_c = ws.cell(row=ri, column=gt_col, value=gt_pct)
                            gt_c.number_format = "0.00%"
                        else:
                            gt_c = ws.cell(row=ri, column=gt_col,
                                           value="NEW" if row_curr_total else None)
                            gt_c.alignment = Alignment(horizontal="center")
                    gt_c.border = thin_border
                    if mode == "abs":
                        gt_val_color = row_curr_total - row_prev_total
                    else:
                        gt_val_color = (row_curr_total - row_prev_total) if row_prev_total else (row_curr_total if row_curr_total else 0)
                    if gt_val_color > 0:
                        gt_c.fill = green_fill
                    elif gt_val_color < 0:
                        gt_c.fill = red_fill

                # Grand Total column totals row
                grand_row = data_start + len(brands_list)
                ws.cell(row=grand_row, column=1, value="Grand Total")
                style_total(ws.cell(row=grand_row, column=1))
                for ci in range(2, gt_col + 1):
                    letter = get_column_letter(ci)
                    if mode == "abs":
                        cell = ws.cell(row=grand_row, column=ci,
                                       value=f"=SUM({letter}{data_start}:{letter}{grand_row - 1})")
                        cell.number_format = "#,##0.00"
                    else:
                        # For %, grand total = average of non-N/A rows would be misleading;
                        # recalculate from raw totals stored as SUM of curr / SUM of prev
                        # We don't have those here, so just leave it empty for the column footer
                        cell = ws.cell(row=grand_row, column=ci, value=None)
                    style_total(cell)

                return grand_row + 3   # first blank row after sub-block

            # Sub-block 1: Growth ₹
            next_r = _write_sub_block(ws, row_offset, f"{comp_title} — Growth (₹)", "abs")
            # Sub-block 2: Growth %
            next_r = _write_sub_block(ws, next_r, f"{comp_title} — Growth (%)", "pct")
            return next_r

        def build_pivot_sheet(ws, curr_pivot: dict, prev_pivot: dict, curr_title: str, prev_title: str, comp_title: str):
            # Union of categories/sub-cats across both months (consistent columns in all blocks)
            all_cols = sorted(
                {col for d in curr_pivot.values() for col in d} |
                {col for d in prev_pivot.values() for col in d}
            )
            brands_list = [b for b in all_brands if b in curr_pivot or b in prev_pivot]

            next_r = _write_pivot_block(ws, 1, curr_pivot, all_cols, curr_title, curr_header_fill, brands_list)
            next_r = _write_pivot_block(ws, next_r, prev_pivot, all_cols, prev_title, prev_header_fill, brands_list)
            _write_pivot_comparison_block(ws, next_r, curr_pivot, prev_pivot, all_cols, comp_title, brands_list)

            ws.column_dimensions["A"].width = 30
            n_cols = len(all_cols) + 2
            for ci in range(2, n_cols + 1):
                ws.column_dimensions[get_column_letter(ci)].width = 18

        # ════════════════════════════════════════════════════════════════════
        # SHEET 2: BWC
        # ════════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet(title=f"(BWC) {month_label}")
        build_pivot_sheet(
            ws2, curr_bwc, prev_bwc,
            curr_title=f"item.CF.Category ({month_label})",
            prev_title=f"item.CF.Category ({prev_month_label})",
            comp_title=f"item.CF.Category ({prev_month_label} → {month_label})",
        )

        # ════════════════════════════════════════════════════════════════════
        # SHEET 3: BWSC
        # ════════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet(title=f"(BWSC) {month_label}")
        build_pivot_sheet(
            ws3, curr_bwsc, prev_bwsc,
            curr_title=f"item.CF.Sub Category ({month_label})",
            prev_title=f"item.CF.Sub Category ({prev_month_label})",
            comp_title=f"item.CF.Sub Category ({prev_month_label} → {month_label})",
        )

        # ════════════════════════════════════════════════════════════════════
        # SHEET 4: Summary — month-over-month growth by brand
        # ════════════════════════════════════════════════════════════════════
        ws4 = wb.create_sheet(title="Summary")

        ws4.cell(row=1, column=1,
                 value=f"Brand Sales Summary — {prev_month_label} vs {month_label}").font = Font(bold=True, size=13)
        ws4.merge_cells("A1:E1")

        sum_headers = [
            "Brand",
            f"{prev_month_label}\nPure Sales (₹)",
            f"{month_label}\nPure Sales (₹)",
            "Growth (₹)",
            "Growth (%)",
        ]
        for ci, h in enumerate(sum_headers, 1):
            cell = ws4.cell(row=3, column=ci, value=h)
            style_header(cell, curr_header_fill)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws4.row_dimensions[3].height = 30

        data_start = 4
        for ri, brand in enumerate(all_brands, data_start):
            prev_val = round(prev_cu.get(brand, 0), 2)
            curr_val = round(curr_cu.get(brand, 0), 2)
            growth_abs = round(curr_val - prev_val, 2)
            if prev_val:
                growth_pct = round((curr_val - prev_val) / prev_val * 100, 2)
            else:
                growth_pct = None   # N/A when no prior-month sales

            ws4.cell(row=ri, column=1, value=brand).border = thin_border

            c2 = ws4.cell(row=ri, column=2, value=prev_val)
            c2.number_format = "#,##0.00"
            c2.border = thin_border

            c3 = ws4.cell(row=ri, column=3, value=curr_val)
            c3.number_format = "#,##0.00"
            c3.border = thin_border

            c4 = ws4.cell(row=ri, column=4, value=growth_abs)
            c4.number_format = "#,##0.00"
            c4.border = thin_border
            if growth_abs > 0:
                c4.fill = green_fill
            elif growth_abs < 0:
                c4.fill = red_fill

            if growth_pct is not None:
                c5 = ws4.cell(row=ri, column=5, value=growth_pct / 100)
                c5.number_format = "0.00%"
                c5.border = thin_border
                if growth_pct > 0:
                    c5.fill = green_fill
                elif growth_pct < 0:
                    c5.fill = red_fill
            else:
                c5 = ws4.cell(row=ri, column=5, value="N/A")
                c5.border = thin_border
                c5.alignment = Alignment(horizontal="center")

        # Grand Total row
        grand_row = data_start + len(all_brands)
        ws4.cell(row=grand_row, column=1, value="Grand Total")
        style_total(ws4.cell(row=grand_row, column=1))

        for col, letter in [(2, "B"), (3, "C"), (4, "D")]:
            cell = ws4.cell(row=grand_row, column=col,
                            value=f"=SUM({letter}{data_start}:{letter}{grand_row - 1})")
            cell.number_format = "#,##0.00"
            style_total(cell)

        # Grand total growth %: (C_total - B_total) / B_total
        pct_cell = ws4.cell(
            row=grand_row, column=5,
            value=f"=IF(B{grand_row}<>0,(C{grand_row}-B{grand_row})/B{grand_row},\"N/A\")"
        )
        pct_cell.number_format = "0.00%"
        style_total(pct_cell)

        ws4.column_dimensions["A"].width = 32
        for col in ["B", "C", "D", "E"]:
            ws4.column_dimensions[col].width = 22

        # ── Stream response ─────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        month_slug = start_date.strftime("%Y-%m")
        filename = f"brand_sales_report_{month_slug}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating brand sales report: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@router.get("/brands")
def get_available_brands(db=Depends(get_database)):
    """
    Get list of available brands.
    Adjust this based on your actual data structure.
    """
    try:
        products_collection = db.get_collection(PRODUCTS_COLLECTION)

        # Get distinct brands
        brands = products_collection.find({"status": "active"}).distinct("brand")
        return {
            "brands": [
                {"value": brand, "label": brand.title()} for brand in brands if brand
            ]
        }

    except Exception as e:
        logger.error(f"Error getting brands: {e}")
        raise HTTPException(status_code=500, detail="Error retrieving brands")


class SalesReportRequest(BaseModel):
    start_date: str
    end_date: str

    @validator("start_date", "end_date")
    def validate_date_format(cls, v):
        try:
            datetime.strptime(v, "%Y-%m-%d")
            return v
        except ValueError:
            raise ValueError("Date must be in YYYY-MM-DD format")

    @validator("end_date")
    def validate_date_range(cls, v, values):
        if "start_date" in values:
            start = datetime.strptime(values["start_date"], "%Y-%m-%d")
            end = datetime.strptime(v, "%Y-%m-%d")
            if start > end:
                raise ValueError("End date must be after start date")
        return v


# Response models
class SalesReportItem(BaseModel):
    item_name: str
    sku_code: str
    units_sold: int
    units_returned: int
    credit_notes: int = 0
    total_amount: float
    closing_stock: int
    total_days_in_stock: int
    total_days_in_stock_any_wh: int = 0
    drr: float
    last_90_days_dates: str = ""  # Optional field for last 90 days in stock date ranges


class SalesReportResponse(BaseModel):
    data: List[SalesReportItem]
    summary: dict
    meta: dict


import asyncio
from concurrent.futures import ThreadPoolExecutor
from typing import Dict, List, Any, Optional
from functools import lru_cache

# Thread pool for parallel operations
executor = ThreadPoolExecutor(max_workers=4)


def format_date_ranges(dates: List[datetime]) -> str:
    """
    Format a list of dates into ranges or individual dates.
    Continuous dates are shown as ranges (e.g., "1 Jan 2025 - 5 Jan 2025")
    Discontinuous dates are shown separated by commas.

    Args:
        dates: List of datetime objects

    Returns:
        Formatted string with date ranges
    """
    if not dates:
        return ""

    # Sort dates
    sorted_dates = sorted(dates)

    ranges = []
    start_date = sorted_dates[0]
    end_date = sorted_dates[0]

    for i in range(1, len(sorted_dates)):
        current_date = sorted_dates[i]
        # Check if current date is consecutive to end_date
        if (current_date - end_date).days == 1:
            end_date = current_date
        else:
            # Save the current range
            if start_date == end_date:
                ranges.append(start_date.strftime("%d %b %Y"))
            else:
                ranges.append(
                    f"{start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}"
                )
            # Start new range
            start_date = current_date
            end_date = current_date

    # Add the last range
    if start_date == end_date:
        ranges.append(start_date.strftime("%d %b %Y"))
    else:
        ranges.append(
            f"{start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}"
        )

    return ", ".join(ranges)


async def fetch_last_n_days_in_stock(
    db, end_datetime: datetime, item_ids: list, n_days: int = 90
) -> Dict:
    """
    Fetch the last N days an item was in stock, regardless of when those days occurred.
    Returns formatted date ranges for each item.

    Args:
        db: Database connection
        end_datetime: End date to look back from
        item_ids: List of zoho_item_ids to fetch stock for
        n_days: Number of days to fetch (default 90)

    Returns:
        Dict mapping item_id to formatted date ranges string
    """
    try:
        stock_collection = db["zoho_warehouse_stock"]

        # CRITICAL OPTIMIZATION: Limit lookback to prevent scanning millions of records
        # Look back maximum 1 year to find the last 90 days in stock
        lookback_limit = end_datetime - timedelta(days=365)

        logger.info(
            f"Fetching last {n_days} days in stock for {len(item_ids)} items from {lookback_limit.date()} to {end_datetime.date()}"
        )

        # Aggregate to get the last N days where stock > 0 for each item
        pipeline = [
            {
                "$match": {
                    "zoho_item_id": {"$in": item_ids},
                    "date": {
                        "$gte": lookback_limit,
                        "$lte": end_datetime,
                    },  # Only last 1 year
                }
            },
            {
                "$project": {
                    "zoho_item_id": 1,
                    "date": 1,
                    "pupscribe_stock": {
                        "$ifNull": [
                            "$warehouses.Pupscribe Enterprises Private Limited",
                            0,
                        ]
                    },
                }
            },
            # Filter only days with stock > 0
            {"$match": {"pupscribe_stock": {"$gt": 0}}},
            # Sort by item and date descending (most recent first)
            {"$sort": {"zoho_item_id": 1, "date": -1}},
            # Group by item and collect last N dates
            {"$group": {"_id": "$zoho_item_id", "stock_dates": {"$push": "$date"}}},
            # Limit to last N days
            {
                "$project": {
                    "_id": 1,
                    "stock_dates": {"$slice": ["$stock_dates", n_days]},
                }
            },
        ]

        def _run_aggregation():
            cursor = stock_collection.aggregate(
                pipeline, allowDiskUse=True, batchSize=10000
            )
            stock_date_ranges = {}
            processed_count = 0
            for doc in cursor:
                processed_count += 1
                item_id = doc["_id"]
                dates = doc.get("stock_dates", [])

                if processed_count <= 3:
                    logger.debug(f"DEBUG: Item {item_id} has {len(dates)} stock dates")

                if dates:
                    formatted_ranges = format_date_ranges(dates)
                    stock_date_ranges[item_id] = formatted_ranges

                    if processed_count == 1:
                        logger.debug(
                            f"DEBUG: First item formatted ranges: {formatted_ranges[:100]}..."
                        )
                else:
                    stock_date_ranges[item_id] = ""

            logger.info(
                f"Fetched last {n_days} days in stock: processed {processed_count} items, returned {len(stock_date_ranges)} items with data"
            )

            if stock_date_ranges:
                sample_keys = list(stock_date_ranges.keys())[:3]
                logger.debug(f"DEBUG: Sample item IDs in result: {sample_keys}")

            return stock_date_ranges

        return await asyncio.to_thread(_run_aggregation)

    except Exception as e:
        logger.error(f"Error fetching last N days in stock: {e}", exc_info=True)
        return {}


async def fetch_stock_data_for_items(
    db,
    start_datetime: datetime,
    end_datetime: datetime,
    item_ids: list,
    fetch_last_90_days: bool = False,
) -> Dict:
    """
    Fetch stock data ONLY for specific items (not all items in database).
    This is MUCH faster than scanning all 4.3M records.

    Args:
        db: Database connection
        start_datetime: Start date for days in stock calculation
        end_datetime: End date for closing stock
        item_ids: List of zoho_item_ids to fetch stock for
        fetch_last_90_days: If True, also fetch last 90 days in stock with date ranges

    Returns:
        Dict mapping item_id to {closing_stock, total_days_in_stock, last_90_days_dates (optional)}
    """
    try:
        stock_collection = db["zoho_warehouse_stock"]

        logger.info(
            f"Fetching stock for {len(item_ids)} specific items up to {end_datetime.date()}"
        )

        # CRITICAL OPTIMIZATION: Only query stock for items that have sales
        pipeline = [
            # Match ONLY the specific items we need
            {
                "$match": {
                    "zoho_item_id": {"$in": item_ids},  # ONLY these items!
                    "date": {"$lte": end_datetime},
                }
            },
            # Project only needed fields early
            {
                "$project": {
                    "zoho_item_id": 1,
                    "date": 1,
                    "pupscribe_stock": {
                        "$ifNull": [
                            "$warehouses.Pupscribe Enterprises Private Limited",
                            0,
                        ]
                    },
                }
            },
            # Sort by item and date descending
            {"$sort": {"zoho_item_id": 1, "date": -1}},
            # Group and calculate using $sum (no arrays)
            {
                "$group": {
                    "_id": "$zoho_item_id",
                    "closing_stock": {"$first": "$pupscribe_stock"},
                    "latest_date": {"$first": "$date"},
                    "total_days_in_stock": {
                        "$sum": {
                            "$cond": {
                                "if": {
                                    "$and": [
                                        {"$gte": ["$date", start_datetime]},
                                        {"$lte": ["$date", end_datetime]},
                                        {"$gt": ["$pupscribe_stock", 0]},
                                    ]
                                },
                                "then": 1,
                                "else": 0,
                            }
                        }
                    },
                }
            },
            {
                "$project": {
                    "_id": 1,
                    "closing_stock": 1,
                    "total_days_in_stock": 1,
                    "latest_date": 1,
                }
            },
        ]

        # Execute with optimized settings - run in thread to avoid blocking event loop
        def _run_aggregation():
            cursor = stock_collection.aggregate(
                pipeline, allowDiskUse=True, batchSize=10000
            )
            result = {}
            for doc in cursor:
                result[doc["_id"]] = {
                    "closing_stock": doc.get("closing_stock", 0),
                    "total_days_in_stock": doc.get("total_days_in_stock", 0),
                }
            return result

        stock_data = await asyncio.to_thread(_run_aggregation)

        # If requested, also fetch last 90 days in stock with date ranges
        if fetch_last_90_days:
            last_90_days_data = await fetch_last_n_days_in_stock(
                db, end_datetime, item_ids, 90
            )
            # Merge the last 90 days data into stock_data
            for item_id, date_ranges in last_90_days_data.items():
                if item_id in stock_data:
                    stock_data[item_id]["last_90_days_dates"] = date_ranges
                else:
                    stock_data[item_id] = {
                        "closing_stock": 0,
                        "total_days_in_stock": 0,
                        "last_90_days_dates": date_ranges,
                    }

        logger.info(f"Fetched stock data for {len(stock_data)} items")
        return stock_data

    except Exception as e:
        logger.error(f"Error fetching stock data for items: {e}")
        return {}


async def fetch_stock_data_for_items_batch(
    db,
    periods: list,
    item_ids: list,
) -> Dict[int, Dict]:
    """
    Fetch stock data (days_in_stock) for ALL lookback periods in a SINGLE aggregation.
    Groups by (item_id, period_index) using $switch on date buckets.

    Args:
        periods: List of (start_dt, end_dt) tuples
        item_ids: List of zoho_item_ids

    Returns:
        Dict[period_index, Dict[item_id, {"total_days_in_stock": int}]]
    """
    if not periods or not item_ids:
        return {}

    try:
        stock_collection = db["zoho_warehouse_stock"]

        overall_start = min(p[0] for p in periods)
        overall_end = max(p[1] for p in periods)

        # zoho_item_id is always stored as string (backfill + cron both use str())
        str_item_ids = [str(iid) for iid in item_ids]

        # Build period_index switch branches using datetime comparison
        period_branches = []
        for i, (p_start, p_end) in enumerate(periods):
            period_branches.append(
                {
                    "case": {
                        "$and": [
                            {"$gte": ["$date", p_start]},
                            {"$lte": ["$date", p_end]},
                        ]
                    },
                    "then": i,
                }
            )

        pipeline = [
            {
                "$match": {
                    "zoho_item_id": {"$in": str_item_ids},
                    "date": {"$gte": overall_start, "$lte": overall_end},
                }
            },
            {
                "$project": {
                    "zoho_item_id": 1,
                    "date": 1,
                    "pupscribe_stock": {
                        "$ifNull": [
                            "$warehouses.Pupscribe Enterprises Private Limited",
                            0,
                        ]
                    },
                }
            },
            # Assign period_index
            {
                "$addFields": {
                    "period_index": {
                        "$switch": {
                            "branches": period_branches,
                            "default": -1,
                        }
                    }
                }
            },
            {"$match": {"period_index": {"$gte": 0}}},
            # Group by (item_id, period_index) — count days with stock > 0
            {
                "$group": {
                    "_id": {"item_id": "$zoho_item_id", "period": "$period_index"},
                    "total_days_in_stock": {
                        "$sum": {"$cond": [{"$gt": ["$pupscribe_stock", 0]}, 1, 0]}
                    },
                }
            },
        ]

        def _run():
            cursor = stock_collection.aggregate(
                pipeline, allowDiskUse=True, batchSize=10000
            )
            results = {}
            for doc in cursor:
                period_idx = doc["_id"]["period"]
                item_id = doc["_id"]["item_id"]
                if period_idx not in results:
                    results[period_idx] = {}
                results[period_idx][item_id] = {
                    "total_days_in_stock": doc.get("total_days_in_stock", 0),
                }
            return results

        stock_data = await asyncio.to_thread(_run)
        logger.info(
            f"Lookback stock batch: fetched data for {len(stock_data)} periods, "
            f"{len(item_ids)} items"
        )
        return stock_data

    except Exception as e:
        logger.error(f"Error fetching stock data batch: {e}")
        return {}


async def fetch_zoho_lookback_sales_batch(
    db,
    periods: list,
    target_item_ids: list,
) -> Dict[int, Dict[str, dict]]:
    """
    Fetch Zoho sales + returns for ALL lookback periods in a SINGLE aggregation.
    Covers the full date range (earliest period start → most recent period end),
    then uses $switch to bucket each document into its period index.

    Args:
        periods: List of (start_dt, end_dt) tuples, one per lookback period
        target_item_ids: List of item_ids to filter for

    Returns:
        Dict[period_index, Dict[item_id_str, {"units_sold": int, "returns": int}]]
    """
    if not periods or not target_item_ids:
        return {}

    try:
        invoices_collection = db[INVOICES_COLLECTION]

        # Full date range spanning all periods
        overall_start = min(p[0] for p in periods)
        overall_end = max(p[1] for p in periods)
        overall_start_str = overall_start.strftime("%Y-%m-%d")
        overall_end_str = overall_end.strftime("%Y-%m-%d")

        # Include both string and int forms — invoices may store item_id as either type
        target_ids_set = []
        for iid in target_item_ids:
            target_ids_set.append(str(iid))
            try:
                target_ids_set.append(int(iid))
            except (ValueError, TypeError):
                pass

        # Build custom pipeline that carries doc_date through composite expansion
        # so we can bucket results by period_index
        cn_start = datetime.strptime(overall_start_str, "%Y-%m-%d")
        cn_end = datetime.strptime(overall_end_str, "%Y-%m-%d").replace(
            hour=23, minute=59, second=59
        )

        invoice_match = {
            "$match": {
                "$and": [
                    {"date": {"$gte": overall_start_str, "$lte": overall_end_str}},
                    {"status": {"$nin": ["draft", "void"]}},
                ]
            }
        }

        credit_note_match_conditions = [
            {
                "$or": [
                    {"date": {"$gte": overall_start_str, "$lte": overall_end_str}},
                    {"date": {"$gte": cn_start, "$lte": cn_end}},
                ]
            },
            {"customer_name": {"$nin": TRANSFER_ORDER_CUSTOMERS}},
        ]

        # Build period_index switch branches
        period_branches = []
        for i, (p_start, p_end) in enumerate(periods):
            ps = p_start.strftime("%Y-%m-%d")
            pe = p_end.strftime("%Y-%m-%d")
            period_branches.append(
                {
                    "case": {
                        "$and": [
                            {"$gte": ["$doc_date", ps]},
                            {"$lte": ["$doc_date", pe]},
                        ]
                    },
                    "then": i,
                }
            )

        custom_pipeline = [
            invoice_match,
            {"$addFields": {"doc_type": "invoice", "doc_date": "$date"}},
            {
                "$unionWith": {
                    "coll": "credit_notes",
                    "pipeline": [
                        {"$match": {"$and": credit_note_match_conditions}},
                        {
                            "$addFields": {
                                "doc_type": "credit_note",
                                # credit_notes may store date as ISODate or string
                                "doc_date": {
                                    "$cond": {
                                        "if": {"$eq": [{"$type": "$date"}, "string"]},
                                        "then": "$date",
                                        "else": {
                                            "$dateToString": {
                                                "format": "%Y-%m-%d",
                                                "date": "$date",
                                            }
                                        },
                                    }
                                },
                            }
                        },
                    ],
                }
            },
            {"$unwind": "$line_items"},
            {
                "$lookup": {
                    "from": "composite_products",
                    "localField": "line_items.item_id",
                    "foreignField": "composite_item_id",
                    "as": "composite_match",
                }
            },
            {
                "$addFields": {
                    "composite_info": {"$arrayElemAt": ["$composite_match", 0]},
                }
            },
            {
                "$project": {
                    "doc_date": 1,
                    "items_to_process": {
                        "$cond": [
                            {"$gt": [{"$size": "$composite_match"}, 0]},
                            {
                                "$map": {
                                    "input": "$composite_info.components",
                                    "as": "component",
                                    "in": {
                                        "doc_type": "$doc_type",
                                        "doc_date": "$doc_date",
                                        "item_id": "$$component.item_id",
                                        "quantity": {
                                            "$multiply": [
                                                "$line_items.quantity",
                                                "$$component.quantity",
                                            ]
                                        },
                                    },
                                }
                            },
                            [
                                {
                                    "doc_type": "$doc_type",
                                    "doc_date": "$doc_date",
                                    "item_id": "$line_items.item_id",
                                    "quantity": "$line_items.quantity",
                                }
                            ],
                        ]
                    },
                }
            },
            {"$unwind": "$items_to_process"},
            {"$replaceRoot": {"newRoot": "$items_to_process"}},
            # Filter to target items
            {"$match": {"item_id": {"$in": target_ids_set}}},
            # Assign period_index
            {
                "$addFields": {
                    "period_index": {
                        "$switch": {
                            "branches": period_branches,
                            "default": -1,
                        }
                    }
                }
            },
            {"$match": {"period_index": {"$gte": 0}}},
            # Group by (item_id, period_index)
            {
                "$group": {
                    "_id": {"item_id": "$item_id", "period": "$period_index"},
                    "total_units_sold": {
                        "$sum": {
                            "$cond": [
                                {"$eq": ["$doc_type", "invoice"]},
                                "$quantity",
                                0,
                            ]
                        }
                    },
                    "total_returns": {
                        "$sum": {
                            "$cond": [
                                {"$eq": ["$doc_type", "credit_note"]},
                                "$quantity",
                                0,
                            ]
                        }
                    },
                }
            },
        ]

        def _run():
            cursor = invoices_collection.aggregate(
                custom_pipeline, allowDiskUse=True, batchSize=5000
            )
            results = {}
            for doc in cursor:
                period_idx = doc["_id"]["period"]
                item_id = str(doc["_id"]["item_id"])
                if period_idx not in results:
                    results[period_idx] = {}
                results[period_idx][item_id] = {
                    "units_sold": doc.get("total_units_sold", 0),
                    "returns": doc.get("total_returns", 0),
                }
            return results

        sales_data = await asyncio.to_thread(_run)
        logger.info(
            f"Lookback sales batch: fetched data for {len(sales_data)} periods, "
            f"{overall_start_str} to {overall_end_str}"
        )
        return sales_data

    except Exception as e:
        logger.error(f"Error fetching lookback sales batch: {e}")
        import traceback

        logger.error(traceback.format_exc())
        return {}


async def fetch_stock_data_optimized(
    db, start_datetime: datetime, end_datetime: datetime
) -> Dict:
    """
    Optimized stock data fetching with better aggregation pipeline.
    - Requires zoho_item_id field in stock records (run migration script first)
    - Closing Stock: Most recent stock value on or before end_date
    - Days in Stock: Count of days within date range where stock > 0

    Performance optimizations:
    - Limits lookback to 730 days (2 years) before start date for closing stock
    - Uses $sum instead of $push to avoid building arrays
    - Early projection to reduce data processing
    """
    try:
        stock_collection = db["zoho_warehouse_stock"]

        # Split into two focused parallel queries instead of one monolithic scan:
        #
        # Q1 — days_in_stock: only the exact report period (e.g. 3 months).
        #      No sort needed, just a $sum. Scans ~period_days × n_items docs.
        #
        # Q2 — closing_stock: 90-day lookback from end_date to find the most
        #      recent stock per item. 90 days is sufficient because Zoho stock
        #      is synced daily; any item with no record in 90 days has 0 stock.
        #      Scans ~(period_days + 90) × n_items docs — far fewer than the
        #      former 2-year (730-day) window which scanned ~1.8M docs.
        #
        # Running both in separate threads lets MongoDB serve them concurrently.

        closing_lookback = end_datetime - timedelta(days=90)

        logger.info(
            f"Fetching stock data: days [{start_datetime.date()}–{end_datetime.date()}], "
            f"closing [{closing_lookback.date()}–{end_datetime.date()}] (parallel)"
        )

        def _fetch_days_in_stock():
            """Count days with inventory within the report period only."""
            pipeline = [
                {
                    "$match": {
                        "date": {"$gte": start_datetime, "$lte": end_datetime},
                        "zoho_item_id": {"$ne": None},
                    }
                },
                {
                    "$project": {
                        "zoho_item_id": 1,
                        "pupscribe_stock": {
                            "$ifNull": [
                                "$warehouses.Pupscribe Enterprises Private Limited",
                                0,
                            ]
                        },
                    }
                },
                {
                    "$group": {
                        "_id": "$zoho_item_id",
                        "total_days_in_stock": {
                            "$sum": {"$cond": [{"$gt": ["$pupscribe_stock", 0]}, 1, 0]}
                        },
                    }
                },
            ]
            result = {}
            for doc in stock_collection.aggregate(
                pipeline, allowDiskUse=True, batchSize=10000
            ):
                result[doc["_id"]] = doc.get("total_days_in_stock", 0)
            return result

        def _fetch_closing_stock():
            """Get the most recent stock per item as of end_date (90-day lookback)."""
            pipeline = [
                {
                    "$match": {
                        "date": {"$gte": closing_lookback, "$lte": end_datetime},
                        "zoho_item_id": {"$ne": None},
                    }
                },
                # Sort BEFORE $project — allows MongoDB to use compound index
                # {zoho_item_id: 1, date: -1} without an in-memory sort.
                {"$sort": {"zoho_item_id": 1, "date": -1}},
                {
                    "$project": {
                        "zoho_item_id": 1,
                        "date": 1,
                        "pupscribe_stock": {
                            "$ifNull": [
                                "$warehouses.Pupscribe Enterprises Private Limited",
                                0,
                            ]
                        },
                    }
                },
                {
                    "$group": {
                        "_id": "$zoho_item_id",
                        "closing_stock": {"$first": "$pupscribe_stock"},
                        "latest_date": {"$first": "$date"},
                    }
                },
            ]
            result = {}
            for doc in stock_collection.aggregate(
                pipeline, allowDiskUse=True, batchSize=10000
            ):
                result[doc["_id"]] = {
                    "closing_stock": doc.get("closing_stock", 0) or 0,
                    "latest_date": doc.get("latest_date"),
                }
            return result

        def _fetch_days_in_stock_any_wh():
            """Count days with stock in ANY warehouse within the report period."""
            pipeline = [
                {
                    "$match": {
                        "date": {"$gte": start_datetime, "$lte": end_datetime},
                        "zoho_item_id": {"$ne": None},
                    }
                },
                {
                    "$project": {
                        "zoho_item_id": 1,
                        "total_wh_stock": {
                            "$reduce": {
                                "input": {"$objectToArray": "$warehouses"},
                                "initialValue": 0,
                                "in": {"$add": ["$$value", "$$this.v"]},
                            }
                        },
                    }
                },
                {
                    "$group": {
                        "_id": "$zoho_item_id",
                        "total_days_in_stock_any_wh": {
                            "$sum": {"$cond": [{"$gt": ["$total_wh_stock", 0]}, 1, 0]}
                        },
                    }
                },
            ]
            result = {}
            for doc in stock_collection.aggregate(pipeline, allowDiskUse=True, batchSize=10000):
                result[doc["_id"]] = doc.get("total_days_in_stock_any_wh", 0)
            return result

        # Run all three aggregations in parallel
        days_data, closing_data, days_any_wh_data = await asyncio.gather(
            asyncio.to_thread(_fetch_days_in_stock),
            asyncio.to_thread(_fetch_closing_stock),
            asyncio.to_thread(_fetch_days_in_stock_any_wh),
        )

        # Merge: union of all item_ids seen in any query
        all_item_ids = set(days_data.keys()) | set(closing_data.keys()) | set(days_any_wh_data.keys())
        stock_data = {}
        for item_id in all_item_ids:
            stock_data[item_id] = {
                "closing_stock": closing_data.get(item_id, {}).get("closing_stock", 0),
                "total_days_in_stock": days_data.get(item_id, 0),
                "total_days_in_stock_any_wh": days_any_wh_data.get(item_id, 0),
            }

        logger.info(
            f"Fetched {len(stock_data)} stock records for date range "
            f"{start_datetime.date()} to {end_datetime.date()}"
        )
        return stock_data

    except Exception as e:
        logger.error(f"Error fetching stock data: {e}")
        return {}


async def fetch_all_products_indexed(db) -> Dict:
    """
    Fetch all products and create an indexed map for O(1) lookups.
    """
    try:
        products_collection = db["products"]

        def _fetch():
            cursor = products_collection.find(
                {
                    "$and": [
                        {"name": {"$ne": ""}},
                        {"name": {"$ne": "amazon"}},
                    ]
                },
                {
                    "item_id": 1,
                    "cf_sku_code": 1,
                    "name": 1,
                    "_id": 0,
                },
                batch_size=5000,
            )
            products_map = {}
            for product in cursor:
                products_map[product.get("item_id")] = product.get("cf_sku_code", "")
            return products_map

        products_map = await asyncio.to_thread(_fetch)
        logger.info(f"Indexed {len(products_map)} products")
        return products_map

    except Exception as e:
        logger.error(f"Error fetching products: {e}")
        return {}


def process_batch(batch: List[Dict], stock_data: Dict, products_map: Dict) -> Dict:
    items = []
    total_units = 0
    total_amount = 0
    total_stock = 0

    for item in batch:
        item_id = item["_id"]
        units_sold = item.get("total_units_sold", 0)
        units_returned = item.get("total_units_returned", 0)
        amount = item.get("total_amount", 0.0)

        # Get SKU and stock data...
        sku_code = products_map.get(item_id, "")
        stock_info = stock_data.get(item_id, {})
        closing_stock = stock_info.get("closing_stock", 0)
        days_in_stock = stock_info.get("total_days_in_stock", 0)
        days_in_stock_any_wh = stock_info.get("total_days_in_stock_any_wh", 0)
        drr = round(units_sold / days_in_stock, 2) if days_in_stock > 0 else 0

        # Only accumulate totals for items with names
        if item.get("item_name") != "":
            total_units += units_sold
            total_amount += amount
            total_stock += closing_stock

            items.append(
                SalesReportItem(
                    item_name=item.get("item_name") or "",
                    sku_code=sku_code,
                    units_sold=units_sold,
                    total_amount=round(amount, 2),
                    closing_stock=closing_stock,
                    total_days_in_stock=days_in_stock,
                    total_days_in_stock_any_wh=days_in_stock_any_wh,
                    drr=drr,
                )
            )

    return {
        "items": items,
        "units": total_units,
        "amount": total_amount,
        "stock": total_stock,
    }


@lru_cache(maxsize=100)
def get_excluded_customer_list() -> str:
    patterns = [
        # "(EC)",
        # "(NA)",
        # "(amzb2b)",
        # "(amz2b2)",
        # "(PUPEV)",
        # "(RS)",
        # "(MKT)",
        # "(SPUR)",
        # "(SSAM)",
        # "(OSAM)",
        # "(OSAMP)",
        # "(DON)",
        # "(Jiomart b2c)",
        # "Blinkit",
        # "KIRANAKART",
        # "Mr. Tikoti Laxman",
        # "Rushil Kalsi",
        # "ETRADE",
        # "Pupscribe",
        # "Flipkart",
    ]

    escaped_patterns = [re.escape(pattern) for pattern in patterns]
    regex_pattern = "|".join(escaped_patterns)

    return regex_pattern


async def get_cached_report(cache_key: str, db) -> Optional[Dict]:
    """
    Get cached report if available and fresh (less than 1 hour old).
    """
    try:
        cache_collection = db.get_collection("report_cache")

        def _fetch():
            return cache_collection.find_one(
                {
                    "_id": cache_key,
                    "created_at": {"$gte": datetime.now() - timedelta(hours=1)},
                }
            )

        cached = await asyncio.to_thread(_fetch)

        if cached:
            return cached.get("data")

    except Exception as e:
        logger.warning(f"Cache retrieval failed: {e}")

    return None


async def cache_report(cache_key: str, data: Dict, db):
    """
    Cache report for future use.
    """
    try:
        cache_collection = db.get_collection("report_cache")

        def _write():
            cache_collection.replace_one(
                {"_id": cache_key},
                {"_id": cache_key, "data": data, "created_at": datetime.now()},
                upsert=True,
            )

        await asyncio.to_thread(_write)

        logger.info(f"Report cached with key: {cache_key}")

    except Exception as e:
        logger.warning(f"Failed to cache report: {e}")


TRANSFER_ORDER_CUSTOMERS = [
    "(amzb2b) Pupscribe Enterprises Pvt Ltd (KA)",
    "(amzb2b) Pupscribe Enterprises Pvt Ltd (MH)",
    "(amzb2b) Pupscribe Enterprises Pvt Ltd (TL)",
    "(amzb2b) Pupscribe Enterprises Pvt Ltd (TN)",
    "(amzb2b) Pupscribe Enterprises Pvt Ltd (WB)",
    "(amzb2b) Pupscribe Enterprises Pvt Ltd (HY)",
    "(amzb2b) Pupscribe Enterprises Pvt Ltd (DL)",
    "Pupscribe Enterprises Private Limited (Blinkit Maharashtra)",
    "Pupscribe Enterprises Private Limited (Blinkit Karnataka)",
    "Pupscribe Enterprises Private Limited (Blinkit Telangana)",
    "Pupscribe Enterprises Private Limited (Blinkit Tamil Nadu)",
    "Pupscribe Enterprises Private Limited (Blinkit Haryana)",
    "Pupscribe Enterprises Private Limited (Blinkit West Bengal)",
    "(amzb2b) Pupscribe Enterprises Pvt Ltd (UP)",
    "(amzb2b) Pupscribe Enterprises Pvt Ltd (GJ)",
    "(amzb2b) Pupscribe Enterprises Pvt Ltd (KL)",
    "(amzb2b) Pupscribe Enterprises Pvt Ltd (AD)",
    "(amzb2b) Pupscribe Enterprises Pvt Ltd (GA)",
    "(amzb2b) Pupscribe Enterprises Pvt Ltd (PB)",
]


def build_optimized_pipeline_with_credit_notes_and_composites(
    start_date: str, end_date: str, excluded_customers: str = None
) -> List[Dict]:
    """
    Build an optimized aggregation pipeline that incorporates both invoices and credit notes,
    and handles composite products by breaking them down into individual components.
    Credit notes are matched using ISODate format (datetime objects) since credit_notes
    collection stores dates as ISODate, while invoices store dates as strings.
    Transfer order customers are excluded from credit notes.

    When excluded_customers is None or empty, no customer filtering is applied.
    """
    # credit_notes stores date as ISODate, invoices as string
    cn_start = datetime.strptime(start_date, "%Y-%m-%d")
    cn_end = datetime.strptime(end_date, "%Y-%m-%d").replace(
        hour=23, minute=59, second=59
    )

    # Build invoice match conditions
    invoice_match_conditions = [
        {"date": {"$gte": start_date, "$lte": end_date}},
        {"status": {"$nin": ["draft", "void"]}},
        # {"invoice_number": {"$regex": "INV/", "$options": "i"}},
    ]

    # Build credit note match conditions (use ISODate for credit_notes collection)
    credit_note_match_conditions = [
        {
            "$or": [
                {"date": {"$gte": start_date, "$lte": end_date}},
                {"date": {"$gte": cn_start, "$lte": cn_end}},
            ]
        },
        # Exclude transfer order customers from credit notes (returns)
        {"customer_name": {"$nin": TRANSFER_ORDER_CUSTOMERS}},
    ]

    # Only add customer exclusion filter if excluded_customers is provided
    if excluded_customers:
        customer_filter = {
            "customer_name": {
                "$not": {
                    "$regex": excluded_customers,
                    "$options": "i",
                }
            }
        }
        invoice_match_conditions.append(customer_filter)
        credit_note_match_conditions.append(customer_filter)

    return [
        {"$match": {"$and": invoice_match_conditions}},
        {"$addFields": {"doc_type": "invoice"}},
        {
            "$unionWith": {
                "coll": "credit_notes",
                "pipeline": [
                    {"$match": {"$and": credit_note_match_conditions}},
                    {"$addFields": {"doc_type": "credit_note"}},
                ],
            }
        },
        {"$unwind": "$line_items"},
        {
            "$lookup": {
                "from": "composite_products",
                "localField": "line_items.item_id",
                "foreignField": "composite_item_id",
                "as": "composite_match",
            }
        },
        {
            "$addFields": {
                "is_composite": {
                    "$or": [
                        {"$gt": [{"$size": "$composite_match"}, 0]},
                        {"$eq": ["$line_items.is_combo_product", True]},
                    ]
                },
                "composite_info": {"$arrayElemAt": ["$composite_match", 0]},
            }
        },
        {
            "$project": {
                "items_to_process": {
                    "$cond": [
                        # If it's a composite item with components found
                        {"$gt": [{"$size": "$composite_match"}, 0]},
                        # Then: create array of component items
                        {
                            "$map": {
                                "input": "$composite_info.components",
                                "as": "component",
                                "in": {
                                    "doc_type": "$doc_type",
                                    "invoice_number": "$invoice_number",
                                    "creditnote_number": "$creditnote_number",
                                    "item_id": "$$component.item_id",
                                    "item_name": "$$component.name",
                                    # Multiply original quantity by component quantity
                                    "quantity": {
                                        "$multiply": [
                                            "$line_items.quantity",
                                            "$$component.quantity",
                                        ]
                                    },
                                    # Distribute amount proportionally
                                    "item_total": {
                                        "$divide": [
                                            "$line_items.item_total",
                                            {"$size": "$composite_info.components"},
                                        ]
                                    },
                                },
                            }
                        },
                        # Else: keep as single regular item
                        [
                            {
                                "doc_type": "$doc_type",
                                "invoice_number": "$invoice_number",
                                "creditnote_number": "$creditnote_number",
                                "item_id": "$line_items.item_id",
                                "item_name": "$line_items.name",
                                "quantity": "$line_items.quantity",
                                "item_total": "$line_items.item_total",
                            }
                        ],
                    ]
                }
            }
        },
        {"$unwind": "$items_to_process"},
        {"$replaceRoot": {"newRoot": "$items_to_process"}},
        {
            "$group": {
                "_id": "$item_id",
                "item_name": {"$first": "$item_name"},
                "invoice_units": {
                    "$sum": {
                        "$cond": [
                            {"$eq": ["$doc_type", "invoice"]},
                            "$quantity",
                            0,
                        ]
                    }
                },
                "credit_note_units": {
                    "$sum": {
                        "$cond": [
                            {"$eq": ["$doc_type", "credit_note"]},
                            "$quantity",
                            0,
                        ]
                    }
                },
                "invoice_amount": {
                    "$sum": {
                        "$cond": [
                            {"$eq": ["$doc_type", "invoice"]},
                            "$item_total",
                            0,
                        ]
                    }
                },
                "credit_note_amount": {
                    "$sum": {
                        "$cond": [
                            {"$eq": ["$doc_type", "credit_note"]},
                            "$item_total",
                            0,
                        ]
                    }
                },
                "invoice_numbers": {
                    "$addToSet": {
                        "$cond": [
                            {"$eq": ["$doc_type", "invoice"]},
                            "$invoice_number",
                            None,
                        ]
                    }
                },
                "credit_note_numbers": {
                    "$addToSet": {
                        "$cond": [
                            {"$eq": ["$doc_type", "credit_note"]},
                            "$creditnote_number",
                            None,
                        ]
                    }
                },
            }
        },
        {
            "$project": {
                "_id": 1,
                "item_name": 1,
                "invoice_units": 1,
                "credit_note_units": 1,
                "invoice_amount": 1,
                "credit_note_amount": 1,
                "total_units_sold": "$invoice_units",
                "total_units_returned": "$credit_note_units",
                "total_amount": "$invoice_amount",
                "invoice_numbers": {
                    "$filter": {
                        "input": "$invoice_numbers",
                        "cond": {"$ne": ["$$this", None]},
                    }
                },
                "credit_note_numbers": {
                    "$filter": {
                        "input": "$credit_note_numbers",
                        "cond": {"$ne": ["$$this", None]},
                    }
                },
            }
        },
    ]


def process_batch_with_credit_notes(
    batch: List[Dict], stock_data: Dict, products_map: Dict
) -> Dict:
    """
    Process a batch of items that includes net calculations from invoices and credit notes.
    Maintains the same output structure as the original process_batch function.
    """
    items = []
    total_units = 0
    total_amount = 0
    total_stock = 0
    total_returns = 0
    for item in batch:
        item_id = item["_id"]

        # Net quantities (already calculated in the pipeline)
        units_sold = item.get("total_units_sold", 0)
        amount = item.get("total_amount", 0.0)

        # Additional data for debugging/logging (can be removed in production)
        invoice_units = item.get("invoice_units", 0)
        credit_note_units = item.get("credit_note_units", 0)
        invoice_numbers = item.get("invoice_numbers", [])
        credit_note_numbers = item.get("credit_note_numbers", [])
        # Get SKU and stock data (same as before)
        sku_code = products_map.get(item_id, "")
        stock_info = stock_data.get(item_id, {})
        closing_stock = stock_info.get("closing_stock", 0)
        days_in_stock = stock_info.get("total_days_in_stock", 0)
        days_in_stock_any_wh = stock_info.get("total_days_in_stock_any_wh", 0)
        last_90_days_dates = stock_info.get(
            "last_90_days_dates", ""
        )  # Get the new field
        drr = round(units_sold / days_in_stock, 2) if days_in_stock > 0 else 0

        # Only accumulate totals for items with names (same filtering as before)
        if item.get("item_name") != "":
            total_units += units_sold
            total_amount += amount
            total_stock += closing_stock
            total_returns += credit_note_units
            # Log credit note activity for items that have it (optional, only for debugging)
            if credit_note_units > 0:
                logger.debug(
                    f"Item {item.get('item_name')} - Invoice units: {invoice_units}, "
                    f"Credit note units: {credit_note_units}, Net units: {units_sold}"
                )

            items.append(
                SalesReportItem(
                    item_name=item.get("item_name") or "",
                    sku_code=sku_code,
                    units_returned=credit_note_units,
                    credit_notes=credit_note_units,
                    units_sold=units_sold,  # This is now net units (invoice - credit notes)
                    total_amount=round(amount, 2),  # This is now net amount
                    closing_stock=closing_stock,
                    total_days_in_stock=days_in_stock,
                    total_days_in_stock_any_wh=days_in_stock_any_wh,
                    drr=drr,
                    last_90_days_dates=last_90_days_dates,  # Add the new field
                )
            )

    return {
        "items": items,
        "units": total_units,
        "amount": total_amount,
        "stock": total_stock,
        "returns": total_returns,
    }


@router.get("/sales-report")
async def get_sales_report_fast(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    any_last_90_days: bool = Query(
        False, description="Include last 90 days in stock with date ranges"
    ),
    exclude_customers: bool = True,
    db=Depends(get_database),
):
    """
    Ultra-optimized sales report that incorporates credit notes.
    Net sales = Invoice quantities - Credit note quantities
    Response structure remains the same for frontend compatibility.

    When any_last_90_days=True, includes a new field 'last_90_days_dates' with formatted date ranges
    showing the last 90 days the item was in stock, regardless of when those days occurred.
    """

    try:
        # Validate dates
        start_datetime = datetime.strptime(start_date, "%Y-%m-%d")
        end_datetime = datetime.strptime(end_date, "%Y-%m-%d")

        if start_datetime > end_datetime:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="End date must be after start date",
            )

        logger.info(
            f"Generating sales report with credit notes for {start_date} to {end_date}, any_last_90_days={any_last_90_days}"
        )
        start_time = datetime.now()

        # Use efficient customer filtering
        excluded_customers = get_excluded_customer_list() if exclude_customers else None

        # Build the pipeline (pure Python — instant)
        invoices_collection = db[INVOICES_COLLECTION]
        pipeline = build_optimized_pipeline_with_credit_notes_and_composites(
            start_date, end_date, excluded_customers
        )

        # Run all three heavy fetches in parallel:
        #   - Invoice aggregation (heavy: $unionWith + $unwind + $lookup + $group)
        #   - Stock data aggregation (heavy: 4-5M records)
        #   - Products map (moderate)
        # Previously the invoice pipeline only started AFTER stock+products finished,
        # wasting the time they overlapped.
        def _fetch_raw_invoice_docs():
            """Run the invoice/credit-note aggregation and return all results as a list.
            The final $group reduces output to one doc per item_id (~hundreds to low thousands),
            so list() is safe here."""
            cursor = invoices_collection.aggregate(
                pipeline,
                allowDiskUse=True,
                batchSize=5000,
            )
            return list(cursor)

        invoice_task = asyncio.create_task(asyncio.to_thread(_fetch_raw_invoice_docs))
        stock_task = asyncio.create_task(
            fetch_stock_data_optimized(db, start_datetime, end_datetime)
        )
        products_task = asyncio.create_task(fetch_all_products_indexed(db))

        logger.info(
            "Executing invoice aggregation, stock fetch, and products fetch in parallel..."
        )
        raw_invoice_docs, stock_data, products_map = await asyncio.gather(
            invoice_task, stock_task, products_task
        )

        # If any_last_90_days is requested, fetch the last 90 days in stock for all items.
        # This is sequential (needs stock_data keys first) but rarely used.
        if any_last_90_days and stock_data:
            item_ids = list(stock_data.keys())
            logger.info(f"Fetching last 90 days in stock for {len(item_ids)} items")
            logger.debug(f"DEBUG: Sample stock_data keys: {item_ids[:3]}")

            last_90_days_data = await fetch_last_n_days_in_stock(
                db, end_datetime, item_ids, 90
            )
            logger.debug(
                f"DEBUG: Received {len(last_90_days_data)} items from fetch_last_n_days_in_stock"
            )

            merged_count = 0
            for item_id, date_ranges in last_90_days_data.items():
                if item_id in stock_data:
                    stock_data[item_id]["last_90_days_dates"] = date_ranges
                    merged_count += 1

            logger.debug(
                f"DEBUG: Merged {merged_count} items with last_90_days_dates into stock_data"
            )

        logger.info(
            f"Processing {len(raw_invoice_docs)} invoice docs, "
            f"{len(stock_data)} stock items, {len(products_map)} products"
        )

        # Process all buffered results in a thread (CPU-bound, fast — data already fetched)
        def _process_all_results():
            processed = process_batch_with_credit_notes(
                raw_invoice_docs, stock_data, products_map
            )
            items = processed["items"]
            items.sort(key=lambda x: x.item_name)
            return (
                items,
                processed["units"],
                processed["returns"],
                processed["amount"],
                processed["stock"],
            )

        (
            sales_report_items,
            total_units,
            total_returns,
            total_amount,
            total_closing_stock,
        ) = await asyncio.to_thread(_process_all_results)

        execution_time = (datetime.now() - start_time).total_seconds()
        logger.info(
            f"Sales report with credit notes generated in {execution_time:.2f} seconds with {len(sales_report_items)} items"
        )

        # Calculate summary (same logic as before, but now with net values)
        avg_drr = (
            sum(item.drr for item in sales_report_items) / len(sales_report_items)
            if sales_report_items
            else 0
        )
        items_with_stock = sum(
            1 for item in sales_report_items if item.closing_stock > 0
        )
        items_out_of_stock = sum(
            1 for item in sales_report_items if item.closing_stock == 0
        )

        # Prepare response (SAME STRUCTURE as before - no frontend changes needed)
        response_data = {
            "data": [item.dict() for item in sales_report_items],
            "summary": {
                "total_items": len(sales_report_items),
                "total_units_sold": total_units,  # Now net units (invoice - credit notes)
                "total_units_returned": total_returns,  # Now net units (invoice - credit notes)
                "total_amount": round(total_amount, 2),  # Now net amount
                "total_closing_stock": total_closing_stock,
                "average_drr": round(avg_drr, 2),
                "items_with_stock": items_with_stock,
                "items_out_of_stock": items_out_of_stock,
                "date_range": {"start_date": start_date, "end_date": end_date},
            },
            "meta": {
                "timestamp": datetime.now().isoformat(),
                "execution_time_seconds": round(execution_time, 2),
                "query_type": "ultra_fast_sales_report_with_credit_notes",
                "from_cache": False,
                "optimizations": [
                    "parallel_aggregations",
                    "indexed_products",
                    "batch_processing",
                    "optimized_customer_filter",
                    "credit_notes_integration",  # NEW
                ],
                "credit_notes_integrated": True,  # NEW field to indicate credit notes are included
            },
        }

        return JSONResponse(content=response_data)

    except Exception as e:
        logger.error(f"Error in sales report with credit notes: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e)
        )


@router.get("/sales-report/download")
async def download_sales_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    any_last_90_days: bool = Query(
        False, description="Include last 90 days in stock with date ranges"
    ),
    db=Depends(get_database),
):
    """
    Download sales report as XLSX file with credit notes integration.
    Net sales = Invoice quantities - Credit note quantities

    When any_last_90_days=True, includes a new column 'Last 90 Days In Stock (Dates)' with formatted date ranges
    showing the last 90 days the item was in stock, regardless of when those days occurred.
    """

    try:
        # Validate dates
        start_datetime = datetime.strptime(start_date, "%Y-%m-%d")
        end_datetime = datetime.strptime(end_date, "%Y-%m-%d")

        if start_datetime > end_datetime:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="End date must be after start date",
            )

        logger.info(
            f"Generating sales report download with credit notes for {start_date} to {end_date}, any_last_90_days={any_last_90_days}"
        )
        start_time = datetime.now()

        # Use the SAME optimized approach as the main API
        stock_task = asyncio.create_task(
            fetch_stock_data_optimized(db, start_datetime, end_datetime)
        )
        products_task = asyncio.create_task(fetch_all_products_indexed(db))

        excluded_customers = get_excluded_customer_list()
        invoices_collection = db[INVOICES_COLLECTION]

        # Use the NEW pipeline with credit notes
        pipeline = build_optimized_pipeline_with_credit_notes_and_composites(
            start_date, end_date, excluded_customers
        )

        # Process exactly the same way as the main API
        stock_data, products_map = await asyncio.gather(stock_task, products_task)

        # If any_last_90_days is requested, fetch the last 90 days in stock for all items
        if any_last_90_days and stock_data:
            item_ids = list(stock_data.keys())
            logger.info(
                f"Fetching last 90 days in stock for download: {len(item_ids)} items"
            )
            logger.debug(f"DEBUG (download): Sample stock_data keys: {item_ids[:3]}")

            last_90_days_data = await fetch_last_n_days_in_stock(
                db, end_datetime, item_ids, 90
            )
            logger.debug(
                f"DEBUG (download): Received {len(last_90_days_data)} items from fetch_last_n_days_in_stock"
            )

            # Merge into stock_data
            merged_count = 0
            for item_id, date_ranges in last_90_days_data.items():
                if item_id in stock_data:
                    stock_data[item_id]["last_90_days_dates"] = date_ranges
                    merged_count += 1

            logger.debug(
                f"DEBUG (download): Merged {merged_count} items with last_90_days_dates into stock_data"
            )

        # Run cursor creation + iteration in thread to avoid blocking
        def _process_download_cursor(collection, pipeline, stock_data, products_map):
            result_cursor = collection.aggregate(
                pipeline,
                allowDiskUse=True,
                batchSize=1000,
            )
            sales_report_items = []
            total_units = 0
            total_returns = 0
            total_amount = 0.0
            total_closing_stock = 0

            batch = []
            batch_size = 500

            for item in result_cursor:
                batch.append(item)
                if len(batch) >= batch_size:
                    processed = process_batch_with_credit_notes(
                        batch, stock_data, products_map
                    )
                    sales_report_items.extend(processed["items"])
                    total_units += processed["units"]
                    total_returns += processed["returns"]
                    total_amount += processed["amount"]
                    total_closing_stock += processed["stock"]
                    batch = []

            if batch:
                processed = process_batch_with_credit_notes(
                    batch, stock_data, products_map
                )
                sales_report_items.extend(processed["items"])
                total_units += processed["units"]
                total_returns += processed["returns"]
                total_amount += processed["amount"]
                total_closing_stock += processed["stock"]

            sales_report_items.sort(key=lambda x: x.item_name)
            return (
                sales_report_items,
                total_units,
                total_returns,
                total_amount,
                total_closing_stock,
            )

        (
            sales_report_items,
            total_units,
            total_returns,
            total_amount,
            total_closing_stock,
        ) = await asyncio.to_thread(
            _process_download_cursor,
            invoices_collection,
            pipeline,
            stock_data,
            products_map,
        )

        execution_time = (datetime.now() - start_time).total_seconds()
        logger.info(
            f"Download data with credit notes prepared in {execution_time:.2f} seconds with {len(sales_report_items)} items"
        )

        # Create Excel file
        output = io.BytesIO()

        # Convert SalesReportItem objects to dictionaries for DataFrame
        sales_data = []
        for item in sales_report_items:
            row_data = {
                "Item Name": item.item_name,
                "SKU Code": item.sku_code,
                "Units Sold": item.units_sold,
                "Units Returned": item.units_returned,
                "Total Amount (₹)": item.total_amount,
                "Closing Stock": item.closing_stock,
                "Days In Stock": item.total_days_in_stock,
                "DRR (Daily Run Rate)": item.drr,
            }
            # Add the Last 90 Days In Stock column only if requested
            if any_last_90_days:
                row_data["Last 90 Days In Stock (Dates)"] = item.last_90_days_dates
            sales_data.append(row_data)

        # Create DataFrame
        df = pd.DataFrame(sales_data)

        # Create Excel file with multiple sheets
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Main data sheet
            df.to_excel(writer, sheet_name="Sales Report", index=False)

            # Summary sheet
            summary_data = {
                "Metric": [
                    "Report Period",
                    "Generated On",
                    "Processing Method",
                    "Credit Notes Integrated",
                    "",
                    "NET SALES SUMMARY",
                    "Total Items",
                    "Net Units Sold",
                    "Net Units Returned",
                    "Net Amount (₹)",
                    "Total Closing Stock",
                    "Items With Stock",
                    "Items Out of Stock",
                    "Average DRR",
                ],
                "Value": [
                    f"{start_date} to {end_date}",
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "Optimized with Credit Notes Integration",
                    "Yes",
                    "",
                    "",
                    len(sales_report_items),
                    total_units,
                    total_returns,
                    f"₹{total_amount:,.2f}",
                    total_closing_stock,
                    sum(1 for item in sales_report_items if item.closing_stock > 0),
                    sum(1 for item in sales_report_items if item.closing_stock == 0),
                    (
                        f"{sum(item.drr for item in sales_report_items) / len(sales_report_items):.2f}"
                        if sales_report_items
                        else "0.00"
                    ),
                ],
            }

            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

            # Format the main sheet
            workbook = writer.book
            worksheet = writer.sheets["Sales Report"]

            # Auto-adjust column widths
            for column in df:
                column_length = max(df[column].astype(str).map(len).max(), len(column))
                col_idx = df.columns.get_loc(column)
                # Convert column index to Excel column letter
                col_letter = (
                    chr(65 + col_idx)
                    if col_idx < 26
                    else chr(65 + col_idx // 26 - 1) + chr(65 + col_idx % 26)
                )
                worksheet.column_dimensions[col_letter].width = min(
                    column_length + 2, 50
                )

        output.seek(0)

        # Generate filename with credit notes indicator
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"net_sales_report_{start_date}_to_{end_date}_{timestamp}.xlsx"

        logger.info(f"Excel file generated successfully: {filename}")

        # Return the Excel file as streaming response
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except HTTPException as e:
        raise e
    except Exception as e:
        logger.error(f"Error generating sales report download with credit notes: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An unexpected error occurred: {str(e)}",
        )


from functools import lru_cache
from typing import Optional, Dict, Any
import asyncio


@router.get("/data-metadata")
async def get_data_metadata(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    db=Depends(get_database),
):
    """
    Get metadata for the SELECTED date range (not overall data availability).
    """

    try:
        # Validate dates
        try:
            start_datetime = datetime.strptime(start_date, "%Y-%m-%d")
            end_datetime = datetime.strptime(end_date, "%Y-%m-%d")

            if start_datetime > end_datetime:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="End date must be after start date",
                )
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid date format. Use YYYY-MM-DD",
            )

        # Get collections
        invoices_collection = db[INVOICES_COLLECTION]
        stock_collection = db["zoho_warehouse_stock"]

        # Query inventory and sales data for the SELECTED date range
        inventory_count = await asyncio.to_thread(
            stock_collection.count_documents,
            {"date": {"$gte": start_datetime, "$lte": end_datetime}},
        )

        invoice_count = await asyncio.to_thread(
            invoices_collection.count_documents,
            {"date": {"$gte": start_date, "$lte": end_date}},
        )

        # Get distinct dates for inventory
        inventory_distinct_dates = await asyncio.to_thread(
            lambda: list(
                stock_collection.distinct(
                    "date", {"date": {"$gte": start_datetime, "$lte": end_datetime}}
                )
            )
        )
        inventory_dates_set = set(
            d.strftime("%Y-%m-%d") if isinstance(d, datetime) else d
            for d in inventory_distinct_dates
        )

        # Get distinct dates for sales
        sales_distinct_dates = await asyncio.to_thread(
            lambda: list(
                invoices_collection.distinct(
                    "date", {"date": {"$gte": start_date, "$lte": end_date}}
                )
            )
        )
        sales_dates_set = set(
            d if isinstance(d, str) else d.strftime("%Y-%m-%d")
            for d in sales_distinct_dates
        )

        # Calculate all dates in the range
        from datetime import timedelta

        all_dates = set()
        current_date = start_datetime
        while current_date <= end_datetime:
            all_dates.add(current_date.strftime("%Y-%m-%d"))
            current_date += timedelta(days=1)

        # Find missing dates
        missing_inventory_dates = sorted(all_dates - inventory_dates_set)
        missing_sales_dates = sorted(all_dates - sales_dates_set)

        # Return metadata for selected date range
        metadata = {
            "inventory_data": {
                "first_inventory_date": start_date,
                "last_inventory_date": end_date,
                "total_stock_records": inventory_count,
                "missing_dates": missing_inventory_dates,
            },
            "sales_data": {
                "first_sales_date": start_date,
                "last_sales_date": end_date,
                "valid_invoices": invoice_count,
                "missing_dates": missing_sales_dates,
            },
            "date_range": {
                "start_date": start_date,
                "end_date": end_date,
                "filtered": True,
            },
            "last_updated": datetime.now().isoformat(),
        }

        return JSONResponse(
            content={
                "data": metadata,
                "meta": {
                    "timestamp": datetime.now().isoformat(),
                    "query_type": "date_range_filtered_metadata",
                    "date_filtered": True,
                },
            }
        )

    except HTTPException as e:
        raise e
    except Exception as e:
        logger.error(f"Error fetching data metadata: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An unexpected error occurred: {str(e)}",
        )


async def get_sales_metadata_optimized(
    collection,
    start_date: Optional[str],
    end_date: Optional[str],
    start_datetime: Optional[datetime],
    end_datetime: Optional[datetime],
) -> Dict[str, Any]:
    """
    Optimized sales metadata retrieval using a single aggregation pipeline.
    """
    try:
        # OPTIMIZATION 1: Single pipeline for all sales metrics
        pipeline = []

        # If dates are consistently formatted as YYYY-MM-DD strings, use string comparison
        if start_date and end_date:
            # Try string comparison first (faster)
            pipeline.append(
                {"$match": {"date": {"$gte": start_date, "$lte": end_date}}}
            )

        # Add facet for parallel processing of different metrics
        pipeline.append(
            {
                "$facet": {
                    # Get total count
                    "total_count": [{"$count": "count"}],
                    # Get valid invoices with dates
                    "valid_invoices": [
                        {"$match": {"status": {"$nin": ["draft", "void"]}}},
                        {
                            "$group": {
                                "_id": None,
                                "count": {"$sum": 1},
                                "first_date": {"$min": "$date"},
                                "last_date": {"$max": "$date"},
                            }
                        },
                    ],
                    # Get date range for all invoices
                    "date_range": [
                        {
                            "$group": {
                                "_id": None,
                                "min_date": {"$min": "$date"},
                                "max_date": {"$max": "$date"},
                            }
                        }
                    ],
                }
            }
        )

        # Execute the pipeline
        result = list(collection.aggregate(pipeline, allowDiskUse=True))

        if result:
            facet_result = result[0]

            # Extract total count
            total_invoices = (
                facet_result["total_count"][0]["count"]
                if facet_result.get("total_count")
                else 0
            )

            # Extract valid invoices data
            valid_data = facet_result.get("valid_invoices", [{}])[0]
            valid_count = valid_data.get("count", 0)

            # Parse dates if they're strings
            first_date = valid_data.get("first_date")
            last_date = valid_data.get("last_date")

            # Convert string dates to datetime if needed
            if first_date and isinstance(first_date, str):
                try:
                    first_date = datetime.strptime(first_date, "%Y-%m-%d")
                except:
                    pass

            if last_date and isinstance(last_date, str):
                try:
                    last_date = datetime.strptime(last_date, "%Y-%m-%d")
                except:
                    pass

            return {
                "first_sales_date": first_date.isoformat() if first_date else None,
                "last_sales_date": last_date.isoformat() if last_date else None,
                "total_invoices": total_invoices,
                "valid_invoices": valid_count,
            }

        # If no results, get basic counts
        total_invoices = await run_in_executor(collection.count_documents, {})
        valid_invoices = await run_in_executor(
            collection.count_documents, {"status": {"$nin": ["draft", "void"]}}
        )

        return {
            "first_sales_date": None,
            "last_sales_date": None,
            "total_invoices": total_invoices,
            "valid_invoices": valid_invoices,
        }

    except Exception as e:
        logger.error(f"Error in optimized sales metadata: {e}")
        return {
            "first_sales_date": None,
            "last_sales_date": None,
            "total_invoices": 0,
            "valid_invoices": 0,
        }


async def get_inventory_metadata_optimized(
    collection,
    start_datetime: Optional[datetime] = None,
    end_datetime: Optional[datetime] = None,
) -> Dict[str, Any]:
    """
    Optimized inventory metadata retrieval.
    Returns the actual date range available in the collection (not filtered by user's date selection).
    Note: start_datetime and end_datetime parameters are kept for API compatibility but not used.
    """
    try:
        # Don't filter by date - show the actual available data range
        # OPTIMIZATION: Single aggregation for all metrics
        pipeline = [
            {"$match": {}},  # No date filter - get all available data
            {
                "$group": {
                    "_id": None,
                    "count": {"$sum": 1},
                    "first_date": {"$min": "$date"},
                    "last_date": {"$max": "$date"},
                }
            },
        ]

        result = list(collection.aggregate(pipeline, allowDiskUse=True))

        if result:
            data = result[0]
            return {
                "first_inventory_date": (
                    data["first_date"].isoformat() if data.get("first_date") else None
                ),
                "last_inventory_date": (
                    data["last_date"].isoformat() if data.get("last_date") else None
                ),
                "total_stock_records": data.get("count", 0),
            }

        return {
            "first_inventory_date": None,
            "last_inventory_date": None,
            "total_stock_records": 0,
        }

    except Exception as e:
        logger.error(f"Error in optimized inventory metadata: {e}")
        return {
            "first_inventory_date": None,
            "last_inventory_date": None,
            "total_stock_records": 0,
        }


# Helper function for async execution
async def run_in_executor(func, *args):
    """Run blocking function in executor."""
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, func, *args)


# CACHED VERSION for frequently accessed metadata
@router.get("/data-metadata-cached")
async def get_data_metadata_cached(
    start_date: str = Query(
        None, description="Start date in YYYY-MM-DD format (optional)"
    ),
    end_date: str = Query(None, description="End date in YYYY-MM-DD format (optional)"),
    db=Depends(get_database),
    force_refresh: bool = Query(False, description="Force refresh cache"),
):
    """
    Cached version of data metadata for better performance.
    Cache is refreshed every 5 minutes or on demand.
    """

    cache_key = f"metadata_{start_date}_{end_date}"

    # Check if we need to refresh cache
    if force_refresh or not hasattr(get_data_metadata_cached, "_cache"):
        get_data_metadata_cached._cache = {}
        get_data_metadata_cached._cache_time = {}

    # Check if cached data exists and is fresh (5 minutes)
    current_time = datetime.now()
    cache_time = get_data_metadata_cached._cache_time.get(cache_key)

    if (
        not force_refresh
        and cache_key in get_data_metadata_cached._cache
        and cache_time
        and (current_time - cache_time).seconds < 300
    ):

        logger.info(f"Returning cached metadata for key: {cache_key}")
        cached_data = get_data_metadata_cached._cache[cache_key]
        cached_data["meta"]["from_cache"] = True
        cached_data["meta"]["cache_age_seconds"] = (current_time - cache_time).seconds
        return JSONResponse(content=cached_data)

    # Get fresh data
    logger.info(f"Fetching fresh metadata for key: {cache_key}")
    response = await get_data_metadata(start_date, end_date, db)

    # Parse response and cache it
    if hasattr(response, "body"):
        content = json.loads(response.body)
        get_data_metadata_cached._cache[cache_key] = content
        get_data_metadata_cached._cache_time[cache_key] = current_time
        content["meta"]["from_cache"] = False
        return JSONResponse(content=content)

    return response


@router.get("/estimates-vs-invoices")
def get_estimates_vs_invoices(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    db=Depends(get_database),
):
    """
    Compare estimates and invoices line items to calculate fill rate per item.

    Composite items are automatically expanded into their individual components.
    For example, if an estimate has 2 units of a combo product containing:
    - 1x Item A
    - 2x Item B

    This will be expanded to:
    - 2 units of Item A
    - 4 units of Item B

    Args:
        start_date: Start date for the report
        end_date: End date for the report

    Returns:
        List of items with estimated quantities, invoiced quantities, and fill rate percentage
    """
    try:
        # Validate date format
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            end = datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(
                status_code=400, detail="Invalid date format. Use YYYY-MM-DD format."
            )

        if start > end:
            raise HTTPException(
                status_code=400, detail="Start date must be before or equal to end date"
            )

        # Get collections
        estimates_collection = db["estimates"]
        invoices_collection = db["invoices"]
        # Query estimates within date range - only invoiced estimates
        # Handle both string and Date object formats in the database
        estimates_query = {
            "customer_name": {"$nin": EXCLUDED_CUSTOMERS_LIST},
            "status": "invoiced",  # Only invoiced estimates
            "$or": [
                {"date": {"$gte": start, "$lte": end}},  # Date objects
                {"date": {"$gte": start_date, "$lte": end_date}},  # Strings
            ],
        }

        # Query invoices within date range
        # Handle both string and Date object formats in the database
        invoices_query = {
            "customer_name": {"$nin": EXCLUDED_CUSTOMERS_LIST},
            "status": {"$nin": ["void"]},
            "$or": [
                {"date": {"$gte": start, "$lte": end}},  # Date objects
                {"date": {"$gte": start_date, "$lte": end_date}},  # Strings
            ],
        }

        logger.info(f"Fetching estimates from {start_date} to {end_date}")
        logger.info(f"Estimates query: {estimates_query}")

        # Use aggregation to deduplicate estimates and handle composite items
        estimates_pipeline = [
            {"$match": estimates_query},
            {"$sort": {"_id": -1}},  # Sort by _id descending (latest first)
            {
                "$group": {
                    "_id": "$estimate_id",  # Group by estimate_id to deduplicate
                    "latest": {"$first": "$$ROOT"},  # Take first (latest) document
                }
            },
            {
                "$replaceRoot": {"newRoot": "$latest"}
            },  # Replace root with the latest document
            {"$unwind": "$line_items"},  # Unwind line items to process each one
            # Lookup composite products by item_id
            {
                "$lookup": {
                    "from": "composite_products",
                    "localField": "line_items.item_id",
                    "foreignField": "composite_item_id",
                    "as": "composite_match",
                }
            },
            # Check if item is composite
            {
                "$addFields": {
                    "is_composite": {
                        "$or": [
                            {"$gt": [{"$size": "$composite_match"}, 0]},
                            {"$eq": ["$line_items.is_combo_product", True]},
                        ]
                    },
                    "composite_info": {"$arrayElemAt": ["$composite_match", 0]},
                }
            },
            # Expand composite items or keep regular items
            {
                "$project": {
                    "items_to_process": {
                        "$cond": [
                            # If it's a composite item with components found
                            {"$gt": [{"$size": "$composite_match"}, 0]},
                            # Then: create array of component items
                            {
                                "$map": {
                                    "input": "$composite_info.components",
                                    "as": "component",
                                    "in": {
                                        "item_id": "$$component.item_id",
                                        "name": "$$component.name",
                                        "sku": "$$component.sku_code",
                                        # Multiply original quantity by component quantity
                                        "quantity": {
                                            "$multiply": [
                                                "$line_items.quantity",
                                                "$$component.quantity",
                                            ]
                                        },
                                        "item_custom_fields": [],
                                    },
                                }
                            },
                            # Else: keep as single regular item
                            [
                                {
                                    "item_id": "$line_items.item_id",
                                    "name": "$line_items.name",
                                    "sku": "$line_items.sku",
                                    "quantity": "$line_items.quantity",
                                    "item_custom_fields": "$line_items.item_custom_fields",
                                }
                            ],
                        ]
                    }
                }
            },
            {"$unwind": "$items_to_process"},
            {"$replaceRoot": {"newRoot": "$items_to_process"}},
        ]
        estimates = list(
            estimates_collection.aggregate(
                estimates_pipeline, allowDiskUse=True, batchSize=1000
            )
        )

        logger.info(f"Fetching invoices from {start_date} to {end_date}")

        # Use aggregation to deduplicate invoices and handle composite items
        invoices_pipeline = [
            {"$match": invoices_query},
            {"$sort": {"_id": -1}},  # Sort by _id descending (latest first)
            {
                "$group": {
                    "_id": "$invoice_id",  # Group by invoice_id to deduplicate
                    "latest": {"$first": "$$ROOT"},  # Take first (latest) document
                }
            },
            {
                "$replaceRoot": {"newRoot": "$latest"}
            },  # Replace root with the latest document
            {"$unwind": "$line_items"},  # Unwind line items to process each one
            # Lookup composite products by item_id
            {
                "$lookup": {
                    "from": "composite_products",
                    "localField": "line_items.item_id",
                    "foreignField": "composite_item_id",
                    "as": "composite_match",
                }
            },
            # Check if item is composite
            {
                "$addFields": {
                    "is_composite": {
                        "$or": [
                            {"$gt": [{"$size": "$composite_match"}, 0]},
                            {"$eq": ["$line_items.is_combo_product", True]},
                        ]
                    },
                    "composite_info": {"$arrayElemAt": ["$composite_match", 0]},
                }
            },
            # Expand composite items or keep regular items
            {
                "$project": {
                    "items_to_process": {
                        "$cond": [
                            # If it's a composite item with components found
                            {"$gt": [{"$size": "$composite_match"}, 0]},
                            # Then: create array of component items
                            {
                                "$map": {
                                    "input": "$composite_info.components",
                                    "as": "component",
                                    "in": {
                                        "item_id": "$$component.item_id",
                                        "name": "$$component.name",
                                        "sku": "$$component.sku_code",
                                        # Multiply original quantity by component quantity
                                        "quantity": {
                                            "$multiply": [
                                                "$line_items.quantity",
                                                "$$component.quantity",
                                            ]
                                        },
                                        "item_custom_fields": [],
                                    },
                                }
                            },
                            # Else: keep as single regular item
                            [
                                {
                                    "item_id": "$line_items.item_id",
                                    "name": "$line_items.name",
                                    "sku": "$line_items.sku",
                                    "quantity": "$line_items.quantity",
                                    "item_custom_fields": "$line_items.item_custom_fields",
                                }
                            ],
                        ]
                    }
                }
            },
            {"$unwind": "$items_to_process"},
            {"$replaceRoot": {"newRoot": "$items_to_process"}},
        ]
        invoices = list(
            invoices_collection.aggregate(
                invoices_pipeline, allowDiskUse=True, batchSize=1000
            )
        )

        logger.info(
            f"Found {len(estimates)} line items from estimates and {len(invoices)} line items from invoices (composite items expanded)"
        )

        # Debug logging for specific item
        logger.info(f"Processing aggregated line items with composite expansion...")

        # Get warehouse stock collection
        warehouse_stock_collection = db["zoho_warehouse_stock"]

        # Helper function to extract cf_sku_code from item_custom_fields
        def get_cf_sku_code(item):
            item_custom_fields = item.get("item_custom_fields", [])
            for field in item_custom_fields:
                if field.get("api_name") == "cf_sku_code":
                    return field.get("value", "")
            return item.get("sku", "")

        # Aggregate line items from estimates (already expanded by pipeline)
        estimate_items = {}
        debug_item_name = ""
        debug_total = 0

        for item in estimates:
            item_id = item.get("item_id")
            if not item_id:
                continue

            quantity = item.get("quantity", 0)
            item_name = item.get("name", "")

            # Debug logging for specific item
            if debug_item_name in item_name:
                debug_total += quantity
                logger.info(
                    f"Found estimate for {item_name}: qty={quantity}, running total={debug_total}"
                )

            if item_id not in estimate_items:
                estimate_items[item_id] = {
                    "item_id": item_id,
                    "item_name": item_name,
                    "sku": get_cf_sku_code(item),
                    "estimated_quantity": 0,
                }

            estimate_items[item_id]["estimated_quantity"] += quantity

        logger.info(f"Total estimated quantity for {debug_item_name}: {debug_total}")

        # Aggregate line items from invoices (already expanded by pipeline)
        invoice_items = {}
        for item in invoices:
            item_id = item.get("item_id")
            if not item_id:
                continue

            quantity = item.get("quantity", 0)

            if item_id not in invoice_items:
                invoice_items[item_id] = {
                    "item_id": item_id,
                    "item_name": item.get("name", ""),
                    "sku": get_cf_sku_code(item),
                    "invoiced_quantity": 0,
                }

            invoice_items[item_id]["invoiced_quantity"] += quantity

        # Get closing stock for all items on end_date
        all_item_ids = set(estimate_items.keys()) | set(invoice_items.keys())

        # Convert end_date string to datetime for comparison
        end_datetime = datetime.combine(end, datetime.min.time())

        logger.info(f"Fetching closing stock for {len(all_item_ids)} items")

        # Use aggregation pipeline to efficiently get the latest stock for each item
        closing_stock_map = {}
        stock_date = None

        try:
            pipeline = [
                # Match only items we care about and dates on or before end_date
                {
                    "$match": {
                        "zoho_item_id": {"$in": list(all_item_ids)},
                        "date": {"$lte": end_datetime},
                    }
                },
                # Sort by date descending to get latest first
                {"$sort": {"date": -1}},
                # Group by item and take the first (latest) record
                {
                    "$group": {
                        "_id": "$zoho_item_id",
                        "warehouses": {"$first": "$warehouses"},
                        "date": {"$first": "$date"},
                    }
                },
            ]

            warehouse_stocks = list(warehouse_stock_collection.aggregate(pipeline))

            logger.info(f"Found stock data for {len(warehouse_stocks)} items")

            # Create map of item_id to total closing stock and capture the stock date
            for stock in warehouse_stocks:
                item_id = stock.get("_id")
                if item_id:
                    warehouses = stock.get("warehouses", {})
                    total_stock = (
                        sum(warehouses.values()) if isinstance(warehouses, dict) else 0
                    )
                    closing_stock_map[item_id] = total_stock

                    # Capture stock date (all should have the same or similar date)
                    if stock_date is None and stock.get("date"):
                        stock_date = (
                            stock.get("date").strftime("%Y-%m-%d")
                            if hasattr(stock.get("date"), "strftime")
                            else str(stock.get("date"))
                        )

        except Exception as e:
            logger.error(f"Error fetching warehouse stock: {str(e)}")
            # Continue without stock data if there's an error
            pass

        # Combine and calculate fill rate
        result = []
        missed_items = 0
        over_delivered_items = 0

        for item_id in all_item_ids:
            estimate_data = estimate_items.get(item_id, {})
            invoice_data = invoice_items.get(item_id, {})

            estimated_qty = estimate_data.get("estimated_quantity", 0)
            invoiced_qty = invoice_data.get("invoiced_quantity", 0)
            closing_stock = closing_stock_map.get(item_id, 0)

            # Calculate fill rate
            fill_rate = 0
            if estimated_qty > 0:
                fill_rate = (invoiced_qty / estimated_qty) * 100

            # Count missed vs over-delivered
            if invoiced_qty < estimated_qty:
                missed_items += 1
            elif invoiced_qty > estimated_qty:
                over_delivered_items += 1

            # Get item details (prefer invoice data if available, else use estimate data)
            item_name = invoice_data.get("item_name") or estimate_data.get(
                "item_name", ""
            )
            sku = invoice_data.get("sku") or estimate_data.get("sku", "")

            result.append(
                {
                    "item_id": item_id,
                    "item_name": item_name,
                    "sku": sku,
                    "estimated_quantity": estimated_qty,
                    "invoiced_quantity": invoiced_qty,
                    "fill_rate": round(fill_rate, 2),
                    "closing_stock": closing_stock,
                    "missed_quantity": max(0, estimated_qty - invoiced_qty),
                }
            )

        # Sort by missed items first (invoiced < estimated), then by missed quantity descending
        result.sort(
            key=lambda x: (
                0 if x["invoiced_quantity"] < x["estimated_quantity"] else 1,
                -x["missed_quantity"],
            )
        )

        result.sort(key=lambda x: (x["item_name"]))
        return {
            "data": result,
            "meta": {
                "start_date": start_date,
                "end_date": end_date,
                "total_items": len(result),
                "total_estimates": len(estimates),
                "total_invoices": len(invoices),
                "missed_items": missed_items,
                "over_delivered_items": over_delivered_items,
                "fully_delivered_items": len(result)
                - missed_items
                - over_delivered_items,
                "stock_date": stock_date,
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating estimates vs invoices report: {str(e)}")
        raise HTTPException(
            status_code=500, detail=f"Failed to generate report: {str(e)}"
        )


def _check_invoice_recency_sync(db) -> dict:
    """
    Check whether the invoices collection has data for the last 15 days.
    Returns a dict with keys: missing (bool), latest_date (str|None), days_behind (int|None).
    Invoice dates are stored as strings in YYYY-MM-DD format.
    """
    invoices_col = db[INVOICES_COLLECTION]
    latest_doc = invoices_col.find_one(
        {"date": {"$type": "string"}},
        sort=[("date", -1)],
        projection={"date": 1},
    )
    if not latest_doc:
        return {"missing": True, "latest_date": None, "days_behind": None}

    latest_date_str = latest_doc["date"]
    try:
        latest_dt = datetime.strptime(latest_date_str[:10], "%Y-%m-%d")
    except ValueError:
        return {"missing": True, "latest_date": latest_date_str, "days_behind": None}

    today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    days_behind = (today - latest_dt).days
    return {
        "missing": days_behind > 15,
        "latest_date": latest_date_str[:10],
        "days_behind": days_behind,
    }


@router.get("/invoice-sync-status")
async def get_invoice_sync_status(db=Depends(get_database)):
    """
    Returns whether the invoices collection is missing the last 15 days of data.
    Used by report pages to warn users when invoices may be re-syncing.
    """
    try:
        result = await asyncio.to_thread(_check_invoice_recency_sync, db)
        return result
    except Exception as e:
        logger.error(f"Error checking invoice sync status: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to check invoice sync status")


# ─── Composite Item Bulk Upload ────────────────────────────────────────────────

_INVENTORY_REFRESH_TOKEN = os.getenv("INVENTORY_REFRESH_TOKEN")
_INVENTORY_ORGANIZATION_ID = os.getenv(
    "INVENTORY_ORGANIZATION_ID", os.getenv("ORGANIZATION_ID", "776755316")
)
_ZOHO_INVENTORY_BASE = "https://www.zohoapis.com/inventory/v1"
_MAX_COMPOSITE_COMPONENTS = 5


def _get_zoho_inventory_token() -> str:
    url = _BOOKS_URL.format(
        clientId=_BOOKS_CLIENT_ID,
        clientSecret=_BOOKS_CLIENT_SECRET,
        grantType="refresh_token",
        books_refresh_token=_INVENTORY_REFRESH_TOKEN,
    )
    r = requests.post(url, timeout=30)
    r.raise_for_status()
    return r.json()["access_token"]


def _create_composite_item_in_zoho(
    name: str, sku: str, rate: float, components: list
) -> dict:
    token = _get_zoho_inventory_token()
    headers = {
        "Authorization": f"Zoho-oauthtoken {token}",
        "Content-Type": "application/json",
    }
    payload = {
        "name": name,
        "sku": sku,
        "sales_rate": rate,
        "item_type":"inventory",
        "is_combo_product": True,
        "mapped_items": [
            {"item_id": c["item_id"], "quantity": c["quantity"]}
            for c in components
        ],
    }
    logger.info(f"Zoho Inventory create composite payload: {json.dumps(payload)}")
    r = requests.post(
        f"{_ZOHO_INVENTORY_BASE}/compositeitems",
        headers=headers,
        params={"organization_id": _INVENTORY_ORGANIZATION_ID},
        json=payload,
        timeout=30,
    )
    logger.info(f"Zoho Inventory response {r.status_code}: {r.text[:500]}")
    if not r.ok:
        raise ValueError(f"Zoho Inventory HTTP {r.status_code}: {r.text[:300]}")
    data = r.json()
    if data.get("code") != 0:
        raise ValueError(f"Zoho Inventory error {data.get('code')}: {data.get('message')}")
    return data["composite_item"]


def _parse_composite_template(file_bytes: bytes) -> list:
    """
    Parse the composite items template Excel.
    Returns list of raw row dicts.
    """
    import openpyxl

    wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Find header row (contains "Composite Name")
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        if row[0] and str(row[0]).strip().lower() == "composite name":
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("Could not find header row (expected 'Composite Name' in column A)")

    rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        name = row[0]
        if name is None or str(name).strip() == "":
            continue
        sku_code = row[1]
        rate = row[2]
        components_raw = []
        for i in range(_MAX_COMPOSITE_COMPONENTS):
            col_sku = 3 + i * 2
            col_qty = 4 + i * 2
            c_sku = row[col_sku] if col_sku < len(row) else None
            c_qty = row[col_qty] if col_qty < len(row) else None
            if c_sku and str(c_sku).strip():
                components_raw.append({
                    "sku_code": str(c_sku).strip(),
                    "quantity": float(c_qty) if c_qty is not None else 1.0,
                })
        rows.append({
            "name": str(name).strip(),
            "sku_code": str(sku_code).strip() if sku_code else "",
            "rate": float(rate) if rate is not None else 0.0,
            "components_raw": components_raw,
        })
    return rows


def _validate_composite_rows(rows: list, db) -> dict:
    """
    Validate parsed rows against the products collection.
    Returns {rows: [...], valid_count, error_count, total_count}.
    """
    # Collect all component SKUs to batch-fetch
    all_skus = set()
    for row in rows:
        for c in row["components_raw"]:
            all_skus.add(c["sku_code"])

    products_col = db[PRODUCTS_COLLECTION]
    product_docs = list(products_col.find(
        {"cf_sku_code": {"$in": list(all_skus)}},
        {"_id": 1, "cf_sku_code": 1, "name": 1, "item_id": 1},
    ))
    sku_map = {p["cf_sku_code"]: p for p in product_docs}

    # Also check if composite SKU already exists
    comp_col = db[COMPOSITE_COLLECTION]
    composite_skus_raw = rows
    existing_skus = set(
        d["sku_code"]
        for d in comp_col.find(
            {"sku_code": {"$in": [r["sku_code"] for r in composite_skus_raw]}},
            {"sku_code": 1},
        )
    )

    result_rows = []
    for i, row in enumerate(rows):
        errors = []
        warnings = []

        if not row["name"]:
            errors.append("Composite Name is required")
        if not row["sku_code"]:
            errors.append("Composite SKU Code is required")
        elif row["sku_code"] in existing_skus:
            errors.append(f"Composite SKU '{row['sku_code']}' already exists in the database")
        if row["rate"] <= 0:
            warnings.append("Rate is 0 or negative")
        if not row["components_raw"]:
            errors.append("At least one component is required")

        resolved_components = []
        for c in row["components_raw"]:
            prod = sku_map.get(c["sku_code"])
            if prod:
                resolved_components.append({
                    "sku_code": c["sku_code"],
                    "quantity": c["quantity"],
                    "name": prod.get("name", ""),
                    "item_id": prod.get("item_id", ""),
                    "product_id": str(prod["_id"]),
                })
            else:
                errors.append(f"Component SKU '{c['sku_code']}' not found in products")
                resolved_components.append({
                    "sku_code": c["sku_code"],
                    "quantity": c["quantity"],
                    "name": None,
                    "item_id": None,
                    "product_id": None,
                })

        result_rows.append({
            "row": i + 2,  # 1-based + header
            "name": row["name"],
            "sku_code": row["sku_code"],
            "rate": row["rate"],
            "components": resolved_components,
            "valid": len(errors) == 0,
            "errors": errors,
            "warnings": warnings,
        })

    valid_count = sum(1 for r in result_rows if r["valid"])
    return {
        "rows": result_rows,
        "valid_count": valid_count,
        "error_count": len(result_rows) - valid_count,
        "total_count": len(result_rows),
    }


@router.get("/composite-upload/template")
def download_composite_template():
    """Download the Excel template for bulk composite item creation."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Composite Items"

    header_fill = PatternFill("solid", fgColor="1E40AF")
    comp_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    note_fill = PatternFill("solid", fgColor="EFF6FF")
    note_font = Font(italic=True, color="1E3A8A", size=9)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Instruction row
    note_cols = 3 + _MAX_COMPOSITE_COMPONENTS * 2
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=note_cols)
    note_cell = ws.cell(row=1, column=1,
        value=(
            "Fill in Composite Name, SKU Code, Rate, then Component SKU + Qty for each component "
            "(up to 5). Leave unused component columns blank. Component SKUs must match existing "
            "product SKU codes exactly."
        )
    )
    note_cell.fill = note_fill
    note_cell.font = note_font
    note_cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 30

    # Headers row 2
    fixed_headers = ["Composite Name", "Composite SKU Code", "Rate (MRP)"]
    comp_headers = []
    for i in range(1, _MAX_COMPOSITE_COMPONENTS + 1):
        comp_headers += [f"Component {i} SKU", f"Component {i} Qty"]
    headers = fixed_headers + comp_headers

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = comp_fill if col > 3 else header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[2].height = 20

    # Column widths
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    for i in range(_MAX_COMPOSITE_COMPONENTS):
        col_letter_sku = chr(ord("D") + i * 2)
        col_letter_qty = chr(ord("E") + i * 2)
        ws.column_dimensions[col_letter_sku].width = 22
        ws.column_dimensions[col_letter_qty].width = 10

    # Freeze header rows
    ws.freeze_panes = "A3"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="composite_items_template.xlsx"'},
    )


@router.post("/composite-upload/validate")
async def validate_composite_upload(file: UploadFile = File(...)):
    """
    Parse and validate an uploaded composite items Excel file.
    Returns rows with validation status; valid rows include resolved component details.
    """
    try:
        file_bytes = await file.read()
        rows = _parse_composite_template(file_bytes)
        if not rows:
            raise HTTPException(status_code=400, detail="No data rows found in the uploaded file")
        db = get_database()
        result = await asyncio.to_thread(_validate_composite_rows, rows, db)
        return JSONResponse(content=result)
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Composite upload validate error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class CompositeCreateRequest(BaseModel):
    items: List[dict]


@router.post("/composite-upload/create")
async def create_composite_items(payload: CompositeCreateRequest):
    """
    Create validated composite items in Zoho Inventory and save to MongoDB.
    Expects items from the validate endpoint (valid=True rows only).
    Returns per-item results with success/failure.
    """
    db = get_database()
    results = []
    created_count = 0

    for item in payload.items:
        name = item.get("name", "")
        sku_code = item.get("sku_code", "")
        rate = item.get("rate", 0.0)
        components = item.get("components", [])

        try:
            zoho_item = await asyncio.to_thread(
                _create_composite_item_in_zoho, name, sku_code, rate, components
            )
            composite_item_id = zoho_item.get("composite_item_id") or zoho_item.get("item_id", "")

            now = datetime.utcnow()
            doc = {
                "composite_item_id": str(composite_item_id),
                "name": name,
                "sku_code": sku_code,
                "rate": rate,
                "components": [
                    {
                        "name": c.get("name", ""),
                        "quantity": c.get("quantity", 1.0),
                        "item_id": c.get("item_id", ""),
                        "sku_code": c.get("sku_code", ""),
                        "product_id": c.get("product_id", ""),
                    }
                    for c in components
                ],
                "last_updated": now,
                "created_at": now,
            }
            await asyncio.to_thread(
                db[COMPOSITE_COLLECTION].insert_one, doc
            )
            results.append({"sku_code": sku_code, "name": name, "success": True, "composite_item_id": str(composite_item_id)})
            created_count += 1
        except Exception as e:
            logger.error(f"Failed to create composite item '{sku_code}': {e}")
            results.append({"sku_code": sku_code, "name": name, "success": False, "error": str(e)})

    return JSONResponse(content={
        "results": results,
        "created_count": created_count,
        "failed_count": len(results) - created_count,
    })


# ─── PIS (Product Information Sheet) ──────────────────────────────────────────

from .design import (  # noqa: E402
    _parse_pis_workbook,
    _PIS_SHEETS,
    _style_header_row,
    _auto_col_width,
    _upload_to_s3 as _pis_upload_to_s3,
    _slugify as _pis_slugify,
    _extract_email_from_token,
)

_PIS_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_PIS_BRAND_ORDERS_COLLECTION = "brand_orders"


@router.get("/pis-template")
def zoho_download_pis_template():
    """Return empty PIS XLSX template."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, headers in _PIS_SHEETS:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
        _style_header_row(ws, 1, len(headers))
        ws.freeze_panes = "A2"
        _auto_col_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=_PIS_CONTENT_TYPE,
        headers={"Content-Disposition": 'attachment; filename="PIS_Template.xlsx"'},
    )


@router.post("/upload-pis")
async def zoho_preview_pis(file: UploadFile = File(...)):
    """Dry-run PIS parse — returns what would be updated without writing."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be an Excel (.xlsx) file")
    data = await file.read()
    db = get_database()
    result = await asyncio.to_thread(_parse_pis_workbook, data, db, True)
    return JSONResponse(content=result)


@router.post("/confirm-pis")
async def zoho_confirm_pis(
    file: UploadFile = File(...),
    order_id: str = Form(...),
    authorization: Optional[str] = Header(default=None),
):
    """Apply PIS XLSX to design_catalogue and attach the file to the given brand order
    so it appears in both the brand_orders page and the design orders page."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be an Excel (.xlsx) file")

    email = _extract_email_from_token(authorization)
    data = await file.read()
    db = get_database()

    result = await asyncio.to_thread(_parse_pis_workbook, data, db, False)

    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_filename = re.sub(r"[^\w.\-]", "_", file.filename)

    def _attach_to_order():
        order = db[_PIS_BRAND_ORDERS_COLLECTION].find_one(
            {"_id": ObjectId(order_id)}, {"brand": 1, "name": 1}
        )
        if not order:
            raise ValueError(f"Order {order_id} not found")

        brand_slug = _pis_slugify(order["brand"])
        name_slug = _pis_slugify(order["name"])
        s3_key = f"brand_orders/{brand_slug}/{name_slug}/pis/{ts}_{safe_filename}"

        _pis_upload_to_s3(data, s3_key, _PIS_CONTENT_TYPE)

        doc_entry = {
            "doc_id": str(ObjectId()),
            "filename": file.filename,
            "s3_key": s3_key,
            "content_type": _PIS_CONTENT_TYPE,
            "size": len(data),
            "uploaded_at": datetime.utcnow(),
            "category": "Product Information Sheet",
            "folder": "",
        }
        designer_doc_entry = {**doc_entry, "doc_id": str(ObjectId())}

        db[_PIS_BRAND_ORDERS_COLLECTION].update_one(
            {"_id": ObjectId(order_id)},
            {
                "$push": {
                    "documents": doc_entry,
                    "designer_documents": designer_doc_entry,
                },
                "$set": {"updated_at": datetime.utcnow()},
            },
        )
        return s3_key, order["brand"], order["name"]

    try:
        s3_key, order_brand, order_name = await asyncio.to_thread(_attach_to_order)
        result["audit"] = {"uploaded_by": email or "unknown", "uploaded_at": ts, "s3_key": s3_key}
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        logger.warning("PIS order attachment failed: %s", exc)
        raise HTTPException(status_code=500, detail="Failed to attach PIS to order")

    asyncio.create_task(fire_trigger("pil_uploaded", {
        "brand": order_brand,
        "order_name": order_name,
        "filename": file.filename,
    }, db))

    return JSONResponse(content=result)
