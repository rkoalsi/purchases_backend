from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form, Body
from fastapi.responses import JSONResponse, StreamingResponse
from datetime import datetime, timedelta
from ..database import get_database, serialize_mongo_document
import asyncio
import io
import os
import ast
import zipfile
from decimal import Decimal, ROUND_HALF_UP
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Optional
from pydantic import BaseModel
import logging
import boto3
import requests
from botocore.config import Config as BotocoreConfig
import xlrd
import xlwt
from xlutils.copy import copy as xl_copy

from .amazon import compute_drr_for_asins_sync, generate_report_by_date_range

logger = logging.getLogger(__name__)
router = APIRouter()

PO_COLLECTION = "vendor_purchase_orders"
MARGINS_COLLECTION = "vendor_margins"
PRODUCTS_COLLECTION = "products"
SKU_MAPPING_COLLECTION = "amazon_sku_mapping"
ZOHO_STOCK_COLLECTION = "zoho_warehouse_stock"
INVENTORY_COLLECTION = "amazon_vendor_inventory"
SALES_COLLECTION = "amazon_vendor_sales"
CUSTOMERS_COLLECTION = "customers"
ESTIMATES_COLLECTION = "estimates"
PACKAGES_COLLECTION = "packages"
TRANSFER_ORDERS_COLLECTION = "transfer_orders"
ASSEMBLIES_COLLECTION = "assemblies"
SALES_ORDERS_COLLECTION = "sales_orders"

ZOHO_BOOKS_BASE = "https://books.zoho.com/api/v3"
ZOHO_INVENTORY_BASE = os.getenv("ZOHO_INVENTORY_BASE", "https://www.zohoapis.com/inventory/v1")
ORGANIZATION_ID = os.getenv("ORGANIZATION_ID", "776755316")

# Fixed warehouse IDs for vendor PO transfer orders
TO_FROM_WAREHOUSE_ID = "3220178000000403010"  # Pupscribe Enterprises Private Limited
TO_TO_WAREHOUSE_ID = "3220178000156676949"    # Mumbai (Amazon)
BOOKS_URL = os.getenv("BOOKS_URL")
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
BOOKS_REFRESH_TOKEN = os.getenv("BOOKS_REFRESH_TOKEN")
INVENTORY_REFRESH_TOKEN = os.getenv("INVENTORY_REFRESH_TOKEN", BOOKS_REFRESH_TOKEN)

COVERAGE_DAYS = 35

_ZOHO_TOKEN_URL = "https://accounts.zoho.com/oauth/v2/token"


def _get_zoho_token() -> str:
    url = BOOKS_URL.format(
        clientId=CLIENT_ID,
        clientSecret=CLIENT_SECRET,
        grantType="refresh_token",
        books_refresh_token=BOOKS_REFRESH_TOKEN,
    )
    r = requests.post(url, timeout=30)
    r.raise_for_status()
    return r.json()["access_token"]


def _get_inventory_token() -> str:
    r = requests.post(
        _ZOHO_TOKEN_URL,
        params={
            "refresh_token": INVENTORY_REFRESH_TOKEN,
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "grant_type": "refresh_token",
        },
        timeout=30,
    )
    r.raise_for_status()
    return r.json()["access_token"]


def _parse_addresses(raw) -> list[dict]:
    if isinstance(raw, list):
        return raw
    if not raw or not isinstance(raw, str):
        return []
    try:
        return ast.literal_eval(raw)
    except (ValueError, SyntaxError):
        return []


# ─── helpers ──────────────────────────────────────────────────────────────────


def _extract_gst(item_tax_preferences: list) -> float:
    if not item_tax_preferences:
        return 0.0
    for pref in item_tax_preferences:
        if pref.get("tax_specification") == "intra":
            return float(pref.get("tax_percentage", 0))
    return float(item_tax_preferences[0].get("tax_percentage", 0))


def _fetch_monthly_amazon_sales(db, asins: list, num_months: int = 5):
    """Return ({asin: [units_m1, ..., units_m5]}, [(year, month, label), ...]) for the last num_months complete calendar months.

    Mirrors the Amazon report page: sums FBA sales (amazon_sales_traffic.salesByAsin.unitsOrdered)
    + VC sales (amazon_vendor_sales.orderedUnits) per month, matching parentAsin / asin respectively.
    """
    import calendar as _cal
    from datetime import date
    today = date.today()
    months = []
    for i in range(1, num_months + 1):
        m = today.month - i
        y = today.year
        if m <= 0:
            m += 12
            y -= 1
        label = f"{_cal.month_abbr[m]} {y}"
        months.append((y, m, label))
    if not months:
        return {}, []
    earliest = datetime(months[-1][0], months[-1][1], 1)

    by_asin: dict = {}

    # FBA sales (Seller Central / Seller Partner API)
    fba_pipeline = [
        {"$match": {"parentAsin": {"$in": asins}, "date": {"$gte": earliest}}},
        {"$group": {
            "_id": {"asin": "$parentAsin", "year": {"$year": "$date"}, "month": {"$month": "$date"}},
            "units_sold": {"$sum": "$salesByAsin.unitsOrdered"},
        }},
    ]
    for doc in db["amazon_sales_traffic"].aggregate(fba_pipeline):
        a = doc["_id"]["asin"]
        key = (doc["_id"]["year"], doc["_id"]["month"])
        by_asin.setdefault(a, {})
        by_asin[a][key] = by_asin[a].get(key, 0) + int(doc["units_sold"])

    # VC sales (Vendor Central)
    vc_pipeline = [
        {"$match": {"asin": {"$in": asins}, "date": {"$gte": earliest}}},
        {"$group": {
            "_id": {"asin": "$asin", "year": {"$year": "$date"}, "month": {"$month": "$date"}},
            "units_sold": {"$sum": "$orderedUnits"},
        }},
    ]
    for doc in db["amazon_vendor_sales"].aggregate(vc_pipeline):
        a = doc["_id"]["asin"]
        key = (doc["_id"]["year"], doc["_id"]["month"])
        by_asin.setdefault(a, {})
        by_asin[a][key] = by_asin[a].get(key, 0) + int(doc["units_sold"])

    result = {a: [by_asin.get(a, {}).get((y, m), 0) for (y, m, _) in months] for a in asins}
    return result, months


def _get_dispatched_costs_by_asin(package_number: str, db) -> dict:
    """Return {asin: {qty, item_total, item_total_gst}} for a linked package's line items."""
    pkg = db[PACKAGES_COLLECTION].find_one({"package_number": package_number})
    if not pkg:
        return {}
    line_items = pkg.get("line_items") or []
    sku_rows = []
    for li in line_items:
        sku = None
        for cf in li.get("item_custom_fields") or []:
            if cf.get("api_name") == "cf_sku_code":
                sku = cf.get("value")
                break
        if sku:
            sku_rows.append({"sku": sku, "qty": float(li.get("quantity") or 0)})
    if not sku_rows:
        return {}
    skus = [r["sku"] for r in sku_rows]
    sku_to_asin = {m["sku_code"]: m["item_id"] for m in db[SKU_MAPPING_COLLECTION].find({"sku_code": {"$in": skus}}, {"sku_code": 1, "item_id": 1})}
    asins = list(set(sku_to_asin.values()))
    asin_to_margin, asin_to_cost_price, asin_to_etrade = {}, {}, {}
    for m in db[MARGINS_COLLECTION].find({"asin": {"$in": asins}}, {"asin": 1, "etrade_asp": 1, "margin": 1, "cost_price_wo_tax": 1}):
        if m.get("etrade_asp") is not None:
            asin_to_etrade[m["asin"]] = float(m["etrade_asp"])
        if m.get("margin") is not None:
            asin_to_margin[m["asin"]] = float(m["margin"])
        if m.get("cost_price_wo_tax") is not None:
            asin_to_cost_price[m["asin"]] = float(m["cost_price_wo_tax"])
    sku_to_gst: dict[str, float] = {}
    for p in db[PRODUCTS_COLLECTION].find({"cf_sku_code": {"$in": skus}}, {"cf_sku_code": 1, "item_tax_preferences": 1}):
        sku_to_gst[p["cf_sku_code"]] = _extract_gst(p.get("item_tax_preferences") or [])
    missing = [s for s in skus if s not in sku_to_gst]
    if missing:
        comp_ids = {c["sku_code"]: c["composite_item_id"] for c in db["composite_products"].find({"sku_code": {"$in": missing}}, {"sku_code": 1, "composite_item_id": 1}) if c.get("sku_code") and c.get("composite_item_id")}
        if comp_ids:
            for p in db[PRODUCTS_COLLECTION].find({"item_id": {"$in": list(comp_ids.values())}}, {"item_id": 1, "item_tax_preferences": 1}):
                for sku, iid in comp_ids.items():
                    if p.get("item_id") == iid:
                        sku_to_gst[sku] = _extract_gst(p.get("item_tax_preferences") or [])
    result: dict = {}
    for r in sku_rows:
        sku, qty = r["sku"], r["qty"]
        asin = sku_to_asin.get(sku)
        if not asin:
            continue
        gst = sku_to_gst.get(sku, 0.0)
        etrade_asp = asin_to_etrade.get(asin)
        margin = asin_to_margin.get(asin)
        mrp_wo_gst = round(etrade_asp / (1 + gst / 100), 2) if (etrade_asp and gst) else etrade_asp
        item_cost = None
        if mrp_wo_gst is not None and margin is not None:
            unit_cost = round(round(mrp_wo_gst, 2) * (1 - round(margin * 100, 2) / 100), 4)
            item_cost = round(unit_cost * qty, 2)
        elif asin in asin_to_cost_price:
            item_cost = round(asin_to_cost_price[asin] * qty, 2)
        item_cost_gst = round(item_cost * (1 + gst / 100), 2) if (item_cost is not None and gst) else item_cost
        if asin in result:
            result[asin]["qty"] = result[asin]["qty"] + qty
            result[asin]["item_total"] = round((result[asin]["item_total"] or 0) + (item_cost or 0), 2)
            result[asin]["item_total_gst"] = round((result[asin]["item_total_gst"] or 0) + (item_cost_gst or 0), 2)
        else:
            result[asin] = {"qty": qty, "item_total": item_cost, "item_total_gst": item_cost_gst}
    return result


def _compute_packages_accepted_costs(
    package_numbers: list[str], db
) -> tuple[float | None, float | None]:
    """Sum accepted_total_cost and accepted_total_cost_gst across all linked packages."""
    total_cost: float = 0.0
    total_cost_gst: float = 0.0
    has_any = False
    for pkg_num in package_numbers:
        try:
            c, cg = _compute_package_accepted_costs(pkg_num, db)
        except ValueError:
            continue
        if c is not None:
            total_cost += c
            has_any = True
        if cg is not None:
            total_cost_gst += cg
    if not has_any:
        return None, None
    return round(total_cost, 2), round(total_cost_gst, 2)


def _compute_package_accepted_costs(
    package_number: str, db
) -> tuple[float | None, float | None]:
    """Compute accepted_total_cost and accepted_total_cost_gst from a linked package.

    Uses the same formula as total_cost / total_cost_gst (ASP as base rate, margin discount),
    but with each package line item's quantity as the accepted qty.
    Returns (accepted_total_cost, accepted_total_cost_gst) or (None, None) if not computable.
    """
    pkg = db[PACKAGES_COLLECTION].find_one({"package_number": package_number})
    if not pkg:
        raise ValueError(f"Package {package_number!r} not found in packages collection")

    line_items = pkg.get("line_items") or []

    # Extract SKUs from item_custom_fields
    sku_qty_pairs: list[tuple[str, float]] = []
    for li in line_items:
        sku = None
        for cf in li.get("item_custom_fields") or []:
            if cf.get("api_name") == "cf_sku_code":
                sku = cf.get("value")
                break
        if sku:
            sku_qty_pairs.append((sku, float(li.get("quantity") or 0)))

    if not sku_qty_pairs:
        return None, None

    skus = [s for s, _ in sku_qty_pairs]

    # SKU → ASIN
    sku_to_asin: dict[str, str] = {}
    for m in db[SKU_MAPPING_COLLECTION].find({"sku_code": {"$in": skus}}, {"sku_code": 1, "item_id": 1}):
        sku_to_asin[m["sku_code"]] = m["item_id"]

    asins = list(sku_to_asin.values())

    # ASIN → etrade_asp, margin, cost_price_wo_tax
    asin_to_margin: dict[str, float] = {}
    asin_to_cost_price: dict[str, float] = {}
    asin_to_etrade: dict[str, float] = {}
    if asins:
        for m in db[MARGINS_COLLECTION].find(
            {"asin": {"$in": asins}},
            {"asin": 1, "etrade_asp": 1, "margin": 1, "cost_price_wo_tax": 1},
        ):
            if m.get("etrade_asp") is not None:
                asin_to_etrade[m["asin"]] = float(m["etrade_asp"])
            if m.get("margin") is not None:
                asin_to_margin[m["asin"]] = float(m["margin"])
            if m.get("cost_price_wo_tax") is not None:
                asin_to_cost_price[m["asin"]] = float(m["cost_price_wo_tax"])

    # SKU → GST (via products)
    sku_to_gst: dict[str, float] = {}
    for p in db[PRODUCTS_COLLECTION].find(
        {"cf_sku_code": {"$in": skus}},
        {"cf_sku_code": 1, "item_tax_preferences": 1},
    ):
        sku_to_gst[p["cf_sku_code"]] = _extract_gst(p.get("item_tax_preferences") or [])

    accepted_total_cost = 0.0
    accepted_total_cost_gst = 0.0
    has_any = False
    for sku, qty in sku_qty_pairs:
        asin = sku_to_asin.get(sku)
        if not asin:
            continue
        gst = sku_to_gst.get(sku, 0.0)
        etrade_asp = asin_to_etrade.get(asin)
        margin = asin_to_margin.get(asin)

        # Use etrade_asp as the base rate (same as _enrich_items: asp_base → mrp_wo_gst)
        asp_base = etrade_asp
        mrp_wo_gst = (
            round(asp_base / (1 + gst / 100), 2) if (asp_base and gst) else asp_base
        ) if asp_base is not None else None

        # Mirror the total_cost formula exactly
        if mrp_wo_gst is not None and margin is not None:
            _rate = round(mrp_wo_gst, 2)
            _discount_factor = 1 - round(margin * 100, 2) / 100
            item_cost = round(_rate * qty * _discount_factor, 2)
        elif asin in asin_to_cost_price:
            item_cost = round(asin_to_cost_price[asin] * qty, 2)
        else:
            continue

        item_cost_gst = (
            round(item_cost * (1 + gst / 100), 2) if gst else item_cost
        )
        accepted_total_cost += item_cost
        accepted_total_cost_gst += item_cost_gst
        has_any = True

    if not has_any:
        return None, None

    return round(accepted_total_cost, 2), round(accepted_total_cost_gst, 2)


def _parse_po_excel(file_bytes: bytes) -> tuple[str, str, list[dict]]:
    """Parse VC PO Excel. Returns (po_number, vendor, items)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("PO file has no data rows")

    # header row: PO, Vendor, Ship to location, ASIN, External Id, External Id Type,
    #             Model Number, Title, Availability, Window Type, Window start,
    #             Window end, Expected date, Quantity Requested, Expected Quantity,
    #             Unit Cost, currency
    po_number = str(rows[1][0]).strip() if rows[1][0] else ""
    vendor = str(rows[1][1]).strip() if rows[1][1] else ""

    items = []
    for row in rows[1:]:
        if not row[0]:
            continue
        items.append(
            {
                "asin": str(row[3]).strip() if row[3] else "",
                "model_number": str(row[6]).strip() if row[6] else "",
                "title": str(row[7]).strip() if row[7] else "",
                "ship_to_location": str(row[2]).strip() if row[2] else "",
                "requested_qty": int(row[13]) if row[13] is not None else 0,
                "supply_qty": None,  # will be set to final_supply_qty after enrichment
                "accepted_qty": 0,
                "received_qty": None,
                "etrade_unit_cost": float(row[15]) if row[15] is not None else 0.0,
            }
        )
    return po_number, vendor, items


async def _compute_drr_for_po_async(db, po_date_str: str, asins: list[str]) -> dict:
    """Compute DRR using the exact same PSR 'all' report function so values always match."""
    if not asins or not po_date_str:
        return {}
    try:
        po_dt = datetime.strptime(po_date_str, "%Y-%m-%d")
    except ValueError:
        return {}
    # Use same 30-day window as a typical PSR run ending on the PO date
    start_date = (po_dt - timedelta(days=30)).strftime("%Y-%m-%d")
    report = await generate_report_by_date_range(start_date, po_date_str, db, report_type="all")
    asin_set = set(asins)
    result = {
        item["asin"]: {"drr": item.get("drr", 0), "drr_flag": item.get("drr_flag", "")}
        for item in report
        if item.get("asin") in asin_set
    }
    # ASINs with no sales in the report window won't appear in the PSR result.
    # Fall back to the sync 90-day inventory lookback for those.
    missing = [a for a in asins if a not in result]
    if missing:
        end_dt = po_dt.replace(hour=23, minute=59, second=59, microsecond=999999)
        fallback = await asyncio.to_thread(compute_drr_for_asins_sync, db, end_dt, missing)
        result.update(fallback)
    return result


def _enrich_items(
    items: list[dict],
    po_number: str,
    db,
    use_stored_stock: bool = False,
    po_date_str: str | None = None,
    drr_map: dict | None = None,
) -> tuple[list[dict], str | None, str | None]:
    """Enrich PO items with product/stock/sales data. Returns (enriched_items, inventory_date_str, zoho_stock_date_str).

    use_stored_stock=True: zoho_stock/current_stock/open_po/last_30_sales are read from the stored
    item fields (frozen at upload time). Only margin/product-derived fields are recomputed.
    use_stored_stock=False: all data is fetched live.
      - If po_date_str provided: zoho_stock and current_stock are fetched from T-2 (po_date - 2 days).
      - supply_qty is set to final_supply_qty.
    """
    asins = [it["asin"] for it in items if it["asin"]]
    model_numbers = [it["model_number"] for it in items if it["model_number"]]

    # Some VC PO exports put the ASIN in the Model Number (SKU) column — detect and include them
    def _looks_like_asin(s: str) -> bool:
        return bool(s) and len(s) == 10 and s[:2].upper() == "B0"

    asin_like_models = [
        m for m in model_numbers if _looks_like_asin(m) and m not in asins
    ]
    if asin_like_models:
        asins = list(set(asins + asin_like_models))

    # detect old-format items that predate stock snapshotting — fall back to live fetch
    has_stored_stock = items and "zoho_stock" in items[0]
    effective_use_stored = use_stored_stock and has_stored_stock

    # zoho_cutoff = PO date (Zoho data is available same-day)
    # stock_cutoff = T-2 from PO date (Amazon vendor inventory has a 2-day lag)
    po_date_dt: datetime | None = None
    zoho_cutoff: datetime | None = None
    stock_cutoff: datetime | None = None
    if po_date_str:
        try:
            po_date_dt = datetime.strptime(po_date_str, "%Y-%m-%d")
        except ValueError:
            pass
    if po_date_dt and not effective_use_stored:
        zoho_cutoff = po_date_dt
        stock_cutoff = po_date_dt - timedelta(days=2)

    # --- batch load products by cf_sku_code (always fresh — MRP/GST/HSN can change) ---
    products_by_model: dict[str, dict] = {}
    for p in db[PRODUCTS_COLLECTION].find(
        {"cf_sku_code": {"$in": model_numbers}},
        {
            "cf_sku_code": 1,
            "item_id": 1,
            "rate": 1,
            "item_tax_preferences": 1,
            "hsn_or_sac": 1,
            "purchase_status": 1,
        },
    ):
        products_by_model[p["cf_sku_code"]] = p

    # --- batch load sku mapping for ASIN → sku_code fallback ---
    sku_map_by_asin: dict[str, str] = {}
    for m in db[SKU_MAPPING_COLLECTION].find(
        {"item_id": {"$in": asins}}, {"item_id": 1, "sku_code": 1}
    ):
        sku_map_by_asin[m["item_id"]] = m["sku_code"]

    extra_skus = [
        sku_map_by_asin[asin]
        for asin in asins
        if asin in sku_map_by_asin and sku_map_by_asin[asin] not in products_by_model
    ]
    if extra_skus:
        for p in db[PRODUCTS_COLLECTION].find(
            {"cf_sku_code": {"$in": extra_skus}},
            {
                "cf_sku_code": 1,
                "item_id": 1,
                "rate": 1,
                "item_tax_preferences": 1,
                "hsn_or_sac": 1,
                "purchase_status": 1,
            },
        ):
            products_by_model[p["cf_sku_code"]] = p

    products_by_asin: dict[str, dict] = {}
    for asin in asins:
        sku = sku_map_by_asin.get(asin, "")
        if sku and sku in products_by_model:
            products_by_asin[asin] = products_by_model[sku]

    # --- batch load vendor margins (always fresh — editable after upload) ---
    margins_by_asin: dict[str, float] = {}
    cost_prices_by_asin: dict[str, float] = {}
    etrade_asp_by_asin: dict[str, float] = {}
    for m in db[MARGINS_COLLECTION].find(
        {"asin": {"$in": asins}},
        {"asin": 1, "margin": 1, "cost_price_wo_tax": 1, "etrade_asp": 1},
    ):
        if m.get("margin") is not None:
            margins_by_asin[m["asin"]] = float(m["margin"])
        if m.get("cost_price_wo_tax") is not None:
            cost_prices_by_asin[m["asin"]] = float(m["cost_price_wo_tax"])
        if m.get("etrade_asp") is not None:
            etrade_asp_by_asin[m["asin"]] = float(m["etrade_asp"])

    # --- live stock/sales — only fetched at upload time (or for old-format items) ---
    zoho_latest: dict[str, int] = {}
    current_stock_by_asin: dict[str, int] = {}
    open_po_by_asin: dict[str, int] = {}
    sales_by_asin: dict[str, int] = {}
    inventory_date_str: str | None = None
    zoho_stock_date_str: str | None = None

    if not effective_use_stored:
        zoho_item_ids = list(
            {p["item_id"] for p in products_by_model.values() if p.get("item_id")}
        )
        if zoho_item_ids:
            zoho_match: dict = {"zoho_item_id": {"$in": zoho_item_ids}}
            if zoho_cutoff:
                zoho_match["date"] = {"$lte": zoho_cutoff}
            for doc in db[ZOHO_STOCK_COLLECTION].aggregate(
                [
                    {"$match": zoho_match},
                    {"$sort": {"date": -1}},
                    {
                        "$group": {
                            "_id": "$zoho_item_id",
                            "warehouses": {"$first": "$warehouses"},
                            "date": {"$first": "$date"},
                        }
                    },
                ]
            ):
                wh = doc.get("warehouses", {})
                zoho_latest[doc["_id"]] = (
                    int(sum(v for v in wh.values() if isinstance(v, (int, float))))
                    if isinstance(wh, dict)
                    else 0
                )
                if doc.get("date"):
                    d = (
                        doc["date"].strftime("%Y-%m-%d")
                        if hasattr(doc["date"], "strftime")
                        else str(doc["date"])[:10]
                    )
                    if zoho_stock_date_str is None or d > zoho_stock_date_str:
                        zoho_stock_date_str = d

        inv_match: dict = {"asin": {"$in": asins}}
        if stock_cutoff:
            # use $lt next-day midnight so records stored at any time on T-2 are included
            inv_match["date"] = {"$lt": stock_cutoff + timedelta(days=1)}
        latest_inv = db[INVENTORY_COLLECTION].find_one(
            inv_match, {"date": 1}, sort=[("date", -1)]
        )
        inventory_date = latest_inv["date"] if latest_inv else None
        inventory_date_str = (
            inventory_date.strftime("%Y-%m-%d") if inventory_date else None
        )
        if inventory_date:
            for doc in db[INVENTORY_COLLECTION].find(
                {"asin": {"$in": asins}, "date": inventory_date},
                {"asin": 1, "sellableOnHandInventoryUnits": 1},
            ):
                current_stock_by_asin[doc["asin"]] = int(
                    doc.get("sellableOnHandInventoryUnits") or 0
                )

        # open PO: processing → supply_qty_override if set, else final_supply_fo if set, else supply_qty if > 0, else requested_qty
        #          packed/closed/intransit → accepted_qty
        for doc in db[PO_COLLECTION].aggregate(
            [
                {
                    "$match": {
                        "po_number": {"$ne": po_number},
                        "po_status": {
                            "$in": ["processing", "packed", "closed", "intransit"]
                        },
                    }
                },
                {"$unwind": "$items"},
                {"$match": {"items.asin": {"$in": asins}}},
                {
                    "$group": {
                        "_id": "$items.asin",
                        "total": {
                            "$sum": {
                                "$cond": {
                                    "if": {"$eq": ["$po_status", "processing"]},
                                    "then": {
                                        "$cond": {
                                            "if": {"$ne": [{"$ifNull": ["$items.supply_qty_override", None]}, None]},
                                            "then": {"$ifNull": ["$items.supply_qty_override", 0]},
                                            "else": {
                                                "$cond": {
                                                    "if": {"$ne": [{"$ifNull": ["$items.final_supply_fo", None]}, None]},
                                                    "then": {"$ifNull": ["$items.final_supply_fo", 0]},
                                                    "else": {
                                                        "$cond": {
                                                            "if": {"$gt": [{"$ifNull": ["$items.supply_qty", 0]}, 0]},
                                                            "then": {"$ifNull": ["$items.supply_qty", 0]},
                                                            "else": {"$ifNull": ["$items.requested_qty", 0]},
                                                        }
                                                    },
                                                }
                                            },
                                        }
                                    },
                                    "else": {"$ifNull": ["$items.accepted_qty", 0]},
                                }
                            }
                        },
                    }
                },
            ]
        ):
            open_po_by_asin[doc["_id"]] = int(doc["total"])

        # Sales: strictly 30 days ending at T-2 (stock_cutoff). Inclusive window: [T-2-29, T-2] = 30 days.
        if stock_cutoff:
            sales_end = stock_cutoff
        else:
            sales_end = datetime.now()
        sales_start = sales_end - timedelta(days=29)
        for doc in db[SALES_COLLECTION].aggregate(
            [
                {
                    "$match": {
                        "asin": {"$in": asins},
                        "date": {"$gte": sales_start, "$lte": sales_end},
                    }
                },
                {"$group": {"_id": "$asin", "total_units": {"$sum": "$orderedUnits"}}},
            ]
        ):
            sales_by_asin[doc["_id"]] = int(doc["total_units"])

    # --- fetch DRR: use pre-computed map from PSR async path if provided, else fall back ---
    if drr_map is None:
        _drr_map: dict = {}
        if asins:
            drr_end = po_date_dt if po_date_dt else datetime.now()
            _drr_map = compute_drr_for_asins_sync(db, drr_end, asins)
    else:
        _drr_map = drr_map

    # --- fetch last 5 months Amazon sales per ASIN ---
    monthly_sales_by_asin: dict = {}
    month_labels: list = []
    if asins:
        monthly_sales_by_asin, month_labels = _fetch_monthly_amazon_sales(db, asins)

    # --- enrich each item ---
    enriched = []
    for item in items:
        asin = item["asin"]
        model = item["model_number"]

        # When VC PO has ASIN in the SKU column, also try sku_map lookup via model_number
        model_as_asin_sku = (
            sku_map_by_asin.get(model) if _looks_like_asin(model) else None
        )
        product = (
            products_by_asin.get(asin)
            or products_by_model.get(model)
            or (products_by_model.get(model_as_asin_sku) if model_as_asin_sku else None)
            or products_by_asin.get(model)
            or {}
        )
        zoho_item_id = product.get("item_id", "")

        mrp = float(product.get("rate") or 0)
        gst = _extract_gst(product.get("item_tax_preferences") or [])
        etrade_asp = etrade_asp_by_asin.get(asin)
        # MRP w/o GST is based on eTrade ASP if available, else falls back to Zoho MRP
        asp_base = etrade_asp if etrade_asp is not None else mrp
        mrp_wo_gst = (
            round(asp_base / (1 + gst / 100), 2) if (asp_base and gst) else asp_base
        )
        margin = margins_by_asin.get(asin, None)
        # Use directly stored cost_price_wo_tax if available, else compute from margin
        if asin in cost_prices_by_asin:
            cost_price_wo_tax = cost_prices_by_asin[asin]
        elif margin is not None:
            cost_price_wo_tax = round(mrp_wo_gst - mrp_wo_gst * margin, 2)
        else:
            cost_price_wo_tax = None

        # For frozen POs (processing/packed/closed), pricing was established at upload time.
        # Use the stored values so costs stay consistent with the Zoho estimate created then.
        if effective_use_stored:
            stored_mrp_wo_gst = item.get("mrp_wo_gst")
            stored_gst = item.get("gst")
            stored_margin = item.get("margin")
            stored_cost_price = item.get("cost_price_wo_tax")
            if stored_mrp_wo_gst is not None:
                mrp_wo_gst = stored_mrp_wo_gst
            if stored_gst is not None:
                gst = stored_gst
            if stored_margin is not None:
                margin = stored_margin
            if stored_cost_price is not None:
                cost_price_wo_tax = stored_cost_price
        etrade = item.get("etrade_unit_cost", 0)
        diff = (
            round(etrade - cost_price_wo_tax, 2)
            if cost_price_wo_tax is not None
            else None
        )

        accepted_qty = item.get("accepted_qty") or 0
        received_qty = item.get("received_qty")

        if effective_use_stored:
            zoho_stock = item.get("zoho_stock", 0)
            current_stock = item.get("current_stock", 0)
            open_po_override = item.get("open_po_override")
            open_po = (
                open_po_override
                if open_po_override is not None
                else item.get("open_po", 0)
            )
            last_30_sales = item.get("last_30_sales", 0)
            supply_qty_override = item.get("supply_qty_override")
            if supply_qty_override is not None:
                supply_qty = supply_qty_override
            else:
                sv = item.get("supply_qty")
                supply_qty = sv if sv is not None else item["requested_qty"]
            # DRR was computed and stored at freeze time — use it so Final Supply Qty
            # doesn't drift when Amazon sales data changes after the PO is locked.
            stored_drr = item.get("final_drr")
            stored_drr_flag = item.get("final_drr_flag", "")
            final_drr = stored_drr if isinstance(stored_drr, (int, float)) else None
            final_drr_flag = (
                stored_drr_flag if not isinstance(stored_drr, (int, float)) else ""
            )
        else:
            zoho_stock = zoho_latest.get(zoho_item_id, 0) if zoho_item_id else 0
            current_stock = current_stock_by_asin.get(asin, 0)
            open_po_override = item.get("open_po_override")
            open_po_computed = open_po_by_asin.get(asin, 0)
            open_po = (
                open_po_override if open_po_override is not None else open_po_computed
            )
            last_30_sales = sales_by_asin.get(asin, 0)
            # compute final_supply_qty and use it as supply_qty
            # Use round-half-up (math.floor(x+0.5)) to match Excel's ROUND() behaviour.
            # Python's built-in round() uses banker's rounding (round-half-to-even) which
            # diverges from Excel on .5 boundaries (e.g. 10.5 → Python=10, Excel=11).
            total_qty_tmp = current_stock + open_po
            ads_tmp = round(last_30_sales / 30, 2)
            target_tmp = int(
                (Decimal(str(ads_tmp)) * Decimal(COVERAGE_DAYS)).quantize(
                    Decimal("1"), rounding=ROUND_HALF_UP
                )
            )
            max_allowed_tmp = target_tmp - total_qty_tmp
            supply_qty_override = item.get("supply_qty_override")
            if supply_qty_override is not None:
                supply_qty = supply_qty_override
            elif total_qty_tmp == 0:
                supply_qty = item["requested_qty"]
            else:
                supply_qty = max(
                    0, min(int(item["requested_qty"]), int(max_allowed_tmp))
                )
            asin_drr = _drr_map.get(asin, {})
            drr_result = asin_drr.get("drr")
            drr_flag = asin_drr.get("drr_flag", "")
            final_drr = drr_result if isinstance(drr_result, (int, float)) else None
            final_drr_flag = (
                drr_flag if not isinstance(drr_result, (int, float)) else ""
            )

        # Match Zoho estimate formula: round(mrp_wo_gst, 2) * qty * (1 - round(margin*100, 2)%)
        # Using cost_price_wo_tax directly introduces per-unit rounding that diverges from Zoho totals.
        if mrp_wo_gst is not None and margin is not None:
            _rate = round(mrp_wo_gst, 2)
            _discount_factor = 1 - round(margin * 100, 2) / 100
            total_cost = round(_rate * supply_qty * _discount_factor, 2)
            total_cost_accepted = round(_rate * accepted_qty * _discount_factor, 2) if accepted_qty else None
        elif cost_price_wo_tax is not None:
            total_cost = round(cost_price_wo_tax * supply_qty, 2)
            total_cost_accepted = round(cost_price_wo_tax * accepted_qty, 2) if accepted_qty else None
        else:
            total_cost = None
            total_cost_accepted = None
        total_cost_gst = (
            round(total_cost * (1 + gst / 100), 2)
            if (total_cost is not None and gst)
            else total_cost
        )
        total_cost_accepted_gst = (
            round(total_cost_accepted * (1 + gst / 100), 2)
            if (total_cost_accepted is not None and gst)
            else total_cost_accepted
        )

        total_qty = current_stock + open_po
        ads = round(last_30_sales / 30, 2)
        target_stock = int(
            (Decimal(str(ads)) * Decimal(COVERAGE_DAYS)).quantize(
                Decimal("1"), rounding=ROUND_HALF_UP
            )
        )
        max_allowed_qty = target_stock - total_qty

        lead_time = item.get("lead_time_override") or 10
        coverage_days_item = item.get("coverage_days_override") or COVERAGE_DAYS
        total_target_days = lead_time + coverage_days_item
        net_total_days = round(total_qty / final_drr, 2) if final_drr else None
        monthly_sales = monthly_sales_by_asin.get(asin, [0] * 5)

        # Final Units (for under-ordering): formula per overstock sheet
        final_units_override = item.get("final_units_override")
        if final_units_override is not None:
            final_units = final_units_override
        elif final_drr and net_total_days is not None:
            if net_total_days < lead_time:
                final_units = int(
                    (Decimal(str(final_drr)) * Decimal(str(total_target_days))).quantize(
                        Decimal("1"), rounding=ROUND_HALF_UP
                    )
                )
            elif net_total_days > total_target_days:
                final_units = 0
            else:
                diff_days = Decimal(str(total_target_days)) - Decimal(str(net_total_days))
                final_units = int(
                    (diff_days * Decimal(str(final_drr))).quantize(
                        Decimal("1"), rounding=ROUND_HALF_UP
                    )
                )
        else:
            final_units = None

        # Final Supply Qty (for over-ordering): min(final_units, requested_qty)
        final_supply_fo_override = item.get("final_supply_fo_override")
        requested_qty = item["requested_qty"]
        if final_supply_fo_override is not None:
            final_supply_fo = final_supply_fo_override
        elif final_units is not None:
            final_supply_fo = min(final_units, requested_qty)
        else:
            final_supply_fo = None

        # Costs based on final_supply_fo (matches the Excel "Supply Qty" column I = AM)
        if final_supply_fo is not None:
            if mrp_wo_gst is not None and margin is not None:
                _rate = round(mrp_wo_gst, 2)
                _discount_factor = 1 - round(margin * 100, 2) / 100
                total_cost_fo = round(_rate * final_supply_fo * _discount_factor, 2)
            elif cost_price_wo_tax is not None:
                total_cost_fo = round(cost_price_wo_tax * final_supply_fo, 2)
            else:
                total_cost_fo = None
            total_cost_fo_gst = (
                round(total_cost_fo * (1 + gst / 100), 2)
                if (total_cost_fo is not None and gst)
                else total_cost_fo
            )
        else:
            total_cost_fo = None
            total_cost_fo_gst = None

        enriched.append(
            {
                **item,
                "supply_qty": supply_qty,
                "accepted_qty": accepted_qty,
                "received_qty": received_qty,
                "zoho_mrp": mrp,
                "etrade_asp": etrade_asp_by_asin.get(asin),
                "gst": gst,
                "mrp_wo_gst": mrp_wo_gst,
                "margin": margin,
                "cost_price_wo_tax": cost_price_wo_tax,
                "total_cost": total_cost,
                "total_cost_gst": total_cost_gst,
                "total_cost_fo": total_cost_fo,
                "total_cost_fo_gst": total_cost_fo_gst,
                "total_cost_accepted": total_cost_accepted,
                "total_cost_accepted_gst": total_cost_accepted_gst,
                "hsn": str(product.get("hsn_or_sac") or ""),
                "diff": diff,
                "zoho_stock": zoho_stock,
                "purchase_status": product.get("purchase_status") or "",
                "current_stock": current_stock,
                "open_po": open_po,
                "open_po_override": item.get("open_po_override"),
                "supply_qty_override": item.get("supply_qty_override"),
                "total_qty": total_qty,
                "last_30_sales": last_30_sales,
                "ads": ads,
                "coverage_days": coverage_days_item,
                "coverage_days_override": item.get("coverage_days_override"),
                "lead_time": lead_time,
                "lead_time_override": item.get("lead_time_override"),
                "net_total_days": net_total_days,
                "total_target_days": total_target_days,
                "target_stock": int(target_stock),
                "max_allowed_qty": int(max_allowed_qty),
                "final_supply_qty": supply_qty,  # final = supply (they are always equal)
                "final_drr": final_drr,
                "final_drr_flag": final_drr_flag,
                "final_units": final_units,
                "final_units_override": item.get("final_units_override"),
                "final_supply_fo": final_supply_fo,
                "final_supply_fo_override": item.get("final_supply_fo_override"),
                "monthly_sales": monthly_sales,
                "month_labels": month_labels,
            }
        )

    return enriched, inventory_date_str, zoho_stock_date_str


# ─── endpoints ────────────────────────────────────────────────────────────────


@router.post("/upload")
async def upload_vendor_po(
    file: UploadFile = File(...),
    po_date: str = Form(...),
    db=Depends(get_database),
):
    """Upload a Vendor Central PO Excel file and store it."""
    try:
        datetime.strptime(po_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="po_date must be YYYY-MM-DD")

    file_bytes = await file.read()

    # Parse Excel outside the thread (pure bytes, no I/O)
    try:
        po_number, vendor, items = await asyncio.to_thread(_parse_po_excel, file_bytes)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if not po_number:
        raise HTTPException(status_code=400, detail="Could not extract PO number from file")

    # Compute DRR using same async path as PSR report
    asins = [it["asin"] for it in items if it.get("asin")]
    drr_map = await _compute_drr_for_po_async(db, po_date, asins)

    def _process():
        existing = db[PO_COLLECTION].find_one({"po_number": po_number})
        if existing:
            current_status = existing.get("po_status", "pending")
            if current_status not in {"pending", "processing", "packed", "closed"}:
                raise ValueError(
                    f"PO {po_number} already exists with status '{current_status}' which cannot be overwritten"
                )
            # Re-enrich with fresh data, then preserve accepted_qty/received_qty per item
            enriched_items, inventory_date, zoho_stock_date = _enrich_items(
                items, po_number, db, po_date_str=po_date, drr_map=drr_map
            )
            existing_by_asin = {it["asin"]: it for it in existing.get("items", [])}
            for item in enriched_items:
                prev = existing_by_asin.get(item["asin"], {})
                item["accepted_qty"] = prev.get(
                    "accepted_qty", item.get("accepted_qty", 0)
                )
                item["received_qty"] = prev.get(
                    "received_qty", item.get("received_qty")
                )
            db[PO_COLLECTION].update_one(
                {"po_number": po_number},
                {
                    "$set": {
                        "vendor": vendor,
                        "po_date": po_date,
                        "po_status": current_status,
                        "inventory_date": inventory_date,
                        "zoho_stock_date": zoho_stock_date,
                        "item_count": len(items),
                        "updated_at": datetime.now(),
                        "items": enriched_items,
                    }
                },
            )
            return po_number, enriched_items, inventory_date, zoho_stock_date

        enriched_items, inventory_date, zoho_stock_date = _enrich_items(
            items, po_number, db, po_date_str=po_date, drr_map=drr_map
        )

        doc = {
            "po_number": po_number,
            "vendor": vendor,
            "po_date": po_date,
            "po_status": "pending",
            "inventory_date": inventory_date,
            "zoho_stock_date": zoho_stock_date,
            "item_count": len(items),
            "created_at": datetime.now(),
            "updated_at": datetime.now(),
            "items": enriched_items,  # enriched items stored (stock/sales frozen at upload time)
        }
        db[PO_COLLECTION].insert_one(doc)
        return po_number, enriched_items, inventory_date, zoho_stock_date

    try:
        po_number, enriched_items, inventory_date, zoho_stock_date = (
            await asyncio.to_thread(_process)
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return JSONResponse(
        status_code=201,
        content={
            "po_number": po_number,
            "inventory_date": inventory_date,
            "zoho_stock_date": zoho_stock_date,
            "items": serialize_mongo_document(enriched_items),
        },
    )


@router.get("/")
async def list_vendor_pos(db=Depends(get_database)):
    """List all purchase orders with total requested/accepted/received qty sums."""

    def _fetch():
        pipeline = [
            {
                "$addFields": {
                    "total_requested_qty": {"$sum": "$items.requested_qty"},
                    "total_accepted_qty": {"$sum": "$items.accepted_qty"},
                    "total_received_qty": {
                        "$ifNull": ["$received_qty", {"$sum": "$items.received_qty"}]
                    },
                    # Supply Qty mirrors Excel col I: supply_qty_override if set, else final_supply_fo
                    "total_supply_qty": {
                        "$sum": {
                            "$map": {
                                "input": "$items",
                                "as": "it",
                                "in": {
                                    "$cond": [
                                        {"$ne": ["$$it.supply_qty_override", None]},
                                        "$$it.supply_qty_override",
                                        {"$ifNull": ["$$it.final_supply_fo", 0]},
                                    ]
                                },
                            }
                        }
                    },
                    "_items_cost": {
                        "$sum": {
                            "$map": {
                                "input": "$items",
                                "as": "it",
                                "in": {"$ifNull": ["$$it.total_cost_fo", {"$ifNull": ["$$it.total_cost", 0]}]},
                            }
                        }
                    },
                    "_items_cost_gst": {
                        "$sum": {
                            "$map": {
                                "input": "$items",
                                "as": "it",
                                "in": {"$ifNull": ["$$it.total_cost_fo_gst", {"$ifNull": ["$$it.total_cost_gst", 0]}]},
                            }
                        }
                    },
                }
            },
            # Join estimate to get authoritative sub_total / total when linked
            {
                "$lookup": {
                    "from": ESTIMATES_COLLECTION,
                    "localField": "zoho_estimate_id",
                    "foreignField": "estimate_id",
                    "as": "_est",
                }
            },
            {
                "$addFields": {
                    "_est": {"$arrayElemAt": ["$_est", 0]},
                }
            },
            # Join sales order by estimate_id to derive packages automatically
            {
                "$lookup": {
                    "from": SALES_ORDERS_COLLECTION,
                    "localField": "zoho_estimate_id",
                    "foreignField": "estimate_id",
                    "as": "_so",
                }
            },
            {
                "$addFields": {
                    "_so": {"$arrayElemAt": ["$_so", 0]},
                }
            },
            {
                "$addFields": {
                    # null = no SO found; [] = SO exists but no packages; [...] = package numbers
                    "so_packages": {
                        "$cond": {
                            "if": {"$gt": ["$_so", None]},
                            "then": {
                                "$map": {
                                    "input": {"$ifNull": ["$_so.packages", []]},
                                    "as": "pkg",
                                    "in": "$$pkg.package_number",
                                }
                            },
                            "else": None,
                        }
                    },
                    "so_number": "$_so.salesorder_number",
                }
            },
            {
                "$addFields": {
                    # Use estimate sub_total/total when available, else fall back to item sums
                    "total_cost": {
                        "$ifNull": ["$_est.sub_total", "$_items_cost"]
                    },
                    "total_cost_gst": {
                        "$ifNull": ["$_est.total", "$_items_cost_gst"]
                    },
                }
            },
            {
                "$project": {
                    "po_number": 1,
                    "vendor": 1,
                    "po_date": 1,
                    "po_status": 1,
                    "item_count": 1,
                    "created_at": 1,
                    "order_file_s3_key": 1,
                    "total_requested_qty": 1,
                    "total_accepted_qty": 1,
                    "total_received_qty": 1,
                    "total_supply_qty": 1,
                    "total_cost": 1,
                    "total_cost_gst": 1,
                    "estimate_number": 1,
                    "zoho_estimate_id": 1,
                    "package_number": 1,
                    "accepted_total_cost": 1,
                    "accepted_total_cost_gst": 1,
                    "transfer_order_number": 1,
                    "transfer_order_id": 1,
                    "bundle_ids": 1,
                    "assembly_numbers": 1,
                    "packages": 1,
                    "sales_order_no": 1,
                    "so_packages": 1,
                    "so_number": 1,
                    "_id": 0,
                }
            },
            {"$sort": {"po_date": -1}},
        ]
        return list(db[PO_COLLECTION].aggregate(pipeline))

    pos = await asyncio.to_thread(_fetch)
    return serialize_mongo_document(pos)


# ─── shipment summary download / upload (must precede /{po_number}/... routes) ─

SHIPMENT_DATE_FIELDS = {
    "appointment_initiated_date",
    "appointment_date",
    "dispatched_date",
    "delivery_date",
}

SHIPMENT_COLUMNS = [
    ("PO Date", "po_date", False),
    ("PO Number", "po_number", False),
    ("Location", "location", False),
    ("Requested Qty", "total_requested_qty", False),
    ("Supply Qty (After Overstock)", "total_supply_qty", False),
    ("Accepted / Dispatched Qty", "total_accepted_qty", False),
    ("Short Supply (Qty)", "short_supply_qty", False),
    ("Short Supply (%)", "short_supply_pct", False),
    ("Reason for Short Supply", "reason_for_short_supply", True),
    ("Box Count", "box_count", True),
    ("Appointment Initiated Date", "appointment_initiated_date", True),
    ("Appointment ID", "appointment_id", True),
    ("Appointment Date", "appointment_date", True),
    ("Dispatched Date", "dispatched_date", True),
    ("Status", "po_status", False),
    ("Delivery Date", "delivery_date", True),
]

_DATE_FMT_IN = ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"]

EDITABLE_FIELDS = {
    "Reason for Short Supply": "reason_for_short_supply",
    "Box Count": "box_count",
    "Appointment Initiated Date": "appointment_initiated_date",
    "Appointment ID": "appointment_id",
    "Appointment Date": "appointment_date",
    "Dispatched Date": "dispatched_date",
    "Delivery Date": "delivery_date",
}


def _to_ddmmyyyy(val: str | None) -> str:
    if not val:
        return ""
    for fmt in _DATE_FMT_IN:
        try:
            return datetime.strptime(val, fmt).strftime("%d/%m/%Y")
        except ValueError:
            pass
    return val


def _parse_date_field(val: str | None) -> str | None:
    """Normalise date strings from an uploaded file to yyyy-mm-dd for DB storage."""
    if not val:
        return None
    val = str(val).strip()
    for fmt in _DATE_FMT_IN:
        try:
            return datetime.strptime(val, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return val


@router.get("/shipment_summary/download")
async def download_shipment_summary(db=Depends(get_database)):
    """Download the Etrade Shipment Summary as an XLSX file."""

    def _fetch():
        pipeline = [
            {
                "$addFields": {
                    "total_requested_qty": {"$sum": "$items.requested_qty"},
                    "total_supply_qty": {"$sum": "$items.final_supply_fo"},
                    "total_accepted_qty": {"$sum": "$items.accepted_qty"},
                    "location": {"$arrayElemAt": ["$items.ship_to_location", 0]},
                }
            },
            {
                "$project": {
                    "po_number": 1,
                    "po_date": 1,
                    "po_status": 1,
                    "location": 1,
                    "total_requested_qty": 1,
                    "total_supply_qty": 1,
                    "total_accepted_qty": 1,
                    "reason_for_short_supply": 1,
                    "box_count": 1,
                    "appointment_initiated_date": 1,
                    "appointment_id": 1,
                    "appointment_date": 1,
                    "dispatched_date": 1,
                    "delivery_date": 1,
                    "_id": 0,
                }
            },
            {"$sort": {"po_date": -1}},
        ]
        return list(db[PO_COLLECTION].aggregate(pipeline))

    raw_rows = await asyncio.to_thread(_fetch)
    rows = serialize_mongo_document(raw_rows)

    for r in rows:
        supply = r.get("total_supply_qty") or 0
        accepted = r.get("total_accepted_qty") or 0
        r["short_supply_qty"] = accepted - supply
        r["short_supply_pct"] = (
            f"{((accepted - supply) / supply * 100):.2f}%" if supply else ""
        )
        for field in ["po_date"] + list(SHIPMENT_DATE_FIELDS):
            r[field] = _to_ddmmyyyy(r.get(field))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shipment Summary"

    header_fill_ro = PatternFill("solid", fgColor="D9E1F2")
    header_fill_ed = PatternFill("solid", fgColor="FFF2CC")
    cell_fill_ed = PatternFill("solid", fgColor="FFFCE5")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (label, _, editable) in enumerate(SHIPMENT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill_ed if editable else header_fill_ro
        cell.alignment = center

    ws.row_dimensions[1].height = 30

    for row_idx, r in enumerate(rows, start=2):
        for col_idx, (_, field, editable) in enumerate(SHIPMENT_COLUMNS, start=1):
            val = r.get(field)
            cell = ws.cell(
                row=row_idx, column=col_idx, value=val if val is not None else ""
            )
            cell.alignment = Alignment(vertical="center")
            if editable:
                cell.fill = cell_fill_ed

    for col_idx, (label, _, _) in enumerate(SHIPMENT_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(label),
            *(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, len(rows) + 2)
            ),
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=etrade_shipment_summary.xlsx"
        },
    )


@router.post("/shipment_summary/upload")
async def upload_shipment_summary(
    file: UploadFile = File(...), db=Depends(get_database)
):
    """Upload an edited Etrade Shipment Summary XLSX to bulk-update editable fields."""
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(
            status_code=400, detail="Could not parse uploaded file as XLSX"
        )

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="File is empty")

    header = [str(h).strip() if h is not None else "" for h in header_row]

    try:
        po_col = header.index("PO Number")
    except ValueError:
        raise HTTPException(
            status_code=400, detail="'PO Number' column not found in uploaded file"
        )

    field_col_map: dict[str, int] = {}
    for label, field in EDITABLE_FIELDS.items():
        try:
            field_col_map[field] = header.index(label)
        except ValueError:
            pass

    errors: list[str] = []

    def _bulk_update(updates: list[tuple[str, dict]]):
        from pymongo import UpdateOne

        ops = [
            UpdateOne(
                {"po_number": po}, {"$set": {**fields, "updated_at": datetime.now()}}
            )
            for po, fields in updates
        ]
        if ops:
            db[PO_COLLECTION].bulk_write(ops, ordered=False)
        return len(ops)

    batch: list[tuple[str, dict]] = []
    for data_row in rows_iter:
        po_number = data_row[po_col]
        if not po_number:
            continue
        po_number = str(po_number).strip()

        fields: dict = {}
        for field, col_idx in field_col_map.items():
            raw = data_row[col_idx] if col_idx < len(data_row) else None
            if raw is None or str(raw).strip() == "":
                fields[field] = None
            elif field == "box_count":
                try:
                    fields[field] = int(float(str(raw).strip()))
                except (ValueError, TypeError):
                    errors.append(f"PO {po_number}: invalid box_count '{raw}'")
                    continue
            elif field in SHIPMENT_DATE_FIELDS:
                fields[field] = _parse_date_field(str(raw).strip())
            else:
                fields[field] = str(raw).strip() or None

        if fields:
            batch.append((po_number, fields))

    updated = await asyncio.to_thread(_bulk_update, batch) if batch else 0

    result: dict = {"updated": updated}
    if errors:
        result["errors"] = errors
    return result


@router.get("/{po_number}/report")
async def get_po_report(po_number: str, db=Depends(get_database)):
    """Generate enriched report for a PO."""

    def _get_doc():
        return db[PO_COLLECTION].find_one({"po_number": po_number})

    doc = await asyncio.to_thread(_get_doc)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"PO {po_number} not found")

    # Compute DRR using same async path as PSR report
    _items = doc.get("items", [])
    _asins = [it["asin"] for it in _items if it.get("asin")]
    _drr_map = await _compute_drr_for_po_async(db, doc.get("po_date", ""), _asins)

    def _fetch():
        items = doc.get("items", [])
        po_status = doc.get("po_status", "pending")
        use_stored = po_status in FREEZE_ON_STATUS
        enriched, inv_date, zoho_date = _enrich_items(
            items,
            po_number,
            db,
            use_stored_stock=use_stored,
            po_date_str=doc.get("po_date"),
            drr_map=_drr_map,
        )
        # Attach per-item dispatched costs from all linked packages (if any)
        pkg_numbers = doc.get("packages") or ([doc["package_number"]] if doc.get("package_number") else [])
        if pkg_numbers:
            merged: dict = {}
            for pn in pkg_numbers:
                for asin, d in _get_dispatched_costs_by_asin(pn, db).items():
                    if asin in merged:
                        merged[asin]["qty"] += d["qty"]
                        merged[asin]["item_total"] += d["item_total"]
                        merged[asin]["item_total_gst"] += d["item_total_gst"]
                    else:
                        merged[asin] = dict(d)
            for it in enriched:
                d = merged.get(it.get("asin"))
                if d:
                    it["dispatched_qty"] = d["qty"]
                    it["total_cost_dispatched"] = d["item_total"]
                    it["total_cost_dispatched_gst"] = d["item_total_gst"]
        return enriched, inv_date, zoho_date

    enriched, live_inv_date, live_zoho_date = await asyncio.to_thread(_fetch)

    # Write computed total_cost back to each item in the DB so the list aggregation reflects it.
    def _write_back_costs():
        for it in enriched:
            asin = it.get("asin")
            db[PO_COLLECTION].update_one(
                {"po_number": po_number, "items.asin": asin},
                {"$set": {"items.$.total_cost": it.get("total_cost"), "items.$.total_cost_gst": it.get("total_cost_gst")}},
            )

    asyncio.create_task(asyncio.to_thread(_write_back_costs))

    po_status = doc["po_status"]
    # For new-format frozen POs, live_inv_date is None (stored snapshot used).
    # For old-format frozen POs (no zoho_stock in items), live_inv_date is freshly computed
    # with T-2 cutoff — prefer it over the stale stored value.
    inv_date = live_inv_date if live_inv_date is not None else doc.get("inventory_date")
    zoho_date = (
        live_zoho_date if live_zoho_date is not None else doc.get("zoho_stock_date")
    )

    return {
        "po_number": doc["po_number"],
        "vendor": doc.get("vendor"),
        "po_date": doc["po_date"],
        "po_status": po_status,
        "inventory_date": inv_date,
        "zoho_stock_date": zoho_date,
        "po_update_date": doc.get("po_update_date"),
        "estimate_number": doc.get("estimate_number"),
        "zoho_estimate_id": doc.get("zoho_estimate_id"),
        "transfer_order_number": doc.get("transfer_order_number"),
        "transfer_order_id": doc.get("transfer_order_id"),
        "bundle_ids": doc.get("bundle_ids") or [],
        "assembly_numbers": doc.get("assembly_numbers") or [],
        "packages": doc.get("packages") or [],
        "sales_order_no": doc.get("sales_order_no"),
        "items": serialize_mongo_document(enriched),
    }


@router.get("/{po_number}/estimate_diff")
async def get_estimate_diff(po_number: str, db=Depends(get_database)):
    """Compare PO items against linked estimate line items."""

    def _fetch():
        doc = db[PO_COLLECTION].find_one({"po_number": po_number})
        if not doc:
            return None, None
        est = None
        if doc.get("zoho_estimate_id"):
            est = db[ESTIMATES_COLLECTION].find_one({"estimate_id": doc["zoho_estimate_id"]})
        if not est and doc.get("estimate_number"):
            est = db[ESTIMATES_COLLECTION].find_one({"estimate_number": doc["estimate_number"]})
        return doc, est

    doc, est = await asyncio.to_thread(_fetch)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"PO {po_number} not found")
    if est is None:
        raise HTTPException(status_code=404, detail="No estimate linked to this PO")

    def _cf_sku(li: dict) -> str:
        for cf in li.get("item_custom_fields", []):
            if cf.get("api_name") == "cf_sku_code":
                return cf.get("value") or ""
        return ""

    # Build estimate lookup by cf_sku_code
    est_by_sku: dict = {}
    for li in est.get("line_items", []):
        sku = _cf_sku(li)
        if sku:
            est_by_sku[sku] = li

    # Compute PO item totals using the estimate formula
    po_items = [it for it in doc.get("items", []) if (it.get("status") or "").lower() != "inactive"]

    # Build ASIN → cf_sku_code map so PO items with ASINs as model_number can match estimate line items keyed by cf_sku_code
    po_model_numbers = [it["model_number"] for it in po_items if it.get("model_number")]
    def _fetch_sku_map():
        return {
            m["item_id"]: m["sku_code"]
            for m in db[SKU_MAPPING_COLLECTION].find(
                {"item_id": {"$in": po_model_numbers}},
                {"item_id": 1, "sku_code": 1},
            )
            if m.get("item_id") and m.get("sku_code")
        }
    asin_to_sku_diff: dict[str, str] = await asyncio.to_thread(_fetch_sku_map)

    rows = []
    for it in po_items:
        model = it.get("model_number") or ""
        # Estimate is created with final_supply_fo (override takes precedence, same as create logic)
        supply = it.get("final_supply_fo_override") if it.get("final_supply_fo_override") is not None else (it.get("final_supply_fo") if it.get("final_supply_fo") is not None else 0)
        mrp_wo_gst = it.get("mrp_wo_gst")
        margin = it.get("margin")
        gst = it.get("gst") or 0

        if mrp_wo_gst is not None and margin is not None:
            rate = round(mrp_wo_gst, 2)
            disc = round(margin * 100, 2) / 100
            po_item_total = round(rate * supply * (1 - disc), 2)
        elif it.get("cost_price_wo_tax") is not None:
            rate = it["cost_price_wo_tax"]
            po_item_total = round(rate * supply, 2)
        else:
            rate = None
            po_item_total = None

        # model_number may be an ASIN — resolve to cf_sku_code for estimate lookup
        sku_key = asin_to_sku_diff.get(model, model)
        eli = est_by_sku.get(sku_key)
        in_estimate = eli is not None
        est_qty = eli.get("quantity") if eli else None
        est_rate = eli.get("rate") if eli else None
        est_item_total = eli.get("item_total") if eli else None

        rows.append({
            "model_number": model,
            "title": it.get("title") or "",
            "asin": it.get("asin") or "",
            "supply_qty": supply,
            "mrp_wo_gst": mrp_wo_gst,
            "margin": margin,
            "gst": gst,
            "po_rate": rate,
            "po_item_total": po_item_total,
            "in_estimate": in_estimate,
            "estimate_qty": est_qty,
            "estimate_rate": est_rate,
            "estimate_item_total": est_item_total,
            "qty_diff": (supply - est_qty) if (in_estimate and est_qty is not None) else None,
            "rate_diff": round(rate - est_rate, 4) if (in_estimate and rate is not None and est_rate is not None) else None,
            "total_diff": round(po_item_total - est_item_total, 2) if (in_estimate and po_item_total is not None and est_item_total is not None) else None,
        })

    # Items in estimate but not in PO
    # Resolve ASIN model_numbers to cf_sku_codes so the comparison uses the same key as est_by_sku
    po_skus = {asin_to_sku_diff.get(it.get("model_number", ""), it.get("model_number", "")) for it in po_items}
    only_in_estimate = [
        {"sku": _cf_sku(li), "name": li.get("name", ""), "quantity": li.get("quantity"), "rate": li.get("rate"), "item_total": li.get("item_total")}
        for li in est.get("line_items", [])
        if _cf_sku(li) and _cf_sku(li) not in po_skus
    ]

    po_total = round(sum(r["po_item_total"] for r in rows if r["po_item_total"] is not None), 2)
    po_total_gst = round(sum(
        round(r["po_item_total"] * (1 + r["gst"] / 100), 2) if r["gst"] else r["po_item_total"]
        for r in rows if r["po_item_total"] is not None
    ), 2)

    return {
        "po_number": po_number,
        "estimate_number": est.get("estimate_number"),
        "estimate_sub_total": est.get("sub_total"),
        "estimate_total": est.get("total"),
        "po_computed_total": po_total,
        "po_computed_total_gst": po_total_gst,
        "items": serialize_mongo_document(rows),
        "only_in_estimate": serialize_mongo_document(only_in_estimate),
    }


@router.get("/{po_number}/package_breakdown")
async def get_package_breakdown(po_number: str, db=Depends(get_database)):
    """Return per-item accepted cost breakdown for the linked package."""

    def _fetch():
        doc = db[PO_COLLECTION].find_one({"po_number": po_number})
        if not doc:
            return None, None, []
        # Derive package numbers from linked sales order
        est_id = doc.get("zoho_estimate_id")
        pkg_numbers: list[str] = []
        if est_id:
            so = db[SALES_ORDERS_COLLECTION].find_one({"estimate_id": est_id}, {"packages": 1})
            if so:
                pkg_numbers = [p["package_number"] for p in (so.get("packages") or []) if p.get("package_number")]
        # Fall back to manually linked packages for backward compatibility
        if not pkg_numbers:
            pkg_numbers = doc.get("packages") or ([doc["package_number"]] if doc.get("package_number") else [])
        if not pkg_numbers:
            return doc, None, []
        pkg = db[PACKAGES_COLLECTION].find_one({"package_number": pkg_numbers[0]})
        return doc, pkg, pkg_numbers

    doc, pkg, pkg_numbers = await asyncio.to_thread(_fetch)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"PO {po_number} not found")
    if pkg is None:
        raise HTTPException(status_code=404, detail="No package linked to this PO")

    package_number = pkg_numbers[0]
    line_items = pkg.get("line_items") or []

    # Extract SKU, qty, name per line item
    sku_rows: list[dict] = []
    for li in line_items:
        sku = None
        for cf in li.get("item_custom_fields") or []:
            if cf.get("api_name") == "cf_sku_code":
                sku = cf.get("value")
                break
        if sku:
            sku_rows.append({
                "sku": sku,
                "name": li.get("name") or "",
                "qty": float(li.get("quantity") or 0),
            })

    if not sku_rows:
        return {
            "po_number": po_number,
            "package_number": package_number,
            "accepted_total_cost": None,
            "accepted_total_cost_gst": None,
            "items": [],
        }

    def _enrich():
        skus = [r["sku"] for r in sku_rows]
        sku_to_asin: dict[str, str] = {}
        for m in db[SKU_MAPPING_COLLECTION].find({"sku_code": {"$in": skus}}, {"sku_code": 1, "item_id": 1}):
            sku_to_asin[m["sku_code"]] = m["item_id"]
        asins = list(sku_to_asin.values())
        asin_to_margin: dict[str, float] = {}
        asin_to_cost_price: dict[str, float] = {}
        asin_to_etrade: dict[str, float] = {}
        if asins:
            for m in db[MARGINS_COLLECTION].find(
                {"asin": {"$in": asins}},
                {"asin": 1, "etrade_asp": 1, "margin": 1, "cost_price_wo_tax": 1},
            ):
                if m.get("etrade_asp") is not None:
                    asin_to_etrade[m["asin"]] = float(m["etrade_asp"])
                if m.get("margin") is not None:
                    asin_to_margin[m["asin"]] = float(m["margin"])
                if m.get("cost_price_wo_tax") is not None:
                    asin_to_cost_price[m["asin"]] = float(m["cost_price_wo_tax"])
        sku_to_gst: dict[str, float] = {}
        for p in db[PRODUCTS_COLLECTION].find(
            {"cf_sku_code": {"$in": skus}, "status": "active"},
            {"cf_sku_code": 1, "item_tax_preferences": 1},
        ):
            sku_to_gst[p["cf_sku_code"]] = _extract_gst(p.get("item_tax_preferences") or [])
        missing_gst_skus = [s for s in skus if s not in sku_to_gst]
        if missing_gst_skus:
            composite_item_ids = {
                c["sku_code"]: c["composite_item_id"]
                for c in db["composite_products"].find(
                    {"sku_code": {"$in": missing_gst_skus}},
                    {"sku_code": 1, "composite_item_id": 1},
                )
                if c.get("sku_code") and c.get("composite_item_id")
            }
            if composite_item_ids:
                for p in db[PRODUCTS_COLLECTION].find(
                    {"item_id": {"$in": list(composite_item_ids.values())}},
                    {"item_id": 1, "item_tax_preferences": 1},
                ):
                    for sku, iid in composite_item_ids.items():
                        if p.get("item_id") == iid:
                            sku_to_gst[sku] = _extract_gst(p.get("item_tax_preferences") or [])
        return sku_to_asin, asin_to_etrade, asin_to_margin, asin_to_cost_price, sku_to_gst

    sku_to_asin, asin_to_etrade, asin_to_margin, asin_to_cost_price, sku_to_gst = await asyncio.to_thread(_enrich)

    items = []
    total_cost = 0.0
    total_cost_gst = 0.0
    for r in sku_rows:
        sku = r["sku"]
        qty = r["qty"]
        asin = sku_to_asin.get(sku)
        gst = sku_to_gst.get(sku, 0.0)
        etrade_asp = asin_to_etrade.get(asin) if asin else None
        margin = asin_to_margin.get(asin) if asin else None

        mrp_wo_gst = None
        if etrade_asp is not None:
            mrp_wo_gst = round(etrade_asp / (1 + gst / 100), 2) if gst else etrade_asp

        item_cost: float | None = None
        unit_cost: float | None = None
        if mrp_wo_gst is not None and margin is not None:
            unit_cost = round(round(mrp_wo_gst, 2) * (1 - round(margin * 100, 2) / 100), 4)
            item_cost = round(unit_cost * qty, 2)
        elif asin and asin in asin_to_cost_price:
            unit_cost = asin_to_cost_price[asin]
            item_cost = round(unit_cost * qty, 2)

        item_cost_gst = round(item_cost * (1 + gst / 100), 2) if (item_cost is not None and gst) else item_cost

        if item_cost is not None:
            total_cost += item_cost
            total_cost_gst += item_cost_gst or item_cost

        items.append({
            "sku": sku,
            "name": r["name"],
            "qty": qty,
            "asin": asin,
            "etrade_asp": etrade_asp,
            "gst": gst,
            "mrp_wo_gst": mrp_wo_gst,
            "margin": round(margin * 100, 2) if margin is not None else None,  # as percentage
            "unit_cost": unit_cost,
            "item_total": item_cost,
            "item_total_gst": item_cost_gst,
        })

    return {
        "po_number": po_number,
        "package_number": package_number,
        "accepted_total_cost": round(total_cost, 2),
        "accepted_total_cost_gst": round(total_cost_gst, 2),
        "items": serialize_mongo_document(items),
    }


VALID_STATUSES = {
    "pending",
    "processing",
    "packed",
    "closed",
    "intransit",
    "delivered",
    "completed",
}
FROZEN_STATUSES = {"packed", "closed", "intransit", "delivered", "completed"}
# Processing also triggers a freeze — stock/sales are locked when PO moves to processing
FREEZE_ON_STATUS = FROZEN_STATUSES | {"processing"}


@router.delete("/{po_number}")
async def delete_vendor_po(po_number: str, db=Depends(get_database)):
    """Delete a purchase order by PO number."""

    def _delete():
        return db[PO_COLLECTION].delete_one({"po_number": po_number}).deleted_count

    count = await asyncio.to_thread(_delete)
    if not count:
        raise HTTPException(status_code=404, detail=f"PO {po_number} not found")
    return {"po_number": po_number, "deleted": True}


@router.patch("/{po_number}/status")
async def update_po_status(po_number: str, po_status: str, db=Depends(get_database)):
    if po_status not in VALID_STATUSES:
        raise HTTPException(
            status_code=400, detail=f"status must be one of {VALID_STATUSES}"
        )

    def _get_doc():
        return db[PO_COLLECTION].find_one({"po_number": po_number})

    doc = await asyncio.to_thread(_get_doc)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"PO {po_number} not found")

    # If transitioning to a frozen status, compute DRR using PSR async path before the thread
    _freeze_drr_map: dict = {}
    if po_status in FREEZE_ON_STATUS and doc.get("po_status") not in FREEZE_ON_STATUS:
        _freeze_asins = [it["asin"] for it in doc.get("items", []) if it.get("asin")]
        if _freeze_asins and doc.get("po_date"):
            _freeze_drr_map = await _compute_drr_for_po_async(db, doc.get("po_date", ""), _freeze_asins)

    def _update():
        update_fields: dict = {"po_status": po_status, "updated_at": datetime.now()}

        # Transitioning into processing or a frozen status: re-enrich with live T-2 data and freeze permanently
        if (
            po_status in FREEZE_ON_STATUS
            and doc.get("po_status") not in FREEZE_ON_STATUS
        ):
            enriched, inv_date, zoho_date = _enrich_items(
                doc.get("items", []),
                po_number,
                db,
                use_stored_stock=False,
                po_date_str=doc.get("po_date"),
                drr_map=_freeze_drr_map,
            )
            update_fields["items"] = enriched
            update_fields["inventory_date"] = inv_date
            update_fields["zoho_stock_date"] = zoho_date
            if po_status == "processing":
                update_fields["po_update_date"] = datetime.now().strftime("%Y-%m-%d")

        db[PO_COLLECTION].update_one({"po_number": po_number}, {"$set": update_fields})
        return po_status

    await asyncio.to_thread(_update)
    return {"po_number": po_number, "po_status": po_status}


@router.patch("/{po_number}/items/{asin}/accepted_qty")
async def update_item_accepted_qty(
    po_number: str, asin: str, accepted_qty: int, db=Depends(get_database)
):
    """Update the accepted quantity for a specific item in a PO."""
    if accepted_qty < 0:
        raise HTTPException(status_code=400, detail="accepted_qty must be >= 0")

    def _update():
        result = db[PO_COLLECTION].update_one(
            {"po_number": po_number, "items.asin": asin},
            {
                "$set": {
                    "items.$.accepted_qty": accepted_qty,
                    "updated_at": datetime.now(),
                }
            },
        )
        return result.matched_count

    matched = await asyncio.to_thread(_update)
    if not matched:
        raise HTTPException(
            status_code=404, detail=f"PO {po_number} / ASIN {asin} not found"
        )
    return {"po_number": po_number, "asin": asin, "accepted_qty": accepted_qty}


@router.patch("/{po_number}/items/{asin}/received_qty")
async def update_item_received_qty(
    po_number: str, asin: str, received_qty: int, db=Depends(get_database)
):
    """Update the received quantity for a specific item in a PO."""
    if received_qty < 0:
        raise HTTPException(status_code=400, detail="received_qty must be >= 0")

    def _update():
        result = db[PO_COLLECTION].update_one(
            {"po_number": po_number, "items.asin": asin},
            {
                "$set": {
                    "items.$.received_qty": received_qty,
                    "updated_at": datetime.now(),
                }
            },
        )
        return result.matched_count

    matched = await asyncio.to_thread(_update)
    if not matched:
        raise HTTPException(
            status_code=404, detail=f"PO {po_number} / ASIN {asin} not found"
        )
    return {"po_number": po_number, "asin": asin, "received_qty": received_qty}


@router.patch("/{po_number}/received_qty")
async def update_po_received_qty(
    po_number: str, received_qty: int, db=Depends(get_database)
):
    """Set total received quantity at the PO level (overrides per-item sum in list view)."""
    if received_qty < 0:
        raise HTTPException(status_code=400, detail="received_qty must be >= 0")

    def _update():
        result = db[PO_COLLECTION].update_one(
            {"po_number": po_number},
            {"$set": {"received_qty": received_qty, "updated_at": datetime.now()}},
        )
        return result.matched_count

    matched = await asyncio.to_thread(_update)
    if not matched:
        raise HTTPException(status_code=404, detail=f"PO {po_number} not found")
    return {"po_number": po_number, "received_qty": received_qty}


@router.patch("/{po_number}/items/{asin}/open_qty")
async def update_item_open_qty(
    po_number: str, asin: str, open_qty: int, db=Depends(get_database)
):
    """Override the open PO quantity for a specific item. Use -1 to clear the override and revert to auto-computed."""

    def _update():
        override_value = None if open_qty < 0 else open_qty
        result = db[PO_COLLECTION].update_one(
            {"po_number": po_number, "items.asin": asin},
            {
                "$set": {
                    "items.$.open_po_override": override_value,
                    "updated_at": datetime.now(),
                }
            },
        )
        return result.matched_count

    matched = await asyncio.to_thread(_update)
    if not matched:
        raise HTTPException(
            status_code=404, detail=f"PO {po_number} / ASIN {asin} not found"
        )
    return {
        "po_number": po_number,
        "asin": asin,
        "open_po_override": None if open_qty < 0 else open_qty,
    }


@router.patch("/{po_number}/items/{asin}/supply_qty")
async def update_item_supply_qty(
    po_number: str, asin: str, supply_qty: int, db=Depends(get_database)
):
    """Override the supply quantity for a specific item. Use -1 to clear the override and revert to auto-computed."""

    def _update():
        override_value = None if supply_qty < 0 else supply_qty
        result = db[PO_COLLECTION].update_one(
            {"po_number": po_number, "items.asin": asin},
            {
                "$set": {
                    "items.$.supply_qty_override": override_value,
                    "updated_at": datetime.now(),
                }
            },
        )
        return result.matched_count

    matched = await asyncio.to_thread(_update)
    if not matched:
        raise HTTPException(
            status_code=404, detail=f"PO {po_number} / ASIN {asin} not found"
        )
    return {
        "po_number": po_number,
        "asin": asin,
        "supply_qty_override": None if supply_qty < 0 else supply_qty,
    }


@router.patch("/{po_number}/items/{asin}/lead_time")
async def update_item_lead_time(
    po_number: str, asin: str, lead_time: int, db=Depends(get_database)
):
    """Set the lead time (days) for a specific item. Use -1 to reset to default (10)."""

    def _update():
        override_value = None if lead_time < 0 else lead_time
        result = db[PO_COLLECTION].update_one(
            {"po_number": po_number, "items.asin": asin},
            {"$set": {"items.$.lead_time_override": override_value, "updated_at": datetime.now()}},
        )
        return result.matched_count

    matched = await asyncio.to_thread(_update)
    if not matched:
        raise HTTPException(status_code=404, detail=f"PO {po_number} / ASIN {asin} not found")
    return {"po_number": po_number, "asin": asin, "lead_time": None if lead_time < 0 else lead_time}


@router.patch("/{po_number}/items/{asin}/coverage_days")
async def update_item_coverage_days(
    po_number: str, asin: str, coverage_days: int, db=Depends(get_database)
):
    """Override coverage days for a specific item. Use -1 to reset to default (35)."""

    def _update():
        override_value = None if coverage_days < 0 else coverage_days
        result = db[PO_COLLECTION].update_one(
            {"po_number": po_number, "items.asin": asin},
            {"$set": {"items.$.coverage_days_override": override_value, "updated_at": datetime.now()}},
        )
        return result.matched_count

    matched = await asyncio.to_thread(_update)
    if not matched:
        raise HTTPException(status_code=404, detail=f"PO {po_number} / ASIN {asin} not found")
    return {"po_number": po_number, "asin": asin, "coverage_days": None if coverage_days < 0 else coverage_days}


@router.patch("/{po_number}/items/{asin}/final_units")
async def update_item_final_units(
    po_number: str, asin: str, final_units: int, db=Depends(get_database)
):
    """Override Final Units (for under-ordering) for a specific item. Use -1 to reset to formula."""

    def _update():
        override_value = None if final_units < 0 else final_units
        result = db[PO_COLLECTION].update_one(
            {"po_number": po_number, "items.asin": asin},
            {"$set": {"items.$.final_units_override": override_value, "updated_at": datetime.now()}},
        )
        return result.matched_count

    matched = await asyncio.to_thread(_update)
    if not matched:
        raise HTTPException(status_code=404, detail=f"PO {po_number} / ASIN {asin} not found")
    return {"po_number": po_number, "asin": asin, "final_units": None if final_units < 0 else final_units}


@router.patch("/{po_number}/items/{asin}/final_supply_fo")
async def update_item_final_supply_fo(
    po_number: str, asin: str, final_supply_fo: int, db=Depends(get_database)
):
    """Override Final Supply Qty (for over-ordering) for a specific item. Use -1 to reset to formula."""

    def _update():
        override_value = None if final_supply_fo < 0 else final_supply_fo
        result = db[PO_COLLECTION].update_one(
            {"po_number": po_number, "items.asin": asin},
            {"$set": {"items.$.final_supply_fo_override": override_value, "updated_at": datetime.now()}},
        )
        return result.matched_count

    matched = await asyncio.to_thread(_update)
    if not matched:
        raise HTTPException(status_code=404, detail=f"PO {po_number} / ASIN {asin} not found")
    return {"po_number": po_number, "asin": asin, "final_supply_fo": None if final_supply_fo < 0 else final_supply_fo}


@router.patch("/{po_number}/items/{asin}/etrade_unit_cost")
async def update_item_etrade_unit_cost(
    po_number: str, asin: str, etrade_unit_cost: float, db=Depends(get_database)
):
    """Update the eTrade unit cost for a specific item and recompute diff."""
    if etrade_unit_cost < 0:
        raise HTTPException(status_code=400, detail="etrade_unit_cost must be >= 0")

    def _update():
        doc = db[PO_COLLECTION].find_one(
            {"po_number": po_number, "items.asin": asin}, {"items.$": 1}
        )
        if not doc or not doc.get("items"):
            return False, None
        cost = doc["items"][0].get("cost_price_wo_tax")
        new_diff = round(etrade_unit_cost - cost, 2) if cost is not None else None
        db[PO_COLLECTION].update_one(
            {"po_number": po_number, "items.asin": asin},
            {
                "$set": {
                    "items.$.etrade_unit_cost": etrade_unit_cost,
                    "items.$.diff": new_diff,
                    "updated_at": datetime.now(),
                }
            },
        )
        return True, new_diff

    found, new_diff = await asyncio.to_thread(_update)
    if not found:
        raise HTTPException(
            status_code=404, detail=f"PO {po_number} / ASIN {asin} not found"
        )
    return {
        "po_number": po_number,
        "asin": asin,
        "etrade_unit_cost": etrade_unit_cost,
        "diff": new_diff,
    }


def _build_po_excel(doc: dict, enriched: list) -> bytes:
    """Build PO report Excel workbook and return bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PO Report"

    yellow_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    inv_date_label = doc.get("inventory_date") or "Latest"
    po_date_str_raw = doc.get("po_date", "")
    try:
        _po_dt = datetime.strptime(po_date_str_raw, "%Y-%m-%d")
        _drr_start_dt = _po_dt - timedelta(days=30)
        _drr_range = (
            f"{_drr_start_dt.strftime('%-d %b %Y')} – {_po_dt.strftime('%-d %b %Y')}"
        )
    except (ValueError, TypeError):
        _drr_range = po_date_str_raw

    # Month labels for last 5 months (from first enriched item if available)
    _month_labels = (enriched[0].get("month_labels") or []) if enriched else []
    _month_headers = [lbl for (_, _, lbl) in _month_labels] if _month_labels else [f"Month {i+1} Sales" for i in range(5)]

    # Column layout (41 cols + 5 monthly = 46 total):
    # A-H: meta+qty, I=Supply(formula=AM), J=Accepted, K=ShortSupply, L=Received, M=Mismatch
    # N=ZohoMRP, O=eTrade ASP, P=GST, Q=MRPwoGST, R=Margin, S=CostPrice
    # T=TotalCost(SupplyQty), U=TotalCostwoGST(SupplyQty), V=TotalCost(AcceptedQty), W=TotalCostwoGST(AcceptedQty)
    # X=HSN, Y=EtradeUnitCost, Z=Diff, AA=ZohoStock, AB=Status, AC=CurrentStock, AD=OpenPO
    # AE=TotalQty, AF=FinalDRR, AG=NetTotalDays, AH=LeadTime, AI=CoverageDays
    # AJ=TotalTargetDays, AK=TargetStock
    # AL=FinalUnits(For Under-ordering), AM=FinalSupplyQty(For Over-ordering)
    # AN-AR=5 months Amazon sales
    headers = [
        ("PO Date", False), ("PO", False), ("PO Status", False), ("Ship to location", False),
        ("ASIN", True), ("Model Number", True), ("Title", True), ("Requested Qty", True),
        ("Supply Qty", False), ("Accepted Qty", False), ("Short Supply\n(Accepted-Supply)", False),
        ("Received QTY", False), ("Mismatch QTY", False),
        ("Zoho MRP", True), ("eTrade ASP", True), ("GST", True), ("MRP w/o GST", True),
        ("Margin (%)", True), ("Cost Price w/o Tax", True),
        ("Total Cost\n(Supply Qty)", True), ("Total cost w/o GST\n(Supply Qty)", True),
        ("Total Cost\n(Accepted Qty)", True), ("Total cost w/o GST\n(Accepted Qty)", True),
        ("HSN", True), ("Etrade Unit Cost", True), ("Diff", True),
        ("Zoho Stock", True), ("Status", True), (f"Current Stock\n({inv_date_label})", True),
        ("Open PO", True), ("Total Qty", True),
        (f"Final DRR\n({_drr_range})", True),
        ("Net Total Days", True), ("Lead Time", True), ("Coverage Days", True),
        ("Total Target Days", True), ("Target Stock", True),
        ("Final Units\n(For Under-ordering)", True), ("Final Supply Qty\n(For Over-ordering)", True),
    ] + [(lbl, True) for lbl in _month_headers]

    for col_idx, (label, highlight) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        if highlight:
            cell.fill = yellow_fill

    pct_format = "0.00%"
    num_format = "#,##0.00"
    int_format = "#,##0"
    days_format = "0.00"

    po_number = doc["po_number"]
    po_date_str = doc["po_date"]
    for row_idx, item in enumerate(enriched, 2):
        r = row_idx
        accepted = item.get("accepted_qty")
        received = item.get("received_qty")
        monthly_sales = item.get("monthly_sales") or []

        static = {
            1: po_date_str,          # A  PO Date
            2: po_number,            # B  PO
            3: doc["po_status"],     # C  PO Status
            4: item["ship_to_location"],  # D
            5: item["asin"],         # E  ASIN
            6: item["model_number"], # F  Model Number
            7: item["title"],        # G  Title
            8: item["requested_qty"], # H  Requested Qty
            10: accepted if accepted is not None else "",  # J  Accepted Qty
            12: received if received is not None else "",  # L  Received QTY
            14: item["zoho_mrp"],    # N  Zoho MRP
            15: item.get("etrade_asp") if item.get("etrade_asp") is not None else "",  # O  eTrade ASP
            16: item["gst"] / 100,   # P  GST
            18: item["margin"] if item["margin"] is not None else "",  # R  Margin
            24: item["hsn"],         # X  HSN
            25: item["etrade_unit_cost"],  # Y  Etrade Unit Cost
            27: item["zoho_stock"],  # AA Zoho Stock
            28: item["purchase_status"],  # AB Status
            29: item["current_stock"],  # AC Current Stock
            30: item["open_po"],     # AD Open PO
            32: (item.get("final_drr") if item.get("final_drr") is not None else (item.get("final_drr_flag") or "")),  # AF Final DRR
            34: item.get("lead_time", 10),   # AH Lead Time
            35: item.get("coverage_days", COVERAGE_DAYS),  # AI Coverage Days
        }
        # Monthly sales cols 40-44 (AN-AR)
        for mi, units in enumerate(monthly_sales[:5]):
            static[40 + mi] = units

        supply_qty_override = item.get("supply_qty_override")
        final_units_override = item.get("final_units_override")
        final_supply_fo_override = item.get("final_supply_fo_override")
        formulas = {
            9:  f"=AM{r}",                                                              # I   Supply Qty = Final Supply Qty (For Over-ordering)
            11: f'=IF(J{r}="","",J{r}-I{r})',                                          # K   Short Supply
            13: f'=IF(L{r}="","",J{r}-L{r})',                                          # M   Mismatch
            17: f'=IF(O{r}="",ROUND(N{r}/(1+P{r}),2),ROUND(O{r}/(1+P{r}),2))',        # Q   MRP w/o GST
            19: f'=IF(R{r}="","",ROUND(Q{r}*(1-R{r}),2))',                             # S   Cost Price w/o Tax
            20: f'=IF(OR(Q{r}="",R{r}="",I{r}=""),"",ROUND(Q{r}*(1-R{r})*I{r},2))',    # T   Total Cost (Supply Qty)
            21: f'=IF(T{r}="","",ROUND(T{r}*(1+P{r}),2))',                             # U   Total cost w/o GST (Supply Qty)
            22: f'=IF(S{r}="","",ROUND(S{r}*J{r},2))',                                 # V   Total Cost (Accepted Qty)
            23: f'=IF(V{r}="","",ROUND(V{r}*(1+P{r}),2))',                             # W   Total cost w/o GST (Accepted Qty)
            26: f'=IF(S{r}="","",Y{r}-S{r})',                                          # Z   Diff
            31: f"=AC{r}+AD{r}",                                                        # AE  Total Qty
            33: f'=IF(AF{r}=0,"",ROUND(AE{r}/AF{r},2))',                               # AG  Net Total Days
            36: f"=AH{r}+AI{r}",                                                        # AJ  Total Target Days
            37: f"=ROUND(AF{r}*AJ{r},0)",                                               # AK  Target Stock = DRR * Total Target Days
            38: f'=IF(AF{r}=0,"",IF(AG{r}<AH{r},ROUND(AF{r}*AJ{r},0),IF(AG{r}>AJ{r},0,ROUND((AJ{r}-AG{r})*AF{r},0))))',  # AL  Final Units (For Under-ordering)
            39: f'=IF(AL{r}="","",MIN(AL{r},H{r}))',                                   # AM  Final Supply Qty (For Over-ordering)
        }
        if supply_qty_override is not None:
            formulas.pop(9, None)
            static[9] = supply_qty_override
        if final_units_override is not None:
            formulas.pop(38, None)
            static[38] = final_units_override
        if final_supply_fo_override is not None:
            formulas.pop(9, None)
            formulas.pop(39, None)
            static[9] = final_supply_fo_override
            static[39] = final_supply_fo_override

        total_cols = 39 + len(_month_headers)
        for col_idx in range(1, total_cols + 1):
            if col_idx in formulas:
                cell = ws.cell(row=r, column=col_idx, value=formulas[col_idx])
            else:
                cell = ws.cell(row=r, column=col_idx, value=static.get(col_idx, ""))
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            if col_idx in (14, 15, 17, 19, 20, 21, 22, 23, 25, 26):  # currency cols
                cell.number_format = num_format
            elif col_idx in (16, 18):  # GST%, Margin%
                cell.number_format = pct_format
            elif col_idx in (8, 10, 12, 27, 29, 30, 31, 34, 35, 36, 37, 38, 39) or col_idx >= 40:
                cell.number_format = int_format
            elif col_idx == 32:  # Final DRR
                cell.number_format = "0.00"
            elif col_idx == 33:  # Net Total Days
                cell.number_format = days_format

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
    ws.column_dimensions["G"].width = 40
    ws.column_dimensions["AF"].width = 26  # Final DRR with date range
    ws.row_dimensions[1].height = 50

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/{po_number}/download")
async def download_po_report(po_number: str, db=Depends(get_database)):
    """Download enriched PO report as Excel."""

    def _get_doc():
        return db[PO_COLLECTION].find_one({"po_number": po_number})

    doc = await asyncio.to_thread(_get_doc)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"PO {po_number} not found")

    _dl_asins = [it["asin"] for it in doc.get("items", []) if it.get("asin")]
    _dl_drr_map = await _compute_drr_for_po_async(db, doc.get("po_date", ""), _dl_asins)

    def _build():
        po_status = doc.get("po_status", "pending")
        use_stored = po_status in FREEZE_ON_STATUS
        enriched, _, __ = _enrich_items(
            doc.get("items", []),
            po_number,
            db,
            use_stored_stock=use_stored,
            po_date_str=doc.get("po_date"),
            drr_map=_dl_drr_map,
        )
        # Write back fresh per-item computed fields so the list aggregation stays current
        item_updates = {}
        for i, item in enumerate(enriched):
            item_updates[f"items.{i}.final_supply_fo"] = item.get("final_supply_fo")
            item_updates[f"items.{i}.total_cost_fo"] = item.get("total_cost_fo")
            item_updates[f"items.{i}.total_cost_fo_gst"] = item.get("total_cost_fo_gst")
        db[PO_COLLECTION].update_one(
            {"po_number": po_number},
            {"$set": item_updates},
        )
        return enriched

    enriched = await asyncio.to_thread(_build)

    excel_bytes = _build_po_excel(doc, enriched)
    po_date_str = doc.get("po_date", "unknown")
    filename = f"PO_Report_{po_number}_{po_date_str}.xlsx"
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/bulk_download")
async def bulk_download_po_reports(po_numbers: list[str], db=Depends(get_database)):
    """Download multiple PO reports as a single zip file."""
    if not po_numbers:
        raise HTTPException(status_code=400, detail="No PO numbers provided")

    # Fetch all docs first, then compute DRR async in parallel, then enrich
    def _fetch_docs():
        return [
            db[PO_COLLECTION].find_one({"po_number": pn})
            for pn in po_numbers
        ]

    raw_docs = await asyncio.to_thread(_fetch_docs)
    docs = [(pn, doc) for pn, doc in zip(po_numbers, raw_docs) if doc]

    drr_maps = await asyncio.gather(*[
        _compute_drr_for_po_async(
            db,
            doc.get("po_date", ""),
            [it["asin"] for it in doc.get("items", []) if it.get("asin")],
        )
        for _, doc in docs
    ])

    def _build_all():
        results = []
        for (pn, doc), drr_map in zip(docs, drr_maps):
            po_status = doc.get("po_status", "pending")
            use_stored = po_status in FREEZE_ON_STATUS
            enriched, _, __ = _enrich_items(
                doc.get("items", []),
                pn,
                db,
                use_stored_stock=use_stored,
                po_date_str=doc.get("po_date"),
                drr_map=drr_map,
            )
            results.append((pn, doc.get("po_date", "unknown"), doc, enriched))
        return results

    results = await asyncio.to_thread(_build_all)
    if not results:
        raise HTTPException(status_code=404, detail="No matching POs found")

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for pn, po_date_str, doc, enriched in results:
            excel_bytes = _build_po_excel(doc, enriched)
            zf.writestr(f"PO_Report_{pn}_{po_date_str}.xlsx", excel_bytes)
    zip_buf.seek(0)

    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="PO_Reports.zip"'},
    )


# ─── upload order (POItemExport) ──────────────────────────────────────────────

S3_BUCKET = os.getenv("S3_BUCKET", "pupscribe-purchases")
AWS_REGION = os.getenv("AWS_REGION", "ap-south-1")

AVAILABILITY_IN_STOCK = "AC - Accepted: In stock"
AVAILABILITY_OUT_OF_STOCK = "OS - Cancelled: Out of stock"

# Column indices in the POItemExport XLS (0-based)
_COL_PO = 0
_COL_ASIN = 5
_COL_AVAILABILITY = 11
_COL_ACCEPTED_QTY = 13


def _validate_and_fill_order_xls(
    file_bytes: bytes,
    expected_po_number: str,
    accepted_by_asin: dict[str, int],
) -> bytes:
    """Validate PO number in every data row, then fill Accepted quantity & Availability."""
    rb = xlrd.open_workbook(file_contents=file_bytes, formatting_info=True)
    rs = rb.sheet_by_index(0)

    if rs.nrows < 2:
        raise ValueError("File has no data rows")

    mismatched = set()
    for row_idx in range(1, rs.nrows):
        po_cell = rs.cell_value(row_idx, _COL_PO)
        po_in_file = str(po_cell).strip() if po_cell else ""
        if po_in_file and po_in_file != expected_po_number:
            mismatched.add(po_in_file)

    if mismatched:
        raise ValueError(
            f"File contains PO number(s) {sorted(mismatched)} but expected {expected_po_number}. "
            "Please upload the correct POItemExport file for this purchase order."
        )

    wb = xl_copy(rb)
    ws = wb.get_sheet(0)

    for row_idx in range(1, rs.nrows):
        asin_cell = rs.cell_value(row_idx, _COL_ASIN)
        asin = str(asin_cell).strip() if asin_cell else ""
        if not asin or asin not in accepted_by_asin:
            continue

        qty = accepted_by_asin[asin]
        ws.write(row_idx, _COL_ACCEPTED_QTY, qty)
        availability = AVAILABILITY_OUT_OF_STOCK if qty == 0 else AVAILABILITY_IN_STOCK
        ws.write(row_idx, _COL_AVAILABILITY, availability)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload_to_s3(file_bytes: bytes, s3_key: str) -> None:
    """Upload bytes to S3 (private object)."""
    client = boto3.client("s3", region_name=AWS_REGION)
    client.put_object(
        Bucket=S3_BUCKET,
        Key=s3_key,
        Body=file_bytes,
        ContentType="application/vnd.ms-excel",
    )


def _presign_s3(s3_key: str, expires: int = 3600) -> str:
    """Generate a presigned GET URL valid for `expires` seconds."""
    client = boto3.client(
        "s3",
        region_name=AWS_REGION,
        endpoint_url=f"https://s3.{AWS_REGION}.amazonaws.com",
        config=BotocoreConfig(signature_version="s3v4"),
    )
    return client.generate_presigned_url(
        "get_object",
        Params={"Bucket": S3_BUCKET, "Key": s3_key},
        ExpiresIn=expires,
    )


@router.post("/{po_number}/upload_order")
async def upload_order_file(
    po_number: str,
    file: UploadFile = File(...),
    db=Depends(get_database),
):
    """Upload a POItemExport XLS, fill Accepted qty & Availability from DB, store in S3."""
    if not file.filename or not file.filename.lower().endswith(".xls"):
        raise HTTPException(
            status_code=400, detail="File must be a .xls POItemExport file"
        )

    file_bytes = await file.read()

    class _NotFound(Exception):
        pass

    class _BadFile(Exception):
        pass

    def _process_safe():
        doc = db[PO_COLLECTION].find_one({"po_number": po_number})
        if not doc:
            raise _NotFound(f"PO {po_number} not found")
        accepted_by_asin: dict[str, int] = {
            item["asin"]: int(item.get("accepted_qty") or 0)
            for item in doc.get("items", [])
            if item.get("asin")
        }
        try:
            filled_bytes = _validate_and_fill_order_xls(
                file_bytes, po_number, accepted_by_asin
            )
        except ValueError as e:
            raise _BadFile(str(e))
        s3_key = f"vendor_purchase_orders/{po_number}/POItemExport_filled_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xls"
        _upload_to_s3(filled_bytes, s3_key)
        db[PO_COLLECTION].update_one(
            {"po_number": po_number},
            {"$set": {"order_file_s3_key": s3_key, "updated_at": datetime.now()}},
        )
        return filled_bytes

    try:
        filled_bytes = await asyncio.to_thread(_process_safe)
    except _NotFound as e:
        raise HTTPException(status_code=404, detail=str(e))
    except _BadFile as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.exception("upload_order failed")
        raise HTTPException(status_code=500, detail=str(e))

    filename = f"POItemExport_{po_number}_filled.xls"
    return StreamingResponse(
        io.BytesIO(filled_bytes),
        media_type="application/vnd.ms-excel",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{po_number}/order_file")
async def get_order_file_url(po_number: str, db=Depends(get_database)):
    """Return a short-lived presigned URL for the stored order file."""

    def _fetch():
        return db[PO_COLLECTION].find_one(
            {"po_number": po_number}, {"order_file_s3_key": 1}
        )

    doc = await asyncio.to_thread(_fetch)
    if not doc:
        raise HTTPException(status_code=404, detail=f"PO {po_number} not found")
    s3_key = doc.get("order_file_s3_key")
    if not s3_key:
        raise HTTPException(
            status_code=404, detail="No order file uploaded for this PO"
        )

    try:
        url = await asyncio.to_thread(_presign_s3, s3_key)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    return {"url": url}


@router.delete("/{po_number}/order_file")
async def delete_order_file(po_number: str, db=Depends(get_database)):
    """Delete the stored order file from S3 and clear the reference in DB."""

    class _NotFound(Exception):
        pass

    def _delete():
        doc = db[PO_COLLECTION].find_one(
            {"po_number": po_number}, {"order_file_s3_key": 1}
        )
        if not doc:
            raise _NotFound(f"PO {po_number} not found")
        s3_key = doc.get("order_file_s3_key")
        if not s3_key:
            raise _NotFound("No order file uploaded for this PO")
        boto3.client("s3", region_name=AWS_REGION).delete_object(
            Bucket=S3_BUCKET, Key=s3_key
        )
        db[PO_COLLECTION].update_one(
            {"po_number": po_number},
            {
                "$unset": {"order_file_s3_key": ""},
                "$set": {"updated_at": datetime.now()},
            },
        )

    try:
        await asyncio.to_thread(_delete)
    except _NotFound as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.exception("delete_order_file failed")
        raise HTTPException(status_code=500, detail=str(e))

    return {"po_number": po_number, "deleted": True}


# ─── shipment summary ─────────────────────────────────────────────────────────


class ShipmentUpdate(BaseModel):
    reason_for_short_supply: Optional[str] = None
    box_count: Optional[int] = None
    appointment_initiated_date: Optional[str] = None
    appointment_id: Optional[str] = None
    appointment_date: Optional[str] = None
    dispatched_date: Optional[str] = None
    delivery_date: Optional[str] = None


@router.get("/shipment_summary")
async def get_shipment_summary(db=Depends(get_database)):
    """Return one row per PO with shipment summary columns for the Etrade Shipment Summary page."""

    def _fetch():
        pipeline = [
            {
                "$addFields": {
                    "total_requested_qty": {"$sum": "$items.requested_qty"},
                    "total_supply_qty": {"$sum": "$items.final_supply_fo"},
                    "total_accepted_qty": {"$sum": "$items.accepted_qty"},
                    "location": {"$arrayElemAt": ["$items.ship_to_location", 0]},
                }
            },
            {
                "$project": {
                    "po_number": 1,
                    "po_date": 1,
                    "po_status": 1,
                    "location": 1,
                    "total_requested_qty": 1,
                    "total_supply_qty": 1,
                    "total_accepted_qty": 1,
                    "reason_for_short_supply": 1,
                    "box_count": 1,
                    "appointment_initiated_date": 1,
                    "appointment_id": 1,
                    "appointment_date": 1,
                    "dispatched_date": 1,
                    "delivery_date": 1,
                    "_id": 0,
                }
            },
            {"$sort": {"po_date": -1}},
        ]
        return list(db[PO_COLLECTION].aggregate(pipeline))

    rows = await asyncio.to_thread(_fetch)
    return serialize_mongo_document(rows)


@router.patch("/{po_number}/shipment")
async def update_shipment_fields(
    po_number: str, body: ShipmentUpdate, db=Depends(get_database)
):
    """Update editable shipment tracking fields on a PO."""
    update: dict = {"updated_at": datetime.now()}
    for field, value in body.model_dump(exclude_unset=True).items():
        update[field] = value

    if len(update) == 1:
        raise HTTPException(status_code=400, detail="No fields to update")

    def _update():
        return (
            db[PO_COLLECTION]
            .update_one({"po_number": po_number}, {"$set": update})
            .matched_count
        )

    matched = await asyncio.to_thread(_update)
    if not matched:
        raise HTTPException(status_code=404, detail=f"PO {po_number} not found")
    return {"po_number": po_number, "updated": True}


# ─── margins ──────────────────────────────────────────────────────────────────


@router.get("/margins")
async def get_margins(db=Depends(get_database)):
    def _fetch():
        return list(db[MARGINS_COLLECTION].find({}, {"_id": 0}))

    margins = await asyncio.to_thread(_fetch)
    return serialize_mongo_document(margins)


@router.put("/margins/{asin}")
async def upsert_margin(
    asin: str,
    margin: Optional[float] = None,
    cost_price_wo_tax: Optional[float] = None,
    etrade_asp: Optional[float] = None,
    etrade_po: Optional[bool] = None,
    etrade_df: Optional[bool] = None,
    db=Depends(get_database),
):
    if all(v is None for v in [margin, cost_price_wo_tax, etrade_asp, etrade_po, etrade_df]):
        raise HTTPException(
            status_code=400,
            detail="At least one field must be provided",
        )
    if margin is not None and not (0 <= margin <= 1):
        raise HTTPException(
            status_code=400, detail="margin must be between 0 and 1 (e.g. 0.35 for 35%)"
        )
    if cost_price_wo_tax is not None and cost_price_wo_tax < 0:
        raise HTTPException(status_code=400, detail="cost_price_wo_tax must be >= 0")
    if etrade_asp is not None and etrade_asp < 0:
        raise HTTPException(status_code=400, detail="etrade_asp must be >= 0")

    fields: dict = {"asin": asin, "updated_at": datetime.now()}
    if margin is not None:
        fields["margin"] = margin
    if cost_price_wo_tax is not None:
        fields["cost_price_wo_tax"] = cost_price_wo_tax
    if etrade_asp is not None:
        fields["etrade_asp"] = etrade_asp
    if etrade_po is not None:
        fields["etrade_po"] = etrade_po
    if etrade_df is not None:
        fields["etrade_df"] = etrade_df

    def _upsert():
        db[MARGINS_COLLECTION].update_one({"asin": asin}, {"$set": fields}, upsert=True)

    await asyncio.to_thread(_upsert)
    return {"asin": asin, **{k: v for k, v in fields.items() if k not in ("asin", "updated_at")}}


@router.post("/bulk_update")
async def bulk_update_vendor_pos(
    file: UploadFile = File(...),
    db=Depends(get_database),
):
    """Bulk-update accepted_qty, received_qty, and po_status for existing POs.

    Expected Excel columns (header row required):
      A: PO Number | B: ASIN | C: Accepted Qty | D: Received Qty | E: PO Status

    Only POs with status pending/processing/packed/closed will be updated.
    PO Status in column E is optional per row; the first non-empty value wins for each PO.
    """
    file_bytes = await file.read()

    def _process():
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise ValueError("File has no data rows")

        # Group rows by PO number
        updates: dict = {}
        for row in rows[1:]:
            if not row[0]:
                continue
            po_num = str(row[0]).strip()
            asin = str(row[1]).strip() if row[1] else ""
            accepted_qty = int(row[2]) if row[2] is not None else None
            received_qty = int(row[3]) if row[3] is not None else None
            po_status_raw = (
                str(row[4]).strip().lower() if len(row) > 4 and row[4] else None
            )

            if po_num not in updates:
                updates[po_num] = {"po_status": po_status_raw, "items": {}}
            elif po_status_raw and updates[po_num]["po_status"] is None:
                updates[po_num]["po_status"] = po_status_raw

            if asin:
                updates[po_num]["items"][asin] = {
                    "accepted_qty": accepted_qty,
                    "received_qty": received_qty,
                }

        updatable_statuses = {"pending", "processing", "packed", "closed"}
        results = []

        for po_num, update_data in updates.items():
            doc = db[PO_COLLECTION].find_one({"po_number": po_num})
            if not doc:
                results.append({"po_number": po_num, "status": "not_found"})
                continue

            current_status = doc.get("po_status", "pending")
            if current_status not in updatable_statuses:
                results.append(
                    {
                        "po_number": po_num,
                        "status": "skipped",
                        "reason": f"status '{current_status}' is not updatable",
                    }
                )
                continue

            update_fields: dict = {"updated_at": datetime.now()}

            new_status = update_data.get("po_status")
            if new_status and new_status in VALID_STATUSES:
                update_fields["po_status"] = new_status

            items = doc.get("items", [])
            item_changes = 0
            for item in items:
                asin = item["asin"]
                if asin in update_data["items"]:
                    item_update = update_data["items"][asin]
                    if item_update["accepted_qty"] is not None:
                        item["accepted_qty"] = item_update["accepted_qty"]
                        item_changes += 1
                    if item_update["received_qty"] is not None:
                        item["received_qty"] = item_update["received_qty"]
                        item_changes += 1

            if item_changes:
                update_fields["items"] = items

            if len(update_fields) > 1:  # more than just updated_at
                db[PO_COLLECTION].update_one(
                    {"po_number": po_num}, {"$set": update_fields}
                )
                results.append(
                    {
                        "po_number": po_num,
                        "status": "updated",
                        "items_changed": item_changes,
                    }
                )
            else:
                results.append({"po_number": po_num, "status": "no_changes"})

        return results

    try:
        results = await asyncio.to_thread(_process)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return JSONResponse(status_code=200, content={"results": results})


@router.get("/margins/bulk")
async def get_margins_for_asins(asins: str, db=Depends(get_database)):
    asin_list = [a.strip() for a in asins.split(",") if a.strip()]

    def _fetch():
        return list(
            db[MARGINS_COLLECTION].find({"asin": {"$in": asin_list}}, {"_id": 0})
        )

    margins = await asyncio.to_thread(_fetch)
    return {m["asin"]: m["margin"] for m in margins}


# ─── estimate ─────────────────────────────────────────────────────────────────


class EstimateCreateRequest(BaseModel):
    billing_address_id: str
    shipping_address_id: str
    date: Optional[str] = None  # YYYY-MM-DD; defaults to today
    skip_inactive: bool = False  # If True, skip truly-inactive items instead of blocking


class EstimateLinkRequest(BaseModel):
    estimate_number: str


@router.get("/etrade_customer")
async def get_etrade_customer(db=Depends(get_database)):
    """Return ETRADE MARKETING contact_id and deduplicated address list."""

    def _fetch():
        customer = db[CUSTOMERS_COLLECTION].find_one(
            {"contact_name": {"$regex": "ETRADE", "$options": "i"}},
            {"contact_id": 1, "addresses": 1},
        )
        if not customer:
            raise ValueError("ETRADE customer not found in customers collection")

        raw_addresses = _parse_addresses(customer.get("addresses", []))
        seen: set[str] = set()
        unique: list[dict] = []
        for addr in raw_addresses:
            aid = addr.get("address_id", "")
            address_text = addr.get("address", "")
            if not aid or not address_text:
                continue
            if aid not in seen:
                seen.add(aid)
                unique.append(
                    {
                        "address_id": aid,
                        "attention": addr.get("attention", ""),
                        "address": address_text,
                        "street2": addr.get("street2", ""),
                        "city": addr.get("city", ""),
                        "state": addr.get("state", ""),
                        "zip": addr.get("zip", ""),
                        "country": addr.get("country", "India"),
                    }
                )

        return {"contact_id": customer["contact_id"], "addresses": unique}

    try:
        result = await asyncio.to_thread(_fetch)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    return result


@router.post("/{po_number}/estimate")
async def create_zoho_estimate(
    po_number: str, body: EstimateCreateRequest, db=Depends(get_database)
):
    """Create a Zoho Books estimate for all active PO items."""

    def _create():
        doc = db[PO_COLLECTION].find_one({"po_number": po_number})
        if not doc:
            raise ValueError(f"PO {po_number} not found")
        if doc.get("zoho_estimate_id"):
            raise ValueError(
                f"PO already has estimate {doc.get('estimate_number')} — unlink it first"
            )

        # Exclude inactive items
        items = [
            it
            for it in doc.get("items", [])
            if (it.get("status") or "").lower() != "inactive"
        ]
        if not items:
            raise ValueError("PO has no active items")

        # --- Step 1: ASIN → sku_code via amazon_sku_mapping ---
        model_numbers = [it["model_number"] for it in items if it.get("model_number")]
        asin_to_sku_code: dict[str, str] = {
            m["item_id"]: m["sku_code"]
            for m in db[SKU_MAPPING_COLLECTION].find(
                {"item_id": {"$in": model_numbers}},
                {"item_id": 1, "sku_code": 1},
            )
            if m.get("item_id") and m.get("sku_code")
        }

        # For each item, lookup_key = sku_code resolved from ASIN, or model_number used directly
        all_lookup_keys = list({
            asin_to_sku_code.get(mn, mn)
            for mn in model_numbers
        })

        # --- Step 2: composite_products first (always active in Zoho Books) ---
        sku_to_composite_id: dict[str, str] = {}
        sku_to_composite_name: dict[str, str] = {}
        for c in db["composite_products"].find(
            {"sku_code": {"$in": all_lookup_keys}},
            {"sku_code": 1, "composite_item_id": 1, "name": 1},
        ):
            if c.get("sku_code") and c.get("composite_item_id"):
                sku_to_composite_id[c["sku_code"]] = c["composite_item_id"]
                sku_to_composite_name[c["sku_code"]] = c.get("name", c["sku_code"])

        # --- Step 3: products fallback for items not found in composite_products ---
        # Prefer active products (dual 5%/12% GST variants — 12% are inactive)
        non_composite_keys = [k for k in all_lookup_keys if k not in sku_to_composite_id]
        sku_to_product_id: dict[str, str] = {}
        sku_to_product_status: dict[str, str] = {}
        sku_to_product_name: dict[str, str] = {}
        for p in db[PRODUCTS_COLLECTION].find(
            {"cf_sku_code": {"$in": non_composite_keys}},
            {"cf_sku_code": 1, "item_id": 1, "status": 1, "name": 1},
        ):
            sku = p.get("cf_sku_code")
            item_id = p.get("item_id")
            if not sku or not item_id:
                continue
            status = p.get("status", "")
            # Active entry wins; only write inactive if nothing stored yet
            if status == "active" or sku not in sku_to_product_id:
                sku_to_product_id[sku] = item_id
                sku_to_product_status[sku] = status
                sku_to_product_name[sku] = p.get("name", sku)

        inactive_items: list[str] = []
        skipped: list[str] = []
        line_items: list[dict] = []
        for it in items:
            mn = it.get("model_number", "")
            lookup_key = asin_to_sku_code.get(mn, mn)  # sku_code if ASIN, else model_number direct

            # Composite wins — always active in Zoho Books
            zoho_item_id = sku_to_composite_id.get(lookup_key)
            if not zoho_item_id:
                # Fall back to regular product
                zoho_item_id = sku_to_product_id.get(lookup_key)
                if not zoho_item_id:
                    skipped.append(mn or it.get("asin", "?"))
                    continue
                # Check if product is inactive/deleted in Zoho
                item_status = sku_to_product_status.get(lookup_key, "active")
                if item_status and item_status != "active":
                    product_name = sku_to_product_name.get(lookup_key, lookup_key)
                    inactive_items.append(f"{mn} — {product_name} (status: {item_status})")
                    continue  # blocked below unless skip_inactive=True
            margin = it.get("margin")
            discount = round(margin * 100, 2) if margin is not None else 0
            qty = it.get("final_supply_fo")
            qty = qty if qty is not None else 0
            line_items.append(
                {
                    "item_id": zoho_item_id,
                    "quantity": qty,
                    "rate": round(it.get("mrp_wo_gst") or 0, 2),
                    "discount": f"{discount}%",
                    "unit": "pcs",
                    "hsn_or_sac": it.get("hsn", ""),
                }
            )

        if inactive_items and not body.skip_inactive:
            raise ValueError(
                "INACTIVE_ITEMS:" + "\n".join(inactive_items)
            )

        if not line_items:
            raise ValueError(
                f"No line items could be built — missing Zoho item_id for all items: {', '.join(skipped)}"
            )

        # ETRADE customer
        customer = db[CUSTOMERS_COLLECTION].find_one(
            {"contact_name": {"$regex": "ETRADE", "$options": "i"}},
            {"contact_id": 1},
        )
        if not customer:
            raise ValueError("ETRADE customer not found")

        token = _get_zoho_token()
        headers = {"Authorization": f"Zoho-oauthtoken {token}"}

        # Compute next estimate number using the same FY counter pattern
        now = datetime.now()
        fy_start = now.year if now.month >= 4 else now.year - 1
        fy_str = f"{str(fy_start)[-2:]}-{str(fy_start + 1)[-2:]}"

        list_r = requests.get(
            f"{ZOHO_BOOKS_BASE}/estimates",
            headers=headers,
            params={
                "organization_id": ORGANIZATION_ID,
                "filter_by": "Status.All",
                "per_page": 200,
                "sort_column": "estimate_number",
                "sort_order": "D",
            },
            timeout=30,
        )
        list_r.raise_for_status()
        all_estimates = list_r.json().get("estimates", [])
        fy_estimates = [
            e for e in all_estimates if f"/{fy_str}/" in e.get("estimate_number", "")
        ]
        if fy_estimates:
            parts = str(fy_estimates[0]["estimate_number"]).split("/")
            last_num = int(parts[-1])
            num_width = len(parts[-1])
            prefix = parts[0]
        else:
            parts = (
                str(all_estimates[0]["estimate_number"]).split("/")
                if all_estimates
                else ["EST"]
            )
            last_num = 0
            num_width = len(parts[-1]) if len(parts) > 1 else 4
            prefix = parts[0]

        counter_id = f"estimate_counter_{fy_str}"
        db.counters.update_one(
            {"_id": counter_id}, {"$max": {"seq": last_num}}, upsert=True
        )
        counter = db.counters.find_one_and_update(
            {"_id": counter_id}, {"$inc": {"seq": 1}}, return_document=True
        )
        new_estimate_number = (
            f"{prefix}/{fy_str}/{str(counter['seq']).zfill(num_width)}"
        )

        estimate_date = body.date or now.strftime("%Y-%m-%d")
        payload: dict = {
            "estimate_number": new_estimate_number,
            "customer_id": customer["contact_id"],
            "reference_number": po_number,
            "date": estimate_date,
            "place_of_supply": "MH",
            "billing_address_id": body.billing_address_id,
            "shipping_address_id": body.shipping_address_id,
            "is_inclusive_tax": False,
            "line_items": line_items,
            "branch_id":"3220178000156681877",
            "branch_name":"Amazon (Mumbai Branch)",
            "salesperson_id": "3220178000000692003",
            "dispatch_from_address": {
                "zip": "401208",
                "country": "India",
                "address": "Gala No. 5 & 7, 1st Floor, Survey No 68, Building No 3",
                "city": "Palghar",
                "address_id": "3220178000541133001",
                "attention": "Mr Akshay Mayekar",
                "street2": "Near Meenakshi Inds. Estate, Naik Pada, Waliv Naka, Vasai E",
                "state": "Maharashtra",
                "state_code": "MH",
            },
        }

        logger.info("Zoho estimate payload: %s", payload)
        r = requests.post(
            f"{ZOHO_BOOKS_BASE}/estimates",
            headers=headers,
            json=payload,
            params={
                "organization_id": ORGANIZATION_ID,
                "ignore_auto_number_generation": "true",
            },
            timeout=30,
        )
        logger.info("Zoho estimate response %s: %s", r.status_code, r.text)
        r.raise_for_status()
        data = r.json()

        if data.get("code") != 0:
            raise ValueError(f"Zoho error: {data.get('message', 'Unknown error')}")

        estimate = data["estimate"]
        db[PO_COLLECTION].update_one(
            {"po_number": po_number},
            {
                "$set": {
                    "zoho_estimate_id": estimate["estimate_id"],
                    "estimate_number": estimate["estimate_number"],
                    "updated_at": datetime.now(),
                }
            },
        )

        return estimate, skipped, inactive_items

    try:
        estimate, skipped, inactive_items = await asyncio.to_thread(_create)
    except ValueError as e:
        msg = str(e)
        if msg.startswith("INACTIVE_ITEMS:"):
            items_list = [line for line in msg[len("INACTIVE_ITEMS:"):].splitlines() if line]
            raise HTTPException(
                status_code=400,
                detail={"type": "inactive_items", "items": items_list},
            )
        raise HTTPException(status_code=400, detail=msg)
    except requests.HTTPError as e:
        body_text = e.response.text if e.response is not None else ""
        raise HTTPException(
            status_code=502, detail=f"Zoho API error: {e} — {body_text}"
        )

    result = {
        "estimate_id": estimate["estimate_id"],
        "estimate_number": estimate["estimate_number"],
        "total": estimate.get("total"),
        "status": estimate.get("status"),
    }
    if skipped:
        result["skipped_items"] = skipped
    if inactive_items:
        result["skipped_inactive"] = inactive_items
    return JSONResponse(status_code=201, content=result)


@router.patch("/{po_number}/estimate")
async def link_zoho_estimate(
    po_number: str, body: EstimateLinkRequest, db=Depends(get_database)
):
    """Link an existing Zoho estimate to a PO by estimate_number (looked up in estimates collection)."""

    def _link():
        if not db[PO_COLLECTION].find_one({"po_number": po_number}):
            raise ValueError(f"PO {po_number} not found")

        est = db[ESTIMATES_COLLECTION].find_one(
            {"estimate_number": body.estimate_number},
            {"estimate_id": 1, "estimate_number": 1},
        )
        if not est:
            raise ValueError(
                f"Estimate {body.estimate_number!r} not found in estimates collection"
            )

        db[PO_COLLECTION].update_one(
            {"po_number": po_number},
            {
                "$set": {
                    "zoho_estimate_id": est.get("estimate_id"),
                    "estimate_number": est["estimate_number"],
                    "updated_at": datetime.now(),
                }
            },
        )
        return est

    try:
        est = await asyncio.to_thread(_link)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return {
        "po_number": po_number,
        "estimate_number": est["estimate_number"],
        "estimate_id": est.get("estimate_id"),
    }


@router.delete("/{po_number}/estimate")
async def unlink_zoho_estimate(po_number: str, db=Depends(get_database)):
    """Remove the estimate link from a PO."""

    def _unlink():
        result = db[PO_COLLECTION].update_one(
            {"po_number": po_number},
            {
                "$unset": {"zoho_estimate_id": "", "estimate_number": ""},
                "$set": {"updated_at": datetime.now()},
            },
        )
        return result.matched_count

    count = await asyncio.to_thread(_unlink)
    if not count:
        raise HTTPException(status_code=404, detail=f"PO {po_number} not found")
    return {"po_number": po_number, "unlinked": True}


class PackageLinkRequest(BaseModel):
    package_number: str


@router.patch("/{po_number}/package")
async def link_package(
    po_number: str, body: PackageLinkRequest, db=Depends(get_database)
):
    """Add a Zoho package to the PO's packages array. Recomputes accepted costs from all packages."""

    def _link():
        doc = db[PO_COLLECTION].find_one({"po_number": po_number})
        if not doc:
            raise ValueError(f"PO {po_number} not found")
        # Validate package exists
        if not db[PACKAGES_COLLECTION].find_one({"package_number": body.package_number}):
            raise ValueError(f"Package {body.package_number!r} not found in packages collection")
        # Add to array (addToSet avoids duplicates)
        db[PO_COLLECTION].update_one(
            {"po_number": po_number},
            {"$addToSet": {"packages": body.package_number}, "$set": {"updated_at": datetime.now()}},
        )
        # Recompute accepted costs across all packages
        updated_doc = db[PO_COLLECTION].find_one({"po_number": po_number}, {"packages": 1})
        all_packages = updated_doc.get("packages") or []
        accepted_total_cost, accepted_total_cost_gst = _compute_packages_accepted_costs(all_packages, db)
        cost_update: dict = {"updated_at": datetime.now()}
        if accepted_total_cost is not None:
            cost_update["accepted_total_cost"] = accepted_total_cost
        if accepted_total_cost_gst is not None:
            cost_update["accepted_total_cost_gst"] = accepted_total_cost_gst
        db[PO_COLLECTION].update_one({"po_number": po_number}, {"$set": cost_update})
        return all_packages, accepted_total_cost, accepted_total_cost_gst

    try:
        packages, accepted_total_cost, accepted_total_cost_gst = await asyncio.to_thread(_link)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return {
        "po_number": po_number,
        "packages": packages,
        "accepted_total_cost": accepted_total_cost,
        "accepted_total_cost_gst": accepted_total_cost_gst,
    }


@router.delete("/{po_number}/package/{pkg_number}")
async def unlink_specific_package(po_number: str, pkg_number: str, db=Depends(get_database)):
    """Remove a specific package from the PO's packages array and recompute accepted costs."""

    def _unlink():
        result = db[PO_COLLECTION].update_one(
            {"po_number": po_number},
            {"$pull": {"packages": pkg_number}, "$set": {"updated_at": datetime.now()}},
        )
        if not result.matched_count:
            return None, None, None
        updated_doc = db[PO_COLLECTION].find_one({"po_number": po_number}, {"packages": 1})
        remaining = updated_doc.get("packages") or []
        if remaining:
            accepted_total_cost, accepted_total_cost_gst = _compute_packages_accepted_costs(remaining, db)
            db[PO_COLLECTION].update_one(
                {"po_number": po_number},
                {"$set": {"accepted_total_cost": accepted_total_cost, "accepted_total_cost_gst": accepted_total_cost_gst}},
            )
        else:
            accepted_total_cost = accepted_total_cost_gst = None
            db[PO_COLLECTION].update_one(
                {"po_number": po_number},
                {"$unset": {"accepted_total_cost": "", "accepted_total_cost_gst": ""}},
            )
        return remaining, accepted_total_cost, accepted_total_cost_gst

    packages, accepted_total_cost, accepted_total_cost_gst = await asyncio.to_thread(_unlink)
    if packages is None:
        raise HTTPException(status_code=404, detail=f"PO {po_number} not found")
    return {"po_number": po_number, "packages": packages, "accepted_total_cost": accepted_total_cost, "accepted_total_cost_gst": accepted_total_cost_gst}


@router.delete("/{po_number}/package")
async def unlink_all_packages(po_number: str, db=Depends(get_database)):
    """Remove all packages and accepted cost fields from a PO."""

    def _unlink():
        result = db[PO_COLLECTION].update_one(
            {"po_number": po_number},
            {
                "$set": {"packages": [], "updated_at": datetime.now()},
                "$unset": {"package_number": "", "accepted_total_cost": "", "accepted_total_cost_gst": ""},
            },
        )
        return result.matched_count

    count = await asyncio.to_thread(_unlink)
    if not count:
        raise HTTPException(status_code=404, detail=f"PO {po_number} not found")
    return {"po_number": po_number, "packages": [], "unlinked": True}


# ─── sales order number ────────────────────────────────────────────────────────


class SalesOrderLinkRequest(BaseModel):
    sales_order_no: str


@router.get("/search/sales_orders")
async def search_sales_orders(q: str = "", db=Depends(get_database)):
    """Search sales_orders collection by salesorder_number (prefix match)."""

    def _search():
        if not q.strip():
            return []
        results = db[SALES_ORDERS_COLLECTION].find(
            {"salesorder_number": {"$regex": q.strip(), "$options": "i"}},
            {"salesorder_number": 1, "salesorder_id": 1, "customer_name": 1, "_id": 0},
        ).limit(10)
        return list(results)

    return await asyncio.to_thread(_search)


@router.patch("/{po_number}/sales_order_no")
async def update_sales_order_no(
    po_number: str, body: SalesOrderLinkRequest, db=Depends(get_database)
):
    """Set or update the sales order number on a PO. Looks up salesorder_id if SO exists in DB."""

    def _update():
        so_number = body.sales_order_no.strip()
        so_doc = db[SALES_ORDERS_COLLECTION].find_one(
            {"salesorder_number": so_number},
            {"salesorder_id": 1, "_id": 0},
        )
        update_fields: dict = {"sales_order_no": so_number, "updated_at": datetime.now()}
        if so_doc:
            update_fields["sales_order_id"] = so_doc["salesorder_id"]
        result = db[PO_COLLECTION].update_one(
            {"po_number": po_number},
            {"$set": update_fields},
        )
        return result.matched_count, so_number, so_doc.get("salesorder_id") if so_doc else None

    count, so_number, so_id = await asyncio.to_thread(_update)
    if not count:
        raise HTTPException(status_code=404, detail=f"PO {po_number} not found")
    return {"po_number": po_number, "sales_order_no": so_number, "sales_order_id": so_id}


# ─── transfer order ────────────────────────────────────────────────────────────


class TransferOrderCreateRequest(BaseModel):
    date: Optional[str] = None  # YYYY-MM-DD; defaults to today


class TransferOrderLinkRequest(BaseModel):
    transfer_order_number: str


@router.post("/{po_number}/transfer_order")
async def create_transfer_order(
    po_number: str, body: TransferOrderCreateRequest, db=Depends(get_database)
):
    """Create a Zoho Inventory transfer order from the linked package's line items."""

    def _create():
        doc = db[PO_COLLECTION].find_one({"po_number": po_number})
        if not doc:
            raise ValueError(f"PO {po_number} not found")
        if doc.get("transfer_order_id"):
            raise ValueError(
                f"PO already has transfer order {doc.get('transfer_order_number')} — unlink it first"
            )

        pkg_number = doc.get("package_number")
        if not pkg_number:
            raise ValueError("No package linked to this PO — link a package first")

        pkg = db[PACKAGES_COLLECTION].find_one({"package_number": pkg_number})
        if not pkg:
            raise ValueError(f"Package {pkg_number!r} not found")

        line_items_raw = pkg.get("line_items") or []
        if not line_items_raw:
            raise ValueError("Package has no line items")

        line_items: list[dict] = []
        for li in line_items_raw:
            item_id = li.get("item_id")
            if not item_id:
                continue
            qty = float(li.get("quantity") or 0)
            if qty <= 0:
                continue
            line_items.append({
                "item_id": item_id,
                "name": li.get("name") or "",
                "quantity_transfer": qty,
                "unit": li.get("unit", "pcs"),
            })

        if not line_items:
            raise ValueError("No valid line items found in package")

        # Compute next transfer order number (numeric sort, not lexicographic)
        agg = list(db[TRANSFER_ORDERS_COLLECTION].aggregate([
            {"$match": {"transfer_order_number": {"$regex": r"^TO-\d+"}}},
            {"$addFields": {"_to_num": {"$toInt": {"$arrayElemAt": [{"$split": ["$transfer_order_number", "-"]}, 1]}}}},
            {"$sort": {"_to_num": -1}},
            {"$limit": 1},
            {"$project": {"_to_num": 1}},
        ]))
        last_num = agg[0]["_to_num"] if agg else 0
        new_to_number = f"TO-{last_num + 1}"

        token = _get_inventory_token()
        headers = {"Authorization": f"Zoho-oauthtoken {token}"}

        to_date = body.date or datetime.now().strftime("%Y-%m-%d")
        payload: dict = {
            "transfer_order_number": new_to_number,
            "date": to_date,
            "from_warehouse_id": TO_FROM_WAREHOUSE_ID,
            "to_warehouse_id": TO_TO_WAREHOUSE_ID,
            "line_items": line_items,
        }

        r = requests.post(
            f"{ZOHO_INVENTORY_BASE}/transferorders",
            headers=headers,
            json=payload,
            params={
                "organization_id": ORGANIZATION_ID,
                "ignore_auto_number_generation": "true",
            },
            timeout=30,
        )
        data = r.json()

        if not r.ok or data.get("code") != 0:
            msg = data.get("message", "Unknown error")
            bad_ids = data.get("error_info") or []
            if bad_ids:
                # Map error item_ids back to names from the line items we built
                id_to_name = {li["item_id"]: li.get("name") or li["item_id"] for li in line_items}
                bad_names = [id_to_name.get(iid, iid) for iid in bad_ids]
                msg = f"{msg} — items: {', '.join(bad_names)}"
            raise ValueError(f"Zoho error {data.get('code')}: {msg}")

        to = data["transfer_order"]
        db[PO_COLLECTION].update_one(
            {"po_number": po_number},
            {
                "$set": {
                    "transfer_order_id": to["transfer_order_id"],
                    "transfer_order_number": to["transfer_order_number"],
                    "updated_at": datetime.now(),
                }
            },
        )
        return to

    try:
        to = await asyncio.to_thread(_create)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except requests.HTTPError as e:
        body_text = e.response.text if e.response is not None else ""
        raise HTTPException(status_code=502, detail=f"Zoho API error: {e} — {body_text}")

    return JSONResponse(
        status_code=201,
        content={
            "transfer_order_id": to["transfer_order_id"],
            "transfer_order_number": to["transfer_order_number"],
            "status": to.get("status"),
        },
    )


@router.patch("/{po_number}/transfer_order")
async def link_transfer_order(
    po_number: str, body: TransferOrderLinkRequest, db=Depends(get_database)
):
    """Link an existing Zoho transfer order to a PO by transfer_order_number."""

    def _link():
        if not db[PO_COLLECTION].find_one({"po_number": po_number}):
            raise ValueError(f"PO {po_number} not found")

        to = db[TRANSFER_ORDERS_COLLECTION].find_one(
            {"transfer_order_number": body.transfer_order_number},
            {"transfer_order_id": 1, "transfer_order_number": 1},
        )
        if not to:
            raise ValueError(
                f"Transfer order {body.transfer_order_number!r} not found in transfer_orders collection"
            )

        db[PO_COLLECTION].update_one(
            {"po_number": po_number},
            {
                "$set": {
                    "transfer_order_id": to.get("transfer_order_id"),
                    "transfer_order_number": to["transfer_order_number"],
                    "updated_at": datetime.now(),
                }
            },
        )
        return to

    try:
        to = await asyncio.to_thread(_link)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return {
        "po_number": po_number,
        "transfer_order_number": to["transfer_order_number"],
        "transfer_order_id": to.get("transfer_order_id"),
    }


@router.delete("/{po_number}/transfer_order")
async def unlink_transfer_order(po_number: str, db=Depends(get_database)):
    """Remove the transfer order link from a PO."""

    def _unlink():
        result = db[PO_COLLECTION].update_one(
            {"po_number": po_number},
            {
                "$unset": {"transfer_order_id": "", "transfer_order_number": ""},
                "$set": {"updated_at": datetime.now()},
            },
        )
        return result.matched_count

    count = await asyncio.to_thread(_unlink)
    if not count:
        raise HTTPException(status_code=404, detail=f"PO {po_number} not found")
    return {"po_number": po_number, "unlinked": True}


# ─── assemblies ────────────────────────────────────────────────────────────────


class AssemblyLinkRequest(BaseModel):
    assembly_number: str  # reference_number in assemblies collection


@router.post("/{po_number}/assemblies")
async def create_assemblies(po_number: str, db=Depends(get_database)):
    """Create Zoho Inventory assemblies for all composite items in the linked package."""

    def _create():
        doc = db[PO_COLLECTION].find_one({"po_number": po_number})
        if not doc:
            raise ValueError(f"PO {po_number} not found")
        if not doc.get("transfer_order_id"):
            raise ValueError("No transfer order linked — link a transfer order first")
        if doc.get("bundle_ids"):
            raise ValueError("Assemblies already created — unlink them first")
        pkg_number = doc.get("package_number")
        if not pkg_number:
            raise ValueError("No package linked — link a package first")

        pkg = db[PACKAGES_COLLECTION].find_one({"package_number": pkg_number})
        if not pkg:
            raise ValueError(f"Package {pkg_number} not found")

        line_items = pkg.get("line_items") or []
        item_ids = [li.get("item_id") for li in line_items if li.get("item_id")]
        if not item_ids:
            raise ValueError("Package has no line items with item_id")

        # Find which line items are composite products (fetch components too)
        composite_by_item_id = {
            c["composite_item_id"]: c
            for c in db["composite_products"].find(
                {"composite_item_id": {"$in": item_ids}},
                {"composite_item_id": 1, "name": 1, "sku_code": 1, "components": 1},
            )
            if c.get("composite_item_id")
        }
        if not composite_by_item_id:
            raise ValueError("No composite products found in this package's line items")

        token = _get_inventory_token()
        headers = {"Authorization": f"Zoho-oauthtoken {token}"}
        today = datetime.now().strftime("%Y-%m-%d")

        created_bundle_ids = []
        created_assembly_numbers = []
        for li in line_items:
            iid = li.get("item_id")
            if iid not in composite_by_item_id:
                continue
            qty = float(li.get("quantity") or 0)
            if qty <= 0:
                continue
            comp = composite_by_item_id[iid]

            # Build component line_items for the bundle
            raw_components = comp.get("components") or []
            if isinstance(raw_components, str):
                try:
                    raw_components = _parse_addresses(raw_components)  # reuse ast.literal_eval helper
                except Exception:
                    raw_components = []
            bundle_line_items = []
            for c in raw_components:
                comp_qty = float(c.get("quantity") or 0)
                if not c.get("item_id") or comp_qty <= 0:
                    continue
                bundle_line_items.append({
                    "item_id": c["item_id"],
                    "name": c.get("name", ""),
                    "quantity_consumed": comp_qty * qty,
                    "warehouse_id": TO_FROM_WAREHOUSE_ID,
                })

            payload = {
                "composite_item_id": iid,
                "composite_item_name": comp.get("name", ""),
                "date": today,
                "quantity_to_bundle": qty,
                "warehouse_id": TO_FROM_WAREHOUSE_ID,
                "line_items": bundle_line_items,
            }
            r = requests.post(
                f"{ZOHO_INVENTORY_BASE}/bundles",
                headers=headers,
                json=payload,
                params={"organization_id": ORGANIZATION_ID},
                timeout=30,
            )
            data = r.json()
            if not r.ok or data.get("code") != 0:
                msg = data.get("message", "Unknown error")
                raise ValueError(f"Zoho bundle error for {comp.get('name', iid)}: {msg}")
            bundle = data.get("bundle", {})
            created_bundle_ids.append(bundle.get("bundle_id", ""))
            created_assembly_numbers.append(bundle.get("reference_number", ""))

        if not created_bundle_ids:
            raise ValueError("No assemblies were created")

        db[PO_COLLECTION].update_one(
            {"po_number": po_number},
            {
                "$set": {
                    "bundle_ids": created_bundle_ids,
                    "assembly_numbers": created_assembly_numbers,
                    "updated_at": datetime.now(),
                }
            },
        )
        return created_bundle_ids, created_assembly_numbers

    try:
        bundle_ids, assembly_numbers = await asyncio.to_thread(_create)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except requests.HTTPError as e:
        body_text = e.response.text if e.response is not None else ""
        raise HTTPException(status_code=502, detail=f"Zoho API error: {e} — {body_text}")

    return JSONResponse(
        status_code=201,
        content={
            "po_number": po_number,
            "bundle_ids": bundle_ids,
            "assembly_numbers": assembly_numbers,
        },
    )


@router.patch("/{po_number}/assemblies")
async def link_assembly(
    po_number: str, body: AssemblyLinkRequest, db=Depends(get_database)
):
    """Link an existing assembly to a PO by its reference_number."""

    def _link():
        doc = db[PO_COLLECTION].find_one({"po_number": po_number})
        if not doc:
            raise ValueError(f"PO {po_number} not found")

        assembly = db[ASSEMBLIES_COLLECTION].find_one(
            {"reference_number": body.assembly_number},
            {"bundle_id": 1, "reference_number": 1},
        )
        if not assembly:
            raise ValueError(
                f"Assembly {body.assembly_number!r} not found in assemblies collection"
            )

        existing_bundle_ids = doc.get("bundle_ids") or []
        existing_assembly_numbers = doc.get("assembly_numbers") or []
        if assembly["bundle_id"] in existing_bundle_ids:
            raise ValueError(f"Assembly {body.assembly_number!r} is already linked")

        existing_bundle_ids.append(assembly["bundle_id"])
        existing_assembly_numbers.append(assembly["reference_number"])

        db[PO_COLLECTION].update_one(
            {"po_number": po_number},
            {
                "$set": {
                    "bundle_ids": existing_bundle_ids,
                    "assembly_numbers": existing_assembly_numbers,
                    "updated_at": datetime.now(),
                }
            },
        )
        return existing_bundle_ids, existing_assembly_numbers

    try:
        bundle_ids, assembly_numbers = await asyncio.to_thread(_link)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return {
        "po_number": po_number,
        "bundle_ids": bundle_ids,
        "assembly_numbers": assembly_numbers,
    }


@router.delete("/{po_number}/assemblies")
async def unlink_assemblies(po_number: str, db=Depends(get_database)):
    """Remove all assembly links from a PO."""

    def _unlink():
        result = db[PO_COLLECTION].update_one(
            {"po_number": po_number},
            {
                "$unset": {"bundle_ids": "", "assembly_numbers": ""},
                "$set": {"updated_at": datetime.now()},
            },
        )
        return result.matched_count

    count = await asyncio.to_thread(_unlink)
    if not count:
        raise HTTPException(status_code=404, detail=f"PO {po_number} not found")
    return {"po_number": po_number, "unlinked": True}
