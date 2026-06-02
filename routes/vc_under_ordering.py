from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse
from datetime import datetime, timedelta
from ..database import get_database
import asyncio
import io
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal, ROUND_HALF_UP
from typing import Optional
from pydantic import BaseModel
import logging

from .amazon import compute_drr_for_asins_sync, generate_report_by_date_range

logger = logging.getLogger(__name__)
router = APIRouter()

OVERRIDES_COLLECTION = "vc_under_ordering_overrides"
SKU_MAPPING_COLLECTION = "amazon_sku_mapping"
MARGINS_COLLECTION = "vendor_margins"
INVENTORY_COLLECTION = "amazon_vendor_inventory"
SALES_COLLECTION = "amazon_vendor_sales"
ZOHO_STOCK_COLLECTION = "zoho_warehouse_stock"

DEFAULT_LEAD_TIME = 10
DEFAULT_COVERAGE_DAYS = 35

MONTH_NAMES = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}


class UnderOrderingOverride(BaseModel):
    open_po_override: Optional[int] = None
    lead_time_override: Optional[int] = None
    coverage_days_override: Optional[int] = None
    final_units_override: Optional[int] = None


async def _compute_drr_async(db, today: datetime, asins: list[str]) -> dict:
    """Compute DRR using the same PSR 'all' report function as vendor_po, so values always match."""
    drr_start = today - timedelta(days=89)
    report = await generate_report_by_date_range(
        drr_start.strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d"), db, report_type="all"
    )
    asin_set = set(asins)
    result = {
        item["asin"]: {"drr": item.get("drr", 0), "drr_flag": item.get("drr_flag", "")}
        for item in report
        if item.get("asin") in asin_set
    }
    missing = [a for a in asins if a not in result]
    if missing:
        end_dt = today.replace(hour=23, minute=59, second=59, microsecond=999999)
        fallback = await asyncio.to_thread(compute_drr_for_asins_sync, db, end_dt, missing)
        result.update(fallback)
    return result


def _fetch_data(db, drr_map: dict) -> tuple:
    today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    drr_start = today - timedelta(days=89)
    drr_period_label = f"{drr_start.strftime('%-d %b %Y')} – {today.strftime('%-d %b %Y')}"

    # 1. All ASINs from vendor_margins with etrade_asp set
    margins_docs = list(db[MARGINS_COLLECTION].find(
        {"etrade_asp": {"$exists": True, "$ne": None}},
        {"asin": 1, "etrade_asp": 1, "_id": 0},
    ))
    asins_with_asp = {d["asin"] for d in margins_docs if d.get("asin")}
    etrade_asp_by_asin: dict[str, float] = {
        d["asin"]: float(d["etrade_asp"]) for d in margins_docs if d.get("asin") and d.get("etrade_asp") is not None
    }

    if not asins_with_asp:
        return []

    # 2. amazon_sku_mapping for these ASINs
    sku_docs = list(db[SKU_MAPPING_COLLECTION].find(
        {"item_id": {"$in": list(asins_with_asp)}},
        {"item_id": 1, "sku_code": 1, "_id": 0},
    ))
    asin_to_sku: dict[str, str] = {d["item_id"]: d["sku_code"] for d in sku_docs if d.get("item_id") and d.get("sku_code")}
    asins = list(asin_to_sku.keys())

    if not asins:
        return []

    # 3. Product details by SKU
    skus = list(set(asin_to_sku.values()))
    product_by_sku: dict[str, dict] = {}
    for p in db["products"].find(
        {"cf_sku_code": {"$in": skus}},
        {"cf_sku_code": 1, "name": 1, "rate": 1, "status": 1, "purchase_status": 1, "item_id": 1, "_id": 0},
    ):
        product_by_sku[p["cf_sku_code"]] = p

    # 5. Current inventory with Etrade (latest date from amazon_vendor_inventory)
    etrade_latest = db[INVENTORY_COLLECTION].find_one(
        {"asin": {"$in": asins}}, {"date": 1}, sort=[("date", -1)]
    )
    etrade_inv_by_asin: dict[str, int] = {}
    inventory_date_str = ""
    if etrade_latest:
        etrade_date = etrade_latest["date"]
        inventory_date_str = etrade_date.strftime("%d %b %Y") if isinstance(etrade_date, datetime) else str(etrade_date)
        for doc in db[INVENTORY_COLLECTION].find(
            {"asin": {"$in": asins}, "date": etrade_date},
            {"asin": 1, "sellableOnHandInventoryUnits": 1},
        ):
            etrade_inv_by_asin[doc["asin"]] = int(doc.get("sellableOnHandInventoryUnits") or 0)

    # 6. Open PO from vendor_purchase_orders — exact same logic as vendor_po.py:
    #    processing → supply_qty_override if set, else final_supply_fo if set,
    #                 else supply_qty if > 0, else requested_qty
    #    packed/closed/intransit → accepted_qty
    open_po_by_asin: dict[str, int] = {}
    for doc in db["vendor_purchase_orders"].aggregate([
        {"$match": {"po_status": {"$in": ["processing", "packed", "closed", "intransit"]}}},
        {"$unwind": "$items"},
        {"$match": {"items.asin": {"$in": asins}}},
        {"$group": {
            "_id": "$items.asin",
            "total": {"$sum": {
                "$cond": {
                    "if": {"$eq": ["$po_status", "processing"]},
                    "then": {
                        "$cond": {
                            "if": {"$ne": [{"$ifNull": ["$items.supply_qty_override", None]}, None]},
                            "then": {"$ifNull": ["$items.supply_qty_override", 0]},
                            "else": {
                                "$cond": {
                                    "if": {"$ne": [{"$ifNull": ["$items.final_supply_fo", None]}, None]},
                                    "then": {"$ifNull": ["$items.final_supply_fo", 0]},
                                    "else": {
                                        "$cond": {
                                            "if": {"$gt": [{"$ifNull": ["$items.supply_qty", 0]}, 0]},
                                            "then": {"$ifNull": ["$items.supply_qty", 0]},
                                            "else": {"$ifNull": ["$items.requested_qty", 0]},
                                        }
                                    },
                                }
                            },
                        }
                    },
                    "else": {"$ifNull": ["$items.accepted_qty", 0]},
                }
            }},
        }},
    ]):
        open_po_by_asin[doc["_id"]] = int(doc["total"] or 0)

    # 7. Zoho stock
    item_id_by_sku: dict[str, str] = {}
    for p in db["products"].find({"cf_sku_code": {"$in": skus}}, {"cf_sku_code": 1, "item_id": 1, "_id": 0}):
        item_id_by_sku[p["cf_sku_code"]] = p.get("item_id", "")
    zoho_item_ids = list({v for v in item_id_by_sku.values() if v})
    zoho_latest: dict[str, int] = {}
    zoho_date_str = ""
    if zoho_item_ids:
        latest_zoho_date = None
        for doc in db[ZOHO_STOCK_COLLECTION].aggregate([
            {"$match": {"zoho_item_id": {"$in": zoho_item_ids}}},
            {"$sort": {"date": -1}},
            {"$group": {"_id": "$zoho_item_id", "warehouses": {"$first": "$warehouses"}, "date": {"$first": "$date"}}},
        ]):
            wh = doc.get("warehouses", {})
            zoho_latest[doc["_id"]] = int(sum(v for v in wh.values() if isinstance(v, (int, float)))) if isinstance(wh, dict) else 0
            d = doc.get("date")
            if d and (latest_zoho_date is None or d > latest_zoho_date):
                latest_zoho_date = d
        if latest_zoho_date:
            zoho_date_str = latest_zoho_date.strftime("%-d %b %Y") if isinstance(latest_zoho_date, datetime) else str(latest_zoho_date)

    # 8. Stock in transit from regular purchase_orders (Zoho POs, status "Issued")
    #    SKU → ASIN via amazon_sku_mapping (sku_code → item_id)
    sku_to_asin: dict[str, str] = {v: k for k, v in asin_to_sku.items()}

    def _build_sit() -> tuple[dict[str, float], dict[str, list[str]]]:
        sit_qty: dict[str, float] = {}
        sit_brands_map: dict[str, set[str]] = {}
        open_pos = list(db["purchase_orders"].find(
            {"order_status_formatted": "Issued"},
            {"purchaseorder_number": 1, "line_items": 1, "vendor_name": 1, "date": 1, "_id": 0},
        ))
        # Deduplicate Zoho-amended POs: PO-XYZ and PO-XYZ-1 are amendments of the
        # same shipment — keep only the latest by date to avoid double-counting.
        def _po_date_str(po: dict) -> str:
            d = po.get("date")
            if isinstance(d, datetime):
                return d.strftime("%Y-%m-%d")
            return str(d) if d else ""

        base_to_best: dict[str, dict] = {}
        for po in open_pos:
            base = re.sub(r"-\d+$", "", po.get("purchaseorder_number", ""))
            existing = base_to_best.get(base)
            if existing is None or _po_date_str(po) > _po_date_str(existing):
                base_to_best[base] = po
        for po in base_to_best.values():
            vendor = po.get("vendor_name", "")
            for li in po.get("line_items", []):
                sku_code = ""
                for cf in li.get("item_custom_fields", []):
                    if cf.get("api_name") == "cf_sku_code":
                        sku_code = cf.get("value", "")
                        break
                if not sku_code:
                    continue
                asin = sku_to_asin.get(sku_code)
                if not asin:
                    continue
                qty = float(li.get("quantity") or 0)
                qty_received = float(li.get("quantity_received") or 0)
                transit = qty - qty_received
                if transit > 0:
                    sit_qty[asin] = sit_qty.get(asin, 0) + transit
                    sit_brands_map.setdefault(asin, set()).add(vendor)
        return sit_qty, {a: sorted(b) for a, b in sit_brands_map.items()}

    sit_qty_by_asin, sit_brands_by_asin = _build_sit()

    # 9. Monthly sales — last 4 calendar months (FBA + VC combined, matching Amazon PSR)
    months: list[tuple[int, int]] = []
    ref = today.replace(day=1)
    for _ in range(4):
        ref = (ref - timedelta(days=1)).replace(day=1)
        months.append((ref.year, ref.month))
    months.reverse()

    earliest = datetime(months[0][0], months[0][1], 1)
    months_set = set(months)

    monthly_sales: dict[str, dict] = {a: {} for a in asins}

    # FBA sales (Seller Central / amazon_sales_traffic)
    for doc in db["amazon_sales_traffic"].aggregate([
        {"$match": {"parentAsin": {"$in": asins}, "date": {"$gte": earliest}}},
        {"$group": {
            "_id": {"asin": "$parentAsin", "year": {"$year": "$date"}, "month": {"$month": "$date"}},
            "total": {"$sum": "$salesByAsin.unitsOrdered"},
        }},
    ]):
        a = doc["_id"]["asin"]
        y, m = doc["_id"]["year"], doc["_id"]["month"]
        if (y, m) in months_set:
            label = f"{MONTH_NAMES[m]} {y}"
            monthly_sales.setdefault(a, {})
            monthly_sales[a][label] = monthly_sales[a].get(label, 0) + int(doc["total"] or 0)

    # VC sales (Vendor Central / amazon_vendor_sales)
    for doc in db[SALES_COLLECTION].aggregate([
        {"$match": {"asin": {"$in": asins}, "date": {"$gte": earliest}}},
        {"$group": {
            "_id": {"asin": "$asin", "year": {"$year": "$date"}, "month": {"$month": "$date"}},
            "total": {"$sum": "$orderedUnits"},
        }},
    ]):
        a = doc["_id"]["asin"]
        y, m = doc["_id"]["year"], doc["_id"]["month"]
        if (y, m) in months_set:
            label = f"{MONTH_NAMES[m]} {y}"
            monthly_sales.setdefault(a, {})
            monthly_sales[a][label] = monthly_sales[a].get(label, 0) + int(doc["total"] or 0)

    month_labels = [f"{MONTH_NAMES[m]} {y}" for (y, m) in months]

    # 10. Load overrides
    overrides: dict[str, dict] = {}
    for doc in db[OVERRIDES_COLLECTION].find({"asin": {"$in": asins}}):
        overrides[doc["asin"]] = doc

    # 11. Assemble rows
    rows = []
    for asin in sorted(asins):
        sku = asin_to_sku.get(asin, "")
        prod = product_by_sku.get(sku, {})
        override = overrides.get(asin, {})

        drr_info = drr_map.get(asin, {})
        drr_val = drr_info.get("drr") if isinstance(drr_info, dict) else None
        drr_flag = drr_info.get("drr_flag", "") if isinstance(drr_info, dict) else ""
        try:
            drr = float(drr_val) if drr_val is not None and str(drr_val) not in ("No data", "No stock") else 0.0
        except (TypeError, ValueError):
            drr = 0.0

        current_inv = etrade_inv_by_asin.get(asin, 0)
        open_po_auto = open_po_by_asin.get(asin, 0)
        open_po_override = override.get("open_po_override")
        open_po = open_po_override if open_po_override is not None else open_po_auto
        total_inv = current_inv + open_po

        net_total_days = round(total_inv / drr, 2) if drr > 0 else None

        lead_time = override.get("lead_time_override") if override.get("lead_time_override") is not None else DEFAULT_LEAD_TIME
        coverage_days = override.get("coverage_days_override") if override.get("coverage_days_override") is not None else DEFAULT_COVERAGE_DAYS
        total_target_days = lead_time + coverage_days
        target_stock = int(
            (Decimal(str(drr)) * Decimal(str(total_target_days))).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        ) if drr > 0 else 0

        # Final Units (under-ordering formula)
        final_units_override = override.get("final_units_override")
        if final_units_override is not None:
            final_units = final_units_override
        elif drr > 0 and net_total_days is not None:
            if net_total_days < lead_time:
                final_units = int(
                    (Decimal(str(drr)) * Decimal(str(total_target_days))).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
                )
            elif net_total_days > total_target_days:
                final_units = 0
            else:
                diff = Decimal(str(total_target_days)) - Decimal(str(net_total_days))
                final_units = int(diff * Decimal(str(drr)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        else:
            final_units = None

        item_id = item_id_by_sku.get(sku, "")
        zoho_stock = zoho_latest.get(item_id, 0)

        # If no Zoho warehouse stock, nothing to ship — force final units to 0
        if zoho_stock == 0 and final_units_override is None:
            final_units = 0

        sit_total = sit_qty_by_asin.get(asin, 0)
        sit_brands = ", ".join(sit_brands_by_asin.get(asin, []))

        rows.append({
            "asin": asin,
            "sku_code": sku,
            "item_name": prod.get("name", ""),
            "current_inv": current_inv,
            "open_po": open_po,
            "open_po_auto": open_po_auto,
            "open_po_overridden": open_po_override is not None,
            "total_inv": total_inv,
            "drr": round(drr, 2),
            "drr_flag": drr_flag,
            "net_total_days": net_total_days,
            "lead_time": lead_time,
            "lead_time_overridden": override.get("lead_time_override") is not None,
            "coverage_days": coverage_days,
            "coverage_days_overridden": override.get("coverage_days_override") is not None,
            "total_target_days": total_target_days,
            "target_stock": target_stock,
            "final_units": final_units,
            "final_units_overridden": final_units_override is not None,
            "zoho_stock": zoho_stock,
            "sit_total": round(sit_total, 2),
            "sit_brands": sit_brands,
            "status": prod.get("purchase_status") or prod.get("status") or "",
            "etrade_asp": etrade_asp_by_asin.get(asin),
            "monthly_sales": monthly_sales.get(asin, {}),
            "month_labels": month_labels,
        })

    return rows, inventory_date_str, drr_period_label, zoho_date_str


# ─── Endpoints ────────────────────────────────────────────────────────────────

@router.get("/data")
async def get_data(db=Depends(get_database)):
    try:
        today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        # Get ASINs first to compute DRR async (matches PSR/vendor_po)
        margins_docs = await asyncio.to_thread(
            lambda: list(db[MARGINS_COLLECTION].find(
                {"etrade_asp": {"$exists": True, "$ne": None}}, {"asin": 1, "_id": 0}
            ))
        )
        asins_with_asp = [d["asin"] for d in margins_docs if d.get("asin")]
        sku_docs = await asyncio.to_thread(
            lambda: list(db[SKU_MAPPING_COLLECTION].find(
                {"item_id": {"$in": asins_with_asp}}, {"item_id": 1, "_id": 0}
            ))
        )
        asins = [d["item_id"] for d in sku_docs if d.get("item_id")]
        drr_map = await _compute_drr_async(db, today, asins) if asins else {}
        result, inv_date, drr_period, zoho_date = await asyncio.to_thread(_fetch_data, db, drr_map)
        return JSONResponse({"rows": result, "inventory_date": inv_date, "drr_period": drr_period, "zoho_date": zoho_date})
    except Exception as e:
        logger.error(f"VC under ordering data error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/overrides/{asin}")
async def save_override(asin: str, body: UnderOrderingOverride, db=Depends(get_database)):
    update: dict = {}
    if body.open_po_override is not None:
        update["open_po_override"] = body.open_po_override
    if body.lead_time_override is not None:
        update["lead_time_override"] = body.lead_time_override
    if body.coverage_days_override is not None:
        update["coverage_days_override"] = body.coverage_days_override
    if body.final_units_override is not None:
        update["final_units_override"] = body.final_units_override

    await asyncio.to_thread(
        lambda: db[OVERRIDES_COLLECTION].update_one(
            {"asin": asin},
            {"$set": {"asin": asin, **update}},
            upsert=True,
        )
    )
    return JSONResponse({"ok": True})


@router.delete("/overrides/{asin}/{field}")
async def clear_override(asin: str, field: str, db=Depends(get_database)):
    allowed = {"open_po_override", "lead_time_override", "coverage_days_override", "final_units_override"}
    if field not in allowed:
        raise HTTPException(status_code=400, detail="Invalid field")
    await asyncio.to_thread(
        lambda: db[OVERRIDES_COLLECTION].update_one({"asin": asin}, {"$unset": {field: ""}})
    )
    return JSONResponse({"ok": True})


@router.get("/download")
async def download_xlsx(db=Depends(get_database)):
    try:
        today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        margins_docs = await asyncio.to_thread(
            lambda: list(db[MARGINS_COLLECTION].find(
                {"etrade_asp": {"$exists": True, "$ne": None}}, {"asin": 1, "_id": 0}
            ))
        )
        asins_with_asp = [d["asin"] for d in margins_docs if d.get("asin")]
        sku_docs = await asyncio.to_thread(
            lambda: list(db[SKU_MAPPING_COLLECTION].find(
                {"item_id": {"$in": asins_with_asp}}, {"item_id": 1, "_id": 0}
            ))
        )
        asins = [d["item_id"] for d in sku_docs if d.get("item_id")]
        drr_map = await _compute_drr_async(db, today, asins) if asins else {}
        result, inv_date, drr_period, zoho_date = await asyncio.to_thread(_fetch_data, db, drr_map)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Under-ordering (For All ASINs)"

    header_fill  = PatternFill("solid", fgColor="1F4E79")
    formula_fill  = PatternFill("solid", fgColor="D6EAF8")   # light blue — auto formula
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap    = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

    month_labels = result[0]["month_labels"] if result else []
    inv_label  = f"Current Inv\n(Etrade{(' – ' + inv_date) if inv_date else ''})"
    drr_label  = f"DRR\n({drr_period})" if drr_period else "DRR"
    zoho_label = f"Zoho Stock\n({zoho_date})" if zoho_date else "Zoho Stock"

    # Column layout (1-indexed):
    # A=1 ASIN  B=2 SKU  C=3 Item Name  D=4 Curr Inv  E=5 Open PO
    # F=6 Total Inv (formula)  G=7 DRR  H=8 Net Days (formula)
    # I=9 Lead Time  J=10 Coverage Days
    # K=11 Total Target Days (formula)  L=12 Target Stock (formula)
    # M=13 Final Units (formula or override)
    # N=14 Zoho Stock  O=15 Status  P+ monthly sales

    headers = [
        "ASIN", "SKU Code", "Item Name",
        inv_label, "Open PO Qty",
        "Total Inventory\n(=Curr Inv + Open PO)",
        drr_label, "Net Total Days\n(=Total Inv ÷ DRR)",
        "Lead Time", "Coverage Days",
        "Total Target Days\n(=Lead + Coverage)",
        "Target Stock\n(=DRR × Total Target Days)",
        "Final Units\n(under-ordering formula)",
        zoho_label,
        "Stock in Transit\n(Open Zoho POs)",
        "SIT Brands",
        "Status",
    ] + month_labels

    # Row 1: headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value     = h
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center
    ws.row_dimensions[1].height = 48

    # Data rows — formula cells use Excel formulas
    DATA_START = 2
    for row_offset, row in enumerate(result):
        r = DATA_START + row_offset
        ms = row.get("monthly_sales", {})

        # Static / input columns
        ws.cell(r, 1,  row["asin"])
        ws.cell(r, 2,  row["sku_code"])
        ws.cell(r, 3,  row["item_name"]).alignment = left_wrap
        ws.cell(r, 4,  row["current_inv"])
        ws.cell(r, 5,  row["open_po"])

        # F: Total Inventory = Current Inv + Open PO
        ws.cell(r, 6,  f"=D{r}+E{r}").fill = formula_fill

        # G: DRR (algorithmically computed — raw value)
        if row.get("drr", 0) == 0:
            ws.cell(r, 7, "Manual input required")
        elif row.get("drr_flag", "").startswith("OK"):
            ws.cell(r, 7, row["drr"])
        else:
            ws.cell(r, 7, row.get("drr_flag", ""))

        # H: Net Total Days = Total Inv / DRR  (blank if DRR=0)
        ws.cell(r, 8,  f'=IF(G{r}=0,"",ROUND(F{r}/G{r},2))').fill = formula_fill

        # I, J: Lead Time, Coverage Days (editable)
        ws.cell(r, 9,  row["lead_time"])
        ws.cell(r, 10, row["coverage_days"])

        # K: Total Target Days = Lead Time + Coverage Days
        ws.cell(r, 11, f"=I{r}+J{r}").fill = formula_fill

        # L: Target Stock = DRR × Total Target Days
        ws.cell(r, 12, f"=IF(G{r}=0,0,ROUND(G{r}*K{r},0))").fill = formula_fill

        # M: Final Units — override as hard value, or formula (0 if Zoho stock=0 or DRR=0)
        if row.get("final_units_overridden") and row.get("final_units") is not None:
            ws.cell(r, 13, row["final_units"])
        else:
            ws.cell(r, 13,
                f'=IF(N{r}=0,0,'
                f'IF(G{r}=0,"",'
                f'IF(H{r}<I{r},ROUND(G{r}*K{r},0),'
                f'IF(H{r}>K{r},0,'
                f'MAX(0,ROUND((K{r}-H{r})*G{r},0))))))'
            ).fill = formula_fill

        # N: Zoho Stock  O: Stock in Transit  P: SIT Brands  Q: Status
        ws.cell(r, 14, row["zoho_stock"])
        ws.cell(r, 15, row["sit_total"])
        ws.cell(r, 16, row["sit_brands"]).alignment = left_wrap
        ws.cell(r, 17, row["status"])

        # Monthly sales
        for lbl_idx, lbl in enumerate(month_labels):
            ws.cell(r, 18 + lbl_idx, ms.get(lbl, 0))

    # Column widths
    col_widths = [14, 14, 42, 12, 10, 14, 10, 14, 10, 12, 16, 18, 18, 10, 14, 36, 12] + [12] * len(month_labels)
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=vc_under_ordering.xlsx"},
    )


@router.post("/upload")
async def upload_overrides(file: UploadFile = File(...), db=Depends(get_database)):
    """Bulk update overrides from Excel. Expected columns: ASIN, Open PO Qty, Lead Time, Coverage Days, Final Units."""
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")

    # Find header row (look for ASIN column)
    header_row_idx = None
    headers: dict[str, int] = {}
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        row_lower = [str(c).strip().lower() if c is not None else "" for c in row]
        if "asin" in row_lower:
            header_row_idx = r_idx
            for c_idx, val in enumerate(row_lower):
                headers[val] = c_idx
            break

    if header_row_idx is None:
        raise HTTPException(status_code=400, detail="Could not find ASIN header row")

    def _col(name: str) -> Optional[int]:
        for k, v in headers.items():
            if name in k:
                return v
        return None

    col_asin = _col("asin")
    col_open_po = _col("open po")
    col_lead = _col("lead time")
    col_cover = _col("coverage")
    col_final = _col("final units")

    if col_asin is None:
        raise HTTPException(status_code=400, detail="ASIN column not found")

    def _safe_int(val):
        try:
            return int(float(val))
        except (TypeError, ValueError):
            return None

    updated = 0
    ops = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        asin = str(row[col_asin]).strip() if row[col_asin] else None
        if not asin or asin.lower() in ("none", "asin", ""):
            continue

        upd: dict = {}
        if col_open_po is not None:
            v = _safe_int(row[col_open_po])
            if v is not None:
                upd["open_po_override"] = v
        if col_lead is not None:
            v = _safe_int(row[col_lead])
            if v is not None:
                upd["lead_time_override"] = v
        if col_cover is not None:
            v = _safe_int(row[col_cover])
            if v is not None:
                upd["coverage_days_override"] = v
        if col_final is not None:
            v = _safe_int(row[col_final])
            if v is not None:
                upd["final_units_override"] = v

        if upd:
            ops.append({"asin": asin, **upd})
            updated += 1

    def _do_upserts():
        for op in ops:
            a = op.pop("asin")
            db[OVERRIDES_COLLECTION].update_one({"asin": a}, {"$set": {"asin": a, **op}}, upsert=True)

    await asyncio.to_thread(_do_upserts)
    return JSONResponse({"updated": updated})
