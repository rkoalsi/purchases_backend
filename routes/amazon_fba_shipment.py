from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse
from datetime import datetime, timedelta, date
from ..helpers.datetime_utils import utcnow
from ..database import get_database
import asyncio
import io
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from typing import Optional, Any
from pydantic import BaseModel
import logging
import re
import requests

from .amazon import compute_drr_for_asins_sync, generate_report_by_date_range, get_amazon_token, make_sp_api_request

logger = logging.getLogger(__name__)
router = APIRouter()

PLANNING_OVERRIDES_COLLECTION = "amazon_fba_shipment_planning_overrides"
PROCESSING_COLLECTION = "amazon_fba_shipment_processing"
SUMMARY_COLLECTION = "amazon_fba_shipment_summary"
SKU_MAPPING_COLLECTION = "amazon_sku_mapping"
FBA_SHIPMENTS_COLLECTION = "amazon_fba_shipments"

DEFAULT_LEAD_TIME = 10
DEFAULT_COVERAGE_DAYS = 30

# ─── Zoho Books (for estimate creation) ──────────────────────────────────────

_ZOHO_BOOKS_BASE = "https://books.zoho.com/api/v3"
_ZOHO_ORG_ID = os.getenv("ORGANIZATION_ID", "776755316")
_ZOHO_BOOKS_URL = os.getenv("BOOKS_URL")
_ZOHO_CLIENT_ID = os.getenv("CLIENT_ID")
_ZOHO_CLIENT_SECRET = os.getenv("CLIENT_SECRET")
_ZOHO_BOOKS_REFRESH_TOKEN = os.getenv("BOOKS_REFRESH_TOKEN")


def _get_zoho_books_token() -> str:
    url = _ZOHO_BOOKS_URL.format(
        clientId=_ZOHO_CLIENT_ID,
        clientSecret=_ZOHO_CLIENT_SECRET,
        grantType="refresh_token",
        books_refresh_token=_ZOHO_BOOKS_REFRESH_TOKEN,
    )
    r = requests.post(url, timeout=30)
    r.raise_for_status()
    return r.json()["access_token"]


MONTH_NAMES = {
    1: "Jan",
    2: "Feb",
    3: "Mar",
    4: "Apr",
    5: "May",
    6: "Jun",
    7: "Jul",
    8: "Aug",
    9: "Sep",
    10: "Oct",
    11: "Nov",
    12: "Dec",
}

EXCLUDED_ASINS = [
    "B0FRLVMP6V",
    "B0C5HPTBMH",
    "B09V2XYPX1",
    "B0D7SDWMMB",
    "B09XV4WXJF",
    "B0GLHF3MN1",
    "B0C3M9PRDK",
    "B0GL77X9NH",
    "B0FMS1Z318",
    "B0B7XWXZRF",
    "B0B7S3WVXD",
    "B0CDH72346",
    "B0B7X8BZ5M",
    "B0DDGT97VL",
    "B0CNPXH2J9",
    "B09B2TMGCM",
    "B0GLHNX5SP",
    "B0BJ8ZRHGW",
    "B0FZB5JXJZ",
    "B0B7S1F22L",
    "B09GR818PH",
    "B0B31VDTRT",
    "B0DN5NTTZ5",
    "B0CPQ3ZGJV",
    "B0FDKSXXND",
    "B0GXSQ9HTT",
    "B0C555B1MP",
    "B08KJ2G1Z3",
    "B0CNXMVPLH",
    "B09PV8KR9F",
    "B0GVQB9K99",
    "B0GMPTX2Y3",
    "B0DHD7PB5G",
    "B0GXSXCGDZ",
    "B0FY3Z7FCM",
    "B09T6XRRY5",
    "B0DDCPMNB4",
    "B09RH7S1FC",
    "B0CC4V41NK",
    "B0DYJDQ6Q7",
    "B09RH8Q2NC",
    "B0FV3JVBBH",
    "B0DKBKWKDL",
    "B0DDKHZKL3",
    "B0D33YLBNR",
    "B0FVXX3JB7",
    "B093C6V33C",
    "B0D4ZBFCLT",
    "B0G2WW1NN5",
    "B0B7S2VHGX",
    "B0GMPHHFKV",
    "B09FZL1L5Z",
    "B0FPDCDLGC",
    "B0D4JKYYX6",
    "B093CB5XV2",
    "B0GPXT2VPD",
    "B0FM8QH3FF",
    "B0CC53TTFM",
    "B093C8RDD2",
    "B08WPX98XR",
    "B0C6R272BG",
    "B0DDT9HMMW",
    "B0FXVMFC9C",
    "B0B81VHB39",
    "B0DD48LN2S",
    "B0G7VZDHR4",
    "B0DTTDG6YK",
    "B09RK8Y73J",
    "B09RK9VZQS",
    "B0DF7H32XY",
    "B0B7XWSYSN",
    "B08XMMVSGB",
    "B09FGH6K7Y",
    "B0CC4YFQWM",
    "B0FBGGQX57",
    "B0CGDWBDXF",
    "B0CNTH2NLK",
    "B0FV3HD63F",
    "B0GTFV7BPC",
    "B0F484C9V3",
    "B0BJ8ZNVDW",
    "B09RPT5XF9",
    "B0DKBKXNXC",
    "B0G5YFSBG5",
    "B09FZKX8XF",
    "B0DKFRSWZS",
    "B0D947RP1X",
    "B0DDH3B9FL",
    "B0C6R2HNJY",
    "B0CXJMW615",
    "B0B7XB3PH1",
    "B0CC52S93K",
    "B093C87K9H",
    "B0D7MTLW9S",
    "B08WPCYNRN",
    "B0DYJ6596P",
    "B0GLH692G5",
    "B0GFNFBJX9",
    "B09RPTM16Y",
    "B093C8P9M8",
    "B09HNKH3GB",
    "B0FDB7HYGH",
    "B09XV6QFH8",
    "B0C69GBXVP",
    "B0DDH9X9P1",
    "B0C3MB3826",
    "B0GRRVZB91",
    "B0D949FQ6F",
    "B09RH62YRR",
    "B0GLDVBWTG",
    "B096SVSNJS",
    "B09FGKFS4T",
    "B0FXVP9ZBM",
    "B09P85TSJB",
    "B0C69JWS33",
    "B09W5PFJWW",
    "B0C23K137X",
    "B0CPSFFWSH",
    "B0G1GYW6P3",
    "B08WPFCXGY",
    "B0B7RSDK88",
    "B0DM5932LJ",
    "B09P874YBJ",
    "B0GVG639SP",
    "B0GN2M9824",
]

# ─── Pydantic models ──────────────────────────────────────────────────────────


class PlanningOverride(BaseModel):
    asin: str
    open_shipment_qty_override: Optional[int] = None
    lead_time: Optional[int] = None
    coverage_days: Optional[int] = None
    final_units_override: Optional[int] = None
    fnsku: Optional[str] = None
    platform: Optional[str] = None
    mrp_override: Optional[float] = None
    sp_override: Optional[float] = None
    status_override: Optional[str] = None


class SummaryUpdate(BaseModel):
    reason_for_short_supply: Optional[str] = None
    appointment_initiated_date: Optional[str] = None
    appointment_date: Optional[str] = None
    dispatched_date: Optional[str] = None
    status: Optional[str] = None


# ─── DRR (same methodology as vendor_po and vc_under_ordering) ───────────────


async def _compute_drr_async(db, today: datetime, asins: list[str]) -> dict:
    """Compute DRR using the PSR 'all' report, matching vendor_po and vc_under_ordering."""
    if not asins:
        return {}
    drr_start = today - timedelta(days=89)
    report = await generate_report_by_date_range(
        drr_start.strftime("%Y-%m-%d"),
        today.strftime("%Y-%m-%d"),
        db,
        report_type="all",
    )
    asin_set = set(asins)
    result = {
        item["asin"]: {"drr": item.get("drr", 0), "drr_flag": item.get("drr_flag", "")}
        for item in report
        if item.get("asin") in asin_set
    }
    missing = [a for a in asins if a not in result]
    if missing:
        end_dt = today.replace(hour=23, minute=59, second=59, microsecond=999999)
        fallback = await asyncio.to_thread(
            compute_drr_for_asins_sync, db, end_dt, missing
        )
        result.update(fallback)
    return result


# ─── Sync helpers (run in thread) ────────────────────────────────────────────


def _get_asins(db) -> list[str]:
    docs = list(
        db[SKU_MAPPING_COLLECTION].find(
            {"item_id": {"$nin": EXCLUDED_ASINS}}, {"item_id": 1, "_id": 0}
        )
    )
    return list({d["item_id"] for d in docs if d.get("item_id")})


def _fetch_planning_data(db, drr_map: dict) -> tuple:
    """Compute full FBA shipment planning rows from MongoDB."""
    today = utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    fba_inv_date_str = ""
    zoho_date_str = ""
    etrade_date_str = ""

    # 1. All ASINs from amazon_sku_mapping
    sku_map_docs = list(
        db[SKU_MAPPING_COLLECTION].find(
            {}, {"item_id": 1, "sku_code": 1, "fnsku": 1, "amazon_status": 1, "_id": 0}
        )
    )
    if not sku_map_docs:
        return []

    asins = list({d["item_id"] for d in sku_map_docs if d.get("item_id")})
    asin_to_sku: dict[str, str] = {
        d["item_id"]: d["sku_code"]
        for d in sku_map_docs
        if d.get("item_id") and d.get("sku_code")
    }
    asin_to_fnsku_default: dict[str, str] = {
        d["item_id"]: d.get("fnsku", "") for d in sku_map_docs if d.get("item_id")
    }
    asin_to_amazon_status: dict[str, str] = {
        d["item_id"]: d.get("amazon_status", "") for d in sku_map_docs if d.get("item_id")
    }

    # Platform derived from vendor_margins flags
    etrade_po_asins: set[str] = set()
    etrade_df_asins: set[str] = set()
    for d in db["vendor_margins"].find(
        {"asin": {"$in": asins}, "$or": [{"etrade_po": True}, {"etrade_df": True}]},
        {"asin": 1, "etrade_po": 1, "etrade_df": 1, "_id": 0},
    ):
        a = d.get("asin")
        if not a:
            continue
        if d.get("etrade_po"):
            etrade_po_asins.add(a)
        elif d.get("etrade_df"):
            etrade_df_asins.add(a)

    # 2. Product details by SKU
    skus = list(asin_to_sku.values())
    product_by_sku: dict[str, dict] = {}
    for p in db["products"].find(
        {"cf_sku_code": {"$in": skus}},
        {
            "cf_sku_code": 1,
            "name": 1,
            "rate": 1,
            "item_tax_preferences": 1,
            "status": 1,
            "purchase_status": 1,
            "item_id": 1,  # needed for zoho_warehouse_stock lookup
            "_id": 0,
        },
    ):
        product_by_sku[p["cf_sku_code"]] = p

    # 3. Current FBA inventory from amazon_ledger (exclude VKSX, SELLABLE only)
    ledger_latest = db["amazon_ledger"].find_one(
        {
            "asin": {"$in": asins},
            "location": {"$ne": "VKSX"},
            "disposition": "SELLABLE",
        },
        {"date": 1},
        sort=[("date", -1)],
    )
    fba_inv_by_asin: dict[str, int] = {}
    if ledger_latest:
        fba_inv_date = ledger_latest["date"]
        fba_inv_date_str = (
            fba_inv_date.strftime("%-d %b %Y")
            if isinstance(fba_inv_date, datetime)
            else str(fba_inv_date)
        )
        for doc in db["amazon_ledger"].aggregate(
            [
                {
                    "$match": {
                        "asin": {"$in": asins},
                        "date": fba_inv_date,
                        "location": {"$ne": "VKSX"},
                        "disposition": "SELLABLE",
                    }
                },
                {
                    "$group": {
                        "_id": "$asin",
                        "total": {"$sum": "$ending_warehouse_balance"},
                    }
                },
            ]
        ):
            fba_inv_by_asin[doc["_id"]] = int(doc["total"] or 0)

    # 4. Open shipment qty from FBA processing sheet (uploaded shipments)
    open_fba_shipment: dict[str, int] = {}
    for doc in db[PROCESSING_COLLECTION].aggregate(
        [
            {"$match": {"asin": {"$in": asins}}},
            {
                "$group": {
                    "_id": "$asin",
                    "total": {"$sum": {"$ifNull": ["$requested_qty", 0]}},
                }
            },
        ]
    ):
        open_fba_shipment[doc["_id"]] = int(doc["total"] or 0)

    # 5. Latest Zoho stock by item_id
    zoho_item_ids = list(
        {p.get("item_id", "") for p in product_by_sku.values() if p.get("item_id")}
    )
    zoho_latest: dict[str, int] = {}
    if zoho_item_ids:
        latest_zoho_dt = None
        for doc in db["zoho_warehouse_stock"].aggregate(
            [
                {"$match": {"zoho_item_id": {"$in": zoho_item_ids}}},
                {"$sort": {"date": -1}},
                {
                    "$group": {
                        "_id": "$zoho_item_id",
                        "warehouses": {"$first": "$warehouses"},
                        "date": {"$first": "$date"},
                    }
                },
            ]
        ):
            wh = doc.get("warehouses", {})
            zoho_latest[doc["_id"]] = (
                int(wh.get("Pupscribe Enterprises Private Limited", 0) or 0)
                if isinstance(wh, dict)
                else 0
            )
            d = doc.get("date")
            if d and (latest_zoho_dt is None or d > latest_zoho_dt):
                latest_zoho_dt = d
        if latest_zoho_dt:
            zoho_date_str = (
                latest_zoho_dt.strftime("%-d %b %Y")
                if isinstance(latest_zoho_dt, datetime)
                else str(latest_zoho_dt)
            )

    item_id_by_sku: dict[str, str] = {}
    for p in db["products"].find(
        {"cf_sku_code": {"$in": skus}}, {"cf_sku_code": 1, "item_id": 1, "_id": 0}
    ):
        item_id_by_sku[p["cf_sku_code"]] = p.get("item_id", "")

    # 6. Current inventory with Etrade (amazon_vendor_inventory)
    etrade_latest_doc = db["amazon_vendor_inventory"].find_one(
        {"asin": {"$in": asins}}, {"date": 1}, sort=[("date", -1)]
    )
    etrade_inv_by_asin: dict[str, int] = {}
    if etrade_latest_doc:
        etrade_date = etrade_latest_doc["date"]
        etrade_date_str = (
            etrade_date.strftime("%-d %b %Y")
            if isinstance(etrade_date, datetime)
            else str(etrade_date)
        )
        for doc in db["amazon_vendor_inventory"].find(
            {"asin": {"$in": asins}, "date": etrade_date},
            {"asin": 1, "sellableOnHandInventoryUnits": 1},
        ):
            etrade_inv_by_asin[doc["asin"]] = int(
                doc.get("sellableOnHandInventoryUnits") or 0
            )

    # 7. Open PO (etrade) — same logic as vc_under_ordering.py:
    #    processing → supply_qty_override → final_supply_fo → supply_qty (if >0) → requested_qty
    #    packed/closed/intransit → accepted_qty
    open_etrade_po: dict[str, int] = {}
    for doc in db["vendor_purchase_orders"].aggregate(
        [
            {
                "$match": {
                    "po_status": {
                        "$in": ["processing", "packed", "closed", "intransit"]
                    }
                }
            },
            {"$unwind": "$items"},
            {"$match": {"items.asin": {"$in": asins}}},
            {
                "$group": {
                    "_id": "$items.asin",
                    "total": {
                        "$sum": {
                            "$cond": {
                                "if": {"$eq": ["$po_status", "processing"]},
                                "then": {
                                    "$cond": {
                                        "if": {
                                            "$ne": [
                                                {
                                                    "$ifNull": [
                                                        "$items.supply_qty_override",
                                                        None,
                                                    ]
                                                },
                                                None,
                                            ]
                                        },
                                        "then": {
                                            "$ifNull": ["$items.supply_qty_override", 0]
                                        },
                                        "else": {
                                            "$cond": {
                                                "if": {
                                                    "$ne": [
                                                        {
                                                            "$ifNull": [
                                                                "$items.final_supply_fo",
                                                                None,
                                                            ]
                                                        },
                                                        None,
                                                    ]
                                                },
                                                "then": {
                                                    "$ifNull": [
                                                        "$items.final_supply_fo",
                                                        0,
                                                    ]
                                                },
                                                "else": {
                                                    "$cond": {
                                                        "if": {
                                                            "$gt": [
                                                                {
                                                                    "$ifNull": [
                                                                        "$items.supply_qty",
                                                                        0,
                                                                    ]
                                                                },
                                                                0,
                                                            ]
                                                        },
                                                        "then": {
                                                            "$ifNull": [
                                                                "$items.supply_qty",
                                                                0,
                                                            ]
                                                        },
                                                        "else": {
                                                            "$ifNull": [
                                                                "$items.requested_qty",
                                                                0,
                                                            ]
                                                        },
                                                    }
                                                },
                                            }
                                        },
                                    }
                                },
                                "else": {"$ifNull": ["$items.accepted_qty", 0]},
                            }
                        }
                    },
                }
            },
        ]
    ):
        open_etrade_po[doc["_id"]] = int(doc["total"] or 0)

    # 8. Monthly sales — last 4 calendar months
    months: list[tuple[int, int]] = []
    ref = today.replace(day=1)
    for _ in range(4):
        ref = (ref - timedelta(days=1)).replace(day=1)
        months.append((ref.year, ref.month))
    months.reverse()  # oldest first

    def month_start_end(y: int, m: int):
        s = datetime(y, m, 1)
        e = datetime(y + 1, 1, 1) if m == 12 else datetime(y, m + 1, 1)
        return s, e

    monthly_sales: dict[str, dict] = {a: {} for a in asins}
    for y, m in months:
        label = f"{MONTH_NAMES[m]} {y}"
        ms, me = month_start_end(y, m)
        for doc in db["amazon_vendor_sales"].aggregate(
            [
                {"$match": {"asin": {"$in": asins}, "date": {"$gte": ms, "$lt": me}}},
                {"$group": {"_id": "$asin", "total": {"$sum": "$orderedUnits"}}},
            ]
        ):
            monthly_sales[doc["_id"]][label] = int(doc["total"] or 0)
        for doc in db["amazon_ledger"].aggregate(
            [
                {"$match": {"asin": {"$in": asins}, "date": {"$gte": ms, "$lt": me}}},
                {
                    "$group": {
                        "_id": "$asin",
                        "total": {"$sum": {"$ifNull": ["$customer_shipments", 0]}},
                    }
                },
            ]
        ):
            existing = monthly_sales[doc["_id"]].get(label, 0)
            monthly_sales[doc["_id"]][label] = existing + int(doc["total"] or 0)

    month_labels = [f"{MONTH_NAMES[m]} {y}" for (y, m) in months]

    # 9. Load overrides
    overrides: dict[str, dict] = {}
    for doc in db[PLANNING_OVERRIDES_COLLECTION].find({"asin": {"$in": asins}}):
        overrides[doc["asin"]] = doc

    # 10. Assemble rows
    rows = []
    for asin in asins:
        sku = asin_to_sku.get(asin, "")
        prod = product_by_sku.get(sku, {})
        override = overrides.get(asin, {})

        # MRP/SP: use override if present, else fall back to product rate
        mrp = (
            override["mrp_override"]
            if override.get("mrp_override") is not None
            else (prod.get("rate") or 0)
        )
        sp = (
            override["sp_override"]
            if override.get("sp_override") is not None
            else (prod.get("rate") or 0)
        )

        # Status: use override if present (falls back to product status for legacy compatibility)
        status = (
            override["status_override"]
            if override.get("status_override") is not None
            else prod.get("status", "")
        )
        purchase_status = prod.get("purchase_status", "")

        current_fba_inv = fba_inv_by_asin.get(asin, 0)

        open_shipment_qty = (
            override["open_shipment_qty_override"]
            if "open_shipment_qty_override" in override
            and override["open_shipment_qty_override"] is not None
            else open_fba_shipment.get(asin, 0)
        )

        total_inventory = current_fba_inv + open_shipment_qty

        drr_info = drr_map.get(asin, {})
        drr_val = drr_info.get("drr") if isinstance(drr_info, dict) else None
        try:
            drr = (
                float(drr_val)
                if drr_val is not None and str(drr_val) != "No data"
                else 0.0
            )
        except (TypeError, ValueError):
            drr = 0.0

        net_total_days = round(total_inventory / drr, 2) if drr > 0 else 0

        lead_time = (
            override.get("lead_time")
            if override.get("lead_time") is not None
            else DEFAULT_LEAD_TIME
        )
        coverage_days = (
            override.get("coverage_days")
            if override.get("coverage_days") is not None
            else DEFAULT_COVERAGE_DAYS
        )
        total_target_days = lead_time + coverage_days
        target_stock = round(drr * total_target_days, 0) if drr > 0 else 0

        if (
            "final_units_override" in override
            and override["final_units_override"] is not None
        ):
            final_units = override["final_units_override"]
        else:
            if drr == 0:
                final_units = 0
            elif net_total_days < lead_time:
                final_units = round(drr * total_target_days)
            elif net_total_days > total_target_days:
                final_units = 0
            else:
                final_units = max(0, round((total_target_days - net_total_days) * drr))

        item_id = item_id_by_sku.get(sku, "")
        zoho_stock = zoho_latest.get(item_id, 0)

        etrade_inv = etrade_inv_by_asin.get(asin, 0)
        open_po = open_etrade_po.get(asin, 0)
        total_qty = etrade_inv + open_po

        sales_by_month = monthly_sales.get(asin, {})

        if asin not in EXCLUDED_ASINS:
            rows.append(
                {
                    "asin": asin,
                    "sku_code": sku,
                    "fnsku": override.get("fnsku")
                    or asin_to_fnsku_default.get(asin, ""),
                    "amazon_status": asin_to_amazon_status.get(asin, "") or None,
                    "item_name": prod.get("name", ""),
                    "mrp": mrp,
                    "sp": sp,
                    "current_fba_inventory": current_fba_inv,
                    "open_shipment_qty": open_shipment_qty,
                    "open_shipment_qty_auto": open_fba_shipment.get(asin, 0),
                    "total_inventory": total_inventory,
                    "drr": round(drr, 2),
                    "net_total_days": net_total_days,
                    "lead_time": lead_time,
                    "coverage_days": coverage_days,
                    "total_target_days": total_target_days,
                    "target_stock": target_stock,
                    "final_units_to_send": max(0, final_units),
                    "zoho_stock": zoho_stock,
                    "status": status,
                    "purchase_status": purchase_status,
                    "platform": override.get("platform")
                    or (
                        "Etrade - PO & DF"
                        if asin in etrade_po_asins
                        else (
                            "Etrade - DF only"
                            if asin in etrade_df_asins
                            else "FBA or SC"
                        )
                    ),
                    "current_inventory_etrade": etrade_inv,
                    "open_po": open_po,
                    "total_qty": total_qty,
                    "monthly_sales": sales_by_month,
                    "month_labels": month_labels,
                    "open_shipment_overridden": "open_shipment_qty_override" in override
                    and override["open_shipment_qty_override"] is not None,
                    "final_units_overridden": "final_units_override" in override
                    and override["final_units_override"] is not None,
                }
            )

    return rows, fba_inv_date_str, zoho_date_str, etrade_date_str


_SHIPMENT_DATE_RE = re.compile(r'\((\d{2}/\d{2}/\d{4})')
_ACTIVE_STATUSES = ["WORKING", "SHIPPED", "IN_TRANSIT", "DELIVERED", "CHECKED_IN", "RECEIVING"]


def _fetch_processing_data(db) -> list[dict]:
    # 1. Manually uploaded rows
    uploaded_rows = list(db[PROCESSING_COLLECTION].find({}, {"_id": 0}))
    uploaded_ids: set[str] = {r["shipment_id"] for r in uploaded_rows if r.get("shipment_id")}

    # Normalise GST: older rows stored as decimal fraction (0.05) instead of percentage (5)
    for row in uploaded_rows:
        g = row.get("gst")
        if g is not None and 0 < g < 1:
            row["gst"] = round(g * 100, 2)

    # Enrich uploaded rows with shipment names from the queue collection
    if uploaded_ids:
        name_by_id: dict[str, str] = {
            doc["ShipmentId"]: doc.get("ShipmentName", "")
            for doc in db[FBA_SHIPMENTS_COLLECTION].find(
                {"ShipmentId": {"$in": list(uploaded_ids)}},
                {"ShipmentId": 1, "ShipmentName": 1, "_id": 0},
            )
            if doc.get("ShipmentId")
        }
        for row in uploaded_rows:
            row["shipment_name"] = name_by_id.get(row.get("shipment_id", ""), "")

    # 2. Active queue shipments not yet covered by an uploaded sheet
    queue_shipments = list(db[FBA_SHIPMENTS_COLLECTION].find(
        {
            "ShipmentId": {"$nin": list(uploaded_ids)},
            "items.0": {"$exists": True},
        },
        {
            "ShipmentId": 1,
            "ShipmentName": 1,
            "DestinationFulfillmentCenterId": 1,
            "items": 1,
            "_id": 0,
        },
    ))

    if not queue_shipments:
        return uploaded_rows

    # 3. Collect all SellerSKUs from queue items for product enrichment
    all_skus = list({
        item["SellerSKU"]
        for s in queue_shipments
        for item in s.get("items", [])
        if item.get("SellerSKU")
    })

    sku_to_asin: dict[str, str] = {
        doc["sku_code"]: doc.get("item_id", "")
        for doc in db[SKU_MAPPING_COLLECTION].find(
            {"sku_code": {"$in": all_skus}},
            {"sku_code": 1, "item_id": 1, "_id": 0},
        )
        if doc.get("sku_code")
    }

    prod_by_sku: dict[str, dict] = {
        p["cf_sku_code"]: p
        for p in db["products"].find(
            {"cf_sku_code": {"$in": all_skus}},
            {"cf_sku_code": 1, "name": 1, "rate": 1, "hsn_or_sac": 1, "item_tax_preferences": 1, "_id": 0},
        )
        if p.get("cf_sku_code")
    }

    def _gst(prod: dict):
        for pref in prod.get("item_tax_preferences") or []:
            if pref.get("tax_specification") == "intra":
                return pref.get("tax_percentage")
        return None

    def _parse_shipment_date(name: str | None):
        if not name:
            return None
        m = _SHIPMENT_DATE_RE.search(name)
        if not m:
            return None
        try:
            return datetime.strptime(m.group(1), "%d/%m/%Y")
        except ValueError:
            return None

    # 4. Build virtual rows for each queue shipment's items
    virtual_rows: list[dict] = []
    for shipment in queue_shipments:
        shipment_id = shipment["ShipmentId"]
        location = shipment.get("DestinationFulfillmentCenterId", "")
        date = _parse_shipment_date(shipment.get("ShipmentName"))
        for item in shipment.get("items", []):
            seller_sku = item.get("SellerSKU", "")
            prod = prod_by_sku.get(seller_sku, {})
            virtual_rows.append({
                "shipment_id": shipment_id,
                "shipment_name": shipment.get("ShipmentName", ""),
                "date": date,
                "location": location,
                "sku_code": seller_sku,
                "asin": sku_to_asin.get(seller_sku, ""),
                "fnsku": item.get("FulfillmentNetworkSKU", ""),
                "item_name": prod.get("name", ""),
                "mrp": prod.get("rate"),
                "sp": prod.get("rate"),
                "requested_qty": item.get("QuantityShipped", 0) or 0,
                "packed_qty": 0,
                "cost_price": None,
                "hsn_code": prod.get("hsn_or_sac", ""),
                "gst": _gst(prod),
                "_from_queue": True,
            })

    return uploaded_rows + virtual_rows


def _fetch_summary_data(db) -> list[dict]:
    pipeline = [
        {
            "$group": {
                "_id": {"shipment_id": "$shipment_id", "location": "$location"},
                "shipment_date": {"$first": "$date"},
                "requested_qty": {"$sum": "$requested_qty"},
                "dispatched_qty": {"$sum": "$packed_qty"},
            }
        },
        {
            "$project": {
                "_id": 0,
                "shipment_id": "$_id.shipment_id",
                "location": "$_id.location",
                "shipment_date": 1,
                "requested_qty": 1,
                "dispatched_qty": 1,
                "short_supply_qty": {
                    "$subtract": ["$requested_qty", "$dispatched_qty"]
                },
                "short_supply_pct": {
                    "$cond": {
                        "if": {"$gt": ["$dispatched_qty", 0]},
                        "then": {
                            "$divide": [
                                {"$subtract": ["$requested_qty", "$dispatched_qty"]},
                                "$dispatched_qty",
                            ]
                        },
                        "else": 0,
                    }
                },
            }
        },
        {"$sort": {"shipment_date": -1, "shipment_id": 1}},
    ]
    aggregated = list(db[PROCESSING_COLLECTION].aggregate(pipeline))

    stored_edits: dict[str, dict] = {}
    for doc in db[SUMMARY_COLLECTION].find({}, {"_id": 0}):
        key = f"{doc.get('shipment_id')}|||{doc.get('location', '')}"
        stored_edits[key] = doc

    rows = []
    for row in aggregated:
        key = f"{row['shipment_id']}|||{row.get('location', '')}"
        edit = stored_edits.get(key, {})
        rows.append(
            {
                **row,
                "reason_for_short_supply": edit.get("reason_for_short_supply", ""),
                "appointment_initiated_date": edit.get(
                    "appointment_initiated_date", ""
                ),
                "appointment_date": edit.get("appointment_date", ""),
                "dispatched_date": edit.get("dispatched_date", ""),
                "status": edit.get("status", "Pending"),
            }
        )
    return rows


def _generate_planning_excel(
    rows: list[dict],
    fba_inv_date: str = "",
    zoho_date: str = "",
    etrade_date: str = "",
    drr_period: str = "",
) -> bytes:
    """Generate Excel matching the FBA shipment format with formulas for computed columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FBA Shipment Planning"

    if not rows:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    month_labels = rows[0].get("month_labels", [])

    # Column layout:
    # A(1) ASIN  B(2) FNSKU  C(3) Amazon Status  D(4) SKU  E(5) Item Name
    # F(6) MRP  G(7) SP  H(8) FBA Inv  I(9) Open Shipment  J(10)=H+I Total Inv
    # K(11) DRR  L(12)=J/K Net Days  M(13) Lead  N(14) Coverage
    # O(15)=M+N Target Days  P(16)=K*O Target Stock  Q(17) Final Units
    # R(18) Zoho  S(19) Status  T(20) Platform
    # U(21) Etrade Inv  V(22) Open PO  W(23)=U+V Total Qty
    # X+(24+) Monthly Sales

    fba_inv_label = (
        f"Current FBA\nInventory\n({fba_inv_date})"
        if fba_inv_date
        else "Current FBA\nInventory"
    )
    drr_label = f"DRR\n({drr_period})" if drr_period else "DRR"
    zoho_label = f"Zoho Stock\n({zoho_date})" if zoho_date else "Zoho Stock"
    etrade_label = (
        f"Current Inventory\nwith Etrade\n({etrade_date})"
        if etrade_date
        else "Current Inventory\nwith Etrade"
    )

    headers = [
        "ASIN",                                          # A  1
        "FNSKU",                                         # B  2
        "Amazon Status",                                 # C  3
        "SKU Code",                                      # D  4
        "Item Name",                                     # E  5
        "MRP",                                           # F  6
        "SP",                                            # G  7
        fba_inv_label,                                   # H  8
        "Open Shipment\nQty",                            # I  9
        "Total Inventory\n(FBA + Open Shipment)",        # J 10  =H+I
        drr_label,                                       # K 11
        "Net Total Days\n(=Total Inv ÷ DRR)",            # L 12  =J/K
        "Lead Time",                                     # M 13
        "Coverage Days",                                 # N 14
        "Total Target Days\n(=Lead + Coverage)",         # O 15  =M+N
        "Target Stock\n(=DRR × Total Target Days)",      # P 16  =K*O
        "Final Units\n(under-ordering formula)",         # Q 17
        zoho_label,                                      # R 18
        "Purchase Status",                               # S 19
        "Platform",                                      # T 20
        etrade_label,                                    # U 21
        "Open PO",                                       # V 22
        "Total Qty\n(Etrade + Open PO)",                 # W 23  =U+V
    ] + [f"{label}\nUnits Sold" for label in month_labels]  # X+ 24+

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    ws.row_dimensions[1].height = 55

    def _num(v):
        """Return numeric value as-is (including 0); return None only for None/missing."""
        return v if v is not None else None

    for row_idx, row in enumerate(rows, 2):
        r = row_idx
        drr_val = row.get("drr")  # keep 0.0 — formulas depend on it

        ws.cell(row=r, column=1,  value=row.get("asin", ""))                        # A ASIN
        ws.cell(row=r, column=2,  value=row.get("fnsku") or None)                   # B FNSKU
        ws.cell(row=r, column=3,  value=row.get("amazon_status") or None)           # C Amazon Status
        ws.cell(row=r, column=4,  value=row.get("sku_code", ""))                    # D SKU Code
        c = ws.cell(row=r, column=5, value=row.get("item_name") or None)            # E Item Name
        c.alignment = left_wrap
        ws.cell(row=r, column=6,  value=_num(row.get("mrp")))                       # F MRP
        ws.cell(row=r, column=7,  value=_num(row.get("sp")))                        # G SP
        ws.cell(row=r, column=8,  value=_num(row.get("current_fba_inventory")))     # H FBA Inv
        ws.cell(row=r, column=9,  value=_num(row.get("open_shipment_qty")))         # I Open Shipment
        ws.cell(row=r, column=10, value=f"=H{r}+I{r}")                             # J Total Inv
        ws.cell(row=r, column=11, value=drr_val)                                    # K DRR
        ws.cell(row=r, column=12, value=f'=IF(K{r}=0,"",ROUND(J{r}/K{r},2))')     # L Net Days
        ws.cell(row=r, column=13, value=row.get("lead_time", DEFAULT_LEAD_TIME))    # M Lead Time
        ws.cell(row=r, column=14, value=row.get("coverage_days", DEFAULT_COVERAGE_DAYS))  # N Coverage
        ws.cell(row=r, column=15, value=f"=M{r}+N{r}")                             # O Target Days
        ws.cell(row=r, column=16, value=f"=IF(K{r}=0,0,ROUND(K{r}*O{r},0))")      # P Target Stock
        ws.cell(
            row=r, column=17,
            value=f'=IF(K{r}=0,"",IF(L{r}<M{r},ROUND(K{r}*O{r},0),IF(L{r}>O{r},0,MAX(0,ROUND((O{r}-L{r})*K{r},0)))))',
        )                                                                            # Q Final Units
        ws.cell(row=r, column=18, value=_num(row.get("zoho_stock")))                # R Zoho Stock
        ws.cell(row=r, column=19, value=row.get("purchase_status") or None)         # S Purchase Status
        ws.cell(row=r, column=20, value=row.get("platform") or None)                # T Platform
        ws.cell(row=r, column=21, value=_num(row.get("current_inventory_etrade")))  # U Etrade Inv
        ws.cell(row=r, column=22, value=_num(row.get("open_po")))                   # V Open PO
        ws.cell(row=r, column=23, value=f"=U{r}+V{r}")                             # W Total Qty

        monthly_sales = row.get("monthly_sales", {})
        for offset, label in enumerate(month_labels):
            val = monthly_sales.get(label)
            ws.cell(row=r, column=24 + offset, value=_num(val))                    # X+ Monthly Sales

    col_widths = [
        14,  # A ASIN
        14,  # B FNSKU
        22,  # C Amazon Status
        18,  # D SKU Code
        50,  # E Item Name
        8,   # F MRP
        8,   # G SP
        12,  # H FBA Inv
        12,  # I Open Shipment
        14,  # J Total Inv
        10,  # K DRR
        12,  # L Net Days
        9,   # M Lead Time
        10,  # N Coverage
        14,  # O Target Days
        14,  # P Target Stock
        14,  # Q Final Units
        10,  # R Zoho Stock
        12,  # S Status
        12,  # T Platform
        15,  # U Etrade Inv
        9,   # V Open PO
        12,  # W Total Qty
    ]
    col_widths += [14] * len(month_labels)
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "F2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── Page 1: Planning ─────────────────────────────────────────────────────────


@router.get("/planning")
async def get_fba_planning(database=Depends(get_database)):
    try:
        today = utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        drr_start = today - timedelta(days=89)
        drr_period = (
            f"{drr_start.strftime('%-d %b %Y')} – {today.strftime('%-d %b %Y')}"
        )
        asins = await asyncio.to_thread(_get_asins, database)
        drr_map = await _compute_drr_async(database, today, asins)
        rows, fba_inv_date, zoho_date, etrade_date = await asyncio.to_thread(
            _fetch_planning_data, database, drr_map
        )
        return {
            "rows": rows,
            "fba_inv_date": fba_inv_date,
            "zoho_date": zoho_date,
            "etrade_date": etrade_date,
            "drr_period": drr_period,
        }
    except Exception as e:
        logger.error(f"Error fetching FBA planning data: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/planning/download")
async def download_fba_planning(database=Depends(get_database)):
    try:
        today = utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        drr_start = today - timedelta(days=89)
        drr_period = (
            f"{drr_start.strftime('%-d %b %Y')} – {today.strftime('%-d %b %Y')}"
        )
        asins = await asyncio.to_thread(_get_asins, database)
        drr_map = await _compute_drr_async(database, today, asins)
        rows, fba_inv_date, zoho_date, etrade_date = await asyncio.to_thread(
            _fetch_planning_data, database, drr_map
        )
        excel_bytes = _generate_planning_excel(
            rows, fba_inv_date, zoho_date, etrade_date, drr_period
        )
        filename = f"fba_shipment_planning_{today.strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        logger.error(f"Error generating FBA planning Excel: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/planning/{asin}")
async def update_planning_override(
    asin: str, payload: PlanningOverride, database=Depends(get_database)
):
    try:
        update_doc = {}
        if payload.open_shipment_qty_override is not None:
            update_doc["open_shipment_qty_override"] = (
                payload.open_shipment_qty_override
            )
        if payload.lead_time is not None:
            update_doc["lead_time"] = payload.lead_time
        if payload.coverage_days is not None:
            update_doc["coverage_days"] = payload.coverage_days
        if payload.final_units_override is not None:
            update_doc["final_units_override"] = payload.final_units_override
        if payload.fnsku is not None:
            update_doc["fnsku"] = payload.fnsku
        if payload.platform is not None:
            update_doc["platform"] = payload.platform
        if payload.mrp_override is not None:
            update_doc["mrp_override"] = payload.mrp_override
        if payload.sp_override is not None:
            update_doc["sp_override"] = payload.sp_override
        if payload.status_override is not None:
            update_doc["status_override"] = payload.status_override

        if update_doc:

            def _upsert(db):
                db[PLANNING_OVERRIDES_COLLECTION].update_one(
                    {"asin": asin},
                    {"$set": {**update_doc, "updated_at": utcnow()}},
                    upsert=True,
                )

            await asyncio.to_thread(_upsert, database)

        return {"success": True}
    except Exception as e:
        logger.error(f"Error updating planning override for {asin}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/planning/{asin}/override")
async def clear_planning_override(
    asin: str, field: str, database=Depends(get_database)
):
    """Clear a specific override field so the auto-computed value is used."""
    allowed = {
        "open_shipment_qty_override",
        "lead_time",
        "coverage_days",
        "final_units_override",
        "mrp_override",
        "sp_override",
        "status_override",
    }
    if field not in allowed:
        raise HTTPException(status_code=400, detail=f"Field {field} is not clearable")

    def _clear(db):
        db[PLANNING_OVERRIDES_COLLECTION].update_one(
            {"asin": asin},
            {"$unset": {field: ""}, "$set": {"updated_at": utcnow()}},
        )

    await asyncio.to_thread(_clear, database)
    return {"success": True}


@router.post("/planning/upload")
async def upload_planning_overrides(
    file: UploadFile = File(...), database=Depends(get_database)
):
    """
    Upload the FBA shipment planning Excel to populate amazon_sku_mapping and overrides.
    Expected columns match the reference format:
      ASIN, SKU Code, FNSKU, Item Name, MRP, SP, Lead Time, Coverage Days, Status, Platform,
      Open Shipment Qty (optional), Final Units (optional)
    """
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=400, detail="Empty file")

        # Find header row
        header_row = None
        header_idx = 0
        for i, row in enumerate(rows):
            row_lower = [str(c).strip().lower() if c is not None else "" for c in row]
            if "asin" in row_lower:
                header_row = row_lower
                header_idx = i
                break

        if header_row is None:
            raise HTTPException(
                status_code=400, detail="Could not find header row with 'ASIN' column"
            )

        def _col(*names):
            for name in names:
                n = name.lower().strip()
                if n in header_row:
                    return header_row.index(n)
                # try partial match for multi-line headers
                for i, h in enumerate(header_row):
                    if n in h:
                        return i
            return None

        idx_asin = _col("asin")
        idx_sku = _col("sku code", "sku_code")
        idx_fnsku = _col("fnsku")
        idx_item_name = _col("item name", "item_name")
        idx_mrp = _col("mrp")
        idx_sp = _col("sp")
        idx_lead = _col("lead time", "lead_time")
        idx_coverage = _col("coverage days", "coverage_days")
        idx_status = _col("status")
        idx_platform = _col("platform")
        idx_open_qty = _col("open shipment qty", "open_shipment_qty")
        idx_final = _col("final units", "final_units_to_send")

        if idx_asin is None:
            raise HTTPException(status_code=400, detail="Missing 'ASIN' column")

        def _int_or_none(val):
            try:
                return (
                    int(float(val))
                    if val is not None and str(val).strip() not in ("", "None", "N/A")
                    else None
                )
            except (TypeError, ValueError):
                return None

        def _float_or_none(val):
            try:
                return (
                    float(val)
                    if val is not None and str(val).strip() not in ("", "None", "N/A")
                    else None
                )
            except (TypeError, ValueError):
                return None

        def _str_or_none(val):
            s = str(val).strip() if val is not None else ""
            return s if s and s.lower() not in ("none", "n/a", "") else None

        updated_sku_mapping = 0
        updated_overrides = 0

        def _process_rows(db):
            nonlocal updated_sku_mapping, updated_overrides
            for row in rows[header_idx + 1 :]:
                if not any(row):
                    continue
                asin_val = _str_or_none(row[idx_asin]) if idx_asin is not None else None
                if not asin_val:
                    continue

                sku_val = _str_or_none(row[idx_sku]) if idx_sku is not None else None
                fnsku_val = (
                    _str_or_none(row[idx_fnsku]) if idx_fnsku is not None else None
                )

                # Upsert amazon_sku_mapping
                if sku_val or fnsku_val:
                    sku_map_update: dict = {"updated_at": utcnow()}
                    if sku_val:
                        sku_map_update["sku_code"] = sku_val
                    if fnsku_val:
                        sku_map_update["fnsku"] = fnsku_val
                    db[SKU_MAPPING_COLLECTION].update_one(
                        {"item_id": asin_val},
                        {"$set": {**sku_map_update, "item_id": asin_val}},
                        upsert=True,
                    )
                    updated_sku_mapping += 1

                # Build planning overrides doc
                override_update: dict = {"updated_at": utcnow()}

                if fnsku_val:
                    override_update["fnsku"] = fnsku_val

                mrp_val = _float_or_none(row[idx_mrp]) if idx_mrp is not None else None
                if mrp_val is not None:
                    override_update["mrp_override"] = mrp_val

                sp_val = _float_or_none(row[idx_sp]) if idx_sp is not None else None
                if sp_val is not None:
                    override_update["sp_override"] = sp_val

                lead_val = _int_or_none(row[idx_lead]) if idx_lead is not None else None
                if lead_val is not None:
                    override_update["lead_time"] = lead_val

                coverage_val = (
                    _int_or_none(row[idx_coverage])
                    if idx_coverage is not None
                    else None
                )
                if coverage_val is not None:
                    override_update["coverage_days"] = coverage_val

                status_val = (
                    _str_or_none(row[idx_status]) if idx_status is not None else None
                )
                if status_val is not None:
                    override_update["status_override"] = status_val

                platform_val = (
                    _str_or_none(row[idx_platform])
                    if idx_platform is not None
                    else None
                )
                if platform_val is not None:
                    override_update["platform"] = platform_val

                open_qty_val = (
                    _int_or_none(row[idx_open_qty])
                    if idx_open_qty is not None
                    else None
                )
                if open_qty_val is not None:
                    override_update["open_shipment_qty_override"] = open_qty_val

                final_val = (
                    _int_or_none(row[idx_final]) if idx_final is not None else None
                )
                if final_val is not None:
                    override_update["final_units_override"] = final_val

                if len(override_update) > 1:  # more than just updated_at
                    db[PLANNING_OVERRIDES_COLLECTION].update_one(
                        {"asin": asin_val},
                        {"$set": override_update},
                        upsert=True,
                    )
                    updated_overrides += 1

        await asyncio.to_thread(_process_rows, database)
        return {
            "success": True,
            "sku_mapping_updated": updated_sku_mapping,
            "overrides_updated": updated_overrides,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading planning overrides: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ─── Page 2: Processing ───────────────────────────────────────────────────────


@router.get("/processing/template")
async def download_fba_processing_template():
    """Return a blank Excel template with the expected column headers for FBA processing upload."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FBA Processing"

    headers = [
        "Shipment ID", "Date", "Location", "SKU Code", "ASIN", "FNSKU",
        "Item Name", "MRP", "SP", "Requested Qty", "Packed Qty",
        "Cost Price", "HSN Code", "GST",
    ]

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # Example row
    ws.append([
        "FBA1234567890", "2024-01-15", "Mumbai WH", "SKU001",
        "B01234567890", "X001ABCDEF", "Sample Product Name",
        999, 899, 100, 100, 450, "33061090", 18,
    ])

    col_widths = [18, 12, 16, 14, 14, 14, 35, 8, 8, 13, 11, 11, 12, 6]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="fba_processing_template.xlsx"'},
    )


@router.get("/processing")
async def get_fba_processing(database=Depends(get_database)):
    try:
        rows = await asyncio.to_thread(_fetch_processing_data, database)
        return {"rows": rows}
    except Exception as e:
        logger.error(f"Error fetching FBA processing data: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/processing/upload")
async def upload_fba_processing(
    file: UploadFile = File(...), database=Depends(get_database)
):
    """
    Upload FBA shipment processing sheet.
    Expected columns: Shipment ID, Date, Location, SKU Code, ASIN, FNSKU, Item Name, MRP, SP,
                      Requested Qty, Packed Qty, Cost Price, HSN Code, GST
    """
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active

        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise HTTPException(status_code=400, detail="Empty file")

        header_row = None
        header_idx = 0
        for i, row in enumerate(all_rows):
            row_clean = [str(c).strip().lower() if c else "" for c in row]
            if "shipment id" in row_clean or "shipment_id" in row_clean:
                header_row = row_clean
                header_idx = i
                break

        if header_row is None:
            raise HTTPException(
                status_code=400,
                detail="Could not find header row with 'Shipment ID' column",
            )

        def _find_col(*names):
            for name in names:
                try:
                    return header_row.index(name.lower())
                except ValueError:
                    pass
            return None

        idx_shipment_id = _find_col("shipment id", "shipment_id")
        idx_date = _find_col("date")
        idx_location = _find_col("location")
        idx_sku = _find_col("sku code", "sku_code")
        idx_asin = _find_col("asin")
        idx_fnsku = _find_col("fnsku")
        idx_item_name = _find_col("item name", "item_name")
        idx_mrp = _find_col("mrp")
        idx_sp = _find_col("sp")
        idx_requested_qty = _find_col("requested qty", "requested_qty")
        idx_packed_qty = _find_col("packed qty", "packed_qty")
        idx_cost_price = _find_col("cost price", "cost_price")
        idx_hsn = _find_col("hsn code", "hsn_code", "hsn")
        idx_gst = _find_col("gst")

        if idx_shipment_id is None:
            raise HTTPException(status_code=400, detail="Missing 'Shipment ID' column")

        def _safe_num(val):
            try:
                return (
                    float(val)
                    if val is not None and str(val).strip() not in ("", "None")
                    else None
                )
            except (TypeError, ValueError):
                return None

        def _safe_int(val):
            try:
                return (
                    int(float(val))
                    if val is not None and str(val).strip() not in ("", "None")
                    else 0
                )
            except (TypeError, ValueError):
                return 0

        def _safe_str(val, idx):
            if idx is None or val is None:
                return ""
            return str(val).strip()

        def _parse_date(val):
            if val is None:
                return None
            if isinstance(val, datetime):
                return val
            if isinstance(val, date):
                return datetime(val.year, val.month, val.day)
            try:
                return datetime.strptime(str(val).strip(), "%Y-%m-%d")
            except ValueError:
                try:
                    return datetime.strptime(str(val).strip(), "%d/%m/%Y")
                except ValueError:
                    return None

        records = []
        for raw in all_rows[header_idx + 1 :]:
            if not any(raw):
                continue
            shipment_id = _safe_str(raw[idx_shipment_id], idx_shipment_id)
            if not shipment_id:
                continue

            record = {
                "shipment_id": shipment_id,
                "date": _parse_date(raw[idx_date] if idx_date is not None else None),
                "location": _safe_str(
                    raw[idx_location] if idx_location is not None else None,
                    idx_location,
                ),
                "sku_code": _safe_str(
                    raw[idx_sku] if idx_sku is not None else None, idx_sku
                ),
                "asin": _safe_str(
                    raw[idx_asin] if idx_asin is not None else None, idx_asin
                ),
                "fnsku": _safe_str(
                    raw[idx_fnsku] if idx_fnsku is not None else None, idx_fnsku
                ),
                "item_name": _safe_str(
                    raw[idx_item_name] if idx_item_name is not None else None,
                    idx_item_name,
                ),
                "mrp": _safe_num(raw[idx_mrp] if idx_mrp is not None else None),
                "sp": _safe_num(raw[idx_sp] if idx_sp is not None else None),
                "requested_qty": _safe_int(
                    raw[idx_requested_qty] if idx_requested_qty is not None else None
                ),
                "packed_qty": _safe_int(
                    raw[idx_packed_qty] if idx_packed_qty is not None else None
                ),
                "cost_price": _safe_num(
                    raw[idx_cost_price] if idx_cost_price is not None else None
                ),
                "hsn_code": _safe_str(
                    raw[idx_hsn] if idx_hsn is not None else None, idx_hsn
                ),
                "gst": _safe_num(raw[idx_gst] if idx_gst is not None else None),
                "uploaded_at": utcnow(),
            }
            records.append(record)

        if not records:
            raise HTTPException(status_code=400, detail="No data rows found in file")

        def _store(db):
            for rec in records:
                db[PROCESSING_COLLECTION].update_one(
                    {
                        "shipment_id": rec["shipment_id"],
                        "sku_code": rec["sku_code"],
                        "asin": rec["asin"],
                    },
                    {"$set": rec},
                    upsert=True,
                )

        await asyncio.to_thread(_store, database)

        # Immediately link any matching DB shipments
        linked = await asyncio.to_thread(_relink_processing_uploads, database)

        return {"success": True, "rows_uploaded": len(records), "db_shipments_linked": linked}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading FBA processing data: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


class ClearProcessingRequest(BaseModel):
    shipment_ids: Optional[list[str]] = None


@router.delete("/processing")
async def clear_fba_processing(
    payload: ClearProcessingRequest = ClearProcessingRequest(),
    database=Depends(get_database),
):
    """Delete processing records. If shipment_ids provided, deletes only those; otherwise deletes all."""

    def _clear(db):
        ids = payload.shipment_ids or []
        query = {"shipment_id": {"$in": ids}} if ids else {}
        # Collect affected IDs before deletion for flag cleanup
        affected = ids if ids else list(
            {doc["shipment_id"] for doc in db[PROCESSING_COLLECTION].find({}, {"shipment_id": 1, "_id": 0})}
        )
        deleted = db[PROCESSING_COLLECTION].delete_many(query).deleted_count
        if affected:
            # Only clear the flag for shipments that now have no remaining processing rows
            still_present = {
                doc["shipment_id"]
                for doc in db[PROCESSING_COLLECTION].find(
                    {"shipment_id": {"$in": affected}}, {"shipment_id": 1, "_id": 0}
                )
            }
            to_unlink = [sid for sid in affected if sid not in still_present]
            if to_unlink:
                db[FBA_SHIPMENTS_COLLECTION].update_many(
                    {"ShipmentId": {"$in": to_unlink}},
                    {"$set": {"has_processing_upload": False}},
                )
        return deleted

    deleted = await asyncio.to_thread(_clear, database)
    return {"success": True, "deleted": deleted}


class SaveProcessingRowsRequest(BaseModel):
    rows: list[Any]


@router.post("/processing/rows")
async def save_processing_rows(
    payload: SaveProcessingRowsRequest,
    database=Depends(get_database),
):
    """Save processing rows as JSON (no Excel upload). Used when initialising from queue-sourced data."""

    def _store(db):
        now = utcnow()
        saved = 0
        for rec in payload.rows:
            shipment_id = rec.get("shipment_id", "")
            sku_code = rec.get("sku_code", "")
            if not shipment_id or not sku_code:
                continue
            doc = {k: v for k, v in rec.items() if k not in ("_from_queue", "_id")}
            # Parse date from ISO string if needed
            raw_date = doc.get("date")
            if isinstance(raw_date, str) and raw_date:
                try:
                    doc["date"] = datetime.fromisoformat(raw_date)
                except ValueError:
                    doc["date"] = None
            doc["uploaded_at"] = now
            db[PROCESSING_COLLECTION].update_one(
                {"shipment_id": shipment_id, "sku_code": sku_code, "asin": rec.get("asin", "")},
                {"$set": doc},
                upsert=True,
            )
            saved += 1
        return saved

    saved = await asyncio.to_thread(_store, database)
    linked = await asyncio.to_thread(_relink_processing_uploads, database)
    return {"success": True, "rows_saved": saved, "db_shipments_linked": linked}


# ─── Processing: per-shipment Excel download ──────────────────────────────────


@router.get("/processing/{shipment_id}/download")
async def download_processing_by_shipment(
    shipment_id: str, database=Depends(get_database)
):
    """Download all processing rows for a specific shipment as Excel."""
    try:
        def _fetch(db):
            all_rows = _fetch_processing_data(db)
            return [r for r in all_rows if r.get("shipment_id") == shipment_id]

        rows = await asyncio.to_thread(_fetch, database)
        if not rows:
            raise HTTPException(
                status_code=404,
                detail=f"No processing data found for shipment {shipment_id}",
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FBA Processing"

        headers = [
            "Shipment ID", "Date", "Location", "SKU Code", "ASIN", "FNSKU",
            "Item Name", "MRP", "SP", "Requested Qty", "Packed Qty",
            "Cost Price", "HSN Code", "GST",
        ]
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        for row_idx, row in enumerate(rows, 2):
            date_val = row.get("date")
            if isinstance(date_val, datetime):
                date_str = date_val.strftime("%Y-%m-%d")
            elif date_val:
                date_str = str(date_val)
            else:
                date_str = None
            ws.cell(row=row_idx, column=1, value=row.get("shipment_id", ""))
            ws.cell(row=row_idx, column=2, value=date_str)
            ws.cell(row=row_idx, column=3, value=row.get("location", ""))
            ws.cell(row=row_idx, column=4, value=row.get("sku_code", ""))
            ws.cell(row=row_idx, column=5, value=row.get("asin", ""))
            ws.cell(row=row_idx, column=6, value=row.get("fnsku", ""))
            ws.cell(row=row_idx, column=7, value=row.get("item_name", ""))
            ws.cell(row=row_idx, column=8, value=row.get("mrp"))
            ws.cell(row=row_idx, column=9, value=row.get("sp"))
            ws.cell(row=row_idx, column=10, value=row.get("requested_qty", 0))
            ws.cell(row=row_idx, column=11, value=row.get("packed_qty", 0))
            ws.cell(row=row_idx, column=12, value=row.get("cost_price"))
            ws.cell(row=row_idx, column=13, value=row.get("hsn_code", ""))
            ws.cell(row=row_idx, column=14, value=row.get("gst"))

        col_widths = [18, 12, 16, 14, 14, 14, 35, 8, 8, 13, 11, 11, 12, 6]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_id = re.sub(r"[^\w\-]", "_", shipment_id)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="fba_processing_{safe_id}.xlsx"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error downloading processing data for {shipment_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ─── Processing: Zoho estimate CRUD ──────────────────────────────────────────


class FbaEstimateCreateRequest(BaseModel):
    billing_address_id: str
    shipping_address_id: str
    date: Optional[str] = None


class FbaEstimateLinkRequest(BaseModel):
    estimate_number: str


@router.get("/processing/{shipment_id}/estimate")
async def get_fba_shipment_estimate(
    shipment_id: str, database=Depends(get_database)
):
    """Return linked estimate info for an FBA shipment (or empty object if none)."""
    def _get(db):
        doc = db[FBA_SHIPMENTS_COLLECTION].find_one(
            {"ShipmentId": shipment_id},
            {"zoho_estimate_id": 1, "estimate_number": 1, "_id": 0},
        )
        if not doc:
            return {}
        est_id = doc.get("zoho_estimate_id")
        est_num = doc.get("estimate_number")
        if not est_id and not est_num:
            return {}
        est = None
        if est_id:
            est = db["estimates"].find_one(
                {"estimate_id": est_id},
                {"estimate_id": 1, "estimate_number": 1, "total": 1, "sub_total": 1, "status": 1, "_id": 0},
            )
        if not est and est_num:
            est = db["estimates"].find_one(
                {"estimate_number": est_num},
                {"estimate_id": 1, "estimate_number": 1, "total": 1, "sub_total": 1, "status": 1, "_id": 0},
            )
        return {
            "zoho_estimate_id": est_id,
            "estimate_number": est_num,
            "sub_total": est.get("sub_total") if est else None,
            "total": est.get("total") if est else None,
            "status": est.get("status") if est else None,
        }

    result = await asyncio.to_thread(_get, database)
    return result


@router.post("/processing/{shipment_id}/estimate")
async def create_fba_shipment_estimate(
    shipment_id: str,
    body: FbaEstimateCreateRequest,
    database=Depends(get_database),
):
    """Create a Zoho Books estimate for all processing rows of an FBA shipment."""

    def _create(db):
        doc = db[FBA_SHIPMENTS_COLLECTION].find_one({"ShipmentId": shipment_id})
        if doc and doc.get("zoho_estimate_id"):
            raise ValueError(
                f"Shipment already has estimate {doc.get('estimate_number')} — unlink it first"
            )

        all_rows = _fetch_processing_data(db)
        rows = [r for r in all_rows if r.get("shipment_id") == shipment_id]
        if not rows:
            raise ValueError(f"No processing data found for shipment {shipment_id}")

        skus = [r["sku_code"] for r in rows if r.get("sku_code")]

        sku_to_composite_id: dict[str, str] = {
            c["sku_code"]: c["composite_item_id"]
            for c in db["composite_products"].find(
                {"sku_code": {"$in": skus}},
                {"sku_code": 1, "composite_item_id": 1},
            )
            if c.get("sku_code") and c.get("composite_item_id")
        }

        sku_to_product_id: dict[str, str] = {}
        sku_to_product_status: dict[str, str] = {}
        sku_to_product_name: dict[str, str] = {}
        non_composite = [s for s in skus if s not in sku_to_composite_id]
        for p in db["products"].find(
            {"cf_sku_code": {"$in": non_composite}},
            {"cf_sku_code": 1, "item_id": 1, "status": 1, "name": 1},
        ):
            sku = p.get("cf_sku_code")
            item_id = p.get("item_id")
            if not sku or not item_id:
                continue
            status = p.get("status", "")
            if status == "active" or sku not in sku_to_product_id:
                sku_to_product_id[sku] = item_id
                sku_to_product_status[sku] = status
                sku_to_product_name[sku] = p.get("name", sku)

        line_items: list[dict] = []
        skipped: list[str] = []

        for row in rows:
            sku = row.get("sku_code", "")
            zoho_item_id = sku_to_composite_id.get(sku)
            if not zoho_item_id:
                zoho_item_id = sku_to_product_id.get(sku)
                if not zoho_item_id:
                    skipped.append(sku)
                    continue
                item_status = sku_to_product_status.get(sku, "active")
                if item_status and item_status != "active":
                    skipped.append(f"{sku} (inactive)")
                    continue

            mrp = row.get("mrp") or 0
            gst = row.get("gst")
            rate = round(mrp / (1 + gst / 100), 2) if gst and gst > 0 else round(float(mrp), 2)
            qty = row.get("requested_qty", 0)
            line_items.append({
                "item_id": zoho_item_id,
                "quantity": qty,
                "rate": rate,
                "unit": "pcs",
                "hsn_or_sac": row.get("hsn_code", ""),
            })

        if not line_items:
            raise ValueError(
                f"No line items could be built (missing Zoho item_id for: {', '.join(skipped) or 'all items'})"
            )

        customer = db["customers"].find_one(
            {"contact_name": {"$regex": "ETRADE", "$options": "i"}},
            {"contact_id": 1},
        )
        if not customer:
            raise ValueError("ETRADE customer not found in customers collection")

        token = _get_zoho_books_token()
        zh = {"Authorization": f"Zoho-oauthtoken {token}"}

        now = datetime.now()
        fy_start = now.year if now.month >= 4 else now.year - 1
        fy_str = f"{str(fy_start)[-2:]}-{str(fy_start + 1)[-2:]}"

        list_r = requests.get(
            f"{_ZOHO_BOOKS_BASE}/estimates",
            headers=zh,
            params={
                "organization_id": _ZOHO_ORG_ID,
                "filter_by": "Status.All",
                "per_page": 200,
                "sort_column": "estimate_number",
                "sort_order": "D",
            },
            timeout=30,
        )
        list_r.raise_for_status()
        all_estimates = list_r.json().get("estimates", [])
        fy_estimates = [e for e in all_estimates if f"/{fy_str}/" in e.get("estimate_number", "")]
        if fy_estimates:
            parts = str(fy_estimates[0]["estimate_number"]).split("/")
            last_num = int(parts[-1])
            num_width = len(parts[-1])
            prefix = parts[0]
        else:
            parts = str(all_estimates[0]["estimate_number"]).split("/") if all_estimates else ["EST"]
            last_num = 0
            num_width = len(parts[-1]) if len(parts) > 1 else 4
            prefix = parts[0]

        counter_id = f"estimate_counter_{fy_str}"
        db.counters.update_one({"_id": counter_id}, {"$max": {"seq": last_num}}, upsert=True)
        counter = db.counters.find_one_and_update(
            {"_id": counter_id}, {"$inc": {"seq": 1}}, return_document=True
        )
        new_num = f"{prefix}/{fy_str}/{str(counter['seq']).zfill(num_width)}"

        payload: dict = {
            "estimate_number": new_num,
            "customer_id": customer["contact_id"],
            "reference_number": shipment_id,
            "date": body.date or now.strftime("%Y-%m-%d"),
            "place_of_supply": "MH",
            "billing_address_id": body.billing_address_id,
            "shipping_address_id": body.shipping_address_id,
            "is_inclusive_tax": False,
            "line_items": line_items,
            "branch_id": "3220178000156681877",
            "branch_name": "Amazon (Mumbai Branch)",
            "salesperson_id": "3220178000000692003",
            "dispatch_from_address_id": "3220178000541133001",
        }
        logger.info("FBA estimate payload for %s: %s", shipment_id, payload)
        r = requests.post(
            f"{_ZOHO_BOOKS_BASE}/estimates",
            headers=zh,
            json=payload,
            params={"organization_id": _ZOHO_ORG_ID, "ignore_auto_number_generation": "true"},
            timeout=30,
        )
        logger.info("FBA estimate response %s: %s", r.status_code, r.text)
        r.raise_for_status()
        data = r.json()
        if data.get("code") != 0:
            raise ValueError(f"Zoho error: {data.get('message', 'Unknown error')}")

        estimate = data["estimate"]
        db[FBA_SHIPMENTS_COLLECTION].update_one(
            {"ShipmentId": shipment_id},
            {"$set": {
                "zoho_estimate_id": estimate["estimate_id"],
                "estimate_number": estimate["estimate_number"],
                "updated_at": utcnow(),
            }},
            upsert=True,
        )
        db["estimates"].update_one(
            {"estimate_id": estimate["estimate_id"]},
            {"$set": estimate},
            upsert=True,
        )
        return estimate, skipped

    try:
        estimate, skipped = await asyncio.to_thread(_create, database)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except requests.HTTPError as e:
        body_text = e.response.text if e.response is not None else ""
        raise HTTPException(status_code=502, detail=f"Zoho API error: {e} — {body_text}")
    except Exception as e:
        logger.error(f"Error creating FBA estimate for {shipment_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

    result: dict = {
        "estimate_id": estimate["estimate_id"],
        "estimate_number": estimate["estimate_number"],
        "sub_total": estimate.get("sub_total"),
        "total": estimate.get("total"),
        "status": estimate.get("status"),
    }
    if skipped:
        result["skipped_items"] = skipped
    return JSONResponse(status_code=201, content=result)


@router.put("/processing/{shipment_id}/estimate")
async def link_fba_shipment_estimate(
    shipment_id: str,
    body: FbaEstimateLinkRequest,
    database=Depends(get_database),
):
    """Link an existing Zoho Books estimate number to an FBA shipment."""

    def _link(db):
        est = db["estimates"].find_one(
            {"estimate_number": body.estimate_number},
            {"estimate_id": 1, "estimate_number": 1, "total": 1, "sub_total": 1, "status": 1, "_id": 0},
        )
        if not est:
            token = _get_zoho_books_token()
            r = requests.get(
                f"{_ZOHO_BOOKS_BASE}/estimates",
                headers={"Authorization": f"Zoho-oauthtoken {token}"},
                params={"organization_id": _ZOHO_ORG_ID, "estimate_number": body.estimate_number},
                timeout=30,
            )
            r.raise_for_status()
            estimates = r.json().get("estimates", [])
            if not estimates:
                raise ValueError(f"Estimate {body.estimate_number} not found in Zoho")
            est = estimates[0]
            db["estimates"].update_one({"estimate_id": est["estimate_id"]}, {"$set": est}, upsert=True)

        db[FBA_SHIPMENTS_COLLECTION].update_one(
            {"ShipmentId": shipment_id},
            {"$set": {
                "zoho_estimate_id": est.get("estimate_id", ""),
                "estimate_number": est.get("estimate_number", body.estimate_number),
                "updated_at": utcnow(),
            }},
            upsert=True,
        )
        return est

    try:
        est = await asyncio.to_thread(_link, database)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except requests.HTTPError as e:
        raise HTTPException(status_code=502, detail=f"Zoho API error: {e}")
    except Exception as e:
        logger.error(f"Error linking estimate for {shipment_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

    return {
        "success": True,
        "zoho_estimate_id": est.get("estimate_id"),
        "estimate_number": est.get("estimate_number"),
        "sub_total": est.get("sub_total"),
        "total": est.get("total"),
    }


@router.delete("/processing/{shipment_id}/estimate")
async def unlink_fba_shipment_estimate(
    shipment_id: str, database=Depends(get_database)
):
    """Remove the estimate link from an FBA shipment."""
    def _unlink(db):
        db[FBA_SHIPMENTS_COLLECTION].update_one(
            {"ShipmentId": shipment_id},
            {"$unset": {"zoho_estimate_id": "", "estimate_number": ""}},
        )

    await asyncio.to_thread(_unlink, database)
    return {"success": True}


@router.get("/processing/{shipment_id}/estimate_diff")
async def get_fba_estimate_diff(
    shipment_id: str, database=Depends(get_database)
):
    """Compare FBA processing rows against the linked Zoho estimate line items."""

    def _diff(db):
        doc = db[FBA_SHIPMENTS_COLLECTION].find_one(
            {"ShipmentId": shipment_id},
            {"zoho_estimate_id": 1, "estimate_number": 1},
        )
        if not doc:
            raise ValueError(f"Shipment {shipment_id} not found")
        est_id = doc.get("zoho_estimate_id")
        est_num = doc.get("estimate_number")
        if not est_id and not est_num:
            raise ValueError("No estimate linked to this shipment")

        est = None
        if est_id:
            est = db["estimates"].find_one({"estimate_id": est_id})
        if not est and est_num:
            est = db["estimates"].find_one({"estimate_number": est_num})
        if not est:
            raise ValueError("Estimate not found in local DB — re-link it to refresh")

        all_rows = _fetch_processing_data(db)
        proc_rows = [r for r in all_rows if r.get("shipment_id") == shipment_id]

        # Build item_id → sku_code maps
        est_line_items = est.get("line_items", [])
        est_item_ids = [li["item_id"] for li in est_line_items if li.get("item_id")]

        item_id_to_sku: dict[str, str] = {}
        for p in db["products"].find(
            {"item_id": {"$in": est_item_ids}},
            {"item_id": 1, "cf_sku_code": 1},
        ):
            if p.get("item_id") and p.get("cf_sku_code"):
                item_id_to_sku[p["item_id"]] = p["cf_sku_code"]
        for c in db["composite_products"].find(
            {"composite_item_id": {"$in": est_item_ids}},
            {"composite_item_id": 1, "sku_code": 1},
        ):
            if c.get("composite_item_id") and c.get("sku_code"):
                item_id_to_sku[c["composite_item_id"]] = c["sku_code"]

        est_by_sku: dict[str, dict] = {}
        for li in est_line_items:
            sku = item_id_to_sku.get(li.get("item_id", ""), "")
            if sku:
                est_by_sku[sku] = li

        items = []
        proc_sub_total = 0.0
        proc_total = 0.0
        proc_total_qty = 0
        proc_skus: set[str] = set()

        for row in proc_rows:
            sku = row.get("sku_code", "")
            proc_skus.add(sku)
            req_qty = row.get("requested_qty", 0)
            mrp = row.get("mrp") or 0
            gst = row.get("gst")
            rate_wo_gst = round(mrp / (1 + gst / 100), 2) if gst and gst > 0 else round(float(mrp), 2)
            proc_item_sub = round(rate_wo_gst * req_qty, 2)
            proc_item_total = round(float(mrp) * req_qty, 2)

            proc_sub_total += proc_item_sub
            proc_total += proc_item_total
            proc_total_qty += req_qty

            eli = est_by_sku.get(sku)
            est_qty = eli.get("quantity") if eli else None
            est_rate = eli.get("rate") if eli else None
            est_item_total = round(float(est_rate) * int(est_qty), 2) if est_rate is not None and est_qty is not None else None

            items.append({
                "sku_code": sku,
                "asin": row.get("asin", ""),
                "item_name": row.get("item_name", ""),
                "processing_qty": req_qty,
                "estimate_qty": est_qty,
                "processing_rate": rate_wo_gst,
                "estimate_rate": est_rate,
                "processing_item_total": proc_item_sub,
                "estimate_item_total": est_item_total,
                "in_estimate": eli is not None,
                "qty_match": est_qty == req_qty if est_qty is not None else False,
                "rate_match": abs(float(est_rate or 0) - rate_wo_gst) < 0.05 if est_rate is not None else False,
            })

        only_in_estimate = [
            {
                "sku_code": item_id_to_sku.get(li.get("item_id", ""), li.get("item_id", "")),
                "item_name": li.get("name", ""),
                "estimate_qty": li.get("quantity"),
                "estimate_rate": li.get("rate"),
                "estimate_item_total": round(float(li.get("rate") or 0) * int(li.get("quantity") or 0), 2),
            }
            for li in est_line_items
            if item_id_to_sku.get(li.get("item_id", ""), "") not in proc_skus
        ]

        est_total_qty = sum(int(li.get("quantity") or 0) for li in est_line_items)

        return {
            "estimate_number": est.get("estimate_number"),
            "zoho_estimate_id": est_id,
            "estimate_sub_total": est.get("sub_total"),
            "estimate_total": est.get("total"),
            "processing_sub_total": round(proc_sub_total, 2),
            "processing_total": round(proc_total, 2),
            "processing_item_count": len(proc_rows),
            "estimate_item_count": len(est_line_items),
            "processing_total_qty": proc_total_qty,
            "estimate_total_qty": est_total_qty,
            "items": items,
            "only_in_estimate": only_in_estimate,
        }

    try:
        return await asyncio.to_thread(_diff, database)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"Error computing FBA estimate diff for {shipment_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ─── Page 3: Summary ──────────────────────────────────────────────────────────


@router.get("/summary")
async def get_fba_summary(database=Depends(get_database)):
    try:
        rows = await asyncio.to_thread(_fetch_summary_data, database)
        return {"rows": rows}
    except Exception as e:
        logger.error(f"Error fetching FBA summary data: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/summary/{shipment_id}")
async def update_fba_summary(
    shipment_id: str,
    location: str,
    payload: SummaryUpdate,
    database=Depends(get_database),
):
    try:
        update_doc: dict = {
            "shipment_id": shipment_id,
            "location": location,
            "updated_at": utcnow(),
        }
        if payload.reason_for_short_supply is not None:
            update_doc["reason_for_short_supply"] = payload.reason_for_short_supply
        if payload.appointment_initiated_date is not None:
            update_doc["appointment_initiated_date"] = (
                payload.appointment_initiated_date
            )
        if payload.appointment_date is not None:
            update_doc["appointment_date"] = payload.appointment_date
        if payload.dispatched_date is not None:
            update_doc["dispatched_date"] = payload.dispatched_date
        if payload.status is not None:
            update_doc["status"] = payload.status

        def _upsert(db):
            db[SUMMARY_COLLECTION].update_one(
                {"shipment_id": shipment_id, "location": location},
                {"$set": update_doc},
                upsert=True,
            )

        await asyncio.to_thread(_upsert, database)
        return {"success": True}
    except Exception as e:
        logger.error(f"Error updating FBA summary: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ─── Amazon SP-API: FBA Inbound Plans (2024-03-20) + Shipments (v0) ────────────

SP_MARKETPLACE_ID = "A21TJRUUN4KGV"  # India
FBA_INBOUND_API_VERSION = "2024-03-20"

ALL_SHIPMENT_STATUSES = [
    "CLOSED",
    "WORKING",
    "SHIPPED",
    "IN_TRANSIT",
    "DELIVERED",
    "CHECKED_IN",
    "RECEIVING",
    "CANCELLED",
    "DELETED",
    "ERROR",
]


# ─── Sync helper (called by cron + manual trigger) ────────────────────────────


def _sync_shipments_to_db(db, days: int = 3650) -> dict:
    """
    Fetch FBA shipments (DATE_RANGE across all statuses) + their per-SKU items
    from SP-API v0, then upsert into amazon_fba_shipments.

    - days: how many days back to query (default 90 for nightly cron;
            use 365 for full initial backfill via the manual trigger endpoint).
    - Skips re-fetching items for CLOSED shipments that are already in DB and
      were last synced within 7 days (they won't change).
    - Cross-references amazon_fba_shipment_processing to flag which SP-API
      shipments have a corresponding manual processing upload.

    Returns a summary dict with total/inserted/updated/errors/skipped counts.
    """
    from datetime import timezone as _tz

    after = (datetime.now(_tz.utc) - timedelta(days=days)).strftime(
        "%Y-%m-%dT%H:%M:%SZ"
    )
    before = datetime.now(_tz.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    # ── 1. Collect all shipment headers ──────────────────────────────────────
    shipments: list[dict] = []
    seen_ids: set[str] = set()

    for status in ALL_SHIPMENT_STATUSES:
        params: dict = {
            "MarketplaceId": SP_MARKETPLACE_ID,
            "QueryType": "DATE_RANGE",
            "ShipmentStatusList": status,
            "LastUpdatedAfter": after,
            "LastUpdatedBefore": before,
        }
        while True:
            try:
                result = make_sp_api_request(
                    endpoint="/fba/inbound/v0/shipments",
                    method="GET",
                    params=params,
                )
            except Exception as exc:
                logger.warning(f"SP-API shipments query ({status}) failed: {exc}")
                break

            payload = result.get("payload") or {}
            for s in payload.get("ShipmentData", []):
                sid = s.get("ShipmentId")
                if sid and sid not in seen_ids:
                    seen_ids.add(sid)
                    shipments.append(s)

            nt = payload.get("NextToken")
            if not nt:
                break
            params = {
                "MarketplaceId": SP_MARKETPLACE_ID,
                "QueryType": "NEXT_TOKEN",
                "NextToken": nt,
            }

    logger.info(f"_sync_shipments_to_db: found {len(shipments)} shipments to process")

    # ── 2. Pre-load existing DB docs (for smart skip of stable CLOSED ones) ──
    all_fetched_ids = [s["ShipmentId"] for s in shipments if s.get("ShipmentId")]
    existing_docs: dict[str, dict] = {
        doc["ShipmentId"]: doc
        for doc in db[FBA_SHIPMENTS_COLLECTION].find(
            {"ShipmentId": {"$in": all_fetched_ids}},
            {"ShipmentId": 1, "ShipmentStatus": 1, "last_synced_at": 1, "_id": 0},
        )
    }

    # ── 3. Pre-load which shipment IDs have processing uploads ───────────────
    processing_ids: set[str] = {
        doc["shipment_id"]
        for doc in db[PROCESSING_COLLECTION].find(
            {"shipment_id": {"$in": all_fetched_ids}},
            {"shipment_id": 1, "_id": 0},
        )
        if doc.get("shipment_id")
    }

    # ── 4. For each shipment, fetch items and upsert ──────────────────────────
    inserted = updated = errors = skipped = 0
    now = utcnow()
    stale_threshold = now - timedelta(days=7)

    for shipment in shipments:
        shipment_id = shipment.get("ShipmentId")
        if not shipment_id:
            continue

        current_status = shipment.get("ShipmentStatus", "")
        existing = existing_docs.get(shipment_id)

        # Skip re-fetching items for stable CLOSED shipments synced recently
        if (
            existing
            and current_status == "CLOSED"
            and existing.get("ShipmentStatus") == "CLOSED"
            and existing.get("last_synced_at")
            and existing["last_synced_at"] > stale_threshold
        ):
            # Still update the header fields (status may differ) but keep old items
            db[FBA_SHIPMENTS_COLLECTION].update_one(
                {"ShipmentId": shipment_id},
                {
                    "$set": {
                        **shipment,
                        "has_processing_upload": shipment_id in processing_ids,
                        "last_synced_at": now,
                    }
                },
            )
            skipped += 1
            continue

        # Fetch per-SKU items from SP-API
        try:
            items_result = make_sp_api_request(
                endpoint=f"/fba/inbound/v0/shipments/{shipment_id}/items",
                method="GET",
                params={"MarketplaceId": SP_MARKETPLACE_ID},
            )
            items: list[dict] = (
                (items_result.get("payload") or {}).get("ItemData") or []
            )
        except Exception as exc:
            logger.warning(f"Failed to fetch items for {shipment_id}: {exc}")
            items = []
            errors += 1

        # Annotate items with derived fields
        for item in items:
            shipped = item.get("QuantityShipped", 0) or 0
            received = item.get("QuantityReceived", 0) or 0
            item["QuantityPending"] = max(0, shipped - received)
            item["FullyInwarded"] = shipped > 0 and received >= shipped

        doc = {
            **shipment,
            "items": items,
            "total_skus": len(items),
            "fully_inwarded_count": sum(1 for i in items if i.get("FullyInwarded")),
            "pending_inward_count": sum(
                1 for i in items if (i.get("QuantityPending") or 0) > 0
            ),
            "has_processing_upload": shipment_id in processing_ids,
            "last_synced_at": now,
        }

        res = db[FBA_SHIPMENTS_COLLECTION].update_one(
            {"ShipmentId": shipment_id},
            {"$set": doc},
            upsert=True,
        )
        if res.upserted_id:
            inserted += 1
        else:
            updated += 1

    # ── 5. After any new processing upload, re-link existing DB shipments ─────
    # (marks has_processing_upload=True on any DB shipment that now has a match)
    newly_linked = db[FBA_SHIPMENTS_COLLECTION].update_many(
        {"ShipmentId": {"$in": list(processing_ids)}, "has_processing_upload": {"$ne": True}},
        {"$set": {"has_processing_upload": True}},
    ).modified_count

    # Ensure indexes exist (idempotent)
    db[FBA_SHIPMENTS_COLLECTION].create_index("ShipmentId", unique=True)
    db[FBA_SHIPMENTS_COLLECTION].create_index("ShipmentStatus")
    db[FBA_SHIPMENTS_COLLECTION].create_index("last_synced_at")
    db[FBA_SHIPMENTS_COLLECTION].create_index("has_processing_upload")

    summary = {
        "total": len(shipments),
        "inserted": inserted,
        "updated": updated,
        "skipped_stable": skipped,
        "errors": errors,
        "newly_linked": newly_linked,
    }
    logger.info(f"_sync_shipments_to_db: done — {summary}")
    return summary


def _relink_processing_uploads(db) -> int:
    """
    Called after a processing upload to immediately set has_processing_upload=True
    on any amazon_fba_shipments doc that matches. Fast and cheap.
    Returns count of DB shipments newly linked.
    """
    processing_ids: set[str] = {
        doc["shipment_id"]
        for doc in db[PROCESSING_COLLECTION].find({}, {"shipment_id": 1, "_id": 0})
        if doc.get("shipment_id")
    }
    if not processing_ids:
        return 0
    return db[FBA_SHIPMENTS_COLLECTION].update_many(
        {"ShipmentId": {"$in": list(processing_ids)}, "has_processing_upload": {"$ne": True}},
        {"$set": {"has_processing_upload": True}},
    ).modified_count


# ─── DB-backed read endpoints ─────────────────────────────────────────────────


@router.get("/db/shipments")
async def list_fba_shipments_from_db(
    status: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    database=Depends(get_database),
):
    """
    Return FBA shipments stored in MongoDB (synced by cron).
    Sorted by the date embedded in ShipmentName ("FBA STA (DD/MM/YYYY HH:MM)-FC"),
    newest first. Falls back to ShipmentId descending for names without a date.
    Supports filtering by ShipmentStatus and pagination.
    """
    try:
        match: dict = {}
        if status:
            match["ShipmentStatus"] = status.upper()

        skip = (page - 1) * page_size

        # Aggregation stage: parse the DD/MM/YYYY HH:MM date from ShipmentName
        _extract_date = {
            "$addFields": {
                "_shipment_dt": {
                    "$let": {
                        "vars": {
                            "m": {
                                "$regexFind": {
                                    "input": {"$ifNull": ["$ShipmentName", ""]},
                                    "regex": r"\((\d{2}/\d{2}/\d{4} \d{2}:\d{2})\)",
                                }
                            }
                        },
                        "in": {
                            "$cond": {
                                "if": {"$ne": ["$$m", None]},
                                "then": {
                                    "$dateFromString": {
                                        "dateString": {
                                            "$arrayElemAt": ["$$m.captures", 0]
                                        },
                                        "format": "%d/%m/%Y %H:%M",
                                        "onError": None,
                                        "onNull": None,
                                    }
                                },
                                "else": None,
                            }
                        },
                    }
                }
            }
        }

        def _query(db):
            count_result = list(
                db[FBA_SHIPMENTS_COLLECTION].aggregate(
                    [{"$match": match}, {"$count": "total"}]
                )
            )
            total = count_result[0]["total"] if count_result else 0

            pipeline = [
                {"$match": match},
                _extract_date,
                {"$sort": {"_shipment_dt": -1, "ShipmentId": -1}},
                {"$skip": skip},
                {"$limit": page_size},
                {"$project": {"_id": 0, "_shipment_dt": 0}},
            ]
            docs = list(db[FBA_SHIPMENTS_COLLECTION].aggregate(pipeline))
            return total, docs

        total, docs = await asyncio.to_thread(_query, database)
        return {
            "shipments": docs,
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": -(-total // page_size),  # ceiling division
        }
    except Exception as e:
        logger.error(f"Error listing FBA shipments from DB: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/db/shipments/{shipment_id}")
async def get_fba_shipment_from_db(
    shipment_id: str, database=Depends(get_database)
):
    """
    Return a single FBA shipment (with embedded items) from MongoDB.
    """
    try:
        def _query(db):
            return db[FBA_SHIPMENTS_COLLECTION].find_one(
                {"ShipmentId": shipment_id}, {"_id": 0}
            )

        doc = await asyncio.to_thread(_query, database)
        if not doc:
            raise HTTPException(
                status_code=404,
                detail=f"Shipment {shipment_id} not found in DB. Run /sp/shipments/sync first.",
            )
        return doc
    except HTTPException:
        raise
    except Exception as e:
        logger.error(
            f"Error fetching shipment {shipment_id} from DB: {e}", exc_info=True
        )
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/sp/shipments/sync")
async def trigger_fba_shipments_sync(
    days: int = 365,
    database=Depends(get_database),
):
    """
    Manually trigger a full SP-API sync of FBA shipments + items into MongoDB.

    - days=365 (default): full historical backfill — use this for the first run.
    - days=90: same as the nightly cron window.

    CLOSED shipments already synced within 7 days won't re-fetch items (skipped).
    Returns insert/update/skipped/error counts.
    """
    try:
        result = await asyncio.to_thread(_sync_shipments_to_db, database, days)
        return {"success": True, **result}
    except Exception as e:
        logger.error(f"Error during manual FBA shipments sync: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sp/inbound-plans")
async def list_inbound_plans(
    page_size: Optional[int] = 20,
    pagination_token: Optional[str] = None,
    status: Optional[str] = None,
    sort_by: Optional[str] = "LAST_UPDATED_TIME",
    sort_order: Optional[str] = "DESC",
):
    """
    List inbound plans via SP-API 2024-03-20.
    Returns plan IDs, status, and whether shipments have been confirmed.
    """
    try:
        params: dict = {"sortBy": sort_by, "sortOrder": sort_order}
        if page_size is not None:
            params["pageSize"] = page_size
        if pagination_token:
            params["paginationToken"] = pagination_token
        if status:
            params["status"] = status

        def _call():
            return make_sp_api_request(
                endpoint=f"/inbound/fba/{FBA_INBOUND_API_VERSION}/inboundPlans",
                method="GET",
                params=params,
            )

        result = await asyncio.to_thread(_call)
        return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error calling listInboundPlans: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sp/inbound-plans/{inbound_plan_id}")
async def get_inbound_plan(inbound_plan_id: str):
    """
    Get full detail for a single inbound plan including shipments array.
    shipments[] will be empty if placement was never confirmed in Seller Central.
    """
    try:
        def _call():
            return make_sp_api_request(
                endpoint=f"/inbound/fba/{FBA_INBOUND_API_VERSION}/inboundPlans/{inbound_plan_id}",
                method="GET",
            )

        result = await asyncio.to_thread(_call)
        return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error calling getPlan {inbound_plan_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sp/inbound-plans/{inbound_plan_id}/items")
async def get_inbound_plan_items(inbound_plan_id: str):
    """
    Get items (SKUs + quantities) for an inbound plan.
    Returns msku, fnsku, asin, quantity per item.
    """
    try:
        def _call():
            return make_sp_api_request(
                endpoint=f"/inbound/fba/{FBA_INBOUND_API_VERSION}/inboundPlans/{inbound_plan_id}/items",
                method="GET",
            )

        result = await asyncio.to_thread(_call)
        return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error calling getPlanItems {inbound_plan_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sp/shipments")
async def list_fba_shipments(
    days: Optional[int] = 365,
    next_token: Optional[str] = None,
):
    """
    List all FBA STA shipments via SP-API v0 using DATE_RANGE query.

    - days: how many days back to look (default 365)
    - next_token: pagination token from a previous response

    STA shipments only appear via DATE_RANGE + CLOSED status — they are not
    returned by simple ShipmentStatusList queries. Active statuses
    (WORKING/SHIPPED/RECEIVING) are also checked and merged in.

    Returns shipments sorted newest first with ShipmentId, ShipmentName,
    ShipmentStatus, DestinationFulfillmentCenterId, ShipFromAddress.
    """
    from datetime import timezone

    try:
        after = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%dT%H:%M:%SZ")
        before = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

        all_statuses = ALL_SHIPMENT_STATUSES

        def _fetch_all():
            shipments = []
            seen_ids = set()

            for status in all_statuses:
                params = {
                    "MarketplaceId": SP_MARKETPLACE_ID,
                    "QueryType": "DATE_RANGE",
                    "ShipmentStatusList": status,
                    "LastUpdatedAfter": after,
                    "LastUpdatedBefore": before,
                }
                if next_token:
                    params["NextToken"] = next_token
                    params["QueryType"] = "NEXT_TOKEN"

                while True:
                    result = make_sp_api_request(
                        endpoint="/fba/inbound/v0/shipments",
                        method="GET",
                        params=params,
                    )
                    payload = result.get("payload") or {}
                    for s in payload.get("ShipmentData", []):
                        sid = s.get("ShipmentId")
                        if sid and sid not in seen_ids:
                            seen_ids.add(sid)
                            shipments.append(s)
                    nt = payload.get("NextToken")
                    if not nt:
                        break
                    params = {
                        "MarketplaceId": SP_MARKETPLACE_ID,
                        "QueryType": "NEXT_TOKEN",
                        "NextToken": nt,
                    }

            return shipments

        shipments = await asyncio.to_thread(_fetch_all)

        # Sort newest first using ShipmentName date (fallback to ID)
        shipments.sort(key=lambda s: s.get("ShipmentName", ""), reverse=True)

        return {"shipments": shipments, "count": len(shipments)}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error calling listShipments v0: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sp/shipments/{shipment_id}/items")
async def get_fba_shipment_items(shipment_id: str):
    """
    Get per-SKU inward status for a shipment via SP-API v0.

    Key fields per item:
      - SellerSKU           — your SKU code
      - FulfillmentNetworkSKU — FNSKU
      - QuantityShipped     — units you sent to Amazon
      - QuantityReceived    — units Amazon has actually inwarded so far
      - QuantityInCase      — case-pack qty (0 if not case-packed)

    If QuantityShipped > QuantityReceived, those units are still pending inward.
    """
    try:
        def _call():
            return make_sp_api_request(
                endpoint=f"/fba/inbound/v0/shipments/{shipment_id}/items",
                method="GET",
                params={"MarketplaceId": SP_MARKETPLACE_ID},
            )

        result = await asyncio.to_thread(_call)
        items = (result.get("payload") or {}).get("ItemData", [])

        # Annotate each item with pending qty for easy reading
        for item in items:
            shipped = item.get("QuantityShipped", 0) or 0
            received = item.get("QuantityReceived", 0) or 0
            item["QuantityPending"] = max(0, shipped - received)
            item["FullyInwarded"] = (shipped > 0 and received >= shipped)

        return {
            "shipment_id": shipment_id,
            "items": items,
            "total_skus": len(items),
            "fully_inwarded": sum(1 for i in items if i["FullyInwarded"]),
            "pending_inward": sum(1 for i in items if i["QuantityPending"] > 0),
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching items for shipment {shipment_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
