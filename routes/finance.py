"""Cash-aware purchase planning.

Two reports that sit between finance and purchase:

* **Working Capital** — how much cash is actually free to spend on inventory, and
  when. Cash on hand + expected collections − committed outflows, bucketed by
  horizon.
* **Brand Order Plan** — the master report's per-brand demand priced at COGS,
  ranked by GMROI, then funded from the working-capital envelope so each brand
  comes out as ORDER FULL / ORDER PARTIAL / DEFER.

Bank data (`bank_accounts`, `bank_transactions`) is written by `order_form_backend`
crons/webhooks into the same database, so it is read directly here.
"""

from fastapi import APIRouter, HTTPException, Depends, Query, status
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel, Field
from bson import ObjectId
from datetime import datetime, timedelta
from typing import Dict, List, Optional
import asyncio
import bisect
import io
import logging
import re as _re

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..database import get_database
from ..services.master_report import _generate_master_report_data

logger = logging.getLogger(__name__)
router = APIRouter()

BANK_ACCOUNTS = "bank_accounts"
BANK_TRANSACTIONS = "bank_transactions"
INVOICES = "invoices"
BILLS = "bills"
PURCHASE_ORDERS = "purchase_orders"
BRAND_LOGISTICS = "brand_logistics"
BRAND_ORDERS = "brand_orders"
NEW_BRAND_RESERVES = "finance_new_brand_reserves"

# Ceiling on how much stock a single order may buy, expressed as days of cover.
# The master report sizes orders from demand alone; without a ceiling a brand
# whose lookback DRR spikes can ask for a year of stock in one go.
DEFAULT_MAX_COVER_DAYS = 120

# A brand needs at least this many days of selling before its run rate is
# trusted enough to plan a purchase from. Jolly Pawps sold 884 units in its
# first two days: divided by two days in stock that is a DRR of 442/day, which
# reads as zero cover (CRITICAL, top of the funding queue) and would authorise
# roughly 53,000 units against the 120-day cap. Two days of a launch is not a
# run rate.
DEFAULT_MIN_SELLING_DAYS = 30

# Average calendar month, for turning a daily run rate into months of cover.
DAYS_PER_MONTH = 30.44

# Zoho keeps voided invoices at their full original balance, and drafts are not
# yet receivable. Including either overstates AR by crores.
AR_STATUSES = ("sent", "overdue", "partially_paid")
AP_STATUSES = ("open", "overdue", "partially_paid")

# A credit card's balance is money owed, not money held.
LIABILITY_ACCOUNT_TYPES = ("credit_card",)
CASH_ACCOUNT_TYPES = ("bank", "cash", "payment_clearing", "other_current_asset")

# Internal movement between our own accounts — counting it inflates both sides
# of any in-vs-out view.
INTERNAL_TXN_TYPES = ("transfer_fund",)

AGEING_BUCKETS = [
    ("overdue", None, 0),
    ("0_30", 0, 30),
    ("31_60", 31, 60),
    ("61_90", 61, 90),
    ("90_plus", 91, None),
]

HORIZONS = (30, 60, 90)

# Last-resort fallbacks, used only when the live rate lookup fails and the user
# has not supplied a rate. Kept deliberately conservative.
DEFAULT_FX = {"USD": 86.0, "CNY": 12.0, "EUR": 93.0, "GBP": 110.0, "INR": 1.0}

FX_API_URL = "https://api.frankfurter.dev/v1/latest"
FX_CURRENCIES = ("USD", "CNY", "EUR", "GBP")
FX_CACHE_TTL_SECONDS = 3600

# Module-level cache so a page refresh (and the plan endpoint, which needs the
# same rates) doesn't hit the FX API on every request.
_fx_cache: Dict[str, object] = {"fetched_at": None, "rates": None, "source": None}


async def _fetch_live_fx() -> Dict:
    """Current INR rates for the currencies our POs are raised in.

    Returns ``{"rates": {...}, "source": "live"|"cache"|"fallback", "as_of": ...}``.
    A failure here must never break the report — purchase still needs the plan
    even if the FX host is down, so we degrade to the cached or default rate and
    say so in the payload.
    """
    now = datetime.now()
    fetched_at = _fx_cache.get("fetched_at")
    if (
        isinstance(fetched_at, datetime)
        and _fx_cache.get("rates")
        and (now - fetched_at).total_seconds() < FX_CACHE_TTL_SECONDS
    ):
        return {
            "rates": dict(_fx_cache["rates"]),  # type: ignore[arg-type]
            "source": "cache",
            "as_of": fetched_at.isoformat(),
        }

    try:
        import httpx

        async with httpx.AsyncClient(timeout=8.0, follow_redirects=True) as client:
            resp = await client.get(
                FX_API_URL,
                params={"base": "INR", "symbols": ",".join(FX_CURRENCIES)},
            )
            resp.raise_for_status()
            body = resp.json()

        # The API quotes INR→foreign; we need foreign→INR, so invert.
        rates = {"INR": 1.0}
        for ccy, per_inr in (body.get("rates") or {}).items():
            if per_inr:
                rates[ccy] = round(1.0 / float(per_inr), 4)
        rates["RMB"] = rates.get("CNY", DEFAULT_FX["CNY"])

        if not rates.get("USD"):
            raise ValueError("FX response missing USD")

        _fx_cache["fetched_at"] = now
        _fx_cache["rates"] = rates
        _fx_cache["source"] = "live"
        return {
            "rates": dict(rates),
            "source": "live",
            "as_of": body.get("date") or now.date().isoformat(),
        }
    except Exception as e:
        logger.warning(f"Live FX lookup failed, falling back: {e}")
        cached = _fx_cache.get("rates")
        if cached:
            return {
                "rates": dict(cached),  # type: ignore[arg-type]
                "source": "stale_cache",
                "as_of": fetched_at.isoformat() if isinstance(fetched_at, datetime) else None,
            }
        return {"rates": dict(DEFAULT_FX), "source": "fallback", "as_of": None}


# ─── Helpers ────────────────────────────────────────────────────────────────────

def _f(val, default: float = 0.0) -> float:
    """Zoho returns numerics as strings often enough that this is unavoidable."""
    if val is None or val == "":
        return default
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return default


def _parse_date(val) -> Optional[datetime]:
    """`due_date` is a 'YYYY-MM-DD' string on invoices but a datetime on bills."""
    if not val:
        return None
    if isinstance(val, datetime):
        return val.replace(hour=0, minute=0, second=0, microsecond=0)
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:19] if " " in s else s, fmt)
        except ValueError:
            continue
    return None


def _bucket_for(due: Optional[datetime], today: datetime) -> str:
    """Which ageing bucket a due date falls into. No date = treat as due now,
    which is the conservative choice for a payable and the honest one for a
    receivable we cannot schedule."""
    if due is None:
        return "overdue"
    days = (due - today).days
    if days < 0:
        return "overdue"
    for name, lo, hi in AGEING_BUCKETS[1:]:
        if (lo is None or days >= lo) and (hi is None or days <= hi):
            return name
    return "90_plus"


def _empty_buckets() -> Dict[str, float]:
    return {name: 0.0 for name, _, _ in AGEING_BUCKETS}


def _within(buckets: Dict[str, float], horizon_days: int) -> float:
    """Sum of everything landing on or before `horizon_days`. Overdue is
    included in every horizon — it is already due."""
    total = buckets.get("overdue", 0.0)
    if horizon_days >= 30:
        total += buckets.get("0_30", 0.0)
    if horizon_days >= 60:
        total += buckets.get("31_60", 0.0)
    if horizon_days >= 90:
        total += buckets.get("61_90", 0.0)
    return round(total, 2)


async def _resolve_fx(
    usd_inr: Optional[float], cny_inr: Optional[float]
) -> Dict:
    """Build the currency→INR map, preferring a user-supplied rate over the live one.

    POs are raised in USD and CNY but every figure in these reports is rupees, so
    a rate is always needed. Finance often wants to plan at the rate they expect
    to actually pay at rather than today's spot, hence the override.
    """
    live = await _fetch_live_fx()
    rates = dict(DEFAULT_FX)
    rates.update(live["rates"])

    overrides = {}
    if usd_inr is not None:
        rates["USD"] = usd_inr
        overrides["USD"] = usd_inr
    if cny_inr is not None:
        rates["CNY"] = cny_inr
        rates["RMB"] = cny_inr
        overrides["CNY"] = cny_inr

    return {
        "rates": rates,
        "meta": {
            "source": "manual" if len(overrides) == 2 else live["source"],
            "as_of": live["as_of"],
            "live_rates": {c: live["rates"].get(c) for c in FX_CURRENCIES},
            "overrides": overrides,
            "applied": {"USD": rates.get("USD"), "CNY": rates.get("CNY")},
        },
    }


@router.get("/fx-rates")
async def get_fx_rates():
    """Current foreign-currency→INR rates, for pre-filling the rate inputs."""
    live = await _fetch_live_fx()
    return JSONResponse(status_code=status.HTTP_200_OK, content=live)


# ─── New-brand reserves ─────────────────────────────────────────────────────────
#
# Brands we are about to take on have no SKUs, no sales and no COGS, so the
# master report cannot see them at all — yet the cash to launch them is spent out
# of the same envelope as everything else. Without a provision the plan hands the
# entire envelope to existing brands and a new-brand launch silently overspends.
# A reserve is set aside off the top and the ranked allocation runs on what is
# left, so both decisions are visible in one place.


class NewBrandReserve(BaseModel):
    brand: str = Field(..., min_length=1, description="Working name of the brand")
    amount: float = Field(..., ge=0, description="Cash to set aside, in INR")
    notes: Optional[str] = None
    expected_order_date: Optional[str] = Field(
        None, description="When the first order is expected, YYYY-MM-DD"
    )
    is_active: bool = True


class NewBrandReserveUpdate(BaseModel):
    brand: Optional[str] = Field(None, min_length=1)
    amount: Optional[float] = Field(None, ge=0)
    notes: Optional[str] = None
    expected_order_date: Optional[str] = None
    is_active: Optional[bool] = None


def _serialise_reserve(doc: Dict) -> Dict:
    return {
        "id": str(doc["_id"]),
        "brand": doc.get("brand", "") or "",
        "amount": round(_f(doc.get("amount")), 2),
        "notes": doc.get("notes") or "",
        "expected_order_date": doc.get("expected_order_date") or "",
        "is_active": bool(doc.get("is_active", True)),
        "created_at": (doc["created_at"].isoformat() if isinstance(doc.get("created_at"), datetime) else None),
        "updated_at": (doc["updated_at"].isoformat() if isinstance(doc.get("updated_at"), datetime) else None),
    }


def _new_brand_reserves_sync(db, active_only: bool = False) -> List[Dict]:
    q = {"is_active": True} if active_only else {}
    docs = list(
        db.get_collection(NEW_BRAND_RESERVES)
        .find(q)
        .sort([("is_active", -1), ("amount", -1), ("_id", -1)])
    )
    return [_serialise_reserve(d) for d in docs]


@router.get("/new-brand-reserves")
async def list_new_brand_reserves(db=Depends(get_database)):
    """Cash set aside for brands that do not exist in the system yet."""
    rows = await asyncio.to_thread(_new_brand_reserves_sync, db, False)
    return JSONResponse(status_code=status.HTTP_200_OK, content={
        "reserves": rows,
        "total_reserved": round(sum(r["amount"] for r in rows if r["is_active"]), 2),
        "count": len(rows),
    })


@router.post("/new-brand-reserves")
async def create_new_brand_reserve(body: NewBrandReserve, db=Depends(get_database)):
    now = datetime.now()
    doc = body.model_dump()
    doc.update({"created_at": now, "updated_at": now})

    def _insert():
        res = db.get_collection(NEW_BRAND_RESERVES).insert_one(doc)
        return db.get_collection(NEW_BRAND_RESERVES).find_one({"_id": res.inserted_id})

    saved = await asyncio.to_thread(_insert)
    return JSONResponse(status_code=status.HTTP_201_CREATED, content=_serialise_reserve(saved))


@router.patch("/new-brand-reserves/{reserve_id}")
async def update_new_brand_reserve(
    reserve_id: str, body: NewBrandReserveUpdate, db=Depends(get_database)
):
    try:
        oid = ObjectId(reserve_id)
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid id")

    changes = {k: v for k, v in body.model_dump(exclude_unset=True).items() if v is not None}
    # is_active=False is a legitimate value that the None filter above would drop.
    if "is_active" in body.model_dump(exclude_unset=True):
        changes["is_active"] = bool(body.is_active)
    if not changes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Nothing to update")
    changes["updated_at"] = datetime.now()

    def _update():
        db.get_collection(NEW_BRAND_RESERVES).update_one({"_id": oid}, {"$set": changes})
        return db.get_collection(NEW_BRAND_RESERVES).find_one({"_id": oid})

    saved = await asyncio.to_thread(_update)
    if not saved:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Reserve not found")
    return JSONResponse(status_code=status.HTTP_200_OK, content=_serialise_reserve(saved))


@router.delete("/new-brand-reserves/{reserve_id}")
async def delete_new_brand_reserve(reserve_id: str, db=Depends(get_database)):
    try:
        oid = ObjectId(reserve_id)
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid id")

    res = await asyncio.to_thread(
        lambda: db.get_collection(NEW_BRAND_RESERVES).delete_one({"_id": oid})
    )
    if not res.deleted_count:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Reserve not found")
    return JSONResponse(status_code=status.HTTP_200_OK, content={"deleted": True})


# ─── Per-brand planning settings ────────────────────────────────────────────────
#
# How much cover an order may buy is a brand decision, not a global one: a brand
# on a 96-day lead time out of China cannot be held to the same ceiling as one
# restocked locally in three weeks. The cap lives on `brand_logistics` next to
# `lead_time` — the collection the plan already reads — so there is one place
# per brand and no second source of truth to drift.


class BrandCoverSetting(BaseModel):
    max_cover_days: Optional[float] = Field(
        None, ge=0, description="Days of cover an order may buy. Null clears the override."
    )


def _brand_settings_sync(db) -> List[Dict]:
    """Every brand we could plan for, with its saved cap where it has one."""
    logistics = {
        (d.get("brand", "") or "").strip().lower(): d
        for d in db.get_collection(BRAND_LOGISTICS).find({}, {"_id": 0})
    }
    brands = {
        b.strip() for b in db.get_collection("products").distinct("brand") if b and b.strip()
    }
    brands.update((d.get("brand", "") or "").strip() for d in logistics.values())

    rows = []
    for brand in sorted(b for b in brands if b):
        s = logistics.get(brand.lower(), {})
        cap = s.get("max_cover_days")
        rows.append({
            "brand": brand,
            "lead_time": _f(s.get("lead_time"), 60.0),
            "max_cover_days": _f(cap) if cap is not None else None,
            "has_override": cap is not None,
        })
    return rows


@router.get("/brand-settings")
async def list_brand_settings(db=Depends(get_database)):
    """Per-brand lead time and order cover cap."""
    rows = await asyncio.to_thread(_brand_settings_sync, db)
    return JSONResponse(status_code=status.HTTP_200_OK, content={
        "brands": rows,
        "default_max_cover_days": DEFAULT_MAX_COVER_DAYS,
        "with_override": sum(1 for r in rows if r["has_override"]),
    })


@router.put("/brand-settings/{brand}")
async def set_brand_cover_days(
    brand: str, body: BrandCoverSetting, db=Depends(get_database)
):
    """Save (or clear) a brand's order cover cap.

    Upserts by brand so a brand with no logistics row yet can still carry a cap;
    `$set`/`$unset` of this one field leaves `lead_time` and the safety-day
    settings owned by the Brand Logistics page untouched.
    """
    name = brand.strip()
    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Brand is required")

    update = (
        {"$set": {"brand": name, "max_cover_days": body.max_cover_days}}
        if body.max_cover_days is not None
        else {"$set": {"brand": name}, "$unset": {"max_cover_days": ""}}
    )

    def _save():
        db.get_collection(BRAND_LOGISTICS).update_one(
            {"brand": {"$regex": f"^{_re.escape(name)}$", "$options": "i"}},
            update,
            upsert=True,
        )
        return db.get_collection(BRAND_LOGISTICS).find_one({"brand": name}, {"_id": 0})

    saved = await asyncio.to_thread(_save)
    return JSONResponse(status_code=status.HTTP_200_OK, content={
        "brand": name,
        "max_cover_days": (saved or {}).get("max_cover_days"),
        "lead_time": _f((saved or {}).get("lead_time"), 60.0),
    })


# ─── Cash position ──────────────────────────────────────────────────────────────

def _cash_position_sync(db) -> Dict:
    """Net cash from `bank_accounts`, split into holdings and liabilities.

    Zoho carries two different numbers per account and they are not the same
    thing:

    * ``balance`` — the **book** balance, i.e. what has been recorded in Books.
    * ``bank_balance`` — what the **bank feed** says is actually in the account.

    For "how much can we spend", the bank feed is the truth. Yes Bank, for
    example, holds ₹58.5L in the feed while Books shows ₹65 because the
    transactions have not been categorised yet. Using ``balance`` alone hides
    real money.

    Only fed accounts populate ``bank_balance`` (the rest sit at 0), so we take
    it per account when it is non-zero and fall back to the book balance
    otherwise — never zeroing out an unfed account like ICICI.

    Accounts flagged `deleted_in_zoho` are excluded — the cron flags rather than
    deletes them so their transactions aren't orphaned, but they are not live cash.
    """
    accounts = list(
        db.get_collection(BANK_ACCOUNTS).find(
            {"deleted_in_zoho": {"$ne": True}},
            {
                "account_id": 1, "account_name": 1, "account_type": 1,
                "balance": 1, "bank_balance": 1, "uncategorized_transactions": 1,
                "currency_code": 1, "is_active": 1, "_id": 0,
            },
        )
    )

    holdings, liabilities, unreconciled_total = 0.0, 0.0, 0.0
    rows = []
    for a in accounts:
        book = _f(a.get("balance"))
        feed = _f(a.get("bank_balance"))
        has_feed = feed != 0
        effective = feed if has_feed else book
        unreconciled = effective - book if has_feed else 0.0

        atype = a.get("account_type", "") or ""
        is_liability = atype in LIABILITY_ACCOUNT_TYPES
        if is_liability:
            liabilities += effective
        elif atype in CASH_ACCOUNT_TYPES:
            holdings += effective
        unreconciled_total += unreconciled

        rows.append({
            "account_id": a.get("account_id", ""),
            "account_name": a.get("account_name", "") or "",
            "account_type": atype,
            "book_balance": round(book, 2),
            "bank_balance": round(feed, 2),
            "has_feed": has_feed,
            "balance": round(effective, 2),
            "unreconciled": round(unreconciled, 2),
            "uncategorized_transactions": int(_f(a.get("uncategorized_transactions"))),
            "is_liability": is_liability,
            # What this account contributes to net cash.
            "net_contribution": round(-effective if is_liability else effective, 2),
            "currency_code": a.get("currency_code", "INR") or "INR",
            "is_active": bool(a.get("is_active", True)),
        })

    rows.sort(key=lambda r: (r["is_liability"], -abs(r["balance"])))

    return {
        "accounts": rows,
        "holdings": round(holdings, 2),
        "liabilities": round(liabilities, 2),
        "net_cash": round(holdings - liabilities, 2),
        "book_net_cash": round(
            sum(r["book_balance"] for r in rows if not r["is_liability"]
                and r["account_type"] in CASH_ACCOUNT_TYPES)
            - sum(r["book_balance"] for r in rows if r["is_liability"]), 2
        ),
        "unreconciled": round(unreconciled_total, 2),
    }


OPEX_BREAKDOWN_LIMIT = 15

# Spend already accounted for elsewhere in the working-capital picture, so
# counting it as a running cost would charge it twice. Custom duty is part of
# the landed cost of an inbound order and is shown per PO in the open
# purchase-order commitment.
OPEX_EXCLUDED_ACCOUNTS = ("Custom Duty",)


def _monthly_opex_sync(db, start: datetime, end: datetime) -> Dict:
    """Average monthly operating outflow, measured over an explicit window.

    Uses `expense` transactions only. `signed_amount` is negative for cash out
    (a credit on the bank account), so the sum is negated to get a positive
    burn figure. Transfers are excluded as internal movement.

    Also returns what the money went on. `offset_account_name` is the account
    the spend was booked against — salaries, loan EMIs, reimbursements — which
    is the only field on the transaction that says *what* it was. It is granular
    (127 distinct names over a 90-day window, largely one per payee), so the
    biggest are listed individually and the tail is rolled into "Other".

    `OPEX_EXCLUDED_ACCOUNTS` is left out entirely: that spend is already charged
    against the open purchase-order commitment.
    """
    days = max(1, (end - start).days)
    match = {
        "transaction_date": {"$gte": start, "$lte": end},
        "transaction_type": {"$in": ["expense", "tds_payment"]},
        "offset_account_name": {"$nin": list(OPEX_EXCLUDED_ACCOUNTS)},
    }

    rows = list(db.get_collection(BANK_TRANSACTIONS).aggregate([
        {"$match": match},
        {"$group": {
            "_id": {"$ifNull": ["$offset_account_name", ""]},
            "out": {"$sum": "$signed_amount"},
            "n": {"$sum": 1},
        }},
        # signed_amount is negative for cash out, so ascending = biggest spend.
        {"$sort": {"out": 1}},
    ]))

    net_out = -sum(r["out"] for r in rows)
    count = sum(r["n"] for r in rows)
    total = max(0.0, net_out)
    per_month = (30.0 / days) if days else 0.0

    breakdown = []
    for r in rows[:OPEX_BREAKDOWN_LIMIT]:
        amount = -r["out"]
        if amount <= 0:
            continue
        breakdown.append({
            "category": (r["_id"] or "").strip() or "Uncategorised",
            "amount": round(amount, 2),
            "monthly": round(amount * per_month, 2),
            "pct": round(amount / total * 100, 1) if total > 0 else 0.0,
            "transactions": r["n"],
        })

    tail = rows[OPEX_BREAKDOWN_LIMIT:]
    tail_amount = -sum(r["out"] for r in tail)
    if tail_amount > 0:
        breakdown.append({
            "category": f"Other ({len(tail)} categories)",
            "amount": round(tail_amount, 2),
            "monthly": round(tail_amount * per_month, 2),
            "pct": round(tail_amount / total * 100, 1) if total > 0 else 0.0,
            "transactions": sum(r["n"] for r in tail),
            "is_other": True,
        })

    return {
        "monthly_opex": round(total / days * 30, 2),
        "window_days": days,
        "window_start": start.date().isoformat(),
        "window_end": end.date().isoformat(),
        "spent_in_window": round(total, 2),
        "transactions_sampled": count,
        "breakdown": breakdown,
        "category_count": len(rows),
        "excluded_accounts": list(OPEX_EXCLUDED_ACCOUNTS),
    }


def _collection_run_rate_sync(db, start: datetime, end: datetime) -> Dict:
    """How fast money actually comes in, measured over an explicit window.

    The invoice ledger tells us exactly *how much* is owed, but not how much of
    it will land inside a planning horizon — plenty of the overdue book has been
    overdue for months. Rather than asking someone to guess a percentage, we
    measure the rate at which customer payments have genuinely been hitting the
    bank and project that forward. Widening or narrowing the window is how you
    plan against a seasonal stretch rather than the trailing quarter.
    """
    days = max(1, (end - start).days)

    pipeline = [
        {"$match": {
            "transaction_date": {"$gte": start, "$lte": end},
            "transaction_type": {"$in": ["customer_payment", "payment_refund"]},
        }},
        {"$group": {"_id": None, "net": {"$sum": "$signed_amount"}, "n": {"$sum": 1}}},
    ]
    res = list(db.get_collection(BANK_TRANSACTIONS).aggregate(pipeline))
    net_in = res[0]["net"] if res else 0.0
    count = res[0]["n"] if res else 0

    return {
        "daily_collection_rate": round(max(0.0, net_in) / days, 2),
        "window_days": days,
        "window_start": start.date().isoformat(),
        "window_end": end.date().isoformat(),
        "collected_in_window": round(net_in, 2),
        "transactions_sampled": count,
    }


# How far back to look when learning how a customer pays, and how many settled
# invoices they need before we trust their own curve over the house average.
AR_BEHAVIOUR_LOOKBACK_DAYS = 365
AR_BEHAVIOUR_MIN_INVOICES = 5


def _lateness_curve(points: List[tuple]) -> Optional[tuple]:
    """Value-weighted CDF of how late invoices get paid.

    `points` is [(days_late, invoice_value)]. Returns `(days[], cumulative[])`
    for bisect. Weighted by value, not count: a customer who pays fifty ₹5k
    invoices on time and one ₹20L invoice six months late is not a prompt payer
    for planning purposes.
    """
    if not points:
        return None
    points.sort()
    total = sum(v for _, v in points) or 1.0
    days, cum, running = [], [], 0.0
    for d, v in points:
        running += v
        days.append(d)
        cum.append(running / total)
    return days, cum


def _curve_at(curve: Optional[tuple], day: int) -> Optional[float]:
    """Fraction of value historically settled by `day` days after the due date."""
    if not curve:
        return None
    days, cum = curve
    i = bisect.bisect_right(days, day) - 1
    return cum[i] if i >= 0 else 0.0


_indexes_ensured = False


def _ensure_indexes(db) -> None:
    """Index the settled-invoice scan. Once per process; idempotent in Mongo.

    Without it the behaviour model is a 67k-document collection scan and takes
    ~10.7s; with it, 1.2s.
    """
    global _indexes_ensured
    if _indexes_ensured:
        return
    try:
        db.get_collection(INVOICES).create_index(
            [("status", 1), ("last_payment_date", 1)],
            name="fin_status_lastpay", background=True,
        )
    except Exception as e:
        logger.warning(f"Could not ensure finance indexes: {e}")
    _indexes_ensured = True


def _ar_payment_curves_sync(db, as_of: datetime) -> tuple:
    """Learn how each customer actually pays, from invoices they have settled.

    Returns `(per_customer_curves, house_curve)`.
    """
    _ensure_indexes(db)
    cutoff = (as_of - timedelta(days=AR_BEHAVIOUR_LOOKBACK_DAYS)).strftime("%Y-%m-%d")
    cursor = db.get_collection(INVOICES).find(
        {
            "status": "paid",
            # due_date/last_payment_date are 'YYYY-MM-DD' strings, so a string
            # comparison is a correct date comparison and stays indexable.
            "last_payment_date": {"$gte": cutoff},
        },
        {"customer_id": 1, "due_date": 1, "last_payment_date": 1, "total": 1, "_id": 0},
    )

    by_customer: Dict[str, List[tuple]] = {}
    house: List[tuple] = []
    for inv in cursor:
        due = _parse_date(inv.get("due_date"))
        paid_on = _parse_date(inv.get("last_payment_date"))
        if not due or not paid_on:
            continue
        value = _f(inv.get("total"))
        if value <= 0:
            continue
        point = ((paid_on - due).days, value)
        house.append(point)
        by_customer.setdefault(inv.get("customer_id") or "", []).append(point)

    curves = {
        cid: _lateness_curve(pts)
        for cid, pts in by_customer.items()
        if cid and len(pts) >= AR_BEHAVIOUR_MIN_INVOICES
    }
    return curves, _lateness_curve(house)


def _expected_collections_sync(db, as_of: datetime, horizons=HORIZONS) -> Dict:
    """What will realistically be collected per horizon, customer by customer.

    Treating every invoice due inside the window as collectable overstates
    badly — back-tested against the 90 days to 2026-08-06 it predicted ₹286.6L
    against ₹195.0L actually collected. The problem is that *who* owes it
    matters more than when it is due: walk-in trade settles the same day, while
    Blink Commerce's median invoice is paid **267 days** after its due date.

    So each open invoice is scored against that customer's own history,
    conditioned on the fact that it has not been paid yet:

        P(paid within horizon | still unpaid today)
            = (F(d + horizon) − F(d)) / (1 − F(d))

    where `d` is how long the invoice is already past due and `F` is the
    customer's value-weighted lateness curve. An invoice that has already
    outlived everything that customer has ever done scores zero rather than
    certainty — the conservative reading, and the honest one for a debt that is
    behaving unlike every other debt they have settled.
    """
    curves, house = _ar_payment_curves_sync(db, as_of)

    invoices = list(db.get_collection(INVOICES).find(
        {"status": {"$in": list(AR_STATUSES)}},
        {"customer_id": 1, "customer_name": 1, "balance": 1, "due_date": 1, "_id": 0},
    ))

    out: Dict[int, Dict] = {}
    for horizon in horizons:
        expected = 0.0
        own_curve_value = 0.0
        doubtful = 0.0
        by_customer: Dict[str, Dict] = {}

        for inv in invoices:
            balance = _f(inv.get("balance"))
            if balance <= 0:
                continue
            due = _parse_date(inv.get("due_date")) or as_of
            past_due = (as_of - due).days
            cid = inv.get("customer_id") or ""
            curve = curves.get(cid)
            has_own = curve is not None
            curve = curve or house

            settled_by_now = _curve_at(curve, past_due)
            if settled_by_now is None or settled_by_now >= 0.999:
                share = 0.0
            else:
                settled_by_then = _curve_at(curve, past_due + horizon) or settled_by_now
                share = max(0.0, (settled_by_then - settled_by_now) / (1 - settled_by_now))

            landing = balance * share
            expected += landing
            doubtful += balance - landing
            if has_own:
                own_curve_value += balance

            name = inv.get("customer_name") or "Unknown"
            row = by_customer.setdefault(name, {
                "customer_name": name, "outstanding": 0.0,
                "expected": 0.0, "has_own_history": has_own,
            })
            row["outstanding"] += balance
            row["expected"] += landing

        rows = sorted(by_customer.values(), key=lambda r: -r["outstanding"])[:20]
        for r in rows:
            r["outstanding"] = round(r["outstanding"], 2)
            r["expected"] = round(r["expected"], 2)
            r["expected_pct"] = (
                round(r["expected"] / r["outstanding"] * 100, 1) if r["outstanding"] > 0 else 0.0
            )

        out[horizon] = {
            "expected": round(expected, 2),
            "not_expected_in_window": round(doubtful, 2),
            "top_customers": rows,
        }

    total_ar = sum(_f(i.get("balance")) for i in invoices if _f(i.get("balance")) > 0)
    return {
        "by_horizon": out,
        "customers_with_own_history": len(curves),
        "invoices_scored": len(invoices),
        "value_on_own_history": round(own_curve_value, 2),
        "value_on_house_average": round(max(0.0, total_ar - own_curve_value), 2),
        "lookback_days": AR_BEHAVIOUR_LOOKBACK_DAYS,
        "min_invoices_for_own_curve": AR_BEHAVIOUR_MIN_INVOICES,
    }


def _cashflow_series_sync(db, days: int = 90) -> List[Dict]:
    """Daily cash in / cash out for a trend chart, transfers excluded."""
    today = datetime.now()
    since = today - timedelta(days=days)

    pipeline = [
        {"$match": {
            "transaction_date": {"$gte": since, "$lte": today},
            "transaction_type": {"$nin": list(INTERNAL_TXN_TYPES)},
        }},
        {"$group": {
            "_id": {"$dateToString": {"format": "%Y-%m-%d", "date": "$transaction_date"}},
            "inflow": {"$sum": {"$cond": [{"$gt": ["$signed_amount", 0]}, "$signed_amount", 0]}},
            "outflow": {"$sum": {"$cond": [{"$lt": ["$signed_amount", 0]}, "$signed_amount", 0]}},
        }},
        {"$sort": {"_id": 1}},
    ]
    return [
        {
            "date": r["_id"],
            "inflow": round(r["inflow"], 2),
            "outflow": round(-r["outflow"], 2),
            "net": round(r["inflow"] + r["outflow"], 2),
        }
        for r in db.get_collection(BANK_TRANSACTIONS).aggregate(pipeline)
    ]


# ─── Receivables / payables ─────────────────────────────────────────────────────

def _receivables_sync(db) -> Dict:
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    buckets = _empty_buckets()
    top: List[Dict] = []

    cursor = db.get_collection(INVOICES).find(
        {"status": {"$in": list(AR_STATUSES)}},
        {"customer_name": 1, "invoice_number": 1, "balance": 1,
         "due_date": 1, "status": 1, "_id": 0},
    )
    by_customer: Dict[str, float] = {}
    for inv in cursor:
        bal = _f(inv.get("balance"))
        if bal <= 0:
            continue
        bucket = _bucket_for(_parse_date(inv.get("due_date")), today)
        buckets[bucket] += bal
        name = inv.get("customer_name", "") or "Unknown"
        by_customer[name] = by_customer.get(name, 0.0) + bal

    top = sorted(
        ({"customer_name": k, "balance": round(v, 2)} for k, v in by_customer.items()),
        key=lambda r: -r["balance"],
    )[:20]

    return {
        "buckets": {k: round(v, 2) for k, v in buckets.items()},
        "total": round(sum(buckets.values()), 2),
        "top_customers": top,
    }


def _payables_sync(db) -> Dict:
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    buckets = _empty_buckets()
    by_vendor: Dict[str, float] = {}

    cursor = db.get_collection(BILLS).find(
        {"status": {"$in": list(AP_STATUSES)}},
        {"vendor_name": 1, "bill_number": 1, "balance": 1,
         "due_date": 1, "status": 1, "_id": 0},
    )
    for bill in cursor:
        bal = _f(bill.get("balance"))
        if bal <= 0:
            continue
        buckets[_bucket_for(_parse_date(bill.get("due_date")), today)] += bal
        name = bill.get("vendor_name", "") or "Unknown"
        by_vendor[name] = by_vendor.get(name, 0.0) + bal

    top = sorted(
        ({"vendor_name": k, "balance": round(v, 2)} for k, v in by_vendor.items()),
        key=lambda r: -r["balance"],
    )[:20]

    return {
        "buckets": {k: round(v, 2) for k, v in buckets.items()},
        "total": round(sum(buckets.values()), 2),
        "top_vendors": top,
    }


def _date_str(v) -> str:
    """Brand-order dates are stored as `YYYY-MM-DD` strings or datetimes."""
    if not v:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    return str(v)[:10]


# (label, amount field, date field, currency). Amounts on a brand order are not
# all in the same currency: the supplier legs are in the PO's currency, the
# landed-cost legs (duty, freight, service providers) are paid locally in INR.
# See the payment report headers in routes/brand_orders.py.
_BO_PAYMENT_LEGS = [
    ("Advance payment", "advance_payment_amount", "advance_payment_date", "PO"),
    ("Custom duty", "custom_duty", "custom_duty_due_date", "INR"),
    ("Shipping charges", "shipping_charges", "shipping_charges_due_date", "INR"),
    ("Balance payment", "balance_payment_amount", "balance_payment_date", "INR"),
    ("Paid to supplier", "total_payment_made_to_supplier",
     "total_payment_made_to_supplier_date", "PO"),
]


def _brand_order_payment_schedule(bo: Dict, po_ccy: str, fx_rate: float) -> List[Dict]:
    """Dated payment legs recorded against a brand order, oldest date first.

    Legs with neither an amount nor a date are dropped — a blank row on the
    brand order says nothing. Undated legs are kept (the money is committed even
    if the date has not been agreed) and sort last.
    """
    legs: List[Dict] = []
    for label, amt_key, date_key, ccy_kind in _BO_PAYMENT_LEGS:
        amount = _f(bo.get(amt_key))
        date = _date_str(bo.get(date_key))
        if amount <= 0 and not date:
            continue
        ccy = po_ccy if ccy_kind == "PO" else "INR"
        rate = fx_rate if ccy_kind == "PO" else 1.0
        legs.append({
            "label": label,
            "amount": round(amount, 2),
            "currency_code": ccy,
            "amount_inr": round(amount * rate, 2),
            "date": date,
        })

    for vp in bo.get("vendor_payments") or []:
        amount = _f(vp.get("amount"))
        date = _date_str(vp.get("date"))
        name = (vp.get("name") or "").strip()
        if amount <= 0 and not date:
            continue
        legs.append({
            "label": name or "Service provider",
            "amount": round(amount, 2),
            "currency_code": "INR",
            "amount_inr": round(amount, 2),
            "date": date,
        })

    # Undated legs last, dated legs chronologically.
    legs.sort(key=lambda p: (p["date"] == "", p["date"]))
    return legs


def _open_po_commitment_sync(db, fx: Dict[str, float]) -> Dict:
    """Value still to be paid on issued purchase orders.

    Open POs are `order_status_formatted == "Issued"` — the same definition the
    master report uses for stock in transit, so the two always agree. Only the
    unreceived portion of each line counts.

    These are foreign-currency orders, so every row carries its own rate and the
    converted total is explicitly labelled rather than silently mixed.

    Each row also carries the payment/milestone schedule the buyers maintain
    against the order on `/brand_orders` (advance, custom duty, shipping,
    balance, plus any ad-hoc vendor payments), so the commitment can be read as
    "what is due, and when" rather than a single lump sum.
    """
    pos = list(
        db.get_collection(PURCHASE_ORDERS).find(
            {"order_status_formatted": "Issued"},
            {"purchaseorder_number": 1, "vendor_name": 1, "vendor_id": 1,
             "currency_code": 1, "date": 1, "delivery_date": 1,
             "line_items": 1, "_id": 0},
        ).sort([("date", 1), ("purchaseorder_number", 1)])
    )

    po_numbers = [p.get("purchaseorder_number") for p in pos if p.get("purchaseorder_number")]
    brand_orders = {
        bo["purchaseorder_number"]: bo
        for bo in db.get_collection(BRAND_ORDERS).find(
            {"purchaseorder_number": {"$in": po_numbers}},
            {"_id": 1, "purchaseorder_number": 1, "brand": 1, "name": 1,
             "po_sub_total": 1, "po_due_date": 1,
             "advance_payment_amount": 1, "advance_payment_date": 1,
             "custom_duty": 1, "custom_duty_due_date": 1,
             "shipping_charges": 1, "shipping_charges_due_date": 1,
             "balance_payment_amount": 1, "balance_payment_date": 1,
             "total_payment_made_to_supplier": 1,
             "total_payment_made_to_supplier_date": 1,
             "vendor_payments": 1,
             "ready_date": 1, "etd_date": 1, "eta_port_date": 1,
             "duty_payment_date": 1, "inward_date": 1},
        )
        if bo.get("purchaseorder_number")
    }

    rows = []
    total_inr = 0.0
    for po in pos:
        ccy = (po.get("currency_code") or "INR").upper()
        rate = fx.get(ccy, 1.0)
        open_val = 0.0
        for li in po.get("line_items", []):
            qty = _f(li.get("quantity"))
            recd = _f(li.get("quantity_received"))
            remaining = qty - recd
            if remaining > 0:
                open_val += remaining * _f(li.get("rate"))
        if open_val <= 0:
            continue
        inr = open_val * rate
        total_inr += inr
        bo = brand_orders.get(po.get("purchaseorder_number") or "") or {}
        schedule = _brand_order_payment_schedule(bo, ccy, rate)
        rows.append({
            "purchaseorder_number": po.get("purchaseorder_number", ""),
            "vendor_name": po.get("vendor_name", "") or "",
            "currency_code": ccy,
            "fx_rate": rate,
            "open_value": round(open_val, 2),
            "open_value_inr": round(inr, 2),
            "date": str(po.get("date", ""))[:10],
            "delivery_date": str(po.get("delivery_date", ""))[:10],
            # From the brand order the buyers maintain against this PO.
            "brand": bo.get("brand", "") or "",
            "order_name": bo.get("name", "") or "",
            "brand_order_id": str(bo["_id"]) if bo.get("_id") else "",
            "po_due_date": _date_str(bo.get("po_due_date")),
            "milestones": [
                {"label": label, "date": d}
                for label, d in (
                    ("Ready", _date_str(bo.get("ready_date"))),
                    ("ETD", _date_str(bo.get("etd_date"))),
                    ("Port ETA", _date_str(bo.get("eta_port_date"))),
                    ("Duty paid", _date_str(bo.get("duty_payment_date"))),
                    ("Inward", _date_str(bo.get("inward_date"))),
                )
                if d
            ],
            "payment_schedule": schedule,
            "scheduled_total_inr": round(sum(p["amount_inr"] for p in schedule), 2),
        })

    rows.sort(key=lambda r: -r["open_value_inr"])
    return {"purchase_orders": rows, "total_inr": round(total_inr, 2), "count": len(rows)}


CREDIT_NOTES = "credit_notes"

# A sale is not counted until it is a real document.
SALES_EXCLUDED_STATUSES = ("void", "draft")


def _last_3_months_range() -> tuple:
    """The three most recent *complete* calendar months.

    Same convention as `sheets_updater._last_3_months_range` so the two reports
    never quote different "last 3 months". Complete months rather than a rolling
    90 days because a part-month tail makes the average read low for no reason —
    on 7 Aug that would be 3.2 months of sales divided by 3.
    """
    today = datetime.now()
    end_dt = today.replace(day=1) - timedelta(days=1)
    start_month, start_year = end_dt.month - 2, end_dt.year
    if start_month <= 0:
        start_month += 12
        start_year -= 1
    start_dt = end_dt.replace(year=start_year, month=start_month, day=1)
    return start_dt, end_dt


def _brand_sales_3m_sync(db) -> Dict:
    """Invoiced sales per brand over the last three complete months, ex-GST.

    Measured straight off the invoice ledger rather than reused from the report
    period, because the report's date range is the user's to change and "last 3
    months" has to mean last 3 months.

    **The amounts are already ex-GST.** `line_items.item_total` sums exactly to
    the invoice `sub_total`, with `tax_total` added on top to reach `total` —
    verified across both `is_inclusive_tax` True and False invoices, where Zoho
    back-calculates the line to its ex-tax amount. Stripping tax again here
    would understate sales by the GST rate.
    """
    start_dt, end_dt = _last_3_months_range()
    start_s, end_s = start_dt.strftime("%Y-%m-%d"), end_dt.strftime("%Y-%m-%d")

    def _by_item(collection: str, date_match: Dict) -> Dict[str, float]:
        pipeline = [
            {"$match": {**date_match, "status": {"$nin": list(SALES_EXCLUDED_STATUSES)}}},
            {"$unwind": "$line_items"},
            {"$group": {
                "_id": "$line_items.item_id",
                "amount": {"$sum": "$line_items.item_total"},
            }},
        ]
        return {
            str(r["_id"]): _f(r["amount"])
            for r in db.get_collection(collection).aggregate(pipeline, allowDiskUse=True)
            if r.get("_id")
        }

    # `invoices.date` is a 'YYYY-MM-DD' string; `credit_notes.date` is a real
    # datetime. Same window, two different comparisons.
    sales_by_item = _by_item(INVOICES, {"date": {"$gte": start_s, "$lte": end_s}})
    returns_by_item = _by_item(
        CREDIT_NOTES,
        {"date": {"$gte": start_dt, "$lte": end_dt.replace(hour=23, minute=59, second=59)}},
    )

    item_ids = set(sales_by_item) | set(returns_by_item)
    brand_by_item = {
        str(p["item_id"]): (p.get("brand") or "").strip()
        for p in db.get_collection("products").find(
            {"item_id": {"$in": list(item_ids)}}, {"item_id": 1, "brand": 1, "_id": 0}
        )
        if p.get("item_id")
    }

    by_brand: Dict[str, Dict[str, float]] = {}
    unmapped = 0.0
    for item_id, amount in sales_by_item.items():
        brand = brand_by_item.get(item_id)
        if not brand:
            unmapped += amount
            continue
        by_brand.setdefault(brand.lower(), {"gross": 0.0, "returns": 0.0})["gross"] += amount
    for item_id, amount in returns_by_item.items():
        brand = brand_by_item.get(item_id)
        if not brand:
            continue
        by_brand.setdefault(brand.lower(), {"gross": 0.0, "returns": 0.0})["returns"] += amount

    return {
        "by_brand": by_brand,
        "start_date": start_s,
        "end_date": end_s,
        "months": 3,
        # Sales on items with no product row (services, adjustments, deleted
        # SKUs). Reported so the brand totals never silently disagree with the
        # ledger.
        "unmapped_sales": round(unmapped, 2),
    }


# Where an arrival date comes from, best first. `purchase_orders.delivery_date`
# is deliberately absent: it is empty on every live open PO.
_ETA_SOURCES = (
    ("eta_port_date", "ETA port"),
    ("etd_date", "ETD + in transit"),
    ("ready_date", "ready at factory"),
    ("po_due_date", "PO due"),
)


def _transit_arrivals_sync(db) -> Dict[str, List[Dict]]:
    """When each brand's in-transit stock is expected, and how much of it.

    Built from the same open POs the master report counts as stock in transit
    (`order_status_formatted == "Issued"`, unreceived quantity only), so the
    units reconcile with `in_transit_units` on the brand row.

    Dates come from the brand order the buyers maintain on `/brand_orders`, not
    from the PO — Zoho's `delivery_date` is empty on every live open PO. The
    date used is labelled with its basis, because they mean different things:
    **ETA port is not availability**, customs clearance and inland transport
    still follow it.

    Line items are attributed to a brand individually rather than assigning the
    whole PO to the brand order's brand, since a PO can span brands.
    """
    pos = list(db.get_collection(PURCHASE_ORDERS).find(
        {"order_status_formatted": "Issued"},
        {"purchaseorder_number": 1, "vendor_name": 1, "line_items": 1, "_id": 0},
    ))
    if not pos:
        return {}

    numbers = [p.get("purchaseorder_number") for p in pos if p.get("purchaseorder_number")]
    orders = {
        bo["purchaseorder_number"]: bo
        for bo in db.get_collection(BRAND_ORDERS).find(
            {"purchaseorder_number": {"$in": numbers}},
            {"_id": 0, "purchaseorder_number": 1, "name": 1, "ready_date": 1,
             "etd_date": 1, "eta_port_date": 1, "po_due_date": 1},
        )
        if bo.get("purchaseorder_number")
    }

    item_ids = {
        str(li.get("item_id")) for p in pos for li in p.get("line_items", []) if li.get("item_id")
    }
    products = {
        str(p["item_id"]): p
        for p in db.get_collection("products").find(
            {"item_id": {"$in": list(item_ids)}},
            {"item_id": 1, "brand": 1, "rate": 1, "_id": 0},
        )
        if p.get("item_id")
    }

    by_brand: Dict[str, List[Dict]] = {}
    for po in pos:
        number = po.get("purchaseorder_number") or ""
        bo = orders.get(number) or {}

        eta, basis = None, "no date set"
        for field, label in _ETA_SOURCES:
            parsed = _parse_date(bo.get(field))
            if parsed:
                eta, basis = parsed, label
                break

        # Roll the PO's unreceived lines up per brand.
        per_brand: Dict[str, Dict[str, float]] = {}
        for li in po.get("line_items", []):
            remaining = _f(li.get("quantity")) - _f(li.get("quantity_received"))
            if remaining <= 0:
                continue
            product = products.get(str(li.get("item_id"))) or {}
            brand = (product.get("brand") or "").strip()
            if not brand:
                continue
            agg = per_brand.setdefault(brand.lower(), {"units": 0.0, "mrp_value": 0.0})
            agg["units"] += remaining
            agg["mrp_value"] += remaining * _f(product.get("rate"))

        for brand_key, agg in per_brand.items():
            by_brand.setdefault(brand_key, []).append({
                "purchaseorder_number": number,
                "vendor_name": po.get("vendor_name", "") or "",
                "order_name": bo.get("name", "") or "",
                "eta_date": eta.date().isoformat() if eta else "",
                "eta_basis": basis,
                "units": round(agg["units"], 0),
                "mrp_value": round(agg["mrp_value"], 2),
            })

    # Soonest first; anything undated sorts last rather than pretending to be
    # imminent.
    for rows in by_brand.values():
        rows.sort(key=lambda r: (not r["eta_date"], r["eta_date"]))
    return by_brand


def _fallback_unit_costs_sync(db, fx: Dict[str, float]) -> Dict[str, float]:
    """Latest purchase-order rate per SKU, in INR.

    Needed because COGS comes from inventory valuation, which by construction
    has no unit cost for a SKU that is out of stock — and a stocked-out brand is
    exactly the one most likely to need ordering. Without this its capital
    requirement silently reads as zero and it drops out of the funding queue.

    `product_unit_prices` would be the better source but is not populated, so we
    fall back to what was actually last paid.
    """
    costs: Dict[str, float] = {}
    cursor = db.get_collection(PURCHASE_ORDERS).find(
        {}, {"line_items": 1, "currency_code": 1, "date": 1, "_id": 0}
    ).sort([("date", 1)])

    # Ascending date so later (more recent) POs overwrite earlier ones.
    for po in cursor:
        rate_fx = fx.get((po.get("currency_code") or "INR").upper(), 1.0)
        for li in po.get("line_items", []):
            sku = ""
            for cf in li.get("item_custom_fields", []) or []:
                if cf.get("api_name") == "cf_sku_code":
                    sku = cf.get("value", "") or ""
                    break
            if not sku:
                continue
            rate = _f(li.get("rate"))
            if rate > 0:
                costs[sku] = round(rate * rate_fx, 2)
    return costs


# ─── Working capital ────────────────────────────────────────────────────────────

def _build_working_capital(
    cash: Dict, receivables: Dict, payables: Dict,
    open_pos: Dict, opex: Dict, collection: Dict,
    ar_behaviour: Optional[Dict] = None,
) -> Dict:
    """Assemble the cash bridge into 30/60/90-day projections.

    Expected collections are bounded by three facts, never a guess:

    1. **Who owes it and how they pay** — each open invoice scored against that
       customer's own settlement history (`_expected_collections_sync`). This is
       the binding constraint in practice and the most accurate: back-tested
       over the 90 days to 2026-08-06 it predicted ₹175.3L against ₹195.0L
       actually collected, where assuming everything due lands predicted ₹286.6L.
    2. **How much is owed and due** inside the horizon, from the invoice ledger.
    3. **How fast money has actually been arriving**, from the bank.

    The lowest applies, and the payload names which one bound it.
    """
    monthly_opex = opex["monthly_opex"]
    net_cash = cash["net_cash"]
    daily_rate = collection["daily_collection_rate"]
    projections = []

    for horizon in HORIZONS:
        ar_due = _within(receivables["buckets"], horizon)
        run_rate_ar = daily_rate * horizon
        behaviour = (ar_behaviour or {}).get("by_horizon", {}).get(horizon)
        behaviour_ar = behaviour["expected"] if behaviour else None

        bounds = {"invoices_due": ar_due, "collection_speed": run_rate_ar}
        if behaviour_ar is not None:
            bounds["payment_behaviour"] = behaviour_ar
        limited_by = min(bounds, key=lambda k: bounds[k])
        expected_ar = bounds[limited_by]

        due_ap = _within(payables["buckets"], horizon)
        # Open POs have no reliable payment date on the record, so the whole
        # remaining commitment is charged to the nearest horizon. Conservative
        # by design: it never lets the plan overspend.
        po_due = open_pos["total_inr"]
        opex_due = monthly_opex * (horizon / 30.0)
        free = net_cash + expected_ar - due_ap - po_due - opex_due
        projections.append({
            "horizon_days": horizon,
            "opening_cash": round(net_cash, 2),
            "expected_collections": round(expected_ar, 2),
            "ar_due_in_window": round(ar_due, 2),
            "run_rate_capacity": round(run_rate_ar, 2),
            "behaviour_estimate": round(behaviour_ar, 2) if behaviour_ar is not None else None,
            # Which constraint bound the estimate: too few invoices to collect,
            # too little collection speed, or customers who simply do not pay
            # that fast?
            "collections_limited_by": limited_by,
            "implied_realisation_pct": round(expected_ar / ar_due * 100, 1) if ar_due > 0 else 0.0,
            "bills_due": round(due_ap, 2),
            "open_po_commitment": round(po_due, 2),
            "operating_expenses": round(opex_due, 2),
            "free_cash": round(free, 2),
        })

    return {
        "cash": cash,
        "receivables": receivables,
        "payables": payables,
        "open_purchase_orders": open_pos,
        "opex": opex,
        "collection": collection,
        "ar_behaviour": ar_behaviour or {},
        "projections": projections,
        # Default envelope. The order plan picks a horizon explicitly — orders
        # placed today are paid over a 60–96 day lead time, not within 30 days,
        # so planning against the 30-day figure understates what is affordable.
        "purchase_envelope": projections[-1]["free_cash"],
        "purchase_envelope_horizon": projections[-1]["horizon_days"],
        "generated_at": datetime.now().isoformat(),
    }


@router.get("/working-capital")
async def get_working_capital(
    usd_inr: Optional[float] = Query(
        None, gt=0, description="USD→INR override; omit to use the live rate"
    ),
    cny_inr: Optional[float] = Query(
        None, gt=0, description="CNY/RMB→INR override; omit to use the live rate"
    ),
    start_date: Optional[str] = Query(
        None,
        description="Start of the window used to measure collection and expense rates "
                    "(YYYY-MM-DD). Defaults to 90 days ago.",
    ),
    end_date: Optional[str] = Query(
        None, description="End of that window (YYYY-MM-DD). Defaults to today."
    ),
    db=Depends(get_database),
):
    """Cash on hand + expected collections − committed outflows, per horizon.

    Balances are always **current** — bank balances, receivables and payables
    describe today, and cannot be rewound. The date range instead sets the window
    over which the **rates** are measured: how fast money is being collected and
    how fast it is being spent. Widen it to smooth out a lumpy month, or point it
    at last year's peak season to plan against that behaviour.

    Sources: `bank_accounts` (bank feed where present, net of credit-card
    liability), `invoices` (AR, excluding void/draft), `bills` (AP), issued
    `purchase_orders` converted to INR at the live or supplied rate, and
    `bank_transactions` for both rates.
    """
    try:
        window_end = (
            datetime.strptime(end_date, "%Y-%m-%d") if end_date else datetime.now()
        )
        window_start = (
            datetime.strptime(start_date, "%Y-%m-%d")
            if start_date
            else window_end - timedelta(days=90)
        )
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="start_date and end_date must be YYYY-MM-DD",
        )
    if window_start >= window_end:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="start_date must be before end_date",
        )

    fx_resolved = await _resolve_fx(usd_inr, cny_inr)
    fx = fx_resolved["rates"]
    try:
        cash, receivables, payables, open_pos, opex, collection, ar_behaviour = await asyncio.gather(
            asyncio.to_thread(_cash_position_sync, db),
            asyncio.to_thread(_receivables_sync, db),
            asyncio.to_thread(_payables_sync, db),
            asyncio.to_thread(_open_po_commitment_sync, db, fx),
            asyncio.to_thread(_monthly_opex_sync, db, window_start, window_end),
            asyncio.to_thread(_collection_run_rate_sync, db, window_start, window_end),
            asyncio.to_thread(_expected_collections_sync, db, datetime.now()),
        )
    except Exception as e:
        logger.error(f"working-capital failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to build working capital view: {e}",
        )

    payload = _build_working_capital(
        cash, receivables, payables, open_pos, opex, collection, ar_behaviour
    )
    payload["fx"] = fx_resolved["meta"]
    payload["rate_window"] = {
        "start_date": window_start.date().isoformat(),
        "end_date": window_end.date().isoformat(),
        "days": (window_end - window_start).days,
    }
    return JSONResponse(status_code=status.HTTP_200_OK, content=payload)


@router.get("/cashflow")
async def get_cashflow(
    days: int = Query(90, ge=7, le=730, description="Trailing window in days"),
    db=Depends(get_database),
):
    """Daily cash in / out, excluding internal transfers."""
    series = await asyncio.to_thread(_cashflow_series_sync, db, days)
    return JSONResponse(status_code=status.HTTP_200_OK, content={
        "days": days,
        "series": series,
        "total_inflow": round(sum(s["inflow"] for s in series), 2),
        "total_outflow": round(sum(s["outflow"] for s in series), 2),
        "net": round(sum(s["net"] for s in series), 2),
    })


# ─── Brand order plan ───────────────────────────────────────────────────────────

def _classify_urgency(
    coverage: float, lead_time: float, drr: float, is_new: bool = False
) -> str:
    """Where a brand sits against its own lead time.

    The 3× lead-time ceiling matches the dead-SKU exclusion the dashboard already
    applies to weighted average cover, so the two pages never disagree about what
    counts as overstocked.

    A brand still inside its selling threshold short-circuits to NEW. Everything
    below this line divides by a run rate, and a brand a few days old does not
    have one — its coverage reads near zero, which would otherwise put it at the
    very top of the funding queue on the strength of a launch week.
    """
    if is_new:
        return "NEW"
    if drr <= 0:
        return "EXCESS"
    if coverage >= 3 * lead_time:
        return "EXCESS"
    if coverage < lead_time:
        return "CRITICAL"
    if coverage < lead_time + 30:
        return "ORDER"
    return "HEALTHY"


URGENCY_RANK = {"CRITICAL": 0, "NEW": 1, "ORDER": 2, "HEALTHY": 3, "EXCESS": 4}


def _unit_cost_for(
    item: Dict, fallback_costs: Dict[str, float], fx: Dict[str, float]
) -> tuple:
    """Best available cost to buy one more unit, and where it came from.

    Precedence, most to least authoritative for a *forward* purchase:

    1. **Managed unit price** (`product_unit_prices`, the Unit Prices page) —
       what the buyer says the next unit will actually cost. Empty today, but the
       moment it is populated it takes over from everything below.
    2. **COGS** — FIFO value of stock on hand. Accurate, but historical, and
       absent entirely for anything out of stock.
    3. **Last purchase-order rate** — what we last actually paid.

    Returns ``(unit_cost_inr, source)``.
    """
    managed = _f(item.get("unit_price"))
    if managed > 0:
        ccy = (item.get("unit_price_currency") or "INR").upper()
        return round(managed * fx.get(ccy, 1.0), 2), "unit_price"

    cogs = _f(item.get("cogs_unit_cost"))
    if cogs > 0:
        return cogs, "cogs"

    last_po = _f(fallback_costs.get(item.get("sku_code", "")))
    if last_po > 0:
        return last_po, "last_po"

    return 0.0, "none"


def _cap_order_qty(
    order_qty: float, drr: float, on_hand: float, in_transit: float,
    max_cover_days: Optional[int],
) -> float:
    """Trim a suggested order so it never buys more than `max_cover_days` of cover.

    The master report sizes an order from demand alone. When a SKU's DRR comes
    from a lookback window — or from a short burst of sales — the suggested
    quantity can be several months of stock, and the plan would then spend real
    cash on it. This ceiling counts what is already on hand and on the water, so
    it only ever removes units that would push total cover past the limit.

    A SKU with no run rate is left alone: there is no cover to compute, and the
    master report has its own reasons for suggesting a quantity there (a demand
    override, typically).
    """
    if not max_cover_days or order_qty <= 0 or drr <= 0:
        return order_qty
    room = drr * max_cover_days - (on_hand + in_transit)
    return float(max(0.0, min(order_qty, round(room))))


def _aggregate_brands(
    items: List[Dict], brand_logistics: Dict,
    fallback_costs: Dict[str, float], fx: Dict[str, float],
    max_cover_days: Optional[int] = None,
    min_selling_days: int = DEFAULT_MIN_SELLING_DAYS,
    period_days: int = 90,
    sales_3m: Optional[Dict] = None,
    transit_arrivals: Optional[Dict] = None,
) -> List[Dict]:
    """Roll master-report SKUs into brand rows priced per `_unit_cost_for`.

    Every SKU in the brand is kept, not just the ones with an order quantity —
    the buyer needs to see the whole brand to judge the recommendation.
    """
    by_brand: Dict[str, List[Dict]] = {}
    for it in items:
        brand = (it.get("brand") or "Unassigned").strip() or "Unassigned"
        by_brand.setdefault(brand, []).append(it)

    rows = []
    for brand, group in by_brand.items():
        settings = brand_logistics.get(brand.lower(), {})
        lead_time = _f(settings.get("lead_time"), 60.0)

        # How long the brand has actually been selling: the longest-running SKU
        # in it. Max, not average — one new SKU inside an established brand does
        # not make the brand new, but the brand can be no older than its oldest
        # SKU's stock history.
        selling_days = max(
            (_f((it.get("combined_metrics") or {}).get("total_days_in_stock")) for it in group),
            default=0.0,
        )
        has_sales = any(
            _f((it.get("combined_metrics") or {}).get("avg_daily_run_rate")) > 0 for it in group
        )
        # No sales at all is not "new", it is dead — that is the EXCESS path.
        is_new_brand = bool(has_sales and 0 < selling_days < min_selling_days)
        # The cap is a per-brand decision saved on brand_logistics; the request
        # parameter is only the default for brands that have not set one.
        saved_cap = settings.get("max_cover_days")
        base_cover_cap = int(_f(saved_cap)) if saved_cap is not None else max_cover_days
        cover_cap_source = "brand" if saved_cap is not None else "default"
        # A launch DRR is a guess, so buy one lead time of cover at most: enough
        # not to stock out before a real reorder, by which point there is real
        # data to size it from.
        brand_cover_cap = (
            min(base_cover_cap or int(lead_time), int(lead_time))
            if is_new_brand
            else base_cover_cap
        )
        if is_new_brand:
            cover_cap_source = "new_brand"

        stock_units = transit_units = 0.0
        stock_value = transit_value = 0.0
        wh_units = 0.0
        wh_mrp_value = transit_mrp_value = 0.0
        skus_missing_mrp = 0
        net_units_sold = 0.0
        gross_units_sold = credit_note_units = 0.0
        order_units = capital_required = 0.0
        suggested_units = suggested_capital = 0.0
        revenue = cogs_sold = 0.0
        drr = 0.0
        skus_to_order = 0
        skus_capped = 0
        missing_cost = 0
        estimated_cost = 0
        sku_rows = []

        for it in group:
            cm = it.get("combined_metrics", {}) or {}
            sku = it.get("sku_code", "")
            unit_cost, cost_source = _unit_cost_for(it, fallback_costs, fx)
            on_hand = _f(it.get("latest_total_stock"))
            transit = _f(it.get("total_stock_in_transit"))
            suggested_qty = _f(it.get("order_qty_plus_extra_qty_rounded"))
            net_units = _f(cm.get("total_sales"))
            sku_drr = _f(cm.get("avg_daily_run_rate"))
            order_qty = _cap_order_qty(
                suggested_qty, sku_drr, on_hand, transit, brand_cover_cap
            )

            drr += sku_drr
            stock_units += on_hand
            transit_units += transit
            stock_value += on_hand * unit_cost
            transit_value += transit * unit_cost
            revenue += _f(cm.get("total_amount"))
            cogs_sold += max(0.0, net_units) * unit_cost

            # Retail-value view, Pupscribe warehouse only. `latest_zoho_stock` is
            # the WH snapshot; `latest_total_stock` above also carries FBA, which
            # is not in the warehouse and must not be counted here.
            mrp = _f(it.get("mrp"))
            wh_qty = _f(it.get("latest_zoho_stock"))
            wh_units += wh_qty
            wh_mrp_value += wh_qty * mrp
            transit_mrp_value += transit * mrp
            if mrp <= 0 and (wh_qty > 0 or transit > 0):
                skus_missing_mrp += 1
            net_units_sold += max(0.0, net_units)
            gross_units_sold += _f(cm.get("total_units_sold"))
            credit_note_units += _f(cm.get("total_credit_notes"))

            suggested_units += suggested_qty
            suggested_capital += suggested_qty * unit_cost
            if suggested_qty > order_qty:
                skus_capped += 1

            if order_qty > 0:
                order_units += order_qty
                capital_required += order_qty * unit_cost
                skus_to_order += 1
                if cost_source == "none":
                    missing_cost += 1
                elif cost_source == "last_po":
                    estimated_cost += 1

            sku_rows.append({
                "sku_code": sku,
                "product_name": (
                    it.get("item_name") or it.get("product_name") or it.get("name") or ""
                ),
                "unit_cost": round(unit_cost, 2),
                "cost_source": cost_source,
                "on_hand": round(on_hand, 2),
                "in_transit": round(transit, 2),
                "drr": round(sku_drr, 3),
                "current_days_coverage": round(_f(it.get("current_days_coverage")), 1),
                # What demand asked for, before the cover ceiling.
                "suggested_qty": round(suggested_qty, 0),
                "order_qty": round(order_qty, 0),
                "capped_units": round(max(0.0, suggested_qty - order_qty), 0),
                "capital_required": round(order_qty * unit_cost, 2),
                "excess_or_order": it.get("excess_or_order", ""),
            })

        coverage = round((stock_units + transit_units) / drr, 1) if drr > 0 else 0.0
        gross_margin = revenue - cogs_sold
        # Inventory at cost, including what is already on the water — that
        # capital is just as committed as what sits in the warehouse.
        avg_inventory = stock_value + transit_value
        gmroi = round(gross_margin / avg_inventory, 2) if avg_inventory > 0 else 0.0
        margin_pct = round(gross_margin / revenue * 100, 1) if revenue > 0 else 0.0
        urgency = _classify_urgency(coverage, lead_time, drr, is_new_brand)

        # What freeing this brand down to a healthy 2× lead time would release.
        cash_trapped = 0.0
        if urgency == "EXCESS" and drr > 0 and coverage > 2 * lead_time:
            excess_units = (coverage - 2 * lead_time) * drr
            unit_val = avg_inventory / (stock_units + transit_units) if (stock_units + transit_units) > 0 else 0.0
            cash_trapped = excess_units * unit_val
        elif urgency == "EXCESS" and drr <= 0:
            cash_trapped = avg_inventory

        # Sell-through DRR: net units over the *calendar* window, not over the
        # days the SKU happened to be in stock. The master report's DRR answers
        # "how fast does it sell when available", which is the right basis for
        # sizing an order; this answers "how fast does the warehouse actually
        # drain", which is the right basis for valuing what is sitting in it. A
        # SKU in stock 10 days of 90 has a flattering days-in-stock DRR and would
        # make the holding look like weeks of cover when it is months.
        sell_through_drr = round(net_units_sold / period_days, 3) if period_days > 0 else 0.0
        monthly_units = sell_through_drr * DAYS_PER_MONTH

        def _months(units: float) -> Optional[float]:
            # No sell-through means no meaningful answer — not "zero months".
            if monthly_units <= 0:
                return None
            return round(units / monthly_units, 1)

        wh_collection_value = wh_mrp_value / 2
        transit_collection_value = transit_mrp_value / 2

        # When the in-transit stock is expected, soonest first. The first leg
        # with a date is the one the buyer is waiting on; an ETA already in the
        # past is flagged rather than quietly shown as "coming".
        arrivals = (transit_arrivals or {}).get(brand.lower(), [])
        today = datetime.now().date()
        for leg in arrivals:
            eta = _parse_date(leg.get("eta_date"))
            leg["days_away"] = (eta.date() - today).days if eta else None
            leg["is_overdue"] = bool(eta and eta.date() < today)
        next_arrival = next((leg for leg in arrivals if leg.get("eta_date")), None)

        # Average monthly sales over the last three complete calendar months,
        # ex-GST — measured off the invoice ledger by `_brand_sales_3m_sync`,
        # deliberately independent of the report's own date range so that
        # "last 3 months" always means last 3 months. Returns come from actual
        # credit notes, not from a per-unit estimate.
        s3 = sales_3m.get(brand.lower(), {})
        sales_3m_gross = _f(s3.get("gross"))
        sales_3m_returns = _f(s3.get("returns"))
        sales_3m_net = max(0.0, sales_3m_gross - sales_3m_returns)
        avg_monthly_sales = sales_3m_gross / 3
        avg_monthly_sales_net = sales_3m_net / 3

        sku_rows.sort(key=lambda r: -r["capital_required"])

        rows.append({
            "brand": brand,
            "sku_count": len(group),
            "skus_to_order": skus_to_order,
            "skus_missing_cost": missing_cost,
            "skus_estimated_cost": estimated_cost,
            "lead_time": lead_time,
            "drr": round(drr, 3),
            "on_hand_units": round(stock_units, 0),
            "in_transit_units": round(transit_units, 0),
            "stock_value": round(stock_value, 2),
            "transit_value": round(transit_value, 2),
            "inventory_at_cost": round(avg_inventory, 2),
            "current_days_coverage": coverage,
            "revenue": round(revenue, 2),
            "cogs_sold": round(cogs_sold, 2),
            "gross_margin": round(gross_margin, 2),
            "margin_pct": margin_pct,
            "gmroi": gmroi,
            "order_units": round(order_units, 0),
            "capital_required": round(capital_required, 2),
            # Demand's ask before the cover ceiling, so the buyer can see what
            # the limit removed rather than just a smaller number.
            "suggested_units": round(suggested_units, 0),
            "suggested_capital": round(suggested_capital, 2),
            "capped_units": round(max(0.0, suggested_units - order_units), 0),
            "capped_capital": round(max(0.0, suggested_capital - capital_required), 2),
            "skus_capped": skus_capped,
            "is_new_brand": is_new_brand,
            "selling_days": round(selling_days, 0),
            "cover_cap_days": brand_cover_cap,
            "cover_cap_source": cover_cap_source,

            # ── Retail-value view (Pupscribe warehouse) ──────────────────────
            "wh_units": round(wh_units, 0),
            "wh_mrp_value": round(wh_mrp_value, 2),
            "wh_collection_value": round(wh_collection_value, 2),
            "transit_mrp_value": round(transit_mrp_value, 2),
            "transit_collection_value": round(transit_collection_value, 2),
            "total_mrp_value": round(wh_mrp_value + transit_mrp_value, 2),
            "total_collection_value": round(wh_collection_value + transit_collection_value, 2),
            "skus_missing_mrp": skus_missing_mrp,

            # ── Sales, ex-GST ────────────────────────────────────────────────
            "avg_monthly_sales_ex_gst": round(avg_monthly_sales, 2),
            "avg_monthly_sales_net_ex_gst": round(avg_monthly_sales_net, 2),
            "sales_3m_gross_ex_gst": round(sales_3m_gross, 2),
            "sales_3m_returns_ex_gst": round(sales_3m_returns, 2),
            "sales_3m_net_ex_gst": round(sales_3m_net, 2),

            # ── Months of cover on measured sell-through ─────────────────────
            "sell_through_drr": sell_through_drr,
            "monthly_units": round(monthly_units, 1),
            "monthly_mrp_value": round(
                monthly_units * (wh_mrp_value / wh_units) if wh_units > 0 else 0.0, 2
            ),
            "months_in_warehouse": _months(wh_units),
            "months_in_transit": _months(transit_units),
            "months_total": _months(wh_units + transit_units),
            "transit_arrivals": arrivals,
            "next_arrival": next_arrival,

            "urgency": urgency,
            "cash_trapped": round(cash_trapped, 2),
            "skus": sku_rows,
        })

    return rows


def _allocate(rows: List[Dict], envelope: float, new_brand_reserve: float = 0.0) -> Dict:
    """Fund brands from the cash envelope, best return on inventory first.

    Ranking by GMROI within urgency tier is the point of the whole report: it
    stops the last of a tight cash position going to a slow, thin-margin brand
    while a fast high-margin one stocks out.

    `new_brand_reserve` is taken off the top before any existing brand is funded.
    A brand we are about to launch has no history for the ranking to work with,
    so it cannot compete on GMROI — it has to be carved out, or it loses to the
    incumbents every time.
    """
    fundable = [r for r in rows if r["urgency"] in ("CRITICAL", "ORDER") and r["capital_required"] > 0]
    fundable.sort(key=lambda r: (URGENCY_RANK[r["urgency"]], -r["gmroi"], -r["drr"]))

    reserve = max(0.0, new_brand_reserve)
    available = max(0.0, envelope - reserve)
    remaining = available
    total_requested = sum(r["capital_required"] for r in fundable)

    for rank, row in enumerate(fundable, start=1):
        need = row["capital_required"]
        row["priority"] = rank
        if remaining <= 0:
            row["funded_amount"] = 0.0
            row["funded_pct"] = 0.0
            row["recommendation"] = "DEFER"
        elif remaining >= need:
            row["funded_amount"] = round(need, 2)
            row["funded_pct"] = 100.0
            row["recommendation"] = "ORDER FULL"
            remaining -= need
        else:
            row["funded_amount"] = round(remaining, 2)
            row["funded_pct"] = round(remaining / need * 100, 1)
            row["recommendation"] = "ORDER PARTIAL"
            remaining = 0.0

        # Trim the brand's SKU list to what the funded amount actually buys,
        # highest-value lines first, so purchase has something to act on.
        budget = row["funded_amount"]
        for sku in row["skus"]:
            cost = sku["capital_required"]
            if cost <= 0 or budget <= 0:
                sku["funded_qty"] = 0
            elif budget >= cost:
                sku["funded_qty"] = sku["order_qty"]
                budget -= cost
            else:
                unit = sku["unit_cost"]
                sku["funded_qty"] = int(budget // unit) if unit > 0 else 0
                budget = 0.0

    for row in rows:
        if "recommendation" not in row:
            row["priority"] = None
            row["funded_amount"] = 0.0
            row["funded_pct"] = 0.0
            if row["urgency"] == "NEW":
                # Deliberately not auto-funded. Ranking is by GMROI over the
                # period, and a brand a few days old has neither a trustworthy
                # GMROI nor a trustworthy run rate — letting it compete would
                # hand it cash ahead of established brands on the strength of a
                # launch week. The buyer decides, and reserves cash for it the
                # same way they would for a brand not yet in the system.
                row["recommendation"] = "NEW - REVIEW"
                row["review_reason"] = (
                    f"Only {row['selling_days']:.0f} day(s) of selling history, so the run rate is "
                    f"not yet reliable enough to size a purchase from. The quantity shown is "
                    f"capped at one lead time ({row['cover_cap_days']} days) of cover as a "
                    f"starting point — confirm it, then reserve the cash for it under New brands."
                )
            elif row["urgency"] in ("CRITICAL", "ORDER"):
                # Needs stock but priced at nothing — never label this
                # "DO NOT ORDER", which is the opposite of the truth. It means
                # the master report suggested no quantity (usually because the
                # SKUs are inactive or discontinued) or we could not cost them.
                row["recommendation"] = "REVIEW"
                row["review_reason"] = (
                    "No costed order quantity — SKUs are likely inactive or "
                    "discontinued, or have no COGS and no purchase history."
                )
            else:
                row["recommendation"] = "HOLD" if row["urgency"] == "HEALTHY" else "DO NOT ORDER"
            for sku in row["skus"]:
                sku["funded_qty"] = 0

    # Display order must equal funding order, or the priority numbers read as
    # wrong. Anything unranked (REVIEW / HOLD / DO NOT ORDER) sorts after every
    # ranked brand, by urgency then GMROI.
    rows.sort(
        key=lambda r: (
            0 if r.get("priority") else 1,
            r.get("priority") or 0,
            URGENCY_RANK[r["urgency"]],
            -(r["gmroi"] or 0),
        )
    )

    return {
        "total_requested": round(total_requested, 2),
        "total_funded": round(sum(r["funded_amount"] for r in fundable), 2),
        "unfunded": round(max(0.0, total_requested - sum(r["funded_amount"] for r in fundable)), 2),
        "envelope": round(envelope, 2),
        "new_brand_reserve": round(reserve, 2),
        # What the ranked allocation actually had to spend, after the carve-out.
        "available_for_existing_brands": round(available, 2),
        "remaining_envelope": round(remaining, 2),
        "cash_trapped_in_excess": round(sum(r["cash_trapped"] for r in rows), 2),
        # Not part of the funded total — new brands are a buyer decision, and
        # this is what saying yes to all of them would cost on top.
        "new_brands_indicative": round(
            sum(r["capital_required"] for r in rows if r["urgency"] == "NEW"), 2
        ),
        "new_brands_count": sum(1 for r in rows if r["urgency"] == "NEW"),
        "capped_units": round(sum(r.get("capped_units", 0) for r in rows), 0),
        "capped_capital": round(sum(r.get("capped_capital", 0.0) for r in rows), 2),
        "suggested_before_cap": round(
            sum(r.get("suggested_capital", 0.0) for r in rows
                if r["urgency"] in ("CRITICAL", "ORDER")), 2
        ),
    }


async def _brand_order_plan_data(
    start_date: str, end_date: str, db,
    usd_inr: Optional[float], cny_inr: Optional[float],
    envelope_override: Optional[float], horizon_days: int,
    max_cover_days: Optional[int] = DEFAULT_MAX_COVER_DAYS,
    include_new_brand_reserves: bool = True,
    min_selling_days: int = DEFAULT_MIN_SELLING_DAYS,
) -> Dict:
    fx_resolved = await _resolve_fx(usd_inr, cny_inr)
    fx = fx_resolved["rates"]

    # COGS is mandatory here — capital_required is meaningless without a unit
    # cost, and product_unit_prices is not populated.
    master_task = _generate_master_report_data(
        start_date, end_date, db, include_cogs=True
    )
    cash_task = asyncio.to_thread(_cash_position_sync, db)
    ar_task = asyncio.to_thread(_receivables_sync, db)
    ap_task = asyncio.to_thread(_payables_sync, db)
    po_task = asyncio.to_thread(_open_po_commitment_sync, db, fx)
    # Measure the cash rates over the same period the demand figures come from,
    # so the plan is internally consistent.
    _win_end = datetime.strptime(end_date, "%Y-%m-%d")
    _win_start = datetime.strptime(start_date, "%Y-%m-%d")
    opex_task = asyncio.to_thread(_monthly_opex_sync, db, _win_start, _win_end)
    collection_task = asyncio.to_thread(_collection_run_rate_sync, db, _win_start, _win_end)
    fallback_task = asyncio.to_thread(_fallback_unit_costs_sync, db, fx)
    behaviour_task = asyncio.to_thread(_expected_collections_sync, db, datetime.now())
    sales_3m_task = asyncio.to_thread(_brand_sales_3m_sync, db)
    arrivals_task = asyncio.to_thread(_transit_arrivals_sync, db)

    (master, cash, receivables, payables, open_pos, opex,
     collection, fallback_costs, ar_behaviour, sales_3m,
     transit_arrivals) = await asyncio.gather(
        master_task, cash_task, ar_task, ap_task, po_task, opex_task,
        collection_task, fallback_task, behaviour_task, sales_3m_task, arrivals_task
    )

    working_capital = _build_working_capital(
        cash, receivables, payables, open_pos, opex, collection, ar_behaviour
    )
    # Fund against the horizon the buyer is actually planning over. A container
    # ordered today lands in 60-96 days, so the cash that pays for it is the cash
    # available over that period, not what is spare this month.
    horizon_projection = next(
        (p for p in working_capital["projections"] if p["horizon_days"] == horizon_days),
        working_capital["projections"][-1],
    )
    envelope = (
        envelope_override
        if envelope_override is not None
        else horizon_projection["free_cash"]
    )

    def _load_logistics():
        return {
            (d.get("brand", "") or "").lower(): d
            for d in db.get_collection(BRAND_LOGISTICS).find({}, {"_id": 0})
        }

    brand_logistics, reserves = await asyncio.gather(
        asyncio.to_thread(_load_logistics),
        asyncio.to_thread(_new_brand_reserves_sync, db, True),
    )
    if not include_new_brand_reserves:
        reserves = []
    reserve_total = sum(r["amount"] for r in reserves)

    items = master.get("combined_data", []) or []
    period_days = max(1, (_win_end - _win_start).days)
    rows = _aggregate_brands(
        items, brand_logistics, fallback_costs, fx,
        max_cover_days, min_selling_days, period_days,
        sales_3m.get("by_brand", {}), transit_arrivals,
    )
    allocation = _allocate(rows, envelope, reserve_total)

    cogs_date = (master.get("latest_stock_dates") or {}).get("cogs", "")
    skus_missing_cost = sum(r["skus_missing_cost"] for r in rows)
    skus_estimated_cost = sum(r["skus_estimated_cost"] for r in rows)

    # The exact SKUs behind each warning, so the buyer can act on them rather
    # than just being told a count.
    cost_issues = [
        {
            "brand": r["brand"],
            "sku_code": s["sku_code"],
            "product_name": s["product_name"],
            "order_qty": s["order_qty"],
            "unit_cost": s["unit_cost"],
            "cost_source": s["cost_source"],
            "capital_required": s["capital_required"],
        }
        for r in rows
        for s in r["skus"]
        if s["order_qty"] > 0 and s["cost_source"] in ("last_po", "none")
    ]
    cost_issues.sort(key=lambda x: (x["cost_source"] != "none", -x["order_qty"]))

    warnings = []
    if skus_estimated_cost:
        warnings.append({
            "level": "info",
            "count": skus_estimated_cost,
            "cost_source": "last_po",
            "message": (
                f"{skus_estimated_cost} SKU(s) had no COGS (usually because they are out "
                f"of stock) and are costed at their most recent purchase-order rate instead."
            ),
        })
    if skus_missing_cost:
        warnings.append({
            "level": "warning",
            "count": skus_missing_cost,
            "cost_source": "none",
            "message": (
                f"{skus_missing_cost} SKU(s) with an order quantity have no managed unit "
                f"price, no COGS and no purchase history, so their capital requirement is "
                f"understated."
            ),
        })

    if allocation["new_brands_count"] > 0:
        new_names = ", ".join(r["brand"] for r in rows if r["urgency"] == "NEW")
        warnings.append({
            "level": "info",
            "count": allocation["new_brands_count"],
            "cost_source": "new_brand",
            "message": (
                f"{new_names} {'has' if allocation['new_brands_count'] == 1 else 'have'} under "
                f"{min_selling_days} days of selling history, so {'it is' if allocation['new_brands_count'] == 1 else 'they are'} "
                f"held out of the automatic ranking and left for you to decide."
            ),
        })

    if allocation["capped_units"] > 0:
        warnings.append({
            "level": "info",
            "count": int(allocation["capped_units"]),
            "cost_source": "order_cap",
            "message": (
                f"Order quantities are capped at {max_cover_days} days of cover: "
                f"{int(allocation['capped_units']):,} unit(s) worth "
                f"₹{int(allocation['capped_capital']):,} were trimmed from what demand asked for."
            ),
        })

    return {
        "period": {"start_date": start_date, "end_date": end_date},
        "working_capital": working_capital,
        "horizon_days": horizon_projection["horizon_days"],
        "max_cover_days": max_cover_days,
        "min_selling_days": min_selling_days,
        "period_days": period_days,
        # Retail-value roll-up across every brand, for the page header.
        "inventory_value": {
            "wh_mrp_value": round(sum(r["wh_mrp_value"] for r in rows), 2),
            "wh_collection_value": round(sum(r["wh_collection_value"] for r in rows), 2),
            "transit_mrp_value": round(sum(r["transit_mrp_value"] for r in rows), 2),
            "transit_collection_value": round(sum(r["transit_collection_value"] for r in rows), 2),
            "total_mrp_value": round(sum(r["total_mrp_value"] for r in rows), 2),
            "total_collection_value": round(sum(r["total_collection_value"] for r in rows), 2),
            "wh_units": round(sum(r["wh_units"] for r in rows), 0),
            "skus_missing_mrp": sum(r["skus_missing_mrp"] for r in rows),
            "sales_3m_gross_ex_gst": round(sum(r["sales_3m_gross_ex_gst"] for r in rows), 2),
            "sales_3m_net_ex_gst": round(sum(r["sales_3m_net_ex_gst"] for r in rows), 2),
            "avg_monthly_sales_ex_gst": round(
                sum(r["avg_monthly_sales_ex_gst"] for r in rows), 2
            ),
            "avg_monthly_sales_net_ex_gst": round(
                sum(r["avg_monthly_sales_net_ex_gst"] for r in rows), 2
            ),
            "sales_window": {
                "start_date": sales_3m.get("start_date", ""),
                "end_date": sales_3m.get("end_date", ""),
                "months": sales_3m.get("months", 3),
            },
            "unmapped_sales": sales_3m.get("unmapped_sales", 0.0),
        },
        "allocation": allocation,
        "new_brand_reserves": reserves,
        "brands": rows,
        "warnings": warnings,
        "cost_issues": cost_issues,
        "cogs_as_of": cogs_date,
        "fx": fx_resolved["meta"],
        "generated_at": datetime.now().isoformat(),
    }


@router.get("/brand-order-plan")
async def get_brand_order_plan(
    start_date: str = Query(..., description="Start date YYYY-MM-DD"),
    end_date: str = Query(..., description="End date YYYY-MM-DD"),
    usd_inr: Optional[float] = Query(None, gt=0, description="USD→INR override; omit for live rate"),
    cny_inr: Optional[float] = Query(None, gt=0, description="CNY→INR override; omit for live rate"),
    envelope_override: Optional[float] = Query(
        None, description="Spend this instead of the computed free cash"
    ),
    horizon_days: int = Query(
        90, description="Planning horizon whose free cash funds the plan (30, 60 or 90)"
    ),
    max_cover_days: Optional[int] = Query(
        DEFAULT_MAX_COVER_DAYS, ge=0,
        description="Never order more than this many days of cover per SKU. 0 disables the cap.",
    ),
    include_new_brand_reserves: bool = Query(
        True, description="Set aside the active new-brand reserves before funding existing brands"
    ),
    min_selling_days: int = Query(
        DEFAULT_MIN_SELLING_DAYS, ge=0,
        description="Days of selling history a brand needs before its run rate is trusted "
                    "enough to auto-fund. Below this it is held out for review. 0 disables.",
    ),
    db=Depends(get_database),
):
    """Which brands to order, ranked by GMROI and funded from available cash.

    Each brand is classified against **its own** lead time — CRITICAL below one
    lead time, ORDER below lead time + 30, EXCESS above 3× lead time — then the
    fundable ones are ranked by GMROI (gross margin per rupee of inventory) and
    paid for out of the working-capital envelope until it runs out.

    Runs the full master report with COGS enabled, so expect ~60–90s.
    """
    try:
        payload = await _brand_order_plan_data(
            start_date, end_date, db, usd_inr, cny_inr,
            envelope_override, horizon_days,
            max_cover_days or None, include_new_brand_reserves, min_selling_days,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"brand-order-plan failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to build brand order plan: {e}",
        )
    return JSONResponse(status_code=status.HTTP_200_OK, content=payload)


# ─── Excel ──────────────────────────────────────────────────────────────────────

_HDR_FILL = PatternFill("solid", fgColor="1F3864")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
_SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")
_MONEY = '#,##0'
_THIN = Border(*(Side(style="thin", color="BFBFBF"),) * 4)

_URGENCY_FILL = {
    "CRITICAL": PatternFill("solid", fgColor="FF6B6B"),
    "NEW": PatternFill("solid", fgColor="D9C2F0"),
    "ORDER": PatternFill("solid", fgColor="FFD966"),
    "HEALTHY": PatternFill("solid", fgColor="C6E0B4"),
    "EXCESS": PatternFill("solid", fgColor="F4B183"),
}


def _write_header(ws, headers: List[str], row: int = 1):
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=name)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN
    ws.row_dimensions[row].height = 30


def _autosize(ws, widths: List[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_plan_xlsx(data: Dict) -> bytes:
    wb = openpyxl.Workbook()

    # ── Sheet 1: cash bridge ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    wc = data["working_capital"]
    alloc = data["allocation"]

    ws["A1"] = "Cash-Aware Purchase Plan"
    ws["A1"].font = Font(bold=True, size=16, color="1F3864")
    ws["A2"] = (
        f"Period {data['period']['start_date']} to {data['period']['end_date']}  ·  "
        f"COGS as of {data.get('cogs_as_of') or 'n/a'}  ·  "
        f"generated {data['generated_at'][:16].replace('T', ' ')}"
    )
    ws["A2"].font = Font(italic=True, size=9, color="666666")

    fx_meta = data.get("fx") or {}
    applied = fx_meta.get("applied") or {}
    ws["A3"] = (
        f"Purchase orders converted to INR at USD {applied.get('USD', 'n/a')} / "
        f"CNY {applied.get('CNY', 'n/a')} "
        f"({fx_meta.get('source', 'unknown')}"
        f"{', as of ' + fx_meta['as_of'] if fx_meta.get('as_of') else ''})"
    )
    ws["A3"].font = Font(italic=True, size=9, color="666666")

    r = 4
    ws.cell(row=r, column=1, value="CASH POSITION").fill = _SECTION_FILL
    ws.cell(row=r, column=1).font = Font(bold=True)
    r += 1
    for label, val in (
        ("Cash & bank holdings", wc["cash"]["holdings"]),
        ("Credit card liabilities", -wc["cash"]["liabilities"]),
        ("Net cash", wc["cash"]["net_cash"]),
    ):
        ws.cell(row=r, column=1, value=label)
        c = ws.cell(row=r, column=2, value=val)
        c.number_format = _MONEY
        if label == "Net cash":
            ws.cell(row=r, column=1).font = Font(bold=True)
            c.font = Font(bold=True)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="PROJECTED FREE CASH").fill = _SECTION_FILL
    ws.cell(row=r, column=1).font = Font(bold=True)
    r += 1
    _write_header(ws, ["Horizon", "Opening cash", "Expected collections",
                       "Bills due", "Open PO commitment", "Operating expenses",
                       "Free cash"], row=r)
    r += 1
    for p in wc["projections"]:
        ws.cell(row=r, column=1, value=f"{p['horizon_days']} days")
        for col, key in enumerate(
            ["opening_cash", "expected_collections", "bills_due",
             "open_po_commitment", "operating_expenses", "free_cash"], start=2
        ):
            c = ws.cell(row=r, column=col, value=p[key])
            c.number_format = _MONEY
            if key == "free_cash":
                c.font = Font(bold=True)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="ALLOCATION").fill = _SECTION_FILL
    ws.cell(row=r, column=1).font = Font(bold=True)
    r += 1
    for label, val in (
        ("Purchase envelope", alloc["envelope"]),
        ("Reserved for new brands", -alloc.get("new_brand_reserve", 0.0)),
        ("Available for existing brands", alloc.get("available_for_existing_brands", alloc["envelope"])),
        ("Total requested by brands", alloc["total_requested"]),
        ("Funded", alloc["total_funded"]),
        ("Unfunded shortfall", alloc["unfunded"]),
        ("Cash trapped in EXCESS brands", alloc["cash_trapped_in_excess"]),
        (f"Trimmed by the {data.get('max_cover_days') or '—'}-day cover cap",
         alloc.get("capped_capital", 0.0)),
        (f"New brands held for review (< {data.get('min_selling_days')} days selling), indicative",
         alloc.get("new_brands_indicative", 0.0)),
    ) + tuple(
        (label, (data.get("inventory_value") or {}).get(key, 0.0))
        for label, key in (
            ("— Warehouse stock at MRP", "wh_mrp_value"),
            ("— Warehouse collection value (MRP ÷ 2)", "wh_collection_value"),
            ("— In-transit stock at MRP", "transit_mrp_value"),
            ("— In-transit collection value (MRP ÷ 2)", "transit_collection_value"),
            (
                "— Avg monthly sales, ex-GST ("
                + f"{(data.get('inventory_value') or {}).get('sales_window', {}).get('start_date', '')} to "
                + f"{(data.get('inventory_value') or {}).get('sales_window', {}).get('end_date', '')})",
                "avg_monthly_sales_ex_gst",
            ),
            ("— Avg monthly sales net of returns, ex-GST", "avg_monthly_sales_net_ex_gst"),
        )
    ):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val).number_format = _MONEY
        r += 1

    if data.get("new_brand_reserves"):
        r += 1
        ws.cell(row=r, column=1, value="NEW BRANDS (RESERVED)").fill = _SECTION_FILL
        ws.cell(row=r, column=1).font = Font(bold=True)
        r += 1
        _write_header(ws, ["Brand", "Reserved", "Expected order date", "Notes"], row=r)
        r += 1
        for nb in data["new_brand_reserves"]:
            ws.cell(row=r, column=1, value=nb["brand"])
            ws.cell(row=r, column=2, value=nb["amount"]).number_format = _MONEY
            ws.cell(row=r, column=3, value=nb.get("expected_order_date") or "")
            ws.cell(row=r, column=4, value=nb.get("notes") or "")
            r += 1

    if data.get("warnings"):
        r += 1
        ws.cell(row=r, column=1, value="WARNINGS").fill = PatternFill("solid", fgColor="FFF2CC")
        ws.cell(row=r, column=1).font = Font(bold=True)
        r += 1
        for w in data["warnings"]:
            ws.cell(row=r, column=1, value=w["message"] if isinstance(w, dict) else w)
            r += 1

    _autosize(ws, [34, 20, 22, 18, 22, 22, 18])

    # ── Sheet 2: brand plan ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Brand Plan")
    headers = [
        "Priority", "Brand", "Urgency", "Recommendation", "Lead Time",
        "Cover Cap (days)", "Days Selling", "DRR", "On Hand", "In Transit", "Days Coverage",
        "WH Units", "WH MRP Value", "WH Collection Value",
        "Transit MRP Value", "Transit Collection Value",
        "Sell-through DRR", "Months in WH", "Months in Transit",
        "Avg Monthly Sales, last 3 months (ex-GST)",
        "Avg Monthly Sales Net of Returns (ex-GST)",
        "Inventory at Cost", "Revenue", "Gross Margin", "Margin %", "GMROI",
        "SKUs to Order", "Suggested Units", "Order Units (capped)", "Units Trimmed",
        "Capital Required", "Funded Amount", "Funded %", "Cash Trapped",
    ]
    _write_header(ws2, headers)
    for i, b in enumerate(data["brands"], start=2):
        vals = [
            b.get("priority"), b["brand"], b["urgency"], b["recommendation"],
            b["lead_time"], b.get("cover_cap_days"), b.get("selling_days"), b["drr"],
            b["on_hand_units"], b["in_transit_units"], b["current_days_coverage"],
            b.get("wh_units"), b.get("wh_mrp_value"), b.get("wh_collection_value"),
            b.get("transit_mrp_value"), b.get("transit_collection_value"),
            b.get("sell_through_drr"), b.get("months_in_warehouse"), b.get("months_in_transit"),
            b.get("avg_monthly_sales_ex_gst"), b.get("avg_monthly_sales_net_ex_gst"),
            b["inventory_at_cost"], b["revenue"],
            b["gross_margin"], b["margin_pct"], b["gmroi"], b["skus_to_order"],
            b.get("suggested_units", b["order_units"]), b["order_units"],
            b.get("capped_units", 0), b["capital_required"], b["funded_amount"],
            b["funded_pct"], b["cash_trapped"],
        ]
        for col, v in enumerate(vals, start=1):
            c = ws2.cell(row=i, column=col, value=v)
            c.border = _THIN
            if col in (13, 14, 15, 16, 20, 21, 22, 23, 24, 31, 32, 34):
                c.number_format = _MONEY
        ws2.cell(row=i, column=3).fill = _URGENCY_FILL.get(b["urgency"], PatternFill())
    ws2.freeze_panes = "C2"
    _autosize(ws2, [
        9, 22, 11, 16, 11, 15, 12, 9, 11, 11, 14,
        11, 16, 18, 17, 20, 15, 13, 15, 20, 24,
        18, 16, 16, 10, 9, 13, 15, 18, 13, 18, 16, 11, 16,
    ])

    # ── Sheet 3: SKU detail for funded brands ───────────────────────────────
    ws3 = wb.create_sheet("SKU Detail")
    sku_headers = [
        "Brand", "Recommendation", "SKU Code", "Product Name", "Unit Cost",
        "Cost Source", "On Hand", "In Transit", "DRR", "Days Coverage",
        "Suggested Qty", "Order Qty (capped)", "Funded Qty", "Capital Required",
        "Excess / Order",
    ]
    _COST_SOURCE_LABEL = {
        "unit_price": "Managed unit price",
        "cogs": "COGS",
        "last_po": "Last PO rate (est)",
        "none": "Unknown",
    }
    _write_header(ws3, sku_headers)
    row = 2
    for b in data["brands"]:
        if b["recommendation"] in ("DO NOT ORDER",):
            continue
        for s in b["skus"]:
            vals = [
                b["brand"], b["recommendation"], s["sku_code"], s["product_name"],
                s["unit_cost"], _COST_SOURCE_LABEL.get(s.get("cost_source", ""), ""),
                s["on_hand"], s["in_transit"], s["drr"],
                s["current_days_coverage"], s.get("suggested_qty", s["order_qty"]),
                s["order_qty"], s.get("funded_qty", 0),
                s["capital_required"], s["excess_or_order"],
            ]
            for col, v in enumerate(vals, start=1):
                c = ws3.cell(row=row, column=col, value=v)
                c.border = _THIN
                if col in (5, 14):
                    c.number_format = _MONEY
            row += 1
    ws3.freeze_panes = "C2"
    _autosize(ws3, [20, 16, 16, 44, 12, 18, 11, 11, 9, 14, 13, 16, 12, 18, 15])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@router.get("/brand-order-plan/download")
async def download_brand_order_plan(
    start_date: str = Query(...),
    end_date: str = Query(...),
    usd_inr: Optional[float] = Query(None, gt=0, description="USD→INR override; omit for live rate"),
    cny_inr: Optional[float] = Query(None, gt=0, description="CNY→INR override; omit for live rate"),
    envelope_override: Optional[float] = Query(None),
    horizon_days: int = Query(90),
    max_cover_days: Optional[int] = Query(DEFAULT_MAX_COVER_DAYS, ge=0),
    include_new_brand_reserves: bool = Query(True),
    min_selling_days: int = Query(DEFAULT_MIN_SELLING_DAYS, ge=0),
    db=Depends(get_database),
):
    """Three-sheet workbook: cash bridge, brand plan, SKU detail."""
    try:
        data = await _brand_order_plan_data(
            start_date, end_date, db, usd_inr, cny_inr,
            envelope_override, horizon_days,
            max_cover_days or None, include_new_brand_reserves, min_selling_days,
        )
        content = await asyncio.to_thread(_build_plan_xlsx, data)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"brand-order-plan download failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to build workbook: {e}",
        )

    filename = f"brand_order_plan_{start_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
