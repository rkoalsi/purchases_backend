from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from datetime import datetime, timedelta, date
from ..database import get_database
import asyncio
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from typing import Optional
from pydantic import BaseModel
import logging

from .blinkit import compute_drr_blinkit, CITY_TO_STATE

logger = logging.getLogger(__name__)
router = APIRouter()

PROCESSING_COLLECTION = "blinkit_shipment_processing"
SUMMARY_COLLECTION = "blinkit_shipment_summary"
PLANNING_OVERRIDES_COLLECTION = "blinkit_shipment_planning_overrides"
SKU_COLLECTION = "blinkit_sku_mapping"

DEFAULT_LEAD_TIME = 10
DEFAULT_COVERAGE_DAYS = 30

MONTH_NAMES = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr",
    5: "May", 6: "Jun", 7: "Jul", 8: "Aug",
    9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}


# ─── Pydantic models ──────────────────────────────────────────────────────────

class PlanningOverride(BaseModel):
    sku_code: str
    open_shipment_qty_override: Optional[int] = None
    lead_time: Optional[int] = None
    coverage_days: Optional[int] = None
    final_units_override: Optional[int] = None


class SummaryUpdate(BaseModel):
    reason_for_short_supply: Optional[str] = None
    appointment_initiated_date: Optional[str] = None
    appointment_date: Optional[str] = None
    dispatched_date: Optional[str] = None
    status: Optional[str] = None


# ─── DRR (SKU-level, matching PSR methodology: per state then summed) ────────

def _compute_drr_per_sku(db, today: datetime, skus: list) -> dict:
    """
    Compute DRR per sku_code by replicating the PSR state-level methodology:
    1. Aggregate inventory by (sku, city, date) — city normalized to state via CITY_TO_STATE
    2. Aggregate sales by (sku, city, date) — city normalized to state
    3. Run compute_drr_blinkit per (sku, state)
    4. Sum state DRRs per sku to get total DRR (consistent with PSR showing per-state DRRs)
    Returns dict: sku_code -> total DRR (float)
    """
    drr_start = today - timedelta(days=89)

    inv_results = list(db["blinkit_inventory"].aggregate([
        {"$match": {"sku_code": {"$in": skus}, "date": {"$gte": drr_start, "$lte": today}}},
        {"$group": {
            "_id": {"sku": "$sku_code", "city": "$city", "date": "$date"},
            "closing_stock": {"$sum": "$warehouse_inventory"},
        }},
    ], allowDiskUse=True))

    sales_results = list(db["blinkit_sales"].aggregate([
        {"$match": {
            "sku_code": {"$in": skus},
            "order_date": {"$gte": drr_start, "$lte": today},
            "source": {"$ne": "return_orders"},
        }},
        {"$group": {
            "_id": {"sku": "$sku_code", "city": "$city", "date": "$order_date"},
            "units_sold": {"$sum": "$quantity"},
        }},
    ], allowDiskUse=True))

    returns_results = list(db["blinkit_sales"].aggregate([
        {"$match": {
            "sku_code": {"$in": skus},
            "order_date": {"$gte": drr_start, "$lte": today},
            "source": "return_orders",
        }},
        {"$group": {
            "_id": {"sku": "$sku_code", "city": "$city"},
            "total": {"$sum": "$quantity"},
        }},
    ], allowDiskUse=True))

    # Build (sku, state) -> {date_str: {date, closing_stock, units_sold}}
    state_daily: dict = {}
    for doc in inv_results:
        sku = doc["_id"]["sku"]
        city = doc["_id"]["city"]
        state = CITY_TO_STATE.get(city, city)
        key = (sku, state)
        date_val = doc["_id"]["date"]
        date_str = date_val.strftime("%Y-%m-%d") if isinstance(date_val, datetime) else str(date_val)[:10]
        if key not in state_daily:
            state_daily[key] = {}
        if date_str not in state_daily[key]:
            state_daily[key][date_str] = {"date": date_val, "closing_stock": 0, "units_sold": 0}
        state_daily[key][date_str]["closing_stock"] += doc["closing_stock"]

    for doc in sales_results:
        sku = doc["_id"]["sku"]
        city = doc["_id"]["city"]
        state = CITY_TO_STATE.get(city, city)
        key = (sku, state)
        date_val = doc["_id"]["date"]
        date_str = date_val.strftime("%Y-%m-%d") if isinstance(date_val, datetime) else str(date_val)[:10]
        if key not in state_daily:
            state_daily[key] = {}
        if date_str not in state_daily[key]:
            state_daily[key][date_str] = {"date": date_val, "closing_stock": 0, "units_sold": 0}
        state_daily[key][date_str]["units_sold"] += doc["units_sold"]

    # Build (sku, state) -> total returns
    state_returns: dict = {}
    for doc in returns_results:
        sku = doc["_id"]["sku"]
        city = doc["_id"]["city"]
        state = CITY_TO_STATE.get(city, city)
        key = (sku, state)
        state_returns[key] = state_returns.get(key, 0) + doc["total"]

    # Compute DRR per (sku, state), then sum per sku
    sku_drr: dict[str, float] = {}
    sku_state_drr: dict[str, dict[str, float]] = {}
    for (sku, state), date_dict in state_daily.items():
        daily_data = list(date_dict.values())
        returns = state_returns.get((sku, state), 0)
        drr_val, _, _, _ = compute_drr_blinkit(daily_data, returns)
        try:
            drr = float(drr_val)
        except (TypeError, ValueError):
            drr = 0.0
        sku_drr[sku] = round(sku_drr.get(sku, 0.0) + drr, 2)
        if drr > 0:
            if sku not in sku_state_drr:
                sku_state_drr[sku] = {}
            sku_state_drr[sku][state] = round(drr, 2)

    return sku_drr, sku_state_drr


# ─── Sync helper: fetch all planning rows ────────────────────────────────────

def _fetch_planning_data(db, today: datetime, drr_by_sku: dict, sku_state_drr: dict | None = None) -> tuple:
    inv_date_str = ""

    # 1. All SKUs from blinkit_sku_mapping
    sku_docs = list(db[SKU_COLLECTION].find({}, {"item_id": 1, "sku_code": 1, "item_name": 1, "_id": 0}))
    if not sku_docs:
        return [], "", ""

    skus = [d["sku_code"] for d in sku_docs if d.get("sku_code")]
    sku_to_item_id: dict[str, int] = {d["sku_code"]: d["item_id"] for d in sku_docs if d.get("sku_code") and d.get("item_id")}
    sku_to_item_name: dict[str, str] = {d["sku_code"]: d.get("item_name", "") for d in sku_docs if d.get("sku_code")}

    # 2. Current blinkit inventory (latest date, summed across all warehouses per SKU)
    latest_inv = db["blinkit_inventory"].find_one(
        {"sku_code": {"$in": skus}}, {"date": 1}, sort=[("date", -1)]
    )
    current_inv_by_sku: dict[str, int] = {}
    if latest_inv:
        inv_date = latest_inv["date"]
        inv_date_str = inv_date.strftime("%-d %b %Y") if isinstance(inv_date, datetime) else str(inv_date)
        for doc in db["blinkit_inventory"].aggregate([
            {"$match": {"sku_code": {"$in": skus}, "date": inv_date}},
            {"$group": {"_id": "$sku_code", "total": {"$sum": "$warehouse_inventory"}}},
        ]):
            current_inv_by_sku[doc["_id"]] = int(doc["total"] or 0)

    # 3. Open shipment qty from blinkit_shipment_processing by status
    #    processing → requested_qty, packed/intransit → packed_qty
    open_shipment_by_sku: dict[str, int] = {}
    for doc in db[PROCESSING_COLLECTION].aggregate([
        {"$match": {"sku_code": {"$in": skus}}},
        {"$group": {
            "_id": "$sku_code",
            "total": {
                "$sum": {
                    "$cond": {
                        "if": {"$eq": ["$status", "processing"]},
                        "then": {"$ifNull": ["$requested_qty", 0]},
                        "else": {
                            "$cond": {
                                "if": {"$in": ["$status", ["packed", "intransit"]]},
                                "then": {"$ifNull": ["$packed_qty", 0]},
                                "else": 0,
                            }
                        },
                    }
                }
            },
        }},
    ]):
        open_shipment_by_sku[doc["_id"]] = int(doc["total"] or 0)

    # 4. Zoho stock — look up products by sku_code, then zoho_warehouse_stock by zoho_item_id
    product_by_sku: dict[str, dict] = {}
    for p in db["products"].find(
        {"cf_sku_code": {"$in": skus}},
        {"cf_sku_code": 1, "status": 1, "item_id": 1, "_id": 0},
    ):
        product_by_sku[p["cf_sku_code"]] = p

    zoho_item_ids = [p["item_id"] for p in product_by_sku.values() if p.get("item_id")]
    zoho_latest: dict[str, int] = {}
    zoho_stock_date_str = ""
    if zoho_item_ids:
        latest_zoho = db["zoho_warehouse_stock"].find_one(
            {"zoho_item_id": {"$in": zoho_item_ids}},
            {"date": 1},
            sort=[("date", -1)],
        )
        if latest_zoho and latest_zoho.get("date"):
            zd = latest_zoho["date"]
            zoho_stock_date_str = zd.strftime("%-d %b %Y") if isinstance(zd, datetime) else str(zd)[:10]
        for doc in db["zoho_warehouse_stock"].aggregate([
            {"$match": {"zoho_item_id": {"$in": zoho_item_ids}}},
            {"$sort": {"date": -1}},
            {"$group": {
                "_id": "$zoho_item_id",
                "warehouses": {"$first": "$warehouses"},
            }},
        ]):
            wh = doc.get("warehouses", {})
            zoho_latest[doc["_id"]] = (
                int(sum(v for v in wh.values() if isinstance(v, (int, float))))
                if isinstance(wh, dict) else 0
            )

    # 5. Monthly sales (last 4 calendar months) from blinkit_sales
    months: list[tuple[int, int]] = []
    ref = today.replace(day=1)
    for _ in range(4):
        ref = (ref - timedelta(days=1)).replace(day=1)
        months.append((ref.year, ref.month))
    months.reverse()

    def month_bounds(y: int, m: int):
        s = datetime(y, m, 1)
        e = datetime(y + 1, 1, 1) if m == 12 else datetime(y, m + 1, 1)
        return s, e

    monthly_sales: dict[str, dict] = {s: {} for s in skus}
    for y, m in months:
        label = f"{MONTH_NAMES[m]} {y}"
        ms, me = month_bounds(y, m)
        for doc in db["blinkit_sales"].aggregate([
            {"$match": {"sku_code": {"$in": skus}, "order_date": {"$gte": ms, "$lt": me}}},
            {"$group": {"_id": "$sku_code", "total": {"$sum": "$quantity"}}},
        ]):
            monthly_sales[doc["_id"]][label] = int(doc["total"] or 0)

    month_labels = [f"{MONTH_NAMES[m]} {y}" for (y, m) in months]

    # 6. Planning overrides
    overrides: dict[str, dict] = {}
    for doc in db[PLANNING_OVERRIDES_COLLECTION].find({"sku_code": {"$in": skus}}):
        overrides[doc["sku_code"]] = doc

    # 7. Assemble rows (drr_by_sku passed in from async caller)
    rows = []
    for sku in skus:
        override = overrides.get(sku, {})
        prod = product_by_sku.get(sku, {})

        current_inv = current_inv_by_sku.get(sku, 0)
        open_shipment_auto = open_shipment_by_sku.get(sku, 0)
        open_shipment = (
            override["open_shipment_qty_override"]
            if "open_shipment_qty_override" in override and override["open_shipment_qty_override"] is not None
            else open_shipment_auto
        )
        total_inventory = current_inv + open_shipment

        drr = drr_by_sku.get(sku, 0.0)

        net_total_days = round(total_inventory / drr, 2) if drr > 0 else 0

        lead_time = override.get("lead_time") if override.get("lead_time") is not None else DEFAULT_LEAD_TIME
        coverage_days = override.get("coverage_days") if override.get("coverage_days") is not None else DEFAULT_COVERAGE_DAYS
        total_target_days = lead_time + coverage_days
        target_stock = round(drr * total_target_days, 0) if drr > 0 else 0

        if "final_units_override" in override and override["final_units_override"] is not None:
            final_units = override["final_units_override"]
        else:
            if drr == 0:
                final_units = 0
            elif net_total_days < lead_time:
                final_units = round(drr * total_target_days)
            elif net_total_days > total_target_days:
                final_units = 0
            else:
                final_units = max(0, round((total_target_days - net_total_days) * drr))

        # Zoho stock via product zoho_item_id
        zoho_item_id = prod.get("item_id", "")
        zoho_stock = zoho_latest.get(zoho_item_id, 0) if zoho_item_id else 0

        state_drr_for_sku = (sku_state_drr or {}).get(sku, {})
        rows.append({
            "item_id": sku_to_item_id.get(sku, ""),
            "sku_code": sku,
            "item_name": sku_to_item_name.get(sku, ""),
            "current_inventory": current_inv,
            "open_shipment_qty": open_shipment,
            "open_shipment_qty_auto": open_shipment_auto,
            "total_inventory": total_inventory,
            "drr": round(drr, 2),
            "state_drr": state_drr_for_sku,
            "net_total_days": net_total_days,
            "lead_time": lead_time,
            "coverage_days": coverage_days,
            "total_target_days": total_target_days,
            "target_stock": target_stock,
            "final_units_to_send": max(0, final_units),
            "zoho_stock": zoho_stock,
            "status": prod.get("status", ""),
            "monthly_sales": monthly_sales.get(sku, {}),
            "month_labels": month_labels,
            "open_shipment_overridden": "open_shipment_qty_override" in override and override["open_shipment_qty_override"] is not None,
            "final_units_overridden": "final_units_override" in override and override["final_units_override"] is not None,
        })

    return rows, inv_date_str, zoho_stock_date_str


def _generate_planning_excel(rows: list[dict], inv_date: str = "", drr_period: str = "") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Blinkit Shipment Planning"

    if not rows:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    month_labels = rows[0].get("month_labels", [])
    inv_label = f"Current Inventory\n({inv_date})" if inv_date else "Current Inventory"
    drr_label = f"DRR\n({drr_period})" if drr_period else "DRR"

    # Collect all unique states with DRR > 0, sorted alphabetically
    all_states = sorted({
        state
        for row in rows
        for state, val in row.get("state_drr", {}).items()
        if val and val > 0
    })

    # Columns: A=Item ID, B=SKU Code, C=Item Name, D=Current Inv, E=Open Shipment,
    # F=Total Inv, G=DRR, H=Net Days, I=Lead Time, J=Coverage, K=Target Days,
    # L=Target Stock, M=Final Units, N=Zoho Stock, O=Status, P+=Monthly Sales, then State DRRs
    headers = [
        "Item ID", "SKU Code", "Item Name",
        inv_label, "Open Shipment Qty",
        "Total Inventory\n(Current + Open Shipment)",
        drr_label, "Net Total Days\n(=Total Inv ÷ DRR)",
        "Lead Time", "Coverage Days",
        "Total Target Days\n(=Lead + Coverage)",
        "Target Stock\n(=DRR × Target Days)",
        "Final Units\n(Under-ordering formula)",
        "Zoho Stock", "Status",
    ] + [f"{label}\nUnits Sold" for label in month_labels] + [f"DRR\n{state}" for state in all_states]

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    ws.row_dimensions[1].height = 55

    for row_idx, row in enumerate(rows, 2):
        r = row_idx
        drr_val = row.get("drr") or None

        ws.cell(row=r, column=1, value=row.get("item_id") or None)
        ws.cell(row=r, column=2, value=row.get("sku_code", ""))
        c = ws.cell(row=r, column=3, value=row.get("item_name", "") or None)
        c.alignment = left_wrap
        ws.cell(row=r, column=4, value=row.get("current_inventory") or None)   # D
        ws.cell(row=r, column=5, value=row.get("open_shipment_qty") or None)   # E
        ws.cell(row=r, column=6, value=f"=D{r}+E{r}")                          # F
        ws.cell(row=r, column=7, value=drr_val)                                # G
        ws.cell(row=r, column=8, value=f'=IF(G{r}=0,"",ROUND(F{r}/G{r},2))')  # H
        ws.cell(row=r, column=9, value=row.get("lead_time", DEFAULT_LEAD_TIME))
        ws.cell(row=r, column=10, value=row.get("coverage_days", DEFAULT_COVERAGE_DAYS))
        ws.cell(row=r, column=11, value=f"=I{r}+J{r}")                         # K
        ws.cell(row=r, column=12, value=f"=IF(G{r}=0,0,ROUND(G{r}*K{r},0))")  # L
        ws.cell(row=r, column=13,
            value=f'=IF(G{r}=0,"",IF(H{r}<I{r},ROUND(G{r}*K{r},0),IF(H{r}>K{r},0,MAX(0,ROUND((K{r}-H{r})*G{r},0)))))')  # M
        ws.cell(row=r, column=14, value=row.get("zoho_stock") or None)
        ws.cell(row=r, column=15, value=row.get("status") or None)

        ms = row.get("monthly_sales", {})
        for offset, label in enumerate(month_labels):
            val = ms.get(label)
            ws.cell(row=r, column=16 + offset, value=val if val else None)

        state_drr = row.get("state_drr", {})
        state_col_start = 16 + len(month_labels)
        for offset, state in enumerate(all_states):
            val = state_drr.get(state)
            ws.cell(row=r, column=state_col_start + offset, value=val if val else None)

    col_widths = [12, 18, 50, 14, 14, 18, 10, 12, 9, 10, 14, 14, 14, 10, 12]
    col_widths += [14] * len(month_labels)
    col_widths += [14] * len(all_states)
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "C2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── Page 1: Planning ─────────────────────────────────────────────────────────

def _fetch_all_planning(db, today: datetime) -> tuple:
    """Single-threaded entry point: compute DRR then assemble rows."""
    skus = [d["sku_code"] for d in db[SKU_COLLECTION].find({}, {"sku_code": 1, "_id": 0}) if d.get("sku_code")]
    drr_by_sku, sku_state_drr = _compute_drr_per_sku(db, today, skus)
    rows, inv_date_str, zoho_stock_date_str = _fetch_planning_data(db, today, drr_by_sku, sku_state_drr)
    return rows, inv_date_str, zoho_stock_date_str


@router.get("/planning")
async def get_blinkit_planning(database=Depends(get_database)):
    try:
        today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        drr_start = today - timedelta(days=89)
        drr_period = f"{drr_start.strftime('%-d %b %Y')} – {today.strftime('%-d %b %Y')}"
        rows, inv_date, zoho_stock_date = await asyncio.to_thread(_fetch_all_planning, database, today)
        return {"rows": rows, "inv_date": inv_date, "drr_period": drr_period, "zoho_stock_date": zoho_stock_date}
    except Exception as e:
        logger.error(f"Error fetching Blinkit planning data: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/planning/download")
async def download_blinkit_planning(database=Depends(get_database)):
    try:
        today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        drr_start = today - timedelta(days=89)
        drr_period = f"{drr_start.strftime('%-d %b %Y')} – {today.strftime('%-d %b %Y')}"
        rows, inv_date, _zoho_date = await asyncio.to_thread(_fetch_all_planning, database, today)
        excel_bytes = _generate_planning_excel(rows, inv_date, drr_period)
        filename = f"blinkit_shipment_planning_{today.strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        logger.error(f"Error generating Blinkit planning Excel: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/planning/{sku_code}")
async def update_planning_override(
    sku_code: str, payload: PlanningOverride, database=Depends(get_database)
):
    try:
        update_doc: dict = {}
        if payload.open_shipment_qty_override is not None:
            update_doc["open_shipment_qty_override"] = payload.open_shipment_qty_override
        if payload.lead_time is not None:
            update_doc["lead_time"] = payload.lead_time
        if payload.coverage_days is not None:
            update_doc["coverage_days"] = payload.coverage_days
        if payload.final_units_override is not None:
            update_doc["final_units_override"] = payload.final_units_override

        if update_doc:
            def _upsert(db):
                db[PLANNING_OVERRIDES_COLLECTION].update_one(
                    {"sku_code": sku_code},
                    {"$set": {**update_doc, "sku_code": sku_code, "updated_at": datetime.utcnow()}},
                    upsert=True,
                )
            await asyncio.to_thread(_upsert, database)

        return {"success": True}
    except Exception as e:
        logger.error(f"Error updating planning override for {sku_code}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/planning/{sku_code}/override")
async def clear_planning_override(sku_code: str, field: str, database=Depends(get_database)):
    allowed = {"open_shipment_qty_override", "lead_time", "coverage_days", "final_units_override"}
    if field not in allowed:
        raise HTTPException(status_code=400, detail=f"Field {field} is not clearable")

    def _clear(db):
        db[PLANNING_OVERRIDES_COLLECTION].update_one(
            {"sku_code": sku_code},
            {"$unset": {field: ""}, "$set": {"updated_at": datetime.utcnow()}},
        )

    await asyncio.to_thread(_clear, database)
    return {"success": True}


# ─── Page 2: Processing ───────────────────────────────────────────────────────

def _fetch_processing_data(db) -> list[dict]:
    return list(db[PROCESSING_COLLECTION].find({}, {"_id": 0}))


@router.get("/processing")
async def get_blinkit_processing(database=Depends(get_database)):
    try:
        rows = await asyncio.to_thread(_fetch_processing_data, database)
        return {"rows": rows}
    except Exception as e:
        logger.error(f"Error fetching Blinkit processing data: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/processing/upload")
async def upload_blinkit_processing(
    file: UploadFile = File(...), database=Depends(get_database)
):
    """
    Upload Blinkit shipment processing sheet.
    Expected columns: RO Number, Date, Location, Item ID, UPC/EAN, SKU Code,
                      Item Name, MRP, SP, Requested Qty, Packed Qty, Status
    """
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active

        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise HTTPException(status_code=400, detail="Empty file")

        header_row = None
        header_idx = 0
        for i, row in enumerate(all_rows):
            row_clean = [str(c).strip().lower() if c else "" for c in row]
            if "ro number" in row_clean or "ro_number" in row_clean:
                header_row = row_clean
                header_idx = i
                break

        if header_row is None:
            raise HTTPException(
                status_code=400,
                detail="Could not find header row with 'RO Number' column",
            )

        def _find_col(*names):
            for name in names:
                try:
                    return header_row.index(name.lower())
                except ValueError:
                    pass
            return None

        idx_ro = _find_col("ro number", "ro_number")
        idx_date = _find_col("date")
        idx_location = _find_col("location")
        idx_item_id = _find_col("item id", "item_id")
        idx_upc = _find_col("upc/ean", "upc", "ean")
        idx_sku = _find_col("sku code", "sku_code")
        idx_item_name = _find_col("item name", "item_name")
        idx_mrp = _find_col("mrp")
        idx_sp = _find_col("sp")
        idx_requested_qty = _find_col("requested qty", "requested_qty")
        idx_packed_qty = _find_col("packed qty", "packed_qty")
        idx_status = _find_col("status")

        if idx_ro is None:
            raise HTTPException(status_code=400, detail="Missing 'RO Number' column")

        def _safe_num(val):
            try:
                return float(val) if val is not None and str(val).strip() not in ("", "None") else None
            except (TypeError, ValueError):
                return None

        def _safe_int(val):
            try:
                return int(float(val)) if val is not None and str(val).strip() not in ("", "None") else 0
            except (TypeError, ValueError):
                return 0

        def _safe_str(val):
            if val is None:
                return ""
            return str(val).strip()

        def _parse_date(val):
            if val is None:
                return None
            if isinstance(val, datetime):
                return val
            if isinstance(val, date):
                return datetime(val.year, val.month, val.day)
            try:
                return datetime.strptime(str(val).strip(), "%Y-%m-%d")
            except ValueError:
                try:
                    return datetime.strptime(str(val).strip(), "%d/%m/%Y")
                except ValueError:
                    return None

        records = []
        for raw in all_rows[header_idx + 1:]:
            if not any(raw):
                continue
            ro_number = _safe_str(raw[idx_ro] if idx_ro is not None else None)
            if not ro_number:
                continue

            record = {
                "ro_number": ro_number,
                "date": _parse_date(raw[idx_date] if idx_date is not None else None),
                "location": _safe_str(raw[idx_location] if idx_location is not None else None),
                "item_id": _safe_str(raw[idx_item_id] if idx_item_id is not None else None),
                "upc_ean": _safe_str(raw[idx_upc] if idx_upc is not None else None),
                "sku_code": _safe_str(raw[idx_sku] if idx_sku is not None else None),
                "item_name": _safe_str(raw[idx_item_name] if idx_item_name is not None else None),
                "mrp": _safe_num(raw[idx_mrp] if idx_mrp is not None else None),
                "sp": _safe_num(raw[idx_sp] if idx_sp is not None else None),
                "requested_qty": _safe_int(raw[idx_requested_qty] if idx_requested_qty is not None else None),
                "packed_qty": _safe_int(raw[idx_packed_qty] if idx_packed_qty is not None else None),
                "status": _safe_str(raw[idx_status] if idx_status is not None else None) or "processing",
                "uploaded_at": datetime.utcnow(),
            }
            records.append(record)

        if not records:
            raise HTTPException(status_code=400, detail="No data rows found in file")

        def _store(db):
            for rec in records:
                db[PROCESSING_COLLECTION].update_one(
                    {
                        "ro_number": rec["ro_number"],
                        "sku_code": rec["sku_code"],
                        "item_id": rec["item_id"],
                    },
                    {"$set": rec},
                    upsert=True,
                )

        await asyncio.to_thread(_store, database)
        return {"success": True, "rows_uploaded": len(records)}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading Blinkit processing data: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/processing")
async def clear_blinkit_processing(database=Depends(get_database)):
    def _clear(db):
        return db[PROCESSING_COLLECTION].delete_many({}).deleted_count

    deleted = await asyncio.to_thread(_clear, database)
    return {"success": True, "deleted": deleted}


# ─── Page 3: Summary ──────────────────────────────────────────────────────────

def _fetch_summary_data(db) -> list[dict]:
    pipeline = [
        {"$group": {
            "_id": {"ro_number": "$ro_number", "location": "$location"},
            "po_date": {"$first": "$date"},
            "requested_qty": {"$sum": "$requested_qty"},
            "accepted_qty": {"$sum": "$packed_qty"},
        }},
        {"$project": {
            "_id": 0,
            "ro_number": "$_id.ro_number",
            "location": "$_id.location",
            "po_date": 1,
            "requested_qty": 1,
            "accepted_qty": 1,
            "short_supply_qty": {"$subtract": ["$requested_qty", "$accepted_qty"]},
            "short_supply_pct": {
                "$cond": {
                    "if": {"$gt": ["$accepted_qty", 0]},
                    "then": {"$divide": [
                        {"$subtract": ["$requested_qty", "$accepted_qty"]},
                        "$accepted_qty",
                    ]},
                    "else": 0,
                }
            },
        }},
        {"$sort": {"po_date": -1, "ro_number": 1}},
    ]
    aggregated = list(db[PROCESSING_COLLECTION].aggregate(pipeline))

    stored_edits: dict[str, dict] = {}
    for doc in db[SUMMARY_COLLECTION].find({}, {"_id": 0}):
        key = f"{doc.get('ro_number')}|||{doc.get('location', '')}"
        stored_edits[key] = doc

    rows = []
    for row in aggregated:
        key = f"{row['ro_number']}|||{row.get('location', '')}"
        edit = stored_edits.get(key, {})
        rows.append({
            **row,
            "reason_for_short_supply": edit.get("reason_for_short_supply", ""),
            "appointment_initiated_date": edit.get("appointment_initiated_date", ""),
            "appointment_date": edit.get("appointment_date", ""),
            "dispatched_date": edit.get("dispatched_date", ""),
            "status": edit.get("status", "Pending"),
        })
    return rows


@router.get("/summary")
async def get_blinkit_summary(database=Depends(get_database)):
    try:
        rows = await asyncio.to_thread(_fetch_summary_data, database)
        return {"rows": rows}
    except Exception as e:
        logger.error(f"Error fetching Blinkit summary data: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/summary/{ro_number}")
async def update_blinkit_summary(
    ro_number: str,
    location: str,
    payload: SummaryUpdate,
    database=Depends(get_database),
):
    try:
        update_doc: dict = {
            "ro_number": ro_number,
            "location": location,
            "updated_at": datetime.utcnow(),
        }
        if payload.reason_for_short_supply is not None:
            update_doc["reason_for_short_supply"] = payload.reason_for_short_supply
        if payload.appointment_initiated_date is not None:
            update_doc["appointment_initiated_date"] = payload.appointment_initiated_date
        if payload.appointment_date is not None:
            update_doc["appointment_date"] = payload.appointment_date
        if payload.dispatched_date is not None:
            update_doc["dispatched_date"] = payload.dispatched_date
        if payload.status is not None:
            update_doc["status"] = payload.status

        def _upsert(db):
            db[SUMMARY_COLLECTION].update_one(
                {"ro_number": ro_number, "location": location},
                {"$set": update_doc},
                upsert=True,
            )

        await asyncio.to_thread(_upsert, database)
        return {"success": True}
    except Exception as e:
        logger.error(f"Error updating Blinkit summary: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
