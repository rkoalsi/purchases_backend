from fastapi import APIRouter, Body, HTTPException, status, Depends, Query, UploadFile, File
from fastapi.responses import JSONResponse, Response, StreamingResponse
from datetime import datetime, timedelta
from typing import List, Dict, Any, Set, Optional
import asyncio
import pandas as pd
import io
import json
import math
import os
import requests
import time
from collections import defaultdict
from ..database import get_database
import logging
import traceback
from ..services.master_service import OptimizedMasterReportService
from ..services.master_report import _generate_master_report_data
from .master_logistics import router as _logistics_router

logger = logging.getLogger(__name__)

_IV_BOOKS_URL = os.getenv("BOOKS_URL")
_IV_CLIENT_ID = os.getenv("CLIENT_ID")
_IV_CLIENT_SECRET = os.getenv("CLIENT_SECRET")
_IV_BOOKS_REFRESH_TOKEN = os.getenv("BOOKS_REFRESH_TOKEN")
_IV_ORG_ID = os.getenv("ORGANIZATION_ID", "776755316")
_PUPSCRIBE_WH_LOCATION_ID = "3220178000143298047"

router = APIRouter()
router.include_router(_logistics_router)

@router.get("/master-report")
async def get_master_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    brand: str = Query(None, description="Filter by brand name (e.g. 'Truelove'). Omit for all brands."),
    db=Depends(get_database),
):
    """
    Generate the master inventory & sales report for a given date range.

    ## Processing Pipeline

    ### 1. Parallel Data Fetching
    All data sources are fetched concurrently to minimise latency:
    - **Zoho** – the **sole source of sales data**. Covers invoices and credit notes
      from the Pupscribe Warehouse for the period. Blinkit and Amazon sales are
      intentionally excluded; all sales are tracked through Zoho.
    - **Amazon FBA closing stock** – fetched from `amazon_ledger` for inventory
      purposes only (not sales). Represents FBA inventory as of `end_date`.
    - **Composite products map** – bundle-to-component expansion rules
    - **Transfer orders** – inter-warehouse transfers for the period (excluded from net sales)
    - **Missed sales** – unfulfilled demand logged for the period

    ### 2. SKU & Product Data Resolution
    All SKU codes extracted from Zoho and FBA stock are batch-loaded from the
    products collection in a single query — fetching name, rate, brand, CBM,
    case pack, purchase status, and stock-in-transit fields.

    ### 3. Per-Source Normalisation
    Zoho records are converted to a canonical structure
    (`{sku_code, units_sold, units_returned, amount, days_in_stock, closing_stock, …}`).
    Zoho separates invoice units from credit-note units explicitly.

    ### 4. Data Combination & Composite Expansion
    Normalised records from all sources are merged by `sku_code`.
    Composite / bundle SKUs are exploded into their components using the
    composite products map; each component's quantities are scaled by its
    `quantity` multiplier. Combined metrics computed per SKU:
    - `total_units_sold`, `total_credit_notes`, `total_amount`
    - `total_closing_stock` = Zoho Pupscribe Warehouse stock + FBA stock
    - `avg_daily_run_rate` (DRR) = net units sold ÷ days in stock (or period days if always in stock)
    - `total_sales` = units_sold − credit_notes − transfer_orders

    ### 5. DRR Lookback (≥ 90-day reports only)
    For SKUs with fewer than 60 days in stock during the current period
    (insufficient data for a reliable DRR), the system looks back up to 6
    previous 90-day windows to find a period where the SKU had ≥ 60 days
    in stock. If found:
    - DRR is replaced with the historical value (`drr_source = "previous_period"`, `highlight = "yellow"`)
    - Missed sales for that SKU are re-fetched from **the same lookback date window** so that
      missed-sales figures are consistent with the DRR being used.
    - **Demand override (green):** if current-period net sales exceed lookback-period net sales,
      `order_qty` is set to current-period net sales, missed sales use the **current-period**
      window (not the lookback window), and `order_qty_demand_override = True` is set on the
      item. The lookback DRR is still used for all coverage / order-quantity calculations.
      (`highlight` remains `"yellow"` in the JSON but the row is coloured **green** in the
      Excel download.)
    - If no valid lookback period is found, the SKU is marked `drr_source = "insufficient_stock"`
      (`highlight = "red"`)

    ### 5a. Missed Sales Sourcing Rules (summary)
    | Row colour | DRR source | Missed Sales period |
    |---|---|---|
    | White (no fill) | `current_period` | Current period |
    | Yellow | `previous_period` (lookback ≤ 180 days back) | **Lookback period** |
    | Orange | `previous_period` (lookback > 180 days back) | **Lookback period** |
    | Red | `insufficient_stock` | Current period |
    | Green | `previous_period` + demand override | **Current period** (override takes precedence) |

    All customers are included in the missed-sales aggregation (no customer-name exclusions).

    ### 6. Brand Logistics Enrichment
    Lead time, safety days, and movement-class thresholds are loaded per brand
    from the `brand_logistics` collection and applied to each SKU.
    Open purchase orders are queried to derive stock-in-transit values
    (used as a fallback when the product record doesn't have manual SIT values).

    ### 7. Movement Classification
    Within each brand group, SKUs are ranked by volume and revenue percentiles:
    - **Fast Mover** – top tier by both volume and revenue
    - **Medium Mover** – mid tier
    - **Slow Mover** – bottom tier or zero movement

    ### 8. Order Quantity Calculations
    For each SKU (skipped for inactive / discontinued items):
    - `target_days` = lead_time + safety_days + order_processing (per-brand, default 10 days)
    - `current_days_coverage` = (closing_stock + stock_in_transit) ÷ DRR
    - `net_target_days` = target_days − current_days_coverage (floored at target_days if already under lead time)
    - `order_qty` = max(0, net_target_days × DRR)
    - `extra_qty` = missed_sales_drr × lead_time (capped at 50 % of DRR to prevent spikes)
    - `order_qty_plus_extra_qty_rounded` = rounded down to nearest case pack multiple
    - `total_cbm` = (rounded_qty ÷ case_pack) × CBM per case

    **Confidence-tier dampening** (applies only to `insufficient_stock` items):

    | Days in stock (current period) | Multiplier | `order_qty_basis` label |
    |---|---|---|
    | < 30 | 0 × (order = 0) | `"Manual buyer input required (0%)"` |
    | 30 – 44 | 0.5 × | `"System order, dampened (50%)"` |
    | 45 – 59 | 0.75 × | `"System order, near-normal (75%)"` |
    | ≥ 60 | 1.0 × (no dampening) | `""` (blank) |

    For green (demand-override) items that are also `insufficient_stock`, both labels are
    combined: e.g. `"Current Period Sales · System order, dampened (50%)"`.
    For demand-override items without dampening the label is simply `"Current Period Sales"`.

    ### 9. Latest Stock Injection
    Real-time stock (latest record in DB, regardless of report period) is attached
    to each SKU (`latest_zoho_stock`, `latest_fba_stock`, `latest_total_stock`).
    All coverage and order calculations are then **recomputed** using this latest
    stock so recommendations reflect current inventory, not end-of-period snapshot.

    ### 10. FBA-Only Stub Rows
    SKUs that have FBA stock as of `end_date` but zero sales in the period are
    injected as stub rows so they appear in the report with their stock and
    order-coverage data even though no sales were recorded.

    ### 11. Growth Rate & Return %
    Calculated using a background-fetched 90-day trailing DRR compared against
    the report-period DRR. Return % = total_units_returned ÷ total_units_sold.

    ## Response
    Returns a JSON object with:
    - **`combined_data`** – array of enriched SKU objects (see field list below)
    - **`summary`** – totals/averages across all SKUs
    - **`individual_reports`** – per-source metadata (success, record count, errors)
    - **`latest_stock_dates`** – date of the most recent stock record in DB for Zoho and FBA
    - **`meta`** – execution time and timestamp

    ## Key Fields on Each SKU in `combined_data`
    | Field | Description |
    |---|---|
    | `sku_code` | Internal SKU identifier |
    | `avg_daily_run_rate` | Units sold per day (DRR) |
    | `drr_source` | `current_period` / `previous_period` / `insufficient_stock` |
    | `drr_lookback_period` | Date range used for lookback DRR (if applicable) |
    | `drr_lookback_days_in_stock` | Days in stock during the lookback period (if applicable) |
    | `drr_lookback_sales` | Gross units sold during the lookback period (if applicable) |
    | `drr_lookback_returns` | Credit note units during the lookback period (if applicable) |
    | `drr_net_lookback_sales` | Net units (sales − returns) used for lookback DRR calculation |
    | `drr_lookback_gt_180` | `true` when the lookback period start is > 180 days before the report start |
    | `highlight` | Row highlight hint: `null` (normal) / `"yellow"` (lookback DRR ≤ 180d) / `"orange"` (lookback DRR > 180d) / `"red"` (insufficient stock) |
    | `order_qty_demand_override` | `true` when current-period net sales exceed lookback net sales — order qty is overridden to current-period net sales; these rows are coloured **green** in the Excel download |
    | `current_period_sales_for_override` | Net sales used as order qty when demand override is active |
    | `total_closing_stock` | Zoho Pupscribe Warehouse stock + FBA stock as of `end_date` |
    | `latest_total_stock` | Zoho Pupscribe Warehouse stock + FBA stock as of latest DB record |
    | `current_days_coverage` | Days of inventory on hand at current DRR |
    | `order_qty_plus_extra_qty_rounded` | Final suggested order quantity (case-pack rounded) |
    | `missed_sales` | Unfulfilled demand units — sourced from the current period for white/red/green rows, and from the lookback window for yellow/orange rows |
    | `order_qty_basis` | Explanation of any non-standard order qty treatment (demand override label, dampening tier with %, or blank for normal rows) |
    | `confidence_multiplier` | Dampening factor applied to order_qty for `insufficient_stock` items (0.0 / 0.5 / 0.75 / 1.0) |
    | `movement` | `Fast Mover` / `Medium Mover` / `Slow Mover` |
    | `excess_or_order` | `ORDER` / `EXCESS` / `NO MOVEMENT` |
    """
    content = await _generate_master_report_data(
        start_date, end_date, db, brand=brand,
    )

    # Strip raw data from individual_reports for the API response (keep metadata only)
    content["individual_reports"] = {
        source: {
            "source": report.get("source", source),
            "success": report.get("success", False),
            "record_count": len(report.get("data", [])),
            "error": report.get("error"),
        }
        for source, report in content["individual_reports"].items()
    }

    return JSONResponse(status_code=status.HTTP_200_OK, content=content)


@router.get("/master-report/download")
async def download_master_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    brand: str = Query(None, description="Filter by brand name (e.g. 'Truelove'). Omit for all brands."),
    include_cogs: bool = Query(False, description="Include COGS columns (Unit Cost, Pupscribe Stock Value, Pupscribe Stock on Hand) from Zoho Books inventory valuation"),
    db=Depends(get_database),
):
    """
    Download the master report as an Excel (.xlsx) file.

    Runs the exact same processing pipeline as `GET /master-report` and writes
    the result to a multi-sheet Excel workbook streamed back as an attachment.

    > **Note on data sources:** Sales data comes exclusively from **Zoho** (Pupscribe
    > Warehouse invoices and credit notes). Blinkit and Amazon sales are not included.
    > Amazon FBA is used only for closing stock / inventory figures.

    ## Excel Workbook Structure

    ### Master Sheet (`{brand} {start} to {end}` or `Master {start} to {end}`)
    One row per SKU. Columns are grouped as follows:

    **Identity**
    - `Purchase Status` – active / inactive / discontinued until stock lasts
    - `Is New` – Yes if the product was created within the report window
    - `SKU Code`, `Item Name`

    **Sales Metrics**
    - `Total Amount` – gross revenue for the period
    - `Total Units Sold` – gross units sold across all channels
    - `Total Units Returned` – customer returns
    - `Transfer Orders` – inter-warehouse transfers (excluded from net sales)
    - `Net Total Sales` – units_sold − returns − transfer_orders
    - `Return %`

    **DRR & Stock Quality**
    - `Days in Stock` – days the SKU was in stock during the period
    - `Lookback Days in Stock` – days in stock during the lookback period (if used)
    - `Lookback Sales` – units sold during the lookback period (if used)
    - `Avg Daily Run Rate` – DRR used for order calculations
    - `Growth Rate (%)` – change vs. trailing 90-day DRR
    - `DRR Source` – `current_period` / `previous_period` / `insufficient_stock`
    - `DRR Lookback Period` – date range of the lookback window (e.g. `2025-09-01 to 2025-12-01`)

    **Stock Snapshot (period end)**
    - `Total Stock ({end_date})` – Zoho Pupscribe Warehouse stock + FBA stock as of `end_date`
    - `Pupscribe WH Stock ({prev_zoho_date})` – Zoho Pupscribe Warehouse stock from the snapshot one day before the latest DB record
    - `FBA Stock ({end_date})`

    **Stock Snapshot (latest DB record)**
    - `Total Stock ({latest_date})` – Zoho Pupscribe Warehouse stock + FBA stock (most recent record in DB, used for order calculations)
    - `Pupscribe WH Stock ({latest_zoho_date})` – Zoho Pupscribe Warehouse stock as of latest DB record
    - `FBA Stock ({latest_fba_date})`

    **Inventory Status**
    - `In Stock` – Yes / No as of the latest record
    - `Movement` – Fast / Medium / Slow Mover (classified within brand group)

    **Order Parameters**
    - `Safety Days`, `Lead Time`, `Order Processing` (per-brand from brand_logistics, default 10 days), `Target Days`
    - `On-Hand Days Coverage` – days of stock based on period-end stock
    - `Stock in Transit 1 / 2 / 3`, `Total Stock in Transit`
    - `Current Days Coverage` – (latest stock + transit) ÷ DRR

    **Missed Sales**
    - `Missed Sales` – unfulfilled demand units. Sourced from the **current period** for white,
      red, and green rows; sourced from the **lookback period** (same window as the DRR) for
      yellow and orange rows. All customers are included.
    - `Missed Sales DRR` – missed units ÷ period days, capped at 50 % of the item's DRR to
      prevent over-ordering from short-term stockout spikes
    - `Extra Qty` – additional buffer = missed_sales_drr × lead_time

    **Order Quantities**
    - `Net Target Days` – target_days − current_days_coverage
    - `Excess / Order` – ORDER / EXCESS / NO MOVEMENT
    - `Order Qty` – net_target_days × DRR (or current-period net sales for demand-override rows),
      then multiplied by the confidence multiplier for `insufficient_stock` items
    - `Order Qty + Extra Qty` – order_qty + extra_qty
    - `CBM`, `Case Pack`
    - `Order Qty + Extra Qty (Rounded down)` – floored to nearest case pack multiple (green fill when
      demand override was applied)
    - `Total CBM` – (rounded_qty ÷ case_pack) × CBM per case
    - `Days Current Order Lasts` – rounded_qty ÷ DRR
    - `Days Total Inventory Lasts` – current_days_coverage + days_current_order_lasts
    - `Order Qty Basis` (col AW) – human-readable label explaining any non-standard treatment:

      | Label | Meaning |
      |---|---|
      | `"Current Period Sales"` | Green demand-override row; order qty = current-period net sales |
      | `"Manual buyer input required (0%)"` | Red row, < 30 days in stock; order qty set to 0 |
      | `"System order, dampened (50%)"` | Red row, 30–44 days in stock; order qty halved |
      | `"System order, near-normal (75%)"` | Red row, 45–59 days in stock; order qty at 75 % |
      | `"Current Period Sales · <tier>"` | Green row that is also dampened (combined label) |
      | *(blank)* | Normal row — no special treatment needed |

    ## Conditional Row Formatting
    | Colour | Condition | Missed Sales source | Meaning |
    |---|---|---|---|
    | Green | `order_qty_demand_override = true` (takes precedence) | Current period | Lookback DRR used but current-period net sales exceed lookback net sales — order qty overridden to current-period net sales |
    | Orange | `drr_source = "previous_period"` and lookback period start > 180 days before report start | Lookback period | DRR from a stale lookback window (> 180 days in the past) |
    | Yellow | `drr_source = "previous_period"` and lookback ≤ 180 days back | Lookback period | DRR from a recent lookback period; SKU had < 60 days in stock in the current period |
    | Red | `drr_source = "insufficient_stock"` | Current period | No reliable DRR — fewer than 60 days in stock across the current period and all 6 lookback windows; order qty dampened by confidence tier |
    | (no fill) | `drr_source = "current_period"` | Current period | Normal row — ≥ 60 days in stock, DRR derived from current period |

    ## Excel Formulas
    Several columns are written as live Excel formulas rather than static values
    so the sheet remains interactive:
    - `Missed Sales DRR` = Missed Sales ÷ period days
    - `Extra Qty` = Missed Sales DRR × Lead Time
    - `Net Target Days` = Target Days − Current Days Coverage
    - `Order Qty` = max(0, Net Target Days × Avg DRR)
    - `Order Qty + Extra Qty` = Order Qty + Extra Qty
    - `Order Qty + Extra Qty (Rounded down)` = FLOOR to case pack
    - `Total CBM` = (Rounded Qty ÷ Case Pack) × CBM
    - `Days Current Order Lasts` = Rounded Qty ÷ DRR
    - `Days Total Inventory Lasts` = Current Days Coverage + Days Current Order Lasts

    ## Response
    Returns a binary `.xlsx` file as a streaming attachment.
    Filename format: `master_report_{start_date}_to_{end_date}_{YYYYMMDD_HHMMSS}.xlsx`
    """
    try:
        # Get the full master report data (with raw individual_reports for Excel sheets)
        content = await _generate_master_report_data(
            start_date=start_date,
            end_date=end_date,
            db=db,
            brand=brand,
            include_cogs=include_cogs,
        )

        combined_data = content.get("combined_data", [])

        if not combined_data:
            if not brand:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="No data found to download",
                )
            # Brand filter active but no data yet (e.g. brand-new brand with no products
            # in DB). Fall through so the Draft Order sheet is generated (blank).

        # Build a readable date-range sheet name (Excel max 31 chars) and stock column labels
        try:
            _sd = datetime.strptime(start_date, "%Y-%m-%d")
            _ed = datetime.strptime(end_date, "%Y-%m-%d")
            if _sd.year == _ed.year:
                _sheet_name = f"{_sd.strftime('%d %b')} - {_ed.strftime('%d %b %Y')}"
            else:
                _sheet_name = f"{_sd.strftime('%d %b %Y')} - {_ed.strftime('%d %b %Y')}"
            _sheet_name = _sheet_name[:31]
            _end_date_label = _ed.strftime("%d %b %Y")
            _period_days = (_ed - _sd).days + 1
        except Exception:
            _sheet_name = "Master Sheet"
            _period_days = 30
            _end_date_label = end_date

        # Latest stock column labels from the actual latest dates in the DB
        latest_stock_dates = content.get("latest_stock_dates", {})
        def _fmt_date(d: str) -> str:
            try:
                return datetime.strptime(d, "%Y-%m-%d").strftime("%d %b %Y")
            except Exception:
                return d or "Latest"
        _latest_zoho_label = _fmt_date(latest_stock_dates.get("zoho", ""))
        _prev_zoho_label   = _fmt_date(latest_stock_dates.get("prev_zoho", ""))
        _latest_fba_label  = _fmt_date(latest_stock_dates.get("fba", ""))
        _prev_fba_label    = _fmt_date(latest_stock_dates.get("prev_fba", ""))
        # For total, use the more recent of the two
        _latest_total_label = _latest_fba_label if latest_stock_dates.get("fba", "") >= latest_stock_dates.get("zoho", "") else _latest_zoho_label
        _prev_total_label   = _prev_fba_label if latest_stock_dates.get("prev_fba", "") >= latest_stock_dates.get("prev_zoho", "") else _prev_zoho_label
        _etrade_inv_label = _fmt_date(latest_stock_dates.get("etrade", "")) or _end_date_label
        _blinkit_inv_label = _fmt_date(latest_stock_dates.get("blinkit", "")) or _end_date_label
        _cogs_date_label = _fmt_date(latest_stock_dates.get("cogs", "")) or _end_date_label

        # Create Excel file
        excel_buffer = io.BytesIO()

        # Track which Excel data rows (0-based) are demand-override (green) rows
        demand_override_row_indices: set = set()
        lookback_gt_180_row_indices: set = set()

        # currency_code → Excel number format using symbol prefix (shared by master + draft sheets)
        _CURRENCY_FORMATS: Dict[str, str] = {
            "USD": '$#,##0.00',
            "EUR": '€#,##0.00',
            "GBP": '£#,##0.00',
            "CNY": '¥#,##0.00',
            "INR": '₹#,##0.00',
        }
        _DEFAULT_NUMBER_FMT = "#,##0.00"

        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            # ── Instructions Sheet (first sheet) ─────────────────────────────────
            from openpyxl import Workbook as _OWB
            from openpyxl.styles import (
                Font as _IFont,
                PatternFill as _IPF,
                Alignment as _IAlign,
                Border as _IBorder,
                Side as _ISide,
            )
            from openpyxl.utils import get_column_letter as _gcl

            _instr_ws = writer.book.create_sheet("Instructions", 0)

            # ── colour palette ────────────────────────────────────────────────────
            _hdr_fill   = _IPF(start_color="1F4E79", end_color="1F4E79", fill_type="solid")   # dark blue
            _sec_fill   = _IPF(start_color="2E75B6", end_color="2E75B6", fill_type="solid")   # mid blue
            _alt_fill   = _IPF(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")   # light blue
            _grn_fill   = _IPF(start_color="90EE90", end_color="90EE90", fill_type="solid")
            _yel_fill   = _IPF(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            _org_fill   = _IPF(start_color="FFA500", end_color="FFA500", fill_type="solid")
            _red_fill   = _IPF(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            _wht_fill   = _IPF(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            _thin  = _ISide(style="thin",   color="AAAAAA")
            _thick = _ISide(style="medium", color="1F4E79")
            _border_all = _IBorder(left=_thin, right=_thin, top=_thin, bottom=_thin)

            def _iw(row, col, value, bold=False, italic=False, size=11,
                    color="000000", fill=None, wrap=True, align="left", border=True):
                cell = _instr_ws.cell(row=row, column=col, value=value)
                cell.font = _IFont(bold=bold, italic=italic, size=size, color=color)
                cell.alignment = _IAlign(wrap_text=wrap, vertical="top", horizontal=align)
                if fill:
                    cell.fill = fill
                if border:
                    cell.border = _border_all
                return cell

            # ── column widths ─────────────────────────────────────────────────────
            _instr_ws.column_dimensions["A"].width = 28
            _instr_ws.column_dimensions["B"].width = 70
            _instr_ws.column_dimensions["C"].width = 18

            # ── merge A1:C1 for the main title ────────────────────────────────────
            _instr_ws.merge_cells("A1:C1")
            _iw(1, 1, "📦  Pupscribe Master Purchase Report — How to Read This File",
                bold=True, size=14, color="FFFFFF", fill=_hdr_fill, align="center", border=False)
            _instr_ws.row_dimensions[1].height = 30

            # ── Section 1: Overview ───────────────────────────────────────────────
            _instr_ws.merge_cells("A2:C2")
            _iw(2, 1, "OVERVIEW", bold=True, size=11, color="FFFFFF", fill=_sec_fill, align="center")

            _overview_rows = [
                ("What this report does",
                 "Shows sales, stock, and recommended order quantities for every active SKU "
                 "across Zoho (Pupscribe WH), Amazon FBA, and Blinkit for the selected date range. "
                 "Each row is one SKU. The rightmost columns contain the system-calculated Order Qty."),
                ("Date range",
                 f"Report period: {start_date} → {end_date}  ({_period_days} days)"),
                ("Sheets in this file",
                 "Instructions (this sheet) · Master data sheet (named after the date range) · "
                 "Draft Order sheet (when a Brand filter is active)"),
            ]
            for _ri, (_lbl, _desc) in enumerate(_overview_rows, start=3):
                _fill = _alt_fill if _ri % 2 == 1 else _wht_fill
                _iw(_ri, 1, _lbl, bold=True, fill=_fill)
                _iw(_ri, 2, _desc, fill=_fill)
                _instr_ws.merge_cells(f"B{_ri}:C{_ri}")
                _instr_ws.row_dimensions[_ri].height = 40

            _next_row = 3 + len(_overview_rows)

            # ── Section 2: Row colour legend ──────────────────────────────────────
            _instr_ws.merge_cells(f"A{_next_row}:C{_next_row}")
            _iw(_next_row, 1, "ROW COLOUR LEGEND", bold=True, size=11,
                color="FFFFFF", fill=_sec_fill, align="center")
            _next_row += 1

            _colour_rows = [
                (_wht_fill, "000000", "White",  "current_period",
                 "DRR is calculated from current-period sales (normal). "
                 "Item was in Pupscribe WH ≥ 60 days."),
                (_yel_fill, "000000", "Yellow", "previous_period (≤ 180 days ago)",
                 "Item was in stock < 60 days this period. DRR is taken from a recent lookback "
                 "window (within the last 180 days). Use with moderate confidence."),
                (_org_fill, "000000", "Orange", "previous_period (> 180 days ago)",
                 "Lookback window is older than 180 days. DRR may not reflect current demand. "
                 "Review manually before ordering."),
                (_grn_fill, "000000", "Green",  "demand_override",
                 "Current-period sales exceed the lookback DRR, so the Order Qty is set to "
                 "current-period Net Sales rather than DRR × Target Days. Demand is accelerating."),
                (_red_fill, "000000", "Red",    "insufficient_stock",
                 "Insufficient stock history to compute a reliable DRR. "
                 "A confidence dampening multiplier is applied to the Order Qty (see DRR Flag column)."),
            ]
            for _cr in _colour_rows:
                _cfill, _cfont, _cname, _csrc, _cdesc = _cr
                _iw(_next_row, 1, _cname, bold=True, fill=_cfill, color=_cfont)
                _iw(_next_row, 2, f"DRR Source: {_csrc}", fill=_cfill, color=_cfont, italic=True)
                _iw(_next_row, 3, "", fill=_cfill, border=True)
                _instr_ws.merge_cells(f"B{_next_row}:C{_next_row}")
                _next_row += 1
                _iw(_next_row, 1, "", fill=_cfill, border=True)
                _instr_ws.cell(row=_next_row, column=1).fill = _cfill
                _iw(_next_row, 2, _cdesc, fill=_wht_fill, wrap=True)
                _instr_ws.merge_cells(f"B{_next_row}:C{_next_row}")
                _instr_ws.row_dimensions[_next_row].height = 45
                _next_row += 1

            # ── Section 3: Key column guide ───────────────────────────────────────
            _instr_ws.merge_cells(f"A{_next_row}:C{_next_row}")
            _iw(_next_row, 1, "KEY COLUMN GUIDE", bold=True, size=11,
                color="FFFFFF", fill=_sec_fill, align="center")
            _next_row += 1

            _col_guide = [
                ("Purchase Status",        "active / inactive / discontinued until stock lasts. "
                                           "Inactive/discontinued rows get Order Qty = 0 (except Missed Sales extra qty)."),
                ("Is New",                 "Yes = item launched within this report period. "
                                           "New items with zero sales show 'NO MOVEMENT'."),
                ("Avg Daily Run Rate",     "Units sold per day. "
                                           "White/Green/Red rows: Net Total Sales ÷ Days in Stock. "
                                           "Yellow/Orange rows: Net Lookback Sales ÷ Lookback Days in Stock."),
                ("Net Total Sales",        "Total Units Sold − Total Units Returned − Transfer Orders. "
                                           "This is the demand figure used for White/Red DRR."),
                ("Net Lookback Sales",     "MAX(0, Lookback Sales − Lookback Returns). "
                                           "Used as the numerator for DRR on Yellow/Orange rows, "
                                           "and for the Growth Rate formula on all row colours."),
                ("Lookback Return %",      "Lookback Returns ÷ Lookback Sales × 100. "
                                           "Return rate during the comparison period."),
                ("Growth Rate (%)",        "(Net Total Sales − Net Lookback Sales) ÷ Net Lookback Sales × 100. "
                                           "Populated for all row colours once comparison-period data is available. "
                                           "Blank when no lookback sales exist."),
                ("DRR Source",             "current_period · previous_period · insufficient_stock. "
                                           "Drives both the row colour and the DRR formula used."),
                ("DRR Lookback Period",    "Date range of the lookback window used for Yellow/Orange rows."),
                ("Days in Stock (PW)",     "Number of days the item was physically in stock at Pupscribe WH "
                                           "during this report period. Used as the DRR denominator for White rows."),
                ("Lookback Days in Stock", "Days in stock during the lookback period. DRR denominator for Yellow/Orange rows."),
                ("On-Hand Days Coverage",  "Latest Total Stock ÷ DRR. How many days the current physical stock will last."),
                ("Current Days Coverage",  "(Latest Total Stock + Total Stock in Transit) ÷ DRR."),
                ("Target Days",            "Lead Time + Safety Days + Order Processing (days). "
                                           "The coverage target the order is sized to achieve."),
                ("Net Target Days",        "IF(Current Coverage < Lead Time, Target Days, Target Days − Current Coverage). "
                                           "The shortfall in days that needs to be replenished."),
                ("Missed Sales",           "Units estimated as lost due to stockouts during this period."),
                ("Missed Sales DRR",       "MIN(Missed Sales ÷ Period Days, 0.5 × DRR). "
                                           "Caps missed-sales uplift at 50 % of DRR to prevent over-ordering."),
                ("Extra Qty",              "Missed Sales DRR × Lead Time. Added on top of Order Qty."),
                ("Excess / Order",         "ORDER = stock below target; EXCESS = already covered; NO MOVEMENT = DRR is 0."),
                ("Order Qty",              "For EXCESS rows: 0. For ORDER rows: MAX(0, Net Target Days × DRR) × Confidence Multiplier. "
                                           "Green rows use Net Total Sales × Confidence Multiplier."),
                ("Confidence Multiplier",  "1.0 for normal rows. Reduced (0 %, 50 %, 75 %) for Red rows based on "
                                           "how little stock history is available (see DRR Flag)."),
                ("DRR Flag",               "Empty for normal rows. Shows dampening label and/or seasonal mismatch warning "
                                           "for Red/uncertain rows."),
                ("Order Qty + Extra Qty (Rounded down)", "Final recommended order quantity, floored to the nearest Case Pack. "
                                           "If the floor would round to 0 and status is ORDER, bumped to 1 case pack."),
                ("Amazon DRR",             "Final DRR (units/day) from the Amazon PSR report for the same date range — "
                                           "identical to the 'Final DRR' column in the Amazon Sales vs Inventory report. "
                                           "Combines VC + FBA sales. 0 when the SKU has no Amazon sales history."),
                ("Days Total Inventory Lasts (Amazon)", "Amazon Total Inventory ÷ Amazon DRR. "
                                           "How many days the combined Amazon stock (VC + FBA) covers at the PSR DRR. "
                                           "0 when Amazon DRR is 0."),
                ("Total CBM",              "Cubic metres for the rounded order: (Order Rounded ÷ Case Pack) × CBM per case."),
            ]
            for _ri2, (_col_name, _col_desc) in enumerate(_col_guide, start=0):
                _rr = _next_row + _ri2
                _cfill2 = _alt_fill if _ri2 % 2 == 0 else _wht_fill
                _iw(_rr, 1, _col_name, bold=True, fill=_cfill2)
                _iw(_rr, 2, _col_desc, fill=_cfill2, wrap=True)
                _instr_ws.merge_cells(f"B{_rr}:C{_rr}")
                _instr_ws.row_dimensions[_rr].height = 45
            _next_row += len(_col_guide)

            # ── Section 4: Ordering logic summary ────────────────────────────────
            _instr_ws.merge_cells(f"A{_next_row}:C{_next_row}")
            _iw(_next_row, 1, "ORDERING LOGIC SUMMARY", bold=True, size=11,
                color="FFFFFF", fill=_sec_fill, align="center")
            _next_row += 1

            _logic_rows = [
                ("Step 1 – Compute DRR",
                 "White/Green/Red: Net Total Sales ÷ Days in Stock (or ÷ Period Days if no stock days). "
                 "Yellow/Orange: Net Lookback Sales ÷ Lookback Days in Stock."),
                ("Step 2 – Check coverage",
                 "Current Days Coverage = (Latest Stock + Transit) ÷ DRR. "
                 "If Coverage ≥ Target Days → EXCESS (no order needed)."),
                ("Step 3 – Size order",
                 "Net Target Days = Target Days − Current Coverage (min: 0 when coverage ≥ Lead Time). "
                 "Order Qty = Net Target Days × DRR × Confidence Multiplier."),
                ("Step 4 – Add missed-sales buffer",
                 "Extra Qty = Missed Sales DRR × Lead Time. "
                 "Final qty = Order Qty + Extra Qty, floored to Case Pack."),
                ("Step 5 – Demand override (Green rows)",
                 "If current-period Net Sales > lookback DRR × period days, the order is set to "
                 "Net Total Sales × Confidence Multiplier instead — demand acceleration takes precedence."),
                ("Inactive / discontinued",
                 "Order Qty is forced to 0. Only the Missed Sales Extra Qty is preserved for "
                 "'discontinued until stock lasts' items."),
            ]
            for _ri3, (_step, _desc3) in enumerate(_logic_rows, start=0):
                _rr3 = _next_row + _ri3
                _cfill3 = _alt_fill if _ri3 % 2 == 0 else _wht_fill
                _iw(_rr3, 1, _step, bold=True, fill=_cfill3)
                _iw(_rr3, 2, _desc3, fill=_cfill3, wrap=True)
                _instr_ws.merge_cells(f"B{_rr3}:C{_rr3}")
                _instr_ws.row_dimensions[_rr3].height = 45
            _next_row += len(_logic_rows)

            # ── Section 5: Stock sources ──────────────────────────────────────────
            _instr_ws.merge_cells(f"A{_next_row}:C{_next_row}")
            _iw(_next_row, 1, "STOCK DATA SOURCES", bold=True, size=11,
                color="FFFFFF", fill=_sec_fill, align="center")
            _next_row += 1

            _stock_rows = [
                ("Pupscribe WH Stock",   "Closing stock from Zoho Inventory at the latest available snapshot date."),
                ("FBA Stock",            "Amazon FBA inventory from the latest FBA snapshot."),
                ("Total Stock",          "Pupscribe WH Stock + FBA Stock."),
                ("Blinkit Inventory",    "Latest Blinkit dark-store inventory (date shown in column header)."),
                ("Etrade Inventory",     "Vendor Central closing stock at the end of the period."),
                ("Stock in Transit 1–3 + Transit PO 1–3", "Open purchase orders in transit (up to 3 POs per vendor, sorted by date). "
                 "Each SIT column shows the pending qty from that vendor's Nth earliest issued PO. "
                 "The adjacent 'Transit N PO' column shows the corresponding PO number. "
                                         "Sorted by expected delivery date, earliest first."),
                ("Total Stock in Transit", "Sum of the three transit columns."),
                ("Net Total Stock",      "Total Stock (latest) + Total Stock in Transit."),
            ]
            for _ri4, (_src, _desc4) in enumerate(_stock_rows, start=0):
                _rr4 = _next_row + _ri4
                _cfill4 = _alt_fill if _ri4 % 2 == 0 else _wht_fill
                _iw(_rr4, 1, _src, bold=True, fill=_cfill4)
                _iw(_rr4, 2, _desc4, fill=_cfill4, wrap=True)
                _instr_ws.merge_cells(f"B{_rr4}:C{_rr4}")
                _instr_ws.row_dimensions[_rr4].height = 35
            _next_row += len(_stock_rows)

            # ── Footer note ───────────────────────────────────────────────────────
            _instr_ws.merge_cells(f"A{_next_row}:C{_next_row}")
            _iw(_next_row, 1,
                "ℹ️  Formula cells in the master sheet reference each other — do not delete "
                "intermediate columns (e.g. Net Lookback Sales, Net Total Sales) even if they appear "
                "redundant. Hidden columns (Credit Notes, Confidence Multiplier, prev-date stock) "
                "are also referenced by formulas.",
                italic=True, size=10, fill=_alt_fill, wrap=True, align="center", bold=False)
            _instr_ws.row_dimensions[_next_row].height = 50

            # freeze the header row
            _instr_ws.freeze_panes = "A2"

            # ── Master Sheet ──────────────────────────────────────────────────────
            combined_df_data = []
            _master_unit_price_currencies: list = []
            for item in combined_data:
                if not isinstance(item, dict):
                    continue
                # Combo products are removed from the report — their sales were
                # redistributed to component items in master_report.py.  Any combo
                # that survived to this point (e.g. missing composite_products entry)
                # is still excluded here to keep the sheet clean.
                if item.get("is_combo_product"):
                    continue
                if item.get("order_qty_demand_override"):
                    demand_override_row_indices.add(len(combined_df_data))
                if item.get("drr_lookback_gt_180"):
                    lookback_gt_180_row_indices.add(len(combined_df_data))

                metrics = item.get("combined_metrics", {})
                _master_unit_price_currencies.append(item.get("unit_price_currency", "") or "")
                combined_df_data.append(
                    {
                        "Purchase Status": item.get("purchase_status", ""),
                        "Is New": "Yes" if item.get("is_new", False) else "No",
                        "Live on Amazon": "Yes" if item.get("live_on_amazon", False) else "No",
                        "Live on Blinkit": "Yes" if item.get("live_on_blinkit", False) else "No",
                        "SKU Code": item.get("sku_code", ""),
                        "Brand": item.get("brand", ""),
                        "Item Name": item.get("item_name", ""),
                        "Manufacturer Code": item.get("manufacturer_code", ""),
                        "MRP": item.get("mrp") or 0,
                        "Unit Price": item.get("unit_price", 0),
                        **(
                            {
                                "Unit Cost (COGS)": item.get("cogs_unit_cost", 0),
                                f"Pupscribe WH Stock Value ({_cogs_date_label})": item.get("cogs_asset_value", 0),
                                f"Pupscribe WH Stock on Hand ({_cogs_date_label})": item.get("cogs_qty", 0),
                            }
                            if include_cogs
                            else {}
                        ),
                        "Total Amount": f"₹{metrics.get('total_amount', 0)}",
                        "Total Units Sold": metrics.get("total_units_sold", 0),
                        "Total Units Returned": metrics.get("total_units_returned", 0),
                        "Total Credit Notes": metrics.get("total_credit_notes", 0),
                        "Transfer Orders": metrics.get("transfer_orders", 0),
                        "Net Total Sales": metrics.get("total_sales", 0),
                        "Return %": item.get("return_pct", 0),
                        "Days in Stock (Pupscribe Warehouse)": metrics.get("total_days_in_stock", 0),
                        "Lookback Days in Stock": item.get("drr_lookback_days_in_stock", 0),
                        "Lookback Sales": item.get("drr_lookback_sales", 0),
                        "Lookback Returns": item.get("drr_lookback_returns", 0),
                        "Net Lookback Sales": item.get("drr_net_lookback_sales", 0),
                        "Lookback Return %": None,  # formula injected below
                        "Avg Daily Run Rate": metrics.get("avg_daily_run_rate", 0),
                        "Growth Rate (%)": None,  # Excel formula injected for all row colours
                        "DRR Source": item.get("drr_source", "current_period"),
                        "DRR Lookback Period": item.get("drr_lookback_period", ""),
                        f"Pupscribe WH Stock ({_prev_zoho_label})": item.get("prev_zoho_stock", 0),
                        f"FBA Stock ({_prev_fba_label})": item.get("prev_fba_stock", 0),
                        f"Total Stock ({_prev_total_label})": 0,
                        f"Pupscribe WH Stock ({_latest_zoho_label})": item.get("latest_zoho_stock", 0),
                        f"FBA Stock ({_latest_fba_label})": item.get("latest_fba_stock", 0),
                        f"Total Stock ({_latest_total_label})": item.get("latest_total_stock", 0),
                        "In Stock": "Yes" if item.get("in_stock", False) else "No",
                        "Movement": item.get("movement", ""),
                        "Safety Days": item.get("safety_days", 0),
                        "Lead Time": item.get("lead_time", 0),
                        "Order Processing": item.get("order_processing", 10),
                        "Target Days": item.get("target_days", 0),
                        "On-Hand Days Coverage": item.get("on_hand_days_coverage", 0),
                        "Stock in Transit 1": item.get("stock_in_transit_1", 0),
                        "Transit 1 PO": item.get("transit_1_po", ""),
                        "Stock in Transit 2": item.get("stock_in_transit_2", 0),
                        "Transit 2 PO": item.get("transit_2_po", ""),
                        "Stock in Transit 3": item.get("stock_in_transit_3", 0),
                        "Transit 3 PO": item.get("transit_3_po", ""),
                        "Total Stock in Transit": item.get("total_stock_in_transit", 0),
                        "Net Total Stock": 0,  # formula: Total Stock (latest) + Total Stock in Transit
                        "Current Days Coverage": item.get("current_days_coverage", 0),
                        "Missed Sales": item.get("missed_sales", 0),
                        "Missed Sales DRR": item.get("missed_sales_drr", 0),
                        "Extra Qty": item.get("extra_qty", 0),
                        "Net Target Days": item.get("net_target_days", 0),
                        "Confidence Multiplier": item.get("confidence_multiplier", 1.0),
                        "Excess / Order": item.get("excess_or_order", ""),
                        "Order Qty": item.get("order_qty", 0),
                        "DRR Flag": item.get("order_qty_basis", ""),
                        "Order Qty + Extra Qty": item.get("order_qty_plus_extra_qty", 0),
                        "CBM": item.get("cbm", 0),
                        "Case Pack": item.get("case_pack", 0),
                        "Order Qty + Extra Qty (Rounded down)": item.get("order_qty_plus_extra_qty_rounded", 0),
                        "Total CBM": item.get("total_cbm", 0),
                        "Days Current Order Lasts": item.get("days_current_order_lasts", 0),
                        "Current Days Coverage (copy)": 0,  # formula: mirrors Current Days Coverage
                        "Days Total Inventory Lasts": item.get("days_total_inventory_lasts", 0),
                        "Additional Carton (Round up)": None,
                        "Amazon Quantity (New Listing)": None,
                        "Blinkit Quantity (New Listing)": None,
                        "Super Tails Quantity (New Listing)": None,
                        "Other Additional Quantity": None,
                        "Final order Qty": None,  # formula injected below
                        "Final order Qty (rounded up/down)": None,  # formula injected below
                        "Final CBM": None,         # formula injected below
                        f"Amazon Total Inventory ({_latest_fba_label})": round(item.get("latest_fba_stock", 0) + metrics.get("etrade_inventory", 0), 2),
                        f"Blinkit Inventory ({_blinkit_inv_label})": item.get("blinkit_inventory", 0),
                        "Amazon DRR": item.get("amazon_final_drr", 0.0),
                        "Days Total Inventory Lasts (Amazon)": 0,  # formula: Amazon Total Inv / Amazon DRR
                        "Total MRP": 0,
                        "Collection Value": 0,
                    }
                )

            if combined_df_data:
                combined_df = pd.DataFrame(combined_df_data)
                combined_df.to_excel(writer, sheet_name=_sheet_name, index=False)

                # Inject Excel formulas for all computed columns
                from openpyxl.utils import get_column_letter
                from openpyxl.styles import PatternFill

                ws = writer.sheets[_sheet_name]
                cols_list = list(combined_df.columns)

                def _col(name):
                    return get_column_letter(cols_list.index(name) + 1)

                _A  = _col("Purchase Status")
                _brand_col = _col("Brand")
                _mrp_col        = _col("MRP")
                _total_mrp_col  = _col("Total MRP")
                _cv_col         = _col("Collection Value")
                _F  = _col("Total Units Sold")
                _G  = _col("Total Units Returned")
                _credit_notes_col = _col("Total Credit Notes")
                _I  = _col("Net Total Sales")
                _J  = _col("Return %")
                _N  = _col("Avg Daily Run Rate")
                _S  = _col(f"Pupscribe WH Stock ({_prev_zoho_label})")
                _T  = _col(f"FBA Stock ({_prev_fba_label})")
                _R  = _col(f"Total Stock ({_prev_total_label})")
                _V  = _col(f"Pupscribe WH Stock ({_latest_zoho_label})")
                _W  = _col(f"FBA Stock ({_latest_fba_label})")
                _U  = _col(f"Total Stock ({_latest_total_label})")
                _AA = _col("Safety Days")
                _AB = _col("Lead Time")
                _AC = _col("Order Processing")
                _AD = _col("Target Days")
                _AE = _col("On-Hand Days Coverage")
                _AF = _col("Stock in Transit 1")
                _AG = _col("Stock in Transit 2")
                _AH = _col("Stock in Transit 3")
                _AI = _col("Total Stock in Transit")
                _NTS = _col("Net Total Stock")
                _AJ = _col("Current Days Coverage")
                _AK = _col("Missed Sales")
                _AL = _col("Missed Sales DRR")
                _AM = _col("Extra Qty")
                _AN = _col("Net Target Days")
                _AX = _col("Confidence Multiplier")
                _AO = _col("Excess / Order")
                _AP = _col("Order Qty")
                _AQ = _col("Order Qty + Extra Qty")
                _AR = _col("CBM")
                _AS = _col("Case Pack")
                _AT = _col("Order Qty + Extra Qty (Rounded down)")
                _AU = _col("Total CBM")
                _AV = _col("Days Current Order Lasts")
                _cdc_copy    = _col("Current Days Coverage (copy)")
                _AW = _col("Days Total Inventory Lasts")
                _add_carton  = _col("Additional Carton (Round up)")
                _amz_new_qty = _col("Amazon Quantity (New Listing)")
                _blk_new_qty = _col("Blinkit Quantity (New Listing)")
                _st_new_qty  = _col("Super Tails Quantity (New Listing)")
                _other_qty   = _col("Other Additional Quantity")
                _final_qty   = _col("Final order Qty")
                _final_qty_rounded = _col("Final order Qty (rounded up/down)")
                _final_cbm   = _col("Final CBM")
                _amz_total_inv = _col(f"Amazon Total Inventory ({_latest_fba_label})")
                _etrade_drr  = _col("Amazon DRR")
                _etrade_days = _col("Days Total Inventory Lasts (Amazon)")

                # Rename the "Current Days Coverage (copy)" header to "Current Days Coverage"
                _cdc_copy_col_idx = cols_list.index("Current Days Coverage (copy)") + 1
                ws.cell(row=1, column=_cdc_copy_col_idx).value = "Current Days Coverage"
                _transfer_col        = _col("Transfer Orders")
                _days_in_stock_col   = _col(f"Days in Stock (Pupscribe Warehouse)")
                _drr_src_col         = _col("DRR Source")
                _lkbk_days_col       = _col("Lookback Days in Stock")
                _lkbk_sales_col      = _col("Lookback Sales")
                _lkbk_returns_col    = _col("Lookback Returns")
                _net_lkbk_col        = _col("Net Lookback Sales")
                _lkbk_ret_pct_col    = _col("Lookback Return %")
                _growth_rate_col     = _col("Growth Rate (%)")
                _is_new_col          = _col("Is New")

                drr_source_col_idx = cols_list.index("DRR Source") + 1  # 1-based
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

                for row_idx in range(2, len(combined_df_data) + 2):  # Skip header row
                    r = row_idx
                    inactive = f'OR({_A}{r}="inactive",{_A}{r}="discontinued until stock lasts")'

                    # Net Lookback Sales = MAX(0, Lookback Sales − Lookback Returns)
                    ws[f"{_net_lkbk_col}{r}"] = (
                        f"=MAX(0,{_lkbk_sales_col}{r}-{_lkbk_returns_col}{r})"
                    )

                    # Lookback Return % = Lookback Returns / Lookback Sales × 100
                    ws[f"{_lkbk_ret_pct_col}{r}"] = (
                        f"=IF({_lkbk_sales_col}{r}>0,{_lkbk_returns_col}{r}/{_lkbk_sales_col}{r}*100,0)"
                    )

                    # Growth Rate (%) — injected for ALL row colours once lookback sales
                    # are populated for white/red rows as well as yellow/orange.
                    # Blank ("") when no comparison data exists.
                    ws[f"{_growth_rate_col}{r}"] = (
                        f'=IF({_net_lkbk_col}{r}>0,({_I}{r}-{_net_lkbk_col}{r})/{_net_lkbk_col}{r}*100,"")'
                    )
                    ws[f"{_growth_rate_col}{r}"].number_format = '0.00"%"'

                    # Net Total Sales = Total Units Sold − Total Units Returned − Transfer Orders
                    # (Total Units Returned used instead of Total Credit Notes — values are equivalent)
                    ws[f"{_I}{r}"] = f"={_F}{r}-{_G}{r}-{_transfer_col}{r}"

                    # Avg Daily Run Rate — simple formula per row colour, injected via Python branching
                    # • Green  (demand override):      =IF(DaysInStock>0, NTS/DaysInStock, IF(NTS=0,0, NTS/period_days))
                    # • Yellow/Orange (prev_period):   =IF(LookbackDays>0, MAX(0, LkbkSales-LkbkReturns)/LookbackDays, 0)
                    # • White  (current_period):       =IF(DaysInStock>0, NTS/DaysInStock, IF(NTS=0,0, NTS/period_days))
                    # • Red    (insufficient_stock):   =IF(DaysInStock>0, NTS/DaysInStock, IF(NTS=0,0, NTS/period_days))
                    #   Dampening for red rows is applied to Order Qty via confidence multiplier, not to DRR.
                    _net_ts = f"{_I}{r}"  # Net Total Sales cell (formula already injected above)
                    _data_row_drr_src = combined_df_data[row_idx - 2].get("DRR Source", "current_period")
                    if (row_idx - 2) in demand_override_row_indices or _data_row_drr_src != "previous_period":
                        # Green / White / Red: Net Total Sales / Days in Stock.
                        # MAX(0, …) floors negative DRR (returns > sales) so the
                        # Excess/Order formula resolves to NO MOVEMENT, not ORDER.
                        ws[f"{_N}{r}"] = (
                            f'=MAX(0,IF({_days_in_stock_col}{r}>0,{_net_ts}/{_days_in_stock_col}{r},'
                            f'IF({_net_ts}=0,0,{_net_ts}/{_period_days})))'
                        )
                    else:
                        # Yellow / Orange: Net Lookback Sales / Lookback Days in Stock
                        # Use the Net Lookback Sales cell directly (already = MAX(0, LkbkSales - LkbkReturns))
                        ws[f"{_N}{r}"] = (
                            f'=IF({_lkbk_days_col}{r}>0,'
                            f'{_net_lkbk_col}{r}/{_lkbk_days_col}{r},0)'
                        )

                    # Return %
                    ws[f"{_J}{r}"] = f"=IF({_F}{r}>0,{_G}{r}/{_F}{r}*100,0)"

                    # Total MRP = MRP × Pupscribe WH Stock (latest)
                    ws[f"{_total_mrp_col}{r}"] = f"={_mrp_col}{r}*{_V}{r}"
                    ws[f"{_total_mrp_col}{r}"].number_format = '₹#,##0.00'

                    # Collection Value = Total MRP / 2
                    ws[f"{_cv_col}{r}"] = f"={_total_mrp_col}{r}/2"
                    ws[f"{_cv_col}{r}"].number_format = '₹#,##0.00'

                    # Total Stock (prev) = prev WH + prev FBA
                    ws[f"{_R}{r}"] = f"={_S}{r}+{_T}{r}"

                    # Total Stock (latest) = WH latest + FBA latest
                    ws[f"{_U}{r}"] = f"={_V}{r}+{_W}{r}"

                    # On-Hand Days Coverage = Latest Total Stock / DRR
                    ws[f"{_AE}{r}"] = f"=IF({_N}{r}>0,{_U}{r}/{_N}{r},0)"

                    # Total Stock in Transit = sum of 3 transit cols
                    ws[f"{_AI}{r}"] = f"={_AF}{r}+{_AG}{r}+{_AH}{r}"

                    # Net Total Stock = Total Stock (latest) + Total Stock in Transit
                    ws[f"{_NTS}{r}"] = f"={_U}{r}+{_AI}{r}"

                    # Current Days Coverage = (Latest Total Stock + Total Transit) / DRR
                    ws[f"{_AJ}{r}"] = f"=IF({_N}{r}>0,({_U}{r}+{_AI}{r})/{_N}{r},0)"

                    # Target Days = Lead Time + Safety Days + Order Processing
                    ws[f"{_AD}{r}"] = f"={_AB}{r}+{_AA}{r}+{_AC}{r}"

                    # Missed Sales DRR = MIN(Missed Sales / period_days, 0.5 * DRR), capped only if DRR > 0
                    ws[f"{_AL}{r}"] = f"=IF({_N}{r}>0,MIN({_AK}{r}/{_period_days},0.5*{_N}{r}),{_AK}{r}/{_period_days})"

                    # Extra Qty = Missed Sales DRR * Lead Time
                    ws[f"{_AM}{r}"] = f"={_AL}{r}*{_AB}{r}"

                    # Net Target Days = IF(Current Coverage < Lead Time, Target Days, Target Days - Current Coverage)
                    ws[f"{_AN}{r}"] = f"=IF({_AJ}{r}<{_AB}{r},{_AD}{r},{_AD}{r}-{_AJ}{r})"

                    # Excess / Order
                    # New items (Is New = Yes) with no DRR show NO MOVEMENT; IFERROR catches any
                    # residual #DIV/0! that could appear on new items with zero stock/sales.
                    ws[f"{_AO}{r}"] = (
                        f'=IFERROR('
                        f'IF(OR({_N}{r}=0,AND({_is_new_col}{r}="Yes",{_N}{r}=0)),"NO MOVEMENT",'
                        f'IF({_AJ}{r}<{_AD}{r},"ORDER","EXCESS")),'
                        f'"NO MOVEMENT")'
                    )

                    # Order Qty (0 for inactive/discontinued, 0 for EXCESS)
                    # For demand-override rows (green), use Net Total Sales × confidence multiplier
                    # For all other rows: MAX(0, Net Target Days × DRR) × confidence multiplier
                    if (row_idx - 2) in demand_override_row_indices:
                        ws[f"{_AP}{r}"] = f'=IF({inactive},0,IF({_AO}{r}="EXCESS",0,{_I}{r}*{_AX}{r}))'
                    else:
                        ws[f"{_AP}{r}"] = f'=IF({inactive},0,IF({_AO}{r}="EXCESS",0,MAX(0,{_AN}{r}*{_N}{r})*{_AX}{r}))'

                    # Order Qty + Extra Qty (only Extra Qty for inactive/discontinued; 0 for EXCESS or NO MOVEMENT)
                    ws[f"{_AQ}{r}"] = f'=IF({inactive},{_AM}{r},IF(OR({_AO}{r}="EXCESS",{_AO}{r}="NO MOVEMENT"),0,{_AP}{r}+{_AM}{r}))'

                    # Order Qty + Extra Qty (Rounded):
                    # • EXCESS or NO MOVEMENT → 0
                    # • ORDER and FLOOR result = 0 → bump to 1 case pack (minimum meaningful order)
                    # • Otherwise → FLOOR to nearest case pack
                    ws[f"{_AT}{r}"] = (
                        f'=IF({inactive},0,'
                        f'IF(OR({_AO}{r}="EXCESS",{_AO}{r}="NO MOVEMENT"),0,'
                        f'IF({_AS}{r}>0,'
                        f'IF(AND(FLOOR({_AQ}{r},{_AS}{r})=0,{_AO}{r}="ORDER"),{_AS}{r},FLOOR({_AQ}{r},{_AS}{r})),'
                        f'ROUND({_AQ}{r},0))))'
                    )

                    # Total CBM
                    ws[f"{_AU}{r}"] = f"=IF({inactive},0,IF({_AS}{r}>0,({_AT}{r}/{_AS}{r})*{_AR}{r},0))"

                    # Days Current Order Lasts
                    ws[f"{_AV}{r}"] = f"=IF({inactive},0,IF({_N}{r}>0,{_AT}{r}/{_N}{r},0))"

                    # Current Days Coverage (copy) — duplicate of Current Days Coverage for reference
                    ws[f"{_cdc_copy}{r}"] = f"={_AJ}{r}"

                    # Days Total Inventory Lasts = Current Days Coverage + Days Current Order Lasts
                    ws[f"{_AW}{r}"] = f"=IF({inactive},{_AJ}{r},{_AJ}{r}+{_AV}{r})"

                    # Final order Qty = Additional Carton + Order Rounded + Amazon/Blinkit/Supertails New Listing + Other Additional
                    # Blank manual cells are treated as 0 by Excel arithmetic — no guards needed
                    ws[f"{_final_qty}{r}"] = (
                        f"={_add_carton}{r}+{_AT}{r}+{_amz_new_qty}{r}+{_blk_new_qty}{r}+{_st_new_qty}{r}+{_other_qty}{r}"
                    )

                    # Final order Qty (rounded up/down) — standard half-up rounding of Final order Qty
                    ws[f"{_final_qty_rounded}{r}"] = f"=ROUND({_final_qty}{r},0)"

                    # Final CBM = (Final order Qty / Case Pack) × CBM
                    ws[f"{_final_cbm}{r}"] = f"=IF({inactive},0,IF({_AS}{r}>0,({_final_qty}{r}/{_AS}{r})*{_AR}{r},0))"

                    # Days Total Inventory Lasts (Amazon) = Amazon Total Inventory / Amazon DRR
                    ws[f"{_etrade_days}{r}"] = f"=IF({_etrade_drr}{r}>0,{_amz_total_inv}{r}/{_etrade_drr}{r},0)"

                    # Apply row highlighting based on DRR source / demand override
                    # row_idx is 1-based (row 2 = first data row); convert to 0-based for the set
                    data_row_idx = row_idx - 2
                    drr_source_val = ws.cell(row=r, column=drr_source_col_idx).value
                    if data_row_idx in demand_override_row_indices:
                        # Green: lookback DRR used but current-period sales are higher
                        # — order qty has been overridden to current-period sales
                        for col_idx in range(1, len(combined_df.columns) + 1):
                            ws.cell(row=r, column=col_idx).fill = green_fill
                    elif drr_source_val == "previous_period":
                        # Orange: lookback period is > 180 days back (stale reference)
                        # Yellow: normal lookback (within 180 days)
                        fill = orange_fill if data_row_idx in lookback_gt_180_row_indices else yellow_fill
                        for col_idx in range(1, len(combined_df.columns) + 1):
                            ws.cell(row=r, column=col_idx).fill = fill
                    elif drr_source_val == "insufficient_stock":
                        for col_idx in range(1, len(combined_df.columns) + 1):
                            ws.cell(row=r, column=col_idx).fill = red_fill

                # Apply currency symbol number format to Unit Price column
                _up_col_letter = _col("Unit Price")
                for i, currency_code in enumerate(_master_unit_price_currencies):
                    num_fmt = _CURRENCY_FORMATS.get(currency_code.upper(), _DEFAULT_NUMBER_FMT)
                    ws[f"{_up_col_letter}{i + 2}"].number_format = num_fmt

                # Hide the Confidence Multiplier column (used by Order Qty formula, not for display)
                ws.column_dimensions[_AX].hidden = True

                # Hide Total Credit Notes column — values are identical to Total Units Returned,
                # so it is redundant for display; formulas that referenced it now use Total Units Returned.
                ws.column_dimensions[_credit_notes_col].hidden = True

                # Hide previous-date stock snapshot columns (redundant — latest snapshots are shown)
                ws.column_dimensions[_S].hidden = True   # Pupscribe WH Stock (prev)
                ws.column_dimensions[_T].hidden = True   # FBA Stock (prev)
                ws.column_dimensions[_R].hidden = True   # Total Stock (prev)

                # Hide individual latest WH/FBA stock components — total stock (=V+W) stays visible
                ws.column_dimensions[_V].hidden = True   # Pupscribe WH Stock (latest)
                ws.column_dimensions[_W].hidden = True   # FBA Stock (latest)

                # Hide metadata columns not needed in the default view
                for _hide_col_name in (
                    "Is New",
                    "Live on Amazon",
                    "Live on Blinkit",
                    "Manufacturer Code",
                    "Unit Price",
                    "Total Amount",
                    "Total Units Sold",
                    "Total Units Returned",
                    "Transfer Orders",
                    "Lookback Days in Stock",
                    "Lookback Sales",
                    "Lookback Returns",
                    "Safety Days",
                    "Lead Time",
                    "Order Processing",
                    "Stock in Transit 1",
                    "Transit 1 PO",
                    "Stock in Transit 2",
                    "Transit 2 PO",
                    "Stock in Transit 3",
                    "Transit 3 PO",
                ):
                    try:
                        ws.column_dimensions[_col(_hide_col_name)].hidden = True
                    except (ValueError, KeyError):
                        pass  # column may not exist when COGS mode is off

                # Hide Brand column for individual brand reports (redundant when filtered to one brand)
                if brand:
                    ws.column_dimensions[_brand_col].hidden = True

                # Auto-fit column widths: header text drives the width; skip formula cells
                for col_cells in ws.columns:
                    col_letter = get_column_letter(col_cells[0].column)
                    max_len = 0
                    for cell in col_cells:
                        if cell.value is None:
                            continue
                        val = str(cell.value)
                        if val.startswith("="):
                            continue  # skip formulas — display value is computed, not the formula string
                        max_len = max(max_len, len(val))
                    ws.column_dimensions[col_letter].width = min(max_len + 3, 30)

            # Draft Order sheet — only when a brand filter is active
            if brand:
                _today_label = datetime.now().strftime("%d %b %Y")
                _draft_sheet_name = f"Draft Order {_today_label}"[:31]

                draft_rows = []
                for item in combined_data:
                    if not isinstance(item, dict):
                        continue
                    if item.get("purchase_status", "") != "active":
                        continue
                    if item.get("is_combo_product", False):
                        continue
                    draft_rows.append({
                        "HSN Code": item.get("hsn_or_sac", ""),
                        "Manufacturer Code": item.get("manufacturer_code", ""),
                        "BBCode": item.get("sku_code", ""),
                        "Item Name": item.get("item_name", ""),
                        "Brand": item.get("brand", ""),
                        "SKU Code": item.get("sku", ""),
                        "Category": item.get("category", ""),
                        "Sub Category": item.get("sub_category", ""),
                        "Series": item.get("series", ""),
                        "MRP": item.get("mrp") or "",
                        "Qty": None,
                        "Unit Price": item.get("unit_price", 0) or None,
                        "Total": None,
                        "Case Pack": item.get("case_pack", 0),
                        "Cartons": None,
                        "CBM": item.get("cbm", 0),
                        "Total CBM": None,
                        "Final order Qty (rounded up/down)": None,
                        "Final CBM": None,
                        "_currency": item.get("unit_price_currency", ""),
                    })

                if not draft_rows:
                    # Brand-new brand with no products yet — write a blank sheet with headers only.
                    from openpyxl.utils import get_column_letter as _gcl
                    _blank_cols = [
                        "HSN Code", "Manufacturer Code", "BBCode", "Item Name",
                        "Brand", "SKU Code", "Category", "Sub Category", "Series", "MRP",
                        "Qty", "Unit Price", "Total", "Case Pack", "Cartons",
                        "CBM", "Total CBM", "Final order Qty (rounded up/down)", "Final CBM",
                    ]
                    blank_df = pd.DataFrame(columns=_blank_cols)
                    blank_df.to_excel(writer, sheet_name=_draft_sheet_name, index=False)
                    ws_blank = writer.sheets[_draft_sheet_name]
                    for _ci, _cn in enumerate(_blank_cols, 1):
                        ws_blank.column_dimensions[_gcl(_ci)].width = max(len(_cn) + 4, 14)
                elif draft_rows:
                    from openpyxl.styles import Border, Side, Font as _Font

                    # Group rows by currency so each currency gets its own sheet
                    _currency_groups: dict[str, list] = {}
                    for _dr in draft_rows:
                        _ckey = (_dr.get("_currency") or "").strip()
                        _currency_groups.setdefault(_ckey, []).append(_dr)

                    _multi_currency = len(_currency_groups) > 1
                    _master_sheet_ref = _sheet_name.replace("'", "''")
                    _master_sku_col = _col("SKU Code")

                    for _ckey, _group_rows in sorted(_currency_groups.items()):
                        if _multi_currency and _ckey:
                            _gs_name = f"Draft Order {_ckey} {_today_label}"[:31]
                        else:
                            _gs_name = _draft_sheet_name

                        draft_df = pd.DataFrame([{k: v for k, v in r.items() if k != "_currency"} for r in _group_rows])
                        draft_df.to_excel(writer, sheet_name=_gs_name, index=False)

                        ws_draft = writer.sheets[_gs_name]
                        draft_cols = list(draft_df.columns)

                        def _dcol(name, _dc=draft_cols):
                            return get_column_letter(_dc.index(name) + 1)

                        _dqty    = _dcol("Qty")
                        _dbbcode = _dcol("BBCode")
                        _dup     = _dcol("Unit Price")
                        _dtot    = _dcol("Total")
                        _dcp     = _dcol("Case Pack")
                        _dcar    = _dcol("Cartons")
                        _dcbm    = _dcol("CBM")
                        _dtcbm   = _dcol("Total CBM")
                        _dmrp    = _dcol("MRP")
                        _dfinalqty_rounded = _dcol("Final order Qty (rounded up/down)")
                        _dfinal_cbm        = _dcol("Final CBM")

                        for r_idx, row in enumerate(_group_rows):
                            r = r_idx + 2  # 1-based, skip header
                            ws_draft[f"{_dqty}{r}"] = (
                                f"=_xlfn.XLOOKUP({_dbbcode}{r},"
                                f"'{_master_sheet_ref}'!{_master_sku_col}:{_master_sku_col},"
                                f"'{_master_sheet_ref}'!{_final_qty}:{_final_qty},0)"
                            )
                            ws_draft[f"{_dtot}{r}"]  = f"={_dqty}{r}*{_dup}{r}"
                            ws_draft[f"{_dcar}{r}"]  = f"=IF({_dcp}{r}>0,{_dqty}{r}/{_dcp}{r},0)"
                            ws_draft[f"{_dtcbm}{r}"] = (
                                f"=_xlfn.XLOOKUP({_dbbcode}{r},"
                                f"'{_master_sheet_ref}'!{_master_sku_col}:{_master_sku_col},"
                                f"'{_master_sheet_ref}'!{_final_cbm}:{_final_cbm},0)"
                            )
                            ws_draft[f"{_dfinalqty_rounded}{r}"] = (
                                f"=_xlfn.XLOOKUP({_dbbcode}{r},"
                                f"'{_master_sheet_ref}'!{_master_sku_col}:{_master_sku_col},"
                                f"'{_master_sheet_ref}'!{_final_qty_rounded}:{_final_qty_rounded},0)"
                            )
                            ws_draft[f"{_dfinal_cbm}{r}"] = (
                                f"=IF({_dcp}{r}>0,({_dfinalqty_rounded}{r}/{_dcp}{r})*{_dcbm}{r},0)"
                            )

                            currency_code = row.get("_currency", "") or ""
                            num_fmt = _CURRENCY_FORMATS.get(currency_code.upper(), _DEFAULT_NUMBER_FMT)
                            ws_draft[f"{_dup}{r}"].number_format = num_fmt
                            ws_draft[f"{_dtot}{r}"].number_format = num_fmt
                            ws_draft[f"{_dmrp}{r}"].number_format = '₹#,##0.00'

                        # ── Totals row ──────────────────────────────────────────
                        _last_data_row = len(_group_rows) + 1
                        _total_row     = _last_data_row + 1

                        _thick_top = Border(top=Side(border_style="medium", color="000000"))
                        _bold_font = _Font(bold=True)

                        _item_col = _dcol("Item Name")
                        _label_cell = ws_draft[f"{_item_col}{_total_row}"]
                        _label_cell.value = "Total"
                        _label_cell.font  = _bold_font
                        _label_cell.border = _thick_top

                        _sum_currency_fmt = _CURRENCY_FORMATS.get(_ckey.upper(), _DEFAULT_NUMBER_FMT)

                        _totals_spec: list[tuple[str, str, str]] = [
                            (_dqty,             f"=SUM({_dqty}{2}:{_dqty}{_last_data_row})",                         "#,##0"),
                            (_dtot,             f"=SUM({_dtot}{2}:{_dtot}{_last_data_row})",                         None),
                            (_dcar,             f"=SUM({_dcar}{2}:{_dcar}{_last_data_row})",                         "#,##0.00"),
                            (_dtcbm,            f"=SUM({_dtcbm}{2}:{_dtcbm}{_last_data_row})",                       '#,##0.00 "m³"'),
                            (_dfinalqty_rounded,f"=SUM({_dfinalqty_rounded}{2}:{_dfinalqty_rounded}{_last_data_row})","#,##0"),
                            (_dfinal_cbm,       f"=SUM({_dfinal_cbm}{2}:{_dfinal_cbm}{_last_data_row})",             '#,##0.00 "m³"'),
                        ]
                        for _col_ltr, _formula, _fmt in _totals_spec:
                            _c = ws_draft[f"{_col_ltr}{_total_row}"]
                            _c.value  = _formula
                            _c.font   = _bold_font
                            _c.border = _thick_top
                            if _col_ltr == _dtot:
                                _c.number_format = _sum_currency_fmt
                            elif _fmt:
                                _c.number_format = _fmt

                        for _dc in draft_cols:
                            _cl = _dcol(_dc)
                            if _cl in {_item_col, _dqty, _dtot, _dcar, _dtcbm, _dfinalqty_rounded, _dfinal_cbm}:
                                continue
                            ws_draft[f"{_cl}{_total_row}"].border = _thick_top

                        # Auto-fit columns
                        for col_cells in ws_draft.columns:
                            col_letter = get_column_letter(col_cells[0].column)
                            max_len = 0
                            for cell in col_cells:
                                if cell.value is None:
                                    continue
                                val = str(cell.value)
                                if val.startswith("="):
                                    continue
                                max_len = max(max_len, len(val))
                            ws_draft.column_dimensions[col_letter].width = min(max_len + 3, 30)

        excel_buffer.seek(0)
        file_bytes = excel_buffer.read()

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"master_report_{start_date}_to_{end_date}_{timestamp}.xlsx"

        return Response(
            content=file_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Length": str(len(file_bytes)),
            },
        )

    except Exception as e:
        logger.error(f"Error generating master report Excel: {e}")
        logger.error(f"Excel generation error traceback: {traceback.format_exc()}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate Excel report: {str(e)}",
        )


def _classify_sku_stock(current_days_coverage: float, lead_time: float, drr: float) -> str:
    """Classify a SKU's stock health.
    Reorder Risk → Healthy → Heavy → Overstock → Dead (5 categories).
    """
    if drr == 0:
        return "dead"
    if current_days_coverage > 3 * lead_time:
        return "dead"
    if current_days_coverage > 2 * lead_time:
        return "overstock"
    if current_days_coverage > 1.5 * lead_time:
        return "heavy"
    if current_days_coverage >= lead_time:
        return "healthy"
    return "reorder_risk"


def _aggregate_brand_kpi(
    brand: str,
    items: list,
    brand_logistics_map: Dict,
    rate_map: Dict[str, float],
    period_days: int,
) -> Dict:
    """Aggregate all KPI metrics for a single brand."""
    bl = brand_logistics_map.get(brand.lower(), {})
    lead_time = bl.get("lead_time", 60)
    safety_days_medium = bl.get("safety_days_medium", 25)
    order_processing = bl.get("order_processing", 10)
    target_days = lead_time + safety_days_medium + order_processing

    units_sold = sum(i.get("combined_metrics", {}).get("total_units_sold", 0) for i in items)
    units_returned = sum(i.get("combined_metrics", {}).get("total_units_returned", 0) for i in items)
    credit_notes = sum(i.get("combined_metrics", {}).get("total_credit_notes", 0) for i in items)
    transfer_orders = sum(i.get("combined_metrics", {}).get("transfer_orders", 0) for i in items)
    net_sales = sum(i.get("combined_metrics", {}).get("total_sales", 0) for i in items)
    revenue = sum(i.get("combined_metrics", {}).get("total_amount", 0) for i in items)
    latest_net_stock = sum(i.get("latest_total_stock", 0) for i in items)
    latest_zoho = sum(i.get("latest_zoho_stock", 0) for i in items)
    latest_fba = sum(i.get("latest_fba_stock", 0) for i in items)
    stock_in_transit = sum(i.get("total_stock_in_transit", 0) for i in items)
    total_cbm = sum(i.get("total_cbm", 0) for i in items)

    # Inventory value and missed sales (need rate per SKU)
    net_inv_value = sum(
        i.get("latest_total_stock", 0) * rate_map.get(i.get("sku_code", ""), 0)
        for i in items
    )
    missed_sales_units = sum(i.get("missed_sales", 0) for i in items)
    missed_sales_daily_units = round(missed_sales_units / period_days, 2) if period_days > 0 else 0.0
    missed_sales_value_total = sum(
        i.get("missed_sales", 0) * rate_map.get(i.get("sku_code", ""), 0)
        for i in items
    )
    missed_sales_daily_value = round(missed_sales_value_total / period_days, 2) if period_days > 0 else 0.0

    # Return %
    return_pct = round(units_returned / units_sold * 100, 2) if units_sold > 0 else 0.0

    # Weighted-average growth rate (weight = per-SKU DRR)
    drr_w_sum = 0.0
    drr_w_total = 0.0
    for i in items:
        item_drr = i.get("combined_metrics", {}).get("avg_daily_run_rate", 0)
        item_gr = i.get("growth_rate")
        if item_gr is not None and item_drr > 0:
            drr_w_sum += item_gr * item_drr
            drr_w_total += item_drr
    growth_rate = round(drr_w_sum / drr_w_total, 2) if drr_w_total > 0 else None

    # Brand-level DRR = sum of per-SKU avg_daily_run_rate.
    # This matches the master report exactly: each SKU's DRR uses
    # (units_sold - credit_notes) / days_in_stock (calendar fallback),
    # so summing them gives total brand demand in units/day on the same basis.
    drr = round(sum(i.get("combined_metrics", {}).get("avg_daily_run_rate", 0) for i in items), 4)
    days_cover = round(latest_net_stock / drr, 2) if drr > 0 else 0.0
    current_days_coverage = round((latest_net_stock + stock_in_transit) / drr, 2) if drr > 0 else 0.0

    # Weighted-average days cover (weight = per-SKU on-hand + in-transit stock).
    # Exclude dead SKUs (drr==0 or coverage > 3× lead time) — they distort the
    # average without providing actionable reorder signal.
    _dead_threshold = 3 * lead_time
    stock_w_total = 0.0
    weighted_doc_sum = 0.0
    for i in items:
        item_drr = i.get("combined_metrics", {}).get("avg_daily_run_rate", 0)
        item_stock = i.get("latest_total_stock", 0) + i.get("total_stock_in_transit", 0)
        item_doc = i.get("current_days_coverage", 0)
        if item_drr > 0 and item_stock > 0 and item_doc <= _dead_threshold:
            weighted_doc_sum += item_stock * item_doc
            stock_w_total += item_stock
    weighted_avg_days_cover = round(weighted_doc_sum / stock_w_total, 2) if stock_w_total > 0 else 0.0

    # Alert level based on current days coverage (on-hand + in-transit).
    # Stock already on the water counts towards covering the lead time, so a
    # brand with a full container in transit is not flagged critical.
    if drr == 0:
        alert_level = 3  # no movement
    elif current_days_coverage < lead_time:
        alert_level = 2  # red – critical
    elif current_days_coverage < target_days:
        alert_level = 1  # yellow – caution
    else:
        alert_level = 0  # green – healthy

    # Stock classification per SKU
    classes = [
        _classify_sku_stock(
            i.get("current_days_coverage", 0),
            lead_time,
            i.get("combined_metrics", {}).get("avg_daily_run_rate", 0),
        )
        for i in items
    ]
    total_skus = len(items)
    class_counts = {
        "reorder_risk": sum(1 for c in classes if c == "reorder_risk"),
        "healthy": sum(1 for c in classes if c == "healthy"),
        "heavy": sum(1 for c in classes if c == "heavy"),
        "overstock": sum(1 for c in classes if c == "overstock"),
        "dead": sum(1 for c in classes if c == "dead"),
    }
    class_pct = {k: round(v / total_skus * 100, 1) if total_skus > 0 else 0.0 for k, v in class_counts.items()}

    # Per-SKU list for lowest/highest 10 (exclude dead/no-movement SKUs from both lists)
    sku_detail = sorted(
        [
            {
                "sku_code": i.get("sku_code", ""),
                "item_name": i.get("item_name", ""),
                "drr": round(i.get("combined_metrics", {}).get("avg_daily_run_rate", 0), 4),
                "net_stock": round(i.get("latest_total_stock", 0), 2),
                "days_cover": round(i.get("current_days_coverage", 0), 1),
                "excess_or_order": i.get("excess_or_order", ""),
                "movement": i.get("movement", ""),
                "stock_class": _classify_sku_stock(
                    i.get("current_days_coverage", 0),
                    lead_time,
                    i.get("combined_metrics", {}).get("avg_daily_run_rate", 0),
                ),
            }
            for i in items
            if i.get("combined_metrics", {}).get("avg_daily_run_rate", 0) > 0
        ],
        key=lambda x: x["days_cover"],
    )
    sku_lowest_10 = sku_detail[:10]
    _lowest_codes = {s["sku_code"] for s in sku_lowest_10}
    sku_highest_10 = [s for s in reversed(sku_detail) if s["sku_code"] not in _lowest_codes][:10]

    return {
        "brand": brand,
        "sku_count": total_skus,
        "units_sold": round(units_sold, 2),
        "units_returned": round(units_returned, 2),
        "credit_notes": round(credit_notes, 2),
        "transfer_orders": round(transfer_orders, 2),
        "net_sales": round(net_sales, 2),
        "revenue": round(revenue, 2),
        "return_pct": return_pct,
        "growth_rate": growth_rate,
        "drr": drr,
        "latest_net_stock": round(latest_net_stock, 2),
        "latest_zoho_stock": round(latest_zoho, 2),
        "latest_fba_stock": round(latest_fba, 2),
        "net_sellable_inventory_value": round(net_inv_value, 2),
        "stock_in_transit": round(stock_in_transit, 2),
        "total_cbm": round(total_cbm, 4),
        "days_cover": days_cover,
        "current_days_coverage": current_days_coverage,
        "weighted_avg_days_cover": weighted_avg_days_cover,
        "lead_time": lead_time,
        "safety_days": safety_days_medium,
        "target_days": target_days,
        "alert_level": alert_level,
        "order_count": sum(1 for i in items if i.get("excess_or_order") == "ORDER"),
        "excess_count": sum(1 for i in items if i.get("excess_or_order") == "EXCESS"),
        "no_movement_count": sum(1 for i in items if i.get("excess_or_order") == "NO MOVEMENT"),
        "fast_mover_count": sum(1 for i in items if i.get("movement") == "Fast Mover"),
        "medium_mover_count": sum(1 for i in items if i.get("movement") == "Medium Mover"),
        "slow_mover_count": sum(1 for i in items if i.get("movement") == "Slow Mover"),
        "missed_sales_units": round(missed_sales_units, 2),
        "missed_sales_daily_units": missed_sales_daily_units,
        "missed_sales_value_total": round(missed_sales_value_total, 2),
        "missed_sales_daily_value": missed_sales_daily_value,
        "stock_classification": {"counts": class_counts, "pct": class_pct},
        "sku_lowest_10": sku_lowest_10,
        "sku_highest_10": sku_highest_10,
    }


def _dashboard_kpi_default_dates() -> tuple[str, str]:
    """Return (start_date, end_date) for the default 90-day window ending on the previous Sunday."""
    today = datetime.now()
    # Python weekday(): Mon=0 … Sun=6.  We want the most recent Sunday that is not today.
    days_since_sunday = 7 if today.weekday() == 6 else (today.weekday() + 1)
    prev_sunday = today - timedelta(days=days_since_sunday)
    end = prev_sunday.strftime("%Y-%m-%d")
    start = (prev_sunday - timedelta(days=89)).strftime("%Y-%m-%d")
    return start, end


@router.get("/dashboard-kpi")
async def get_dashboard_kpi(
    start_date: Optional[str] = Query(None, description="Start date YYYY-MM-DD (default: 90 days ending previous Sunday)"),
    end_date: Optional[str] = Query(None, description="End date YYYY-MM-DD (default: previous Sunday)"),
    db=Depends(get_database),
):
    """
    Stock Cover KPI Dashboard – last 90 days, aggregated by brand.

    Runs the full master-report pipeline for the last 90 days (Zoho sales + latest stock)
    and aggregates the result by brand.

    ## Response Structure

    ### `period`
    - `start_date`, `end_date`, `days` – the 90-day window used

    ### `brands` (array, sorted most-critical first)
    One object per brand. Key fields:

    | Field | Description |
    |---|---|
    | `brand` | Brand name |
    | `sku_count` | Number of active SKUs in the brand |
    | `units_sold` | Gross units sold in the period |
    | `units_returned` | Gross units returned |
    | `credit_notes` | Credit-note units |
    | `transfer_orders` | Inter-warehouse transfer units |
    | `net_sales` | units_sold − credit_notes − transfer_orders |
    | `revenue` | Gross revenue (₹) |
    | `return_pct` | units_returned ÷ units_sold × 100 |
    | `growth_rate` | Weighted-average growth rate vs. trailing DRR (% change) |
    | `drr` | Brand-level DRR = sum of per-SKU avg_daily_run_rate |
    | `days_cover` | latest_net_stock ÷ DRR |
    | `current_days_coverage` | (latest_net_stock + stock_in_transit) ÷ DRR |
    | `weighted_avg_days_cover` | Stock-weighted average days cover across SKUs |
    | `latest_net_stock` | Zoho WH stock + FBA stock (most recent DB record) |
    | `latest_zoho_stock` | Zoho Pupscribe Warehouse stock (most recent DB record) |
    | `latest_fba_stock` | Amazon FBA stock (most recent DB record) |
    | `net_sellable_inventory_value` | latest_net_stock × per-SKU rate (₹) |
    | `stock_in_transit` | Open purchase-order stock in transit |
    | `total_cbm` | Total cubic metres for suggested orders |
    | `lead_time` | Brand lead time (days) from brand_logistics |
    | `safety_days` | Medium-mover safety days from brand_logistics |
    | `target_days` | lead_time + safety_days + order_processing (per-brand, default 10) |
    | `alert_level` | Colour-coded health indicator (see below) |
    | `order_count` | SKUs flagged as ORDER |
    | `excess_count` | SKUs flagged as EXCESS |
    | `no_movement_count` | SKUs flagged as NO MOVEMENT |
    | `fast_mover_count` / `medium_mover_count` / `slow_mover_count` | Movement breakdown |
    | `missed_sales_units` | Total unfulfilled demand units in the period |
    | `missed_sales_daily_units` | missed_sales_units ÷ period_days |
    | `missed_sales_value_total` | missed_sales_units × rate (₹) |
    | `missed_sales_daily_value` | missed_sales_value_total ÷ period_days (₹/day) |
    | `stock_classification` | `{ counts: {...}, pct: {...} }` with keys reorder_risk / healthy / heavy / overstock / dead |
    | `sku_lowest_10` | 10 SKUs with the lowest days cover (excluding dead/no-movement) |
    | `sku_highest_10` | 10 SKUs with the highest days cover (excluding those already in lowest 10) |

    ### `alert_level` – Brand Health Colour
    Based on `current_days_coverage` (on-hand + in-transit stock), so containers
    already on the water count towards covering the lead time.
    | Value | Colour | Condition |
    |---|---|---|
    | `0` | **Green** – Healthy | `current_days_coverage ≥ target_days` (brand has sufficient stock for the full replenishment cycle) |
    | `1` | **Yellow** – Caution | `lead_time ≤ current_days_coverage < target_days` (coverage spans lead time but is below target buffer) |
    | `2` | **Red** – Critical | `current_days_coverage < lead_time` (stock incl. transit will run out before a new order can arrive) |
    | `3` | **Grey** – No Movement | `drr = 0` (no sales recorded in the period) |

    ### SKU Stock Classification (per `stock_classification`)
    | Class | Condition |
    |---|---|
    | `reorder_risk` | `0 < days_cover < lead_time` |
    | `healthy` | `lead_time ≤ days_cover < 2 × lead_time` |
    | `heavy` | `2 × lead_time ≤ days_cover < 3 × lead_time` |
    | `overstock` | `days_cover ≥ 3 × lead_time` |
    | `dead` | `drr = 0` (no movement) |

    ### `totals`
    Cross-brand aggregates for all the same fields listed above.

    ### `global_stock_classification`
    Aggregate counts and percentages across all brands.

    ### `latest_stock_dates`
    - `zoho` – date of the most recent Zoho WH stock record in DB
    - `fba` – date of the most recent FBA stock record in DB
    """
    # Resolve dates – fall back to default 90-day window ending previous Sunday
    if not start_date or not end_date:
        start_date, end_date = _dashboard_kpi_default_dates()

    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        period_days = max(1, (end_dt - start_dt).days + 1)

        service = OptimizedMasterReportService(db)

        content, brand_logistics_map = await asyncio.gather(
            _generate_master_report_data(start_date, end_date, db, dashboard_mode=True),
            service.get_brand_logistics(),
        )

        combined_data = content.get("combined_data", [])

        # Batch-load product rates (for inventory value + missed-sales value)
        all_skus = {i.get("sku_code") for i in combined_data if i.get("sku_code")}
        product_data_map = await service.batch_load_all_product_data(all_skus) if all_skus else {}
        rate_map: Dict[str, float] = {
            sku: float(pdata.get("rate") or 0) for sku, pdata in product_data_map.items()
        }

        # Group by brand — skip products with no brand set or not active
        brand_groups: Dict[str, list] = defaultdict(list)
        for item in combined_data:
            if (item.get("purchase_status") or "").strip() != "active":
                continue
            brand = (item.get("brand") or "").strip()
            if not brand:
                continue
            brand_groups[brand].append(item)

        brands_kpi = [
            _aggregate_brand_kpi(brand, items, brand_logistics_map, rate_map, period_days)
            for brand, items in brand_groups.items()
        ]

        # Sort: most critical first, then alphabetical
        brands_kpi.sort(key=lambda x: (-x["alert_level"], x["brand"]))

        # Global totals
        total_net_sales = sum(b["net_sales"] for b in brands_kpi)
        total_net_stock = sum(b["latest_net_stock"] for b in brands_kpi)
        total_units_sold = sum(b["units_sold"] for b in brands_kpi)
        total_units_returned = sum(b["units_returned"] for b in brands_kpi)
        total_drr = round(sum(b["drr"] for b in brands_kpi), 4)
        total_days_cover = round(total_net_stock / total_drr, 2) if total_drr > 0 else 0.0
        total_transit = sum(b["stock_in_transit"] for b in brands_kpi)
        total_current_coverage = round((total_net_stock + total_transit) / total_drr, 2) if total_drr > 0 else 0.0

        # Global weighted avg days cover (weight = brand net stock, exclude dead brands)
        global_stock_w = sum(b["latest_net_stock"] for b in brands_kpi if b["drr"] > 0)
        global_weighted_doc = sum(
            b["latest_net_stock"] * b["weighted_avg_days_cover"]
            for b in brands_kpi if b["drr"] > 0 and b["latest_net_stock"] > 0
        )
        global_weighted_avg_doc = round(global_weighted_doc / global_stock_w, 2) if global_stock_w > 0 else 0.0

        # Global stock classification counts
        global_class_counts: Dict[str, int] = {"reorder_risk": 0, "healthy": 0, "heavy": 0, "overstock": 0, "dead": 0}
        for b in brands_kpi:
            for k, v in b["stock_classification"]["counts"].items():
                global_class_counts[k] = global_class_counts.get(k, 0) + v
        total_global_skus = sum(global_class_counts.values())
        global_class_pct = {
            k: round(v / total_global_skus * 100, 1) if total_global_skus > 0 else 0.0
            for k, v in global_class_counts.items()
        }

        response_payload = {
            "period": {"start_date": start_date, "end_date": end_date, "days": period_days},
            "brands": brands_kpi,
            "totals": {
                "brand_count": len(brands_kpi),
                "sku_count": sum(b["sku_count"] for b in brands_kpi),
                "units_sold": round(total_units_sold, 2),
                "units_returned": round(total_units_returned, 2),
                "credit_notes": round(sum(b["credit_notes"] for b in brands_kpi), 2),
                "transfer_orders": round(sum(b["transfer_orders"] for b in brands_kpi), 2),
                "net_sales": round(total_net_sales, 2),
                "revenue": round(sum(b["revenue"] for b in brands_kpi), 2),
                "return_pct": round(total_units_returned / total_units_sold * 100, 2) if total_units_sold > 0 else 0.0,
                "drr": total_drr,
                "latest_net_stock": round(total_net_stock, 2),
                "latest_zoho_stock": round(sum(b["latest_zoho_stock"] for b in brands_kpi), 2),
                "latest_fba_stock": round(sum(b["latest_fba_stock"] for b in brands_kpi), 2),
                "net_sellable_inventory_value": round(sum(b["net_sellable_inventory_value"] for b in brands_kpi), 2),
                "stock_in_transit": round(total_transit, 2),
                "total_cbm": round(sum(b["total_cbm"] for b in brands_kpi), 4),
                "days_cover": total_days_cover,
                "current_days_coverage": total_current_coverage,
                "weighted_avg_days_cover": global_weighted_avg_doc,
                "missed_sales_units": round(sum(b["missed_sales_units"] for b in brands_kpi), 2),
                "missed_sales_daily_units": round(sum(b["missed_sales_daily_units"] for b in brands_kpi), 2),
                "missed_sales_value_total": round(sum(b["missed_sales_value_total"] for b in brands_kpi), 2),
                "missed_sales_daily_value": round(sum(b["missed_sales_daily_value"] for b in brands_kpi), 2),
            },
            "global_stock_classification": {
                "counts": global_class_counts,
                "pct": global_class_pct,
            },
            "latest_stock_dates": content.get("latest_stock_dates", {}),
            "meta": content.get("meta", {}),
        }
        return JSONResponse(status_code=200, content=response_payload)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating dashboard KPI: {e}")
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Failed to generate dashboard KPI: {str(e)}")


@router.get("/dashboard-kpi/download")
async def download_dashboard_kpi(
    breakdown: str = Query("brand", description="'brand' for brand-level or 'product' for SKU-level"),
    start_date: Optional[str] = Query(None, description="Start date YYYY-MM-DD (default: 90 days ending previous Sunday)"),
    end_date: Optional[str] = Query(None, description="End date YYYY-MM-DD (default: previous Sunday)"),
    db=Depends(get_database),
):
    """
    Download the Stock Cover KPI as an Excel (.xlsx) file.

    Runs the same 90-day master-report pipeline as `GET /dashboard-kpi` and writes
    the aggregated result to a single-sheet Excel workbook.

    ## Query Parameters
    - `breakdown=brand` *(default)* – one row per brand, aggregated KPIs
    - `breakdown=product` – one row per active SKU, individual KPIs

    ## Excel Sheet: `Brand KPI` (breakdown=brand)
    One row per brand. Columns:

    **Identity & Movers**
    - `Brand`, `SKUs`, `Fast Movers`, `Medium Movers`, `Slow Movers`

    **Sales**
    - `Units Sold`, `Units Returned`, `Return %`
    - `Credit Notes`, `Transfer Orders`, `Net Sales`, `Revenue (₹)`
    - `Growth Rate (%)`

    **Inventory & DRR**
    - `DRR (Net Sales / 90)` – brand-level DRR (sum of per-SKU DRR)
    - `Zoho WH Stock ({latest_date})`, `FBA Stock ({latest_date})`, `Latest Net Stock`
    - `Net Inv. Value (₹)` – latest stock × per-SKU rate
    - `Stock in Transit`, `Total CBM`

    **Coverage**
    - `Days Cover (Stock / DRR)` – latest_net_stock ÷ DRR
    - `Weighted Avg Days Cover` – stock-weighted average days cover across SKUs
    - `Current Days Coverage (Stock+Transit/DRR)`
    - `Lead Time (days)`, `Safety Days (Medium)`, `Target Days`

    **Stock Classification**
    - `SKUs - Reorder Risk / Healthy / Heavy / Overstock / Dead`
    - `% Reorder Risk / Healthy / Heavy / Overstock / Dead`
    - `SKUs Needing Order`, `SKUs in Excess`, `SKUs No Movement`

    **Missed Sales**
    - `Missed Sales (units, 90d)`, `Missed Sales (units/day)`
    - `Missed Sales Value (₹, 90d)`, `Missed Sales Value (₹/day)`

    ### Conditional Row Formatting (Brand KPI sheet)
    Based on Current Days Coverage (on-hand + in-transit stock), matching `alert_level`.
    | Colour | Condition | Meaning |
    |---|---|---|
    | Red | `0 < Current Days Coverage < Lead Time` | Critical – stock incl. transit will run out before a new order can arrive |
    | Yellow | `Lead Time ≤ Current Days Coverage < Target Days` | Caution – below the target safety buffer |
    | (no fill) | `Current Days Coverage ≥ Target Days` or no movement | Healthy / no movement – no action required |

    ## Excel Sheet: `Product KPI` (breakdown=product)
    One row per active SKU. Columns:

    **Identity**
    - `Brand`, `Purchase Status`, `Movement`, `Stock Class`
    - `SKU Code`, `Item Name`

    **Sales**
    - `Units Sold`, `Units Returned`, `Return %`
    - `Credit Notes`, `Transfer Orders`, `Net Sales`, `Revenue (₹)`
    - `Growth Rate (%)`, `Avg Daily Run Rate`

    **Inventory**
    - `Zoho WH Stock ({latest_date})`, `FBA Stock ({latest_date})`, `Latest Net Stock`
    - `Net Inv. Value (₹)`, `Stock in Transit`

    **Coverage & Orders**
    - `Days Cover (Stock / DRR)`, `Current Days Coverage`
    - `Lead Time (days)`, `Safety Days`, `Target Days`
    - `Excess / Order`, `Order Qty`, `Order Qty + Extra Qty`, `Order Qty (Rounded)`
    - `Total CBM`, `Days Total Inventory Lasts`

    **Missed Sales**
    - `Missed Sales (units)`, `Missed Sales DRR (units/day)`
    - `Missed Sales Value (₹)`, `Missed Sales Daily Value (₹)`

    ## Response
    Returns a binary `.xlsx` file as a streaming attachment.
    Filename format: `stock_cover_kpi_{breakdown}_{start_date}_to_{end_date}_{YYYYMMDD_HHMMSS}.xlsx`
    """
    try:
        if not start_date or not end_date:
            start_date, end_date = _dashboard_kpi_default_dates()
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        period_days = max(1, (end_dt - start_dt).days + 1)

        service = OptimizedMasterReportService(db)
        content, brand_logistics_map = await asyncio.gather(
            _generate_master_report_data(start_date, end_date, db, dashboard_mode=True),
            service.get_brand_logistics(),
        )

        combined_data = content.get("combined_data", [])
        latest_stock_dates = content.get("latest_stock_dates", {})

        # Build rate_map from mrp already attached to each combined_data item by
        # _generate_master_report_data — avoids a redundant products collection query.
        rate_map: Dict[str, float] = {
            item.get("sku_code"): float(item.get("mrp") or 0)
            for item in combined_data
            if item.get("sku_code")
        }

        def _fmt_date(d: str) -> str:
            try:
                return datetime.strptime(d, "%Y-%m-%d").strftime("%d %b %Y")
            except Exception:
                return d or "Latest"

        _latest_zoho_label = _fmt_date(latest_stock_dates.get("zoho", ""))
        _latest_fba_label = _fmt_date(latest_stock_dates.get("fba", ""))
        _etrade_inv_label = _fmt_date(latest_stock_dates.get("etrade", "")) or _fmt_date(end_date)

        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            if breakdown == "product":
                rows = []
                for item in combined_data:
                    if (item.get("purchase_status") or "").strip() != "active":
                        continue
                    m = item.get("combined_metrics", {})
                    sku = item.get("sku_code", "")
                    rate = rate_map.get(sku, 0)
                    lead_time_item = item.get("lead_time", 60)
                    item_drr = m.get("avg_daily_run_rate", 0)
                    item_doc = item.get("current_days_coverage", 0)
                    rows.append({
                        "Brand": item.get("brand", ""),
                        "Purchase Status": item.get("purchase_status", ""),
                        "Movement": item.get("movement", ""),
                        "Stock Class": _classify_sku_stock(item_doc, lead_time_item, item_drr),
                        "SKU Code": sku,
                        "Item Name": item.get("item_name", ""),
                        "Units Sold": m.get("total_units_sold", 0),
                        "Units Returned": m.get("total_units_returned", 0),
                        "Return %": round(m.get("total_units_returned", 0) / m.get("total_units_sold", 1) * 100, 2) if m.get("total_units_sold", 0) > 0 else 0,
                        "Credit Notes": m.get("total_credit_notes", 0),
                        "Transfer Orders": m.get("transfer_orders", 0),
                        "Net Sales": m.get("total_sales", 0),
                        "Revenue (₹)": m.get("total_amount", 0),
                        "Growth Rate (%)": item.get("growth_rate"),
                        "Avg Daily Run Rate": item_drr,
                        f"Zoho WH Stock ({_latest_zoho_label})": item.get("latest_zoho_stock", 0),
                        f"FBA Stock ({_latest_fba_label})": item.get("latest_fba_stock", 0),
                        "Latest Net Stock": item.get("latest_total_stock", 0),
                        "Net Inv. Value (₹)": round(item.get("latest_total_stock", 0) * rate, 2),
                        "Stock in Transit": item.get("total_stock_in_transit", 0),
                        "Days Cover (Stock / DRR)": m.get("avg_days_of_coverage", 0),
                        "Current Days Coverage": item_doc,
                        "Lead Time (days)": lead_time_item,
                        "Safety Days": item.get("safety_days", 0),
                        "Target Days": item.get("target_days", 0),
                        "Excess / Order": item.get("excess_or_order", ""),
                        "Order Qty": item.get("order_qty", 0),
                        "Order Qty + Extra Qty": item.get("order_qty_plus_extra_qty", 0),
                        "Order Qty (Rounded)": item.get("order_qty_plus_extra_qty_rounded", 0),
                        "Total CBM": item.get("total_cbm", 0),
                        "Days Total Inventory Lasts": item.get("days_total_inventory_lasts", 0),
                        f"Etrade Inventory ({_etrade_inv_label})": m.get("etrade_inventory", 0),
                        "Etrade DRR": m.get("etrade_drr", 0),
                        "Days Total Inventory Lasts (Etrade)": m.get("etrade_days_inventory_lasts", 0),
                        "Missed Sales (units)": item.get("missed_sales", 0),
                        "Missed Sales DRR (units/day)": item.get("missed_sales_drr", 0),
                        "Missed Sales Value (₹)": round(item.get("missed_sales", 0) * rate, 2),
                        "Missed Sales Daily Value (₹)": round(item.get("missed_sales_drr", 0) * rate, 2),
                    })
                if rows:
                    pd.DataFrame(rows).to_excel(writer, sheet_name="Product KPI", index=False)

            else:
                # Brand aggregation — skip products with no brand set or not active
                brand_groups: Dict[str, list] = defaultdict(list)
                for item in combined_data:
                    if (item.get("purchase_status") or "").strip() != "active":
                        continue
                    brand = (item.get("brand") or "").strip()
                    if not brand:
                        continue
                    brand_groups[brand].append(item)

                rows = []
                for brand in sorted(brand_groups.keys()):
                    items = brand_groups[brand]
                    kpi = _aggregate_brand_kpi(brand, items, brand_logistics_map, rate_map, period_days)
                    cc = kpi["stock_classification"]["counts"]
                    cp = kpi["stock_classification"]["pct"]
                    rows.append({
                        "Brand": brand,
                        "SKUs": kpi["sku_count"],
                        "Fast Movers": kpi["fast_mover_count"],
                        "Medium Movers": kpi["medium_mover_count"],
                        "Slow Movers": kpi["slow_mover_count"],
                        "Units Sold": kpi["units_sold"],
                        "Units Returned": kpi["units_returned"],
                        "Return %": kpi["return_pct"],
                        "Credit Notes": kpi["credit_notes"],
                        "Transfer Orders": kpi["transfer_orders"],
                        "Net Sales": kpi["net_sales"],
                        "Revenue (₹)": kpi["revenue"],
                        "Growth Rate (%)": kpi["growth_rate"],
                        "DRR (Net Sales / 90)": kpi["drr"],
                        f"Zoho WH Stock ({_latest_zoho_label})": kpi["latest_zoho_stock"],
                        f"FBA Stock ({_latest_fba_label})": kpi["latest_fba_stock"],
                        "Latest Net Stock": kpi["latest_net_stock"],
                        "Net Inv. Value (₹)": kpi["net_sellable_inventory_value"],
                        "Stock in Transit": kpi["stock_in_transit"],
                        "Total CBM": kpi["total_cbm"],
                        "Days Cover (Stock / DRR)": kpi["days_cover"],
                        "Weighted Avg Days Cover": kpi["weighted_avg_days_cover"],
                        "Current Days Coverage (Stock+Transit/DRR)": kpi["current_days_coverage"],
                        "Lead Time (days)": kpi["lead_time"],
                        "Safety Days (Medium)": kpi["safety_days"],
                        "Target Days": kpi["target_days"],
                        "SKUs - Reorder Risk": cc.get("reorder_risk", 0),
                        "% Reorder Risk": cp.get("reorder_risk", 0),
                        "SKUs - Healthy": cc.get("healthy", 0),
                        "% Healthy": cp.get("healthy", 0),
                        "SKUs - Heavy": cc.get("heavy", 0),
                        "% Heavy": cp.get("heavy", 0),
                        "SKUs - Overstock": cc.get("overstock", 0),
                        "% Overstock": cp.get("overstock", 0),
                        "SKUs - Dead": cc.get("dead", 0),
                        "% Dead": cp.get("dead", 0),
                        "SKUs Needing Order": kpi["order_count"],
                        "SKUs in Excess": kpi["excess_count"],
                        "SKUs No Movement": kpi["no_movement_count"],
                        "Missed Sales (units, 90d)": kpi["missed_sales_units"],
                        "Missed Sales (units/day)": kpi["missed_sales_daily_units"],
                        "Missed Sales Value (₹, 90d)": kpi["missed_sales_value_total"],
                        "Missed Sales Value (₹/day)": kpi["missed_sales_daily_value"],
                    })

                if rows:
                    df = pd.DataFrame(rows)
                    df.to_excel(writer, sheet_name="Brand KPI", index=False)

                    from openpyxl.styles import PatternFill
                    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    ws = writer.sheets["Brand KPI"]
                    # Highlight on coverage incl. transit so it matches alert_level
                    dc_col = list(df.columns).index("Current Days Coverage (Stock+Transit/DRR)") + 1
                    lt_col = list(df.columns).index("Lead Time (days)") + 1
                    tg_col = list(df.columns).index("Target Days") + 1

                    for row_idx in range(2, len(rows) + 2):
                        dc = ws.cell(row=row_idx, column=dc_col).value or 0
                        lt = ws.cell(row=row_idx, column=lt_col).value or 60
                        tg = ws.cell(row=row_idx, column=tg_col).value or 95
                        fill = red_fill if dc > 0 and dc < lt else (yellow_fill if dc > 0 and dc < tg else None)
                        if fill:
                            for col_idx in range(1, len(df.columns) + 1):
                                ws.cell(row=row_idx, column=col_idx).fill = fill

        excel_buffer.seek(0)
        file_bytes = excel_buffer.read()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"stock_cover_kpi_{breakdown}_{start_date}_to_{end_date}_{timestamp}.xlsx"

        return Response(
            content=file_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Length": str(len(file_bytes)),
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating dashboard KPI download: {e}")
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Failed to generate KPI download: {str(e)}")

