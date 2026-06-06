import pandas as pd
from io import BytesIO
from bson import ObjectId
from fastapi import APIRouter, UploadFile, File, HTTPException, status, Depends, Query
from fastapi.responses import JSONResponse, StreamingResponse
from datetime import datetime, timedelta, timezone
from pymongo import ASCENDING
from pymongo.errors import PyMongoError
from ..database import get_database, serialize_mongo_document
import os
import requests
import json
import time, io
import gzip
import re
from typing import Optional, Dict, Any, List
from dotenv import load_dotenv
import logging
import asyncio
from .amazon_vc import router as amazon_vendor_router

# --- Configuration ---
# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ENV
load_dotenv()

SKU_COLLECTION = "amazon_sku_mapping"
SALES_COLLECTION = "amazon_sales_traffic"
INVENTORY_COLLECTION = "amazon_ledger"
FBA_RETURNS_COLLECTION = "amazon_fba_returns"
SC_RETURNS_COLLECTION = "amazon_seller_central_returns"
SELLER_FLEX_RETURNS_COLLECTION = "amazon_seller_flex_returns"
VENDOR_CENTRAL_RETURNS_COLLECTION = "amazon_vendor_central_returns"

router = APIRouter()

router.include_router(amazon_vendor_router, prefix="/vendor")


def format_date_ranges(dates: List[datetime]) -> str:
    """
    Format a list of dates into ranges or individual dates.
    Continuous dates are shown as ranges (e.g., "1 Jan 2025 - 5 Jan 2025")
    Discontinuous dates are shown separated by commas.

    Args:
        dates: List of datetime objects

    Returns:
        Formatted string with date ranges
    """
    if not dates:
        return ""

    # Sort dates
    sorted_dates = sorted(dates)

    ranges = []
    start_date = sorted_dates[0]
    end_date = sorted_dates[0]

    for i in range(1, len(sorted_dates)):
        current_date = sorted_dates[i]
        # Check if current date is consecutive to end_date
        if (current_date - end_date).days == 1:
            end_date = current_date
        else:
            # Save the current range
            if start_date == end_date:
                ranges.append(start_date.strftime("%d %b %Y"))
            else:
                ranges.append(f"{start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}")
            # Start new range
            start_date = current_date
            end_date = current_date

    # Add the last range
    if start_date == end_date:
        ranges.append(start_date.strftime("%d %b %Y"))
    else:
        ranges.append(f"{start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}")

    return ", ".join(ranges)


async def fetch_last_n_days_amazon_ledger(
    db, end_datetime: datetime, asins: list, n_days: int = 90, report_type: str = "fba+seller_flex"
) -> Dict:
    """
    Fetch the last N days an item was in stock for Amazon ledger (FBA/Seller Flex).

    Args:
        db: Database connection
        end_datetime: End date to look back from
        asins: List of ASINs
        n_days: Number of days to fetch (default 90)
        report_type: Type of report (fba, seller_flex, fba+seller_flex)

    Returns:
        Dict mapping ASIN to formatted date ranges string
    """
    try:
        ledger_collection = db["amazon_ledger"]
        lookback_limit = end_datetime - timedelta(days=730)  # Look back 2 years

        logger.info(f"Fetching last {n_days} days in stock (Amazon Ledger) for {len(asins)} ASINs from {lookback_limit.date()} to {end_datetime.date()}")

        # Build match conditions based on report type
        match_condition = {
            "asin": {"$in": asins},
            "date": {"$gte": lookback_limit, "$lte": end_datetime},
            "disposition": "SELLABLE",
        }

        if report_type == "fba":
            match_condition["location"] = {"$ne": "VKSX"}
        elif report_type == "seller_flex":
            match_condition["location"] = "VKSX"

        pipeline = [
            {"$match": match_condition},
            {
                "$project": {
                    "asin": 1,
                    "date": 1,
                    "stock": "$ending_warehouse_balance",
                }
            },
            {"$match": {"stock": {"$gt": 0}}},
            {"$sort": {"asin": 1, "date": -1}},
            {
                "$group": {
                    "_id": "$asin",
                    "stock_dates": {"$push": "$date"}
                }
            },
            {
                "$project": {
                    "_id": 1,
                    "stock_dates": {"$slice": ["$stock_dates", n_days]}
                }
            }
        ]

        def _run_aggregation():
            cursor = ledger_collection.aggregate(pipeline, allowDiskUse=True)
            stock_date_ranges = {}
            for doc in cursor:
                asin = doc["_id"]
                dates = doc.get("stock_dates", [])
                if dates:
                    stock_date_ranges[asin] = format_date_ranges(dates)
                else:
                    stock_date_ranges[asin] = ""
            return stock_date_ranges

        stock_date_ranges = await asyncio.to_thread(_run_aggregation)
        logger.info(f"Fetched last {n_days} days (Amazon Ledger): {len(stock_date_ranges)} ASINs with data")
        return stock_date_ranges

    except Exception as e:
        logger.error(f"Error fetching last N days (Amazon Ledger): {e}", exc_info=True)
        return {}


async def fetch_last_n_days_vendor_inventory(
    db, end_datetime: datetime, asins: list, n_days: int = 90
) -> Dict:
    """
    Fetch the last N days an item was in stock for Amazon Vendor Central inventory.

    Args:
        db: Database connection
        end_datetime: End date to look back from
        asins: List of ASINs
        n_days: Number of days to fetch (default 90)

    Returns:
        Dict mapping ASIN to formatted date ranges string
    """
    try:
        inventory_collection = db["amazon_vendor_inventory"]
        lookback_limit = end_datetime - timedelta(days=730)  # Look back 2 years

        logger.info(f"Fetching last {n_days} days in stock (Vendor) for {len(asins)} ASINs from {lookback_limit.date()} to {end_datetime.date()}")

        pipeline = [
            {
                "$match": {
                    "asin": {"$in": asins},
                    "date": {"$gte": lookback_limit, "$lte": end_datetime},
                }
            },
            {
                "$project": {
                    "asin": 1,
                    "date": 1,
                    "stock": "$sellableOnHandInventoryUnits",
                }
            },
            {"$match": {"stock": {"$gt": 0}}},
            {"$sort": {"asin": 1, "date": -1}},
            {
                "$group": {
                    "_id": "$asin",
                    "stock_dates": {"$push": "$date"}
                }
            },
            {
                "$project": {
                    "_id": 1,
                    "stock_dates": {"$slice": ["$stock_dates", n_days]}
                }
            }
        ]

        def _run_aggregation():
            cursor = inventory_collection.aggregate(pipeline, allowDiskUse=True)
            stock_date_ranges = {}
            for doc in cursor:
                asin = doc["_id"]
                dates = doc.get("stock_dates", [])
                if dates:
                    stock_date_ranges[asin] = format_date_ranges(dates)
                else:
                    stock_date_ranges[asin] = ""
            return stock_date_ranges

        stock_date_ranges = await asyncio.to_thread(_run_aggregation)
        logger.info(f"Fetched last {n_days} days (Vendor): {len(stock_date_ranges)} ASINs with data")
        return stock_date_ranges

    except Exception as e:
        logger.error(f"Error fetching last N days (Vendor): {e}", exc_info=True)
        return {}

# SP API Configuration
REFRESH_TOKEN = os.getenv("SP_REFRESH_TOKEN")
GRANT_TYPE = os.getenv("SP_GRANT_TYPE", "refresh_token")
CLIENT_ID = os.getenv("SP_CLIENT_ID")
CLIENT_SECRET = os.getenv("SP_CLIENT_SECRET")
LOGIN_URL = os.getenv("SP_LOGIN_URL", "https://api.amazon.com/auth/o2/token")
MARKETPLACE_ID = os.getenv("SP_MARKETPLACE_ID", "A21TJRUUN4KGV")  # IN marketplace
SP_API_BASE_URL = os.getenv(
    "SP_API_BASE_URL", "https://sellingpartnerapi-na.amazon.com"
)

# Global variable to store access token and expiry
_sc_access_token = None
_sc_token_expires_at = None


def get_amazon_token() -> str:
    """
    Get a new access token from Amazon SP API using refresh token
    """
    global _sc_access_token, _sc_token_expires_at

    # Check if current token is still valid (with 5 minute buffer)
    if (
        _sc_access_token
        and _sc_token_expires_at
        and datetime.now() < (_sc_token_expires_at - timedelta(minutes=5))
    ):
        logger.info("Using cached access token")
        return _sc_access_token

    logger.info("Requesting new access token from Amazon")
    try:
        payload = {
            "grant_type": GRANT_TYPE,
            "refresh_token": REFRESH_TOKEN,
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
        }

        headers = {"Content-Type": "application/x-www-form-urlencoded"}

        response = requests.post(LOGIN_URL, data=payload, headers=headers)
        response.raise_for_status()

        token_data = response.json()
        _sc_access_token = token_data.get("access_token")
        expires_in = token_data.get("expires_in", 3600)  # Default 1 hour
        _sc_token_expires_at = datetime.now() + timedelta(seconds=expires_in)

        logger.info(f"Successfully obtained new Amazon SP API token (expires in {expires_in}s)")
        logger.info(f"Token prefix: {_sc_access_token[:20]}..." if _sc_access_token else "No token received")
        return _sc_access_token

    except requests.exceptions.RequestException as e:
        # Log detailed error information
        logger.error(f"Error getting Amazon token: {e}")

        # Try to get error response body if available
        if hasattr(e, 'response') and e.response is not None:
            try:
                error_body = e.response.json()
                logger.error(f"Token error response: {json.dumps(error_body, indent=2)}")
            except:
                logger.error(f"Token error response text: {e.response.text}")

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to authenticate with Amazon SP API: {str(e)}",
        )
    except Exception as e:
        logger.error(f"Unexpected error getting Amazon token: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An unexpected error occurred during authentication: {str(e)}",
        )


def make_sp_api_request(
    endpoint: str, method: str = "GET", params: Dict = None, data: Dict = None
) -> Dict:
    """
    Make a request to Amazon SP API with proper authentication
    """
    access_token = get_amazon_token()

    headers = {"x-amz-access-token": access_token, "Content-Type": "application/json"}

    url = f"{SP_API_BASE_URL}{endpoint}"

    try:
        if method.upper() == "GET":
            response = requests.get(url, headers=headers, params=params)
        elif method.upper() == "POST":
            response = requests.post(url, headers=headers, json=data, params=params)
        else:
            raise ValueError(f"Unsupported HTTP method: {method}")

        response.raise_for_status()
        return response.json()

    except requests.exceptions.RequestException as e:
        # Log detailed error information
        error_details = {
            "error": str(e),
            "url": url,
            "method": method,
        }

        # Try to get error response body if available
        if hasattr(e, 'response') and e.response is not None:
            try:
                error_body = e.response.json()
                error_details["response_body"] = error_body
                logger.error(f"SP API request failed: {e}")
                logger.error(f"Response body: {json.dumps(error_body, indent=2)}")
            except:
                error_details["response_text"] = e.response.text
                logger.error(f"SP API request failed: {e}")
                logger.error(f"Response text: {e.response.text}")
        else:
            logger.error(f"SP API request failed: {e}")

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Amazon SP API request failed: {str(e)}",
        )


def create_report(
    report_type: str,
    start_date: str,
    end_date: str,
    marketplace_ids: List[str] = None,
    report_options: Dict = None,
) -> str:
    """
    Create a report request and return the report document ID
    """
    if marketplace_ids is None:
        marketplace_ids = [MARKETPLACE_ID]

    endpoint = "/reports/2021-06-30/reports"

    report_data = {
        "reportType": report_type,
        "dataStartTime": start_date,
        "dataEndTime": end_date,
        "marketplaceIds": marketplace_ids,
    }

    # Add report options if provided (for ledger reports)
    if report_options:
        report_data["reportOptions"] = report_options

    response = make_sp_api_request(endpoint, method="POST", data=report_data)
    return response.get("reportId")


def get_report_status(report_document_id: str) -> Dict:
    """
    Check the status of a report
    """
    endpoint = f"/reports/2021-06-30/reports/{report_document_id}"
    return make_sp_api_request(endpoint)


def download_report_data(report_document_id: str, is_gzipped: bool = False) -> str:
    """
    Download report data using the document ID
    """
    endpoint = f"/reports/2021-06-30/documents/{report_document_id}"
    document_info = make_sp_api_request(endpoint)

    # Get the download URL
    download_url = document_info.get("url")
    if not download_url:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="No download URL found in report document",
        )

    # Download the actual report data
    response = requests.get(download_url)
    response.raise_for_status()

    # Handle gzipped content (for ledger reports)
    if is_gzipped:
        try:
            return gzip.decompress(response.content).decode("utf-8")
        except Exception as e:
            logger.error(f"Error decompressing gzipped content: {e}")
            # Fallback to regular content if not actually gzipped
            return response.text

    return response.text


def get_amazon_sales_traffic_data(
    start_date: str, end_date: str, db: Any, marketplace_ids: List[str] = None
) -> List[Dict]:
    """
    Fetch sales and traffic data from Amazon SP API (ASIN-wise daily data)
    """
    if marketplace_ids is None:
        marketplace_ids = [MARKETPLACE_ID]

    try:
        # Create sales and traffic report
        report_id = create_report(
            report_type="GET_SALES_AND_TRAFFIC_REPORT",
            start_date=start_date,
            end_date=end_date,
            marketplace_ids=marketplace_ids,
        )

        # Add validation for report_id
        if not report_id:
            raise ValueError("Failed to create report - no report ID returned")

        # Wait for report to be ready (with timeout)
        max_wait_time = 300  # 5 minutes
        wait_time = 0

        while wait_time < max_wait_time:
            report_status = get_report_status(report_id)

            # Add validation for report_status
            if not report_status:
                raise ValueError("Failed to get report status - empty response")

            processing_status = report_status.get("processingStatus")
            report_document_id = report_status.get("reportDocumentId")

            if processing_status == "DONE":
                break
            elif processing_status == "FATAL":
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="Sales and traffic report processing failed",
                )

            time.sleep(10)
            wait_time += 10

        if wait_time >= max_wait_time:
            raise HTTPException(
                status_code=status.HTTP_408_REQUEST_TIMEOUT,
                detail="Sales and traffic report processing timed out",
            )

        # Validate report_document_id before downloading
        if not report_document_id:
            raise ValueError("No report document ID available")

        # Download and parse report data (JSON format)
        report_data = download_report_data(report_document_id, is_gzipped=True)

        # Add validation for report_data
        if not report_data:
            raise ValueError("Downloaded report data is empty")

        # Parse JSON data with better error handling
        try:
            json_data = json.loads(report_data)
        except json.JSONDecodeError as json_error:
            raise ValueError(f"Failed to parse JSON report data: {json_error}")

        # Validate json_data structure
        if not isinstance(json_data, dict):
            raise ValueError("Invalid JSON structure - expected dictionary")

        # Extract ASIN-wise daily data
        sales_traffic_data = []

        # Process salesAndTrafficByAsin data
        if "salesAndTrafficByAsin" in json_data:
            asin_data_list = json_data["salesAndTrafficByAsin"]

            if not isinstance(asin_data_list, list):
                raise ValueError("salesAndTrafficByAsin is not a list")

            for asin_data in asin_data_list:
                asin = asin_data["parentAsin"]
                asin_data["date"] = datetime.fromisoformat(start_date)
                asin_data["created_at"] = datetime.now()
                result = db[SALES_COLLECTION].find_one(
                    {"parentAsin": asin, "date": asin_data["date"]}
                )
                if not result:
                    sales_traffic_data.append(asin_data)
        else:
            logger.warning("No 'salesAndTrafficByAsin' data found in report")

        return sales_traffic_data

    except HTTPException:
        # Re-raise HTTPExceptions as-is
        raise
    except ValueError as ve:
        logger.error(f"Validation error in sales and traffic data: {ve}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Data validation error: {str(ve)}",
        )
    except Exception as e:
        logger.error(f"Unexpected error fetching sales and traffic data: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch sales and traffic data: {str(e)}",
        )


def get_amazon_ledger_data(
    start_date: str,
    end_date: str,
    db: Any,
    marketplace_ids: List[str] = None,
    aggregate_by_location: str = "COUNTRY",
    aggregated_by_time_period: str = "DAILY",
) -> List[Dict]:
    logger.info(f"Processing dates - start: {start_date}, end: {end_date}")
    """
    Fetch ledger summary data from Amazon SP API
    """
    if marketplace_ids is None:
        marketplace_ids = [MARKETPLACE_ID]

    try:
        # Report options for ledger data
        report_options = {
            "aggregateByLocation": aggregate_by_location,
            "aggregatedByTimePeriod": aggregated_by_time_period,
        }

        # Create ledger report
        report_id = create_report(
            report_type="GET_LEDGER_SUMMARY_VIEW_DATA",
            start_date=start_date,
            end_date=end_date,
            marketplace_ids=marketplace_ids,
            report_options=report_options,
        )

        # Wait for report to be ready (with timeout)
        max_wait_time = 300  # 5 minutes
        wait_time = 0
        report_document_id = ""
        while wait_time < max_wait_time:
            report_status = get_report_status(report_id)
            processing_status = report_status.get("processingStatus")
            report_document_id = report_status.get("reportDocumentId")

            if processing_status == "DONE":
                break
            elif processing_status == "FATAL":
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="Ledger report processing failed",
                )

            time.sleep(10)
            wait_time += 10

        if wait_time >= max_wait_time:
            raise HTTPException(
                status_code=status.HTTP_408_REQUEST_TIMEOUT,
                detail="Ledger report processing timed out",
            )

        # Download and parse report data (TSV format, gzipped)
        report_data = download_report_data(report_document_id, is_gzipped=True)

        # Convert TSV data to list of dictionaries
        df = pd.read_csv(BytesIO(report_data.encode("utf-8")), sep="\t")
        ledger_data = df.to_dict("records")
        # Add metadata
        normalized_data = []
        for record in ledger_data:
            normalized_record = {
                key.lower().replace(" ", "_"): value for key, value in record.items()
            }
            normalized_record["date"] = datetime.fromisoformat(start_date)
            normalized_record["created_at"] = datetime.now()
            normalized_record["report_type"] = "ledger"
            result = db.amazon_ledger.find_one(
                {
                    "date": normalized_record["date"],
                    "asin": normalized_record["asin"],
                    "location": normalized_record["location"],
                    "ending_warehouse_balance": normalized_record[
                        "ending_warehouse_balance"
                    ],
                }
            )
            if not result:
                normalized_data.append(normalized_record)

        return normalized_data

    except Exception as e:
        logger.error(f"Error fetching ledger data: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch ledger data: {str(e)}",
        )


def get_amazon_inventory_data() -> List[Dict]:
    """
    Fetch inventory data from Amazon SP API
    """
    try:
        endpoint = "/fba/inventory/v1/summaries"
        params = {
            "granularityType": "Marketplace",
            "granularityId": MARKETPLACE_ID,
            "marketplaceIds": MARKETPLACE_ID,
        }

        response = make_sp_api_request(endpoint, params=params)
        inventory_summaries = response.get("inventorySummaries", [])

        # Add metadata
        for record in inventory_summaries:
            record["created_at"] = datetime.now()
            record["_report_type"] = "inventory"

        return inventory_summaries

    except Exception as e:
        logger.error(f"Error fetching inventory data: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch inventory data: {str(e)}",
        )


async def insert_data_to_db(
    collection_name: str, data: List[Dict], db=Depends(get_database)
) -> int:
    """
    Insert data into MongoDB collection
    """
    try:
        if not data:
            return 0

        collection = db[collection_name]

        # Insert data
        result = collection.insert_many(data)
        logger.info(
            f"Inserted {len(result.inserted_ids)} records into {collection_name}"
        )

        return len(result.inserted_ids)

    except PyMongoError as e:
        logger.error(f"Database error inserting data: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database error: {str(e)}",
        )
    except Exception as e:
        logger.error(f"Unexpected error inserting data: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An unexpected error occurred: {str(e)}",
        )


def get_field_mapping_for_amazon_field(amazon_field: str) -> str:
    """
    Dynamically map Amazon field names to our expected field names
    """
    normalized = normalize_field_name(amazon_field)

    # Define mapping rules based on common patterns
    field_mappings = {
        # Order fields
        "order_id": "order_id",
        "orderid": "order_id",
        "amazon_order_id": "order_id",
        "amazonorderid": "order_id",
        # Dates
        "order_date": "order_date",
        "orderdate": "order_date",
        "return_request_date": "return_request_date",
        "returnrequestdate": "return_request_date",
        "return_delivery_date": "return_delivery_date",
        "returndeliverydate": "return_delivery_date",
        # Return information
        "return_request_status": "return_request_status",
        "returnrequeststatus": "return_request_status",
        "return_quantity": "return_quantity",
        "returnquantity": "return_quantity",
        "quantity": "return_quantity",
        "return_reason": "return_reason",
        "returnreason": "return_reason",
        "return_type": "return_type",
        "returntype": "return_type",
        # RMA fields
        "amazon_rma_id": "amazon_rma_id",
        "amazonrmaid": "amazon_rma_id",
        "merchant_rma_id": "merchant_rma_id",
        "merchantrmaid": "merchant_rma_id",
        # Shipping/Label fields
        "label_type": "label_type",
        "labeltype": "label_type",
        "label_cost": "label_cost",
        "labelcost": "label_cost",
        "currency_code": "currency_code",
        "currencycode": "currency_code",
        "return_carrier": "return_carrier",
        "returncarrier": "return_carrier",
        "tracking_id": "tracking_id",
        "trackingid": "tracking_id",
        "label_to_be_paid_by": "label_to_be_paid_by",
        "labeltobepaidby": "label_to_be_paid_by",
        # Claims and prime
        "a_to_z_claim": "a_to_z_claim",
        "atozclaim": "a_to_z_claim",
        "is_prime": "is_prime",
        "isprime": "is_prime",
        # Product fields
        "asin": "asin",
        "merchant_sku": "merchant_sku",
        "merchantsku": "merchant_sku",
        "sku": "merchant_sku",
        "item_name": "item_name",
        "itemname": "item_name",
        "product_name": "item_name",
        "productname": "item_name",
        "category": "category",
        # Policy and resolution
        "in_policy": "in_policy",
        "inpolicy": "in_policy",
        "resolution": "resolution",
        # Financial fields
        "order_amount": "order_amount",
        "orderamount": "order_amount",
        "order_quantity": "order_quantity",
        "orderquantity": "order_quantity",
        "refunded_amount": "refunded_amount",
        "refundedamount": "refunded_amount",
        # Other fields
        "invoice_number": "invoice_number",
        "invoicenumber": "invoice_number",
        "order_item_id": "order_item_id",
        "orderitemid": "order_item_id",
    }

    return field_mappings.get(normalized, normalized)


def normalize_amazon_sc_fields(record: Dict) -> Dict:
    """
    Normalize Amazon SC returns record to our expected structure
    """
    normalized_record = {}

    for amazon_field, value in record.items():
        # Get the mapped field name
        our_field = get_field_mapping_for_amazon_field(amazon_field)

        # Handle null/empty values
        if pd.isna(value) or value == "" or value == " ":
            normalized_record[our_field] = None
        else:
            # Handle special data type conversions
            if our_field in [
                "order_date",
                "return_request_date",
                "return_delivery_date",
            ]:
                try:
                    normalized_record[our_field] = pd.to_datetime(value).to_pydatetime()
                except:
                    normalized_record[our_field] = value
            elif our_field in ["label_cost"]:
                try:
                    normalized_record[our_field] = float(value)
                except:
                    normalized_record[our_field] = value
            elif our_field in ["return_quantity"]:
                try:
                    normalized_record[our_field] = int(float(value))
                except:
                    normalized_record[our_field] = value
            elif our_field == "tracking_id":
                try:
                    # Handle long tracking IDs
                    normalized_record[our_field] = int(float(value))
                except:
                    normalized_record[our_field] = value
            else:
                normalized_record[our_field] = value

    return normalized_record


def normalize_field_name(field_name: str) -> str:
    """
    Convert Amazon field names to our snake_case format
    """
    if not field_name:
        return ""

    # Remove extra spaces and convert to lowercase
    field_name = field_name.strip().lower()

    # Replace spaces, hyphens, and other separators with underscores
    field_name = re.sub(r"[\s\-\.]+", "_", field_name)

    # Remove special characters but keep underscores
    field_name = re.sub(r"[^a-z0-9_]", "", field_name)

    # Remove multiple consecutive underscores
    field_name = re.sub(r"_+", "_", field_name)

    # Remove leading/trailing underscores
    field_name = field_name.strip("_")

    return field_name


def normalize_amazon_fba_fields(record: Dict) -> Dict:
    """
    Normalize Amazon FBA returns record to our expected structure
    """
    normalized_record = {}

    # FBA field mappings
    fba_field_mappings = {
        "return_date": "return_date",
        "returndate": "return_date",
        "order_id": "amazon_order_id",
        "orderid": "amazon_order_id",
        "amazon_order_id": "amazon_order_id",
        "amazonorderid": "amazon_order_id",
        "sku": "sku_code",
        "sku_code": "sku_code",
        "skucode": "sku_code",
        "asin": "asin",
        "fnsku": "fnsku",
        "product_name": "product_name",
        "productname": "product_name",
        "item_name": "product_name",
        "itemname": "product_name",
        "quantity": "quantity",
        "fulfillment_center_id": "fulfillment_center_id",
        "fulfillmentcenterid": "fulfillment_center_id",
        "detailed_disposition": "detailed_disposition",
        "detaileddisposition": "detailed_disposition",
        "reason": "reason",
        "license_plate_number": "license_plate_number",
        "licenseplatenumber": "license_plate_number",
        "customer_comments": "customer_comments",
        "customercomments": "customer_comments",
    }

    for amazon_field, value in record.items():
        # Get normalized field name
        normalized_field = normalize_field_name(amazon_field)

        # Get our target field name
        our_field = fba_field_mappings.get(normalized_field, normalized_field)

        # Handle null/empty values
        if pd.isna(value) or value == "" or value == " ":
            normalized_record[our_field] = None
        else:
            # Handle special data type conversions
            if our_field == "return_date":
                try:
                    normalized_record[our_field] = pd.to_datetime(value).to_pydatetime()
                except:
                    normalized_record[our_field] = value
            elif our_field == "quantity":
                try:
                    normalized_record[our_field] = int(float(value))
                except:
                    normalized_record[our_field] = value
            else:
                normalized_record[our_field] = value

    return normalized_record


def get_amazon_fba_returns_data(
    start_date: str, end_date: str, db: Any, marketplace_ids: List[str] = None
) -> List[Dict]:
    """
    Fetch FBA returns data from Amazon SP API
    """
    if marketplace_ids is None:
        marketplace_ids = [MARKETPLACE_ID]

    try:
        # Create FBA returns report
        report_id = create_report(
            report_type="GET_FBA_FULFILLMENT_CUSTOMER_RETURNS_DATA",
            start_date=start_date,
            end_date=end_date,
            marketplace_ids=marketplace_ids,
        )

        # Add validation for report_id
        if not report_id:
            raise ValueError(
                "Failed to create FBA returns report - no report ID returned"
            )

        # Wait for report to be ready (with timeout)
        max_wait_time = 300  # 5 minutes
        wait_time = 0

        while wait_time < max_wait_time:
            report_status = get_report_status(report_id)

            # Add validation for report_status
            if not report_status:
                raise ValueError(
                    "Failed to get FBA returns report status - empty response"
                )

            processing_status = report_status.get("processingStatus")
            report_document_id = report_status.get("reportDocumentId")

            if processing_status == "DONE":
                break
            elif processing_status == "FATAL":
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="FBA returns report processing failed",
                )

            time.sleep(10)
            wait_time += 10

        if wait_time >= max_wait_time:
            raise HTTPException(
                status_code=status.HTTP_408_REQUEST_TIMEOUT,
                detail="FBA returns report processing timed out",
            )

        # Validate report_document_id before downloading
        if not report_document_id:
            raise ValueError("No FBA returns report document ID available")

        # Download and parse report data (TSV format)
        report_data = download_report_data(report_document_id, is_gzipped=False)

        # Add validation for report_data
        if not report_data:
            raise ValueError("Downloaded FBA returns report data is empty")

        # Convert TSV data to list of dictionaries
        try:
            df = pd.read_csv(io.StringIO(report_data), sep="\t", low_memory=False)
            fba_returns_raw = df.to_dict("records")
        except Exception as parse_error:
            raise ValueError(f"Failed to parse FBA returns TSV data: {parse_error}")

        # Process and normalize data according to your document structure
        normalized_data = []
        for record in fba_returns_raw:
            # Apply dynamic field mapping
            normalized_record = normalize_amazon_fba_fields(record)

            # Check for duplicate using amazon_order_id + sku_code + return_date
            should_insert = True
            if all(
                key in normalized_record and normalized_record[key] is not None
                for key in ["amazon_order_id", "sku_code", "return_date"]
            ):
                duplicate_check = {
                    "amazon_order_id": normalized_record["amazon_order_id"],
                    "sku_code": normalized_record["sku_code"],
                    "return_date": normalized_record["return_date"],
                }

                existing_record = db[FBA_RETURNS_COLLECTION].find_one(duplicate_check)
                if existing_record:
                    should_insert = False
                    logger.debug(
                        f"Skipping duplicate FBA return: {normalized_record['amazon_order_id']} - {normalized_record['sku_code']}"
                    )

            if should_insert:
                normalized_data.append(normalized_record)

        logger.info(
            f"Processed {len(fba_returns_raw)} FBA returns records, {len(normalized_data)} new records to insert"
        )
        return normalized_data

    except HTTPException:
        # Re-raise HTTPExceptions as-is
        raise
    except ValueError as ve:
        logger.error(f"Validation error in FBA returns data: {ve}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"FBA returns data validation error: {str(ve)}",
        )
    except Exception as e:
        logger.error(f"Unexpected error fetching FBA returns data: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch FBA returns data: {str(e)}",
        )


def get_amazon_sc_returns_data(
    start_date: str, end_date: str, db: Any, marketplace_ids: List[str] = None
) -> List[Dict]:
    """
    Fetch Seller Central returns data from Amazon SP API
    """
    if marketplace_ids is None:
        marketplace_ids = [MARKETPLACE_ID]

    try:
        # Create Seller Central returns report
        report_id = create_report(
            report_type="GET_FLAT_FILE_RETURNS_DATA_BY_RETURN_DATE",
            start_date=start_date,
            end_date=end_date,
            marketplace_ids=marketplace_ids,
        )

        # Add validation for report_id
        if not report_id:
            raise ValueError(
                "Failed to create SC returns report - no report ID returned"
            )

        # Wait for report to be ready (with timeout)
        max_wait_time = 300  # 5 minutes
        wait_time = 0

        while wait_time < max_wait_time:
            report_status = get_report_status(report_id)

            # Add validation for report_status
            if not report_status:
                raise ValueError(
                    "Failed to get SC returns report status - empty response"
                )

            processing_status = report_status.get("processingStatus")
            report_document_id = report_status.get("reportDocumentId")

            if processing_status == "DONE":
                break
            elif processing_status == "FATAL":
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="SC returns report processing failed",
                )

            time.sleep(10)
            wait_time += 10

        if wait_time >= max_wait_time:
            raise HTTPException(
                status_code=status.HTTP_408_REQUEST_TIMEOUT,
                detail="SC returns report processing timed out",
            )

        # Validate report_document_id before downloading
        if not report_document_id:
            raise ValueError("No SC returns report document ID available")

        # Download and parse report data (TSV format)
        report_data = download_report_data(report_document_id, is_gzipped=False)

        # Add validation for report_data
        if not report_data:
            raise ValueError("Downloaded SC returns report data is empty")

        # Convert TSV data to list of dictionaries
        try:
            df = pd.read_csv(io.StringIO(report_data), sep="\t", low_memory=False)
            sc_returns_raw = df.to_dict("records")
        except Exception as parse_error:
            raise ValueError(f"Failed to parse SC returns TSV data: {parse_error}")

        # Process and normalize data according to your document structure
        normalized_data = []
        for record in sc_returns_raw:
            # Map Amazon column names to your document structure
            normalized_record = {}

            # Apply dynamic field mapping
            normalized_record = normalize_amazon_sc_fields(record)

            # Check for duplicate using order_id + merchant_sku + return_request_date
            should_insert = True
            if all(
                key in normalized_record and normalized_record[key] is not None
                for key in ["order_id", "merchant_sku", "return_request_date"]
            ):
                duplicate_check = {
                    "order_id": normalized_record["order_id"],
                    "merchant_sku": normalized_record["merchant_sku"],
                    "return_request_date": normalized_record["return_request_date"],
                }

                existing_record = db[SC_RETURNS_COLLECTION].find_one(duplicate_check)
                if existing_record:
                    should_insert = False
                    logger.debug(
                        f"Skipping duplicate SC return: {normalized_record['order_id']} - {normalized_record['merchant_sku']}"
                    )

            if should_insert:
                normalized_data.append(normalized_record)

        logger.info(
            f"Processed {len(sc_returns_raw)} SC returns records, {len(normalized_data)} new records to insert"
        )
        return normalized_data

    except HTTPException:
        # Re-raise HTTPExceptions as-is
        raise
    except ValueError as ve:
        logger.error(f"Validation error in SC returns data: {ve}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"SC returns data validation error: {str(ve)}",
        )
    except Exception as e:
        logger.error(f"Unexpected error fetching SC returns data: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch SC returns data: {str(e)}",
        )


def normalize_seller_flex_return(record: Dict) -> Dict:
    """
    Flatten an External Fulfillment Returns API record into a MongoDB-friendly document.
    """
    metadata = record.get("returnMetadata") or {}
    shipping = record.get("returnShippingInfo") or {}
    channel = record.get("marketplaceChannelDetails") or {}
    forward_tracking = shipping.get("forwardTrackingInfo") or {}
    reverse_tracking = shipping.get("reverseTrackingInfo") or {}
    marketplace_channel = channel.get("marketplaceChannel") or {}
    invoice_info = metadata.get("invoiceInformation") or {}

    def parse_dt(val):
        if not val:
            return None
        try:
            return datetime.fromisoformat(val.replace("Z", "+00:00")).replace(tzinfo=None)
        except Exception:
            return val

    return {
        "return_id": record.get("id"),
        "merchant_sku": record.get("merchantSku"),
        "number_of_units": record.get("numberOfUnits"),
        "status": record.get("status"),
        "return_type": record.get("returnType"),
        "return_sub_type": record.get("returnSubType"),
        "fulfillment_location_id": record.get("fulfillmentLocationId"),
        "return_location_id": record.get("returnLocationId"),
        "package_delivery_mode": record.get("packageDeliveryMode"),
        "creation_datetime": parse_dt(record.get("creationDateTime")),
        "last_updated_datetime": parse_dt(record.get("lastUpdatedDateTime")),
        # returnMetadata
        "return_reason": metadata.get("returnReason"),
        "rma_id": metadata.get("rmaId"),
        "fulfillment_order_id": metadata.get("fulfillmentOrderId"),
        "invoice_id": invoice_info.get("id"),
        # shipping
        "forward_carrier": forward_tracking.get("carrierName"),
        "forward_tracking_id": forward_tracking.get("trackingId"),
        "reverse_carrier": reverse_tracking.get("carrierName"),
        "reverse_tracking_id": reverse_tracking.get("trackingId"),
        "pickup_datetime": parse_dt(shipping.get("pickupDateTime")),
        "delivery_datetime": parse_dt(shipping.get("deliveryDateTime")),
        # marketplace channel
        "marketplace_name": marketplace_channel.get("marketplaceName"),
        "channel_name": marketplace_channel.get("channelName"),
        "merchant_id": channel.get("merchantId"),
        "shipment_id": channel.get("shipmentId"),
        "customer_order_id": channel.get("customerOrderId"),
        "channel_sku": channel.get("channelSku"),
        "exchange_order_id": channel.get("exchangeOrderId"),
    }


def get_seller_flex_returns_data(
    created_since: str,
    created_until: str,
    db: Any,
    return_location_id: Optional[str] = None,
    status_filter: Optional[str] = None,
) -> List[Dict]:
    """
    Fetch Seller Flex returns via the External Fulfillment Returns API (v2024-09-11).
    Paginates through all pages and deduplicates against existing DB records by return_id.
    """
    collection = db[SELLER_FLEX_RETURNS_COLLECTION]
    existing_ids = set(collection.distinct("return_id"))
    logger.info(f"Found {len(existing_ids)} existing seller flex return IDs in DB")

    all_records: List[Dict] = []
    next_token: Optional[str] = None
    page = 0

    while True:
        page += 1
        params: Dict[str, Any] = {
            "createdSince": created_since,
            "createdUntil": created_until,
            "maxResults": 100,
        }
        if return_location_id:
            params["returnLocationId"] = return_location_id
        if status_filter:
            params["status"] = status_filter
        if next_token:
            params["nextToken"] = next_token

        logger.info(f"Fetching seller flex returns page {page} (nextToken={'yes' if next_token else 'none'})")
        response = make_sp_api_request("/externalFulfillment/2024-09-11/returns", params=params)

        items = response.get("returns") or response.get("items") or []
        logger.info(f"Page {page}: received {len(items)} records")

        for item in items:
            normalized = normalize_seller_flex_return(item)
            return_id = normalized.get("return_id")
            if return_id and return_id in existing_ids:
                logger.debug(f"Skipping duplicate seller flex return: {return_id}")
                continue
            if return_id:
                existing_ids.add(return_id)
            normalized["created_at"] = datetime.now()
            all_records.append(normalized)

        next_token = response.get("nextToken")
        if not next_token:
            break

    logger.info(f"Total new seller flex return records to insert: {len(all_records)}")
    return all_records


# API Endpoints


@router.post("/auth/token")
def refresh_amazon_token():
    """
    Refresh Amazon SP API token
    """
    try:
        token = get_amazon_token()
        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "message": "Token refreshed successfully",
                "token_expires_at": (
                    _sc_token_expires_at.isoformat() if _sc_token_expires_at else None
                ),
            },
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e)
        )


@router.post("/sync/sales-traffic")
async def sync_sales_traffic_data(
    start_date: str,
    end_date: str,
    marketplace_ids: Optional[List[str]] = None,
    db=Depends(get_database),
):
    """
    Fetch ASIN-wise daily sales and traffic data from Amazon and insert into database
    Format: YYYY-MM-DD
    """
    try:
        # Validate date format
        datetime.strptime(start_date, "%Y-%m-%d")
        datetime.strptime(end_date, "%Y-%m-%d")

        # Convert to ISO format for API
        start_datetime = f"{start_date}T00:00:00Z"
        end_datetime = f"{end_date}T23:59:59Z"

        # Fetch sales and traffic data
        sales_traffic_data = get_amazon_sales_traffic_data(
            start_datetime, end_datetime, db, marketplace_ids
        )
        # Insert into database
        inserted_count = await insert_data_to_db(
            SALES_COLLECTION, sales_traffic_data, db
        )

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "message": "Sales and traffic data synced successfully",
                "records_inserted": inserted_count,
                "date_range": f"{start_date} to {end_date}",
                "marketplace_ids": marketplace_ids or [MARKETPLACE_ID],
            },
        )

    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD",
        )
    except Exception as e:
        logger.error(f"Error syncing sales and traffic data: {e}")
        raise


@router.post("/sync/ledger")
async def sync_ledger_data(
    start_date: str,
    end_date: str,
    marketplace_ids: Optional[List[str]] = None,
    aggregate_by_location: str = "FC",
    aggregated_by_time_period: str = "DAILY",
    db=Depends(get_database),
):
    """
    Fetch ledger summary data from Amazon and insert into database
    Format: YYYY-MM-DD
    """
    try:
        # Validate date format
        datetime.strptime(start_date, "%Y-%m-%d")
        datetime.strptime(end_date, "%Y-%m-%d")

        # Convert to ISO format for API
        start_datetime = f"{start_date}T00:00:00Z"
        end_datetime = f"{end_date}T00:00:00Z"
        # Fetch ledger data
        ledger_data = get_amazon_ledger_data(
            start_datetime,
            end_datetime,
            db,
            marketplace_ids,
            aggregate_by_location,
            aggregated_by_time_period,
        )

        # Insert into database
        inserted_count = await insert_data_to_db(INVENTORY_COLLECTION, ledger_data, db)

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "message": "Ledger data synced successfully",
                "records_inserted": inserted_count,
                "date_range": f"{start_date} to {end_date}",
                "marketplace_ids": marketplace_ids or [MARKETPLACE_ID],
                "aggregate_by_location": aggregate_by_location,
                "aggregated_by_time_period": aggregated_by_time_period,
            },
        )

    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD",
        )
    except Exception as e:
        logger.error(f"Error syncing ledger data: {e}")
        raise


@router.get("/status")
async def sync_status(
    start_date: str, end_date: str, report_type: str, db=Depends(get_database)
):
    try:
        # Parse date range
        start = datetime.strptime(start_date, "%Y-%m-%d").replace(
            hour=0, minute=0, second=0, microsecond=0, tzinfo=timezone.utc
        )
        end = datetime.strptime(end_date, "%Y-%m-%d").replace(
            hour=23, minute=59, second=59, microsecond=999999, tzinfo=timezone.utc
        )

        async def get_data_status(sales_collection, inventory_collection):
            """Helper function to get status for a collection pair"""
            # Get first and last sales dates in the date range
            sales_date_pipeline = [
                {"$match": {"date": {"$gte": start, "$lte": end}}},
                {
                    "$group": {
                        "_id": None,
                        "first_date": {"$min": "$date"},
                        "last_date": {"$max": "$date"},
                    }
                },
            ]

            sales_dates = list(db[sales_collection].aggregate(sales_date_pipeline))
            first_sales_date = None
            last_sales_date = None
            if sales_dates:
                first_sales_date = (
                    sales_dates[0]["first_date"].isoformat()
                    if sales_dates[0]["first_date"]
                    else None
                )
                last_sales_date = (
                    sales_dates[0]["last_date"].isoformat()
                    if sales_dates[0]["last_date"]
                    else None
                )

            # Get first and last inventory dates in the date range
            inventory_date_pipeline = [
                {"$match": {"date": {"$gte": start, "$lte": end}}},
                {
                    "$group": {
                        "_id": None,
                        "first_date": {"$min": "$date"},
                        "last_date": {"$max": "$date"},
                    }
                },
            ]

            inventory_dates = list(
                db[inventory_collection].aggregate(inventory_date_pipeline)
            )
            first_inventory_date = None
            last_inventory_date = None
            if inventory_dates:
                first_inventory_date = (
                    inventory_dates[0]["first_date"].isoformat()
                    if inventory_dates[0]["first_date"]
                    else None
                )
                last_inventory_date = (
                    inventory_dates[0]["last_date"].isoformat()
                    if inventory_dates[0]["last_date"]
                    else None
                )

            # Get total counts
            sales_count = db[sales_collection].count_documents(
                {"date": {"$gte": start, "$lte": end}}
            )
            inventory_count = db[inventory_collection].count_documents(
                {"date": {"$gte": start, "$lte": end}}
            )

            # Get distinct dates for sales
            sales_distinct_dates = list(db[sales_collection].distinct(
                "date", {"date": {"$gte": start, "$lte": end}}
            ))
            sales_dates_set = set(d.strftime("%Y-%m-%d") if isinstance(d, datetime) else d for d in sales_distinct_dates)

            # Get distinct dates for inventory
            inventory_distinct_dates = list(db[inventory_collection].distinct(
                "date", {"date": {"$gte": start, "$lte": end}}
            ))
            inventory_dates_set = set(d.strftime("%Y-%m-%d") if isinstance(d, datetime) else d for d in inventory_distinct_dates)

            # Calculate all dates in the range
            from datetime import timedelta
            all_dates = set()
            current_date = start.replace(hour=0, minute=0, second=0, microsecond=0)
            end_check = end.replace(hour=0, minute=0, second=0, microsecond=0)
            while current_date <= end_check:
                all_dates.add(current_date.strftime("%Y-%m-%d"))
                current_date += timedelta(days=1)

            # Find missing dates
            missing_sales_dates = sorted(all_dates - sales_dates_set)
            missing_inventory_dates = sorted(all_dates - inventory_dates_set)

            return {
                "sales_data": {
                    "records_count": sales_count,
                    "first_sales_date": first_sales_date,
                    "last_sales_date": last_sales_date,
                    "missing_dates": missing_sales_dates,
                },
                "inventory_data": {
                    "records_count": inventory_count,
                    "first_inventory_date": first_inventory_date,
                    "last_inventory_date": last_inventory_date,
                    "missing_dates": missing_inventory_dates,
                },
            }

        # Handle different report types
        if report_type == "all":
            # Get both vendor_central and fba/seller_flex data
            vendor_status = await get_data_status(
                "amazon_vendor_sales", "amazon_vendor_inventory"
            )
            fba_status = await get_data_status(SALES_COLLECTION, INVENTORY_COLLECTION)

            return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={
                    "date_range": {"start_date": start_date, "end_date": end_date},
                    "vendor_central": vendor_status,
                    "fba_seller_flex": fba_status,
                    "report_type": report_type,
                },
            )

        elif report_type == "vendor_central":
            status_data = await get_data_status(
                "amazon_vendor_sales", "amazon_vendor_inventory"
            )
        else:
            # fba, seller_flex, fba+seller_flex
            status_data = await get_data_status(SALES_COLLECTION, INVENTORY_COLLECTION)

        # For single report types, return the original structure
        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "date_range": {"start_date": start_date, "end_date": end_date},
                "sales_data": status_data["sales_data"],
                "inventory_data": status_data["inventory_data"],
                "report_type": report_type,
            },
        )

    except Exception as e:
        logger.error(f"Error getting sync status: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e)
        )


@router.post("/sync/fba-returns")
async def sync_fba_returns_data(
    start_date: str,
    end_date: str,
    marketplace_ids: Optional[List[str]] = None,
    db=Depends(get_database),
):
    """
    Fetch FBA returns data from Amazon and insert into database
    Format: YYYY-MM-DD
    """
    try:
        # Validate date format
        datetime.strptime(start_date, "%Y-%m-%d")
        datetime.strptime(end_date, "%Y-%m-%d")

        # Convert to ISO format for API
        start_datetime = f"{start_date}T00:00:00Z"
        end_datetime = f"{end_date}T23:59:59Z"

        # Fetch FBA returns data
        fba_returns_data = get_amazon_fba_returns_data(
            start_datetime, end_datetime, db, marketplace_ids
        )

        # Insert into database
        inserted_count = await insert_data_to_db(
            FBA_RETURNS_COLLECTION, fba_returns_data, db
        )

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "message": "FBA returns data synced successfully",
                "records_inserted": inserted_count,
                "date_range": f"{start_date} to {end_date}",
                "marketplace_ids": marketplace_ids or [MARKETPLACE_ID],
            },
        )

    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD",
        )
    except Exception as e:
        logger.error(f"Error syncing FBA returns data: {e}")
        raise


@router.post("/sync/sc-returns")
async def sync_sc_returns_data(
    start_date: str,
    end_date: str,
    marketplace_ids: Optional[List[str]] = None,
    db=Depends(get_database),
):
    """
    Fetch Seller Central returns data from Amazon and insert into database
    Format: YYYY-MM-DD
    """
    try:
        # Validate date format
        datetime.strptime(start_date, "%Y-%m-%d")
        datetime.strptime(end_date, "%Y-%m-%d")

        # Convert to ISO format for API
        start_datetime = f"{start_date}T00:00:00Z"
        end_datetime = f"{end_date}T23:59:59Z"

        # Fetch SC returns data
        sc_returns_data = get_amazon_sc_returns_data(
            start_datetime, end_datetime, db, marketplace_ids
        )

        # Insert into database
        inserted_count = await insert_data_to_db(
            SC_RETURNS_COLLECTION, sc_returns_data, db
        )

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "message": "Seller Central returns data synced successfully",
                "records_inserted": inserted_count,
                "date_range": f"{start_date} to {end_date}",
                "marketplace_ids": marketplace_ids or [MARKETPLACE_ID],
            },
        )

    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD",
        )
    except Exception as e:
        logger.error(f"Error syncing SC returns data: {e}")
        raise


@router.post("/sync/daily-returns")
async def sync_daily_returns_data(
    db=Depends(get_database),
):
    """
    Fetch both FBA and SC returns data for a specific date (default: yesterday)
    This endpoint is designed for daily cron jobs
    Format: YYYY-MM-DD
    """
    try:
        start_date = datetime.now().replace(day=1).strftime("%Y-%m-%d")
        end_date = datetime.now().strftime("%Y-%m-%d")

        # Use the same date for start and end to get that specific day's data
        start_datetime = f"{start_date}T00:00:00Z"
        end_datetime = f"{end_date}T23:59:59Z"

        # Fetch both types of returns data
        fba_returns_data = get_amazon_fba_returns_data(start_datetime, end_datetime, db)

        sc_returns_data = get_amazon_sc_returns_data(start_datetime, end_datetime, db)

        # Insert into respective databases
        fba_inserted_count = await insert_data_to_db(
            FBA_RETURNS_COLLECTION, fba_returns_data, db
        )

        sc_inserted_count = await insert_data_to_db(
            SC_RETURNS_COLLECTION, sc_returns_data, db
        )

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "message": f"Daily returns data synced successfully for {start_date}",
                "fba_returns": {
                    "records_inserted": fba_inserted_count,
                },
                "sc_returns": {
                    "records_inserted": sc_inserted_count,
                },
                "start_date": start_date,
                "end_date": end_date,
                "marketplace_ids": [MARKETPLACE_ID],
            },
        )

    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD",
        )
    except Exception as e:
        logger.error(f"Error syncing daily returns data: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to sync daily returns data: {str(e)}",
        )


@router.post("/sync/monthly-returns")
async def sync_monthly_returns_data(
    year: int,
    month: int,
    marketplace_ids: Optional[List[str]] = None,
    db=Depends(get_database),
):
    """
    Fetch both FBA and SC returns data for the entire month till current date
    """
    try:
        # Calculate date range for the month
        from calendar import monthrange

        # Get first day of the month
        start_date = datetime(year, month, 1)

        # Get current date or last day of month, whichever is earlier
        current_date = datetime.now()
        last_day_of_month = monthrange(year, month)[1]

        if year == current_date.year and month == current_date.month:
            # Current month - sync till today
            end_date = current_date
        else:
            # Past month - sync entire month
            end_date = datetime(year, month, last_day_of_month)

        # Format dates for API
        start_date_str = start_date.strftime("%Y-%m-%d")
        end_date_str = end_date.strftime("%Y-%m-%d")
        start_datetime = start_date.strftime("%Y-%m-%dT00:00:00Z")
        end_datetime = end_date.strftime("%Y-%m-%dT23:59:59Z")

        # Fetch both types of returns data
        fba_returns_data = get_amazon_fba_returns_data(
            start_datetime, end_datetime, db, marketplace_ids
        )

        sc_returns_data = get_amazon_sc_returns_data(
            start_datetime, end_datetime, db, marketplace_ids
        )

        # Insert into respective databases
        fba_inserted_count = await insert_data_to_db(
            FBA_RETURNS_COLLECTION, fba_returns_data, db
        )

        sc_inserted_count = await insert_data_to_db(
            SC_RETURNS_COLLECTION, sc_returns_data, db
        )

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "message": "Monthly returns data synced successfully",
                "fba_returns": {
                    "records_inserted": fba_inserted_count,
                },
                "sc_returns": {
                    "records_inserted": sc_inserted_count,
                },
                "date_range": f"{start_date_str} to {end_date_str}",
                "month": f"{year}-{month:02d}",
                "marketplace_ids": marketplace_ids or [MARKETPLACE_ID],
            },
        )

    except Exception as e:
        logger.error(f"Error syncing monthly returns data: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to sync monthly returns data: {str(e)}",
        )


@router.post("/sync/seller-flex-returns")
async def sync_seller_flex_returns(
    start_date: str,
    end_date: str,
    return_location_id: Optional[str] = None,
    status_filter: Optional[str] = None,
    marketplace_ids: Optional[List[str]] = None,
    db=Depends(get_database),
):
    """
    Fetch Seller Flex returns via the External Fulfillment Returns API and store in DB.
    Uses createdSince/createdUntil filter. Format: YYYY-MM-DD.

    Optional filters:
    - return_location_id: SmartConnect location ID
    - status_filter: CREATED | CARRIER_NOTIFIED_TO_PICK_UP_FROM_CUSTOMER | DELIVERED | PROCESSED
    """
    try:
        datetime.strptime(start_date, "%Y-%m-%d")
        datetime.strptime(end_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid date format. Use YYYY-MM-DD")

    created_since = f"{start_date}T00:00:00Z"
    created_until = f"{end_date}T23:59:59Z"

    try:
        records = get_seller_flex_returns_data(
            created_since=created_since,
            created_until=created_until,
            db=db,
            return_location_id=return_location_id,
            status_filter=status_filter,
        )

        inserted_count = await insert_data_to_db(SELLER_FLEX_RETURNS_COLLECTION, records, db)

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "message": "Seller Flex returns synced successfully",
                "records_inserted": inserted_count,
                "date_range": f"{start_date} to {end_date}",
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error syncing Seller Flex returns: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to sync Seller Flex returns: {str(e)}",
        )


@router.get("/sales-traffic/{asin}")
def get_asin_sales_data(
    asin: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db=Depends(get_database),
):
    """
    Get sales and traffic data for a specific ASIN
    """
    try:
        query = {"$or": [{"parentAsin": asin}, {"childAsin": asin}]}

        # Add date range filter if provided
        if start_date and end_date:
            query["date"] = {"$gte": start_date, "$lte": end_date}

        collection = db[SALES_COLLECTION]
        results = list(collection.find(query).sort("date", 1))

        # Serialize MongoDB documents
        serialized_results = [serialize_mongo_document(doc) for doc in results]

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "asin": asin,
                "records_found": len(serialized_results),
                "data": serialized_results,
            },
        )

    except Exception as e:
        logger.error(f"Error fetching ASIN data: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e)
        )


@router.get("/ledger/summary")
def get_ledger_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    transaction_type: Optional[str] = None,
    db=Depends(get_database),
):
    """
    Get ledger summary data with optional filters
    """
    try:
        query = {}

        # Add date range filter if provided
        if start_date and end_date:
            query["Date"] = {"$gte": start_date, "$lte": end_date}

        # Add transaction type filter if provided
        if transaction_type:
            query["Transaction Type"] = transaction_type

        collection = db[INVENTORY_COLLECTION]
        results = list(collection.find(query).sort("Date", 1))

        # Serialize MongoDB documents
        serialized_results = [serialize_mongo_document(doc) for doc in results]

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "records_found": len(serialized_results),
                "filters": {
                    "start_date": start_date,
                    "end_date": end_date,
                    "transaction_type": transaction_type,
                },
                "data": serialized_results,
            },
        )

    except Exception as e:
        logger.error(f"Error fetching ledger summary: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e)
        )


@router.get("/get_amazon_sku_mapping")
async def get_sku_mapping(database=Depends(get_database)):
    try:
        def _fetch():
            sku_collection = database.get_collection(SKU_COLLECTION)
            docs = list(sku_collection.find({}).sort("item_name", -1))
            if not docs:
                return []

            # Batch-lookup the most recent fnsku for each ASIN from amazon_ledger
            asins = [d["item_id"] for d in docs if d.get("item_id")]
            ledger_collection = database.get_collection("amazon_ledger")
            pipeline = [
                {
                    "$match": {
                        "asin": {"$in": asins},
                        "fnsku": {"$exists": True, "$nin": [None, ""]},
                    }
                },
                {"$sort": {"date": -1}},
                {"$group": {"_id": "$asin", "fnsku": {"$first": "$fnsku"}}},
            ]
            fnsku_map = {
                d["_id"]: d["fnsku"]
                for d in ledger_collection.aggregate(pipeline)
            }

            for doc in docs:
                ledger_fnsku = fnsku_map.get(doc.get("item_id"))
                if ledger_fnsku:
                    doc["fnsku"] = ledger_fnsku
                # else keep whatever is already stored in the mapping doc

            return docs

        docs = await asyncio.to_thread(_fetch)
        return JSONResponse(content=serialize_mongo_document(docs))

    except PyMongoError as e:  # Catch database-specific errors
        logger.info(f"Database error retrieving SKU mapping: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database error retrieving SKU mapping: {e}",
        )
    except Exception as e:  # Catch any other unexpected errors
        logger.info(f"Unexpected error retrieving SKU mapping: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An unexpected error occurred: {e}",
        )


_VALID_AMAZON_STATUSES = {"Active", "Inactive", "Discontinued on Amazon"}


@router.put("/sku-mapping/{asin}/status")
async def update_sku_amazon_status(
    asin: str,
    amazon_status: str = Query(...),
    database=Depends(get_database),
):
    """Update the amazon_status field on an amazon_sku_mapping document."""
    if amazon_status not in _VALID_AMAZON_STATUSES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid status. Must be one of: {', '.join(_VALID_AMAZON_STATUSES)}",
        )

    def _update():
        return database.get_collection(SKU_COLLECTION).update_one(
            {"item_id": asin},
            {"$set": {"amazon_status": amazon_status}},
        )

    result = await asyncio.to_thread(_update)
    if result.matched_count == 0:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="ASIN not found")
    return {"message": "Status updated"}


@router.post("/upload-sku-mapping")
async def upload_sku_mapping(
    file: UploadFile = File(...), database=Depends(get_database)
):
    """
    Uploads SKU mapping data from an Excel file (.xlsx).
    Clears existing mapping and inserts new data.
    Expected columns: 'ASIN', 'SKU', 'Item Name'.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file type. Please upload an Excel file (.xlsx or .xls).",
        )

    try:
        file_content = await file.read()
        df = pd.read_excel(
            BytesIO(file_content), sheet_name="Sheet1"
        )

        # Validate required columns
        required_cols = ["ASIN", "SKU", "Item Name"]
        if not all(col in df.columns for col in required_cols):
            missing = [col for col in required_cols if col not in df.columns]
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Missing required columns: {', '.join(missing)}",
            )

        sku_collection = database.get_collection(SKU_COLLECTION)

        # Optional: Clear existing mapping or implement upsert logic
        # Clearing is simpler if the Excel is the source of truth for the full mapping
        delete_result = sku_collection.delete_many({})
        logger.info(f"Deleted {delete_result.deleted_count} existing SKU mappings.")

        data_by_item_id = {}
        for _, row in df.iterrows():
            item_id = str(row["ASIN"]).strip() if pd.notna(row["ASIN"]) else None
            sku_code = str(row["SKU"]).strip() if pd.notna(row["SKU"]) else None
            item_name = (
                str(row["Item Name"]).strip() if pd.notna(row["Item Name"]) else None
            )

            if (
                item_id and sku_code and item_name
            ):  # Ensure essential fields are present
                data_by_item_id[item_id] = {
                    "item_id": item_id,
                    "sku_code": sku_code,
                    "item_name": item_name,
                }
            else:
                logger.info(f"Skipping SKU row due to missing data: {row.to_dict()}")

        data_to_insert = list(data_by_item_id.values())

        if not data_to_insert:
            return {"message": "No valid SKU mapping data found to insert."}

        insert_result = sku_collection.insert_many(data_to_insert)
        logger.info(f"Inserted {len(insert_result.inserted_ids)} new SKU mappings.")

        # Create index for efficient lookup
        sku_collection.create_index([("item_id", ASCENDING)], unique=True)

        return {
            "message": f"Successfully uploaded and stored {len(insert_result.inserted_ids)} SKU mappings."
        }

    except HTTPException as e:
        raise e  # Re-raise intended HTTP exceptions
    except Exception as e:
        logger.info(f"Error uploading SKU mapping: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An error occurred processing the file: {e}",
        )


MARGINS_COLLECTION = "vendor_margins"


@router.put("/sku-mapping/{asin}/fnsku")
async def update_sku_fnsku(
    asin: str,
    fnsku: str = Query(...),
    database=Depends(get_database),
):
    """Update the fnsku field on an amazon_sku_mapping document."""
    def _update():
        return database.get_collection(SKU_COLLECTION).update_one(
            {"item_id": asin},
            {"$set": {"fnsku": fnsku.strip()}},
        )

    result = await asyncio.to_thread(_update)
    if result.matched_count == 0:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="ASIN not found")
    return {"message": "FNSKU updated"}


@router.post("/upload-etrade-margins")
async def upload_etrade_margins(
    file: UploadFile = File(...), database=Depends(get_database)
):
    """
    Upload Amazon item data from an Excel file.
    Handles: FNSKU, Amazon Status (→ amazon_sku_mapping)
             ASP, New Margin, Cost Price w/o Tax, Etrade PO, Etrade DF (→ vendor_margins)
    ASIN column is required. All other columns are optional — blank cells are skipped.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file type. Please upload an Excel file (.xlsx or .xls).",
        )

    try:
        file_content = await file.read()
        df = pd.read_excel(BytesIO(file_content), sheet_name=0)

        if "ASIN" not in df.columns:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Missing required column: ASIN",
            )

        def _parse_bool(val) -> bool | None:
            if not pd.notna(val):
                return None
            s = str(val).strip().lower()
            if s in ("true", "yes", "y", "1"):
                return True
            if s in ("false", "no", "n", "0"):
                return False
            return None

        margins_collection = database.get_collection(MARGINS_COLLECTION)
        sku_collection = database.get_collection(SKU_COLLECTION)
        upserted = 0
        skipped = 0

        for _, row in df.iterrows():
            asin = str(row["ASIN"]).strip() if pd.notna(row["ASIN"]) else None
            if not asin:
                skipped += 1
                continue

            # ── vendor_margins fields ──────────────────────────────────────────
            margin_fields: dict = {}
            if "ASP" in df.columns and pd.notna(row["ASP"]):
                margin_fields["etrade_asp"] = float(row["ASP"])
            if "New Margin" in df.columns and pd.notna(row["New Margin"]):
                margin_fields["margin"] = float(row["New Margin"])
            if "Cost Price w/o Tax" in df.columns and pd.notna(row["Cost Price w/o Tax"]):
                margin_fields["cost_price_wo_tax"] = float(row["Cost Price w/o Tax"])
            if "Etrade PO" in df.columns:
                v = _parse_bool(row["Etrade PO"])
                if v is not None:
                    margin_fields["etrade_po"] = v
            if "Etrade DF" in df.columns:
                v = _parse_bool(row["Etrade DF"])
                if v is not None:
                    margin_fields["etrade_df"] = v

            if margin_fields:
                margins_collection.update_one(
                    {"asin": asin},
                    {"$set": {"asin": asin, **margin_fields}},
                    upsert=True,
                )

            # ── amazon_sku_mapping fields ──────────────────────────────────────
            sku_fields: dict = {}
            if "FNSKU" in df.columns and pd.notna(row["FNSKU"]):
                fnsku_val = str(row["FNSKU"]).strip()
                if fnsku_val:
                    sku_fields["fnsku"] = fnsku_val
            if "Amazon Status" in df.columns and pd.notna(row["Amazon Status"]):
                status_val = str(row["Amazon Status"]).strip()
                if status_val in _VALID_AMAZON_STATUSES:
                    sku_fields["amazon_status"] = status_val

            if sku_fields:
                sku_collection.update_one(
                    {"item_id": asin},
                    {"$set": sku_fields},
                )

            if margin_fields or sku_fields:
                upserted += 1
            else:
                skipped += 1

        return {
            "message": f"Successfully updated {upserted} records.",
            "upserted": upserted,
            "skipped": skipped,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.info(f"Error uploading amazon item data: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An error occurred processing the file: {e}",
        )


@router.get("/download-etrade-margins-template")
def download_etrade_margins_template(database=Depends(get_database)):
    """
    Download an Excel template pre-filled with existing item data.
    Columns: ASIN, Item Name, SKU Code, FNSKU, Amazon Status,
             ASP, New Margin, Cost Price w/o Tax, Etrade PO, Etrade DF.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        sku_collection = database.get_collection(SKU_COLLECTION)
        margins_collection = database.get_collection(MARGINS_COLLECTION)

        skus = list(sku_collection.find(
            {},
            {"item_id": 1, "item_name": 1, "sku_code": 1, "fnsku": 1, "amazon_status": 1, "_id": 0},
        ))
        margins_list = list(margins_collection.find(
            {},
            {"asin": 1, "etrade_asp": 1, "margin": 1, "cost_price_wo_tax": 1, "etrade_po": 1, "etrade_df": 1, "_id": 0},
        ))
        margins_by_asin = {m["asin"]: m for m in margins_list}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Amazon Items"

        headers = [
            "ASIN", "Item Name", "SKU Code", "FNSKU", "Amazon Status",
            "ASP", "New Margin", "Cost Price w/o Tax", "Etrade PO", "Etrade DF",
        ]
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_font = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, sku in enumerate(skus, 2):
            asin = sku.get("item_id", "")
            m = margins_by_asin.get(asin, {})
            ws.cell(row=row_idx, column=1,  value=asin)
            ws.cell(row=row_idx, column=2,  value=sku.get("item_name", ""))
            ws.cell(row=row_idx, column=3,  value=sku.get("sku_code", ""))
            ws.cell(row=row_idx, column=4,  value=sku.get("fnsku") or "")
            ws.cell(row=row_idx, column=5,  value=sku.get("amazon_status") or "")
            ws.cell(row=row_idx, column=6,  value=m.get("etrade_asp"))
            ws.cell(row=row_idx, column=7,  value=m.get("margin"))
            ws.cell(row=row_idx, column=8,  value=m.get("cost_price_wo_tax"))
            ws.cell(row=row_idx, column=9,  value="Yes" if m.get("etrade_po") else ("No" if "etrade_po" in m else ""))
            ws.cell(row=row_idx, column=10, value="Yes" if m.get("etrade_df") else ("No" if "etrade_df" in m else ""))

        col_widths = [16, 45, 18, 14, 26, 10, 12, 22, 10, 10]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=amazon_items_template.xlsx"},
        )
    except Exception as e:
        logger.info(f"Error generating amazon items template: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An error occurred generating the template: {e}",
        )


SHEET_ID      = "1tn_Lj3KR0zXY8B-8ZUkSznZgE4YzyjtAkcpdHzBCgt4"
SHEET_CSV_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv"

# Google Sheet column names (row 2 is the real header)
GS_COL_ASIN  = "Amazon ASIN"
GS_COL_SKU   = "SKU Code (Final)"
GS_COL_NAME  = "Amazon Final Title"


def _fetch_sheet_rows() -> dict:
    """Return {asin: {sku_code, sheet_name}} from the master Google Sheet."""
    import csv as _csv, io as _io
    resp = requests.get(SHEET_CSV_URL, timeout=30)
    resp.raise_for_status()
    lines = resp.text.splitlines()
    if len(lines) < 2:
        raise ValueError("Google Sheet has fewer than 2 rows")
    reader = _csv.DictReader(_io.StringIO("\n".join(lines[1:])))
    result = {}
    for row in reader:
        asin     = (row.get(GS_COL_ASIN) or "").strip()
        sku_code = (row.get(GS_COL_SKU) or "").strip()
        name     = (row.get(GS_COL_NAME) or "").strip()
        if asin and sku_code:
            result[asin] = {"sku_code": sku_code, "sheet_name": name}
    return result


def _fetch_sp_listings() -> list:
    """Fetch all active listings via GET_MERCHANT_LISTINGS_ALL_DATA report."""
    import csv as _csv, io as _io

    # 1. Request report
    report_resp = make_sp_api_request(
        "/reports/2021-06-30/reports",
        method="POST",
        data={"reportType": "GET_MERCHANT_LISTINGS_ALL_DATA", "marketplaceIds": [MARKETPLACE_ID]},
    )
    report_id = report_resp.get("reportId")
    if not report_id:
        raise ValueError(f"No reportId returned: {report_resp}")

    # 2. Poll
    max_wait, elapsed = 600, 0
    doc_id = None
    while elapsed < max_wait:
        status_resp = make_sp_api_request(f"/reports/2021-06-30/reports/{report_id}")
        ps = status_resp.get("processingStatus", "")
        if ps == "DONE":
            doc_id = status_resp.get("reportDocumentId")
            break
        elif ps in ("FATAL", "CANCELLED"):
            raise ValueError(f"Report ended with status {ps}")
        time.sleep(15)
        elapsed += 15
    if not doc_id:
        raise TimeoutError("Listings report timed out")

    # 3. Download
    doc_info = make_sp_api_request(f"/reports/2021-06-30/documents/{doc_id}")
    dl_url = doc_info.get("url")
    if not dl_url:
        raise ValueError("No download URL in report document")
    raw_resp = requests.get(dl_url, timeout=120)
    raw_resp.raise_for_status()
    raw = raw_resp.content
    if doc_info.get("compressionAlgorithm") == "GZIP":
        raw = gzip.decompress(raw)
    text = None
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = raw.decode("latin-1")

    # 4. Parse TSV
    reader = _csv.DictReader(_io.StringIO(text), delimiter="\t")
    listings = []
    for row in reader:
        asin = (row.get("asin1") or "").strip()
        if asin:
            listings.append({
                "asin":       asin,
                "seller_sku": (row.get("seller-sku") or "").strip(),
                "item_name":  (row.get("item-name") or "").strip(),
            })
    return listings


@router.post("/sync-sku-mapping")
async def sync_sku_mapping(database=Depends(get_database)):
    """
    Syncs amazon_sku_mapping from SP-API + Google Sheet.
    SP-API provides item_name; Google Sheet provides sku_code.
    Also backfills fnsku from amazon_ledger for docs that don't have one yet.
    """
    try:
        sp_listings, sheet_by_asin = await asyncio.gather(
            asyncio.to_thread(_fetch_sp_listings),
            asyncio.to_thread(_fetch_sheet_rows),
        )

        collection = database.get_collection(SKU_COLLECTION)

        def _upsert():
            collection.create_index([("item_id", ASCENDING)], unique=True)
            inserted = modified = unchanged = 0
            for listing in sp_listings:
                sheet = sheet_by_asin.get(listing["asin"])
                # Prefer sheet sku_code; fall back to seller_sku from SP-API
                sku_code = sheet["sku_code"] if sheet else listing["seller_sku"]
                set_fields = {
                    "item_id":    listing["asin"],
                    "seller_sku": listing["seller_sku"],
                    "sku_code":   sku_code,
                }
                # Only overwrite item_name if SP-API returned a non-empty value;
                # otherwise preserve whatever is already stored in the DB.
                item_name = listing["item_name"] or (sheet and sheet.get("sheet_name")) or ""
                if item_name:
                    set_fields["item_name"] = item_name
                result = collection.update_one(
                    {"item_id": listing["asin"]},
                    {"$set": set_fields},
                    upsert=True,
                )
                if result.upserted_id:
                    inserted += 1
                elif result.modified_count:
                    modified += 1
                else:
                    unchanged += 1
            return inserted, modified, unchanged

        def _backfill_fnsku():
            """
            Build ASIN→FNSKU map from amazon_ledger, then update amazon_sku_mapping docs
            that are missing fnsku (preserving any manually-set values).
            Returns count of docs updated.
            """
            ledger_col = database.get_collection(INVENTORY_COLLECTION)
            pipeline = [
                {"$match": {"fnsku": {"$exists": True, "$nin": [None, ""]}}},
                {"$group": {"_id": "$asin", "fnsku": {"$first": "$fnsku"}}},
            ]
            fnsku_map = {
                doc["_id"]: doc["fnsku"]
                for doc in ledger_col.aggregate(pipeline)
                if doc.get("_id") and doc.get("fnsku")
            }
            if not fnsku_map:
                return 0
            # Update only docs where fnsku is not yet set
            updated = 0
            for asin, fnsku in fnsku_map.items():
                result = collection.update_one(
                    {
                        "item_id": asin,
                        "$or": [{"fnsku": {"$exists": False}}, {"fnsku": None}, {"fnsku": ""}],
                    },
                    {"$set": {"fnsku": fnsku}},
                )
                if result.modified_count:
                    updated += 1
            return updated

        inserted, modified, unchanged = await asyncio.to_thread(_upsert)
        fnsku_backfilled = await asyncio.to_thread(_backfill_fnsku)

        return {
            "message": "Sync complete",
            "inserted": inserted,
            "updated": modified,
            "unchanged": unchanged,
            "total_sp_listings": len(sp_listings),
            "fnsku_backfilled": fnsku_backfilled,
        }

    except Exception as e:
        logger.error(f"sync-sku-mapping error: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e),
        )


@router.post("/create_single_item")
def create_single_item(body: dict):
    try:
        item_name = body.get("item_name")
        item_id = body.get("item_id")
        sku_code = body.get("sku_code")
        db = get_database()
        result = db[SKU_COLLECTION].find_one(
            {"item_name": item_name, "item_id": item_id}
        )
        if result:
            raise HTTPException(
                status_code=400, detail="Item Name and Item Id already Exists"
            )
        else:
            db[SKU_COLLECTION].insert_one(
                {"item_name": item_name, "item_id": str(item_id), "sku_code": sku_code}
            )
            return "Item Created"
    except HTTPException as e:
        raise e
    except Exception as e:
        logger.info(f"Error uploading sales data: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An error occurred creating the item: {e}",
        )


@router.delete("/delete_item/{item_id}")
def delete_item(item_id: str):
    try:
        db = get_database()
        result = db[SKU_COLLECTION].delete_one({"_id": ObjectId(item_id)})
        if result.deleted_count == 1:
            return JSONResponse(status_code=200, content="Item Deleted Successfully")

    except HTTPException as e:
        raise e
    except Exception as e:
        logger.info(f"Error uploading sales data: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An error occurred processing the file: {e}",
        )


def _parse_daily_date(date_val):
    """Parse a date from an ISO string or datetime to a date object (for DRR)."""
    from datetime import date as date_type
    if isinstance(date_val, date_type) and not isinstance(date_val, datetime):
        return date_val
    if isinstance(date_val, datetime):
        return date_val.date()
    if isinstance(date_val, str):
        try:
            return datetime.fromisoformat(date_val.replace("Z", "+00:00")).date()
        except ValueError:
            return datetime.strptime(date_val[:10], "%Y-%m-%d").date()
    return None


def compute_drr_v3(daily_data: list, total_returns: float):
    """
    DRR per Amazon_DRR_v7.xlsx:
    - Select 30 most recent in-stock days (closing_stock > 0) from 90-day window
    - Raw mean = total_gross / 30
    - Spike threshold = 2x raw mean; spike days replaced with 2x raw mean
    - Net units = max(0, cleaned_gross - total_returns)
    - Final DRR = net_units / 30, rounded to 1 decimal
    Returns: (drr, lookback_label, lookback_sales, drr_flag)
    drr may be a string for error states.
    """
    if not daily_data:
        msg = "Manual Input Required - No Sales History"
        return msg, None, 0, msg

    parsed = []
    for d in daily_data:
        dk = _parse_daily_date(d.get("date"))
        if dk:
            parsed.append({
                "date": dk,
                "closing_stock": d.get("closing_stock", 0) or 0,
                "units_sold": max(0, d.get("units_sold", 0) or 0),
            })

    if not parsed:
        msg = "Manual Input Required - No Sales History"
        return msg, None, 0, msg

    parsed.sort(key=lambda x: x["date"], reverse=True)

    in_stock = [d for d in parsed if d["closing_stock"] > 0 or d["units_sold"] > 0]
    total_in_stock = len(in_stock)

    if total_in_stock == 0:
        msg = "Manual Input Required - No Sales History"
        return msg, None, 0, msg

    if total_in_stock < 30:
        msg = f"Manual Input Required - Only {total_in_stock} Days Found"
        flag = f"Manual Input Required - Only {total_in_stock} In-Stock Days Found"
        return msg, None, 0, flag

    selected = in_stock[:30]
    oldest = min(d["date"] for d in selected)
    newest = max(d["date"] for d in selected)
    lookback_label = f"{oldest.strftime('%-d %b %Y')} – {newest.strftime('%-d %b %Y')}"

    total_gross = sum(d["units_sold"] for d in selected)
    lookback_sales = total_gross

    raw_mean = round(total_gross / 30, 2)
    spike_threshold = max(raw_mean * 2, 5)

    # Replace spike days with 2x mean (spike threshold)
    spike_excess = sum(
        d["units_sold"] - spike_threshold
        for d in selected
        if d["units_sold"] > spike_threshold
    )
    cleaned_units = total_gross - spike_excess
    spike_count = sum(1 for d in selected if d["units_sold"] > spike_threshold)

    # Subtract total returns from cleaned gross units
    total_ret = max(0, total_returns or 0)
    net_units = max(0.0, cleaned_units - total_ret)

    final_drr = round(net_units / 30, 2)

    if spike_count > 0:
        flag = f"OK - {spike_count} Spike Day(s) Replaced with 2x Mean"
    else:
        flag = "OK - Full Confidence"

    return final_drr, lookback_label, lookback_sales, flag


async def _fetch_vc_drr_daily(database, end: datetime, asins: list) -> dict:
    """Fetch up to 90 days of {date, closing_stock, units_sold} from VC collections for DRR."""
    drr_start = end - timedelta(days=89)
    pipeline = [
        {
            "$match": {
                "asin": {"$in": asins},
                "date": {"$gte": drr_start, "$lte": end},
            }
        },
        {
            "$group": {
                "_id": {"asin": "$asin", "date": "$date"},
                "closing_stock": {"$sum": "$sellableOnHandInventoryUnits"},
            }
        },
        {
            "$lookup": {
                "from": "amazon_vendor_sales",
                "let": {"a": "$_id.asin", "d": "$_id.date"},
                "pipeline": [
                    {
                        "$match": {
                            "$expr": {
                                "$and": [
                                    {"$eq": ["$asin", "$$a"]},
                                    {"$eq": ["$date", "$$d"]},
                                ]
                            }
                        }
                    },
                    {"$group": {"_id": None, "units_sold": {"$sum": "$orderedUnits"}}},
                ],
                "as": "sales",
            }
        },
        {
            "$addFields": {
                "units_sold": {"$ifNull": [{"$first": "$sales.units_sold"}, 0]}
            }
        },
        {
            "$group": {
                "_id": "$_id.asin",
                "daily_data": {
                    "$push": {
                        "date": "$_id.date",
                        "closing_stock": "$closing_stock",
                        "units_sold": "$units_sold",
                    }
                },
            }
        },
    ]

    def _run():
        return list(
            database.get_collection("amazon_vendor_inventory").aggregate(pipeline, allowDiskUse=True)
        )

    results = await asyncio.to_thread(_run)
    return {doc["_id"]: serialize_mongo_document(doc["daily_data"]) for doc in results}


async def _fetch_fba_drr_daily(database, end: datetime, asins: list, report_type: str) -> dict:
    """Fetch up to 90 days of {date, closing_stock, units_sold} from FBA/ledger for DRR."""
    drr_start = end - timedelta(days=89)
    ledger_filter: dict = {
        "asin": {"$in": asins},
        "date": {"$gte": drr_start, "$lte": end},
        "disposition": "SELLABLE",
    }
    if report_type == "fba":
        ledger_filter["location"] = {"$ne": "VKSX"}
    elif report_type == "seller_flex":
        ledger_filter["location"] = "VKSX"

    pipeline = [
        {"$match": ledger_filter},
        {
            "$group": {
                "_id": {"asin": "$asin", "date": "$date"},
                "closing_stock": {"$sum": "$ending_warehouse_balance"},
            }
        },
        {
            "$lookup": {
                "from": SALES_COLLECTION,
                "let": {"a": "$_id.asin", "d": "$_id.date"},
                "pipeline": [
                    {
                        "$match": {
                            "$expr": {
                                "$and": [
                                    {"$eq": ["$parentAsin", "$$a"]},
                                    {"$eq": ["$date", "$$d"]},
                                ]
                            }
                        }
                    },
                    {
                        "$group": {
                            "_id": None,
                            "units_sold": {"$sum": "$salesByAsin.unitsOrdered"},
                        }
                    },
                ],
                "as": "sales",
            }
        },
        {
            "$addFields": {
                "units_sold": {"$ifNull": [{"$first": "$sales.units_sold"}, 0]}
            }
        },
        {
            "$group": {
                "_id": "$_id.asin",
                "daily_data": {
                    "$push": {
                        "date": "$_id.date",
                        "closing_stock": "$closing_stock",
                        "units_sold": "$units_sold",
                    }
                },
            }
        },
    ]

    def _run():
        return list(
            database.get_collection("amazon_ledger").aggregate(pipeline, allowDiskUse=True)
        )

    results = await asyncio.to_thread(_run)
    return {doc["_id"]: serialize_mongo_document(doc["daily_data"]) for doc in results}


def compute_drr_for_asins_sync(db, end: datetime, asins: list) -> dict:
    """
    Compute DRR for a list of ASINs exactly as the Amazon 'all' report does:
    merges VC + FBA daily data (90-day window) and subtracts combined returns.
    Returns {asin: {"drr": value_or_str, "drr_flag": str}}.
    """
    drr_start = end - timedelta(days=89)

    # VC daily
    vc_pipeline = [
        {"$match": {"asin": {"$in": asins}, "date": {"$gte": drr_start, "$lte": end}}},
        {"$group": {"_id": {"asin": "$asin", "date": "$date"}, "closing_stock": {"$sum": "$sellableOnHandInventoryUnits"}}},
        {
            "$lookup": {
                "from": "amazon_vendor_sales",
                "let": {"a": "$_id.asin", "d": "$_id.date"},
                "pipeline": [
                    {"$match": {"$expr": {"$and": [{"$eq": ["$asin", "$$a"]}, {"$eq": ["$date", "$$d"]}]}}},
                    {"$group": {"_id": None, "units_sold": {"$sum": "$orderedUnits"}}},
                ],
                "as": "sales",
            }
        },
        {"$addFields": {"units_sold": {"$ifNull": [{"$first": "$sales.units_sold"}, 0]}}},
        {"$group": {"_id": "$_id.asin", "daily_data": {"$push": {"date": "$_id.date", "closing_stock": "$closing_stock", "units_sold": "$units_sold"}}}},
    ]
    vc_results = list(db.get_collection("amazon_vendor_inventory").aggregate(vc_pipeline, allowDiskUse=True))
    vc_daily = {doc["_id"]: doc["daily_data"] for doc in vc_results}

    # FBA daily (all dispositions — fba+seller_flex combined, same as 'all' report)
    fba_pipeline = [
        {"$match": {"asin": {"$in": asins}, "date": {"$gte": drr_start, "$lte": end}, "disposition": "SELLABLE"}},
        {"$group": {"_id": {"asin": "$asin", "date": "$date"}, "closing_stock": {"$sum": "$ending_warehouse_balance"}}},
        {
            "$lookup": {
                "from": SALES_COLLECTION,
                "let": {"a": "$_id.asin", "d": "$_id.date"},
                "pipeline": [
                    {"$match": {"$expr": {"$and": [{"$eq": ["$parentAsin", "$$a"]}, {"$eq": ["$date", "$$d"]}]}}},
                    {"$group": {"_id": None, "units_sold": {"$sum": "$salesByAsin.unitsOrdered"}}},
                ],
                "as": "sales",
            }
        },
        {"$addFields": {"units_sold": {"$ifNull": [{"$first": "$sales.units_sold"}, 0]}}},
        {"$group": {"_id": "$_id.asin", "daily_data": {"$push": {"date": "$_id.date", "closing_stock": "$closing_stock", "units_sold": "$units_sold"}}}},
    ]
    fba_results = list(db.get_collection("amazon_ledger").aggregate(fba_pipeline, allowDiskUse=True))
    fba_daily = {doc["_id"]: doc["daily_data"] for doc in fba_results}

    # Returns: VC customer returns + FBA returns over the 30-day window ending at `end`.
    # Using 30 days to match the Amazon report methodology (returns are scoped to the
    # same period as the selected in-stock days, not the full 90-day lookback window).
    returns_start = end - timedelta(days=29)
    vc_ret_results = list(db.get_collection("amazon_vendor_sales").aggregate([
        {"$match": {"asin": {"$in": asins}, "date": {"$gte": returns_start, "$lte": end}}},
        {"$group": {"_id": "$asin", "total": {"$sum": {"$ifNull": ["$customerReturns", 0]}}}},
    ]))
    vc_returns = {doc["_id"]: doc["total"] for doc in vc_ret_results}

    fba_ret_results = list(db.get_collection(FBA_RETURNS_COLLECTION).aggregate([
        {"$match": {"asin": {"$in": asins}, "return_date": {"$gte": returns_start, "$lte": end}}},
        {"$group": {"_id": "$asin", "total": {"$sum": {"$ifNull": ["$quantity", 0]}}}},
    ]))
    fba_returns = {doc["_id"]: doc["total"] for doc in fba_ret_results}

    def _merge(a_list, b_list):
        m = {}
        for d in a_list:
            k = d["date"]
            m[k] = {"date": k, "closing_stock": d.get("closing_stock", 0) or 0, "units_sold": d.get("units_sold", 0) or 0}
        for d in b_list:
            k = d["date"]
            if k in m:
                m[k]["closing_stock"] = max(m[k]["closing_stock"], d.get("closing_stock", 0) or 0)
                m[k]["units_sold"] += d.get("units_sold", 0) or 0
            else:
                m[k] = {"date": k, "closing_stock": d.get("closing_stock", 0) or 0, "units_sold": d.get("units_sold", 0) or 0}
        return list(m.values())

    result = {}
    for asin in set(vc_daily.keys()) | set(fba_daily.keys()):
        merged = _merge(vc_daily.get(asin, []), fba_daily.get(asin, []))
        total_ret = (vc_returns.get(asin, 0) or 0) + (fba_returns.get(asin, 0) or 0)
        drr, _, _, flag = compute_drr_v3(merged, total_ret)
        result[asin] = {"drr": drr, "drr_flag": flag}
    return result


async def generate_report_by_date_range(
    start_date: str, end_date: str, database: Any, report_type: str = "fba+seller_flex", any_last_90_days: bool = False
):
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d").replace(
            hour=0, minute=0, second=0, microsecond=0, tzinfo=timezone.utc
        )
        end = datetime.strptime(end_date, "%Y-%m-%d").replace(
            hour=23, minute=59, second=59, microsecond=999999, tzinfo=timezone.utc
        )

        # Handle the new "all" report type
        if report_type == "all":
            vendor_data = await generate_vendor_central_data(
                start, end, database, any_last_90_days, keep_daily_data=True
            )
            fba_seller_flex_data = await generate_amazon_data(
                start, end, database, "fba+seller_flex", any_last_90_days, keep_daily_data=True
            )
            combined_data = combine_report_data(vendor_data, fba_seller_flex_data)
            # Strip internal daily data from final output
            for item in combined_data:
                item.pop("_daily_data", None)
            return combined_data

        elif report_type == "vendor_central":
            return await generate_vendor_central_data(start, end, database, any_last_90_days)

        else:
            # Handle fba, seller_flex, fba+seller_flex
            return await generate_amazon_data(start, end, database, report_type, any_last_90_days)

    except Exception as e:
        import traceback

        logger.info(str(e))
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An error occurred generating Amazon report: {str(e)}",
        )


async def generate_vendor_central_data(start, end, database, any_last_90_days: bool = False, keep_daily_data: bool = False):
    """Generate vendor central report data"""
    vendor_pipeline = [
        {
            "$match": {
                "date": {"$gte": start, "$lte": end},
            }
        },
        {
            "$group": {
                "_id": {"asin": "$asin", "date": "$date"},
                "units_sold": {"$sum": "$orderedUnits"},
                "amount": {"$sum": "$orderedRevenue.amount"},
                # Handle missing customerReturns field with $ifNull
                "customer_returns": {"$sum": {"$ifNull": ["$customerReturns", 0]}},
            }
        },
        {
            "$lookup": {
                "from": "amazon_vendor_inventory",
                "let": {"sales_asin": "$_id.asin", "sales_date": "$_id.date"},
                "pipeline": [
                    {
                        "$match": {
                            "$expr": {
                                "$and": [
                                    {"$eq": ["$asin", "$$sales_asin"]},
                                    {"$eq": ["$date", "$$sales_date"]},
                                ]
                            }
                        }
                    },
                    {
                        "$group": {
                            "_id": {"asin": "$asin", "date": "$date"},
                            "closing_stock": {"$sum": "$sellableOnHandInventoryUnits"},
                        }
                    },
                ],
                "as": "inventory_data",
            }
        },
        {"$addFields": {"inventory_data": {"$last": "$inventory_data"}}},
        {
            "$lookup": {
                "from": "amazon_sku_mapping",
                "localField": "_id.asin",
                "foreignField": "item_id",
                "as": "item_info",
            }
        },
        {"$addFields": {"item_info": {"$last": "$item_info"}}},
        {"$sort": {"_id.date": -1}},
        {
            "$group": {
                "_id": {"asin": "$_id.asin"},
                "total_units_sold": {"$sum": "$units_sold"},
                "total_amount": {"$sum": "$amount"},
                "vc_returns_total": {"$sum": "$customer_returns"},
                "last_day_closing_stock": {
                    "$last": {"$ifNull": ["$inventory_data.closing_stock", 0]}
                },
                "item_name": {"$last": "$item_info.item_name"},
                "sku_code": {"$last": "$item_info.sku_code"},
                "daily_data": {
                    "$push": {
                        "date": "$_id.date",
                        "closing_stock": {
                            "$ifNull": ["$inventory_data.closing_stock", 0]
                        },
                        "units_sold": "$units_sold",
                    }
                },
            }
        },
        {
            "$addFields": {
                "total_days_in_stock": {
                    "$cond": {
                        "if": {
                            "$and": [
                                {"$ne": ["$daily_data", None]},
                                {"$gt": [{"$size": "$daily_data"}, 0]},
                            ]
                        },
                        "then": {
                            "$let": {
                                "vars": {
                                    "sorted_data": {
                                        "$sortArray": {
                                            "input": "$daily_data",
                                            "sortBy": {"date": 1},
                                        }
                                    }
                                },
                                "in": {
                                    "$size": {
                                        "$filter": {
                                            "input": "$$sorted_data",
                                            "cond": {
                                                "$gt": [
                                                    "$$this.closing_stock",
                                                    0,
                                                ]
                                            },
                                        }
                                    }
                                },
                            }
                        },
                        "else": 0,
                    }
                }
            }
        },
        {
            "$project": {
                "_id": 0,
                "asin": "$_id.asin",
                "item_name": "$item_name",
                "sku_code": "$sku_code",
                "units_sold": "$total_units_sold",
                "closing_stock": "$last_day_closing_stock",
                "total_amount": {"$round": ["$total_amount", 2]},
                "stock": "$last_day_closing_stock",
                "total_days_in_stock": "$total_days_in_stock",
                "vc_returns": {"$ifNull": ["$vc_returns_total", 0]},
                "daily_data": "$daily_data",
                "data_source": {"$literal": "vendor_central"},
            }
        },
        {
            "$sort": {
                "sku_code": -1,
                "asin": -1,
            }
        },
    ]

    collection = database.get_collection("amazon_vendor_sales")

    def _run_vendor_aggregation():
        return list(collection.aggregate(vendor_pipeline))

    cursor = await asyncio.to_thread(_run_vendor_aggregation)
    result = serialize_mongo_document(cursor)

    # Fetch 90-day DRR daily data (separate from the report-range daily data)
    asins_for_drr = [item["asin"] for item in result if item.get("asin")]
    drr_daily_map: dict = {}
    if asins_for_drr:
        drr_daily_map = await _fetch_vc_drr_daily(database, end, asins_for_drr)

    for item in result:
        daily_data = item.pop("daily_data", [])
        vc_ret = item.get("vc_returns", 0) or 0
        item["total_returns"] = vc_ret
        item["vc_units_sold"] = item.get("units_sold", 0)

        drr_daily = drr_daily_map.get(item.get("asin"), [])
        drr, lookback_label, lookback_sales, drr_flag = compute_drr_v3(drr_daily, vc_ret)
        item["drr"] = drr
        item["drr_lookback"] = lookback_label or ""
        item["drr_lookback_sales"] = lookback_sales
        item["drr_flag"] = drr_flag

        if keep_daily_data:
            item["_daily_data"] = daily_data
            item["_drr_daily_data"] = drr_daily

    # Fetch last 90 days in stock if requested
    if any_last_90_days and result:
        asins = [item["asin"] for item in result if "asin" in item]
        if asins:
            logger.info(f"Fetching last 90 days in stock for {len(asins)} ASINs (Vendor Central)")
            last_90_days_data = await fetch_last_n_days_vendor_inventory(database, end, asins, 90)

            for item in result:
                asin = item.get("asin")
                if asin:
                    item["last_90_days_dates"] = last_90_days_data.get(asin, "")

    return result


async def generate_amazon_data(start, end, database, report_type, any_last_90_days: bool = False, keep_daily_data: bool = False):
    """Generate FBA/Seller Flex report data with returns information (FBA only)"""
    ledger_match_conditions = [
        {"$eq": ["$asin", "$$sales_asin"]},
        {"$eq": ["$date", "$$sales_date"]},
        {"$eq": ["$disposition", "SELLABLE"]},
    ]

    if report_type == "fba":
        ledger_match_conditions.append({"$ne": ["$location", "VKSX"]})
    elif report_type == "seller_flex":
        ledger_match_conditions.append({"$eq": ["$location", "VKSX"]})

    base_pipeline = [
        {
            "$match": {
                "date": {"$gte": start, "$lte": end},
            }
        },
        {
            "$group": {
                "_id": {"asin": "$parentAsin", "date": "$date"},
                "units_sold": {"$sum": "$salesByAsin.unitsOrdered"},
                "amount": {"$sum": "$salesByAsin.orderedProductSales.amount"},
                "sessions": {
                    "$sum": {
                        "$add": [
                            {"$ifNull": ["$trafficByAsin.sessions", 0]},
                            {"$ifNull": ["$trafficByAsin.sessionsB2B", 0]},
                        ]
                    }
                },
            }
        },
        {
            "$lookup": {
                "from": "amazon_ledger",
                "let": {"sales_asin": "$_id.asin", "sales_date": "$_id.date"},
                "pipeline": [
                    {"$match": {"$expr": {"$and": ledger_match_conditions}}},
                    {
                        "$group": {
                            "_id": {"asin": "$asin", "date": "$date"},
                            "closing_stock": {"$sum": "$ending_warehouse_balance"},
                            "item_name": {"$last": "$title"},
                            "warehouses": {"$addToSet": "$location"},
                        }
                    },
                ],
                "as": "ledger_data",
            }
        },
        {"$addFields": {"ledger_data": {"$first": "$ledger_data"}}},
        {
            "$lookup": {
                "from": "amazon_sku_mapping",
                "localField": "_id.asin",
                "foreignField": "item_id",
                "as": "item_info",
            }
        },
        {"$addFields": {"item_info": {"$last": "$item_info"}}},
    ]

    # Add FBA returns lookup for FBA reports (not seller_flex)
    if report_type in ["fba", "fba+seller_flex", "all"]:
        fba_returns_lookup_stage = {
            "$lookup": {
                "from": FBA_RETURNS_COLLECTION,
                "let": {"sales_asin": "$_id.asin"},
                "pipeline": [
                    {
                        "$match": {
                            "$expr": {
                                "$and": [
                                    {"$eq": ["$asin", "$$sales_asin"]},
                                    {"$gte": ["$return_date", start]},
                                    {"$lte": ["$return_date", end]},
                                ]
                            }
                        }
                    },
                    {
                        "$group": {
                            "_id": "$asin",
                            "fba_returns": {
                                "$sum": {"$toInt": {"$ifNull": ["$quantity", 1]}}
                            },
                        }
                    },
                ],
                "as": "fba_returns_data",
            }
        }
        base_pipeline.append(fba_returns_lookup_stage)
        base_pipeline.append(
            {"$addFields": {"fba_returns_data": {"$first": "$fba_returns_data"}}}
        )

    # Add Seller Flex returns lookup for all report types
    sf_returns_lookup_stage = {
        "$lookup": {
            "from": SELLER_FLEX_RETURNS_COLLECTION,
            "let": {"sales_asin": "$_id.asin"},
            "pipeline": [
                {
                    "$match": {
                        "$expr": {
                            "$and": [
                                {"$eq": ["$asin", "$$sales_asin"]},
                                {"$gte": ["$pickup_date", start]},
                                {"$lte": ["$pickup_date", end]},
                                {"$ne": ["$units", None]},
                            ]
                        }
                    }
                },
                {
                    "$group": {
                        "_id": "$asin",
                        "sf_returns": {"$sum": {"$ifNull": ["$units", 0]}},
                    }
                },
            ],
            "as": "sf_returns_data",
        }
    }
    base_pipeline.append(sf_returns_lookup_stage)
    base_pipeline.append(
        {"$addFields": {"sf_returns_data": {"$first": "$sf_returns_data"}}}
    )

    # Continue with the rest of the pipeline
    group_stage = {
        "$group": {
            "_id": {"asin": "$_id.asin"},
            "total_units_sold": {"$sum": "$units_sold"},
            "total_amount": {"$sum": "$amount"},
            "total_sessions": {"$sum": "$sessions"},
            "total_closing_stock": {
                "$last": {"$ifNull": ["$ledger_data.closing_stock", 0]}
            },
            "item_name": {"$last": "$item_info.item_name"},
            "sku_code": {"$last": "$item_info.sku_code"},
            "all_warehouses": {"$addToSet": "$ledger_data.warehouses"},
            "daily_data": {
                "$push": {
                    "date": "$_id.date",
                    "closing_stock": {"$ifNull": ["$ledger_data.closing_stock", 0]},
                    "units_sold": "$units_sold",
                    "sessions": "$sessions",
                }
            },
        }
    }

    # Track FBA and SF returns separately
    if report_type in ["fba", "fba+seller_flex"]:
        group_stage["$group"]["fba_returns_total"] = {
            "$last": {"$ifNull": ["$fba_returns_data.fba_returns", 0]}
        }
    group_stage["$group"]["sf_returns_total"] = {
        "$last": {"$ifNull": ["$sf_returns_data.sf_returns", 0]}
    }

    base_pipeline.append(group_stage)

    # Add remaining pipeline stages
    base_pipeline.extend(
        [
            {
                "$addFields": {
                    "total_days_in_stock": {
                        "$cond": {
                            "if": {
                                "$and": [
                                    {"$ne": ["$daily_data", None]},
                                    {"$gt": [{"$size": "$daily_data"}, 0]},
                                ]
                            },
                            "then": {
                                "$let": {
                                    "vars": {
                                        "sorted_data": {
                                            "$sortArray": {
                                                "input": "$daily_data",
                                                "sortBy": {"date": 1},
                                            }
                                        }
                                    },
                                    "in": {
                                        "$size": {
                                            "$filter": {
                                                "input": "$$sorted_data",
                                                "cond": {
                                                    "$gt": ["$$this.closing_stock", 0]
                                                },
                                            }
                                        }
                                    },
                                }
                            },
                            "else": 0,
                        }
                    },
                    "warehouses": {
                        "$reduce": {
                            "input": "$all_warehouses",
                            "initialValue": [],
                            "in": {"$setUnion": ["$$value", "$$this"]},
                        }
                    },
                }
            },
        ]
    )

    # Build project stage — include daily_data for DRR v3, computed in Python
    project_stage = {
        "$project": {
            "_id": 0,
            "asin": "$_id.asin",
            "sku_code": "$sku_code",
            "item_name": "$item_name",
            "warehouses": "$warehouses",
            "units_sold": "$total_units_sold",
            "total_amount": "$total_amount",
            "sessions": "$total_sessions",
            "closing_stock": "$total_closing_stock",
            "fba_returns": {"$ifNull": ["$fba_returns_total", 0]},
            "seller_flex_returns": {"$ifNull": ["$sf_returns_total", 0]},
            "total_days_in_stock": "$total_days_in_stock",
            "daily_data": "$daily_data",
            "data_source": {"$literal": report_type},
        }
    }

    base_pipeline.extend(
        [
            project_stage,
            {
                "$sort": {
                    "sku_code": -1,
                    "asin": -1,
                }
            },
        ]
    )

    collection = database.get_collection(SALES_COLLECTION)

    def _run_fba_aggregation():
        return list(collection.aggregate(base_pipeline))

    cursor = await asyncio.to_thread(_run_fba_aggregation)
    result = serialize_mongo_document(cursor)

    # Fetch 90-day DRR daily data (separate from the report-range daily data)
    asins_for_drr = [item["asin"] for item in result if item.get("asin")]
    drr_daily_map: dict = {}
    if asins_for_drr:
        drr_daily_map = await _fetch_fba_drr_daily(database, end, asins_for_drr, report_type)

    for item in result:
        daily_data = item.pop("daily_data", [])
        fba_ret = item.get("fba_returns", 0) or 0
        sf_ret = item.get("seller_flex_returns", 0) or 0
        item["total_returns"] = fba_ret + sf_ret
        item["fba_units_sold"] = item.get("units_sold", 0)

        drr_daily = drr_daily_map.get(item.get("asin"), [])
        drr, lookback_label, lookback_sales, drr_flag = compute_drr_v3(drr_daily, item["total_returns"])
        item["drr"] = drr
        item["drr_lookback"] = lookback_label or ""
        item["drr_lookback_sales"] = lookback_sales
        item["drr_flag"] = drr_flag

        if keep_daily_data:
            item["_daily_data"] = daily_data
            item["_drr_daily_data"] = drr_daily

    # Fetch last 90 days in stock if requested
    if any_last_90_days and result:
        asins = [item["asin"] for item in result if "asin" in item]
        if asins:
            logger.info(f"Fetching last 90 days in stock for {len(asins)} ASINs (Amazon Ledger - {report_type})")
            last_90_days_data = await fetch_last_n_days_amazon_ledger(database, end, asins, 90, report_type)

            for item in result:
                asin = item.get("asin")
                if asin:
                    item["last_90_days_dates"] = last_90_days_data.get(asin, "")

    return result


def combine_report_data(vendor_data, fba_seller_flex_data):
    """Combine vendor central and fba/seller flex data"""
    # Create lookup dictionaries
    vendor_lookup = {}
    fba_lookup = {}

    # Index vendor data by sku_code and asin
    for item in vendor_data:
        key = (item.get("sku_code"), item.get("asin"))
        vendor_lookup[key] = item

    # Index fba data by sku_code and asin
    for item in fba_seller_flex_data:
        key = (item.get("sku_code"), item.get("asin"))
        fba_lookup[key] = item

    combined_results = []
    processed_keys = set()

    # Combine matching records
    for key in vendor_lookup:
        if key in fba_lookup:
            vendor_item = vendor_lookup[key]
            fba_item = fba_lookup[key]

            def _merge_daily(a_list, b_list):
                m: dict = {}
                for d in a_list:
                    key2 = d["date"]
                    m[key2] = {"date": key2, "closing_stock": d.get("closing_stock", 0) or 0, "units_sold": d.get("units_sold", 0) or 0}
                for d in b_list:
                    key2 = d["date"]
                    if key2 in m:
                        m[key2]["closing_stock"] = max(m[key2]["closing_stock"], d.get("closing_stock", 0) or 0)
                        m[key2]["units_sold"] += d.get("units_sold", 0) or 0
                    else:
                        m[key2] = {"date": key2, "closing_stock": d.get("closing_stock", 0) or 0, "units_sold": d.get("units_sold", 0) or 0}
                return list(m.values())

            vc_us = vendor_item.get("units_sold", 0) or 0
            fba_us = fba_item.get("units_sold", 0) or 0

            # Merge 90-day DRR daily data from both platforms
            merged_drr = _merge_daily(
                vendor_item.get("_drr_daily_data", []),
                fba_item.get("_drr_daily_data", []),
            )
            drr, lookback_label, lookback_sales, drr_flag = compute_drr_v3(
                merged_drr, (
                    (fba_item.get("fba_returns", 0) or 0)
                    + (fba_item.get("seller_flex_returns", 0) or 0)
                    + (vendor_item.get("vc_returns", 0) or 0)
                )
            )

            combined_item = {
                "asin": vendor_item.get("asin") or fba_item.get("asin"),
                "sku_code": vendor_item.get("sku_code") or fba_item.get("sku_code"),
                "item_name": vendor_item.get("item_name") or fba_item.get("item_name"),
                "warehouses": fba_item.get("warehouses", []),
                "vc_units_sold": vc_us,
                "fba_units_sold": fba_us,
                "units_sold": vc_us + fba_us,
                "total_amount": round(
                    (
                        vendor_item.get("total_amount", 0)
                        + fba_item.get("total_amount", 0)
                    ),
                    2,
                ),
                "sessions": fba_item.get("sessions", 0),
                "closing_stock": (
                    vendor_item.get("closing_stock", 0)
                    + fba_item.get("closing_stock", 0)
                ),
                "stock": (
                    vendor_item.get("stock", 0) + fba_item.get("closing_stock", 0)
                ),
                "vendor_days_in_stock": vendor_item.get("total_days_in_stock", 0),
                "fba_days_in_stock": fba_item.get("total_days_in_stock", 0),
                "total_days_in_stock": max(
                    vendor_item.get("total_days_in_stock", 0),
                    fba_item.get("total_days_in_stock", 0),
                ),
                "vc_stock": vendor_item.get("closing_stock", 0),
                "fba_stock": fba_item.get("closing_stock", 0),
                "fba_returns": fba_item.get("fba_returns", 0) or 0,
                "seller_flex_returns": fba_item.get("seller_flex_returns", 0) or 0,
                "vc_returns": vendor_item.get("vc_returns", 0) or 0,
                "total_returns": (
                    (fba_item.get("fba_returns", 0) or 0)
                    + (fba_item.get("seller_flex_returns", 0) or 0)
                    + (vendor_item.get("vc_returns", 0) or 0)
                ),
                "drr": drr,
                "drr_lookback": lookback_label or "",
                "drr_lookback_sales": lookback_sales,
                "drr_flag": drr_flag,
                "data_source": "combined",
                # Keep merged report-range daily for calendar (will be stripped after download)
                "_daily_data": _merge_daily(
                    vendor_item.get("_daily_data", []),
                    fba_item.get("_daily_data", []),
                ),
            }

            combined_results.append(combined_item)
            processed_keys.add(key)

    # Add vendor-only records
    for key, vendor_item in vendor_lookup.items():
        if key not in processed_keys:
            vendor_item["warehouses"] = []
            vendor_item["sessions"] = 0
            if "stock" not in vendor_item:
                vendor_item["stock"] = vendor_item.get("closing_stock", 0)
            vendor_item["vc_stock"] = vendor_item.get("closing_stock", 0)
            vendor_item["fba_stock"] = 0
            vendor_item.setdefault("fba_returns", 0)
            vendor_item.setdefault("seller_flex_returns", 0)
            vendor_item.setdefault("vc_returns", 0)
            vendor_item.setdefault("total_returns", vendor_item.get("vc_returns", 0))
            vendor_item.setdefault("vc_units_sold", vendor_item.get("units_sold", 0))
            vendor_item.setdefault("fba_units_sold", 0)
            vendor_item.setdefault("drr_lookback", "")
            vendor_item.setdefault("drr_lookback_sales", 0)
            vendor_item.setdefault("drr_flag", "")
            vendor_item["vendor_days_in_stock"] = vendor_item.get("total_days_in_stock", 0)
            vendor_item["fba_days_in_stock"] = 0
            vendor_item["data_source"] = "vendor_only"
            combined_results.append(vendor_item)

    # Add FBA-only records
    for key, fba_item in fba_lookup.items():
        if key not in processed_keys:
            fba_item["stock"] = fba_item.get("closing_stock", 0)
            fba_item["vc_stock"] = 0
            fba_item["fba_stock"] = fba_item.get("closing_stock", 0)
            fba_item.setdefault("vc_returns", 0)
            fba_item.setdefault("vc_units_sold", 0)
            fba_item.setdefault("fba_units_sold", fba_item.get("units_sold", 0))
            fba_item.setdefault("drr_lookback", "")
            fba_item.setdefault("drr_lookback_sales", 0)
            fba_item.setdefault("drr_flag", "")
            fba_item["vendor_days_in_stock"] = 0
            fba_item["fba_days_in_stock"] = fba_item.get("total_days_in_stock", 0)
            fba_item["data_source"] = "fba_only"
            combined_results.append(fba_item)

    # Sort the results
    combined_results.sort(
        key=lambda x: (x.get("sku_code") or "", x.get("asin") or ""), reverse=True
    )

    return combined_results


@router.get("/get_report_data_by_date_range")
async def get_report_data_by_date_range(
    start_date: str,
    end_date: str,
    report_type: str = "fba+seller_flex",
    any_last_90_days: bool = False,
    database=Depends(get_database),
):
    try:
        # Updated valid types to include "all"
        valid_types = ["fba+seller_flex", "fba", "seller_flex", "vendor_central", "all"]
        if report_type not in valid_types:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid type. Must be one of: {', '.join(valid_types)}",
            )

        report_response = await generate_report_by_date_range(
            start_date=start_date,
            end_date=end_date,
            database=database,
            report_type=report_type,
            any_last_90_days=any_last_90_days,
        )
        logger.info(f"Retrieved dynamic report data of type: {report_type}")
        return report_response

    except HTTPException as e:
        raise e
    except Exception as e:
        logger.info(f"Error retrieving report data: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An error occurred retrieving the report data: {e}",
        )


def format_column_name(column_name):
    """Convert snake_case column names to proper formatted names"""
    # Replace underscores with spaces and title case
    formatted = column_name.replace("_", " ").title()

    # Handle specific cases for better formatting
    replacements = {
        "Asin": "ASIN",
        "Sku Code": "SKU Code",
        "Item Name": "Item Name",
        "Units Sold": "Units Sold",
        "Total Amount": "Total Amount",
        "Closing Stock": "Closing Stock",
        "Sessions": "Sessions",
        "Warehouses": "Warehouses",
        "Last 90 Days Dates": "Last 90 Days Dates",
        "Drr": "DRR",
        "Drr Flag": "DRR Flag",
        "Drr Lookback": "DRR Lookback Period",
        "Drr Lookback Sales": "DRR Lookback Sales",
    }

    return replacements.get(formatted, formatted)


async def _get_last_stock_dates(database, start: datetime, end: datetime):
    """
    Query the actual latest date that has stock data in each inventory collection
    within the given range. Returns (vc_date_label, fba_date_label) strings.
    """
    def _query_vc():
        doc = database.get_collection("amazon_vendor_inventory").find_one(
            {"date": {"$gte": start, "$lte": end}},
            sort=[("date", -1)],
            projection={"date": 1},
        )
        return doc["date"] if doc else None

    def _query_fba():
        doc = database.get_collection("amazon_ledger").find_one(
            {"date": {"$gte": start, "$lte": end}, "disposition": "SELLABLE"},
            sort=[("date", -1)],
            projection={"date": 1},
        )
        return doc["date"] if doc else None

    vc_dt, fba_dt = await asyncio.gather(
        asyncio.to_thread(_query_vc),
        asyncio.to_thread(_query_fba),
    )
    fmt = lambda dt: dt.strftime("%d %b %Y") if dt else None
    return fmt(vc_dt), fmt(fba_dt)


def _parse_date_str(date_val):
    """Parse a date from an ISO string or datetime to a date object."""
    if isinstance(date_val, datetime):
        return date_val.date()
    if isinstance(date_val, str):
        try:
            return datetime.fromisoformat(date_val.replace("Z", "+00:00")).date()
        except ValueError:
            return datetime.strptime(date_val[:10], "%Y-%m-%d").date()
    return None


def _build_active_days_df(vc_daily_by_asin: dict, fba_daily_by_asin: dict, report_data: list, all_dates: list) -> pd.DataFrame:
    """
    Active Days sheet: one row per product, columns = dates.
    Cell = total units sold (VC + FBA) if sales > 0,
           0 if in stock on any platform but no sale,
           blank (None) if no stock at all.
    """
    asin_info = {
        item["asin"]: {"sku_code": item.get("sku_code", ""), "item_name": item.get("item_name", "")}
        for item in report_data
        if item.get("asin")
    }

    all_asins = sorted(set(vc_daily_by_asin.keys()) | set(fba_daily_by_asin.keys()))
    date_cols = [d.strftime("%d %b") for d in all_dates]
    rows = []

    for asin in all_asins:
        vc_by_date = {_parse_date_str(d["date"]): d for d in vc_daily_by_asin.get(asin, []) if _parse_date_str(d.get("date"))}
        fba_by_date = {_parse_date_str(d["date"]): d for d in fba_daily_by_asin.get(asin, []) if _parse_date_str(d.get("date"))}
        info = asin_info.get(asin, {})

        row = {"ASIN": asin, "SKU Code": info.get("sku_code", ""), "Item Name": info.get("item_name", "")}
        days_in_stock = 0
        for dt, col in zip(all_dates, date_cols):
            vc_day = vc_by_date.get(dt) or {}
            fba_day = fba_by_date.get(dt) or {}
            vc_stock = vc_day.get("closing_stock", 0) or 0
            fba_stock = fba_day.get("closing_stock", 0) or 0
            vc_sold = vc_day.get("units_sold", 0) or 0
            fba_sold = fba_day.get("units_sold", 0) or 0
            total_sold = vc_sold + fba_sold
            in_stock = vc_stock > 0 or fba_stock > 0

            if total_sold > 0:
                row[col] = total_sold
                days_in_stock += 1
            elif in_stock:
                row[col] = 0
                days_in_stock += 1
            else:
                row[col] = None  # blank

        row["Total Days In Stock"] = days_in_stock
        ordered = {k: row[k] for k in ["ASIN", "SKU Code", "Item Name", "Total Days In Stock"]}
        ordered.update({col: row[col] for col in date_cols})
        rows.append(ordered)

    return pd.DataFrame(rows) if rows else None


def _fetch_sf_fba_inventory_latest_sync(db, end):
    """Returns ({asin: {"sf": qty|None, "fba": qty|None}}, date_str). SF=VKSX, FBA=all others."""
    latest_doc = db["amazon_ledger"].find_one(
        {"date": {"$lte": end}}, sort=[("date", -1)], projection={"date": 1}
    )
    if not latest_doc:
        return {}, None
    latest_date = latest_doc["date"]
    date_str = latest_date.strftime("%-d %b %Y") if hasattr(latest_date, "strftime") else str(latest_date)
    inv: dict = {}
    for doc in db["amazon_ledger"].aggregate([
        {"$match": {"date": latest_date}},
        {"$group": {
            "_id": {"asin": "$asin", "is_sf": {"$eq": ["$location_key", "VKSX"]}},
            "stock": {"$sum": "$ending_warehouse_balance"},
        }},
    ]):
        asin = doc["_id"]["asin"]
        is_sf = doc["_id"]["is_sf"]
        stock = int(doc.get("stock") or 0)
        if asin not in inv:
            inv[asin] = {"sf": None, "fba": None}
        if is_sf:
            inv[asin]["sf"] = stock
        else:
            inv[asin]["fba"] = stock
    return inv, date_str


def _fetch_vc_inventory_latest_sync(db, end):
    """Returns ({asin: qty}, date_str) from amazon_vendor_inventory."""
    latest_doc = db["amazon_vendor_inventory"].find_one(
        {"date": {"$lte": end}}, sort=[("date", -1)], projection={"date": 1}
    )
    if not latest_doc:
        return {}, None
    latest_date = latest_doc["date"]
    date_str = latest_date.strftime("%-d %b %Y") if hasattr(latest_date, "strftime") else str(latest_date)
    inv: dict = {}
    for doc in db["amazon_vendor_inventory"].aggregate([
        {"$match": {"date": latest_date}},
        {"$group": {"_id": "$asin", "stock": {"$sum": "$sellableOnHandInventoryUnits"}}},
    ]):
        inv[doc["_id"]] = int(doc.get("stock") or 0)
    return inv, date_str


def _fetch_df_inventory_latest_sync(db, end):
    """Returns ({asin: qty}, date_str) from amazon_df_inventory (empty until populated)."""
    latest_doc = db["amazon_df_inventory"].find_one(
        {"date": {"$lte": end}}, sort=[("date", -1)], projection={"date": 1}
    )
    if not latest_doc:
        return {}, None
    latest_date = latest_doc["date"]
    date_str = latest_date.strftime("%-d %b %Y") if hasattr(latest_date, "strftime") else str(latest_date)
    inv: dict = {}
    for doc in db["amazon_df_inventory"].aggregate([
        {"$match": {"date": latest_date}},
        {"$group": {"_id": "$asin", "stock": {"$sum": "$sellableOnHandInventoryUnits"}}},
    ]):
        inv[doc["_id"]] = int(doc.get("stock") or 0)
    return inv, date_str


def _fetch_zoho_stock_for_asins_sync(db, asins):
    """Returns ({asin: qty}, date_str) from zoho_warehouse_stock, Pupscribe WH only."""
    sku_by_asin: dict = {}
    for doc in db["amazon_sku_mapping"].find(
        {"item_id": {"$in": asins}}, {"item_id": 1, "sku_code": 1, "_id": 0}
    ):
        sku_by_asin[doc["item_id"]] = doc.get("sku_code", "")
    skus = list(set(sku_by_asin.values()))
    if not skus:
        return {}, None
    item_id_by_sku: dict = {}
    for doc in db["products"].find(
        {"cf_sku_code": {"$in": skus}}, {"cf_sku_code": 1, "item_id": 1, "_id": 0}
    ):
        item_id_by_sku[doc["cf_sku_code"]] = doc.get("item_id", "")
    item_id_by_asin = {
        asin: item_id_by_sku[sku]
        for asin, sku in sku_by_asin.items()
        if sku in item_id_by_sku and item_id_by_sku[sku]
    }
    zoho_item_ids = list(set(item_id_by_asin.values()))
    if not zoho_item_ids:
        return {}, None
    latest_doc = db["zoho_warehouse_stock"].find_one(
        {"zoho_item_id": {"$in": zoho_item_ids}},
        sort=[("date", -1)],
        projection={"date": 1},
    )
    if not latest_doc:
        return {}, None
    latest_date = latest_doc["date"]
    date_str = latest_date.strftime("%-d %b %Y") if hasattr(latest_date, "strftime") else str(latest_date)
    stock_by_item_id: dict = {}
    for doc in db["zoho_warehouse_stock"].aggregate([
        {"$match": {"zoho_item_id": {"$in": zoho_item_ids}, "date": latest_date}},
        {"$group": {
            "_id": "$zoho_item_id",
            "stock": {"$first": {"$ifNull": ["$warehouses.Pupscribe Enterprises Private Limited", 0]}},
        }},
    ]):
        stock_by_item_id[doc["_id"]] = int(doc.get("stock") or 0)
    stock_by_asin: dict = {}
    for asin, item_id in item_id_by_asin.items():
        stock_by_asin[asin] = stock_by_item_id.get(item_id, 0)
    return stock_by_asin, date_str


def _fetch_vc_open_po_qty_sync(db, asins):
    """Returns {asin: total_open_qty} from vendor_purchase_orders."""
    result: dict = {}
    for doc in db["vendor_purchase_orders"].aggregate([
        {"$match": {"po_status": {"$in": ["processing", "packed", "closed", "intransit"]}}},
        {"$unwind": "$items"},
        {"$match": {"items.asin": {"$in": asins}}},
        {"$project": {
            "asin": "$items.asin",
            "qty": {"$switch": {
                "branches": [{"case": {"$eq": ["$po_status", "processing"]}, "then": {
                    "$cond": [
                        {"$ifNull": ["$items.supply_qty_override", False]},
                        "$items.supply_qty_override",
                        {"$cond": [
                            {"$ifNull": ["$items.final_supply_fo", False]},
                            "$items.final_supply_fo",
                            {"$cond": [
                                {"$gt": [{"$ifNull": ["$items.supply_qty", 0]}, 0]},
                                "$items.supply_qty",
                                {"$ifNull": ["$items.requested_qty", 0]},
                            ]},
                        ]},
                    ],
                }}],
                "default": {"$ifNull": ["$items.accepted_qty", 0]},
            }},
        }},
        {"$group": {"_id": "$asin", "total": {"$sum": "$qty"}}},
    ]):
        result[doc["_id"]] = int(doc.get("total") or 0)
    return result


def _fetch_fba_sit_qty_sync(db, asins):
    """Returns {asin: open_shipment_qty} from amazon_fba_shipment_processing."""
    result: dict = {}
    for doc in db["amazon_fba_shipment_processing"].aggregate([
        {"$match": {"asin": {"$in": asins}}},
        {"$group": {"_id": "$asin", "total": {"$sum": {"$ifNull": ["$requested_qty", 0]}}}},
    ]):
        result[doc["_id"]] = int(doc.get("total") or 0)
    return result


def _build_drr_calculator_sheet(ws, report_data, drr_vc_daily_by_asin, drr_fba_daily_by_asin, drr_dates, report_type, end_date_label, purchase_status_map=None, extra_inventory=None):
    """
    Build Active Days sheet matching reference layout.

    Layout:
      Row 1  – title with end date
      Row 2  – section group headers
      Row 3  – column labels (D1-D90, H1-H90, calc columns, new inventory cols)
      Row 4  – actual dates for D1-D90 and H1-H90
      Row 5  – blank separator
      Row 6+ – one data row per ASIN with values + Excel formulas
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule

    if purchase_status_map is None:
        purchase_status_map = {}
    if extra_inventory is None:
        extra_inventory = {}

    sf_fba_inv  = extra_inventory.get("sf_fba", {})
    vc_inv      = extra_inventory.get("vc", {})
    df_inv      = extra_inventory.get("df", {})
    zoho_stock  = extra_inventory.get("zoho", {})
    open_po_qty = extra_inventory.get("open_po", {})
    fba_sit_qty = extra_inventory.get("fba_sit", {})
    sf_date_str   = extra_inventory.get("sf_date")
    fba_date_str  = extra_inventory.get("fba_date")
    vc_date_str   = extra_inventory.get("vc_date")
    df_date_str   = extra_inventory.get("df_date")
    zoho_date_str = extra_inventory.get("zoho_date")

    def _live_status(inv_val):
        if inv_val is None:
            return "Not Listed"
        return "Live" if inv_val > 0 else "Not Live"

    def _dated_hdr(label, date_str):
        return f"{label}\n({date_str})" if date_str else label

    # ── Column index constants (1-based) ────────────────────────────────
    COL_STATUS   = 1    # Item Status
    COL_ASIN     = 2    # ASIN
    COL_SKU      = 3    # SKU Code
    COL_SF_LIVE  = 4    # SF Live (Date)           [Section 1 – new]
    COL_FBA_LIVE = 5    # FBA Live (Date)          [Section 1 – new]
    COL_VC_LIVE  = 6    # VC Live (Date)           [Section 1 – new]
    COL_DF_LIVE  = 7    # DF Live (Date)           [Section 1 – new]
    COL_NAME     = 8    # Item Name
    COL_B        = 9    # FBA Sales (incl. SF)
    COL_C        = 10   # VC Sales
    COL_D        = 11   # FBA Returns
    COL_E        = 12   # SF Returns
    COL_F        = 13   # VC Returns
    COL_G        = 14   # Total Returns (=D+E+F)
    COL_H        = 15   # blank spacer
    COL_D1       = 16   # D1 – most recent day
    COL_D90      = 105  # D90 – oldest day  (16 + 89)
    COL_H1       = 106  # H1 – first helper
    COL_H90      = 195  # H90 – last helper (106 + 89)
    COL_GF       = 196  # Total Active Days (max 30)
    COL_GG       = 197  # Selected Sum (30 days)
    COL_GH       = 198  # Raw Mean (÷30)
    COL_GI       = 199  # Spike Threshold (2× Mean)
    COL_GJ       = 200  # Cleaned Units (spikes→threshold)
    COL_GK       = 201  # Spike Days Found
    COL_GL       = 202  # Net Units (after returns)
    COL_GM       = 203  # blank spacer
    COL_GN       = 204  # Final DRR (units/day)
    COL_GO       = 205  # DRR Flag
    # Sections 2–4 (new)
    COL_ZOHO     = 206  # Zoho Stock (Date)
    COL_FBA_INV  = 207  # FBA inventory (Date)
    COL_SF_INV   = 208  # SF inventory (Date)
    COL_VC_INV   = 209  # VC inventory (Date)
    COL_DF_INV   = 210  # DF inventory (Date)
    COL_DAYS_FBA = 211  # Days Until Stock Lasts – FBA
    COL_DAYS_SF  = 212  # Days Until Stock Lasts – SF
    COL_DAYS_VC  = 213  # Days Until Stock Lasts – VC
    COL_DAYS_DF  = 214  # Days Until Stock Lasts – DF
    COL_OPEN_PO  = 215  # Open PO Qty
    COL_DAYS_PO  = 216  # Days Until Stock Lasts – Open PO + VC
    COL_FBA_SIT  = 217  # FBA Stock in Transit
    COL_DAYS_SIT = 218  # Days Until Stock Lasts – FBA + SIT

    R_TITLE      = 1
    R_SECTION    = 2
    R_HEADERS    = 3
    R_DATES      = 4
    R_DATA_START = 6  # row 5 left blank

    L = get_column_letter

    # ── Row 2: section group headers ────────────────────────────────────
    ws.cell(row=R_SECTION, column=COL_ASIN).value    = "RAW INPUTS"
    ws.cell(row=R_SECTION, column=COL_SF_LIVE).value = "Item Status"
    ws.cell(row=R_SECTION, column=COL_D1).value = (
        f"90 DAILY COLUMNS [D1 = {end_date_label}]  —  Blank=OOS  |  0=In Stock No Sale  |"
        "  Number=Units Sold  |  Day 1 = Most Recent"
    )
    ws.cell(row=R_SECTION, column=COL_H1).value  = "HELPER ROW — Cumulative count of non-blank days. Selected if count ≤ 30."
    ws.cell(row=R_SECTION, column=COL_GF).value  = "CALCULATIONS"
    ws.cell(row=R_SECTION, column=COL_GL).value  = "NET UNITS"
    ws.cell(row=R_SECTION, column=COL_GN).value  = "FINAL DRR"
    ws.cell(row=R_SECTION, column=COL_GO).value  = "DRR FLAG"
    ws.cell(row=R_SECTION, column=COL_ZOHO).value    = "INVENTORY"
    ws.cell(row=R_SECTION, column=COL_DAYS_FBA).value = "DAYS UNTIL STOCK LASTS"
    ws.cell(row=R_SECTION, column=COL_OPEN_PO).value  = "OPEN PO / IN TRANSIT"
    ws.row_dimensions[R_SECTION].height = 30

    # ── Row 3: column headers ────────────────────────────────────────────
    fixed_headers = {
        COL_STATUS:   "Item Status",
        COL_ASIN:     "ASIN",
        COL_SKU:      "SKU Code",
        COL_SF_LIVE:  _dated_hdr("SF Live", sf_date_str),
        COL_FBA_LIVE: _dated_hdr("FBA Live", fba_date_str),
        COL_VC_LIVE:  _dated_hdr("VC Live", vc_date_str),
        COL_DF_LIVE:  _dated_hdr("DF Live", df_date_str),
        COL_NAME:     "Item Name",
        COL_B:        "FBA Sales\n(incl. SF)",
        COL_C:        "VC Sales",
        COL_D:        "FBA\nReturns",
        COL_E:        "SF\nReturns",
        COL_F:        "VC\nReturns",
        COL_G:        "Total\nReturns",
        COL_GF:       "Total Active\nDays (max 30)",
        COL_GG:       "Selected Sum\n(30 days)",
        COL_GH:       "Raw Mean\n(÷30)",
        COL_GI:       "Spike Threshold\n(2× Mean)",
        COL_GJ:       "Cleaned Units\n(spikes→2× Mean)",
        COL_GK:       "Spike Days\nFound",
        COL_GL:       "Net Units\n(after returns)",
        COL_GN:       "Final DRR\n(units/day)",
        COL_GO:       "DRR Flag",
        COL_ZOHO:     _dated_hdr("Zoho Stock", zoho_date_str),
        COL_FBA_INV:  _dated_hdr("FBA", fba_date_str),
        COL_SF_INV:   _dated_hdr("SF", sf_date_str),
        COL_VC_INV:   _dated_hdr("VC", vc_date_str),
        COL_DF_INV:   _dated_hdr("DF", df_date_str),
        COL_DAYS_FBA: "FBA",
        COL_DAYS_SF:  "SF",
        COL_DAYS_VC:  "VC",
        COL_DAYS_DF:  "DF",
        COL_OPEN_PO:  "Open PO Qty",
        COL_DAYS_PO:  "Days Until Stock\nLasts (Open PO + VC)",
        COL_FBA_SIT:  "FBA Stock\nin Transit",
        COL_DAYS_SIT: "Days Until Stock\nLasts (FBA + SIT)",
    }
    for k in range(90):
        fixed_headers[COL_D1 + k] = f"D{k + 1}"
        fixed_headers[COL_H1 + k] = f"H{k + 1}"
    for col_idx, label in fixed_headers.items():
        c = ws.cell(row=R_HEADERS, column=col_idx)
        c.value = label
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # ── Row 4: actual dates for D1-D90 and H1-H90 ──────────────────────
    for k in range(90):
        date_str = drr_dates[89 - k].strftime("%d %b %Y")
        ws.cell(row=R_DATES, column=COL_D1 + k).value = date_str
        ws.cell(row=R_DATES, column=COL_H1 + k).value = date_str

    SPIKE_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # ── Data rows ────────────────────────────────────────────────────────
    asin_to_item = {item["asin"]: item for item in report_data if item.get("asin")}
    report_asins = {item["asin"] for item in report_data if item.get("asin")}
    all_asins = sorted(report_asins | set(drr_vc_daily_by_asin.keys()) | set(drr_fba_daily_by_asin.keys()))

    for row_offset, asin in enumerate(all_asins):
        r = R_DATA_START + row_offset
        item = asin_to_item.get(asin, {})

        # Item Status
        sku  = item.get("sku_code", "")
        ws.cell(row=r, column=COL_STATUS).value = purchase_status_map.get(sku, "")

        # Section 1 – Live status per platform
        asin_ledger = sf_fba_inv.get(asin, {})
        sf_qty  = asin_ledger.get("sf")
        fba_qty = asin_ledger.get("fba")
        vc_qty  = vc_inv.get(asin) if asin in vc_inv else None
        df_qty  = df_inv.get(asin) if asin in df_inv else None

        ws.cell(row=r, column=COL_SF_LIVE).value  = _live_status(sf_qty)
        ws.cell(row=r, column=COL_FBA_LIVE).value = _live_status(fba_qty)
        ws.cell(row=r, column=COL_VC_LIVE).value  = _live_status(vc_qty)
        ws.cell(row=r, column=COL_DF_LIVE).value  = _live_status(df_qty)

        # ASIN / SKU / Item Name
        name = item.get("item_name", "")
        ws.cell(row=r, column=COL_ASIN).value = asin
        ws.cell(row=r, column=COL_SKU).value  = sku
        ws.cell(row=r, column=COL_NAME).value = name

        # B – FBA Sales (incl. SF)
        if report_type == "all":
            fba_sales = item.get("fba_units_sold", 0) or 0
        elif report_type == "vendor_central":
            fba_sales = 0
        else:
            fba_sales = item.get("units_sold", 0) or 0
        ws.cell(row=r, column=COL_B).value = fba_sales

        # C – VC Sales
        if report_type == "all":
            vc_sales = item.get("vc_units_sold", 0) or 0
        elif report_type == "vendor_central":
            vc_sales = item.get("units_sold", 0) or 0
        else:
            vc_sales = 0
        ws.cell(row=r, column=COL_C).value = vc_sales

        # D – FBA Returns
        fba_ret = 0 if report_type == "vendor_central" else (item.get("fba_returns", 0) or 0)
        ws.cell(row=r, column=COL_D).value = fba_ret

        # E – SF Returns
        sf_ret = 0 if report_type == "vendor_central" else (item.get("seller_flex_returns", 0) or 0)
        ws.cell(row=r, column=COL_E).value = sf_ret

        # F – VC Returns
        ws.cell(row=r, column=COL_F).value = item.get("vc_returns", 0) or 0

        # G – Total Returns formula
        ws.cell(row=r, column=COL_G).value = f"={L(COL_D)}{r}+{L(COL_E)}{r}+{L(COL_F)}{r}"

        # Build date-keyed daily lookups
        vc_by_date  = {_parse_date_str(d.get("date")): d for d in drr_vc_daily_by_asin.get(asin, []) if _parse_date_str(d.get("date"))}
        fba_by_date = {_parse_date_str(d.get("date")): d for d in drr_fba_daily_by_asin.get(asin, []) if _parse_date_str(d.get("date"))}

        # D1-D90 – merged daily: blank=OOS, 0=in-stock-no-sale, N=units sold
        for k in range(90):
            dt = drr_dates[89 - k]
            vc_day  = vc_by_date.get(dt)  or {}
            fba_day = fba_by_date.get(dt) or {}
            vc_stock  = vc_day.get("closing_stock",  0) or 0
            fba_stock = fba_day.get("closing_stock", 0) or 0
            vc_sold   = max(0, vc_day.get("units_sold",  0) or 0)
            fba_sold  = max(0, fba_day.get("units_sold", 0) or 0)
            if vc_stock > 0 or fba_stock > 0 or vc_sold > 0 or fba_sold > 0:
                ws.cell(row=r, column=COL_D1 + k).value = vc_sold + fba_sold

        # Spike highlighting
        gi_ltr  = L(COL_GI)
        d1_ltr  = L(COL_D1)
        d90_ltr = L(COL_D90)
        ws.conditional_formatting.add(
            f"{d1_ltr}{r}:{d90_ltr}{r}",
            FormulaRule(
                formula=[f"AND(NOT(ISBLANK({d1_ltr}{r})),{d1_ltr}{r}>${gi_ltr}{r})"],
                fill=SPIKE_FILL,
            ),
        )

        # H1-H90 – cumulative non-blank count
        for k in range(90):
            d_ltr = L(COL_D1 + k)
            if k == 0:
                formula = f"=IF(NOT(ISBLANK({d_ltr}{r})),1,0)"
            else:
                ph = L(COL_H1 + k - 1)
                formula = f"=IF(NOT(ISBLANK({d_ltr}{r})),{ph}{r}+1,{ph}{r})"
            ws.cell(row=r, column=COL_H1 + k).value = formula

        # Calculation columns (GF–GO)
        ge = L(COL_H90); ct = L(COL_H1)
        h  = L(COL_D1);  cs = L(COL_D90)
        gf = L(COL_GF);  gg = L(COL_GG); gh = L(COL_GH); gi = L(COL_GI)
        gj = L(COL_GJ);  gk = L(COL_GK); gl = L(COL_GL); f_ = L(COL_G)
        gn = L(COL_GN)

        ws.cell(row=r, column=COL_GF).value = f"=MIN({ge}{r},30)"
        ws.cell(row=r, column=COL_GG).value = (
            f"=IFERROR(SUMPRODUCT(({ct}{r}:{ge}{r}<=30)"
            f"*(NOT(ISBLANK({h}{r}:{cs}{r})))"
            f"*IF(ISBLANK({h}{r}:{cs}{r}),0,{h}{r}:{cs}{r})),0)"
        )
        ws.cell(row=r, column=COL_GH).value = f"=IFERROR(IF({gf}{r}<30,0,ROUND({gg}{r}/30,2)),0)"
        ws.cell(row=r, column=COL_GI).value = f"=MAX({gh}{r}*2,5)"
        ws.cell(row=r, column=COL_GJ).value = (
            f"=IFERROR(ROUND({gg}{r}-SUMPRODUCT(({ct}{r}:{ge}{r}<=30)"
            f"*(NOT(ISBLANK({h}{r}:{cs}{r})))"
            f"*(IF(ISBLANK({h}{r}:{cs}{r}),0,{h}{r}:{cs}{r})>{gi}{r})"
            f"*(IF(ISBLANK({h}{r}:{cs}{r}),0,{h}{r}:{cs}{r})-{gi}{r})),2)"
            f',\"Check Inputs\")'
        )
        ws.cell(row=r, column=COL_GK).value = (
            f"=IFERROR(SUMPRODUCT(({ct}{r}:{ge}{r}<=30)"
            f"*(NOT(ISBLANK({h}{r}:{cs}{r})))"
            f"*(IF(ISBLANK({h}{r}:{cs}{r}),0,{h}{r}:{cs}{r})>{gi}{r})*1),0)"
        )
        ws.cell(row=r, column=COL_GL).value = f'=IFERROR(MAX(0,{gj}{r}-{f_}{r}),"Check Inputs")'
        ws.cell(row=r, column=COL_GN).value = (
            f'=IFERROR(IF({gf}{r}=0,"Manual Input Required - No Sales History",'
            f'IF({gf}{r}<30,"Manual Input Required - Only "&{gf}{r}&" In-Stock Days Found",'
            f'ROUND({gl}{r}/30,2))),"Check Inputs")'
        )
        ws.cell(row=r, column=COL_GO).value = (
            f'=IFERROR(IF({gf}{r}=0,"Manual Input Required - No Sales History",'
            f'IF({gf}{r}<30,"Manual Input Required - Only "&{gf}{r}&" In-Stock Days Found",'
            f'IF({gk}{r}>0,"OK - "&{gk}{r}&" Spike Day(s) Replaced with 2x Mean",'
            f'"OK - Full Confidence"))),"Check Inputs")'
        )

        # Section 2 – Inventory values
        ws.cell(row=r, column=COL_ZOHO).value    = zoho_stock.get(asin, 0)
        ws.cell(row=r, column=COL_FBA_INV).value = fba_qty if fba_qty is not None else 0
        ws.cell(row=r, column=COL_SF_INV).value  = sf_qty  if sf_qty  is not None else 0
        ws.cell(row=r, column=COL_VC_INV).value  = vc_qty  if vc_qty  is not None else 0
        ws.cell(row=r, column=COL_DF_INV).value  = df_qty  if df_qty  is not None else 0

        # Section 3 – Days Until Stock Lasts formulas (inv / Final DRR)
        def _days_formula(inv_col):
            iv = L(inv_col)
            return (
                f'=IFERROR(IF(ISNUMBER({gn}{r}),IF({gn}{r}=0,"",ROUND({iv}{r}/{gn}{r},1)),""),"")'
            )
        ws.cell(row=r, column=COL_DAYS_FBA).value = _days_formula(COL_FBA_INV)
        ws.cell(row=r, column=COL_DAYS_SF).value  = _days_formula(COL_SF_INV)
        ws.cell(row=r, column=COL_DAYS_VC).value  = _days_formula(COL_VC_INV)
        ws.cell(row=r, column=COL_DAYS_DF).value  = _days_formula(COL_DF_INV)

        # Section 4 – Open PO / SIT
        ws.cell(row=r, column=COL_OPEN_PO).value = open_po_qty.get(asin, 0)
        vc_inv_ltr = L(COL_VC_INV); po_ltr = L(COL_OPEN_PO)
        ws.cell(row=r, column=COL_DAYS_PO).value = (
            f'=IFERROR(IF(ISNUMBER({gn}{r}),IF({gn}{r}=0,"",ROUND(({vc_inv_ltr}{r}+{po_ltr}{r})/{gn}{r},1)),""),"")'
        )
        ws.cell(row=r, column=COL_FBA_SIT).value = fba_sit_qty.get(asin, 0)
        fba_inv_ltr = L(COL_FBA_INV); sit_ltr = L(COL_FBA_SIT)
        ws.cell(row=r, column=COL_DAYS_SIT).value = (
            f'=IFERROR(IF(ISNUMBER({gn}{r}),IF({gn}{r}=0,"",ROUND(({fba_inv_ltr}{r}+{sit_ltr}{r})/{gn}{r},1)),""),"")'
        )

    # ── Styling ──────────────────────────────────────────────────────────
    C_BLUE_D   = "366092"
    C_RED_D    = "C00000"
    C_GOLD_D   = "8B6914"
    C_GRAY_D   = "595959"
    C_GREEN_D  = "375623"
    C_GOLD_S   = "BF9000"
    C_GRAY_S   = "808080"
    C_GREEN_S  = "1E4620"
    C_RED_S    = "ED0000"
    C_DATE_BG  = "FFF8DC"

    # Row 1 – title
    t = ws.cell(row=R_TITLE, column=COL_ASIN)
    t.value = f"Active Days — Report Ending {end_date_label}"
    t.font  = Font(bold=True, size=13, color="FFFFFF")
    t.fill  = PatternFill(start_color=C_BLUE_D, end_color=C_BLUE_D, fill_type="solid")
    ws.row_dimensions[R_TITLE].height = 26

    # Row 2 – section group headers
    sec_map = {
        COL_ASIN:     (C_BLUE_D,  "FFFFFF"),
        COL_SF_LIVE:  (C_RED_S,   "FFFFFF"),
        COL_D1:       (C_GOLD_S,  "FFFFFF"),
        COL_H1:       (C_GRAY_S,  "FFFFFF"),
        COL_GF:       (C_GREEN_S, "FFFFFF"),
        COL_GL:       (C_GREEN_S, "FFFFFF"),
        COL_GN:       (C_GREEN_S, "FFFFFF"),
        COL_GO:       (C_GREEN_S, "FFFFFF"),
        COL_ZOHO:     (C_RED_S,   "FFFFFF"),
        COL_DAYS_FBA: (C_RED_S,   "FFFFFF"),
        COL_OPEN_PO:  (C_RED_S,   "FFFFFF"),
    }
    for col_idx, (bg, fg) in sec_map.items():
        c = ws.cell(row=R_SECTION, column=col_idx)
        c.font  = Font(bold=True, color=fg)
        c.fill  = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        c.alignment = Alignment(wrap_text=True, vertical="center")

    # Row 3 – column label styling
    yellow_fill  = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_hdr_fill = PatternFill(start_color=C_RED_D,  end_color=C_RED_D,  fill_type="solid")
    for col_idx in range(1, COL_DAYS_SIT + 1):
        c = ws.cell(row=R_HEADERS, column=col_idx)
        if not c.value:
            continue
        if col_idx == COL_STATUS:
            c.font = Font(bold=True, color="000000")
            c.fill = yellow_fill
        elif col_idx in (COL_SF_LIVE, COL_FBA_LIVE, COL_VC_LIVE, COL_DF_LIVE):
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = red_hdr_fill
        elif col_idx <= COL_G:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color=C_BLUE_D, end_color=C_BLUE_D, fill_type="solid")
        elif col_idx <= COL_D90:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color=C_GOLD_D, end_color=C_GOLD_D, fill_type="solid")
        elif col_idx <= COL_H90:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color=C_GRAY_D, end_color=C_GRAY_D, fill_type="solid")
        elif col_idx <= COL_GO:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color=C_GREEN_D, end_color=C_GREEN_D, fill_type="solid")
        else:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = red_hdr_fill
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws.row_dimensions[R_HEADERS].height = 42

    # Row 4 – date labels
    d_date_fill = PatternFill(start_color=C_DATE_BG, end_color=C_DATE_BG, fill_type="solid")
    h_date_fill = PatternFill(start_color="E8E8E8",  end_color="E8E8E8",  fill_type="solid")
    for k in range(90):
        for col_base, fill in ((COL_D1, d_date_fill), (COL_H1, h_date_fill)):
            c = ws.cell(row=R_DATES, column=col_base + k)
            c.fill      = fill
            c.font      = Font(size=8, italic=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[R_DATES].height = 18

    # ── Column widths ────────────────────────────────────────────────────
    ws.column_dimensions[L(COL_STATUS)].width   = 18
    ws.column_dimensions[L(COL_ASIN)].width     = 16
    ws.column_dimensions[L(COL_SKU)].width      = 18
    for ci in (COL_SF_LIVE, COL_FBA_LIVE, COL_VC_LIVE, COL_DF_LIVE):
        ws.column_dimensions[L(ci)].width       = 14
    ws.column_dimensions[L(COL_NAME)].width     = 40
    for ci in range(COL_B, COL_H):
        ws.column_dimensions[L(ci)].width       = 10
    ws.column_dimensions[L(COL_H)].width        = 2
    for k in range(90):
        ws.column_dimensions[L(COL_D1 + k)].width = 11
    for k in range(90):
        ws.column_dimensions[L(COL_H1 + k)].width = 4
    for ci in (COL_GF, COL_GG, COL_GH, COL_GI, COL_GJ, COL_GK, COL_GL):
        ws.column_dimensions[L(ci)].width       = 12
    ws.column_dimensions[L(COL_GM)].width       = 2
    ws.column_dimensions[L(COL_GN)].width       = 15
    ws.column_dimensions[L(COL_GO)].width       = 38
    ws.column_dimensions[L(COL_ZOHO)].width     = 16
    for ci in (COL_FBA_INV, COL_SF_INV, COL_VC_INV, COL_DF_INV):
        ws.column_dimensions[L(ci)].width       = 12
    for ci in (COL_DAYS_FBA, COL_DAYS_SF, COL_DAYS_VC, COL_DAYS_DF):
        ws.column_dimensions[L(ci)].width       = 11
    ws.column_dimensions[L(COL_OPEN_PO)].width  = 13
    ws.column_dimensions[L(COL_DAYS_PO)].width  = 20
    ws.column_dimensions[L(COL_FBA_SIT)].width  = 14
    ws.column_dimensions[L(COL_DAYS_SIT)].width = 20

    # Freeze: cols A–O visible (daily cols start at P/col 16)
    ws.freeze_panes = f"{L(COL_D1)}6"


def _style_sheet(worksheet, df, header_color="366092"):
    """Apply standard header styling and auto-width to a worksheet."""
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    for column in worksheet.columns:
        max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
        worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)


def _style_calendar_sheet(worksheet, df):
    """Style a calendar sheet: header + alternating row colors for units/stock rows."""
    from openpyxl.styles import Font, PatternFill, Alignment
    _style_sheet(worksheet, df, header_color="366092")

    units_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    stock_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    metric_col_idx = None
    for i, cell in enumerate(worksheet[1], 1):
        if cell.value == "Metric":
            metric_col_idx = i
            break

    for row_num in range(2, len(df) + 2):
        if metric_col_idx:
            metric_val = worksheet.cell(row=row_num, column=metric_col_idx).value or ""
            fill = units_fill if "Units" in metric_val else stock_fill
        else:
            fill = units_fill if row_num % 2 == 0 else stock_fill
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

    # Freeze the first 4 identifier columns
    worksheet.freeze_panes = "E2"


def _build_inventory_sheet(ws, daily_by_asin, report_data, all_dates, title, platform, purchase_status_map=None, mode="sales"):
    """
    Calendar-style sheet: one row per ASIN, one column per day.
    mode="sales"     : cell = units_sold (green), 0 if in-stock/no-sale (blue), blank if OOS.
    mode="inventory" : cell = closing_stock (blue if > 0), blank if OOS.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    if purchase_status_map is None:
        purchase_status_map = {}
    if not all_dates:
        return

    is_vc = platform == "VC"
    C_TITLE_BG  = "1F3864" if is_vc else "1E4620"
    C_HDR_BG    = "2E75B6" if is_vc else "375623"
    C_DATE_BG   = "4472C4" if is_vc else "548235"
    C_INFO_BG   = "D6E4F7" if is_vc else "E2EFDA"
    C_SOLD_BG   = "C6EFCE"
    C_STOCK_BG  = "DEEBF7"
    C_ALT_ROW   = "F7F7F7"

    R_TITLE = 1
    R_INFO  = 2
    R_HDR   = 3
    R_DATA  = 4

    COL_ASIN  = 1
    COL_SKU   = 2
    COL_NAME  = 3
    COL_TOTAL = 4
    COL_DIS   = 5
    COL_STOCK = 6
    COL_DATE1 = 7

    L = get_column_letter
    date_col_count = len(all_dates)
    last_col = COL_DATE1 + date_col_count - 1

    # ── Row 1: title ────────────────────────────────────────────────────
    ws.merge_cells(start_row=R_TITLE, start_column=1, end_row=R_TITLE, end_column=last_col)
    tc = ws.cell(row=R_TITLE, column=1)
    tc.value = title
    tc.font  = Font(bold=True, size=13, color="FFFFFF")
    tc.fill  = PatternFill(start_color=C_TITLE_BG, end_color=C_TITLE_BG, fill_type="solid")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[R_TITLE].height = 26

    # ── Row 2: info / legend ─────────────────────────────────────────────
    date_range_str = f"{all_dates[0].strftime('%d %b %Y')} — {all_dates[-1].strftime('%d %b %Y')}"
    ws.merge_cells(start_row=R_INFO, start_column=1, end_row=R_INFO, end_column=last_col)
    ic = ws.cell(row=R_INFO, column=1)
    if mode == "inventory":
        legend_text = "Blue = In Stock (closing stock shown)     Blank = Out of Stock"
    else:
        legend_text = "Green = Units Sold     Blue = In Stock / No Sale     Blank = Out of Stock"
    ic.value = f"Date Range: {date_range_str}     {legend_text}"
    ic.font  = Font(italic=True, size=10, color=C_TITLE_BG)
    ic.fill  = PatternFill(start_color=C_INFO_BG, end_color=C_INFO_BG, fill_type="solid")
    ic.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[R_INFO].height = 20

    # ── Row 3: column headers ────────────────────────────────────────────
    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill(start_color=C_HDR_BG, end_color=C_HDR_BG, fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    date_fill = PatternFill(start_color=C_DATE_BG, end_color=C_DATE_BG, fill_type="solid")

    # Find the actual last date with closing stock data across all ASINs
    _stocked_dates = set()
    for _asin_data in daily_by_asin.values():
        for _d in _asin_data:
            _dt = _parse_date_str(_d.get("date"))
            if _dt and (_d.get("closing_stock") or 0) > 0:
                _stocked_dates.add(_dt)
    stock_date = max(_stocked_dates) if _stocked_dates else (all_dates[-1] - timedelta(days=2))

    fixed_hdrs = {
        COL_ASIN:  "ASIN",
        COL_SKU:   "SKU Code",
        COL_NAME:  "Item Name",
        COL_TOTAL: "Avg Daily\nStock" if mode == "inventory" else "Total Units\nSold",
        COL_DIS:   "Days In\nStock",
        COL_STOCK: f"Closing Stock\n({stock_date.strftime('%d %b %Y')})",
    }
    for ci, label in fixed_hdrs.items():
        c = ws.cell(row=R_HDR, column=ci)
        c.value = label
        c.font  = hdr_font
        c.fill  = hdr_fill
        c.alignment = hdr_align

    for i, dt in enumerate(all_dates):
        c = ws.cell(row=R_HDR, column=COL_DATE1 + i)
        c.value = dt.strftime("%d %b")
        c.font  = Font(bold=True, color="FFFFFF", size=9)
        c.fill  = date_fill
        c.alignment = Alignment(horizontal="center", vertical="bottom", text_rotation=45)
    ws.row_dimensions[R_HDR].height = 52

    # ── Data rows ────────────────────────────────────────────────────────
    asin_info = {
        item["asin"]: {"sku_code": item.get("sku_code", ""), "item_name": item.get("item_name", "")}
        for item in report_data if item.get("asin")
    }
    all_asins = sorted(set(daily_by_asin.keys()))
    if not all_asins:
        all_asins = [item["asin"] for item in report_data if item.get("asin")]

    sold_fill  = PatternFill(start_color=C_SOLD_BG,  end_color=C_SOLD_BG,  fill_type="solid")
    stock_fill = PatternFill(start_color=C_STOCK_BG, end_color=C_STOCK_BG, fill_type="solid")
    alt_fill   = PatternFill(start_color=C_ALT_ROW,  end_color=C_ALT_ROW,  fill_type="solid")
    no_fill    = PatternFill(fill_type=None)

    for row_offset, asin in enumerate(all_asins):
        r = R_DATA + row_offset
        row_bg = alt_fill if row_offset % 2 == 1 else no_fill
        info   = asin_info.get(asin, {})
        sku    = info.get("sku_code", "")
        name   = info.get("item_name", "")

        by_date = {
            _parse_date_str(d.get("date")): d
            for d in daily_by_asin.get(asin, [])
            if _parse_date_str(d.get("date"))
        }

        total_sold    = 0
        total_stock   = 0
        days_in_stock = 0
        latest_stock  = 0

        for i, dt in enumerate(all_dates):
            day   = by_date.get(dt) or {}
            sold  = day.get("units_sold",    0) or 0
            stock = day.get("closing_stock", 0) or 0
            in_stock = stock > 0 or sold > 0
            if dt == stock_date:
                latest_stock = stock

            c = ws.cell(row=r, column=COL_DATE1 + i)

            if mode == "inventory":
                if stock > 0:
                    total_stock   += stock
                    days_in_stock += 1
                    c.value     = stock
                    c.fill      = stock_fill
                    c.font      = Font(size=9, color="1F497D")
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.fill = row_bg
            else:
                total_sold    += sold
                days_in_stock += 1 if in_stock else 0
                if sold > 0:
                    c.value     = sold
                    c.fill      = sold_fill
                    c.font      = Font(bold=True, size=9)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif in_stock:
                    c.value     = 0
                    c.fill      = stock_fill
                    c.font      = Font(size=9, color="888888")
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.fill = row_bg

        def _set(col, val, bold=False, align="left", color="000000", size=10):
            c = ws.cell(row=r, column=col)
            c.value     = val
            c.font      = Font(bold=bold, size=size, color=color)
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.fill      = row_bg

        if mode == "inventory":
            avg_stock = round(total_stock / days_in_stock) if days_in_stock > 0 else 0
            col_total_val = avg_stock
        else:
            col_total_val = total_sold

        _set(COL_ASIN,  asin,           size=9,  color="595959")
        _set(COL_SKU,   sku,            bold=True)
        _set(COL_NAME,  name)
        _set(COL_TOTAL, col_total_val,  bold=True, align="center")
        _set(COL_DIS,   days_in_stock,  align="center")
        _set(COL_STOCK, latest_stock,   align="center")
        ws.row_dimensions[r].height = 18

    # ── Column widths ────────────────────────────────────────────────────
    ws.column_dimensions[L(COL_ASIN)].width  = 14
    ws.column_dimensions[L(COL_SKU)].width   = 16
    ws.column_dimensions[L(COL_NAME)].width  = 36
    ws.column_dimensions[L(COL_TOTAL)].width = 11
    ws.column_dimensions[L(COL_DIS)].width   = 9
    ws.column_dimensions[L(COL_STOCK)].width = 16
    for i in range(date_col_count):
        ws.column_dimensions[L(COL_DATE1 + i)].width = 7

    # Freeze identifier columns + header rows
    ws.freeze_panes = f"{L(COL_DATE1)}{R_DATA}"


@router.get("/download_report_by_date_range")
async def download_report_by_date_range(
    start_date: str,
    end_date: str,
    report_type: str = "fba+seller_flex",  # "fba+seller_flex", "fba", "seller_flex", "vendor_central", or "all"
    any_last_90_days: bool = False,
    database=Depends(get_database),
):

    try:
        # Validate report_type parameter
        valid_types = ["fba+seller_flex", "fba", "seller_flex", "vendor_central", "all"]
        if report_type not in valid_types:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid report_type. Must be one of: {', '.join(valid_types)}",
            )

        start = datetime.strptime(start_date, "%Y-%m-%d").replace(
            hour=0, minute=0, second=0, microsecond=0, tzinfo=timezone.utc
        )
        end = datetime.strptime(end_date, "%Y-%m-%d").replace(
            hour=23, minute=59, second=59, microsecond=999999, tzinfo=timezone.utc
        )
        end_label = datetime.strptime(end_date, "%Y-%m-%d").strftime("%d %b %Y")

        # Resolve actual last-data dates from the inventory collections
        vc_date_label, fba_date_label = await _get_last_stock_dates(database, start, end)
        vc_date_label = vc_date_label or end_label
        fba_date_label = fba_date_label or end_label

        # All dates in the selected range (for calendar sheets)
        all_dates = []
        cur = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_d = datetime.strptime(end_date, "%Y-%m-%d").date()
        while cur <= end_d:
            all_dates.append(cur)
            cur += timedelta(days=1)

        # 90-day DRR window ending at end_date (for Active Days sheet)
        drr_dates = [end_d - timedelta(days=i) for i in range(89, -1, -1)]

        # --- Fetch data with daily detail retained ---
        vc_daily_by_asin = {}
        fba_daily_by_asin = {}
        # DRR-window (90-day) daily data per platform for Active Days sheet
        drr_vc_daily_by_asin: dict = {}
        drr_fba_daily_by_asin: dict = {}

        if report_type == "all":
            vendor_raw = await generate_vendor_central_data(start, end, database, any_last_90_days, keep_daily_data=True)
            fba_raw = await generate_amazon_data(start, end, database, "fba+seller_flex", any_last_90_days, keep_daily_data=True)
            # Collect per-platform daily data before combining
            for item in vendor_raw:
                vc_daily_by_asin[item["asin"]] = item.get("_drr_daily_data", [])
                drr_vc_daily_by_asin[item["asin"]] = item.get("_drr_daily_data", [])
            for item in fba_raw:
                fba_daily_by_asin[item["asin"]] = item.get("_daily_data", [])
                drr_fba_daily_by_asin[item["asin"]] = item.get("_drr_daily_data", [])
            report_data = combine_report_data(vendor_raw, fba_raw)
            for item in report_data:
                item.pop("_daily_data", None)
                item.pop("_drr_daily_data", None)

        elif report_type == "vendor_central":
            vendor_raw = await generate_vendor_central_data(start, end, database, any_last_90_days, keep_daily_data=True)
            for item in vendor_raw:
                drr_data = item.pop("_drr_daily_data", [])
                item.pop("_daily_data", None)
                vc_daily_by_asin[item["asin"]] = drr_data
                drr_vc_daily_by_asin[item["asin"]] = drr_data
            report_data = vendor_raw

        else:  # fba, seller_flex, fba+seller_flex
            fba_raw = await generate_amazon_data(start, end, database, report_type, any_last_90_days, keep_daily_data=True)
            for item in fba_raw:
                fba_daily_by_asin[item["asin"]] = item.pop("_daily_data", [])
                drr_fba_daily_by_asin[item["asin"]] = item.pop("_drr_daily_data", [])
            report_data = fba_raw

        # --- Include all ASINs from sku_mapping even if no sales in the selected range ---
        known_asins = {item["asin"] for item in report_data if item.get("asin")}

        def _fetch_all_sku_mapping():
            return list(database.get_collection("amazon_sku_mapping").find(
                {}, {"item_id": 1, "sku_code": 1, "item_name": 1, "_id": 0}
            ))

        all_sku_docs = await asyncio.to_thread(_fetch_all_sku_mapping)
        missing_asin_ids = []
        for doc in all_sku_docs:
            asin = doc.get("item_id", "")
            if not asin or asin in known_asins:
                continue
            missing_asin_ids.append(asin)
            report_data.append({
                "asin": asin,
                "sku_code": doc.get("sku_code", ""),
                "item_name": doc.get("item_name", ""),
                "units_sold": 0,
                "vc_units_sold": 0,
                "fba_units_sold": 0,
                "fba_returns": 0,
                "seller_flex_returns": 0,
                "vc_returns": 0,
                "total_returns": 0,
                "closing_stock": 0,
                "total_amount": 0,
                "drr": None,
                "drr_flag": "no_data",
            })

        if missing_asin_ids:
            if report_type in ("vendor_central", "all"):
                missing_vc_drr = await _fetch_vc_drr_daily(database, end, missing_asin_ids)
                drr_vc_daily_by_asin.update(missing_vc_drr)
                vc_daily_by_asin.update(missing_vc_drr)
            if report_type != "vendor_central":
                fba_type = "fba+seller_flex" if report_type == "all" else report_type
                missing_fba_drr = await _fetch_fba_drr_daily(database, end, missing_asin_ids, fba_type)
                drr_fba_daily_by_asin.update(missing_fba_drr)
                fba_daily_by_asin.update(missing_fba_drr)

        # For "all" reports: ASINs with FBA sales but no VC sales won't be in
        # vc_daily_by_asin. Fetch their VC inventory data so the VC Inventory
        # sheet shows the correct closing stock.
        if report_type == "all":
            fba_only_asins = [
                item["asin"] for item in report_data
                if item.get("asin") and item["asin"] not in vc_daily_by_asin
            ]
            if fba_only_asins:
                extra_vc = await _fetch_vc_drr_daily(database, end, fba_only_asins)
                vc_daily_by_asin.update(extra_vc)

        if not report_data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No data found for the specified date range and report type",
            )

        # --- Fetch Section 1/2/4 inventory data in parallel ---
        all_asin_list = [item["asin"] for item in report_data if item.get("asin")]
        (
            (sf_fba_inv, sf_fba_date),
            (vc_inv_data, vc_inv_date),
            (df_inv_data, df_inv_date),
            (zoho_data,   zoho_date),
            vc_open_po,
            fba_sit,
        ) = await asyncio.gather(
            asyncio.to_thread(_fetch_sf_fba_inventory_latest_sync, database, end),
            asyncio.to_thread(_fetch_vc_inventory_latest_sync,     database, end),
            asyncio.to_thread(_fetch_df_inventory_latest_sync,     database, end),
            asyncio.to_thread(_fetch_zoho_stock_for_asins_sync,    database, all_asin_list),
            asyncio.to_thread(_fetch_vc_open_po_qty_sync,          database, all_asin_list),
            asyncio.to_thread(_fetch_fba_sit_qty_sync,             database, all_asin_list),
        )
        extra_inventory = {
            "sf_fba":    sf_fba_inv,
            "vc":        vc_inv_data,
            "df":        df_inv_data,
            "zoho":      zoho_data,
            "open_po":   vc_open_po,
            "fba_sit":   fba_sit,
            "sf_date":   sf_fba_date,
            "fba_date":  sf_fba_date,
            "vc_date":   vc_inv_date,
            "df_date":   df_inv_date,
            "zoho_date": zoho_date,
        }

        # --- Build main DataFrame ---
        df = pd.DataFrame(report_data)

        # Standard column rename
        column_mapping = {col: format_column_name(col) for col in df.columns}

        # Override stock columns with actual-last-data date labels
        if report_type == "all":
            vc_stock_col = f"VC Stock (as of {vc_date_label})"
            fba_stock_col = f"FBA Stock (as of {fba_date_label})"
            stock_col = f"Total Stock (as of {vc_date_label})"
            column_mapping["vc_stock"] = vc_stock_col
            column_mapping["fba_stock"] = fba_stock_col
            column_mapping["closing_stock"] = stock_col
        elif report_type == "vendor_central":
            stock_col = f"VC Stock (as of {vc_date_label})"
            column_mapping["closing_stock"] = stock_col
        else:
            stock_col = f"FBA Stock (as of {fba_date_label})"
            column_mapping["closing_stock"] = stock_col

        # Drop the raw duplicate 'stock' field
        column_mapping.pop("stock", None)

        df = df.rename(columns=column_mapping)
        if "Stock" in df.columns:
            df = df.drop(columns=["Stock"])

        # Warehouses: list → string
        if "Warehouses" in df.columns:
            df["Warehouses"] = df["Warehouses"].apply(
                lambda x: ", ".join(x) if isinstance(x, list) else str(x)
            )

        # --- Column order ---
        if report_type == "vendor_central":
            preferred_order = [
                "ASIN", "SKU Code", "Item Name",
                "Total Units Sold", "Total Amount",
                stock_col,
                "VC Returns", "Total Returns",
                "Total Days In Stock", "DRR", "DRR Flag", "DRR Lookback Period", "DRR Lookback Sales",
            ]
        elif report_type == "all":
            preferred_order = [
                "ASIN", "SKU Code", "Item Name", "Warehouses",
                "VC Units Sold", "FBA/SF Units Sold", "Total Units Sold",
                "Total Amount", "Sessions",
                vc_stock_col, fba_stock_col, stock_col,
                "FBA Returns", "SF Returns", "VC Returns", "Total Returns",
                "VC Days In Stock", "FBA Days In Stock", "Total Days In Stock",
                "DRR", "DRR Flag", "DRR Lookback Period", "DRR Lookback Sales",
                "Data Source",
            ]
        else:
            preferred_order = [
                "ASIN", "SKU Code", "Item Name", "Warehouses",
                "Total Units Sold", "Total Amount", "Sessions",
                stock_col,
                "FBA Returns", "SF Returns", "Total Returns",
                "Total Days In Stock", "DRR", "DRR Flag", "DRR Lookback Period", "DRR Lookback Sales",
            ]

        if any_last_90_days:
            preferred_order.append("Last 90 Days Dates")

        column_order = [col for col in preferred_order if col in df.columns]
        remaining_cols = [col for col in df.columns if col not in column_order]
        df = df[column_order + remaining_cols]

        # --- Fetch purchase_status from products collection ---
        sku_codes = [item.get("sku_code", "") for item in report_data if item.get("sku_code")]
        purchase_status_map = {}
        if sku_codes:
            products_col = database.get_collection("products")
            docs = await asyncio.to_thread(
                lambda: list(products_col.find(
                    {"cf_sku_code": {"$in": sku_codes}},
                    {"cf_sku_code": 1, "purchase_status": 1, "_id": 0}
                ))
            )
            for doc in docs:
                purchase_status_map[doc["cf_sku_code"]] = doc.get("purchase_status", "")

        # --- Write Excel ---
        excel_buffer = io.BytesIO()

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Active Days"
        _build_drr_calculator_sheet(
            ws,
            report_data,
            drr_vc_daily_by_asin,
            drr_fba_daily_by_asin,
            drr_dates,
            report_type,
            end_label,
            purchase_status_map,
            extra_inventory,
        )

        # 90-day date range label for sheet titles
        drr_start_label = drr_dates[0].strftime("%d %b %Y")
        drr_end_label   = drr_dates[-1].strftime("%d %b %Y")
        drr_range_str   = f"{drr_start_label} to {drr_end_label}"

        # Vendor Sales → Vendor Inventory → FBA Sales → FBA Inventory
        if drr_vc_daily_by_asin:
            ws_vc_sales = wb.create_sheet("Vendor Sales")
            _build_inventory_sheet(
                ws_vc_sales,
                drr_vc_daily_by_asin,
                report_data,
                drr_dates,
                f"Vendor Central — Sales Report  ({drr_range_str})",
                "VC",
                purchase_status_map,
                mode="sales",
            )

        if drr_vc_daily_by_asin:
            ws_vc_inv = wb.create_sheet("Vendor Inventory")
            _build_inventory_sheet(
                ws_vc_inv,
                drr_vc_daily_by_asin,
                report_data,
                drr_dates,
                f"Vendor Central — Inventory Report  ({drr_range_str})",
                "VC",
                purchase_status_map,
                mode="inventory",
            )

        if drr_fba_daily_by_asin:
            ws_fba_sales = wb.create_sheet("FBA Sales")
            _build_inventory_sheet(
                ws_fba_sales,
                drr_fba_daily_by_asin,
                report_data,
                drr_dates,
                f"FBA — Sales Report  ({drr_range_str})",
                "FBA",
                purchase_status_map,
                mode="sales",
            )

        if drr_fba_daily_by_asin:
            ws_fba_inv = wb.create_sheet("FBA Inventory")
            _build_inventory_sheet(
                ws_fba_inv,
                drr_fba_daily_by_asin,
                report_data,
                drr_dates,
                f"FBA — Inventory Report  ({drr_range_str})",
                "FBA",
                purchase_status_map,
                mode="inventory",
            )

        wb.save(excel_buffer)

        excel_buffer.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        type_suffix = f"_{report_type}" if report_type != "fba+seller_flex" else ""
        filename = f"inventory_report_{start_date}_to_{end_date}{type_suffix}_{timestamp}.xlsx"

        response = StreamingResponse(
            io.BytesIO(excel_buffer.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

        logger.info(f"Generated Excel report for report_type: {report_type}, rows: {len(df)}")
        return response

    except HTTPException as e:
        logger.info(str(e))
        raise e
    except Exception as e:
        logger.info(f"Error generating Excel report: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An error occurred generating the Excel report: {e}",
        )


def format_column_name(column_name):
    """Convert snake_case column names to proper formatted names"""
    formatted = column_name.replace("_", " ").title()
    replacements = {
        "Asin": "ASIN",
        "Sku Code": "SKU Code",
        "Item Name": "Item Name",
        "Units Sold": "Total Units Sold",
        "Vc Units Sold": "VC Units Sold",
        "Fba Units Sold": "FBA/SF Units Sold",
        "Total Amount": "Total Amount",
        "Closing Stock": "Closing Stock",
        "Sessions": "Sessions",
        "Warehouses": "Warehouses",
        "Data Source": "Data Source",
        "Total Days In Stock": "Total Days In Stock",
        "Drr": "DRR",
        "Drr Lookback": "DRR Lookback Period",
        "Drr Lookback Sales": "DRR Lookback Sales",
        "Fba Returns": "FBA Returns",
        "Seller Flex Returns": "SF Returns",
        "Vc Returns": "VC Returns",
        "Total Returns": "Total Returns",
        "Stock": "Stock",
        "Vendor Days In Stock": "VC Days In Stock",
        "Fba Days In Stock": "FBA Days In Stock",
        "Last 90 Days Dates": "Last 90 Days Dates",
    }
    return replacements.get(formatted, formatted)

@router.post("/sync/settlements")
def sync_settlement_reports(
    days_back: Optional[int] = 7,
    db=Depends(get_database),
):
    try:
        logger.info(f"Starting settlement reports sync for last {days_back} days")
        from io import StringIO
        import pandas as pd
        import time
        import requests
        from datetime import datetime, timedelta
        from fastapi.responses import JSONResponse
        from fastapi import status, HTTPException

        marketplace_endpoints = {
            "A21TJRUUN4KGV": "https://sellingpartnerapi-eu.amazon.com",
            "A1F83G8C2ARO7P": "https://sellingpartnerapi-eu.amazon.com",
            "A1PA6795UKMFR9": "https://sellingpartnerapi-eu.amazon.com",
            "ATVPDKIKX0DER": "https://sellingpartnerapi-na.amazon.com",
            "A2EUQ1WTGCTBG2": "https://sellingpartnerapi-na.amazon.com",
            "A1VC38T7YXB528": "https://sellingpartnerapi-fe.amazon.com",
        }

        base_url = marketplace_endpoints.get(
            MARKETPLACE_ID, "https://sellingpartnerapi-eu.amazon.com"
        )

        # Get access token
        access_token = get_amazon_token()

        # MongoDB collection
        collection = db["amazon_settlements"]

        # Get settlement reports
        reports_url = f"{base_url}/reports/2021-06-30/reports"
        headers = {"x-amz-access-token": access_token}

        created_since = datetime.now() - timedelta(days=days_back)

        params = {
            "reportTypes": "GET_V2_SETTLEMENT_REPORT_DATA_FLAT_FILE_V2",
            "createdSince": created_since.strftime("%Y-%m-%dT%H:%M:%SZ"),
            "pageSize": 100,
        }

        response = requests.get(reports_url, headers=headers, params=params)
        response.raise_for_status()

        reports = response.json().get("reports", [])

        if not reports:
            logger.info("No settlement reports found")
            return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={
                    "message": "No settlement reports found",
                    "reports_processed": 0,
                    "records_inserted": 0,
                },
            )

        # Filter completed reports
        done_reports = [r for r in reports if r.get("processingStatus") == "DONE"]

        logger.info(f"Found {len(done_reports)} completed settlement reports")

        total_inserted = 0
        reports_processed = 0

        # Process each report
        for report in done_reports:
            report_id = report.get("reportId")
            report_date = report.get("dataStartTime", "")[:10]

            try:
                # Get report document ID
                report_url = f"{base_url}/reports/2021-06-30/reports/{report_id}"
                report_response = requests.get(report_url, headers=headers)
                report_response.raise_for_status()

                report_data = report_response.json()
                document_id = report_data.get("reportDocumentId")

                if not document_id:
                    logger.warning(f"No document ID for report {report_id}")
                    continue

                # Get document download URL
                doc_url = f"{base_url}/reports/2021-06-30/documents/{document_id}"
                doc_response = requests.get(doc_url, headers=headers)
                doc_response.raise_for_status()

                download_url = doc_response.json().get("url")

                # Download report data
                data_response = requests.get(download_url)
                data_response.raise_for_status()

                # Parse TSV data
                df = pd.read_csv(
                    StringIO(data_response.text),
                    sep="\t",
                    encoding="utf-8",
                    low_memory=False,
                )

                # Extract required fields
                required_fields = [
                    "order-id",
                    "amount-description",
                    "amount",
                    "sku",
                    "quantity-purchased",
                    "posted-date",
                ]

                existing_fields = [
                    field for field in required_fields if field in df.columns
                ]

                if not existing_fields:
                    logger.warning(f"No required fields found in report {report_id}")
                    continue

                extracted_df = df[existing_fields].copy()
                records = extracted_df.to_dict("records")

                # Process records
                processed_records = []

                for record in records:
                    new_record = {}

                    # Map fields
                    field_mapping = {
                        "order-id": "order_id",
                        "amount-description": "amount_description",
                        "amount": "amount",
                        "sku": "sku",
                        "quantity-purchased": "quantity_purchased",
                        "posted-date": "posted_date",
                    }

                    for old_key, new_key in field_mapping.items():
                        if old_key in record:
                            value = record[old_key]
                            new_record[new_key] = None if pd.isna(value) or value == "" else value

                    # Skip if order_id is missing
                    if not new_record.get("order_id"):
                        continue

                    # Add metadata
                    new_record["report_id"] = report_id
                    new_record["report_date"] = report_date
                    new_record["created_at"] = datetime.now()

                    # Data type conversions
                    try:
                        if new_record.get("amount") is not None:
                            new_record["amount"] = float(new_record["amount"])
                    except:
                        pass

                    try:
                        if new_record.get("posted_date") is not None:
                            new_record["posted_date"] = datetime.strptime(
                                new_record["posted_date"], "%d.%m.%Y"
                            )
                    except:
                        pass

                    try:
                        new_record["report_date"] = datetime.strptime(report_date, "%Y-%m-%d")
                    except:
                        pass

                    try:
                        if new_record.get("quantity_purchased") is not None:
                            new_record["quantity_purchased"] = int(new_record["quantity_purchased"])
                    except:
                        pass

                    processed_records.append(new_record)

                if not processed_records:
                    logger.warning(f"No valid records in report {report_id}")
                    continue

                # 🚀 Insert all records (no duplicate checking)
                result = collection.insert_many(processed_records, ordered=False)
                inserted = len(result.inserted_ids)
                total_inserted += inserted
                reports_processed += 1

                logger.info(f"Report {report_id}: Inserted {inserted} records (duplicates allowed)")

                # Small delay to avoid rate limiting
                time.sleep(0.5)

            except Exception as e:
                logger.error(f"Error processing report {report_id}: {e}")
                continue

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "message": "Settlement reports sync completed",
                "reports_processed": reports_processed,
                "records_inserted": total_inserted,
                "days_back": days_back,
            },
        )

    except Exception as e:
        logger.error(f"Error syncing settlement reports: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to sync settlement reports: {str(e)}",
        )


@router.get("/settlements/pivot")
def get_settlements_pivot(
    start_date: Optional[str] = Query(
        None, description="Start date in YYYY-MM-DD format"
    ),
    end_date: Optional[str] = Query(None, description="End date in YYYY-MM-DD format"),
    sku: Optional[str] = Query(None, description="Filter by SKU"),
    order_id: Optional[str] = Query(None, description="Filter by order ID"),
    format: str = Query("json", description="Response format: 'json' or 'excel'"),
    db=Depends(get_database),
):
    """Get Amazon settlements data in pivot table format."""
    try:
        collection = db["amazon_settlements"]

        # Build MongoDB query
        query = {}

        # Step 1: If date range is specified, first find all order_ids within that range
        if start_date or end_date:
            date_query = {}
            if start_date:
                try:
                    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                    date_query["$gte"] = start_dt
                except ValueError:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Invalid start_date format. Use YYYY-MM-DD",
                    )
            if end_date:
                try:
                    end_dt = datetime.strptime(end_date, "%Y-%m-%d")
                    date_query["$lte"] = end_dt
                except ValueError:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Invalid end_date format. Use YYYY-MM-DD",
                    )

            # Find all order_ids that have at least one record in the date range
            orders_in_range = collection.distinct("order_id", {"posted_date": date_query})

            # Now fetch ALL records for those orders (regardless of date)
            query["order_id"] = {"$in": orders_in_range}

        if sku:
            query["sku"] = sku

        if order_id:
            # If specific order_id is provided, override the date-based order filtering
            query["order_id"] = order_id

        # Fetch data from MongoDB
        cursor = collection.find(query)
        settlements = list(cursor)

        if not settlements:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No settlement records found for the given criteria",
            )

        # Convert to DataFrame
        df = pd.DataFrame(settlements)
        df = df[
            [
                "order_id",
                "sku",
                "quantity_purchased",
                "amount_description",
                "amount",
            ]
        ]

        # Group by order+sku+amount_description and sum quantities and amounts
        # This aggregates all records across ALL dates for the same order
        grouping_cols = ["order_id", "sku", "amount_description"]
        df = df.groupby(grouping_cols, as_index=False).agg({
            "quantity_purchased": "sum",
            "amount": "sum"
        })

        # Get total quantity per order (max across all amount_descriptions to handle any discrepancies)
        quantity_per_order = df.groupby(["order_id", "sku"])["quantity_purchased"].max().reset_index()

        # Create pivot table - one row per order+sku
        pivot_df = df.pivot_table(
            index=["order_id", "sku"],
            columns="amount_description",
            values="amount",
            aggfunc="sum",
            fill_value=None,
        )

        pivot_df = pivot_df.reset_index()

        # Merge the quantity back in
        pivot_df = pivot_df.merge(
            quantity_per_order,
            on=["order_id", "sku"],
            how="left"
        )

        # Calculate Grand Total (sum of all amount columns)
        amount_columns = [
            col
            for col in pivot_df.columns
            if col not in ["order_id", "sku", "quantity_purchased"]
        ]
        pivot_df["Grand Total"] = pivot_df[amount_columns].sum(axis=1)

        # 🔧 FIX: Replace NaN/inf values with None for JSON serialization
        pivot_df = pivot_df.replace([float("nan"), float("inf"), float("-inf")], None)

        # Sort by order_id
        pivot_df = pivot_df.sort_values("order_id", ascending=False)

        # Format response based on requested format
        if format.lower() == "excel":
            output = BytesIO()

            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                pivot_df.to_excel(writer, sheet_name="Sheet1", index=False)

                workbook = writer.book
                worksheet = writer.sheets["Sheet1"]

                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            output.seek(0)

            filename = "amazon_settlements_pivot"
            if start_date:
                filename += f"_{start_date}"
            if end_date:
                filename += f"_to_{end_date}"
            filename += ".xlsx"

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"},
            )

        else:
            # Return as JSON
            result = pivot_df.to_dict(orient="records")

            return {
                "success": True,
                "count": len(result),
                "filters": {
                    "start_date": start_date,
                    "end_date": end_date,
                    "sku": sku,
                    "order_id": order_id,
                },
                "data": result,
            }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating settlements pivot: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate settlements pivot: {str(e)}",
        )


@router.get("/settlements/pivot/columns")
def get_pivot_columns(db=Depends(get_database)):
    """
    Get list of all unique amount_description values (column names in pivot table).

    This is useful for understanding what charge types exist in your data.
    """
    try:
        collection = db["amazon_settlements"]

        # Get distinct amount_description values
        columns = collection.distinct("amount_description")

        # Remove None values
        columns = [col for col in columns if col is not None]

        # Sort alphabetically
        columns.sort()

        return {"success": True, "count": len(columns), "columns": columns}

    except Exception as e:
        logger.error(f"Error fetching pivot columns: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch pivot columns: {str(e)}",
        )


@router.get("/settlements/summary")
def get_settlements_summary(
    start_date: Optional[str] = Query(
        None, description="Start date in YYYY-MM-DD format"
    ),
    end_date: Optional[str] = Query(None, description="End date in YYYY-MM-DD format"),
    group_by: str = Query(
        "amount_description",
        description="Group by: 'amount_description', 'sku', or 'date'",
    ),
    db=Depends(get_database),
):
    """
    Get summary statistics for settlements data.

    Query Parameters:
    - start_date: Filter by posted date (start)
    - end_date: Filter by posted date (end)
    - group_by: How to group the summary - 'amount_description', 'sku', or 'date'

    Returns summary totals grouped by the specified field.
    """
    try:
        collection = db["amazon_settlements"]

        # Build match stage
        match_stage = {}

        # Step 1: If date range is specified, first find all order_ids within that range
        if start_date or end_date:
            date_query = {}
            if start_date:
                start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                date_query["$gte"] = start_dt
            if end_date:
                end_dt = datetime.strptime(end_date, "%Y-%m-%d")
                date_query["$lte"] = end_dt

            # Find all order_ids that have at least one record in the date range
            orders_in_range = collection.distinct("order_id", {"posted_date": date_query})

            # Now match ALL records for those orders (regardless of date)
            match_stage["order_id"] = {"$in": orders_in_range}

        # Build aggregation pipeline

        group_field_map = {
            "amount_description": "$amount_description",
            "sku": "$sku",
            "date": {"$dateToString": {"format": "%Y-%m-%d", "date": "$posted_date"}},
        }

        if group_by not in group_field_map:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid group_by value. Must be one of: {', '.join(group_field_map.keys())}",
            )

        pipeline = []

        if match_stage:
            pipeline.append({"$match": match_stage})

        pipeline.extend(
            [
                {
                    "$group": {
                        "_id": group_field_map[group_by],
                        "total_amount": {"$sum": "$amount"},
                        "count": {"$sum": 1},
                        "avg_amount": {"$avg": "$amount"},
                    }
                },
                {"$sort": {"total_amount": -1}},
            ]
        )

        results = list(collection.aggregate(pipeline))

        # Format results
        formatted_results = []
        for result in results:
            formatted_results.append(
                {
                    group_by: result["_id"],
                    "total_amount": (
                        round(result["total_amount"], 2)
                        if result["total_amount"]
                        else 0
                    ),
                    "count": result["count"],
                    "average_amount": (
                        round(result["avg_amount"], 2) if result["avg_amount"] else 0
                    ),
                }
            )

        return {
            "success": True,
            "group_by": group_by,
            "filters": {"start_date": start_date, "end_date": end_date},
            "count": len(formatted_results),
            "data": formatted_results,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating settlements summary: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate settlements summary: {str(e)}",
        )


# ---------------------------------------------------------------------------
# Seller Flex Returns CSV Upload
# ---------------------------------------------------------------------------

SELLER_FLEX_RECON_DATE_FORMAT = "%d-%b-%Y"  # e.g. "01-Feb-2026"


def _parse_recon_date(val: Any) -> Optional[datetime]:
    if not val or (isinstance(val, float) and pd.isna(val)):
        return None
    s = str(val).strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, SELLER_FLEX_RECON_DATE_FORMAT)
    except ValueError:
        return None


def _str_or_none(val: Any) -> Optional[str]:
    """Return stripped string, or None for empty/NaN values."""
    if val is None:
        return None
    if isinstance(val, float) and pd.isna(val):
        return None
    s = str(val).strip()
    return s if s else None


def normalize_seller_flex_recon_row(row: Dict) -> Dict:
    """Map CSV column names to MongoDB-friendly field names."""
    units_raw = row.get("Units")
    units = None
    if units_raw is not None and not (isinstance(units_raw, float) and pd.isna(units_raw)):
        try:
            units = int(float(str(units_raw)))
        except (ValueError, TypeError):
            units = None

    return {
        "return_type": _str_or_none(row.get("Return Type")),
        "customer_order_id": _str_or_none(row.get("Customer Order ID")),
        "shipment_id": _str_or_none(row.get("Shipment ID")),
        "sku": _str_or_none(row.get("SKU")),
        "msku": _str_or_none(row.get("mSKU")),
        "asin": _str_or_none(row.get("ASIN")),
        "external_id1": _str_or_none(row.get("External ID1")),
        "external_id2": _str_or_none(row.get("External ID2")),
        "external_id3": _str_or_none(row.get("External ID3")),
        "units": units,
        "forward_leg_tracking_id": _str_or_none(row.get("Forward Leg Tracking ID")),
        "reverse_leg_tracking_id": _str_or_none(row.get("Reverse Leg Tracking ID")),
        "rma_id": _str_or_none(row.get("RMA ID")),
        "return_status": _str_or_none(row.get("Return Status")),
        "carrier": _str_or_none(row.get("Carrier")),
        "pickup_date": _parse_recon_date(row.get("Pick -up date")),
        "last_updated_on": _parse_recon_date(row.get("Last Updated On")),
        "returned_with_otp": _str_or_none(row.get("Returned with OTP")),
        "days_in_transit": _str_or_none(row.get("Days In-transit")),
        "days_since_return_complete": _str_or_none(row.get("Days Since Return Complete")),
        "return_reason": _str_or_none(row.get("Return Reason")),
        "lpn": _str_or_none(row.get("LPN")),
        "created_at": datetime.utcnow(),
    }


SELLER_FLEX_EDITABLE_FIELDS = {
    "received_date", "condition", "qty_received_wh", "gdrive_link",
    "entry_in_zoho", "transfer_orders_inventory_adj", "safe_t_claim_raise",
}


def _clean_flex_doc(doc: Dict) -> Dict:
    import math
    out = {}
    for k, v in doc.items():
        if k == "_id":
            out["_id"] = str(v)
        elif isinstance(v, datetime):
            out[k] = v.strftime("%Y-%m-%d")
        elif isinstance(v, float) and math.isnan(v):
            out[k] = None
        else:
            out[k] = v
    return out


async def _enrich_flex_with_product_names(docs: List[Dict], db) -> List[Dict]:
    """Add product_name via amazon_sku_mapping.item_id (= ASIN) → item_name."""
    asins = {d.get("asin") for d in docs if d.get("asin")}
    if not asins:
        for doc in docs:
            doc["product_name"] = ""
        return docs

    sku_map_col = db[SKU_COLLECTION]
    sku_docs = await asyncio.to_thread(
        lambda: list(sku_map_col.find(
            {"item_id": {"$in": list(asins)}},
            {"item_id": 1, "item_name": 1, "_id": 0}
        ))
    )
    asin_to_name = {d["item_id"]: d.get("item_name", "") for d in sku_docs}

    for doc in docs:
        doc["product_name"] = asin_to_name.get(doc.get("asin"), "")

    return docs


@router.get("/seller-flex-returns")
async def get_seller_flex_returns(
    start_date: str = Query(..., description="YYYY-MM-DD"),
    end_date: str = Query(..., description="YYYY-MM-DD"),
    db=Depends(get_database),
):
    """Return seller flex recon records filtered by pickup_date range."""
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    query = {"pickup_date": {"$gte": start, "$lte": end}}
    collection = db[SELLER_FLEX_RETURNS_COLLECTION]
    docs = await asyncio.to_thread(
        lambda: list(collection.find(query, {"created_at": 0}).sort("pickup_date", 1))
    )

    docs = [_clean_flex_doc(d) for d in docs]
    docs = await _enrich_flex_with_product_names(docs, db)
    return JSONResponse(content={"records": docs, "total": len(docs)})


SELLER_FLEX_COLUMN_ORDER = [
    "return_type", "customer_order_id", "shipment_id", "sku", "msku", "asin",
    "product_name",
    "external_id1", "external_id2", "external_id3", "units",
    "forward_leg_tracking_id", "reverse_leg_tracking_id", "rma_id",
    "return_status", "carrier", "pickup_date", "last_updated_on",
    "returned_with_otp", "days_in_transit", "days_since_return_complete",
    "return_reason", "lpn",
    # editable columns
    "received_date", "condition", "qty_received_wh", "gdrive_link",
    "entry_in_zoho", "transfer_orders_inventory_adj", "safe_t_claim_raise",
]

SELLER_FLEX_HEADER_MAP = {
    "return_type": "Return Type", "customer_order_id": "Customer Order ID",
    "shipment_id": "Shipment ID", "sku": "SKU", "msku": "mSKU", "asin": "ASIN",
    "product_name": "Product Name",
    "external_id1": "External ID1", "external_id2": "External ID2",
    "external_id3": "External ID3", "units": "Units",
    "forward_leg_tracking_id": "Forward Leg Tracking ID",
    "reverse_leg_tracking_id": "Reverse Leg Tracking ID",
    "rma_id": "RMA ID", "return_status": "Return Status", "carrier": "Carrier",
    "pickup_date": "Pick-up Date", "last_updated_on": "Last Updated On",
    "returned_with_otp": "Returned with OTP", "days_in_transit": "Days In-transit",
    "days_since_return_complete": "Days Since Return Complete",
    "return_reason": "Return Reason", "lpn": "LPN",
    "received_date": "Received Date of Return",
    "condition": "Condition of Product",
    "qty_received_wh": "Qty Received in WH",
    "gdrive_link": "GDrive Link (Images/CCTV)",
    "entry_in_zoho": "Entry in Zoho",
    "transfer_orders_inventory_adj": "Transfer Orders / Inventory Adjustment",
    "safe_t_claim_raise": "Safe-t Claim Raise",
}


@router.get("/seller-flex-returns/download")
async def download_seller_flex_returns(
    start_date: str = Query(...),
    end_date: str = Query(...),
    db=Depends(get_database),
):
    """Download seller flex recon records as an XLSX file (filtered by pickup_date)."""
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    query = {"pickup_date": {"$gte": start, "$lte": end}}
    collection = db[SELLER_FLEX_RETURNS_COLLECTION]
    docs = await asyncio.to_thread(
        lambda: list(collection.find(query, {"created_at": 0}).sort("pickup_date", 1))
    )

    docs = [_clean_flex_doc(d) for d in docs]
    docs = await _enrich_flex_with_product_names(docs, db)

    rows = []
    for doc in docs:
        row = {}
        for col in SELLER_FLEX_COLUMN_ORDER:
            val = doc.get(col)
            if isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d")
            row[SELLER_FLEX_HEADER_MAP[col]] = val
        rows.append(row)

    df = pd.DataFrame(rows, columns=[SELLER_FLEX_HEADER_MAP[c] for c in SELLER_FLEX_COLUMN_ORDER])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Seller Flex Returns")
    buf.seek(0)

    filename = f"seller_flex_returns_{start_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/upload/seller-flex-returns")
async def upload_seller_flex_returns(
    file: UploadFile = File(...),
    db=Depends(get_database),
):
    """
    Upload a Seller Flex Returns Reconciliation CSV.
    Detects the date range from pickup_date / last_updated_on, deletes any
    existing records in that range, then inserts the uploaded records.
    """
    if not file.filename.endswith(".csv"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Please upload a CSV file.",
        )

    try:
        contents = await file.read()
        df = pd.read_csv(BytesIO(contents), low_memory=False, dtype=str)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse CSV: {e}",
        )

    if df.empty:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV file is empty.",
        )

    records = df.to_dict("records")
    normalized: List[Dict] = [normalize_seller_flex_recon_row(r) for r in records]

    # Determine date range from pickup_date (fallback to last_updated_on)
    all_dates: List[datetime] = []
    for rec in normalized:
        d = rec.get("pickup_date") or rec.get("last_updated_on")
        if d:
            all_dates.append(d)

    if not all_dates:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No valid dates found in the CSV. Expected 'Pick -up date' or 'Last Updated On' in DD-Mon-YYYY format.",
        )

    range_start = min(all_dates)
    range_end = max(all_dates)

    collection = db[SELLER_FLEX_RETURNS_COLLECTION]

    # Check for existing records in this date range
    date_query = {
        "$or": [
            {"pickup_date": {"$gte": range_start, "$lte": range_end}},
            {"last_updated_on": {"$gte": range_start, "$lte": range_end}},
        ]
    }
    existing_count = await asyncio.to_thread(collection.count_documents, date_query)

    deleted_count = 0
    if existing_count > 0:
        delete_result = await asyncio.to_thread(collection.delete_many, date_query)
        deleted_count = delete_result.deleted_count
        logger.info(f"Deleted {deleted_count} existing seller flex recon records for range {range_start} – {range_end}")

    insert_result = await asyncio.to_thread(collection.insert_many, normalized)
    inserted_count = len(insert_result.inserted_ids)
    logger.info(f"Inserted {inserted_count} seller flex recon records")

    return JSONResponse(
        status_code=status.HTTP_200_OK,
        content={
            "message": "Seller Flex returns uploaded successfully.",
            "date_range": {
                "start": range_start.strftime("%Y-%m-%d"),
                "end": range_end.strftime("%Y-%m-%d"),
            },
            "existing_deleted": deleted_count,
            "records_inserted": inserted_count,
        },
    )


@router.patch("/seller-flex-returns/{record_id}")
async def update_seller_flex_return(
    record_id: str,
    payload: Dict[str, Any],
    db=Depends(get_database),
):
    """Update editable fields on a single seller flex return record."""
    try:
        oid = ObjectId(record_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid record ID")

    update_data = {k: v for k, v in payload.items() if k in SELLER_FLEX_EDITABLE_FIELDS}
    if not update_data:
        raise HTTPException(status_code=400, detail="No valid editable fields provided")

    # Parse received_date string → datetime if provided
    if "received_date" in update_data and update_data["received_date"]:
        raw = str(update_data["received_date"]).split(".")[0]
        parsed = None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y"):
            try:
                parsed = datetime.strptime(raw, fmt)
                break
            except ValueError:
                continue
        if parsed is None:
            raise HTTPException(status_code=400, detail="received_date must be YYYY-MM-DD")
        update_data["received_date"] = parsed

    collection = db[SELLER_FLEX_RETURNS_COLLECTION]
    result = await asyncio.to_thread(
        collection.update_one, {"_id": oid}, {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Record not found")

    return JSONResponse(content={"message": "Record updated", "modified": result.modified_count})


@router.post("/seller-flex-returns/bulk-update")
async def bulk_update_seller_flex_returns(
    file: UploadFile = File(...),
    db=Depends(get_database),
):
    """
    Upload an XLSX exported from the seller flex returns download to bulk-update
    the editable columns (Received Date, Condition, Qty Received, GDrive Link,
    Entry in Zoho, Transfer Orders, Safe-t Claim). Rows are matched by
    Customer Order ID + SKU.
    """
    if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="Please upload an XLSX file.")

    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents), dtype=str)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    if df.empty:
        raise HTTPException(status_code=400, detail="File is empty.")

    # Reverse-map header → field
    header_to_field = {v: k for k, v in SELLER_FLEX_HEADER_MAP.items()}
    df.rename(columns=header_to_field, inplace=True)

    required = {"customer_order_id", "sku"}
    missing = required - set(df.columns)
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {missing}")

    editable_present = [f for f in SELLER_FLEX_EDITABLE_FIELDS if f in df.columns]
    if not editable_present:
        raise HTTPException(status_code=400, detail="No editable columns found in the file.")

    collection = db[SELLER_FLEX_RETURNS_COLLECTION]
    updated = 0
    for _, row in df.iterrows():
        order_id = str(row.get("customer_order_id") or "").strip()
        sku = str(row.get("sku") or "").strip()
        if not order_id or not sku or order_id == "nan" or sku == "nan":
            continue

        update_data = {}
        for field in editable_present:
            val = row.get(field)
            if pd.isna(val) if isinstance(val, float) else (val is None or str(val).strip() == "nan"):
                continue
            val = str(val).strip()
            if not val:
                continue
            if field == "received_date":
                parsed = None
                for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y"):
                    try:
                        parsed = datetime.strptime(val.split(".")[0], fmt)
                        break
                    except ValueError:
                        continue
                if parsed is None:
                    continue
                val = parsed
            elif field == "qty_received_wh":
                try:
                    val = int(float(val))
                except (ValueError, TypeError):
                    continue
            update_data[field] = val

        if not update_data:
            continue

        result = await asyncio.to_thread(
            collection.update_many,
            {"customer_order_id": order_id, "sku": sku},
            {"$set": update_data},
        )
        updated += result.modified_count

    return JSONResponse(content={"message": "Bulk update complete", "records_updated": updated})


# ---------------------------------------------------------------------------
# FBA Returns – read-only endpoints (data synced via SP-API elsewhere)
# ---------------------------------------------------------------------------

def _clean_doc(doc: Dict) -> Dict:
    import math
    out = {}
    for k, v in doc.items():
        if isinstance(v, datetime):
            out[k] = v.strftime("%Y-%m-%d")
        elif isinstance(v, float) and math.isnan(v):
            out[k] = None
        else:
            out[k] = v
    return out


@router.get("/fba-returns")
async def get_fba_returns(
    start_date: str = Query(..., description="YYYY-MM-DD"),
    end_date: str = Query(..., description="YYYY-MM-DD"),
    db=Depends(get_database),
):
    """Return FBA returns records for the given date range (filtered on return_date)."""
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    query = {"return_date": {"$gte": start, "$lte": end}}
    collection = db[FBA_RETURNS_COLLECTION]
    docs = await asyncio.to_thread(
        lambda: list(collection.find(query, {"_id": 0, "created_at": 0}).sort("return_date", 1))
    )
    docs = [_clean_doc(d) for d in docs]
    return JSONResponse(content={"records": docs, "total": len(docs)})


@router.get("/fba-returns/download")
async def download_fba_returns(
    start_date: str = Query(...),
    end_date: str = Query(...),
    db=Depends(get_database),
):
    """Download FBA returns records as an XLSX file."""
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    query = {"return_date": {"$gte": start, "$lte": end}}
    collection = db[FBA_RETURNS_COLLECTION]
    docs = await asyncio.to_thread(
        lambda: list(collection.find(query, {"_id": 0, "created_at": 0}).sort("return_date", 1))
    )

    COLUMN_ORDER = [
        "return_date", "amazon_order_id", "sku_code", "asin", "fnsku",
        "product_name", "quantity", "fulfillment_center_id",
        "detailed_disposition", "reason", "license_plate_number", "customer_comments",
    ]
    HEADER_MAP = {
        "return_date": "Return Date", "amazon_order_id": "Order ID",
        "sku_code": "SKU", "asin": "ASIN", "fnsku": "FNSKU",
        "product_name": "Product Name", "quantity": "Quantity",
        "fulfillment_center_id": "FC ID", "detailed_disposition": "Disposition",
        "reason": "Reason", "license_plate_number": "LPN",
        "customer_comments": "Customer Comments",
    }

    rows = []
    for doc in docs:
        row = {}
        for col in COLUMN_ORDER:
            val = doc.get(col)
            if isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d")
            row[HEADER_MAP[col]] = val
        rows.append(row)

    df = pd.DataFrame(rows, columns=[HEADER_MAP[c] for c in COLUMN_ORDER])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="FBA Returns")
    buf.seek(0)

    filename = f"fba_returns_{start_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# FBA Warehouse Returns – CSV upload + editable tracking
# ---------------------------------------------------------------------------

FBA_WAREHOUSE_RETURNS_COLLECTION = "amazon_fba_warehouse_returns"

FBA_WAREHOUSE_EDITABLE_FIELDS = {
    "received_date", "condition", "qty_received_wh", "gdrive_link",
    "entry_in_zoho", "transfer_orders_inventory_adj", "safe_t_claim_raise",
}


def _parse_fba_wh_date(val: Any) -> Optional[datetime]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, float) and pd.isna(val):
        return None
    try:
        if hasattr(val, "to_pydatetime"):
            return val.to_pydatetime()
    except Exception:
        pass
    s = str(val).strip()
    if not s:
        return None
    # ISO timestamps from Amazon e.g. "2026-03-04T05:11:34+01:00"
    try:
        from dateutil import parser as dateutil_parser
        return dateutil_parser.parse(s).replace(tzinfo=None)
    except Exception:
        pass
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt)
        except ValueError:
            continue
    return None


def _num_or_none_wh(val: Any) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, float) and pd.isna(val):
        return None
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return None


def normalize_fba_warehouse_return(row: Dict) -> Dict:
    return {
        "request_date": _parse_fba_wh_date(row.get("request-date")),
        "order_id": _str_or_none(row.get("order-id")),
        "order_source": _str_or_none(row.get("order-source")),
        "order_type": _str_or_none(row.get("order-type")),
        "service_speed": _str_or_none(row.get("service-speed")),
        "order_status": _str_or_none(row.get("order-status")),
        "last_updated_date": _parse_fba_wh_date(row.get("last-updated-date")),
        "sku": _str_or_none(row.get("sku")),
        "fnsku": _str_or_none(row.get("fnsku")),
        "disposition": _str_or_none(row.get("disposition")),
        "requested_quantity": _num_or_none_wh(row.get("requested-quantity")),
        "cancelled_quantity": _num_or_none_wh(row.get("cancelled-quantity")),
        "disposed_quantity": _num_or_none_wh(row.get("disposed-quantity")),
        "shipped_quantity": _num_or_none_wh(row.get("shipped-quantity")),
        "in_process_quantity": _num_or_none_wh(row.get("in-process-quantity")),
        "removal_fee": _num_or_none_wh(row.get("removal-fee")),
        "currency": _str_or_none(row.get("currency")),
        "created_at": datetime.utcnow(),
    }


def _clean_fba_wh_doc(doc: Dict) -> Dict:
    import math
    out = {}
    for k, v in doc.items():
        if k == "_id":
            out["_id"] = str(v)
        elif isinstance(v, datetime):
            out[k] = v.strftime("%Y-%m-%d")
        elif isinstance(v, float) and math.isnan(v):
            out[k] = None
        else:
            out[k] = v
    return out


async def _enrich_fba_wh_with_product_names(docs: List[Dict], db) -> List[Dict]:
    """Look up product names from 'products' collection by cf_sku_code."""
    skus = {d.get("sku") for d in docs if d.get("sku")}
    if not skus:
        for doc in docs:
            doc["product_name"] = ""
        return docs

    products_col = db["products"]
    product_docs = await asyncio.to_thread(
        lambda: list(products_col.find(
            {"cf_sku_code": {"$in": list(skus)}},
            {"cf_sku_code": 1, "name": 1, "_id": 0}
        ))
    )
    sku_to_name = {d["cf_sku_code"]: d.get("name", "") for d in product_docs}

    for doc in docs:
        doc["product_name"] = sku_to_name.get(doc.get("sku"), "")

    return docs


FBA_WAREHOUSE_COLUMN_ORDER = [
    "request_date", "order_id", "order_source", "order_type", "service_speed",
    "order_status", "last_updated_date", "sku", "fnsku", "product_name",
    "disposition", "requested_quantity", "cancelled_quantity", "disposed_quantity",
    "shipped_quantity", "in_process_quantity", "removal_fee", "currency",
    # editable columns
    "received_date", "condition", "qty_received_wh", "gdrive_link",
    "entry_in_zoho", "transfer_orders_inventory_adj", "safe_t_claim_raise",
]

FBA_WAREHOUSE_HEADER_MAP = {
    "request_date": "Request Date",
    "order_id": "Order ID",
    "order_source": "Order Source",
    "order_type": "Order Type",
    "service_speed": "Service Speed",
    "order_status": "Order Status",
    "last_updated_date": "Last Updated Date",
    "sku": "SKU",
    "fnsku": "FNSKU",
    "product_name": "Product Name",
    "disposition": "Disposition",
    "requested_quantity": "Requested Qty",
    "cancelled_quantity": "Cancelled Qty",
    "disposed_quantity": "Disposed Qty",
    "shipped_quantity": "Shipped Qty",
    "in_process_quantity": "In-Process Qty",
    "removal_fee": "Removal Fee",
    "currency": "Currency",
    "received_date": "Received Date of Return",
    "condition": "Condition of Product",
    "qty_received_wh": "Qty Received in WH",
    "gdrive_link": "GDrive Link (Images/CCTV)",
    "entry_in_zoho": "Entry in Zoho",
    "transfer_orders_inventory_adj": "Transfer Orders / Inventory Adjustment",
    "safe_t_claim_raise": "Safe-t Claim Raise",
}


@router.post("/upload/fba-warehouse-returns")
async def upload_fba_warehouse_returns(
    file: UploadFile = File(...),
    db=Depends(get_database),
):
    """Upload an Amazon FBA removal/return order CSV."""
    if not file.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="Please upload a CSV file.")

    try:
        contents = await file.read()
        df = pd.read_csv(BytesIO(contents), low_memory=False, dtype=str)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse CSV: {e}")

    if df.empty:
        raise HTTPException(status_code=400, detail="CSV file is empty.")

    records = df.to_dict("records")
    normalized: List[Dict] = [normalize_fba_warehouse_return(r) for r in records]

    all_dates: List[datetime] = [r["request_date"] for r in normalized if r.get("request_date")]
    if not all_dates:
        raise HTTPException(
            status_code=400,
            detail="No valid dates found in 'request-date' column.",
        )

    range_start = min(all_dates)
    range_end = max(all_dates)

    collection = db[FBA_WAREHOUSE_RETURNS_COLLECTION]
    date_query = {"request_date": {"$gte": range_start, "$lte": range_end}}
    existing_count = await asyncio.to_thread(collection.count_documents, date_query)

    deleted_count = 0
    if existing_count > 0:
        delete_result = await asyncio.to_thread(collection.delete_many, date_query)
        deleted_count = delete_result.deleted_count

    insert_result = await asyncio.to_thread(collection.insert_many, normalized)
    inserted_count = len(insert_result.inserted_ids)

    return JSONResponse(
        status_code=status.HTTP_200_OK,
        content={
            "message": "FBA warehouse returns uploaded successfully.",
            "date_range": {
                "start": range_start.strftime("%Y-%m-%d"),
                "end": range_end.strftime("%Y-%m-%d"),
            },
            "existing_deleted": deleted_count,
            "records_inserted": inserted_count,
        },
    )


@router.get("/fba-warehouse-returns")
async def get_fba_warehouse_returns(
    start_date: str = Query(..., description="YYYY-MM-DD"),
    end_date: str = Query(..., description="YYYY-MM-DD"),
    db=Depends(get_database),
):
    """Return FBA warehouse return records filtered by request_date."""
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    collection = db[FBA_WAREHOUSE_RETURNS_COLLECTION]
    docs = await asyncio.to_thread(
        lambda: list(collection.find(
            {"request_date": {"$gte": start, "$lte": end}},
            {"created_at": 0}
        ).sort("request_date", 1))
    )
    docs = [_clean_fba_wh_doc(d) for d in docs]
    docs = await _enrich_fba_wh_with_product_names(docs, db)
    return JSONResponse(content={"records": docs, "total": len(docs)})


@router.get("/fba-warehouse-returns/download")
async def download_fba_warehouse_returns(
    start_date: str = Query(...),
    end_date: str = Query(...),
    db=Depends(get_database),
):
    """Download FBA warehouse returns as XLSX with editable columns included."""
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    collection = db[FBA_WAREHOUSE_RETURNS_COLLECTION]
    docs = await asyncio.to_thread(
        lambda: list(collection.find(
            {"request_date": {"$gte": start, "$lte": end}},
            {"created_at": 0}
        ).sort("request_date", 1))
    )
    docs = [_clean_fba_wh_doc(d) for d in docs]
    docs = await _enrich_fba_wh_with_product_names(docs, db)

    rows = []
    for doc in docs:
        row = {}
        for col in FBA_WAREHOUSE_COLUMN_ORDER:
            val = doc.get(col)
            if isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d")
            row[FBA_WAREHOUSE_HEADER_MAP[col]] = val
        rows.append(row)

    df = pd.DataFrame(rows, columns=[FBA_WAREHOUSE_HEADER_MAP[c] for c in FBA_WAREHOUSE_COLUMN_ORDER])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="FBA Warehouse Returns")
    buf.seek(0)

    filename = f"fba_warehouse_returns_{start_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.patch("/fba-warehouse-returns/{record_id}")
async def update_fba_warehouse_return(
    record_id: str,
    payload: Dict[str, Any],
    db=Depends(get_database),
):
    """Update editable fields on a single FBA warehouse return record."""
    try:
        oid = ObjectId(record_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid record ID")

    update_data = {k: v for k, v in payload.items() if k in FBA_WAREHOUSE_EDITABLE_FIELDS}
    if not update_data:
        raise HTTPException(status_code=400, detail="No valid editable fields provided")

    if "received_date" in update_data and update_data["received_date"]:
        raw = str(update_data["received_date"]).split(".")[0]
        parsed = None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y"):
            try:
                parsed = datetime.strptime(raw, fmt)
                break
            except ValueError:
                continue
        if parsed is None:
            raise HTTPException(status_code=400, detail="received_date must be YYYY-MM-DD")
        update_data["received_date"] = parsed

    collection = db[FBA_WAREHOUSE_RETURNS_COLLECTION]
    result = await asyncio.to_thread(
        collection.update_one, {"_id": oid}, {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Record not found")

    return JSONResponse(content={"message": "Record updated", "modified": result.modified_count})


@router.post("/fba-warehouse-returns/bulk-update")
async def bulk_update_fba_warehouse_returns(
    file: UploadFile = File(...),
    db=Depends(get_database),
):
    """
    Upload the downloaded XLSX and bulk-update editable columns.
    Rows are matched by Order ID + SKU.
    """
    if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="Please upload an XLSX file.")

    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents), dtype=str)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    if df.empty:
        raise HTTPException(status_code=400, detail="File is empty.")

    header_to_field = {v: k for k, v in FBA_WAREHOUSE_HEADER_MAP.items()}
    df.rename(columns=header_to_field, inplace=True)

    required = {"order_id", "sku"}
    missing = required - set(df.columns)
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {missing}")

    editable_present = [f for f in FBA_WAREHOUSE_EDITABLE_FIELDS if f in df.columns]
    if not editable_present:
        raise HTTPException(status_code=400, detail="No editable columns found in the file.")

    collection = db[FBA_WAREHOUSE_RETURNS_COLLECTION]
    updated = 0
    for _, row in df.iterrows():
        order_id = str(row.get("order_id") or "").strip()
        sku = str(row.get("sku") or "").strip()
        if not order_id or not sku or order_id == "nan" or sku == "nan":
            continue

        update_data = {}
        for field in editable_present:
            val = row.get(field)
            if pd.isna(val) if isinstance(val, float) else (val is None or str(val).strip() == "nan"):
                continue
            val = str(val).strip()
            if not val:
                continue
            if field == "received_date":
                parsed = None
                for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y"):
                    try:
                        parsed = datetime.strptime(val.split(".")[0], fmt)
                        break
                    except ValueError:
                        continue
                if parsed is None:
                    continue
                val = parsed
            elif field == "qty_received_wh":
                try:
                    val = int(float(val))
                except (ValueError, TypeError):
                    continue
            update_data[field] = val

        if not update_data:
            continue

        result = await asyncio.to_thread(
            collection.update_many,
            {"order_id": order_id, "sku": sku},
            {"$set": update_data},
        )
        updated += result.modified_count

    return JSONResponse(content={"message": "Bulk update complete", "records_updated": updated})


# ---------------------------------------------------------------------------
# Vendor Central Returns – upload XLSX + read endpoints
# ---------------------------------------------------------------------------

def _parse_vc_date(val: Any) -> Optional[datetime]:
    """Parse dates from Vendor Central returns report (DD/MM/YYYY or datetime objects)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, float) and pd.isna(val):
        return None
    # pandas may give a pd.Timestamp
    try:
        if hasattr(val, "to_pydatetime"):
            return val.to_pydatetime()
    except Exception:
        pass
    s = str(val).strip()
    if not s:
        return None
    # Strip trailing time component if present (e.g. "2025-03-15 00:00:00")
    s_date_only = s[:10]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s_date_only, fmt)
        except ValueError:
            continue
    # Try full string as fallback
    for fmt in ("%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _num_or_none(val: Any) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, float) and pd.isna(val):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _ci_get(row: Dict, *keys: str) -> Any:
    """Case-insensitive dict lookup, tries each key in order."""
    lower = {k.lower(): v for k, v in row.items()}
    for key in keys:
        val = lower.get(key.lower())
        if val is not None:
            return val
    return None


def normalize_vendor_central_return_row(row: Dict) -> Dict:
    return {
        "shipment_id": _str_or_none(_ci_get(row, "Shipment ID")),
        "return_id": _str_or_none(_ci_get(row, "Return ID")),
        "vendor_code": _str_or_none(_ci_get(row, "Vendor Code", "Vendor code")),
        "return_date": _parse_vc_date(_ci_get(row, "Return Date", "Return date")),
        "purchase_order": _str_or_none(_ci_get(row, "Purchase Order", "Purchase order")),
        "warehouse": _str_or_none(_ci_get(row, "Warehouse")),
        "asin": _str_or_none(_ci_get(row, "ASIN")),
        "product": _str_or_none(_ci_get(row, "Product")),
        "quantity": _num_or_none(_ci_get(row, "Quantity")),
        # CSV export uses "Cost per unit"; XLSX uses "Price per unit"
        "price_per_unit": _num_or_none(_ci_get(row, "Price per unit", "Cost per unit")),
        # CSV export uses "Total cost"; XLSX uses "Net Amount"
        "net_amount": _num_or_none(_ci_get(row, "Net Amount", "Total cost")),
        "currency_code": _str_or_none(_ci_get(row, "Currency Code", "Currency code")),
        "ship_to_state": _str_or_none(_ci_get(row, "Ship To State", "Ship to state")),
        "ship_from_state": _str_or_none(_ci_get(row, "Ship From State", "Ship from state")),
        "system_ref_no": _str_or_none(_ci_get(row, "System Ref No")),
        "document_number": _str_or_none(_ci_get(row, "Document Number")),
        "document_date": _parse_vc_date(_ci_get(row, "Document Date")),
        "document_type": _str_or_none(_ci_get(row, "Document Type")),
        "original_document_number": _str_or_none(_ci_get(row, "Original Document Number")),
        "original_document_date": _parse_vc_date(_ci_get(row, "Original Document Date")),
        "hsn": _str_or_none(_ci_get(row, "HSN")),
        "vendor_invoice": _str_or_none(_ci_get(row, "Vendor Invoice")),
        "vendor_invoice_date": _parse_vc_date(_ci_get(row, "Vendor Invoice Date")),
        "igst_tax_rate": _num_or_none(_ci_get(row, "IGST Tax Rate")),
        "igst_tax_amount": _num_or_none(_ci_get(row, "IGST Tax Amount")),
        "cess_tax_rate": _num_or_none(_ci_get(row, "CESS Tax Rate")),
        "cess_tax_amount": _num_or_none(_ci_get(row, "CESS Tax Amount")),
        "cgst_tax_rate": _num_or_none(_ci_get(row, "CGST Tax Rate")),
        "cgst_tax_amount": _num_or_none(_ci_get(row, "CGST Tax Amount")),
        "sgst_tax_rate": _num_or_none(_ci_get(row, "SGST Tax Rate")),
        "sgst_tax_amount": _num_or_none(_ci_get(row, "SGST Tax Amount")),
        "total_amount": _num_or_none(_ci_get(row, "Total Amount", "Total amount")),
        "created_at": datetime.utcnow(),
    }


@router.get("/vendor-central-returns")
async def get_vendor_central_returns(
    start_date: str = Query(..., description="YYYY-MM-DD"),
    end_date: str = Query(..., description="YYYY-MM-DD"),
    db=Depends(get_database),
):
    """Return Vendor Central returns records for the given date range (filtered on document_date)."""
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    query = {"document_date": {"$gte": start, "$lte": end}}
    collection = db[VENDOR_CENTRAL_RETURNS_COLLECTION]
    docs = await asyncio.to_thread(
        lambda: list(collection.find(query, {"created_at": 0}).sort("document_date", 1))
    )
    cleaned = []
    for d in docs:
        cd = _clean_doc(d)
        oid = d.get("_id")
        if oid is not None:
            cd["id"] = str(oid)
        cd.pop("_id", None)
        cleaned.append(cd)
    # sent_to_accounts_team=False rows first, then others
    cleaned.sort(key=lambda r: (0 if r.get("sent_to_accounts_team") is False else 1))
    return JSONResponse(content={"records": cleaned, "total": len(cleaned)})


@router.patch("/vendor-central-returns/{record_id}")
async def patch_vendor_central_return(
    record_id: str,
    payload: Dict,
    db=Depends(get_database),
):
    """Update the 3 editable fields on a Vendor Central return record."""
    allowed = {"entry_in_zoho", "transfer_orders_inventory_adjustment", "sent_to_accounts_team"}
    update_fields = {k: v for k, v in payload.items() if k in allowed}
    if not update_fields:
        raise HTTPException(status_code=400, detail="No valid fields to update.")

    try:
        oid = ObjectId(record_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid record ID.")

    collection = db[VENDOR_CENTRAL_RETURNS_COLLECTION]
    result = await asyncio.to_thread(
        collection.update_one, {"_id": oid}, {"$set": update_fields}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Record not found.")
    return JSONResponse(content={"updated": True})


@router.get("/vendor-central-returns/download")
async def download_vendor_central_returns(
    start_date: str = Query(...),
    end_date: str = Query(...),
    db=Depends(get_database),
):
    """Download Vendor Central returns records as an XLSX file."""
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    query = {"document_date": {"$gte": start, "$lte": end}}
    collection = db[VENDOR_CENTRAL_RETURNS_COLLECTION]
    docs = await asyncio.to_thread(
        lambda: list(collection.find(query, {"_id": 0, "created_at": 0}).sort("document_date", 1))
    )
    docs.sort(key=lambda r: (0 if r.get("sent_to_accounts_team") is False else 1))

    COLUMN_ORDER = [
        "shipment_id", "return_id", "vendor_code", "return_date", "purchase_order",
        "warehouse", "asin", "product", "quantity", "price_per_unit", "net_amount",
        "currency_code", "ship_to_state", "ship_from_state", "system_ref_no",
        "document_number", "document_date", "document_type", "original_document_number",
        "original_document_date", "hsn", "vendor_invoice", "vendor_invoice_date",
        "igst_tax_rate", "igst_tax_amount", "cess_tax_rate", "cess_tax_amount",
        "cgst_tax_rate", "cgst_tax_amount", "sgst_tax_rate", "sgst_tax_amount",
        "total_amount", "entry_in_zoho", "transfer_orders_inventory_adjustment",
        "sent_to_accounts_team",
    ]
    HEADER_MAP = {
        "shipment_id": "Shipment ID", "return_id": "Return ID",
        "vendor_code": "Vendor Code", "return_date": "Return Date",
        "purchase_order": "Purchase Order", "warehouse": "Warehouse",
        "asin": "ASIN", "product": "Product", "quantity": "Quantity",
        "price_per_unit": "Price per unit", "net_amount": "Net Amount",
        "currency_code": "Currency Code", "ship_to_state": "Ship To State",
        "ship_from_state": "Ship From State", "system_ref_no": "System Ref No",
        "document_number": "Document Number", "document_date": "Document Date",
        "document_type": "Document Type",
        "original_document_number": "Original Document Number",
        "original_document_date": "Original Document Date",
        "hsn": "HSN", "vendor_invoice": "Vendor Invoice",
        "vendor_invoice_date": "Vendor Invoice Date",
        "igst_tax_rate": "IGST Tax Rate", "igst_tax_amount": "IGST Tax Amount",
        "cess_tax_rate": "CESS Tax Rate", "cess_tax_amount": "CESS Tax Amount",
        "cgst_tax_rate": "CGST Tax Rate", "cgst_tax_amount": "CGST Tax Amount",
        "sgst_tax_rate": "SGST Tax Rate", "sgst_tax_amount": "SGST Tax Amount",
        "total_amount": "Total Amount",
        "entry_in_zoho": "Entry in Zoho",
        "transfer_orders_inventory_adjustment": "Transfer Orders / Inventory Adjustment",
        "sent_to_accounts_team": "Sent to Accounts Team",
    }

    rows = []
    for doc in docs:
        row = {}
        for col in COLUMN_ORDER:
            val = doc.get(col)
            if isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d")
            row[HEADER_MAP[col]] = val
        rows.append(row)

    df = pd.DataFrame(rows, columns=[HEADER_MAP[c] for c in COLUMN_ORDER])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Vendor Central Returns")
    buf.seek(0)

    filename = f"vendor_central_returns_{start_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/upload/vendor-central-returns")
async def upload_vendor_central_returns(
    file: UploadFile = File(...),
    db=Depends(get_database),
):
    """
    Upload a Vendor Central Returns XLSX (Invoice Report sheet).
    Detects the date range from return_date, deletes existing records in that
    range, then inserts the uploaded records.
    """
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Please upload an Excel file (.xlsx or .xls).",
        )

    try:
        contents = await file.read()
        xf = pd.ExcelFile(BytesIO(contents))
        sheet_map = {s.strip().lower(): s for s in xf.sheet_names}
        matched_sheet = sheet_map.get("invoice report")
        if matched_sheet is None:
            raise ValueError(
                f"Worksheet named 'Invoice Report' not found. Available sheets: {xf.sheet_names}"
            )
        df = pd.read_excel(xf, sheet_name=matched_sheet, dtype=str)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse Excel file: {e}",
        )

    if df.empty:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file is empty.",
        )

    records = df.to_dict("records")
    normalized: List[Dict] = [normalize_vendor_central_return_row(r) for r in records]

    all_dates: List[datetime] = [r["document_date"] for r in normalized if r.get("document_date")]

    if not all_dates:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No valid dates found in the file. Expected a 'Return Date' or 'Return date' column.",
        )

    range_start = min(all_dates)
    range_end = max(all_dates)

    collection = db[VENDOR_CENTRAL_RETURNS_COLLECTION]

    def _replace_records():
        # Delete all existing records in the detected date range, then insert fresh
        collection.delete_many({
            "document_date": {"$gte": range_start, "$lte": range_end.replace(hour=23, minute=59, second=59)}
        })
        for rec in normalized:
            rec.setdefault("entry_in_zoho", None)
            rec.setdefault("transfer_orders_inventory_adjustment", None)
            rec.setdefault("sent_to_accounts_team", False)
        if normalized:
            collection.insert_many(normalized)
        return len(normalized)

    inserted_count = await asyncio.to_thread(_replace_records)
    logger.info(f"VC returns upload: {inserted_count} inserted")

    return JSONResponse(
        status_code=status.HTTP_200_OK,
        content={
            "message": "Vendor Central returns uploaded successfully.",
            "date_range": {
                "start": range_start.strftime("%Y-%m-%d"),
                "end": range_end.strftime("%Y-%m-%d"),
            },
            "records_inserted": inserted_count,
        },
    )


@router.post("/upload/vendor-central-returns-bulk-update")
async def bulk_update_vendor_central_returns(
    file: UploadFile = File(...),
    db=Depends(get_database),
):
    """
    Upload the downloaded Vendor Central Returns XLSX to bulk-update only the
    three editable columns: Entry in Zoho, Transfer Orders / Inventory Adjustment,
    Sent to Accounts Team. All other columns are ignored.
    Matches rows by Return ID (falls back to Document Number + ASIN).
    """
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls).")

    try:
        contents = await file.read()
        xf = pd.ExcelFile(BytesIO(contents))
        sheet_map = {s.strip().lower(): s for s in xf.sheet_names}
        matched_sheet = sheet_map.get("vendor central returns") or xf.sheet_names[0]
        df = pd.read_excel(xf, sheet_name=matched_sheet, dtype=str)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {e}")

    if df.empty:
        raise HTTPException(status_code=400, detail="Excel file is empty.")

    collection = db[VENDOR_CENTRAL_RETURNS_COLLECTION]

    def _bulk_update():
        updated = 0
        skipped = 0
        for _, row in df.iterrows():
            return_id = _str_or_none(row.get("Return ID"))
            doc_number = _str_or_none(row.get("Document Number"))
            asin = _str_or_none(row.get("ASIN"))

            if return_id:
                filt = {"return_id": return_id}
            elif doc_number and asin:
                filt = {"document_number": doc_number, "asin": asin}
            else:
                skipped += 1
                continue

            sent_raw = str(row.get("Sent to Accounts Team") or "").strip().lower()
            sent = True if sent_raw in ("true", "yes", "1") else (False if sent_raw in ("false", "no", "0", "") else None)

            patch = {
                "entry_in_zoho": _str_or_none(row.get("Entry in Zoho")),
                "transfer_orders_inventory_adjustment": _str_or_none(row.get("Transfer Orders / Inventory Adjustment")),
                "sent_to_accounts_team": sent if sent is not None else False,
            }

            result = collection.update_many(filt, {"$set": patch})
            if result.matched_count:
                updated += result.matched_count
            else:
                skipped += 1
        return updated, skipped

    updated_count, skipped_count = await asyncio.to_thread(_bulk_update)
    logger.info(f"VC returns bulk update: {updated_count} updated, {skipped_count} skipped")

    return JSONResponse(
        status_code=200,
        content={
            "message": "Bulk update completed.",
            "records_updated": updated_count,
            "records_skipped": skipped_count,
        },
    )
