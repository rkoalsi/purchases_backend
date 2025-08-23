from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
import io
import re
import logging
import pandas as pd
from functools import lru_cache
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import List, Dict
from ..database import get_database

router = APIRouter()
logger = logging.getLogger(__name__)


@router.get("/download_template")
async def download_template():
    """
    Generate and return an XLSX template file with CI and PL sheets
    """
    try:
        # Create a new workbook
        wb = Workbook()

        # Remove the default sheet
        wb.remove(wb.active)

        # Create CI sheet with specified columns
        ci_sheet = wb.create_sheet("CI")
        ci_sheet.append(["Name", "HSN", "Price"])

        # Optional: Set column widths for better formatting
        ci_sheet.column_dimensions["A"].width = 20  # Name column
        ci_sheet.column_dimensions["B"].width = 15  # HSN column
        ci_sheet.column_dimensions["C"].width = 15  # Price column

        # Create PL sheet with specified column
        pl_sheet = wb.create_sheet("PL")
        pl_sheet.append(["Name"])

        # Optional: Set column width for better formatting
        pl_sheet.column_dimensions["A"].width = 20  # Name column

        # Save workbook to bytes
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        # Generate filename with timestamp (optional)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Template_{timestamp}.xlsx"

        # Return as streaming response
        return StreamingResponse(
            io.BytesIO(file_stream.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error generating template: {str(e)}"
        )


def validate_file(file_content: bytes) -> dict:
    """
    Checks if the given Excel file contains both 'PL' and 'CI' sheets.

    Args:
        file_content (bytes): File content as bytes.

    Returns:
        dict: Response indicating whether the required sheets are present or not.
    """
    try:
        # Load the workbook from bytes
        wb = load_workbook(io.BytesIO(file_content))

        # Get all sheet names
        sheet_names = wb.sheetnames

        # Check if both 'PL' and 'CI' sheets are present
        if "PL" not in sheet_names or "CI" not in sheet_names:
            missing_sheets = []
            if "PL" not in sheet_names:
                missing_sheets.append("PL")
            if "CI" not in sheet_names:
                missing_sheets.append("CI")

            # Return an error response if any sheet is missing
            return {
                "status": "error",
                "message": f"Missing sheets: {', '.join(missing_sheets)}",
            }

        # If both sheets are found, return a success response
        return {"status": "success", "message": "Both PL and CI sheets are present."}

    except Exception as e:
        # Handle any other errors (e.g., invalid file format)
        return {"status": "error", "message": f"An error occurred: {str(e)}"}


def save_combined_sheet(matched_ci, unmatched_ci, matched_pl, unmatched_pl):
    """
    Saves four DataFrames to two sheets in a combined Excel file in memory.

    Args:
        matched_ci (pandas.DataFrame): DataFrame containing matched CI data.
        unmatched_ci (pandas.DataFrame): DataFrame containing unmatched CI data.
        matched_pl (pandas.DataFrame): DataFrame containing matched PL data.
        unmatched_pl (pandas.DataFrame): DataFrame containing unmatched PL data.

    Returns:
        bytes: The combined Excel file content in memory.
    """
    try:
        # Create a workbook
        wb = Workbook()

        # Sheet 1: Write matched and unmatched CI data
        ws1 = wb.active
        ws1.title = "CI Data"

        # Add "Matched CI" title
        ws1.append(["Matched CI"])
        for row in dataframe_to_rows(matched_ci, index=None, header=True):
            ws1.append(row)

        # Add a gap of two rows
        ws1.append([])
        ws1.append([])

        # Add "Unmatched CI" title
        ws1.append(["Unmatched CI"])
        for row in dataframe_to_rows(unmatched_ci, index=None, header=True):
            ws1.append(row)

        # Sheet 2: Write matched and unmatched PL data
        ws2 = wb.create_sheet(title="PL Data")

        # Add "Matched PL" title
        ws2.append(["Matched PL"])
        for row in dataframe_to_rows(matched_pl, index=None, header=True):
            ws2.append(row)

        # Add a gap of two rows
        ws2.append([])
        ws2.append([])

        # Add "Unmatched PL" title
        ws2.append(["Unmatched PL"])
        for row in dataframe_to_rows(unmatched_pl, index=None, header=True):
            ws2.append(row)

        # Create an in-memory buffer
        output_buffer = io.BytesIO()

        # Save the workbook to the buffer
        wb.save(output_buffer)

        # Reset the buffer position to the beginning
        output_buffer.seek(0)

        return output_buffer.getvalue()

    except Exception as e:
        logger.info(f"Error saving combined sheet: {e}")
        return None


@lru_cache(maxsize=None)
def compare_strings(s1, s2):
    """Compare two strings after normalizing them"""
    # remove whitespace, double spaces, hyphens and brackets
    s1 = str(s1).replace(",", "").replace(" ", "").replace("--", "").casefold()
    s2 = str(s2).replace(",", "").replace(" ", "").replace("--", "").casefold()
    # compare strings
    return s1 == s2


def extract_table_data(file_content: bytes, sheet_name: str):
    """
    Extracts table data from a given sheet using pandas.
    """
    # Read the sheet
    df = pd.read_excel(io.BytesIO(file_content), sheet_name=sheet_name)
    # Drop rows that are entirely NaN
    df = df.dropna(how="all")
    # Reset the index
    df.reset_index(drop=True, inplace=True)
    return df


def get_purchase_orders_from_db(items: List[Dict]) -> List[Dict]:
    """
    Get purchase orders from MongoDB database instead of API calls.

    Args:
        items: List of items to search for

    Returns:
        List of found items with rates
    """
    try:
        db = get_database()
        # Query purchase orders from MongoDB
        # Adjust the query based on your MongoDB schema
        purchase_orders = list(db.purchase_orders.find({"status": {"$ne": "draft"}}))

        found_items = []
        found_names = set()

        logger.info("Processing POs from MongoDB")

        # Process each purchase order
        for purchase_order in purchase_orders:
            line_items = purchase_order.get("line_items", [])

            # Check for matching items in line_items
            for item in items:
                for line_item in line_items:
                    item_name = line_item.get("name")
                    rate = line_item.get("rate", 0)

                    if (
                        compare_strings(item_name, item["name"])
                        and item_name not in found_names
                    ):
                        found_items.append({"rate": rate, "name": item_name})
                        found_names.add(item_name)
                        break

        # Add items not found with rate 0
        for item in items:
            if item["name"] not in found_names:
                found_items.append({"rate": 0, "name": item["name"]})
        logger.info("Done Processing POs from MongoDB")
        return found_items

    except Exception as e:
        logger.info(f"Error querying MongoDB: {e}")
        return [{"rate": 0, "name": item["name"]} for item in items]


def get_items_from_db(search_text: str) -> List[Dict]:
    """
    Get items from MongoDB database instead of API calls.

    Args:
        search_text: Text to search for

    Returns:
        List of matching items
    """
    try:
        db = get_database()
        # Query items from MongoDB - adjust based on your schema
        # Using regex for case-insensitive search
        items = list(
            db.products.find(
                {"item_name": {"$regex": re.escape(search_text), "$options": "i"}}
            )
        )
        return items
    except Exception as e:
        logger.info(f"Error querying items from MongoDB: {e}")
        return []


def process_upload_data(file_content: bytes) -> bytes:
    """
    Process the uploaded file and return Excel data.

    Args:
        file_content: The uploaded file content as bytes

    Returns:
        bytes: The processed Excel file content
    """

    # Preprocessing function to remove "cm" and "inch"
    def clean_name(name):
        pattern = r"((SIZE\s*\d+x\d+CM)|(\s*-?\s*\d+\s*cm\s*/\s*\d+\s*inch\s*))[\s/]+.*$|(\d+\s*cm\s*/\s*\d+\s*inch)"
        cleaned_name = re.sub(pattern, "", str(name), flags=re.IGNORECASE)
        return cleaned_name.strip()

    # Extract table data from both sheets
    pl_sheet = extract_table_data(file_content, "PL")
    ci_sheet = extract_table_data(file_content, "CI")

    # Clean "Name" column in both sheets
    pl_sheet["Name"] = pl_sheet["Name"].apply(clean_name)
    ci_sheet["Name"] = ci_sheet["Name"].apply(clean_name)

    logger.info(f"PL Data count: {len(pl_sheet['Name'])}")
    pl_data = [x for x in pl_sheet["Name"] if pd.notna(x)]

    ci_data = [
        {
            "name": row["Name"],
            "hsn": str(int(row["HSN"])) if pd.notna(row["HSN"]) else "",
            "price": row["Price"] if pd.notna(row["Price"]) else 0,
        }
        for _, row in ci_sheet.iterrows()
        if isinstance(row["Name"], str) and pd.notna(row["Name"])
    ]

    # Initialize result lists
    matched_pl, unmatched_pl, matched_ci, unmatched_ci = [], [], [], []

    # Get purchase orders data from MongoDB
    purchase_data = get_purchase_orders_from_db(ci_data)

    logger.info(f"Processing PL Data: {len(pl_data)}")
    # Process PL data
    for item in pl_data:
        items_from_db = get_items_from_db(item)
        if len(items_from_db) > 0:
            product = items_from_db[0]
            product_name = product.get("item_name")
            if compare_strings(item, product_name):
                matched_pl.append({"name": product_name})
            else:
                unmatched_pl.append({"name": item})
        else:
            unmatched_pl.append({"name": item})

    logger.info("Done Processing PL Data")
    logger.info(f"Processing CI Data: {len(ci_data)}")

    # Process CI data
    for item in ci_data:
        name = str(item.get("name")).strip()
        code = item.get("hsn")
        price = item.get("price")
        rounded_price = round(price, 6)

        items_from_db = get_items_from_db(name)
        if len(items_from_db) > 0:
            product = items_from_db[0]
            product_name = product.get("item_name")
            product_code = product.get("hsn_or_sac", "")
            product_price = next(
                (
                    entry["rate"]
                    for entry in purchase_data
                    if compare_strings(entry["name"], product_name)
                ),
                0,
            )

            if (
                compare_strings(name, product_name)
                and compare_strings(code, product_code)
                and compare_strings(rounded_price, product_price)
            ):
                matched_ci.append(
                    {
                        "name": name,
                        "hsn": code,
                        "price": price,
                    }
                )
            else:
                reasons = []
                if not compare_strings(name, product_name):
                    reasons.append(f"Name {name} not matched with {product_name}")
                if not compare_strings(code, product_code):
                    reasons.append(f"HSN {code} not matched with {product_code}")
                if not compare_strings(rounded_price, product_price):
                    reasons.append(
                        f"Price {rounded_price} not matched with {product_price}"
                    )

                reason = "; ".join(reasons)
                unmatched_ci.append(
                    {
                        "name": name,
                        "hsn": code,
                        "price": price,
                        "reason": reason,
                    }
                )
        else:
            reason = f"{name} Not found in Zoho"
            unmatched_ci.append(
                {
                    "name": name,
                    "hsn": code,
                    "price": price,
                    "reason": reason,
                }
            )

    logger.info("Done Processing CI Data")

    # Create DataFrames and sort them
    matched_pl_df = pd.DataFrame(sorted(matched_pl, key=lambda x: str(x["name"])))
    unmatched_pl_df = pd.DataFrame(sorted(unmatched_pl, key=lambda x: str(x["name"])))
    matched_ci_df = pd.DataFrame(sorted(matched_ci, key=lambda x: x.get("hsn", "")))
    unmatched_ci_df = pd.DataFrame(sorted(unmatched_ci, key=lambda x: x.get("hsn", "")))

    # Generate the combined Excel file
    workbook_bytes = save_combined_sheet(
        matched_ci_df, unmatched_ci_df, matched_pl_df, unmatched_pl_df
    )

    return workbook_bytes


@router.post("/upload_template")
async def upload_template(file: UploadFile = File(...)):
    """
    Process uploaded Excel file with CI and PL sheets.
    Validates the file, processes data against MongoDB, and returns processed Excel file.
    """
    try:
        # Check if file is provided
        if not file:
            raise HTTPException(status_code=400, detail="No file provided")

        if not file.filename:
            raise HTTPException(status_code=400, detail="No file selected")

        # Check file extension
        if not file.filename.lower().endswith((".xlsx", ".xls")):
            raise HTTPException(
                status_code=400,
                detail="Invalid file format. Please upload an Excel file (.xlsx or .xls)",
            )

        # Read file content
        file_content = await file.read()

        # Validate file structure
        validation_result = validate_file(file_content)
        if validation_result["status"] == "error":
            raise HTTPException(
                status_code=400,
                detail=f"File validation failed: {validation_result['message']}",
            )

        # Process the file
        processed_excel_bytes = process_upload_data(file_content)

        if not processed_excel_bytes:
            raise HTTPException(status_code=500, detail="Error processing file data")

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"CI_PL_Processed_{timestamp}.xlsx"

        # Return the processed Excel file
        return StreamingResponse(
            io.BytesIO(processed_excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")
