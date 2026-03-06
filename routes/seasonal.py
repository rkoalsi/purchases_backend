import asyncio
import io
import logging
from datetime import datetime
from typing import Dict, List

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..database import get_database

logger = logging.getLogger(__name__)
router = APIRouter()

MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# ---------------------------------------------------------------------------
# Data fetchers
# ---------------------------------------------------------------------------

async def _fetch_products(db, brand: str = None) -> Dict[str, Dict]:
    """Return {sku_code: {name, item_id, brand}} optionally filtered by brand."""
    products_collection = db["products"]
    query: dict = {}
    if brand:
        query["brand"] = {"$regex": f"^{brand}$", "$options": "i"}

    def _run():
        return list(products_collection.find(
            query,
            {"cf_sku_code": 1, "item_id": 1, "name": 1, "brand": 1, "purchase_status": 1, "_id": 0},
        ))

    docs = await asyncio.to_thread(_run)
    result: Dict[str, Dict] = {}
    for doc in docs:
        sku = doc.get("cf_sku_code")
        item_id = doc.get("item_id")
        if sku and item_id:
            result[sku] = {
                "name": doc.get("name", "Unknown"),
                "item_id": str(item_id),
                "brand": doc.get("brand", ""),
                "purchase_status": doc.get("purchase_status", ""),
            }
    return result


def _composite_expansion_stages(target_ids: List[str]) -> list:
    """Shared pipeline stages that unwind line items and expand composites."""
    return [
        {"$unwind": "$line_items"},
        {
            "$lookup": {
                "from": "composite_products",
                "localField": "line_items.item_id",
                "foreignField": "composite_item_id",
                "as": "composite_match",
            }
        },
        {
            "$addFields": {
                "composite_info": {"$arrayElemAt": ["$composite_match", 0]},
            }
        },
        {
            "$project": {
                "date": 1,
                "items_to_process": {
                    "$cond": [
                        {"$gt": [{"$size": "$composite_match"}, 0]},
                        {
                            "$map": {
                                "input": "$composite_info.components",
                                "as": "comp",
                                "in": {
                                    "item_id": "$$comp.item_id",
                                    "quantity": {
                                        "$multiply": [
                                            "$line_items.quantity",
                                            "$$comp.quantity",
                                        ]
                                    },
                                    "date": "$date",
                                },
                            }
                        },
                        [
                            {
                                "item_id": "$line_items.item_id",
                                "quantity": "$line_items.quantity",
                                "date": "$date",
                            }
                        ],
                    ]
                },
            }
        },
        {"$unwind": "$items_to_process"},
        {"$replaceRoot": {"newRoot": "$items_to_process"}},
        {"$match": {"item_id": {"$in": target_ids}}},
    ]


async def _fetch_base_drr(
    db,
    sku_to_item_id: Dict[str, str],
    start_date: str,
    end_date: str,
    period_days: int,
) -> Dict[str, float]:
    """Base DRR = total Zoho invoice units in the report period / period_days."""
    if not sku_to_item_id or period_days <= 0:
        return {}

    invoices_collection = db["invoices"]
    target_ids = list(sku_to_item_id.values())

    pipeline = [
        {
            "$match": {
                "date": {"$gte": start_date, "$lte": end_date},
                "status": {"$nin": ["draft", "void"]},
            }
        },
        *_composite_expansion_stages(target_ids),
        {
            "$group": {
                "_id": "$item_id",
                "total_units": {"$sum": "$quantity"},
            }
        },
    ]

    def _run():
        return list(invoices_collection.aggregate(pipeline, allowDiskUse=True))

    docs = await asyncio.to_thread(_run)
    item_id_to_sku = {v: k for k, v in sku_to_item_id.items()}

    result: Dict[str, float] = {}
    for doc in docs:
        item_id = str(doc["_id"])
        sku = item_id_to_sku.get(item_id)
        if sku:
            units = float(doc.get("total_units", 0))
            result[sku] = round(units / period_days, 4)
    return result


async def _fetch_monthly_sales_by_year(
    db,
    sku_to_item_id: Dict[str, str],
    end_date: str,
) -> Dict[str, Dict[int, Dict[int, float]]]:
    """
    For each SKU return {year: {month_num: units}}.
    Looks back 2 full years from end_date.
    """
    if not sku_to_item_id:
        return {}

    invoices_collection = db["invoices"]
    end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    lookback_start = end_dt.replace(year=end_dt.year - 2).strftime("%Y-%m-%d")
    target_ids = [str(iid) for iid in sku_to_item_id.values()]

    pipeline = [
        {
            "$match": {
                "date": {"$gte": lookback_start, "$lte": end_date},
                "status": {"$nin": ["draft", "void"]},
            }
        },
        *_composite_expansion_stages(target_ids),
        {
            "$addFields": {
                "parsed_date": {
                    "$dateFromString": {"dateString": "$date", "format": "%Y-%m-%d"}
                }
            }
        },
        {
            "$group": {
                "_id": {
                    "item_id": "$item_id",
                    "year": {"$year": "$parsed_date"},
                    "month": {"$month": "$parsed_date"},
                },
                "units": {"$sum": "$quantity"},
            }
        },
    ]

    def _run():
        raw: Dict[str, Dict[int, Dict[int, float]]] = {}
        for doc in invoices_collection.aggregate(pipeline, allowDiskUse=True):
            item_id = str(doc["_id"]["item_id"])
            year = doc["_id"]["year"]
            month = doc["_id"]["month"]
            units = float(doc.get("units", 0))
            raw.setdefault(item_id, {}).setdefault(year, {})[month] = units
        return raw

    raw = await asyncio.to_thread(_run)
    item_id_to_sku = {str(v): k for k, v in sku_to_item_id.items()}

    result: Dict[str, Dict[int, Dict[int, float]]] = {}
    for item_id, year_data in raw.items():
        sku = item_id_to_sku.get(item_id)
        if not sku:
            continue
        result[sku] = {yr: dict(months) for yr, months in year_data.items()}

    return result


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

# Static column definitions: (header label, width, colour)
_STATIC_COLS_BEFORE = [
    ("SKU Code",  14, "1F4E79"),
    ("Item Name", 32, "1F4E79"),
    ("Brand",     16, "1F4E79"),
    ("Status",    18, "1F4E79"),
    ("Month",      8, "1F4E79"),
]
_STATIC_COLS_AFTER = [
    ("Monthly Avg\n= AVG(Year Cols)",                     18, "2E75B6"),
    ("__OVERALL_AVG__",                                   22, "2E75B6"),  # label set dynamically in _build_excel
    ("Seasonal Index\n= Month Avg / Overall Avg",          18, "375623"),
    ("Base DRR\n(Period Units / Period Days)",              18, "C55A11"),
    ("Seasonal DRR\n= Base DRR × Seasonal Index",          20, "7030A0"),
]

# Alternating row shades for even/odd SKU blocks
_SHADE_EVEN = "EBF3FB"
_SHADE_ODD  = "FDFEFE"


def _cell_fill(colour: str) -> PatternFill:
    return PatternFill(start_color=colour, end_color=colour, fill_type="solid")


def _build_excel(
    products: Dict[str, Dict],
    base_drr_by_sku: Dict[str, float],
    monthly_sales_by_year: Dict[str, Dict[int, Dict[int, float]]],
    start_date: str,
    end_date: str,
) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Seasonal Report"

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Show one column per calendar year covered by the selected date range
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    end_dt   = datetime.strptime(end_date,   "%Y-%m-%d")
    display_years = list(range(start_dt.year, end_dt.year + 1))
    year_cols = [(f"{yr} Units\n(Jan–Dec)", 16, "2E75B6") for yr in display_years]

    # Build full column list; override the dynamic overall-avg label
    yr_range = f"{display_years[0]}" if len(display_years) == 1 else f"{display_years[0]}–{display_years[-1]}"
    overall_avg_label = f"Avg Monthly Demand\n({yr_range})"
    static_after = [
        (overall_avg_label if lbl == "__OVERALL_AVG__" else lbl, w, c)
        for lbl, w, c in _STATIC_COLS_AFTER
    ]
    all_cols = _STATIC_COLS_BEFORE + year_cols + static_after
    total_cols = len(all_cols)

    # Column index references (1-based)
    year_col_start    = len(_STATIC_COLS_BEFORE) + 1              # e.g. 6
    year_col_end      = year_col_start + len(display_years) - 1   # e.g. 7
    monthly_avg_col   = year_col_end + 1                           # e.g. 8
    overall_avg_col   = monthly_avg_col + 1                        # e.g. 9
    seasonal_idx_col  = overall_avg_col + 1                        # e.g. 10
    base_drr_col      = seasonal_idx_col + 1                       # e.g. 11
    seasonal_drr_col  = base_drr_col + 1                           # e.g. 12

    yr_start_ltr      = get_column_letter(year_col_start)
    yr_end_ltr        = get_column_letter(year_col_end)
    monthly_avg_ltr   = get_column_letter(monthly_avg_col)
    overall_avg_ltr   = get_column_letter(overall_avg_col)
    seasonal_idx_ltr  = get_column_letter(seasonal_idx_col)
    base_drr_ltr      = get_column_letter(base_drr_col)

    # ── header row ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 48
    for col_idx, (label, width, colour) in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = _cell_fill(colour)
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header
    ws.freeze_panes = "A2"

    # ── legend rows ──────────────────────────────────────────────────────────
    legend_fill = _cell_fill("FFF2CC")
    legend_font = Font(italic=True, size=9, color="595959")

    yrs_str = " & ".join(str(y) for y in display_years)
    legend_lines = [
        f"STEP 1: Year columns ({yrs_str})  → Zoho invoice units for that calendar month in each respective year.",
        f"STEP 2: Monthly Avg               → =AVERAGE({yrs_str} cols)  (average units for that month across both years).",
        f"STEP 3: {overall_avg_label.replace(chr(10), ' ')}  → =SUM(Jan Avg … Dec Avg)/12  (average of the 12 monthly averages).",
        "STEP 4: Seasonal Index            → =Month Avg / Avg Monthly Demand   (1.0 = normal; >1.0 = above average; <1.0 = below average).",
        "STEP 5: Base DRR                  → Total Zoho units sold in the selected period / number of days in the period.",
        "STEP 6: Seasonal DRR              → =Base DRR × Seasonal Index   (projected daily run rate adjusted for seasonality).",
    ]
    for i, line in enumerate(legend_lines):
        r = 2 + i
        ws.row_dimensions[r].height = 15
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
        cell = ws.cell(row=r, column=1, value=line)
        cell.font      = legend_font
        cell.fill      = legend_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── data rows ────────────────────────────────────────────────────────────
    DATA_START = 2 + len(legend_lines)

    sorted_skus = sorted(
        products.keys(),
        key=lambda s: (products[s].get("brand", ""), products[s].get("name", "")),
    )

    current_row = DATA_START
    for sku_idx, sku in enumerate(sorted_skus):
        pdata           = products[sku]
        name            = pdata.get("name", "Unknown")
        brand           = pdata.get("brand", "")
        purchase_status = pdata.get("purchase_status", "")
        base_drr        = base_drr_by_sku.get(sku, 0.0)
        sku_year_data   = monthly_sales_by_year.get(sku, {})

        block_start = current_row
        block_end   = current_row + 11  # 12 rows inclusive

        shade    = _SHADE_EVEN if sku_idx % 2 == 0 else _SHADE_ODD
        row_fill = _cell_fill(shade)

        for month_num in range(1, 13):
            row = current_row + month_num - 1

            ws.cell(row=row, column=1, value=sku)
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=brand)
            ws.cell(row=row, column=4, value=purchase_status)
            ws.cell(row=row, column=5, value=MONTH_NAMES[month_num - 1])

            # Year columns (2025, 2026)
            for yr_offset, yr in enumerate(display_years):
                col = year_col_start + yr_offset
                units = sku_year_data.get(yr, {}).get(month_num, 0.0)
                ws.cell(row=row, column=col, value=round(units, 4))

            # Monthly Avg = AVERAGE of the two year columns
            ws.cell(
                row=row, column=monthly_avg_col,
                value=f"=AVERAGE({yr_start_ltr}{row}:{yr_end_ltr}{row})"
            )

            # Avg Monthly Demand: sum of Monthly Avg block / 12
            ws.cell(
                row=row, column=overall_avg_col,
                value=f"=SUM(${monthly_avg_ltr}${block_start}:${monthly_avg_ltr}${block_end})/12"
            )

            # Seasonal Index
            ws.cell(
                row=row, column=seasonal_idx_col,
                value=f"=IF({overall_avg_ltr}{row}=0,1,{monthly_avg_ltr}{row}/{overall_avg_ltr}{row})"
            )

            # Base DRR (same for all months of this SKU)
            ws.cell(row=row, column=base_drr_col, value=round(base_drr, 4))

            # Seasonal DRR
            ws.cell(
                row=row, column=seasonal_drr_col,
                value=f"={base_drr_ltr}{row}*{seasonal_idx_ltr}{row}"
            )

            # Styling
            for col in range(1, total_cols + 1):
                cell           = ws.cell(row=row, column=col)
                cell.fill      = row_fill
                cell.alignment = (left if col == 2 else center)
                cell.font      = Font(size=10)

        # Number formatting
        for row in range(block_start, block_end + 1):
            ws.cell(row=row, column=monthly_avg_col).number_format   = "0.0000"
            ws.cell(row=row, column=overall_avg_col).number_format   = "0.0000"
            ws.cell(row=row, column=seasonal_idx_col).number_format  = "0.0000"
            ws.cell(row=row, column=seasonal_drr_col).number_format  = "0.0000"

        # Thick bottom border to visually separate SKU blocks
        try:
            from openpyxl.styles import Border, Side
            thick = Side(style="medium", color="595959")
            for col in range(1, total_cols + 1):
                cell = ws.cell(row=block_end, column=col)
                existing = cell.border
                cell.border = Border(
                    top=existing.top if existing else None,
                    left=existing.left if existing else None,
                    right=existing.right if existing else None,
                    bottom=thick,
                )
        except Exception:
            pass  # non-critical styling

        current_row += 12

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------------------------
# Endpoint
# ---------------------------------------------------------------------------

@router.get("/download")
async def download_seasonal_report(
    start_date: str = Query(..., description="Start date YYYY-MM-DD"),
    end_date: str = Query(..., description="End date YYYY-MM-DD"),
    brand: str = Query(None, description="Filter by brand (optional)"),
    db=Depends(get_database),
):
    """
    Download the Seasonal DRR report as an Excel file.

    For each SKU, shows 12 rows (one per calendar month) with:
    - Historical average units sold (same month averaged across last 2 years)
    - Overall monthly average (Excel formula)
    - Seasonal Index (Excel formula)
    - Base DRR for the selected period
    - Seasonal DRR (Excel formula)
    """
    try:
        start_dt   = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt     = datetime.strptime(end_date, "%Y-%m-%d")
        period_days = (end_dt - start_dt).days + 1

        if period_days <= 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="end_date must be on or after start_date",
            )

        # 1. Products
        products = await _fetch_products(db, brand)
        if not products:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No products found for the given filters",
            )

        sku_to_item_id = {sku: pdata["item_id"] for sku, pdata in products.items()}

        # 2. Base DRR + historical monthly sales by year (parallel)
        base_drr_by_sku, monthly_sales_by_year = await asyncio.gather(
            _fetch_base_drr(db, sku_to_item_id, start_date, end_date, period_days),
            _fetch_monthly_sales_by_year(db, sku_to_item_id, end_date),
        )

        # 3. Build Excel (CPU work off the event loop)
        excel_buffer = await asyncio.to_thread(
            _build_excel,
            products,
            base_drr_by_sku,
            monthly_sales_by_year,
            start_date,
            end_date,
        )

        filename = f"seasonal_report_{start_date}_to_{end_date}"
        if brand:
            filename += f"_{brand.replace(' ', '_')}"
        filename += ".xlsx"

        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating seasonal report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Internal server error: {str(e)}",
        )
