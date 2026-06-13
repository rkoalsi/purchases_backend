import io
import math
import os
import re
import uuid
import zipfile
import logging
import asyncio
from datetime import datetime, timedelta
from ..helpers.datetime_utils import utcnow
from typing import Any, List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import boto3
from botocore.config import Config as BotocoreConfig
from bson import ObjectId
from fastapi import APIRouter, Depends, File, Form, Header, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import JSONResponse, StreamingResponse
from jose import jwt, JWTError
from pydantic import BaseModel
from pymongo.errors import PyMongoError

from ..database import get_database, serialize_mongo_document

router = APIRouter()
logger = logging.getLogger(__name__)

PRODUCTS_COLLECTION = "products"
BRAND_ORDERS_COLLECTION = "brand_orders"
CATEGORIES_COLLECTION = "designer_upload_categories"
PO_COLLECTION = "purchase_orders"
VENDORS_COLLECTION = "vendors"
BRANDS_COLLECTION = "brands"
DESIGN_CATALOGUE_COLLECTION = "design_catalogue"
S3_BUCKET = os.getenv("S3_BUCKET", "pupscribe-purchases")
AWS_REGION = os.getenv("AWS_REGION", "ap-south-1")
S3_PUBLIC_URL = os.getenv("S3_PUBLIC_URL", f"https://{os.getenv('S3_BUCKET', 'pupscribe-purchases')}.s3.{os.getenv('AWS_REGION', 'ap-south-1')}.amazonaws.com")
DESIGNER_URL_EXPIRY = 7 * 24 * 3600  # 1 week
DEFAULT_CATEGORIES = ["products", "catalogue"]
MAX_IMAGE_MB = 10
MAX_VIDEO_MB = 200

# Public order-form bucket for product images/videos
PUBLIC_S3_BUCKET = os.getenv("PUBLIC_S3_BUCKET", "order-form")
PUBLIC_S3_REGION = os.getenv("PUBLIC_S3_REGION", "ap-south-1")
PUBLIC_S3_URL = os.getenv("PUBLIC_S3_URL", "https://assets.pupscribe.in")


class AddCategoryRequest(BaseModel):
    name: str


class CreateFolderRequest(BaseModel):
    name: str
    parent_path: Optional[str] = ""


class RenameFolderRequest(BaseModel):
    name: str


def _sanitize_folder_name(name: str) -> str:
    """Convert folder name to a safe path segment (preserves readability)."""
    name = name.strip()
    safe = re.sub(r"[^\w\s.\-]", "", name)
    safe = re.sub(r"\s+", "_", safe).strip("_")
    return safe or "folder"


def _slugify(text: str) -> str:
    text = text.lower().strip()
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[\s_-]+", "_", text)
    return text


def _s3_client():
    return boto3.client(
        "s3",
        region_name=AWS_REGION,
        endpoint_url=f"https://s3.{AWS_REGION}.amazonaws.com",
        config=BotocoreConfig(signature_version="s3v4"),
    )


def _public_s3_client():
    return boto3.client(
        "s3",
        region_name=PUBLIC_S3_REGION,
        aws_access_key_id=os.getenv("PUBLIC_S3_ACCESS_KEY"),
        aws_secret_access_key=os.getenv("PUBLIC_S3_SECRET_KEY"),
    )


def _upload_to_s3(file_bytes: bytes, s3_key: str, content_type: str) -> None:
    boto3.client("s3", region_name=AWS_REGION).put_object(
        Bucket=S3_BUCKET,
        Key=s3_key,
        Body=file_bytes,
        ContentType=content_type,
    )


def _presign_s3(s3_key: str, expires: int = DESIGNER_URL_EXPIRY) -> str:
    return _s3_client().generate_presigned_url(
        "get_object",
        Params={"Bucket": S3_BUCKET, "Key": s3_key},
        ExpiresIn=expires,
    )


def _delete_from_s3(s3_key: str) -> None:
    boto3.client("s3", region_name=AWS_REGION).delete_object(
        Bucket=S3_BUCKET, Key=s3_key
    )


# ─── new-items (existing endpoint) ───────────────────────────────────────────

_CATALOGUE_LOOKUP_FIELDS = {
    # product_name, product_category, hsn_code intentionally excluded —
    # those always come from the products collection (Zoho source of truth).
    "bb_code": 1, "sub_category": 1,
    "mrp": 1, "dimensions": 1, "material": 1, "features": 1,
    "image_links": 1, "squeaker": 1, "catnip": 1,
    "age_group": 1, "pet_size": 1, "chewing_style": 1,
    "size_chart": 1, "ingredient_list": 1, "nutritional_analysis": 1,
    "gst_percentage": 1,
}


@router.get("/new-items")
def get_new_items(
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    search: str = Query(None, description="Search by product name, SKU, or BB code"),
    brand: str = Query(None, description="Filter by brand"),
    zoho_status: str = Query(None, description="Filter by Zoho status (active/inactive)"),
    purchase_status: str = Query(None, description="Filter by purchase status"),
    category: str = Query(None, description="Filter by catalogue product_category"),
):
    """Returns non-combo products sorted by created_at desc, joined with catalogue data."""
    try:
        db = get_database()
        col = db[PRODUCTS_COLLECTION]

        query_filter: dict = {"is_combo_product": False}

        if search:
            rgx = {"$regex": re.escape(search), "$options": "i"}
            cat_ids = db[DESIGN_CATALOGUE_COLLECTION].distinct(
                "product_id", {"bb_code": rgx}
            )
            query_filter["$or"] = [
                {"name": rgx},
                {"cf_sku_code": rgx},
                {"sku": rgx},
                {"_id": {"$in": cat_ids}},
            ]

        if brand:
            query_filter["brand"] = brand
        if zoho_status:
            query_filter["status"] = zoho_status.lower()
        if purchase_status:
            query_filter["purchase_status"] = purchase_status
        if category:
            cat_ids_for_cat = db[DESIGN_CATALOGUE_COLLECTION].distinct(
                "product_id", {"product_category": category}
            )
            query_filter["_id"] = {"$in": cat_ids_for_cat}

        total_count = col.count_documents(query_filter)
        total_pages = math.ceil(total_count / limit) if total_count > 0 else 1
        skip = (page - 1) * limit

        pipeline = [
            {"$match": query_filter},
            {"$sort": {"created_at": -1, "_id": -1}},
            {"$skip": skip},
            {"$limit": limit},
            {"$lookup": {
                "from": DESIGN_CATALOGUE_COLLECTION,
                "localField": "_id",
                "foreignField": "product_id",
                "as": "_cat",
                "pipeline": [{"$project": _CATALOGUE_LOOKUP_FIELDS}],
            }},
            {"$addFields": {"catalogue": {"$arrayElemAt": ["$_cat", 0]}}},
            {"$project": {"_cat": 0}},
        ]
        products = serialize_mongo_document(list(col.aggregate(pipeline)))

        return JSONResponse(content={
            "products": products,
            "pagination": {
                "currentPage": page,
                "totalPages": total_pages,
                "totalProducts": total_count,
                "limit": limit,
            },
        })

    except PyMongoError as e:
        logger.error(f"DB error in design/new-items: {e}")
        raise HTTPException(status_code=500, detail=f"Database error: {e}")
    except Exception as e:
        logger.error(f"Unexpected error in design/new-items: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ─── design catalogue ────────────────────────────────────────────────────────

class CatalogueItemPatch(BaseModel):
    image_links: Optional[List[str]] = None
    video_links: Optional[List[str]] = None
    features: Optional[List[Any]] = None
    ingredient_list: Optional[str] = None
    nutritional_analysis: Optional[str] = None
    age_group: Optional[str] = None
    pet_size: Optional[str] = None
    chewing_style: Optional[str] = None
    material: Optional[str] = None
    squeaker: Optional[bool] = None
    catnip: Optional[bool] = None
    size_chart: Optional[str] = None
    dimensions: Optional[dict] = None


@router.get("/catalogue-items")
def get_catalogue_items(
    page: int = Query(1, ge=1),
    limit: int = Query(24, ge=1, le=100),
    search: str = Query(None),
    category: str = Query(None),
    has_images: bool = Query(None),
):
    try:
        db = get_database()
        col = db[DESIGN_CATALOGUE_COLLECTION]

        q: dict = {}
        if search:
            rgx = {"$regex": re.escape(search), "$options": "i"}
            q["$or"] = [{"product_name": rgx}, {"bb_code": rgx}]
        if category:
            q["product_category"] = category
        if has_images is True:
            q["image_links.0"] = {"$exists": True}
        elif has_images is False:
            q["$or"] = [{"image_links": {"$exists": False}}, {"image_links": []}]

        total = col.count_documents(q)
        skip  = (page - 1) * limit

        pipeline = [
            {"$match": q},
            {"$sort": {"bb_code": 1}},
            {"$skip": skip},
            {"$limit": limit},
            {"$lookup": {
                "from": PRODUCTS_COLLECTION,
                "localField": "product_id",
                "foreignField": "_id",
                "as": "_prod",
                "pipeline": [{"$project": {"images": 1, "image_url": 1}}],
            }},
            {"$addFields": {
                "product_images":    {"$ifNull": [{"$arrayElemAt": ["$_prod.images", 0]}, []]},
                "product_image_url": {"$arrayElemAt": ["$_prod.image_url", 0]},
            }},
            {"$project": {"_prod": 0}},
        ]
        docs = list(col.aggregate(pipeline))

        return JSONResponse(content={
            "items": serialize_mongo_document(docs),
            "pagination": {
                "currentPage": page,
                "totalPages": math.ceil(total / limit) if total else 1,
                "total": total,
                "limit": limit,
            },
        })
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/catalogue-items/categories")
def get_catalogue_categories():
    try:
        db  = get_database()
        cats = db[DESIGN_CATALOGUE_COLLECTION].distinct("product_category")
        return JSONResponse(content={"categories": sorted(c for c in cats if c)})
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/catalogue-items/{bb_code}")
def patch_catalogue_item(bb_code: str, body: CatalogueItemPatch):
    try:
        db = get_database()
        update = body.dict(exclude_unset=True)
        if not update:
            raise HTTPException(status_code=400, detail="No fields provided")
        update["updated_at"] = utcnow()
        res = db[DESIGN_CATALOGUE_COLLECTION].update_one(
            {"bb_code": bb_code}, {"$set": update}
        )
        if res.matched_count == 0:
            raise HTTPException(status_code=404, detail="Catalogue item not found")
        return JSONResponse(content={"ok": True})
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/products/{product_id}/images")
def patch_product_images(product_id: str, body: dict):
    """Update image_url (single), images array, and/or videos array for a product."""
    try:
        db = get_database()
        update: dict = {"updated_at": utcnow()}
        if "image_url" in body:
            update["image_url"] = body["image_url"] or None
        if "images" in body:
            if not isinstance(body["images"], list):
                raise HTTPException(status_code=400, detail="images must be a list")
            update["images"] = body["images"]
        if "videos" in body:
            if not isinstance(body["videos"], list):
                raise HTTPException(status_code=400, detail="videos must be a list")
            update["videos"] = body["videos"]
        if len(update) == 1:
            raise HTTPException(status_code=400, detail="No fields to update")
        db[PRODUCTS_COLLECTION].update_one(
            {"_id": ObjectId(product_id)},
            {"$set": update},
        )
        return JSONResponse(content={"ok": True})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class ProductDetailsPatch(BaseModel):
    category: Optional[str] = None
    sub_category: Optional[str] = None
    series: Optional[str] = None


@router.patch("/products/{product_id}/details")
def patch_product_details(product_id: str, body: ProductDetailsPatch):
    """Update product-level category, sub_category, series fields."""
    try:
        db = get_database()
        update: dict = {k: v for k, v in body.dict().items() if v is not None}
        if not update:
            raise HTTPException(status_code=400, detail="No fields to update")
        update["updated_at"] = utcnow()
        db[PRODUCTS_COLLECTION].update_one(
            {"_id": ObjectId(product_id)},
            {"$set": update},
        )
        return JSONResponse(content={"ok": True})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/products/{product_id}/catalogue-details")
def patch_product_catalogue_details(product_id: str, body: CatalogueItemPatch):
    """Upsert catalogue details by product_id — works with or without an existing catalogue entry."""
    try:
        db = get_database()
        update: dict = {k: v for k, v in body.dict().items() if v is not None}
        if not update:
            raise HTTPException(status_code=400, detail="No fields to update")
        update["updated_at"] = utcnow()
        db[DESIGN_CATALOGUE_COLLECTION].update_one(
            {"product_id": ObjectId(product_id)},
            {"$set": update},
            upsert=True,
        )
        return JSONResponse(content={"ok": True})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/products/{product_id}/upload-image")
async def upload_product_image(product_id: str, file: UploadFile = File(...)):
    """Upload an image to the public S3 bucket and return its public URL."""
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="File must be an image.")
    data = await file.read()
    if len(data) > MAX_IMAGE_MB * 1024 * 1024:
        raise HTTPException(status_code=400, detail=f"Image exceeds {MAX_IMAGE_MB} MB.")
    db = get_database()
    product = db[PRODUCTS_COLLECTION].find_one({"_id": ObjectId(product_id)}, {"item_id": 1, "images": 1, "image_url": 1})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found.")
    item_id = product.get("item_id") or product_id
    # sequential numbering: count existing images to pick next index
    existing_images = product.get("images") or []
    n = len(existing_images) + (1 if product.get("image_url") else 0) + 1
    ext = os.path.splitext(file.filename or "")[1] or ".jpg"
    key = f"product_images/{item_id}_{n}{ext}"
    try:
        _public_s3_client().put_object(
            Bucket=PUBLIC_S3_BUCKET, Key=key, Body=data,
            ContentType=file.content_type, ACL="public-read",
        )
    except Exception as e:
        logger.exception("S3 image upload failed: %s", e)
        raise HTTPException(status_code=500, detail=f"S3 upload failed: {e}")
    return JSONResponse(content={"url": f"{PUBLIC_S3_URL}/{key}"})


@router.post("/products/{product_id}/upload-video")
async def upload_product_video(product_id: str, file: UploadFile = File(...)):
    """Upload a video to the public S3 bucket and return its public URL."""
    allowed = {"video/mp4", "video/quicktime", "video/webm", "video/x-msvideo", "video/mpeg"}
    if not file.content_type or file.content_type not in allowed:
        raise HTTPException(status_code=400, detail="File must be a supported video format (mp4, mov, webm, avi, mpeg).")
    data = await file.read()
    if len(data) > MAX_VIDEO_MB * 1024 * 1024:
        raise HTTPException(status_code=400, detail=f"Video exceeds {MAX_VIDEO_MB} MB.")
    db = get_database()
    product = db[PRODUCTS_COLLECTION].find_one({"_id": ObjectId(product_id)}, {"item_id": 1, "videos": 1})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found.")
    item_id = product.get("item_id") or product_id
    n = len(product.get("videos") or []) + 1
    ext = os.path.splitext(file.filename or "")[1] or ".mp4"
    key = f"product_videos/{item_id}_{n}{ext}"
    try:
        _public_s3_client().put_object(
            Bucket=PUBLIC_S3_BUCKET, Key=key, Body=data,
            ContentType=file.content_type, ACL="public-read",
        )
    except Exception as e:
        logger.exception("S3 video upload failed: %s", e)
        raise HTTPException(status_code=500, detail=f"S3 upload failed: {e}")
    return JSONResponse(content={"url": f"{PUBLIC_S3_URL}/{key}"})


# ─── XLSX helpers ─────────────────────────────────────────────────────────────

def _style_header_row(ws, row: int, col_count: int, fill_hex: str = "6B21A8"):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color="FFFFFF", size=10)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_col_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)


def _make_unified_xlsx(products: list) -> io.BytesIO:
    """XLSX combining products fields + joined catalogue fields."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = [
        "S.No", "Name", "SKU Code", "BB Code", "Brand",
        "Category", "Sub Category", "MRP (₹)",
        "Zoho Status", "Purchase Status",
        "Material", "Age Group", "Pet Size", "Chewing Style",
        "Squeaker", "Catnip",
        "L w/o Pkg (cm)", "B w/o Pkg (cm)", "H w/o Pkg (cm)", "Net Weight (g)",
        "Features",
        "Drive Image Links",
        "Image URL", "OF Images Count", "OF Videos Count",
        "Created", "Updated",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, p in enumerate(products, 1):
        cat = p.get("catalogue") or {}
        d_no = (cat.get("dimensions") or {}).get("without_packaging") or {}
        features = [f for f in (cat.get("features") or []) if f]
        of_imgs = p.get("images") or []
        of_vids = p.get("videos") or []
        ws.append([
            i,
            p.get("name") or "",
            p.get("cf_sku_code") or "",
            cat.get("bb_code") or "",
            p.get("brand") or "",
            p.get("category") or "",
            cat.get("sub_category") or "",
            p.get("rate"),
            p.get("status") or "",
            p.get("purchase_status") or "",
            cat.get("material") or "",
            cat.get("age_group") or "",
            cat.get("pet_size") or "",
            cat.get("chewing_style") or "",
            "Yes" if cat.get("squeaker") else "",
            "Yes" if cat.get("catnip") else "",
            d_no.get("length_cm"),
            d_no.get("breadth_cm"),
            d_no.get("height_cm"),
            d_no.get("net_weight_g"),
            " | ".join(features),
            " | ".join(cat.get("image_links") or []),
            p.get("image_url") or "",
            len(of_imgs),
            len(of_vids),
            str(p.get("created_at") or "")[:10],
            str(p.get("updated_at") or "")[:10],
        ])
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=i + 1, column=col_idx).border = border
            ws.cell(row=i + 1, column=col_idx).alignment = Alignment(vertical="center")

    _auto_col_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _make_catalogue_xlsx(items: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catalogue"

    headers = [
        "S.No", "BB Code", "Product Name", "Category", "Sub Category",
        "MRP (₹)", "HSN Code", "GST %",
        "Material", "Age Group", "Pet Size", "Chewing Style",
        "Squeaker", "Catnip",
        "L w/o Pkg (cm)", "B w/o Pkg (cm)", "H w/o Pkg (cm)", "Net Weight (g)",
        "L w/ Pkg (cm)", "B w/ Pkg (cm)", "H w/ Pkg (cm)", "Gross Weight (g)",
        "Features",
        "Ingredient List", "Nutritional Analysis",
        "Image Links (Drive)", "Product Image URL",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, it in enumerate(items, 1):
        d_no = (it.get("dimensions") or {}).get("without_packaging") or {}
        d_wi = (it.get("dimensions") or {}).get("with_packaging") or {}
        features = [f for f in (it.get("features") or []) if f]
        ws.append([
            i,
            it.get("bb_code") or "",
            it.get("product_name") or "",
            it.get("product_category") or "",
            it.get("sub_category") or "",
            it.get("mrp"),
            it.get("hsn_code") or "",
            it.get("gst_percentage"),
            it.get("material") or "",
            it.get("age_group") or "",
            it.get("pet_size") or "",
            it.get("chewing_style") or "",
            "Yes" if it.get("squeaker") else "",
            "Yes" if it.get("catnip") else "",
            d_no.get("length_cm"),
            d_no.get("breadth_cm"),
            d_no.get("height_cm"),
            d_no.get("net_weight_g"),
            d_wi.get("length_cm"),
            d_wi.get("breadth_cm"),
            d_wi.get("height_cm"),
            d_wi.get("gross_weight_g"),
            " | ".join(features),
            it.get("ingredient_list") or "",
            it.get("nutritional_analysis") or "",
            " | ".join(it.get("image_links") or []),
            it.get("product_image_url") or "",
        ])
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=i + 1, column=col)
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=False)

    _auto_col_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── XLSX download endpoints ──────────────────────────────────────────────────

@router.get("/new-items/download")
def download_products_xlsx(
    search: str = Query(None),
    brand: str = Query(None),
    zoho_status: str = Query(None),
    purchase_status: str = Query(None),
    category: str = Query(None),
):
    """Download all matching products (with catalogue join) as XLSX."""
    try:
        db = get_database()
        col = db[PRODUCTS_COLLECTION]
        q: dict = {"is_combo_product": False}
        if search:
            rgx = {"$regex": re.escape(search), "$options": "i"}
            cat_ids = db[DESIGN_CATALOGUE_COLLECTION].distinct("product_id", {"bb_code": rgx})
            q["$or"] = [{"name": rgx}, {"cf_sku_code": rgx}, {"sku": rgx}, {"_id": {"$in": cat_ids}}]
        if brand:
            q["brand"] = brand
        if zoho_status:
            q["status"] = zoho_status.lower()
        if purchase_status:
            q["purchase_status"] = purchase_status
        if category:
            cat_ids_for_cat = db[DESIGN_CATALOGUE_COLLECTION].distinct("product_id", {"product_category": category})
            q["_id"] = {"$in": cat_ids_for_cat}

        pipeline = [
            {"$match": q},
            {"$sort": {"created_at": -1, "_id": -1}},
            {"$lookup": {
                "from": DESIGN_CATALOGUE_COLLECTION,
                "localField": "_id",
                "foreignField": "product_id",
                "as": "_cat",
                "pipeline": [{"$project": _CATALOGUE_LOOKUP_FIELDS}],
            }},
            {"$addFields": {"catalogue": {"$arrayElemAt": ["$_cat", 0]}}},
            {"$project": {"_cat": 0}},
        ]
        products = serialize_mongo_document(list(col.aggregate(pipeline)))
        buf = _make_unified_xlsx(products)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="products.xlsx"'},
        )
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/catalogue-items/download")
def download_catalogue_xlsx(
    search: str = Query(None),
    category: str = Query(None),
    has_images: bool = Query(None),
):
    """Download all matching catalogue items as XLSX (ignores pagination)."""
    try:
        db = get_database()
        col = db[DESIGN_CATALOGUE_COLLECTION]
        q: dict = {}
        if search:
            rgx = {"$regex": re.escape(search), "$options": "i"}
            q["$or"] = [{"product_name": rgx}, {"bb_code": rgx}]
        if category:
            q["product_category"] = category
        if has_images is True:
            q["image_links.0"] = {"$exists": True}
        elif has_images is False:
            q["$or"] = [{"image_links": {"$exists": False}}, {"image_links": []}]

        pipeline = [
            {"$match": q},
            {"$sort": {"bb_code": 1}},
            {"$lookup": {
                "from": PRODUCTS_COLLECTION,
                "localField": "product_id",
                "foreignField": "_id",
                "as": "_prod",
                "pipeline": [{"$project": {"image_url": 1}}],
            }},
            {"$addFields": {
                "product_image_url": {"$arrayElemAt": ["$_prod.image_url", 0]},
            }},
            {"$project": {"_prod": 0}},
        ]
        items = serialize_mongo_document(list(col.aggregate(pipeline)))
        buf = _make_catalogue_xlsx(items)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="catalogue.xlsx"'},
        )
    except PyMongoError as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── Product Information Sheet (PIS) upload ───────────────────────────────────

# Normalise a header string for lookup
def _norm_header(h: str | None) -> str:
    return (h or "").strip().lower()


# Maps normalised header → internal key.
# Keys starting with "_" are handled specially (not direct catalogue fields).
# None means intentionally ignored.
_PIS_COL_MAP: dict[str, str | None] = {
    "bb code": "__bb_code",
    "item name": "product_name",
    "product name": "product_name",
    "hsn": "hsn_code",
    "ean/upc": None,
    "category": "product_category",
    "sub-category": "sub_category",
    "series": "series",
    "mrp": "mrp",
    "gst": "gst_percentage",
    # dimensions parsed separately
    "product weight w/ packaging (g)": "_weight_with",
    "product weight w/o packaging(g)": "_weight_without",
    "product weight w/o packaging(g) ": "_weight_without",
    "product weight w/o packaging(g) - single piece": "_weight_without",
    "product dimensions w/ packaging (cm)": "_dims_with",
    "product dimensions w/o packaging (cm)": "_dims_without",
    "product dimensions w/o packaging (cm) - single piece": "_dims_without",
    "product dimensions w/o packaging (cm) - single piece ": "_dims_without",
    # features
    "features 1": "_feat_1",
    "features 2": "_feat_2",
    "features 3": "_feat_3",
    "features 4": "_feat_4",
    "features 5": "_feat_5",
    # attributes
    "materials": "material",
    "material": "material",
    "squeaker": "_squeaker",
    "catnip (local/imported)": "_catnip",
    "catnip": "_catnip",
    "dog size": "pet_size",
    "dog age": "age_group",
    "chewing strength": "chewing_style",
    "images": "image_links",
    # treats-specific
    "ingredient list %": "ingredient_list",
    "nutrition analysis": "nutritional_analysis",
    # grooming-specific — pet_size/age_group variants
    "dog size / breed / coat": "pet_size",
    "dog size / breed / coat ": "pet_size",
    "dog age (puppy/adult / senior)": "age_group",
    # grooming — composition maps to ingredient_list
    "composition / ingredients": "ingredient_list",
    # grooming — text attributes
    "scented / unscented": "_g_scented",
    "essential oils": "_g_essential_oils",
    # grooming — boolean certifications
    " ifra-certified": "_g_ifra_certified",
    "ifra-certified": "_g_ifra_certified",
    "slicone free": "_g_silicone_free",
    "silicone free": "_g_silicone_free",
    "sls free": "_g_sls_free",
    "paraben-free": "_g_paraben_free",
    "ph balanced": "_g_ph_balanced",
    "hypoallergenic": "_g_hypoallergenic",
    "soap-free": "_g_soap_free",
    "tear-free": "_g_tear_free",
    "dye-free / colorant-free": "_g_dye_free",
    "artificial fragrance-free": "_g_fragrance_free",
    "alcohol-free": "_g_alcohol_free",
    "dea-free": "_g_dea_free",
    "peg-free": "_g_peg_free",
    "phthalate-free": "_g_phthalate_free",
    "formaldehyde-free": "_g_formaldehyde_free",
    "vegan": "_g_vegan",
    "cruelty-free": "_g_cruelty_free",
    "all natural ingredients": "_g_all_natural",
    "dermatologically tested": "_g_dermatologically_tested",
    # explicitly skipped
    "item code (manufacturer)": None,
    "sr no": None,
    "case pack": None,
    "cleaning instructions": None,
    "padding (plush toys)": None,
    "absorption capacity": None,
    "composition": None,
    "shelf life": None,
    "natural ingredients (yes/no)": None,
    "grain free (yes/no)": None,
    "gluten free (yes/no)": None,
    "functional treat (specify the function, ex: dental)": None,
    "human grade ingredients (yes/no)": None,
    "main animal source": None,
    "form used": None,
    "answer": None,
    "answer ": None,
    "source of ingredients ": None,
    "starch source": None,
    "plant protein source": None,
    "glycerin type & source": None,
    "other animal source": None,
    "special additives": None,
    "animal body parts used": None,
    "feeding guide": None,
    "lab test reports (if any)": None,
}


_GROOMING_BOOL_ATTRS: dict[str, str] = {
    "_g_ifra_certified": "ifra_certified",
    "_g_silicone_free": "silicone_free",
    "_g_sls_free": "sls_free",
    "_g_paraben_free": "paraben_free",
    "_g_ph_balanced": "ph_balanced",
    "_g_hypoallergenic": "hypoallergenic",
    "_g_soap_free": "soap_free",
    "_g_tear_free": "tear_free",
    "_g_dye_free": "dye_free",
    "_g_fragrance_free": "fragrance_free",
    "_g_alcohol_free": "alcohol_free",
    "_g_dea_free": "dea_free",
    "_g_peg_free": "peg_free",
    "_g_phthalate_free": "phthalate_free",
    "_g_formaldehyde_free": "formaldehyde_free",
    "_g_vegan": "vegan",
    "_g_cruelty_free": "cruelty_free",
    "_g_all_natural": "all_natural_ingredients",
    "_g_dermatologically_tested": "dermatologically_tested",
}

_GROOMING_TEXT_ATTRS: dict[str, str] = {
    "_g_scented": "scent_type",
    "_g_essential_oils": "essential_oils",
}


def _parse_dims_str(s: Any) -> tuple[float, float, float] | None:
    if not s:
        return None
    parts = re.split(r"\s*[*xX×]\s*", str(s).strip())
    if len(parts) != 3:
        return None
    try:
        return (float(parts[0]), float(parts[1]), float(parts[2]))
    except (ValueError, TypeError):
        return None


def _to_bool(val: Any) -> bool | None:
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ("yes", "y", "1", "true", "local", "imported", "local/imported"):
        return True
    if s in ("no", "n", "0", "false", ""):
        return False
    return None


def _to_float(val: Any) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _parse_image_links(val: Any) -> list[str] | None:
    if not val:
        return None
    raw = str(val).strip()
    if not raw:
        return None
    # could be newline-or-comma-separated
    parts = [p.strip() for p in re.split(r"[\n,]+", raw) if p.strip()]
    return parts if parts else None


def _process_pis_sheet(ws, sheet_name: str, db, dry_run: bool = False) -> dict:
    """Parse one PIS sheet. If dry_run=True, checks matches without writing. Returns result dict."""
    updated = []
    not_found = []
    skipped = []

    headers_raw = [cell.value for cell in ws[1]]
    headers_norm = [_norm_header(h) for h in headers_raw]

    # Build col index → internal key mapping
    col_keys: list[str | None] = []
    for h in headers_norm:
        col_keys.append(_PIS_COL_MAP.get(h, "__unknown__"))

    has_bb_code_col = "__bb_code" in col_keys

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip blank rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        raw: dict[str, Any] = {}
        for col_i, val in enumerate(row):
            key = col_keys[col_i] if col_i < len(col_keys) else "__unknown__"
            if key is None or key == "__unknown__":
                continue
            # For duplicate headers (e.g. two "Category" cols) keep first non-empty
            if key in raw and raw[key] is not None:
                continue
            raw[key] = val

        bb_code: str | None = str(raw.get("__bb_code", "") or "").strip() or None
        product_name: str | None = str(raw.get("product_name", "") or "").strip() or None

        if not bb_code and not product_name:
            skipped.append({"row": row_idx, "sheet": sheet_name, "reason": "No BB code or product name"})
            continue

        # Build catalogue $set payload
        set_fields: dict[str, Any] = {}
        fields_updated: list[str] = []

        # Direct scalar fields
        for key in ("product_name", "product_category", "sub_category", "series",
                    "material", "pet_size", "age_group", "chewing_style",
                    "ingredient_list", "nutritional_analysis", "hsn_code"):
            v = str(raw[key]).strip() if raw.get(key) is not None else None
            if v:
                set_fields[key] = v
                fields_updated.append(key)

        for key in ("mrp", "gst_percentage"):
            v = _to_float(raw.get(key))
            if v is not None:
                set_fields[key] = v
                fields_updated.append(key)

        # Boolean fields
        squeaker = _to_bool(raw.get("_squeaker"))
        if squeaker is not None:
            set_fields["squeaker"] = squeaker
            fields_updated.append("squeaker")

        catnip = _to_bool(raw.get("_catnip"))
        if catnip is not None:
            set_fields["catnip"] = catnip
            fields_updated.append("catnip")

        # Features
        features = [
            str(raw[f"_feat_{i}"]).strip()
            for i in range(1, 6)
            if raw.get(f"_feat_{i}") and str(raw[f"_feat_{i}"]).strip()
        ]
        if features:
            set_fields["features"] = features
            fields_updated.append("features")

        # Image links
        links = _parse_image_links(raw.get("image_links"))
        if links:
            set_fields["image_links"] = links
            fields_updated.append("image_links")

        # Grooming attributes (boolean certifications + text attrs)
        grooming_attrs: dict[str, Any] = {}
        for internal_key, attr_name in _GROOMING_BOOL_ATTRS.items():
            val = _to_bool(raw.get(internal_key))
            if val is not None:
                grooming_attrs[attr_name] = val
        for internal_key, attr_name in _GROOMING_TEXT_ATTRS.items():
            v = str(raw[internal_key]).strip() if raw.get(internal_key) is not None else None
            if v:
                grooming_attrs[attr_name] = v
        if grooming_attrs:
            set_fields["grooming_attributes"] = grooming_attrs
            fields_updated.append("grooming_attributes")

        # Dimensions
        dims_with = _parse_dims_str(raw.get("_dims_with"))
        weight_with = _to_float(raw.get("_weight_with"))
        dims_without = _parse_dims_str(raw.get("_dims_without"))
        weight_without = _to_float(raw.get("_weight_without"))

        with_pkg: dict[str, Any] = {}
        without_pkg: dict[str, Any] = {}
        if dims_with:
            with_pkg.update({"length_cm": dims_with[0], "breadth_cm": dims_with[1], "height_cm": dims_with[2]})
        if weight_with is not None:
            with_pkg["gross_weight_g"] = weight_with
        if dims_without:
            without_pkg.update({"length_cm": dims_without[0], "breadth_cm": dims_without[1], "height_cm": dims_without[2]})
        if weight_without is not None:
            without_pkg["net_weight_g"] = weight_without

        if with_pkg or without_pkg:
            set_fields["dimensions"] = {
                "with_packaging": with_pkg or None,
                "without_packaging": without_pkg or None,
            }
            fields_updated.append("dimensions")

        if not set_fields:
            skipped.append({"row": row_idx, "sheet": sheet_name, "reason": "No mappable data"})
            continue

        # Snapshot values before adding timestamp (used in preview tooltips)
        fields_values: dict[str, Any] = dict(set_fields)

        set_fields["updated_at"] = utcnow()

        # Match catalogue doc
        if bb_code:
            filter_q = {"bb_code": bb_code}
        else:
            filter_q = {"product_name": {"$regex": f"^{re.escape(product_name)}$", "$options": "i"}}

        if dry_run:
            existing = db[DESIGN_CATALOGUE_COLLECTION].find_one(filter_q, {"_id": 1, "product_name": 1})
            matched = existing is not None
        else:
            res = db[DESIGN_CATALOGUE_COLLECTION].update_one(filter_q, {"$set": set_fields})
            matched = res.matched_count > 0

        if not matched:
            not_found.append({
                "identifier": bb_code or product_name,
                "product_name": product_name,
                "sheet": sheet_name,
            })
        else:
            updated.append({
                "bb_code": bb_code or "—",
                "product_name": product_name or set_fields.get("product_name", ""),
                "sheet": sheet_name,
                "fields_updated": fields_updated,
                "fields_values": fields_values,
            })

    return {"updated": updated, "not_found": not_found, "skipped": skipped}


def _parse_pis_workbook(data: bytes, db, dry_run: bool) -> dict:
    """Load and process all sheets from PIS XLSX bytes. Returns combined result dict."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read Excel file")

    all_updated: list[dict] = []
    all_not_found: list[dict] = []
    all_skipped: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue
        result = _process_pis_sheet(ws, sheet_name, db, dry_run=dry_run)
        all_updated.extend(result["updated"])
        all_not_found.extend(result["not_found"])
        all_skipped.extend(result["skipped"])

    return {
        "summary": {
            "total_rows": len(all_updated) + len(all_not_found) + len(all_skipped),
            "updated": len(all_updated),
            "not_found": len(all_not_found),
            "skipped": len(all_skipped),
        },
        "updated": all_updated,
        "not_found": all_not_found,
        "skipped": all_skipped,
    }


def _extract_email_from_token(authorization: str | None) -> str | None:
    """Return email (sub) from Bearer JWT, or None if missing/invalid."""
    if not authorization or not authorization.startswith("Bearer "):
        return None
    token = authorization.split(" ", 1)[1]
    try:
        secret = os.getenv("SECRET_KEY") or ""
        algo = os.getenv("ALGORITHM") or "HS256"
        payload = jwt.decode(token, secret, algorithms=[algo], options={"verify_aud": False, "verify_exp": False})
        return payload.get("sub")
    except (JWTError, Exception):
        return None


@router.post("/new-items/upload-pis")
async def preview_pis(file: UploadFile = File(...)):
    """Dry-run parse of a PIS XLSX — returns what would be updated without writing to DB."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be an Excel (.xlsx) file")

    data = await file.read()
    db = get_database()
    result = await asyncio.to_thread(_parse_pis_workbook, data, db, True)
    return JSONResponse(content=result)


@router.post("/new-items/confirm-pis")
async def confirm_pis(
    file: UploadFile = File(...),
    authorization: str | None = Header(default=None),
):
    """Apply a PIS XLSX to the DB and archive the file to S3 with uploader info."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be an Excel (.xlsx) file")

    email = _extract_email_from_token(authorization)
    data = await file.read()
    db = get_database()

    result = await asyncio.to_thread(_parse_pis_workbook, data, db, False)

    # Upload to S3 for audit trail
    ts = utcnow().strftime("%Y%m%d_%H%M%S")
    safe_email = re.sub(r"[^\w@.\-]", "_", email or "unknown")
    safe_filename = re.sub(r"[^\w.\-]", "_", file.filename)
    s3_key = f"pis_uploads/{safe_email}/{ts}_{safe_filename}"
    try:
        await asyncio.to_thread(_upload_to_s3, data, s3_key, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        result["audit"] = {"uploaded_by": email or "unknown", "uploaded_at": ts, "s3_key": s3_key}
    except Exception as exc:
        logger.warning("PIS S3 upload failed: %s", exc)
        result["audit"] = {"uploaded_by": email or "unknown", "uploaded_at": ts, "s3_key": None}

    return JSONResponse(content=result)


# PIS sheet definitions: (sheet_name, [column_headers])
_PIS_SHEETS: list[tuple[str, list[str]]] = [
    ("Toys", [
        "Item Code (Manufacturer)", "BB code", "Item Name", "HSN", "EAN/UPC",
        "Category", "Sub-Category", "Series",
        "Product weight w/ packaging (g)", "Product weight w/o packaging(g)",
        "Product dimensions w/ packaging (cm)", "Product dimensions w/o packaging (cm)",
        "Features 1", "Features 2", "Features 3", "Features 4", "Features 5",
        "Materials", "Cleaning Instructions", "Padding (Plush Toys)",
        "Squeaker", "Catnip (Local/Imported)",
        "Dog Size", "Dog Age", "Chewing Strength", "Images",
    ]),
    ("Hygeine", [
        "Item Code (Manufacturer)", "BB code", "Item Name", "HSN", "EAN/UPC",
        "Category", "Sub-Category", "Series",
        "Product weight w/ packaging (g)", "Product weight w/o packaging(g)",
        "Product dimensions w/ packaging (cm)", "Product dimensions w/o packaging (cm) - Single Piece",
        "Product weight w/o packaging(g) - Single Piece",
        "Absorption Capacity", "Composition", "Scented / Unscented", "Shelf Life",
        "Features 1", "Features 2", "Features 3", "Features 4", "Features 5",
        "Materials", "Dog Size", "Images",
    ]),
    ("Outdoor Gear", [
        "Item Code (Manufacturer)", "BB code", "Item Name", "HSN", "EAN/UPC",
        "Category", "Sub-Category", "Series",
        "Product weight w/ packaging (g)", "Product weight w/o packaging(g)",
        "Product dimensions w/ packaging (cm)", "Product dimensions w/o packaging (cm)",
        "Features 1", "Features 2", "Features 3", "Features 4", "Features 5",
        "Materials", "Cleaning Instructions", "Images",
    ]),
    ("Treats", [
        "Sr No", "Item Code (Manufacturer)", "BB code", "Item Name", "HSN", "EAN/UPC",
        "Category", "Sub-Category", "Series",
        "Product weight w/ packaging (g)", "Product weight w/o packaging(g)",
        "Product dimensions w/ packaging (cm)", "Product dimensions w/o packaging (cm) - Single piece",
        "Product weight w/o packaging(g) - Single Piece",
        "Ingredient List %", "Nutrition Analysis",
        "Natural Ingredients (Yes/No)", "Grain Free (Yes/No)", "Gluten Free (Yes/No)",
        "Functional Treat (Specify the function, ex: Dental)", "Human Grade Ingredients (Yes/No)",
        "Main Animal Source", "Form Used",
        "Shelf Life", "feeding guide", "Images", "Lab Test Reports (If Any)",
    ]),
    ("Grooming", [
        "Item Code (Manufacturer)", "BB code", "Item Name", "HSN", "EAN/UPC",
        "Category", "Sub-Category", "Series",
        "Product weight w/ packaging (g)", "Product weight w/o packaging(g) ",
        "Product dimensions w/ packaging (cm)",
        "Composition / Ingredients", "Scented / Unscented", " IFRA-certified",
        "Slicone Free", "SLS Free", "Paraben-Free", "pH Balanced", "Hypoallergenic",
        "Soap-Free", "Tear-Free", "Dye-Free / Colorant-Free", "Artificial Fragrance-Free",
        "Alcohol-Free", "DEA-Free", "PEG-Free", "Phthalate-Free", "Formaldehyde-Free",
        "Vegan", "Cruelty-Free", "All Natural Ingredients", "Essential Oils",
        "Shelf Life",
        "Features 1", "Features 2", "Features 3", "Features 4", "Features 5",
        "Materials", "Dog Size / Breed / Coat ", "Dog Age (Puppy/Adult / Senior)",
        "Dermatologically Tested", "Images",
    ]),
]


@router.get("/new-items/pis-template")
def download_pis_template():
    """Return an empty Product Information Sheet XLSX template with all column headers."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for sheet_name, headers in _PIS_SHEETS:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
        _style_header_row(ws, 1, len(headers))
        ws.freeze_panes = "A2"
        _auto_col_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="PIS_Template.xlsx"'},
    )


# ─── designer orders ──────────────────────────────────────────────────────────

@router.get("/documents/search")
async def search_designer_documents(q: str, db=Depends(get_database)):
    """Search designer_documents filenames across all brand orders."""
    def _search():
        pipeline = [
            {"$unwind": "$designer_documents"},
            {"$match": {"designer_documents.filename": {"$regex": re.escape(q.strip()), "$options": "i"}}},
            {"$project": {
                "order_id": {"$toString": "$_id"},
                "order_name": "$name",
                "brand": "$brand",
                "doc": "$designer_documents",
            }},
            {"$limit": 25},
        ]
        return list(db[BRAND_ORDERS_COLLECTION].aggregate(pipeline))

    results = await asyncio.to_thread(_search)
    return serialize_mongo_document(results)


@router.get("/vendor-brands")
async def get_vendor_brands(db=Depends(get_database)):
    """Return a map of vendor_id → [brand names] from the brands collection."""
    def _fetch():
        result: dict[str, list[str]] = {}
        for brand in db[BRANDS_COLLECTION].find({"vendor_id": {"$exists": True, "$ne": None}}, {"name": 1, "vendor_id": 1}):
            vid = brand.get("vendor_id")
            name = brand.get("name")
            if not vid or not name:
                continue
            result.setdefault(vid, [])
            if name not in result[vid]:
                result[vid].append(name)
        for vid in result:
            result[vid].sort()
        return result
    return await asyncio.to_thread(_fetch)


@router.get("/orders")
async def list_designer_orders(db=Depends(get_database)):
    """Return all brand orders with designer_documents, omitting purchase team documents."""
    def _fetch():
        pipeline = [
            {"$addFields": {"doc_count": {"$size": {"$ifNull": ["$designer_documents", []]}}}},
            {"$project": {"documents": 0}},
            {"$lookup": {
                "from": PO_COLLECTION,
                "localField": "purchaseorder_number",
                "foreignField": "purchaseorder_number",
                "as": "_po",
            }},
            {"$lookup": {
                "from": VENDORS_COLLECTION,
                "localField": "vendor_id",
                "foreignField": "contact_id",
                "as": "_vendor",
            }},
            {"$addFields": {
                "vendor_name": {"$arrayElemAt": ["$_vendor.contact_name", 0]},
                "po_status": {
                    "$let": {
                        "vars": {
                            "fmt": {"$arrayElemAt": ["$_po.order_status_formatted", 0]},
                            "raw": {"$arrayElemAt": ["$_po.order_status", 0]},
                        },
                        "in": {
                            "$cond": {
                                "if": {"$gt": [{"$strLenCP": {"$ifNull": ["$$fmt", ""]}}, 0]},
                                "then": "$$fmt",
                                "else": {
                                    "$cond": {
                                        "if": {"$gt": [{"$strLenCP": {"$ifNull": ["$$raw", ""]}}, 0]},
                                        "then": {"$concat": [
                                            {"$toUpper": {"$substrCP": ["$$raw", 0, 1]}},
                                            {"$substrCP": ["$$raw", 1, {"$subtract": [{"$strLenCP": "$$raw"}, 1]}]},
                                        ]},
                                        "else": None,
                                    }
                                },
                            }
                        },
                    }
                },
            }},
            {"$project": {"_vendor": 0, "_po": 0}},
            {"$addFields": {
                "_sort_num": {
                    "$convert": {
                        "input": {"$trim": {"input": {"$arrayElemAt": [{"$split": ["$name", "#"]}, 1]}}},
                        "to": "int",
                        "onError": 0,
                        "onNull": 0,
                    }
                }
            }},
            {"$sort": {"brand": 1, "_sort_num": -1}},
            {"$project": {"_sort_num": 0}},
        ]
        return list(db[BRAND_ORDERS_COLLECTION].aggregate(pipeline))

    orders = await asyncio.to_thread(_fetch)
    return serialize_mongo_document(orders)


@router.get("/categories")
async def get_designer_categories(db=Depends(get_database)):
    def _fetch():
        cats = list(db[CATEGORIES_COLLECTION].find({}, {"name": 1, "_id": 0}))
        if not cats:
            db[CATEGORIES_COLLECTION].insert_many([{"name": n} for n in DEFAULT_CATEGORIES])
            return DEFAULT_CATEGORIES
        return [c["name"] for c in cats]
    return await asyncio.to_thread(_fetch)


@router.post("/categories")
async def add_designer_category(request: AddCategoryRequest, db=Depends(get_database)):
    name = request.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Category name required")
    def _add():
        if not db[CATEGORIES_COLLECTION].find_one({"name": name}):
            db[CATEGORIES_COLLECTION].insert_one({"name": name})
    await asyncio.to_thread(_add)
    return {"name": name}


@router.delete("/categories/{name}")
async def delete_designer_category(name: str, db=Depends(get_database)):
    def _delete():
        result = db[CATEGORIES_COLLECTION].delete_one({"name": name})
        return result.deleted_count
    deleted = await asyncio.to_thread(_delete)
    if not deleted:
        raise HTTPException(status_code=404, detail="Category not found")
    return {"name": name, "deleted": True}


@router.patch("/categories/{name}")
async def rename_designer_category(name: str, request: Request, db=Depends(get_database)):
    body = await request.json()
    new_name = (body.get("new_name") or "").strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="new_name required")
    def _rename():
        if db[CATEGORIES_COLLECTION].find_one({"name": new_name}):
            raise HTTPException(status_code=409, detail="Category already exists")
        result = db[CATEGORIES_COLLECTION].update_one({"name": name}, {"$set": {"name": new_name}})
        if not result.matched_count:
            raise HTTPException(status_code=404, detail="Category not found")
        # Update all documents in orders that used the old category name
        db[BRAND_ORDERS_COLLECTION].update_many(
            {"designer_documents.category": name},
            {"$set": {"designer_documents.$[elem].category": new_name}},
            array_filters=[{"elem.category": name}]
        )
    await asyncio.to_thread(_rename)
    return {"old_name": name, "new_name": new_name}


@router.get("/{order_id}/line-items")
async def get_order_line_items(order_id: str, db=Depends(get_database)):
    def _fetch():
        order = db[BRAND_ORDERS_COLLECTION].find_one(
            {"_id": ObjectId(order_id)}, {"purchaseorder_number": 1}
        )
        if not order or not order.get("purchaseorder_number"):
            return []
        po = db[PO_COLLECTION].find_one(
            {"purchaseorder_number": order["purchaseorder_number"]},
            {"line_items": 1, "date": 1}
        )
        if not po:
            return []
        line_items = po.get("line_items", [])
        po_date = po.get("date")
        if isinstance(po_date, str):
            try:
                po_date = datetime.strptime(po_date, "%Y-%m-%d")
            except ValueError:
                po_date = None
        if po_date and line_items:
            item_ids = [li.get("item_id") for li in line_items if li.get("item_id")]
            products = {
                str(p["item_id"]): p.get("created_at")
                for p in db["products"].find(
                    {"item_id": {"$in": item_ids}},
                    {"item_id": 1, "created_at": 1, "_id": 0},
                )
                if p.get("item_id")
            }
            for li in line_items:
                prod_created = products.get(str(li.get("item_id")))
                if isinstance(prod_created, datetime) and isinstance(po_date, datetime):
                    li["is_new"] = abs((prod_created - po_date).days) <= 30
                else:
                    li["is_new"] = False
        return line_items
    items = await asyncio.to_thread(_fetch)
    return serialize_mongo_document(items)


# ─── designer folders ─────────────────────────────────────────────────────────

@router.get("/{order_id}/designer-folders")
async def list_designer_folders(order_id: str, db=Depends(get_database)):
    def _fetch():
        doc = db[BRAND_ORDERS_COLLECTION].find_one({"_id": ObjectId(order_id)}, {"designer_folders": 1})
        if not doc:
            raise ValueError("Order not found")
        return doc.get("designer_folders", [])
    try:
        folders = await asyncio.to_thread(_fetch)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return serialize_mongo_document(folders)


@router.post("/{order_id}/designer-folders")
async def create_designer_folder(order_id: str, body: CreateFolderRequest, db=Depends(get_database)):
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Folder name required")
    parent = (body.parent_path or "").strip().strip("/")
    safe_name = _sanitize_folder_name(name)
    path = f"{parent}/{safe_name}" if parent else safe_name

    def _create():
        doc = db[BRAND_ORDERS_COLLECTION].find_one({"_id": ObjectId(order_id)}, {"designer_folders": 1})
        if not doc:
            raise ValueError("Order not found")
        if any(f.get("path") == path for f in doc.get("designer_folders", [])):
            raise ValueError(f"A folder named '{name}' already exists here")
        folder_entry = {
            "folder_id": str(ObjectId()),
            "name": name,
            "path": path,
            "parent_path": parent,
            "created_at": datetime.now(),
        }
        db[BRAND_ORDERS_COLLECTION].update_one(
            {"_id": ObjectId(order_id)},
            {"$push": {"designer_folders": folder_entry}, "$set": {"updated_at": datetime.now()}},
        )
        return folder_entry

    try:
        folder = await asyncio.to_thread(_create)
    except ValueError as e:
        code = 409 if "already exists" in str(e) else 404
        raise HTTPException(status_code=code, detail=str(e))
    return serialize_mongo_document(folder)


@router.patch("/{order_id}/designer-folders/{folder_id}")
async def rename_designer_folder(order_id: str, folder_id: str, body: RenameFolderRequest, db=Depends(get_database)):
    new_name = body.name.strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="Folder name required")

    def _rename():
        doc = db[BRAND_ORDERS_COLLECTION].find_one(
            {"_id": ObjectId(order_id)}, {"designer_folders": 1, "designer_documents": 1}
        )
        if not doc:
            raise ValueError("Order not found")
        folders = doc.get("designer_folders", [])
        folder = next((f for f in folders if f.get("folder_id") == folder_id), None)
        if not folder:
            raise ValueError("Folder not found")

        old_path = folder["path"]
        parent = folder.get("parent_path", "")
        safe_name = _sanitize_folder_name(new_name)
        new_path = f"{parent}/{safe_name}" if parent else safe_name

        if new_path == old_path:
            return {**folder, "name": new_name, "path": new_path}

        db[BRAND_ORDERS_COLLECTION].update_one(
            {"_id": ObjectId(order_id), "designer_folders.folder_id": folder_id},
            {"$set": {"designer_folders.$.name": new_name, "designer_folders.$.path": new_path, "updated_at": datetime.now()}},
        )
        for f in folders:
            f_path = f.get("path", "")
            if f_path.startswith(old_path + "/"):
                new_child_path = new_path + f_path[len(old_path):]
                new_parent = f.get("parent_path", "")
                if new_parent.startswith(old_path):
                    new_parent = new_path + new_parent[len(old_path):]
                db[BRAND_ORDERS_COLLECTION].update_one(
                    {"_id": ObjectId(order_id), "designer_folders.folder_id": f["folder_id"]},
                    {"$set": {"designer_folders.$.path": new_child_path, "designer_folders.$.parent_path": new_parent}},
                )
        for d in doc.get("designer_documents", []):
            d_folder = d.get("folder", "")
            if d_folder == old_path:
                db[BRAND_ORDERS_COLLECTION].update_one(
                    {"_id": ObjectId(order_id), "designer_documents.doc_id": d["doc_id"]},
                    {"$set": {"designer_documents.$.folder": new_path}},
                )
            elif d_folder.startswith(old_path + "/"):
                db[BRAND_ORDERS_COLLECTION].update_one(
                    {"_id": ObjectId(order_id), "designer_documents.doc_id": d["doc_id"]},
                    {"$set": {"designer_documents.$.folder": new_path + d_folder[len(old_path):]}},
                )
        return {**folder, "name": new_name, "path": new_path}

    try:
        result = await asyncio.to_thread(_rename)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return serialize_mongo_document(result)


@router.delete("/{order_id}/designer-folders/{folder_id}")
async def delete_designer_folder(order_id: str, folder_id: str, db=Depends(get_database)):
    """Delete a designer folder (and subfolders); documents inside are moved to parent."""
    def _delete():
        doc = db[BRAND_ORDERS_COLLECTION].find_one(
            {"_id": ObjectId(order_id)}, {"designer_folders": 1, "designer_documents": 1}
        )
        if not doc:
            raise ValueError("Order not found")
        folders = doc.get("designer_folders", [])
        folder = next((f for f in folders if f.get("folder_id") == folder_id), None)
        if not folder:
            raise ValueError("Folder not found")

        path = folder["path"]
        parent_path = folder.get("parent_path", "")

        for d in doc.get("designer_documents", []):
            d_folder = d.get("folder", "")
            if d_folder == path or d_folder.startswith(path + "/"):
                db[BRAND_ORDERS_COLLECTION].update_one(
                    {"_id": ObjectId(order_id), "designer_documents.doc_id": d["doc_id"]},
                    {"$set": {"designer_documents.$.folder": parent_path}},
                )

        ids_to_remove = [folder_id] + [
            f["folder_id"] for f in folders if f.get("path", "").startswith(path + "/")
        ]
        db[BRAND_ORDERS_COLLECTION].update_one(
            {"_id": ObjectId(order_id)},
            {
                "$pull": {"designer_folders": {"folder_id": {"$in": ids_to_remove}}},
                "$set": {"updated_at": datetime.now()},
            },
        )
        return True

    try:
        await asyncio.to_thread(_delete)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return {"folder_id": folder_id, "deleted": True}


@router.post("/{order_id}/designer-documents")
async def upload_designer_document(
    order_id: str,
    file: UploadFile = File(...),
    relative_path: Optional[str] = Form(None),
    category: Optional[str] = Form(None),
    folder: Optional[str] = Form(None),
    item_id: Optional[str] = Form(None),
    item_name: Optional[str] = Form(None),
    db=Depends(get_database),
):
    file_bytes = await file.read()
    filename = file.filename or "file"
    content_type = file.content_type or "application/octet-stream"
    cat = (category or "general").strip()
    folder_path = (folder or "").strip().strip("/")

    def _process():
        doc = db[BRAND_ORDERS_COLLECTION].find_one(
            {"_id": ObjectId(order_id)}, {"brand": 1, "name": 1}
        )
        if not doc:
            raise ValueError("Order not found")

        brand_slug = _slugify(doc["brand"])
        name_slug = _slugify(doc["name"])
        cat_slug = _slugify(cat)

        if relative_path:
            parts = relative_path.replace("\\", "/").strip("/").split("/")
            safe_parts = [re.sub(r"[^\w.\-]", "_", p) for p in parts if p]
            if folder_path:
                safe_folder = "/".join(re.sub(r"[^\w.\-]", "_", p) for p in folder_path.split("/") if p)
                s3_key = f"designer_uploads/{brand_slug}/{name_slug}/{safe_folder}/{'/'.join(safe_parts)}"
            else:
                s3_key = f"designer_uploads/{brand_slug}/{name_slug}/{cat_slug}/{'/'.join(safe_parts)}"
        else:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_filename = re.sub(r"[^\w.\-]", "_", filename)
            if folder_path:
                safe_folder = "/".join(re.sub(r"[^\w.\-]", "_", p) for p in folder_path.split("/") if p)
                s3_key = f"designer_uploads/{brand_slug}/{name_slug}/{safe_folder}/{ts}_{safe_filename}"
            else:
                s3_key = f"designer_uploads/{brand_slug}/{name_slug}/{cat_slug}/{ts}_{safe_filename}"

        _upload_to_s3(file_bytes, s3_key, content_type)

        doc_entry: dict = {
            "doc_id": str(ObjectId()),
            "filename": filename,
            "s3_key": s3_key,
            "content_type": content_type,
            "size": len(file_bytes),
            "uploaded_at": datetime.now(),
            "category": cat,
            "folder": folder_path,
        }
        if item_id:
            doc_entry["item_id"] = item_id
        if item_name:
            doc_entry["item_name"] = item_name

        db[BRAND_ORDERS_COLLECTION].update_one(
            {"_id": ObjectId(order_id)},
            {
                "$push": {"designer_documents": doc_entry},
                "$set": {"updated_at": datetime.now()},
            },
        )
        return doc_entry

    try:
        doc_entry = await asyncio.to_thread(_process)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.exception("upload_designer_document failed")
        raise HTTPException(status_code=500, detail=str(e))

    return serialize_mongo_document(doc_entry)


@router.get("/{order_id}/designer-documents/zip")
async def download_designer_documents_zip(
    order_id: str,
    item_id: Optional[str] = Query(None),
    db=Depends(get_database),
):
    def _fetch_docs():
        doc = db[BRAND_ORDERS_COLLECTION].find_one(
            {"_id": ObjectId(order_id)},
            {"designer_documents": 1, "name": 1, "brand": 1},
        )
        if not doc:
            raise ValueError("Order not found")
        return doc

    try:
        order_doc = await asyncio.to_thread(_fetch_docs)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    documents = order_doc.get("designer_documents", [])
    if item_id:
        documents = [d for d in documents if d.get("item_id") == item_id]
    if not documents:
        raise HTTPException(status_code=404, detail="No designer documents to download")

    def _build_zip():
        s3 = boto3.client("s3", region_name=AWS_REGION)
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for d in documents:
                s3_key = d.get("s3_key")
                if not s3_key:
                    continue
                try:
                    obj = s3.get_object(Bucket=S3_BUCKET, Key=s3_key)
                    data = obj["Body"].read()
                    arcname = d.get("filename") or s3_key.split("/")[-1]
                    zf.writestr(arcname, data)
                except Exception:
                    logger.warning("Skipping %s in zip — download failed", s3_key)
        buf.seek(0)
        return buf

    zip_buf = await asyncio.to_thread(_build_zip)
    safe_name = re.sub(r"[^\w.\-]", "_", order_doc.get("name", order_id))
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_designer.zip"'},
    )


@router.patch("/{order_id}/designer-documents/{doc_id}")
async def update_designer_document(order_id: str, doc_id: str, body: dict, db=Depends(get_database)):
    updates: dict = {}
    category = (body.get("category") or "").strip()
    if category:
        updates["designer_documents.$.category"] = category
    if "folder" in body:
        updates["designer_documents.$.folder"] = (body["folder"] or "").strip().strip("/")
    if not updates:
        raise HTTPException(status_code=400, detail="Nothing to update — provide category or folder")
    updates["updated_at"] = datetime.now()

    def _update():
        result = db[BRAND_ORDERS_COLLECTION].update_one(
            {"_id": ObjectId(order_id), "designer_documents.doc_id": doc_id},
            {"$set": updates},
        )
        return result.matched_count

    matched = await asyncio.to_thread(_update)
    if not matched:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"doc_id": doc_id, **{k.replace("designer_documents.$.", ""): v for k, v in updates.items() if k != "updated_at"}}


@router.get("/{order_id}/designer-documents/{doc_id}/url")
async def get_designer_document_url(order_id: str, doc_id: str, db=Depends(get_database)):
    def _fetch():
        doc = db[BRAND_ORDERS_COLLECTION].find_one(
            {"_id": ObjectId(order_id), "designer_documents.doc_id": doc_id},
            {"designer_documents.$": 1},
        )
        if not doc or not doc.get("designer_documents"):
            return None
        return doc["designer_documents"][0].get("s3_key")

    s3_key = await asyncio.to_thread(_fetch)
    if not s3_key:
        raise HTTPException(status_code=404, detail="Document not found")

    try:
        url = await asyncio.to_thread(_presign_s3, s3_key)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    return {"url": url}


@router.delete("/{order_id}/designer-documents/{doc_id}")
async def delete_designer_document(order_id: str, doc_id: str, db=Depends(get_database)):
    def _delete():
        doc = db[BRAND_ORDERS_COLLECTION].find_one(
            {"_id": ObjectId(order_id), "designer_documents.doc_id": doc_id},
            {"designer_documents.$": 1},
        )
        if not doc or not doc.get("designer_documents"):
            return False
        s3_key = doc["designer_documents"][0].get("s3_key")
        if s3_key:
            try:
                _delete_from_s3(s3_key)
            except Exception:
                pass
        db[BRAND_ORDERS_COLLECTION].update_one(
            {"_id": ObjectId(order_id)},
            {
                "$pull": {"designer_documents": {"doc_id": doc_id}},
                "$set": {"updated_at": datetime.now()},
            },
        )
        return True

    found = await asyncio.to_thread(_delete)
    if not found:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"order_id": order_id, "doc_id": doc_id, "deleted": True}


# ─── Brand Catalogues (shared with order-form `catalogues` collection) ────────

CATALOGUES_COLLECTION = "catalogues"


@router.get("/brand-catalogues")
async def list_brand_catalogues(
    page: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db=Depends(get_database),
):
    def _fetch():
        pipeline = [
            {"$sort": {"created_at": -1, "_id": -1}},
            {"$skip": page * limit},
            {"$limit": limit},
        ]
        total = db[CATALOGUES_COLLECTION].count_documents({})
        docs = [serialize_mongo_document(d) for d in db[CATALOGUES_COLLECTION].aggregate(pipeline)]
        return {"catalogues": docs, "total": total, "page": page, "limit": limit}

    return await asyncio.to_thread(_fetch)


@router.post("/brand-catalogues/upload")
async def upload_brand_catalogue(file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are allowed.")
    key = f"catalogues/{uuid.uuid4()}.pdf"
    data = await file.read()
    _public_s3_client().put_object(
        Bucket=PUBLIC_S3_BUCKET,
        Key=key,
        Body=data,
        ContentType="application/pdf",
    )
    return {"file_url": f"{PUBLIC_S3_URL}/{key}"}


class BrandCatalogueBody(BaseModel):
    name: str
    brands: Optional[List[str]] = None
    image_url: str
    is_active: bool = True


@router.post("/brand-catalogues")
async def create_brand_catalogue(body: BrandCatalogueBody, db=Depends(get_database)):
    def _insert():
        doc = {**body.model_dump(), "created_at": datetime.now()}
        result = db[CATALOGUES_COLLECTION].insert_one(doc)
        return serialize_mongo_document(db[CATALOGUES_COLLECTION].find_one({"_id": result.inserted_id}))

    return await asyncio.to_thread(_insert)


@router.put("/brand-catalogues/{catalogue_id}")
async def update_brand_catalogue(catalogue_id: str, body: BrandCatalogueBody, db=Depends(get_database)):
    def _update():
        update = {**body.model_dump(), "updated_at": datetime.now()}
        result = db[CATALOGUES_COLLECTION].update_one(
            {"_id": ObjectId(catalogue_id)}, {"$set": update}
        )
        if result.matched_count == 0:
            return None
        return serialize_mongo_document(db[CATALOGUES_COLLECTION].find_one({"_id": ObjectId(catalogue_id)}))

    doc = await asyncio.to_thread(_update)
    if doc is None:
        raise HTTPException(status_code=404, detail="Catalogue not found")
    return doc


@router.delete("/brand-catalogues/{catalogue_id}")
async def toggle_brand_catalogue(catalogue_id: str, db=Depends(get_database)):
    def _toggle():
        doc = db[CATALOGUES_COLLECTION].find_one({"_id": ObjectId(catalogue_id)})
        if not doc:
            return None
        new_state = not doc.get("is_active", True)
        db[CATALOGUES_COLLECTION].update_one(
            {"_id": ObjectId(catalogue_id)}, {"$set": {"is_active": new_state}}
        )
        return new_state

    result = await asyncio.to_thread(_toggle)
    if result is None:
        raise HTTPException(status_code=404, detail="Catalogue not found")
    return {"is_active": result}
