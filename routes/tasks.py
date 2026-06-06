import asyncio
import io
import logging
import os
import re
import zipfile
from datetime import datetime, timedelta
from ..helpers.datetime_utils import utcnow
from typing import List, Optional

import boto3
import openpyxl
from botocore.config import Config as BotocoreConfig
from bson import ObjectId
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel
from pymongo.errors import PyMongoError

from ..database import get_database, serialize_mongo_document
from ..helpers.scheduler import send_task_assignment_notification

router = APIRouter()
logger = logging.getLogger(__name__)

TASKS_COLLECTION = "tasks"
NOTIFICATIONS_COLLECTION = "notifications"
S3_BUCKET = os.getenv("S3_BUCKET", "pupscribe-purchases")
AWS_REGION = os.getenv("AWS_REGION", "ap-south-1")
# Pre-signed URLs valid for 7 days — only the s3_key is persisted in MongoDB
ATTACHMENT_URL_EXPIRY = 7 * 24 * 3600

VALID_PRIORITIES = {"urgent", "high", "medium", "low"}
VALID_STATUSES = {"todo", "in_progress", "review", "done"}
PRIORITY_ORDER = {"urgent": 0, "high": 1, "medium": 2, "low": 3}


# ── Pydantic models ───────────────────────────────────────────────────────────

class CreateTaskRequest(BaseModel):
    title: str
    description: Optional[str] = ""
    priority: Optional[str] = "medium"
    status: Optional[str] = "todo"
    assigned_to: Optional[List[str]] = []
    assigned_to_names: Optional[List[str]] = []
    assigned_to_departments: Optional[List[str]] = []
    deadline: Optional[str] = None
    tags: Optional[List[str]] = []
    created_by: str
    created_by_name: str
    creator_department: Optional[str] = None


class UpdateTaskRequest(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    priority: Optional[str] = None
    status: Optional[str] = None
    assigned_to: Optional[List[str]] = None
    assigned_to_names: Optional[List[str]] = None
    assigned_to_departments: Optional[List[str]] = None
    deadline: Optional[str] = None
    tags: Optional[List[str]] = None
    is_hidden: Optional[bool] = None
    actor_id: Optional[str] = None
    actor_name: Optional[str] = None
    notify_assignees: Optional[bool] = False


class AddCommentRequest(BaseModel):
    text: str
    author_id: str
    author_name: str


# ── S3 helpers ────────────────────────────────────────────────────────────────

def _s3_client():
    return boto3.client(
        "s3",
        region_name=AWS_REGION,
        endpoint_url=f"https://s3.{AWS_REGION}.amazonaws.com",
        config=BotocoreConfig(signature_version="s3v4"),
    )


def _presign_s3(s3_key: str, expires: int = ATTACHMENT_URL_EXPIRY, download: bool = False, filename: str = "") -> str:
    """Generate a pre-signed GET URL valid for `expires` seconds (default 7 days).
    If download=True, adds Content-Disposition: attachment so the browser downloads the file."""
    params: dict = {"Bucket": S3_BUCKET, "Key": s3_key}
    if download:
        safe_name = filename or s3_key.split("/")[-1]
        params["ResponseContentDisposition"] = f'attachment; filename="{safe_name}"'
    return _s3_client().generate_presigned_url(
        "get_object",
        Params=params,
        ExpiresIn=expires,
    )


def _upload_to_s3(file_bytes: bytes, s3_key: str, content_type: str) -> None:
    boto3.client("s3", region_name=AWS_REGION).put_object(
        Bucket=S3_BUCKET,
        Key=s3_key,
        Body=file_bytes,
        ContentType=content_type,
    )


def _delete_from_s3(s3_key: str) -> None:
    boto3.client("s3", region_name=AWS_REGION).delete_object(
        Bucket=S3_BUCKET, Key=s3_key
    )


# ── Activity helper ───────────────────────────────────────────────────────────

def _activity_entry(
    activity_type: str,
    actor_id: str,
    actor_name: str,
    field: Optional[str] = None,
    old_value: Optional[str] = None,
    new_value: Optional[str] = None,
    detail: Optional[str] = None,
) -> dict:
    return {
        "activity_id": str(ObjectId()),
        "type": activity_type,
        "actor_id": actor_id,
        "actor_name": actor_name,
        "field": field,
        "old_value": old_value,
        "new_value": new_value,
        "detail": detail,
        "timestamp": utcnow().isoformat(),
    }


def _fan_out_sync(
    db,
    task_id: str,
    task_title: str,
    activity_id: str,
    activity_type: str,
    actor_id: str,
    actor_name: str,
    recipient_ids: list,
    snippet: str,
) -> None:
    """Insert one notification per unique recipient, excluding the actor."""
    now = utcnow()
    seen = set()
    docs = []
    for uid in recipient_ids:
        if not uid or uid == actor_id or uid in seen:
            continue
        seen.add(uid)
        docs.append({
            "user_id": uid,
            "source": "task",
            "task_id": task_id,
            "task_title": task_title,
            "activity_id": activity_id,
            "type": activity_type,
            "actor_name": actor_name,
            "snippet": snippet,
            "read": False,
            "created_at": now,
        })
    if docs:
        db[NOTIFICATIONS_COLLECTION].insert_many(docs)


def _visibility_filter(viewer_id: Optional[str], viewer_role: Optional[str]) -> dict:
    """Return a MongoDB match fragment enforcing per-role visibility rules.
    Admin/manager see everything; regular users only see tasks they created or are assigned to.
    If viewer_id is absent (e.g. old Design Tasks callers), no filter is applied.
    """
    if not viewer_id or viewer_role in ("admin", "manager"):
        return {}
    return {"$or": [{"assigned_to": viewer_id}, {"created_by": viewer_id}]}


# ── List tasks ────────────────────────────────────────────────────────────────

@router.get("")
async def list_tasks(
    viewer_id: Optional[str] = Query(None),
    viewer_role: Optional[str] = Query(None),
    status_filter: Optional[str] = Query(None, alias="status"),
    priority: Optional[str] = Query(None),
    assigned_to: Optional[str] = Query(None),
    department: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: Optional[str] = Query("created_at", description="created_at | updated_at | deadline | priority | title"),
    sort_dir: Optional[str] = Query("desc", description="asc | desc"),
    show_hidden: bool = Query(False),
    db=Depends(get_database),
):
    def _fetch():
        query: dict = {**_visibility_filter(viewer_id, viewer_role), "is_deleted": {"$ne": True}}
        if status_filter and status_filter in VALID_STATUSES:
            query["status"] = status_filter
        if priority and priority in VALID_PRIORITIES:
            query["priority"] = priority
        if assigned_to:
            query["assigned_to"] = assigned_to
        if department:
            dept_or = [
                {"creator_department": department},
                {"assigned_to_departments": department},
            ]
            existing_or = query.pop("$or", None)
            if existing_or:
                query["$and"] = [{"$or": existing_or}, {"$or": dept_or}]
            else:
                query["$or"] = dept_or
        if search:
            search_or = [
                {"title": {"$regex": re.escape(search), "$options": "i"}},
                {"description": {"$regex": re.escape(search), "$options": "i"}},
                {"tags": {"$regex": re.escape(search), "$options": "i"}},
            ]
            existing_or = query.pop("$or", None)
            if existing_or:
                query["$and"] = [{"$or": existing_or}, {"$or": search_or}]
            else:
                query["$or"] = search_or

        if not show_hidden:
            two_days_ago = (utcnow() - timedelta(days=2)).isoformat()
            hide_extra = [
                # Exclude manually hidden tasks
                {"is_hidden": {"$ne": True}},
                # Exclude done tasks whose updated_at is older than 2 days (auto-hidden)
                {"$or": [{"status": {"$ne": "done"}}, {"updated_at": {"$gt": two_days_ago}}]},
            ]
            if "$and" in query:
                query["$and"].extend(hide_extra)
            else:
                query["$and"] = hide_extra

        direction = -1 if sort_dir == "desc" else 1

        # Priority sort needs special handling (urgent > high > medium > low)
        if sort_by == "priority":
            pipeline = [
                {"$match": query},
                {"$addFields": {
                    "_priority_order": {
                        "$switch": {
                            "branches": [
                                {"case": {"$eq": ["$priority", "urgent"]}, "then": 0},
                                {"case": {"$eq": ["$priority", "high"]}, "then": 1},
                                {"case": {"$eq": ["$priority", "medium"]}, "then": 2},
                                {"case": {"$eq": ["$priority", "low"]}, "then": 3},
                            ],
                            "default": 4,
                        }
                    }
                }},
                {"$sort": {"_priority_order": direction, "created_at": -1}},
                {"$project": {"_priority_order": 0}},
            ]
            return list(db[TASKS_COLLECTION].aggregate(pipeline))

        sort_field = sort_by if sort_by in ("created_at", "updated_at", "deadline", "title") else "created_at"
        return list(db[TASKS_COLLECTION].find(query).sort(sort_field, direction))

    tasks = await asyncio.to_thread(_fetch)
    return serialize_mongo_document(tasks)


# ── Stats (admin breakdown) ───────────────────────────────────────────────────

@router.get("/stats")
async def get_task_stats(
    viewer_id: Optional[str] = Query(None),
    viewer_role: Optional[str] = Query(None),
    db=Depends(get_database),
):
    vis = {**_visibility_filter(viewer_id, viewer_role), "is_deleted": {"$ne": True}}

    def _fetch():
        now = utcnow().isoformat()
        pipeline = [
            {"$match": vis},
            {
                "$facet": {
                    "by_status": [
                        {"$group": {"_id": "$status", "count": {"$sum": 1}}},
                    ],
                    "by_priority": [
                        {"$group": {"_id": "$priority", "count": {"$sum": 1}}},
                    ],
                    "by_department": [
                        {"$group": {"_id": "$creator_department", "count": {"$sum": 1}}},
                    ],
                    "overdue": [
                        {
                            "$match": {
                                "deadline": {"$ne": None, "$lt": now},
                                "status": {"$nin": ["done"]},
                            }
                        },
                        {"$count": "count"},
                    ],
                    "by_assignee": [
                        {"$unwind": {"path": "$assigned_to", "preserveNullAndEmptyArrays": False, "includeArrayIndex": "_assignee_idx"}},
                        {
                            "$group": {
                                "_id": {
                                    "user_id": "$assigned_to",
                                    "status": "$status",
                                },
                                "count": {"$sum": 1},
                                "name": {"$first": {"$arrayElemAt": ["$assigned_to_names", "$_assignee_idx"]}},
                            }
                        },
                    ],
                    "recent_activity": [
                        {"$unwind": "$activity"},
                        {"$sort": {"activity.timestamp": -1}},
                        {"$limit": 20},
                        {"$project": {
                            "activity": 1,
                            "task_id": {"$toString": "$_id"},
                            "task_title": "$title",
                        }},
                    ],
                }
            }
        ]
        result = list(db[TASKS_COLLECTION].aggregate(pipeline))
        return result[0] if result else {}

    raw = await asyncio.to_thread(_fetch)

    by_status = {s["_id"]: s["count"] for s in raw.get("by_status", []) if s["_id"]}
    by_priority = {p["_id"]: p["count"] for p in raw.get("by_priority", []) if p["_id"]}
    overdue = raw.get("overdue", [{}])[0].get("count", 0) if raw.get("overdue") else 0
    total = sum(by_status.values())

    # Build per-assignee map — the pipeline doesn't easily denormalize names
    # so we do a second query to get user→name mapping from the tasks themselves
    def _assignee_names():
        results = db[TASKS_COLLECTION].find(
            {"assigned_to": {"$exists": True, "$ne": []}},
            {"assigned_to": 1, "assigned_to_names": 1},
        )
        uid_to_name: dict = {}
        for doc in results:
            for uid, name in zip(doc.get("assigned_to", []), doc.get("assigned_to_names", [])):
                uid_to_name[uid] = name
        return uid_to_name

    assignee_map: dict = {}
    for entry in raw.get("by_assignee", []):
        uid = (entry.get("_id") or {}).get("user_id")
        stat = (entry.get("_id") or {}).get("status", "unknown")
        if not uid:
            continue
        if uid not in assignee_map:
            assignee_map[uid] = {"user_id": uid, "name": uid, "total": 0, "by_status": {}}
        assignee_map[uid]["total"] += entry.get("count", 0)
        assignee_map[uid]["by_status"][stat] = entry.get("count", 0)

    uid_to_name = await asyncio.to_thread(_assignee_names)
    for uid, data in assignee_map.items():
        data["name"] = uid_to_name.get(uid, uid)

    recent_activity = serialize_mongo_document(raw.get("recent_activity", []))

    by_department = {d["_id"] or "None": d["count"] for d in raw.get("by_department", [])}

    return {
        "total": total,
        "by_status": by_status,
        "by_priority": by_priority,
        "by_department": by_department,
        "overdue": overdue,
        "by_assignee": list(assignee_map.values()),
        "recent_activity": recent_activity,
    }


# ── Create task ───────────────────────────────────────────────────────────────

@router.post("")
async def create_task(request: CreateTaskRequest, db=Depends(get_database)):
    if request.priority not in VALID_PRIORITIES:
        raise HTTPException(status_code=400, detail=f"Invalid priority: {request.priority}")
    if request.status not in VALID_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status: {request.status}")

    def _insert():
        now = utcnow()
        activity = _activity_entry(
            "created",
            request.created_by,
            request.created_by_name,
            detail=f"Task created with status '{request.status}' and priority '{request.priority}'",
        )
        doc = {
            "title": request.title.strip(),
            "description": request.description or "",
            "priority": request.priority,
            "status": request.status,
            "assigned_to": request.assigned_to or [],
            "assigned_to_names": request.assigned_to_names or [],
            "assigned_to_departments": request.assigned_to_departments or [],
            "deadline": request.deadline,
            "tags": request.tags or [],
            "created_by": request.created_by,
            "created_by_name": request.created_by_name,
            "creator_department": request.creator_department,
            "comments": [],
            "attachments": [],
            "activity": [activity],
            "created_at": now,
            "updated_at": now,
        }
        result = db[TASKS_COLLECTION].insert_one(doc)
        return db[TASKS_COLLECTION].find_one({"_id": result.inserted_id})

    task = await asyncio.to_thread(_insert)

    if request.assigned_to:
        new_assignees = [
            {"name": name, "department": dept}
            for name, dept in zip(
                request.assigned_to_names or [],
                request.assigned_to_departments or [],
            )
        ]
        asyncio.create_task(
            asyncio.to_thread(
                send_task_assignment_notification,
                request.title.strip(),
                request.created_by_name,
                request.created_by_name,
                new_assignees,
            )
        )

    return JSONResponse(status_code=201, content=serialize_mongo_document(task))


# ── Activity Report download ──────────────────────────────────────────────────

@router.get("/report/download")
async def download_task_report(
    start_date: str = Query(..., description="YYYY-MM-DD"),
    end_date: str = Query(..., description="YYYY-MM-DD"),
    department: Optional[str] = Query(None),
    user_id: Optional[str] = Query(None),
    db=Depends(get_database),
):
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    except ValueError:
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")

    def _fetch():
        query: dict = {
            "$or": [
                {"created_at": {"$gte": start_dt, "$lte": end_dt}},
                {"updated_at": {"$gte": start_dt, "$lte": end_dt}},
            ]
        }
        if department:
            dept_cond = [
                {"creator_department": department},
                {"assigned_to_departments": department},
            ]
            query = {"$and": [query, {"$or": dept_cond}]}
        if user_id:
            user_cond = [{"assigned_to": user_id}, {"created_by": user_id}]
            existing_and = query.get("$and")
            if existing_and:
                query["$and"].append({"$or": user_cond})
            else:
                query = {"$and": [query, {"$or": user_cond}]}
        return list(db[TASKS_COLLECTION].find(query).sort("created_at", -1))

    tasks = await asyncio.to_thread(_fetch)

    # ── Build Excel workbook ──────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────────────────────
    header_fill   = PatternFill("solid", fgColor="1E3A5F")
    summary_fill  = PatternFill("solid", fgColor="2563EB")
    alt_fill      = PatternFill("solid", fgColor="EFF6FF")
    done_fill     = PatternFill("solid", fgColor="D1FAE5")
    overdue_fill  = PatternFill("solid", fgColor="FEE2E2")
    header_font   = Font(bold=True, color="FFFFFF", size=10)
    bold_font     = Font(bold=True, size=10)
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align    = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _header_row(ws, cols: list[tuple[str, int]]):
        for ci, (label, width) in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = width
        ws.row_dimensions[1].height = 22

    def _fmt_dt(val) -> str:
        if not val:
            return ""
        if isinstance(val, str):
            try:
                val = datetime.fromisoformat(val)
            except ValueError:
                return val
        return val.strftime("%d %b %Y %H:%M")

    def _fmt_date(val) -> str:
        if not val:
            return ""
        if isinstance(val, str):
            try:
                val = datetime.fromisoformat(val[:10])
            except ValueError:
                return val
        return val.strftime("%d %b %Y")

    now_iso = utcnow().isoformat()

    # ── Sheet 1: Summary per assignee ─────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    sum_cols = [
        ("Name", 22), ("Department", 18), ("Total Assigned", 14),
        ("To Do", 10), ("In Progress", 12), ("Review", 10), ("Done", 10),
        ("Completion %", 13), ("Overdue", 10),
    ]
    _header_row(ws_sum, sum_cols)

    # Build per-assignee stats
    assignee_stats: dict = {}
    for task in tasks:
        for uid, name, dept in zip(
            task.get("assigned_to", []),
            task.get("assigned_to_names", []),
            task.get("assigned_to_departments", []),
        ):
            if uid not in assignee_stats:
                assignee_stats[uid] = {
                    "name": name, "department": dept or "",
                    "total": 0, "todo": 0, "in_progress": 0, "review": 0,
                    "done": 0, "overdue": 0,
                }
            s = assignee_stats[uid]
            s["total"] += 1
            st = task.get("status", "todo")
            if st in s:
                s[st] += 1
            deadline = task.get("deadline")
            if deadline and st != "done" and deadline < now_iso:
                s["overdue"] += 1

    rows = sorted(assignee_stats.values(), key=lambda x: (x["department"], x["name"]))
    for ri, r in enumerate(rows, 2):
        completion = round(r["done"] / r["total"] * 100, 1) if r["total"] else 0
        values = [
            r["name"], r["department"], r["total"],
            r["todo"], r["in_progress"], r["review"], r["done"],
            f"{completion}%", r["overdue"],
        ]
        fill = alt_fill if ri % 2 == 0 else None
        for ci, val in enumerate(values, 1):
            cell = ws_sum.cell(row=ri, column=ci, value=val)
            cell.alignment = center_align if ci > 2 else left_align
            if fill:
                cell.fill = fill
        ws_sum.row_dimensions[ri].height = 18

    # Totals row
    if rows:
        tr = len(rows) + 2
        total_tasks = sum(r["total"] for r in rows)
        total_done  = sum(r["done"] for r in rows)
        total_comp  = round(total_done / total_tasks * 100, 1) if total_tasks else 0
        totals = [
            "TOTAL", "",
            total_tasks,
            sum(r["todo"] for r in rows),
            sum(r["in_progress"] for r in rows),
            sum(r["review"] for r in rows),
            total_done,
            f"{total_comp}%",
            sum(r["overdue"] for r in rows),
        ]
        for ci, val in enumerate(totals, 1):
            cell = ws_sum.cell(row=tr, column=ci, value=val)
            cell.font = bold_font
            cell.fill = summary_fill
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = center_align if ci > 2 else left_align

    ws_sum.freeze_panes = "A2"

    # ── Sheet 2: Tasks Detail ─────────────────────────────────────────────────
    ws_det = wb.create_sheet("Tasks Detail")

    det_cols = [
        ("Title", 36), ("Status", 13), ("Priority", 11),
        ("Assigned To", 28), ("Department(s)", 22),
        ("Created By", 18), ("Created Date", 18), ("Updated Date", 18),
        ("Deadline", 14), ("Overdue", 10),
        ("Tags", 22), ("Description", 40),
    ]
    _header_row(ws_det, det_cols)

    STATUS_LABELS = {
        "todo": "To Do", "in_progress": "In Progress",
        "review": "Review", "done": "Done",
    }
    PRIORITY_LABELS = {
        "urgent": "Urgent", "high": "High", "medium": "Medium", "low": "Low",
    }

    for ri, task in enumerate(tasks, 2):
        st = task.get("status", "")
        deadline = task.get("deadline", "")
        overdue = bool(deadline and st != "done" and deadline < now_iso)

        depts = list(dict.fromkeys(
            [d for d in task.get("assigned_to_departments", []) if d]
            or ([task.get("creator_department")] if task.get("creator_department") else [])
        ))

        values = [
            task.get("title", ""),
            STATUS_LABELS.get(st, st),
            PRIORITY_LABELS.get(task.get("priority", ""), task.get("priority", "")),
            ", ".join(task.get("assigned_to_names", [])),
            ", ".join(depts),
            task.get("created_by_name", ""),
            _fmt_dt(task.get("created_at")),
            _fmt_dt(task.get("updated_at")),
            _fmt_date(deadline),
            "Yes" if overdue else "No",
            ", ".join(task.get("tags", [])),
            task.get("description", ""),
        ]

        row_fill = (
            done_fill if st == "done"
            else overdue_fill if overdue
            else (alt_fill if ri % 2 == 0 else None)
        )

        for ci, val in enumerate(values, 1):
            cell = ws_det.cell(row=ri, column=ci, value=val)
            cell.alignment = left_align
            if row_fill:
                cell.fill = row_fill
        ws_det.row_dimensions[ri].height = 18

    ws_det.freeze_panes = "A2"

    # ── Sheet 3: Department Summary ───────────────────────────────────────────
    ws_dept = wb.create_sheet("By Department")

    dept_cols = [
        ("Department", 22), ("Total Tasks", 13), ("To Do", 10),
        ("In Progress", 13), ("Review", 10), ("Done", 10),
        ("Completion %", 14), ("Overdue", 10),
    ]
    _header_row(ws_dept, dept_cols)

    dept_stats: dict = {}
    for task in tasks:
        depts = list(dict.fromkeys(
            [d for d in task.get("assigned_to_departments", []) if d]
            or ([task.get("creator_department")] if task.get("creator_department") else ["No Department"])
        ))
        st = task.get("status", "todo")
        deadline = task.get("deadline", "")
        overdue = bool(deadline and st != "done" and deadline < now_iso)
        for dept in depts:
            if dept not in dept_stats:
                dept_stats[dept] = {"total": 0, "todo": 0, "in_progress": 0, "review": 0, "done": 0, "overdue": 0}
            dept_stats[dept]["total"] += 1
            if st in dept_stats[dept]:
                dept_stats[dept][st] += 1
            if overdue:
                dept_stats[dept]["overdue"] += 1

    for ri, (dept, ds) in enumerate(sorted(dept_stats.items()), 2):
        comp = round(ds["done"] / ds["total"] * 100, 1) if ds["total"] else 0
        values = [
            dept, ds["total"], ds["todo"], ds["in_progress"],
            ds["review"], ds["done"], f"{comp}%", ds["overdue"],
        ]
        fill = alt_fill if ri % 2 == 0 else None
        for ci, val in enumerate(values, 1):
            cell = ws_dept.cell(row=ri, column=ci, value=val)
            cell.alignment = center_align if ci > 1 else left_align
            if fill:
                cell.fill = fill
        ws_dept.row_dimensions[ri].height = 18

    ws_dept.freeze_panes = "A2"

    # ── Stream response ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    period_label = f"{start_date}_to_{end_date}"
    filename = f"tasks_report_{period_label}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Get single task ───────────────────────────────────────────────────────────

@router.get("/{task_id}")
async def get_task(task_id: str, db=Depends(get_database)):
    def _fetch():
        return db[TASKS_COLLECTION].find_one({"_id": ObjectId(task_id), "is_deleted": {"$ne": True}})

    task = await asyncio.to_thread(_fetch)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return serialize_mongo_document(task)


def _spawn_on_complete_tasks_sync(db, templates: list, now: datetime) -> None:
    """Create follow-up tasks from on_complete_templates when a task is marked done."""
    USERS_COLLECTION = "purchase_users"
    for tmpl in templates:
        emails = tmpl.get("assignee_emails") or []
        if not emails:
            continue
        users_found = list(db[USERS_COLLECTION].find(
            {"email": {"$in": emails}, "status": "active"}
        ))
        if not users_found:
            logger.warning("on_complete spawn: none of %s found or active, skipping", emails)
            continue
        offset = int(tmpl.get("due_date_offset_days", 7))
        task_doc = {
            "title": tmpl.get("title", ""),
            "description": tmpl.get("description", ""),
            "priority": tmpl.get("priority", "medium"),
            "status": "todo",
            "assigned_to": [str(u["_id"]) for u in users_found],
            "assigned_to_names": [u.get("name", u["email"]) for u in users_found],
            "assigned_to_departments": [u.get("department", "") for u in users_found],
            "deadline": (now + timedelta(days=offset)).strftime("%Y-%m-%d"),
            "tags": list(tmpl.get("tags", [])),
            "created_by": "system",
            "created_by_name": "System",
            "creator_department": "",
            "comments": [],
            "attachments": [],
            "activity": [{
                "type": "created",
                "actor_id": "system",
                "actor_name": "System",
                "detail": "Task auto-created on completion of a linked task.",
                "timestamp": now.isoformat(),
            }],
            "is_hidden": False,
            "is_deleted": False,
            "created_at": now,
            "updated_at": now,
        }
        db[TASKS_COLLECTION].insert_one(task_doc)
        try:
            send_task_assignment_notification(
                task_title=task_doc["title"],
                created_by_name="System",
                assigned_by_name="System",
                new_assignees=[
                    {"name": n, "department": d}
                    for n, d in zip(task_doc["assigned_to_names"], task_doc["assigned_to_departments"])
                ],
            )
        except Exception as _e:
            logger.warning("on_complete Slack notification failed: %s", _e)


# ── Update task ───────────────────────────────────────────────────────────────

@router.put("/{task_id}")
async def update_task(task_id: str, request: UpdateTaskRequest, db=Depends(get_database)):
    actor_id = request.actor_id or "system"
    actor_name = request.actor_name or "System"

    notify_assignees = request.notify_assignees or False
    update_fields = {k: v for k, v in request.model_dump().items()
                     if v is not None and k not in ("actor_id", "actor_name", "notify_assignees")}

    if "priority" in update_fields and update_fields["priority"] not in VALID_PRIORITIES:
        raise HTTPException(status_code=400, detail=f"Invalid priority: {update_fields['priority']}")
    if "status" in update_fields and update_fields["status"] not in VALID_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status: {update_fields['status']}")
    if not update_fields:
        raise HTTPException(status_code=400, detail="No fields to update")

    old_assignees_snapshot: list = []

    def _update():
        nonlocal old_assignees_snapshot
        existing = db[TASKS_COLLECTION].find_one({"_id": ObjectId(task_id)})
        if not existing:
            return None

        old_assignees_snapshot = list(existing.get("assigned_to") or [])

        activities = []
        # Track meaningful field changes for the activity log
        tracked = {
            "status": ("status", lambda v: v.replace("_", " ").title()),
            "priority": ("priority", lambda v: v.title()),
            "title": ("title", lambda v: v),
            "deadline": ("deadline", lambda v: v or "removed"),
            "assigned_to_names": ("assignees", lambda v: ", ".join(v) if v else "none"),
        }
        for field, (label, fmt) in tracked.items():
            if field not in update_fields:
                continue
            old_raw = existing.get(field)
            new_raw = update_fields[field]
            if old_raw != new_raw:
                activities.append(_activity_entry(
                    f"{field}_changed",
                    actor_id,
                    actor_name,
                    field=label,
                    old_value=fmt(old_raw) if old_raw is not None else "—",
                    new_value=fmt(new_raw),
                ))

        now = utcnow()
        update_fields["updated_at"] = now

        ops: dict = {"$set": update_fields}
        if activities:
            ops["$push"] = {"activity": {"$each": activities}}

        db[TASKS_COLLECTION].update_one({"_id": ObjectId(task_id)}, ops)

        # Spawn on-completion follow-up tasks when transitioning to done
        if (update_fields.get("status") == "done"
                and existing.get("status") != "done"
                and existing.get("on_complete_templates")):
            _spawn_on_complete_tasks_sync(db, existing["on_complete_templates"], now)

        # Fan-out notifications for meaningful changes
        title = existing.get("title", "")
        created_by = existing.get("created_by", "")
        current_assignees = list(existing.get("assigned_to") or [])
        for act in activities:
            if act["type"] == "status_changed":
                recipients = [*current_assignees, created_by]
                _fan_out_sync(db, task_id, title, act["activity_id"], act["type"],
                              actor_id, actor_name, recipients,
                              f"changed status to {act['new_value']}")
            elif act["type"] == "assigned_to_names_changed":
                old_ids = set(existing.get("assigned_to") or [])
                added_ids = [uid for uid in (update_fields.get("assigned_to") or []) if uid not in old_ids]
                _fan_out_sync(db, task_id, title, act["activity_id"], act["type"],
                              actor_id, actor_name, added_ids,
                              "assigned you to")

        return db[TASKS_COLLECTION].find_one({"_id": ObjectId(task_id)})

    task = await asyncio.to_thread(_update)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    # Fire Slack notification for newly added assignees — only when the frontend
    # explicitly confirms (notify_assignees=True), i.e. the user clicked "Done"
    if notify_assignees and "assigned_to" in update_fields:
        old_ids = set(old_assignees_snapshot)
        new_ids = update_fields.get("assigned_to") or []
        new_names = update_fields.get("assigned_to_names") or []
        new_depts = update_fields.get("assigned_to_departments") or []

        added_assignees = [
            {"name": name, "department": dept}
            for uid, name, dept in zip(new_ids, new_names, new_depts)
            if uid not in old_ids
        ]
        if added_assignees:
            asyncio.create_task(
                asyncio.to_thread(
                    send_task_assignment_notification,
                    task.get("title", ""),
                    task.get("created_by_name", ""),
                    actor_name,
                    added_assignees,
                )
            )

    return serialize_mongo_document(task)


# ── Delete task ───────────────────────────────────────────────────────────────

@router.delete("/{task_id}")
async def delete_task(task_id: str, db=Depends(get_database)):
    def _delete():
        task = db[TASKS_COLLECTION].find_one({"_id": ObjectId(task_id), "is_deleted": {"$ne": True}}, {"status": 1})
        if not task:
            return False
        if task.get("status") == "done":
            raise HTTPException(status_code=403, detail="Completed tasks cannot be deleted")
        db[TASKS_COLLECTION].update_one(
            {"_id": ObjectId(task_id)},
            {"$set": {"is_deleted": True, "deleted_at": utcnow().isoformat()}},
        )
        return True

    deleted = await asyncio.to_thread(_delete)
    if not deleted:
        raise HTTPException(status_code=404, detail="Task not found")
    return {"task_id": task_id, "deleted": True}


# ── Comments ──────────────────────────────────────────────────────────────────

@router.post("/{task_id}/comments")
async def add_comment(task_id: str, request: AddCommentRequest, db=Depends(get_database)):
    comment = {
        "comment_id": str(ObjectId()),
        "text": request.text.strip(),
        "author_id": request.author_id,
        "author_name": request.author_name,
        "created_at": utcnow().isoformat(),
    }
    activity = _activity_entry(
        "comment_added",
        request.author_id,
        request.author_name,
        detail=request.text[:80] + ("…" if len(request.text) > 80 else ""),
    )

    def _add():
        task = db[TASKS_COLLECTION].find_one(
            {"_id": ObjectId(task_id)},
            {"assigned_to": 1, "created_by": 1, "title": 1},
        )
        if not task:
            return False
        db[TASKS_COLLECTION].update_one(
            {"_id": ObjectId(task_id)},
            {
                "$push": {"comments": comment, "activity": activity},
                "$set": {"updated_at": utcnow()},
            },
        )
        snippet = f'commented: {request.text[:60]}{"…" if len(request.text) > 60 else ""}'
        recipients = [*(task.get("assigned_to") or []), task.get("created_by", "")]
        _fan_out_sync(db, task_id, task.get("title", ""), activity["activity_id"],
                      activity["type"], request.author_id, request.author_name,
                      recipients, snippet)
        return True

    matched = await asyncio.to_thread(_add)
    if not matched:
        raise HTTPException(status_code=404, detail="Task not found")
    return serialize_mongo_document(comment)


@router.delete("/{task_id}/comments/{comment_id}")
async def delete_comment(task_id: str, comment_id: str, db=Depends(get_database)):
    actor_id: str = ""
    actor_name: str = "Unknown"

    def _delete():
        task = db[TASKS_COLLECTION].find_one(
            {"_id": ObjectId(task_id)},
            {"comments": 1},
        )
        if not task:
            return False
        comment = next((c for c in task.get("comments", []) if c.get("comment_id") == comment_id), None)
        nonlocal actor_id, actor_name
        if comment:
            actor_id = comment.get("author_id", "")
            actor_name = comment.get("author_name", "Unknown")

        activity = _activity_entry("comment_deleted", actor_id, actor_name)
        result = db[TASKS_COLLECTION].update_one(
            {"_id": ObjectId(task_id)},
            {
                "$pull": {"comments": {"comment_id": comment_id}},
                "$push": {"activity": activity},
                "$set": {"updated_at": utcnow()},
            },
        )
        return result.matched_count > 0

    deleted = await asyncio.to_thread(_delete)
    if not deleted:
        raise HTTPException(status_code=404, detail="Task or comment not found")
    return {"comment_id": comment_id, "deleted": True}


# ── Attachments ───────────────────────────────────────────────────────────────

@router.post("/{task_id}/attachments")
async def upload_attachment(
    task_id: str,
    file: UploadFile = File(...),
    uploaded_by: str = Form(...),
    uploaded_by_name: str = Form(...),
    db=Depends(get_database),
):
    """
    Upload a file to S3 under tasks/{task_id}/ and store only the s3_key.
    Pre-signed URLs (7 days) are generated on demand via the /url endpoint.
    """
    file_bytes = await file.read()
    filename = file.filename or "file"
    content_type = file.content_type or "application/octet-stream"

    def _process():
        task = db[TASKS_COLLECTION].find_one({"_id": ObjectId(task_id)}, {"_id": 1})
        if not task:
            raise ValueError("Task not found")

        ts = utcnow().strftime("%Y%m%d_%H%M%S")
        safe_filename = re.sub(r"[^\w.\-]", "_", filename)
        s3_key = f"tasks/{task_id}/{ts}_{safe_filename}"

        _upload_to_s3(file_bytes, s3_key, content_type)

        attachment = {
            "file_id": str(ObjectId()),
            "filename": filename,
            "s3_key": s3_key,           # Only the key is stored; URL is presigned on demand
            "content_type": content_type,
            "size": len(file_bytes),
            "uploaded_at": utcnow().isoformat(),
            "uploaded_by": uploaded_by,
            "uploaded_by_name": uploaded_by_name,
        }
        activity = _activity_entry(
            "attachment_added",
            uploaded_by,
            uploaded_by_name,
            detail=filename,
        )
        db[TASKS_COLLECTION].update_one(
            {"_id": ObjectId(task_id)},
            {
                "$push": {"attachments": attachment, "activity": activity},
                "$set": {"updated_at": utcnow()},
            },
        )
        return attachment

    try:
        attachment = await asyncio.to_thread(_process)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.exception("upload_attachment failed")
        raise HTTPException(status_code=500, detail=str(e))

    return serialize_mongo_document(attachment)


@router.get("/{task_id}/attachments/{file_id}/url")
async def get_attachment_url(
    task_id: str,
    file_id: str,
    download: bool = Query(False, description="If true, URL forces a file download instead of inline view"),
    db=Depends(get_database),
):
    """Return a 7-day pre-signed S3 GET URL for the given attachment.
    Pass ?download=true to get a URL that triggers a browser download (Content-Disposition: attachment)."""
    def _fetch():
        task = db[TASKS_COLLECTION].find_one(
            {"_id": ObjectId(task_id)},
            {"attachments": 1},
        )
        if not task:
            return None, None
        for att in task.get("attachments", []):
            if att.get("file_id") == file_id:
                return att.get("s3_key"), att.get("filename", "")
        return None, None

    s3_key, filename = await asyncio.to_thread(_fetch)
    if not s3_key:
        raise HTTPException(status_code=404, detail="Attachment not found")

    try:
        url = await asyncio.to_thread(_presign_s3, s3_key, ATTACHMENT_URL_EXPIRY, download, filename or "")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    return {"url": url, "expires_in_seconds": ATTACHMENT_URL_EXPIRY}


@router.get("/{task_id}/attachments/download-all")
async def download_all_attachments(task_id: str, db=Depends(get_database)):
    """Download all attachments for a task as a single ZIP file streamed to the client."""
    def _fetch_task():
        return db[TASKS_COLLECTION].find_one(
            {"_id": ObjectId(task_id)},
            {"attachments": 1, "title": 1},
        )

    task = await asyncio.to_thread(_fetch_task)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    attachments = task.get("attachments", [])
    if not attachments:
        raise HTTPException(status_code=404, detail="No attachments found for this task")

    s3 = _s3_client()

    def _build_zip() -> bytes:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            seen_names: dict[str, int] = {}
            for att in attachments:
                s3_key = att.get("s3_key")
                original_name = att.get("filename") or s3_key.split("/")[-1] if s3_key else "file"
                if not s3_key:
                    continue
                # De-duplicate filenames inside the ZIP
                if original_name in seen_names:
                    seen_names[original_name] += 1
                    base, _, ext = original_name.rpartition(".")
                    arc_name = f"{base}_{seen_names[original_name]}.{ext}" if ext else f"{original_name}_{seen_names[original_name]}"
                else:
                    seen_names[original_name] = 0
                    arc_name = original_name
                try:
                    obj = s3.get_object(Bucket=S3_BUCKET, Key=s3_key)
                    file_bytes = obj["Body"].read()
                    zf.writestr(arc_name, file_bytes)
                except Exception:
                    logger.warning("Skipping attachment %s in download-all zip (S3 error)", s3_key)
        buf.seek(0)
        return buf.read()

    try:
        zip_bytes = await asyncio.to_thread(_build_zip)
    except Exception as e:
        logger.exception("download_all_attachments failed")
        raise HTTPException(status_code=500, detail=str(e))

    task_title = re.sub(r"[^\w\s\-]", "", task.get("title", task_id))[:40].strip().replace(" ", "_")
    zip_filename = f"{task_title}_attachments.zip" if task_title else f"task_{task_id}_attachments.zip"

    return StreamingResponse(
        io.BytesIO(zip_bytes),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{zip_filename}"'},
    )


@router.delete("/{task_id}/attachments/{file_id}")
async def delete_attachment(task_id: str, file_id: str, db=Depends(get_database)):
    def _delete():
        task = db[TASKS_COLLECTION].find_one({"_id": ObjectId(task_id)}, {"attachments": 1})
        if not task:
            return False, None, None
        att = next((a for a in task.get("attachments", []) if a.get("file_id") == file_id), None)
        if not att:
            return False, None, None
        s3_key = att.get("s3_key")
        uploader = att.get("uploaded_by", "")
        uploader_name = att.get("uploaded_by_name", "Unknown")
        if s3_key:
            try:
                _delete_from_s3(s3_key)
            except Exception:
                logger.warning("Failed to delete S3 key %s", s3_key)
        activity = _activity_entry("attachment_deleted", uploader, uploader_name, detail=att.get("filename", ""))
        result = db[TASKS_COLLECTION].update_one(
            {"_id": ObjectId(task_id)},
            {
                "$pull": {"attachments": {"file_id": file_id}},
                "$push": {"activity": activity},
                "$set": {"updated_at": utcnow()},
            },
        )
        return result.matched_count > 0, uploader, uploader_name

    deleted, _, _ = await asyncio.to_thread(_delete)
    if not deleted:
        raise HTTPException(status_code=404, detail="Task or attachment not found")
    return {"file_id": file_id, "deleted": True}
