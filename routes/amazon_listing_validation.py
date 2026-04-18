import csv as _csv
import io
import logging
import asyncio
from typing import Optional

import requests
from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import PatternFill, Font

from ..database import get_database

logger = logging.getLogger(__name__)
router = APIRouter()

# GST tax code mapping (Amazon code → percentage)
TAX_CODE_MAP = {
    "A_GEN_STANDARD": 18.0,
    "A_GEN_SUPERREDUCED": 5.0,
    "A_GEN_REDUCED": 12.0,
}

COMBO_SHEET_ID  = "1tn_Lj3KR0zXY8B-8ZUkSznZgE4YzyjtAkcpdHzBCgt4"
COMBO_SHEET_GID = "1878034212"
COMBO_SHEET_CSV_URL = (
    f"https://docs.google.com/spreadsheets/d/{COMBO_SHEET_ID}"
    f"/export?format=csv&gid={COMBO_SHEET_GID}"
)

# Column header names in the combo Google Sheet (row 2 is the real header)
COMBO_COL_SKU = "SKU Code"
COMBO_COL_MRP = "MRP"
COMBO_COL_SP  = "SP"
COMBO_COL_HSN = "HSN Code"
COMBO_COL_GST  = "GST %"
COMBO_COL_NAME = "Item Amazon Name"


# Seller Central Template column letters
SC_COLS = {
    "sku": "A",
    "hsn": "DD",
    "tax_code": "EJ",
    "mrp": "FL",
    "sp": "FK",
}

# Vendor Central Template column letters
VC_COLS = {
    "sku": "B",
    "hsn": "DI",
    "mrp": "EQ",
}


def _clean(val) -> str:
    """Normalise a cell value to a stripped string."""
    if val is None:
        return ""
    return str(val).strip().strip("'")


def _find_template_sheet(wb):
    """
    Return the Template sheet from the workbook.
    Prefers an exact case-insensitive match of 'Template' first,
    then falls back to any sheet whose name starts with 'Template'.
    """
    # Exact match first (e.g. "Template" or "Template-PET_TOY")
    for name in wb.sheetnames:
        if name.strip().lower() == "template":
            return wb[name]
    # Starts-with match (e.g. "Template-PET_TOY")
    for name in wb.sheetnames:
        if name.strip().lower().startswith("template"):
            return wb[name]
    raise ValueError(f"No Template sheet found. Available sheets: {wb.sheetnames}")


def _col_idx(letter: str) -> int:
    return column_index_from_string(letter)


def _parse_seller_central(file_bytes: bytes) -> list[dict]:
    """
    Parse the Seller Central xlsm file.
    Header is on row 4, data starts on row 6.
    Returns list of dicts with keys: sku, hsn, tax_code, mrp, sp.
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, keep_vba=True)
    ws = _find_template_sheet(wb)

    col_map = {k: _col_idx(v) for k, v in SC_COLS.items()}
    rows = []

    for row in ws.iter_rows(min_row=6, values_only=True):
        sku = _clean(row[col_map["sku"] - 1])
        if not sku:
            continue
        rows.append(
            {
                "sku": sku,
                "hsn": _clean(row[col_map["hsn"] - 1]),
                "tax_code": _clean(row[col_map["tax_code"] - 1]),
                "mrp": _clean(row[col_map["mrp"] - 1]),
                "sp": _clean(row[col_map["sp"] - 1]),
            }
        )
    return rows


def _parse_vendor_central(file_bytes: bytes) -> list[dict]:
    """
    Parse the Vendor Central xlsm file.
    Header is on row 3, data starts on row 6.
    Returns list of dicts with keys: sku, hsn, mrp.
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, keep_vba=True)
    ws = _find_template_sheet(wb)

    col_map = {k: _col_idx(v) for k, v in VC_COLS.items()}
    rows = []

    for row in ws.iter_rows(min_row=6, values_only=True):
        sku = _clean(row[col_map["sku"] - 1])
        if not sku:
            continue
        rows.append(
            {
                "sku": sku,
                "hsn": _clean(row[col_map["hsn"] - 1]),
                "mrp": _clean(row[col_map["mrp"] - 1]),
            }
        )
    return rows


def _load_products_by_sku(skus: set[str]) -> dict[str, dict]:
    """Batch-load products from MongoDB keyed by cf_sku_code."""
    db = get_database()
    products = list(
        db["products"].find(
            {"cf_sku_code": {"$in": list(skus)}},
            {
                "_id": 0,
                "cf_sku_code": 1,
                "hsn_or_sac": 1,
                "rate": 1,
                "item_tax_preferences": 1,
                "item_name": 1,
            },
        )
    )
    return {str(p["cf_sku_code"]).strip(): p for p in products}


def _fetch_combo_sheet_data() -> dict[str, dict]:
    """
    Fetch combo product reference data from the Google Sheet.
    Row 1 is empty, row 2 is the header — so we skip line 0 and use line 1 as header.
    Returns {sku_code: {mrp, sp, hsn}} with string values.
    """
    resp = requests.get(COMBO_SHEET_CSV_URL, timeout=30)
    resp.raise_for_status()
    lines = resp.text.splitlines()
    if len(lines) < 2:
        raise ValueError("Combo Google Sheet has fewer than 2 rows")
    reader = _csv.DictReader(io.StringIO("\n".join(lines[1:])))
    result = {}
    for row in reader:
        sku = str(row.get(COMBO_COL_SKU) or "").strip()
        if not sku:
            continue
        result[sku] = {
            "mrp":       str(row.get(COMBO_COL_MRP)  or "").strip(),
            "sp":        str(row.get(COMBO_COL_SP)   or "").strip(),
            "hsn":       str(row.get(COMBO_COL_HSN)  or "").strip(),
            "gst":       str(row.get(COMBO_COL_GST)  or "").strip(),
            "item_name": str(row.get(COMBO_COL_NAME) or "").strip(),
        }
    return result


def _load_composite_products(skus: set[str]) -> dict[str, dict]:
    """
    Return composite product docs keyed by sku_code, for the given SKU set.
    Fetches item_tax_preferences so GST can be validated the same as single products.
    """
    db = get_database()
    docs = db["composite_products"].find(
        {"sku_code": {"$in": list(skus)}},
        {"sku_code": 1, "item_tax_preferences": 1, "_id": 0},
    )
    return {str(d["sku_code"]).strip(): d for d in docs}


def _get_db_gst(product: dict) -> Optional[float]:
    """Extract the intra GST percentage from item_tax_preferences."""
    prefs = product.get("item_tax_preferences") or []
    for pref in prefs:
        spec = pref.get("tax_specification", "")
        if spec == "intra":
            return pref.get("tax_percentage")
    # Fallback: return first available percentage
    if prefs:
        return prefs[0].get("tax_percentage")
    return None


def _compare_hsn(file_val: str, db_val: str) -> bool:
    """Compare HSN values loosely – strip spaces, commas, leading zeros."""
    def norm(v):
        return v.replace(" ", "").replace(",", "").lstrip("0").lower()
    return norm(file_val) == norm(db_val)


def _compare_numeric(file_val: str, db_val) -> bool:
    """Compare numeric values after rounding to 2 decimal places."""
    try:
        return round(float(file_val), 2) == round(float(db_val), 2)
    except (TypeError, ValueError):
        return str(file_val).strip() == str(db_val).strip()


def _validate_seller_central(
    rows: list[dict],
    products: dict[str, dict],
    combo_skus: set[str],
    combo_data: dict[str, dict],
) -> list[dict]:
    """Return ALL rows with per-field match status. Skips Amazon example placeholder rows."""
    results = []
    for row in rows:
        sku = row["sku"]
        # Skip Amazon template example placeholder row
        if sku.upper() == "ABC123":
            continue

        is_combo = sku in combo_skus

        if is_combo:
            combo = combo_data.get(sku)
            if not combo:
                results.append({
                    "source": "Seller Central",
                    "sku": sku,
                    "item_name": "—",
                    "found": False,
                    "is_combo": True,
                    "has_mismatch": True,
                    "hsn": None,
                    "gst": None,
                    "mrp": None,
                    "sp": {"file": row.get("sp", "")},
                })
                continue

            file_hsn = row["hsn"]
            sheet_hsn = combo["hsn"]
            hsn_match = not file_hsn or not sheet_hsn or _compare_hsn(file_hsn, sheet_hsn)

            file_mrp = row["mrp"]
            sheet_mrp = combo["mrp"]
            mrp_match = not file_mrp or not sheet_mrp or _compare_numeric(file_mrp, sheet_mrp)

            file_sp = row.get("sp", "")
            sheet_sp = combo["sp"]
            sp_match = not file_sp or not sheet_sp or _compare_numeric(file_sp, sheet_sp)

            file_tax_code = row["tax_code"]
            file_tax_pct = TAX_CODE_MAP.get(file_tax_code)
            sheet_gst_str = combo["gst"]
            if file_tax_code and file_tax_code not in TAX_CODE_MAP:
                gst_match = False
                gst_issue = "Unrecognised tax code"
                gst_file_display = file_tax_code
            elif file_tax_pct is not None and sheet_gst_str:
                try:
                    gst_match = file_tax_pct == float(sheet_gst_str.rstrip("%"))
                except ValueError:
                    gst_match = True
                gst_issue = None
                gst_file_display = f"{file_tax_code} ({file_tax_pct}%)"
            else:
                gst_match = True
                gst_issue = None
                gst_file_display = f"{file_tax_code} ({file_tax_pct}%)" if file_tax_pct else file_tax_code

            results.append({
                "source": "Seller Central",
                "sku": sku,
                "item_name": combo.get("item_name") or "—",
                "found": True,
                "is_combo": True,
                "has_mismatch": not hsn_match or not gst_match or not mrp_match or not sp_match,
                "hsn": {"file": file_hsn, "db": sheet_hsn or "—", "match": hsn_match},
                "gst": {
                    "file": gst_file_display,
                    "db": (f"{sheet_gst_str}%" if not sheet_gst_str.endswith("%") else sheet_gst_str) if sheet_gst_str else "—",
                    "match": gst_match,
                    "issue": gst_issue,
                },
                "mrp": {
                    "file": file_mrp,
                    "db": sheet_mrp or "—",
                    "match": mrp_match,
                },
                "sp": {"file": file_sp, "db": sheet_sp or "—", "mrp_file": file_mrp, "match": sp_match},
            })
            continue

        product = products.get(sku)

        if not product:
            results.append({
                "source": "Seller Central",
                "sku": sku,
                "item_name": "—",
                "found": False,
                "is_combo": False,
                "has_mismatch": True,
                "hsn": None,
                "gst": None,
                "mrp": None,
                "sp": {"file": row.get("sp", "")},
            })
            continue

        db_hsn = str(product.get("hsn_or_sac") or "").strip()
        db_mrp = product.get("rate")
        db_gst = _get_db_gst(product)
        item_name = product.get("item_name", "")

        file_hsn = row["hsn"]
        hsn_match = not file_hsn or _compare_hsn(file_hsn, db_hsn)

        file_tax_code = row["tax_code"]
        file_tax_pct = TAX_CODE_MAP.get(file_tax_code)
        if file_tax_code and file_tax_code not in TAX_CODE_MAP:
            gst_match = False
            gst_issue = "Unrecognised tax code"
            gst_file_display = file_tax_code
        elif file_tax_pct is not None and db_gst is not None:
            gst_match = file_tax_pct == db_gst
            gst_issue = None
            gst_file_display = f"{file_tax_code} ({file_tax_pct}%)"
        else:
            gst_match = True
            gst_issue = None
            gst_file_display = f"{file_tax_code} ({file_tax_pct}%)" if file_tax_pct else file_tax_code

        file_mrp = row["mrp"]
        mrp_match = not file_mrp or db_mrp is None or _compare_numeric(file_mrp, db_mrp)

        file_sp = row.get("sp", "")
        sp_match = True
        if file_sp and file_mrp:
            try:
                sp_match = round(float(file_sp), 2) <= round(float(file_mrp), 2)
            except (ValueError, TypeError):
                sp_match = True

        results.append({
            "source": "Seller Central",
            "sku": sku,
            "item_name": item_name,
            "found": True,
            "is_combo": False,
            "has_mismatch": not hsn_match or not gst_match or not mrp_match or not sp_match,
            "hsn": {"file": file_hsn, "db": db_hsn or "—", "match": hsn_match},
            "gst": {
                "file": gst_file_display,
                "db": f"{db_gst}%" if db_gst is not None else "—",
                "match": gst_match,
                "issue": gst_issue,
            },
            "mrp": {
                "file": file_mrp,
                "db": str(db_mrp) if db_mrp is not None else "—",
                "match": mrp_match,
            },
            "sp": {"file": file_sp, "mrp_file": file_mrp, "match": sp_match},
        })
    return results


def _validate_vendor_central(
    rows: list[dict],
    products: dict[str, dict],
    combo_skus: set[str],
    combo_data: dict[str, dict],
) -> list[dict]:
    """Return ALL rows with per-field match status. Skips Amazon example placeholder rows."""
    results = []
    for row in rows:
        sku = row["sku"]
        if sku.upper() == "ABC123":
            continue

        is_combo = sku in combo_skus

        if is_combo:
            combo = combo_data.get(sku)
            if not combo:
                results.append({
                    "source": "Vendor Central",
                    "sku": sku,
                    "item_name": "—",
                    "found": False,
                    "is_combo": True,
                    "has_mismatch": True,
                    "hsn": None,
                    "mrp": None,
                })
                continue

            file_hsn = row["hsn"]
            sheet_hsn = combo["hsn"]
            hsn_match = not file_hsn or not sheet_hsn or _compare_hsn(file_hsn, sheet_hsn)

            file_mrp = row["mrp"]
            sheet_mrp = combo["mrp"]
            mrp_match = not file_mrp or not sheet_mrp or _compare_numeric(file_mrp, sheet_mrp)

            results.append({
                "source": "Vendor Central",
                "sku": sku,
                "item_name": "—",
                "found": True,
                "is_combo": True,
                "has_mismatch": not hsn_match or not mrp_match,
                "hsn": {"file": file_hsn, "db": sheet_hsn or "—", "match": hsn_match},
                "mrp": {
                    "file": file_mrp,
                    "db": sheet_mrp or "—",
                    "match": mrp_match,
                },
            })
            continue

        product = products.get(sku)

        if not product:
            results.append({
                "source": "Vendor Central",
                "sku": sku,
                "item_name": "—",
                "found": False,
                "is_combo": False,
                "has_mismatch": True,
                "hsn": None,
                "mrp": None,
            })
            continue

        db_hsn = str(product.get("hsn_or_sac") or "").strip()
        db_mrp = product.get("rate")
        item_name = product.get("item_name", "")

        file_hsn = row["hsn"]
        file_mrp = row["mrp"]
        hsn_match = not file_hsn or _compare_hsn(file_hsn, db_hsn)
        mrp_match = not file_mrp or db_mrp is None or _compare_numeric(file_mrp, db_mrp)

        results.append({
            "source": "Vendor Central",
            "sku": sku,
            "item_name": item_name,
            "found": True,
            "is_combo": False,
            "has_mismatch": not hsn_match or not mrp_match,
            "hsn": {"file": file_hsn, "db": db_hsn or "—", "match": hsn_match},
            "mrp": {
                "file": file_mrp,
                "db": str(db_mrp) if db_mrp is not None else "—",
                "match": mrp_match,
            },
        })
    return results


def _build_excel(sc_results: list[dict], vc_results: list[dict]) -> bytes:
    wb = Workbook()

    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    header_font = Font(bold=True)

    def _status_fill(match):
        return green_fill if match else red_fill

    SC_HEADERS = [
        "SKU", "Item Name", "Type", "Status",
        "HSN (File)", "HSN (Ref)", "HSN",
        "GST (File)", "GST (Ref)", "GST",
        "MRP (File)", "MRP (Ref)", "MRP",
        "SP (File)", "SP (Ref)", "SP",
    ]
    VC_HEADERS = [
        "SKU", "Item Name", "Type", "Status",
        "HSN (File)", "HSN (Ref)", "HSN",
        "MRP (File)", "MRP (Ref)", "MRP",
    ]

    def _write_sheet(ws, headers, data, is_sc: bool):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font

        for row in data:
            row_type = "Combo" if row.get("is_combo") else "Single"
            if not row.get("found"):
                if is_sc:
                    ws.append([row["sku"], row["item_name"], row_type, "Not Found",
                                "—", "—", "Not Found", "—", "—", "Not Found",
                                "—", "—", "Not Found", "—", "—", "Not Found"])
                else:
                    ws.append([row["sku"], row["item_name"], row_type, "Not Found",
                                "—", "—", "Not Found", "—", "—", "Not Found"])
                for cell in ws[ws.max_row]:
                    cell.fill = red_fill
            else:
                hsn = row.get("hsn") or {}
                mrp = row.get("mrp") or {}
                status = "Mismatch" if row["has_mismatch"] else "Match"

                if is_sc:
                    gst = row.get("gst") or {}
                    sp = row.get("sp") or {}
                    sp_ref = sp.get("db", "")
                    if sp_ref:
                        sp_status = "Match" if sp.get("match", True) else f"Mismatch (ref: {sp_ref})"
                    else:
                        sp_status = "OK" if sp.get("match", True) else f"SP > MRP ({sp.get('mrp_file', '')})"
                    ws.append([
                        row["sku"], row["item_name"], row_type, status,
                        hsn.get("file", ""), hsn.get("db", ""), "Match" if hsn.get("match") else "Mismatch",
                        gst.get("file", ""), gst.get("db", ""), "Match" if gst.get("match") else "Mismatch",
                        mrp.get("file", ""), mrp.get("db", ""), "Match" if mrp.get("match") else "Mismatch",
                        sp.get("file", ""), sp_ref, sp_status,
                    ])
                    excel_row = ws[ws.max_row]
                    excel_row[3].fill = red_fill if row["has_mismatch"] else green_fill
                    excel_row[6].fill = _status_fill(hsn.get("match"))
                    excel_row[9].fill = _status_fill(gst.get("match"))
                    excel_row[12].fill = _status_fill(mrp.get("match"))
                    excel_row[15].fill = _status_fill(sp.get("match", True))
                else:
                    ws.append([
                        row["sku"], row["item_name"], row_type, status,
                        hsn.get("file", ""), hsn.get("db", ""), "Match" if hsn.get("match") else "Mismatch",
                        mrp.get("file", ""), mrp.get("db", ""), "Match" if mrp.get("match") else "Mismatch",
                    ])
                    excel_row = ws[ws.max_row]
                    excel_row[3].fill = red_fill if row["has_mismatch"] else green_fill
                    excel_row[6].fill = _status_fill(hsn.get("match"))
                    excel_row[9].fill = _status_fill(mrp.get("match"))

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    if sc_results:
        ws_sc = wb.create_sheet(title="Seller Central")
        _write_sheet(ws_sc, SC_HEADERS, sc_results, is_sc=True)

    if vc_results:
        ws_vc = wb.create_sheet(title="Vendor Central")
        _write_sheet(ws_vc, VC_HEADERS, vc_results, is_sc=False)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


async def _run_validation(
    seller_central_file: Optional[UploadFile],
    vendor_central_file: Optional[UploadFile],
) -> tuple[list, list]:
    """
    Core validation logic. Either file may be None.
    Returns (sc_results, vc_results).
    """
    if not seller_central_file and not vendor_central_file:
        raise HTTPException(
            status_code=400,
            detail="At least one file (Seller Central or Vendor Central) must be provided.",
        )

    sc_rows: list[dict] = []
    vc_rows: list[dict] = []

    if seller_central_file:
        sc_bytes = await seller_central_file.read()
        try:
            sc_rows = await asyncio.to_thread(_parse_seller_central, sc_bytes)
        except Exception as e:
            raise HTTPException(
                status_code=400, detail=f"Failed to parse Seller Central file: {e}"
            )

    if vendor_central_file:
        vc_bytes = await vendor_central_file.read()
        try:
            vc_rows = await asyncio.to_thread(_parse_vendor_central, vc_bytes)
        except Exception as e:
            raise HTTPException(
                status_code=400, detail=f"Failed to parse Vendor Central file: {e}"
            )

    all_skus = {r["sku"] for r in sc_rows} | {r["sku"] for r in vc_rows}
    if not all_skus:
        raise HTTPException(
            status_code=400,
            detail="No SKU data found in the uploaded file(s). Ensure data rows start at row 6.",
        )

    products, composite_products, combo_data = await asyncio.gather(
        asyncio.to_thread(_load_products_by_sku, all_skus),
        asyncio.to_thread(_load_composite_products, all_skus),
        asyncio.to_thread(_fetch_combo_sheet_data),
    )
    combo_skus = set(composite_products.keys())

    sc_results = _validate_seller_central(sc_rows, products, combo_skus, combo_data) if sc_rows else []
    vc_results = _validate_vendor_central(vc_rows, products, combo_skus, combo_data) if vc_rows else []

    return sc_results, vc_results


@router.post("/validate")
async def validate_amazon_listings(
    seller_central_file: Optional[UploadFile] = File(None),
    vendor_central_file: Optional[UploadFile] = File(None),
):
    """
    Validate Amazon listing files against the products collection in MongoDB.
    Either file is optional — at least one must be supplied.
    Returns all rows with per-field match status.
    """
    sc_results, vc_results = await _run_validation(
        seller_central_file, vendor_central_file
    )
    return {
        "summary": {
            "seller_central_rows": len(sc_results),
            "vendor_central_rows": len(vc_results),
            "seller_central_mismatches": sum(1 for r in sc_results if r["has_mismatch"]),
            "vendor_central_mismatches": sum(1 for r in vc_results if r["has_mismatch"]),
        },
        "seller_central": sc_results,
        "vendor_central": vc_results,
    }


@router.post("/validate/download")
async def download_amazon_listing_validation(
    seller_central_file: Optional[UploadFile] = File(None),
    vendor_central_file: Optional[UploadFile] = File(None),
):
    """
    Same as /validate but returns a highlighted Excel file with all rows.
    Either file is optional — at least one must be supplied.
    """
    sc_results, vc_results = await _run_validation(
        seller_central_file, vendor_central_file
    )
    excel_bytes = await asyncio.to_thread(_build_excel, sc_results, vc_results)
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=amazon_listing_validation.xlsx"
        },
    )
