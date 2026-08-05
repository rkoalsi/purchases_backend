from fastapi import APIRouter, HTTPException, Depends, Query, Body, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from datetime import datetime, timedelta
from typing import Optional
import io
import logging
import os
import re
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..database import get_database, serialize_mongo_document

logger = logging.getLogger(__name__)
router = APIRouter()

PRODUCTS_COLLECTION = "products"
PURCHASE_ORDERS_COLLECTION = "purchase_orders"
BILLS_COLLECTION = "bills"
BRAND_ORDERS_COLLECTION = "brand_orders"
INVENTORY_ADJUSTMENTS_COLLECTION = "inventory_adjustments"

# Bill creation defaults. "Advance" and "Due on Receipt" are both 0-day terms in Zoho — only the
# label separates them, so payment_terms_label must always be sent alongside payment_terms.
DEFAULT_PAYMENT_TERMS_LABEL = "Advance"
# Zoho custom field "Original Bill Date" (the vendor's own invoice date, distinct from the bill's
# posting date). Matched by api_name; customfield_id 3220178000122463035.
ORIGINAL_BILL_DATE_FIELD = "cf_original_bill_date"
BASE_CURRENCY = "INR"

ZOHO_BOOKS_BASE = "https://books.zoho.com/api/v3"
ZOHO_INVENTORY_BASE = os.getenv("ZOHO_INVENTORY_BASE", "https://www.zohoapis.com/inventory/v1")
ORGANIZATION_ID = os.getenv("ORGANIZATION_ID", "776755316")
BOOKS_URL = os.getenv("BOOKS_URL")
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
BOOKS_REFRESH_TOKEN = os.getenv("BOOKS_REFRESH_TOKEN")
INVENTORY_REFRESH_TOKEN = os.getenv("INVENTORY_REFRESH_TOKEN", BOOKS_REFRESH_TOKEN)
_ZOHO_TOKEN_URL = "https://accounts.zoho.com/oauth/v2/token"

DEFAULT_ACCOUNT_ID = "3220178000000034001"  # Inventory Asset
DEFAULT_WAREHOUSE_ID = "3220178000000403010"  # Pupscribe Enterprises Private Limited

# Line-item-level custom field definitions (distinct from item-master custom fields).
# Confirmed from a live bill (SH26E0558) where a fully-populated line item carried these.
BILL_LINE_ITEM_CF_MFR_CODE = "3220178000000075182"  # api_name cf_item_code, label "Manufacturer Code"
BILL_LINE_ITEM_CF_SKU_CODE = "3220178000000075204"  # api_name cf_sku_code, label "SKU Code"


# Zoho access tokens live ~1 hour. Zoho rate-limits the OAuth token endpoint hard
# ("You have made too many requests continuously"), so we MUST NOT refresh on every
# request — cache each token in-process and reuse until shortly before it expires.
_books_access_token: Optional[str] = None
_books_token_expires_at: Optional[datetime] = None
_inventory_access_token: Optional[str] = None
_inventory_token_expires_at: Optional[datetime] = None
_TOKEN_REFRESH_BUFFER = timedelta(minutes=5)


def _get_zoho_token() -> str:
    global _books_access_token, _books_token_expires_at
    if (
        _books_access_token
        and _books_token_expires_at
        and datetime.now() < (_books_token_expires_at - _TOKEN_REFRESH_BUFFER)
    ):
        return _books_access_token

    url = BOOKS_URL.format(
        clientId=CLIENT_ID,
        clientSecret=CLIENT_SECRET,
        grantType="refresh_token",
        books_refresh_token=BOOKS_REFRESH_TOKEN,
    )
    r = requests.post(url, timeout=30)
    r.raise_for_status()
    data = r.json()
    _books_access_token = data["access_token"]
    _books_token_expires_at = datetime.now() + timedelta(seconds=data.get("expires_in", 3600))
    return _books_access_token


def _get_inventory_token() -> str:
    global _inventory_access_token, _inventory_token_expires_at
    if (
        _inventory_access_token
        and _inventory_token_expires_at
        and datetime.now() < (_inventory_token_expires_at - _TOKEN_REFRESH_BUFFER)
    ):
        return _inventory_access_token

    r = requests.post(
        _ZOHO_TOKEN_URL,
        params={
            "refresh_token": INVENTORY_REFRESH_TOKEN,
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "grant_type": "refresh_token",
        },
        timeout=30,
    )
    r.raise_for_status()
    data = r.json()
    _inventory_access_token = data["access_token"]
    _inventory_token_expires_at = datetime.now() + timedelta(seconds=data.get("expires_in", 3600))
    return _inventory_access_token


_INVENTORY_TRANSIT_ACCOUNT_NAME = "Inventory Transit (Loss/Gain)"
_inventory_transit_account_id_cache: Optional[str] = None


def _get_inventory_transit_account_id() -> Optional[str]:
    """Resolve the 'Inventory Transit (Loss/Gain)' chart-of-accounts id (used on adjustment line
    items), looked up by name and cached for the process lifetime — avoids hardcoding an id that
    could silently point at the wrong GL account if it's ever recreated."""
    global _inventory_transit_account_id_cache
    if _inventory_transit_account_id_cache:
        return _inventory_transit_account_id_cache
    try:
        token = _get_zoho_token()
        r = requests.get(
            f"{ZOHO_BOOKS_BASE}/chartofaccounts",
            headers={"Authorization": f"Zoho-oauthtoken {token}"},
            params={"organization_id": ORGANIZATION_ID, "account_name": _INVENTORY_TRANSIT_ACCOUNT_NAME},
            timeout=30,
        )
        r.raise_for_status()
        accounts = r.json().get("chartofaccounts", [])
        for acc in accounts:
            if acc.get("account_name") == _INVENTORY_TRANSIT_ACCOUNT_NAME:
                _inventory_transit_account_id_cache = acc.get("account_id")
                return _inventory_transit_account_id_cache
    except Exception as e:
        logger.warning("Could not resolve '%s' account id: %s", _INVENTORY_TRANSIT_ACCOUNT_NAME, e)
    return None


def _load_products_by_item_id(db, item_ids: list[str]) -> dict[str, dict]:
    item_ids = [i for i in set(item_ids) if i]
    if not item_ids:
        return {}
    return {
        p["item_id"]: p
        for p in db.get_collection(PRODUCTS_COLLECTION).find(
            {"item_id": {"$in": item_ids}},
            {"item_id": 1, "cf_item_code": 1, "cf_sku_code": 1, "hsn_or_sac": 1},
        )
    }


def _line_item_custom_fields(prod: dict) -> list[dict]:
    return [
        {"customfield_id": BILL_LINE_ITEM_CF_MFR_CODE, "value": (prod or {}).get("cf_item_code", "") or ""},
        {"customfield_id": BILL_LINE_ITEM_CF_SKU_CODE, "value": (prod or {}).get("cf_sku_code", "") or ""},
    ]


def _unbilled_receive_lines(headers: dict, po_data: dict) -> list[dict]:
    """Receive line items on this PO that are received but not yet billed.

    Once a PO has un-billed receives, Zoho refuses a plain PO→bill conversion
    (error 36510 "The purchase order(s) have un-billed receive(s) …") — the bill must be
    raised against the receives. Each returned line carries the receive_id/receive_item_id
    pair Zoho expects on the bill line item.
    """
    receives = po_data.get("purchasereceives") or []
    open_receives = [r for r in receives if (r.get("billed_status") or "") != "billed"]
    if not open_receives:
        return []

    lines = []
    for rcv in open_receives:
        receive_id = rcv.get("receive_id") or rcv.get("purchasereceive_id")
        if not receive_id:
            continue
        r = requests.get(
            f"{ZOHO_BOOKS_BASE}/purchasereceives/{receive_id}",
            headers=headers,
            params={"organization_id": ORGANIZATION_ID},
            timeout=30,
        )
        r.raise_for_status()
        detail = r.json().get("purchasereceive") or {}
        for rli in detail.get("line_items", []):
            qty = (rli.get("quantity") or 0) - (rli.get("quantity_billed") or 0)
            if qty <= 0:
                continue
            lines.append(
                {
                    "item_id": rli.get("item_id"),
                    "quantity": qty,
                    "receive_id": receive_id,
                    "receive_item_id": rli.get("line_item_id"),
                }
            )
    return lines


def _needs_fix(bill: dict) -> bool:
    # Only inventory lines (those with an item_id) can carry SKU/Manufacturer-Code custom
    # fields — and only those can be repaired by fix-custom-fields. Non-product lines
    # (freight, "Samples", other service/charge lines with no item_id) legitimately have
    # none, so flagging them would show an amber warning that the Fix button can never clear.
    return any(li.get("item_id") and not li.get("item_custom_fields") for li in bill.get("line_items", []))


# ---------------------------------------------------------------------------
# PO ↔ Bill discrepancy comparison
# ---------------------------------------------------------------------------
# Zoho does NOT retro-apply a PO edit to a bill that was already raised against it, so a
# bill can silently drift from its PO (e.g. PO-JOLLYPAWPS01: one line's rate was corrected
# on the PO after the bill existed, leaving the bill $23.48 over). Nothing in Zoho surfaces
# that, hence this comparison.

_AMOUNT_EPS = 0.005  # Zoho rounds money to 2dp; anything smaller is float noise, not a discrepancy
_QTY_EPS = 0.0005
_VOID_BILL_STATUSES = {"void", "cancelled"}


def _f(value) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _po_ids_for_bill(bill: dict) -> list[str]:
    """PO ids this bill is raised against.

    `purchaseorder_ids` is the header-level link, but bills built from receives also carry
    `purchaseorder_id` per line — and on some older bills only the line-level link is set.
    """
    po_ids: list[str] = []
    for pid in bill.get("purchaseorder_ids") or []:
        if pid and pid not in po_ids:
            po_ids.append(pid)
    for li in bill.get("line_items") or []:
        pid = li.get("purchaseorder_id")
        if pid and pid not in po_ids:
            po_ids.append(pid)
    return po_ids


def _line_label(li: dict) -> dict:
    return {
        "item_id": li.get("item_id"),
        "name": li.get("name") or li.get("description") or "",
        "sku": li.get("sku") or "",
    }


def _match_bill_lines_to_po_lines(po_lines: list[dict], bill_lines: list[dict]) -> tuple[dict[str, list[dict]], list[dict]]:
    """Group bill lines under the PO line they bill.

    One PO line commonly maps to SEVERAL bill lines — a receive-derived line per receipt plus
    a balance line for the un-received remainder — so this is many-to-one, not one-to-one.
    Matching prefers Zoho's own `purchaseorder_item_id` back-reference; where that is absent
    (manually keyed bills) it falls back to item_id, consuming PO lines in order so a repeated
    item still maps to distinct PO lines.
    """
    by_po_line_id = {pl["line_item_id"]: pl for pl in po_lines if pl.get("line_item_id")}
    grouped: dict[str, list[dict]] = {pl_id: [] for pl_id in by_po_line_id}
    unmatched: list[dict] = []

    po_lines_by_item: dict[str, list[dict]] = {}
    for pl in po_lines:
        if pl.get("item_id"):
            po_lines_by_item.setdefault(pl["item_id"], []).append(pl)
    fallback_cursor: dict[str, int] = {}

    for bl in bill_lines:
        po_item_id = bl.get("purchaseorder_item_id")
        if po_item_id and po_item_id in by_po_line_id:
            grouped[po_item_id].append(bl)
            continue

        candidates = po_lines_by_item.get(bl.get("item_id") or "") or []
        if not candidates:
            unmatched.append(bl)
            continue
        # Reuse the last candidate once exhausted: extra bill lines for an item that IS on the
        # PO are an over-billing of that PO line, not an unrelated line.
        idx = min(fallback_cursor.get(bl["item_id"], 0), len(candidates) - 1)
        fallback_cursor[bl["item_id"]] = idx + 1
        grouped[candidates[idx]["line_item_id"]].append(bl)

    return grouped, unmatched


def _compare_bill_to_pos(db, bill: dict) -> dict:
    """Line-by-line diff of a bill against the PO(s) it is raised against."""
    po_ids = _po_ids_for_bill(bill)
    if not po_ids:
        return {"linked": False, "has_discrepancy": False, "reason": "This bill is not linked to a purchase order."}

    pos = list(db.get_collection(PURCHASE_ORDERS_COLLECTION).find({"purchaseorder_id": {"$in": po_ids}}))
    if not pos:
        return {"linked": False, "has_discrepancy": False, "reason": "Linked purchase order not found in the local sync."}

    po_lines: list[dict] = []
    for po in pos:
        for pl in po.get("line_items") or []:
            po_lines.append({**pl, "_po_number": po.get("purchaseorder_number")})

    bill_lines = bill.get("line_items") or []
    grouped, unmatched = _match_bill_lines_to_po_lines(po_lines, bill_lines)

    lines: list[dict] = []
    for pl in po_lines:
        billed = grouped.get(pl.get("line_item_id"), [])
        po_qty = _f(pl.get("quantity"))
        po_rate = _f(pl.get("rate"))
        po_amount = _f(pl.get("item_total"))
        bill_qty = sum(_f(b.get("quantity")) for b in billed)
        bill_amount = sum(_f(b.get("item_total")) for b in billed)
        # A PO line split across receives can legitimately produce several bill lines; only the
        # distinct RATES matter, and only when they disagree with the PO.
        bill_rates = sorted({round(_f(b.get("rate")), 6) for b in billed})

        qty_diff = round(bill_qty - po_qty, 4)
        amount_diff = round(bill_amount - po_amount, 2)
        rate_mismatch = [r for r in bill_rates if abs(r - po_rate) > 1e-9]

        issues: list[str] = []
        if billed and rate_mismatch:
            issues.append("rate")
        if abs(qty_diff) > _QTY_EPS:
            issues.append("over_billed" if qty_diff > 0 else ("not_billed" if not billed else "under_billed"))
        if abs(amount_diff) > _AMOUNT_EPS and not issues:
            # Amount moved without qty or rate moving — discount/tax-inclusive settings differ.
            issues.append("amount")

        lines.append(
            {
                **_line_label(pl),
                "po_number": pl.get("_po_number"),
                "po_line_item_id": pl.get("line_item_id"),
                "po_quantity": round(po_qty, 4),
                "bill_quantity": round(bill_qty, 4),
                "quantity_diff": qty_diff,
                "po_rate": round(po_rate, 6),
                "bill_rates": bill_rates,
                "rate_diff": round(rate_mismatch[0] - po_rate, 6) if rate_mismatch else 0.0,
                "po_amount": round(po_amount, 2),
                "bill_amount": round(bill_amount, 2),
                "amount_diff": amount_diff,
                "bill_line_count": len(billed),
                "issues": issues,
            }
        )

    extra_lines = [
        {
            **_line_label(bl),
            "bill_quantity": round(_f(bl.get("quantity")), 4),
            "bill_rate": round(_f(bl.get("rate")), 6),
            "bill_amount": round(_f(bl.get("item_total")), 2),
        }
        for bl in unmatched
    ]

    # Other bills against the same PO(s) — without this, a PO billed in two parts reads as a
    # large "under-billed" discrepancy on each bill when nothing is actually wrong.
    other_bills = list(
        db.get_collection(BILLS_COLLECTION).find(
            {"purchaseorder_ids": {"$in": po_ids}, "bill_id": {"$ne": bill.get("bill_id")}},
            {"bill_id": 1, "bill_number": 1, "status": 1, "total": 1, "currency_code": 1},
        )
    )
    other_bills = [b for b in other_bills if (b.get("status") or "").lower() not in _VOID_BILL_STATUSES]

    po_total = round(sum(_f(p.get("total")) for p in pos), 2)
    po_sub_total = round(sum(_f(p.get("sub_total")) for p in pos), 2)
    po_tax_total = round(sum(_f(p.get("tax_total")) for p in pos), 2)
    po_quantity = round(sum(_f(pl.get("quantity")) for pl in po_lines), 4)
    bill_total = round(_f(bill.get("total")), 2)
    bill_quantity = round(sum(_f(b.get("quantity")) for b in bill_lines), 4)

    po_currency = (pos[0].get("currency_code") or "").upper()
    bill_currency = (bill.get("currency_code") or "").upper()
    po_fx = round(_f(pos[0].get("exchange_rate")), 6)
    bill_fx = round(_f(bill.get("exchange_rate")), 6)

    header_issues: list[dict] = []
    if po_currency and bill_currency and po_currency != bill_currency:
        header_issues.append({"field": "Currency", "po": po_currency, "bill": bill_currency, "severity": "error"})
    if po_fx and bill_fx and abs(po_fx - bill_fx) > 1e-6:
        # Normal FX drift between PO date and bill date — informational, never an error.
        header_issues.append({"field": "Exchange rate", "po": po_fx, "bill": bill_fx, "severity": "info"})
    if abs(po_tax_total - round(_f(bill.get("tax_total")), 2)) > _AMOUNT_EPS:
        header_issues.append(
            {"field": "Tax total", "po": po_tax_total, "bill": round(_f(bill.get("tax_total")), 2), "severity": "warn"}
        )
    if abs(_f(bill.get("adjustment"))) > _AMOUNT_EPS:
        header_issues.append(
            {"field": "Adjustment on bill", "po": 0, "bill": round(_f(bill.get("adjustment")), 2), "severity": "warn"}
        )

    flagged = [ln for ln in lines if ln["issues"]]
    has_discrepancy = bool(flagged or extra_lines) or any(h["severity"] != "info" for h in header_issues)

    return {
        "linked": True,
        "has_discrepancy": has_discrepancy,
        "purchase_orders": [
            {"purchaseorder_id": p.get("purchaseorder_id"), "purchaseorder_number": p.get("purchaseorder_number"), "status": p.get("status")}
            for p in pos
        ],
        "currency_code": bill_currency or po_currency,
        "totals": {
            "po_total": po_total,
            "bill_total": bill_total,
            "total_diff": round(bill_total - po_total, 2),
            "po_sub_total": po_sub_total,
            "bill_sub_total": round(_f(bill.get("sub_total")), 2),
            "po_tax_total": po_tax_total,
            "bill_tax_total": round(_f(bill.get("tax_total")), 2),
            "po_quantity": po_quantity,
            "bill_quantity": bill_quantity,
            "quantity_diff": round(bill_quantity - po_quantity, 4),
        },
        "lines": lines,
        "flagged_count": len(flagged),
        "extra_lines": extra_lines,
        "header_issues": header_issues,
        "other_bills": serialize_mongo_document(other_bills),
        "other_bills_total": round(sum(_f(b.get("total")) for b in other_bills), 2),
    }


# Zoho's bill-update API only accepts a small subset of the fields it returns on GET —
# echoing the full GET response back (e.g. purchaseorder_details, item_type_formatted,
# rate_formatted, ...) trips its own field-length/type validation on unrelated read-only
# fields. Keep only what's needed to preserve the line item and set the custom fields.
_LINE_ITEM_WRITABLE_FIELDS = (
    "line_item_id",
    "item_id",
    "name",
    "description",
    "quantity",
    "rate",
    "unit",
    "account_id",
    "hsn_or_sac",
    "tax_id",
    "item_custom_fields",
)


def _writable_line_item(li: dict) -> dict:
    return {k: li[k] for k in _LINE_ITEM_WRITABLE_FIELDS if k in li}


def _upsert_bill(db, bill: dict):
    if not bill or not bill.get("bill_id"):
        return
    db.get_collection(BILLS_COLLECTION).update_one(
        {"bill_id": bill["bill_id"]},
        {"$set": bill},
        upsert=True,
    )


def _upsert_inventory_adjustment(db, adjustment: dict):
    if not adjustment or not adjustment.get("inventory_adjustment_id"):
        return
    db.get_collection(INVENTORY_ADJUSTMENTS_COLLECTION).update_one(
        {"inventory_adjustment_id": adjustment["inventory_adjustment_id"]},
        {"$set": adjustment},
        upsert=True,
    )


def _link_adjustment_to_bill(db, bill_id: str, adjustment_id: str):
    db.get_collection(BILLS_COLLECTION).update_one(
        {"bill_id": bill_id},
        {"$addToSet": {"linked_inventory_adjustment_ids": adjustment_id}},
        upsert=True,
    )


def _resolve_bill_and_po_numbers(db, bill_id: str) -> tuple[Optional[str], Optional[str]]:
    """Live-fetch the bill from Zoho for its bill_number + linked PO number — doesn't depend on a
    local `bills` doc existing, since plain GETs never upsert one."""
    try:
        bill = _fetch_bill(bill_id)
    except Exception as e:
        logger.warning("Could not fetch bill %s to resolve numbers: %s", bill_id, e)
        return None, None

    bill_number = bill.get("bill_number")
    po_number = None
    po_ids = bill.get("purchaseorder_ids") or []
    if po_ids:
        po_doc = db.get_collection(PURCHASE_ORDERS_COLLECTION).find_one(
            {"purchaseorder_id": po_ids[0]}, {"purchaseorder_number": 1}
        )
        po_number = (po_doc or {}).get("purchaseorder_number")
    return bill_number, po_number


# Reference numbers shorter than this are too generic to auto-discover on. Vendor bill numbers are
# often a bare counter ("1"), which as a substring matches half the adjustments in the system
# (e.g. bill "1" matched "CN/26-27/0241"). Real linking conventions use the full PO/shipment number.
_MIN_DISCOVERY_REF_LEN = 4


def _discovery_refs(*refs: Optional[str]) -> list[str]:
    """The subset of candidate reference numbers distinctive enough to auto-link on."""
    return [r for r in refs if r and len(r.strip()) >= _MIN_DISCOVERY_REF_LEN]


def _ref_regex(ref: str) -> str:
    """Match `ref` only as a whole token, so "PO-X01" doesn't match "PO-X011"."""
    return rf"(?<![A-Za-z0-9]){re.escape(ref.strip())}(?![A-Za-z0-9])"


def _ref_matches(ref: str, reference_number: Optional[str]) -> bool:
    return bool(reference_number) and re.search(_ref_regex(ref), reference_number, re.IGNORECASE) is not None


def _linked_adjustments_for_bill(db, bill_id: str) -> list[dict]:
    bill_doc = db.get_collection(BILLS_COLLECTION).find_one({"bill_id": bill_id}, {"linked_inventory_adjustment_ids": 1})
    ids = set((bill_doc or {}).get("linked_inventory_adjustment_ids") or [])

    # Auto-discover adjustments created manually in Zoho whose reference_number embeds this bill's
    # number or its PO's number as a whole token (the convention used historically, e.g.
    # "SH26E0558 ( PO-PETZOO066 )"), even though no explicit link action was ever taken for them.
    bill_number, po_number = _resolve_bill_and_po_numbers(db, bill_id)
    or_clauses = [{"reference_number": {"$regex": _ref_regex(ref), "$options": "i"}} for ref in _discovery_refs(bill_number, po_number)]
    discovered_ids: set[str] = set()
    if or_clauses:
        for doc in db.get_collection(INVENTORY_ADJUSTMENTS_COLLECTION).find({"$or": or_clauses}, {"inventory_adjustment_id": 1}):
            discovered_ids.add(doc["inventory_adjustment_id"])

    new_ids = discovered_ids - ids
    if new_ids:
        db.get_collection(BILLS_COLLECTION).update_one(
            {"bill_id": bill_id},
            {"$addToSet": {"linked_inventory_adjustment_ids": {"$each": list(new_ids)}}},
            upsert=True,
        )

    all_ids = list(ids | discovered_ids)
    if not all_ids:
        return []

    docs = list(
        db.get_collection(INVENTORY_ADJUSTMENTS_COLLECTION).find(
            {"inventory_adjustment_id": {"$in": all_ids}},
            {
                "inventory_adjustment_id": 1,
                "reference_number": 1,
                "date": 1,
                "reason": 1,
                "status": 1,
                "adjustment_type": 1,
                "description": 1,
            },
        )
    )
    return serialize_mongo_document(docs)


def _vendor_brand_map(db) -> dict[str, list[str]]:
    """{vendor contact_id: [brand names]} for every vendor assigned to at least one brand."""
    result: dict[str, set[str]] = {}
    for b in db.get_collection("brands").find({}, {"name": 1, "vendor_ids": 1, "vendor_id": 1}):
        vids = b.get("vendor_ids") or ([b["vendor_id"]] if b.get("vendor_id") else [])
        name = b.get("name")
        if not name:
            continue
        for vid in vids:
            if not vid:
                continue
            result.setdefault(vid, set()).add(name)
    return {vid: sorted(names) for vid, names in result.items()}


@router.get("/list")
def list_bills(
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(25, ge=1, le=200),
    db=Depends(get_database),
):
    def _fetch():
        token = _get_zoho_token()
        headers = {"Authorization": f"Zoho-oauthtoken {token}"}
        params = {
            "organization_id": ORGANIZATION_ID,
            "page": page,
            "per_page": per_page,
            "sort_column": "date",
            "sort_order": "D",
        }
        if search:
            params["search_text"] = search
        r = requests.get(f"{ZOHO_BOOKS_BASE}/bills", headers=headers, params=params, timeout=30)
        r.raise_for_status()
        data = r.json()
        if data.get("code") != 0:
            raise ValueError(f"Zoho error: {data.get('message', 'Unknown error')}")
        return data.get("bills", []), data.get("page_context", {})

    try:
        bills, page_context = _fetch()
    except Exception as e:
        logger.error("Failed to list bills: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

    return {"bills": bills, "page_context": page_context}


@router.get("/po-groups")
def po_groups(search: Optional[str] = Query(None), po: Optional[str] = Query(None), db=Depends(get_database)):
    """Vendors assigned to at least one brand, with their PO/order numbers and any linked bill(s)."""
    vendor_brands = _vendor_brand_map(db)
    vendor_ids = list(vendor_brands.keys())
    if not vendor_ids:
        return {"vendors": []}

    and_clauses: list[dict] = [{"vendor_id": {"$in": vendor_ids}}]
    if search:
        and_clauses.append(
            {
                "$or": [
                    {"purchaseorder_number": {"$regex": search, "$options": "i"}},
                    {"vendor_name": {"$regex": search, "$options": "i"}},
                ]
            }
        )
    query = {"$and": and_clauses} if len(and_clauses) > 1 else and_clauses[0]

    projection = {
        "purchaseorder_number": 1,
        "purchaseorder_id": 1,
        "vendor_id": 1,
        "vendor_name": 1,
        "date": 1,
        "total": 1,
        "currency_code": 1,
        "exchange_rate": 1,
        "status": 1,
    }
    pos = list(db.get_collection(PURCHASE_ORDERS_COLLECTION).find(query, projection))

    # Deep-link guarantee: make sure the requested PO is present even if it fell outside the search filter.
    if po and not any(p.get("purchaseorder_number") == po for p in pos):
        extra = db.get_collection(PURCHASE_ORDERS_COLLECTION).find_one({"purchaseorder_number": po}, projection)
        if extra:
            pos.append(extra)

    # Join brand_orders (by purchaseorder_number) for the order's own name/number and created_at —
    # created_at drives sorting, matching BrandOrders.tsx's ordersByVendor/vendorList sort.
    po_numbers = [p["purchaseorder_number"] for p in pos if p.get("purchaseorder_number")]
    order_by_po: dict[str, dict] = {}
    if po_numbers:
        for o in db.get_collection(BRAND_ORDERS_COLLECTION).find(
            {"purchaseorder_number": {"$in": po_numbers}},
            {"purchaseorder_number": 1, "name": 1, "created_at": 1},
        ):
            pon = o.get("purchaseorder_number")
            if pon and pon not in order_by_po:
                order_by_po[pon] = o

    po_number_by_id = {p["purchaseorder_id"]: p.get("purchaseorder_number") for p in pos if p.get("purchaseorder_id")}
    po_ids = list(po_number_by_id.keys())
    bill_docs: list[dict] = []
    if po_ids:
        bill_docs = list(
            db.get_collection(BILLS_COLLECTION).find(
                {"purchaseorder_ids": {"$in": po_ids}},
                {
                    "bill_id": 1,
                    "bill_number": 1,
                    "status": 1,
                    "total": 1,
                    "total_formatted": 1,
                    "currency_code": 1,
                    "currency_symbol": 1,
                    "purchaseorder_ids": 1,
                    "line_items": 1,
                    "linked_inventory_adjustment_ids": 1,
                },
            )
        )

    # Batch-detect inventory adjustments per bill — same whole-token reference_number convention
    # as _linked_adjustments_for_bill, but done once for the whole page (one query) instead of
    # per-bill, so it shows up here without needing to open each bill.
    all_candidate_refs: set[str] = set()
    bill_candidate_refs: dict[str, set[str]] = {}
    for b in bill_docs:
        refs = set(_discovery_refs(b.get("bill_number")))
        refs.update(_discovery_refs(*(po_number_by_id.get(pid) for pid in b.get("purchaseorder_ids") or [])))
        bill_candidate_refs[b["bill_id"]] = refs
        all_candidate_refs.update(refs)

    matched_adjustments: list[dict] = []
    if all_candidate_refs:
        or_clauses = [{"reference_number": {"$regex": _ref_regex(ref), "$options": "i"}} for ref in all_candidate_refs]
        matched_adjustments = list(
            db.get_collection(INVENTORY_ADJUSTMENTS_COLLECTION).find({"$or": or_clauses}, {"inventory_adjustment_id": 1, "reference_number": 1})
        )

    def _adjustment_count_for(bill: dict) -> int:
        ids = set(bill.get("linked_inventory_adjustment_ids") or [])
        refs = bill_candidate_refs.get(bill["bill_id"], set())
        for adj in matched_adjustments:
            if any(_ref_matches(ref, adj.get("reference_number")) for ref in refs):
                ids.add(adj["inventory_adjustment_id"])
        return len(ids)

    bills_by_po: dict[str, list[dict]] = {}
    for b in bill_docs:
        entry = {
            "bill_id": b.get("bill_id"),
            "bill_number": b.get("bill_number"),
            "status": b.get("status"),
            "total": b.get("total"),
            "total_formatted": b.get("total_formatted"),
            "currency_code": b.get("currency_code"),
            "currency_symbol": b.get("currency_symbol"),
            "needs_fix": _needs_fix(b),
            "adjustment_count": _adjustment_count_for(b),
        }
        for pid in b.get("purchaseorder_ids") or []:
            bills_by_po.setdefault(pid, []).append(entry)

    groups: dict[str, dict] = {}
    for p in pos:
        vid = p.get("vendor_id") or "__unknown__"
        group = groups.setdefault(
            vid,
            {
                "vendor_id": vid,
                "vendor_name": p.get("vendor_name") or "Unknown Vendor",
                "currency_code": p.get("currency_code"),
                "brands": vendor_brands.get(vid, []),
                "purchase_orders": [],
            },
        )
        order_info = order_by_po.get(p.get("purchaseorder_number")) or {}
        real_ts = order_info.get("created_at") or ""  # only counts if backed by an actual brand_orders doc
        sort_ts = real_ts or p.get("date") or ""  # PO date is just a display-order fallback within the vendor
        po_bills = bills_by_po.get(p.get("purchaseorder_id"), [])

        # Cheap header-level discrepancy flag for the collapsed row: total of all live bills vs
        # the PO total. Deliberately NOT the full line-by-line diff — that needs every PO's line
        # items loaded for every vendor on the page. Opening the bill runs the real comparison.
        live_bills = [b for b in po_bills if (b.get("status") or "").lower() not in _VOID_BILL_STATUSES]
        billed_total = round(sum(_f(b.get("total")) for b in live_bills), 2)
        same_currency = all(
            not b.get("currency_code") or not p.get("currency_code") or b["currency_code"] == p["currency_code"]
            for b in live_bills
        )
        bill_total_diff = (
            round(billed_total - _f(p.get("total")), 2)
            if live_bills and same_currency and abs(billed_total - _f(p.get("total"))) > _AMOUNT_EPS
            else None
        )

        group["purchase_orders"].append(
            {
                "purchaseorder_number": p.get("purchaseorder_number"),
                "purchaseorder_id": p.get("purchaseorder_id"),
                "order_name": order_info.get("name"),
                "date": p.get("date"),
                "total": p.get("total"),
                "currency_code": p.get("currency_code"),
                "exchange_rate": p.get("exchange_rate"),
                "status": p.get("status"),
                "bills": po_bills,
                "billed_total": billed_total if live_bills else None,
                "bill_total_diff": bill_total_diff,
                "_sort_ts": sort_ts,
                "_real_ts": real_ts,
            }
        )

    # Mirror BrandOrders.tsx's ordersByVendor/vendorList sort: most-recent order created_at first
    # (falling back to PO date within a vendor's own list when there's no matching brand_orders doc),
    # tie-break by vendor name asc. Vendor-level ranking uses ONLY real brand_orders created_at —
    # exactly like vendorList's aLatest/bLatest — so vendors whose POs never went through the Brand
    # Orders flow don't get bumped up the list by an unrelated PO date.
    for group in groups.values():
        group["purchase_orders"].sort(key=lambda e: str(e.get("_sort_ts") or ""), reverse=True)

    vendors = sorted(groups.values(), key=lambda g: g["vendor_name"] or "")
    vendors.sort(
        key=lambda g: max((str(po.get("_real_ts") or "") for po in g["purchase_orders"]), default=""),
        reverse=True,
    )
    for v in vendors:
        for e in v["purchase_orders"]:
            e.pop("_sort_ts", None)
            e.pop("_real_ts", None)

    return {"vendors": serialize_mongo_document(vendors)}


@router.get("/po-options")
def po_options(
    search: Optional[str] = Query(None),
    vendor_id: Optional[str] = Query(None),
    limit: int = Query(20, ge=1, le=100),
    db=Depends(get_database),
):
    query: dict = {}
    if search:
        query["purchaseorder_number"] = {"$regex": search, "$options": "i"}
    if vendor_id:
        query["vendor_id"] = vendor_id

    docs = list(
        db.get_collection(PURCHASE_ORDERS_COLLECTION)
        .find(
            query,
            {
                "purchaseorder_number": 1,
                "purchaseorder_id": 1,
                "vendor_id": 1,
                "vendor_name": 1,
                "date": 1,
                "total": 1,
                # drive the Create Bill modal's conversion-rate field (shown only for foreign
                # currency; the PO's own rate is the prefill)
                "currency_code": 1,
                "exchange_rate": 1,
            },
        )
        .sort([("date", -1), ("_id", -1)])
        .limit(limit)
    )
    return {"options": serialize_mongo_document(docs)}


@router.get("/{bill_id}")
def get_bill(bill_id: str, db=Depends(get_database)):
    def _fetch():
        token = _get_zoho_token()
        headers = {"Authorization": f"Zoho-oauthtoken {token}"}
        r = requests.get(
            f"{ZOHO_BOOKS_BASE}/bills/{bill_id}",
            headers=headers,
            params={"organization_id": ORGANIZATION_ID},
            timeout=30,
        )
        r.raise_for_status()
        data = r.json()
        if data.get("code") != 0:
            raise ValueError(f"Zoho error: {data.get('message', 'Unknown error')}")
        return data["bill"]

    try:
        bill = _fetch()
    except Exception as e:
        logger.error("Failed to fetch bill %s: %s", bill_id, e)
        raise HTTPException(status_code=500, detail=str(e))

    # Folded into the detail response rather than its own endpoint: the panel that renders it
    # opens with the bill, so a second round-trip would only add latency.
    try:
        comparison = _compare_bill_to_pos(db, bill)
    except Exception as e:
        logger.error("Failed to compare bill %s to its PO: %s", bill_id, e)
        comparison = {"linked": False, "has_discrepancy": False, "reason": "Could not compare against the purchase order."}

    return {"bill": bill, "needs_fix": _needs_fix(bill), "comparison": comparison}


@router.post("/{bill_id}/fix-custom-fields")
def fix_custom_fields(bill_id: str, db=Depends(get_database)):
    def _do_fix():
        token = _get_zoho_token()
        headers = {"Authorization": f"Zoho-oauthtoken {token}"}

        r = requests.get(
            f"{ZOHO_BOOKS_BASE}/bills/{bill_id}",
            headers=headers,
            params={"organization_id": ORGANIZATION_ID},
            timeout=30,
        )
        r.raise_for_status()
        data = r.json()
        if data.get("code") != 0:
            raise ValueError(f"Zoho error: {data.get('message', 'Unknown error')}")
        bill = data["bill"]
        line_items = bill.get("line_items", [])

        empty_item_ids = [li["item_id"] for li in line_items if not li.get("item_custom_fields") and li.get("item_id")]
        if not empty_item_ids:
            return bill, 0

        prod_by_id = _load_products_by_item_id(db, empty_item_ids)

        fixed_count = 0
        for li in line_items:
            if li.get("item_custom_fields"):
                continue
            prod = prod_by_id.get(li.get("item_id"), {})
            li["item_custom_fields"] = _line_item_custom_fields(prod)
            if not li.get("hsn_or_sac"):
                li["hsn_or_sac"] = prod.get("hsn_or_sac", "")
            fixed_count += 1

        if fixed_count == 0:
            return bill, 0

        put_line_items = [_writable_line_item(li) for li in line_items]

        put_r = requests.put(
            f"{ZOHO_BOOKS_BASE}/bills/{bill_id}",
            headers=headers,
            json={"line_items": put_line_items},
            params={"organization_id": ORGANIZATION_ID},
            timeout=30,
        )
        logger.info("Zoho bill update response %s: %s", put_r.status_code, put_r.text)
        put_r.raise_for_status()
        put_data = put_r.json()
        if put_data.get("code") != 0:
            raise ValueError(f"Zoho error: {put_data.get('message', 'Unknown error')}")

        return put_data["bill"], fixed_count

    try:
        bill, fixed_count = _do_fix()
    except Exception as e:
        logger.error("Failed to fix custom fields for bill %s: %s", bill_id, e)
        raise HTTPException(status_code=500, detail=str(e))

    _upsert_bill(db, bill)
    return {"bill": bill, "fixed_count": fixed_count, "needs_fix": _needs_fix(bill)}


@router.post("/create")
def create_bill(
    purchaseorder_number: str = Body(..., embed=True),
    bill_number: str = Body(..., embed=True),
    date: str = Body(..., embed=True),
    original_bill_date: Optional[str] = Body(None, embed=True),
    payment_terms: int = Body(0, embed=True),
    payment_terms_label: str = Body(DEFAULT_PAYMENT_TERMS_LABEL, embed=True),
    exchange_rate: Optional[float] = Body(None, embed=True),
    db=Depends(get_database),
):
    po = db.get_collection(PURCHASE_ORDERS_COLLECTION).find_one(
        {"purchaseorder_number": purchaseorder_number},
        {"purchaseorder_id": 1, "vendor_id": 1, "currency_code": 1},
    )
    if not po or not po.get("purchaseorder_id"):
        raise HTTPException(status_code=404, detail=f"Purchase order {purchaseorder_number} not found")

    # Zoho rejects an exchange_rate on base-currency bills.
    if exchange_rate is not None:
        if (po.get("currency_code") or BASE_CURRENCY) == BASE_CURRENCY:
            raise HTTPException(
                status_code=400,
                detail=f"Purchase order {purchaseorder_number} is in {BASE_CURRENCY}; a conversion rate does not apply.",
            )
        if exchange_rate <= 0:
            raise HTTPException(status_code=400, detail="Conversion rate must be greater than zero.")

    brand_order = db.get_collection(BRAND_ORDERS_COLLECTION).find_one(
        {"purchaseorder_number": purchaseorder_number}, {"name": 1}
    )
    order_number = (brand_order or {}).get("name") or purchaseorder_number

    def _do_create():
        token = _get_zoho_token()
        headers = {"Authorization": f"Zoho-oauthtoken {token}"}

        po_r = requests.get(
            f"{ZOHO_BOOKS_BASE}/purchaseorders/{po['purchaseorder_id']}",
            headers=headers,
            params={"organization_id": ORGANIZATION_ID},
            timeout=30,
        )
        po_r.raise_for_status()
        po_data = po_r.json()
        if po_data.get("code") != 0:
            raise ValueError(f"Zoho error: {po_data.get('message', 'Unknown error')}")
        po_line_items = po_data["purchaseorder"].get("line_items", [])
        if not po_line_items:
            raise ValueError("Purchase order has no line items")

        prod_by_id = _load_products_by_item_id(db, [li.get("item_id") for li in po_line_items])

        def _bill_line(pli: dict, quantity, extra: Optional[dict] = None) -> dict:
            prod = prod_by_id.get(pli.get("item_id"), {})
            return {
                "item_id": pli["item_id"],
                "quantity": quantity,
                "rate": pli.get("rate", 0),
                "account_id": DEFAULT_ACCOUNT_ID,
                "hsn_or_sac": pli.get("hsn_or_sac") or prod.get("hsn_or_sac", ""),
                "item_custom_fields": _line_item_custom_fields(prod),
                **(extra or {}),
            }

        receive_lines = _unbilled_receive_lines(headers, po_data["purchaseorder"])
        if receive_lines:
            # Bill the receives, not the PO lines — Zoho rejects the PO-level conversion while
            # un-billed receives exist. Match each receive line back to its PO line by item_id
            # (consuming PO lines in order so a repeated item maps to distinct PO lines).
            po_lines_by_item: dict[str, list[dict]] = {}
            for pli in po_line_items:
                po_lines_by_item.setdefault(pli.get("item_id"), []).append(pli)
            used: dict[str, int] = {}

            line_items = []
            for rl in receive_lines:
                candidates = po_lines_by_item.get(rl["item_id"]) or []
                if not candidates:
                    raise ValueError(
                        f"Receive line item {rl['item_id']} is not on purchase order {purchaseorder_number}"
                    )
                idx = min(used.get(rl["item_id"], 0), len(candidates) - 1)
                used[rl["item_id"]] = idx + 1
                pli = candidates[idx]
                line_items.append(
                    _bill_line(
                        pli,
                        rl["quantity"],
                        {
                            "purchaseorder_id": po["purchaseorder_id"],
                            "purchaseorder_item_id": pli.get("line_item_id"),
                            "receive_id": rl["receive_id"],
                            "receive_item_id": rl["receive_item_id"],
                        },
                    )
                )
        else:
            line_items = [
                _bill_line(pli, pli.get("quantity_ordered") or pli.get("quantity") or 0)
                for pli in po_line_items
            ]

        payload = {
            "vendor_id": po.get("vendor_id"),
            "purchaseorder_ids": [po["purchaseorder_id"]],
            "bill_number": bill_number,
            "reference_number": order_number,
            "date": date,
            "payment_terms": payment_terms,
            "payment_terms_label": payment_terms_label,
            "line_items": line_items,
        }

        if original_bill_date:
            payload["custom_fields"] = [{"api_name": ORIGINAL_BILL_DATE_FIELD, "value": original_bill_date}]

        # Without an explicit rate Zoho silently applies its own daily one, which is what drifted
        # this bill's total off the PO's (95.2 vs 95.360558).
        if exchange_rate is not None:
            payload["exchange_rate"] = exchange_rate

        logger.info("Zoho bill create payload: %s", payload)
        r = requests.post(
            f"{ZOHO_BOOKS_BASE}/bills",
            headers=headers,
            json=payload,
            params={
                "organization_id": ORGANIZATION_ID,
                "ignore_auto_number_generation": "true",
            },
            timeout=30,
        )
        logger.info("Zoho bill create response %s: %s", r.status_code, r.text)
        if r.status_code >= 400:
            # Surface Zoho's own message (e.g. "un-billed receive(s) …") instead of the bare
            # "400 Client Error" requests puts in HTTPError.
            try:
                msg = r.json().get("message")
            except ValueError:
                msg = None
            raise ValueError(f"Zoho error: {msg or r.text}")
        data = r.json()
        if data.get("code") != 0:
            raise ValueError(f"Zoho error: {data.get('message', 'Unknown error')}")
        return data["bill"]

    try:
        bill = _do_create()
    except Exception as e:
        logger.error("Failed to create bill for PO %s: %s", purchaseorder_number, e)
        raise HTTPException(status_code=500, detail=str(e))

    _upsert_bill(db, bill)
    return {"bill": bill}


# ─── Inventory Adjustments (warehouse short/extra count vs. a bill) ───────────

_IA_SHEETS = [("Short Received", "Short Qty"), ("Extra Received", "Extra Qty")]
_IA_HEADER_FILL = "FFC000"


def _style_ia_header(ws, col_count: int):
    fill = PatternFill("solid", fgColor=_IA_HEADER_FILL)
    font = Font(bold=True, size=11)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_col_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 10), 50)


def _fetch_bill(bill_id: str) -> dict:
    token = _get_zoho_token()
    headers = {"Authorization": f"Zoho-oauthtoken {token}"}
    r = requests.get(
        f"{ZOHO_BOOKS_BASE}/bills/{bill_id}",
        headers=headers,
        params={"organization_id": ORGANIZATION_ID},
        timeout=30,
    )
    r.raise_for_status()
    data = r.json()
    if data.get("code") != 0:
        raise ValueError(f"Zoho error: {data.get('message', 'Unknown error')}")
    return data["bill"]


@router.get("/{bill_id}/inventory-adjustment/template")
def inventory_adjustment_template(bill_id: str):
    try:
        bill = _fetch_bill(bill_id)
    except Exception as e:
        logger.error("Failed to fetch bill %s for IA template: %s", bill_id, e)
        raise HTTPException(status_code=500, detail=str(e))

    line_items = bill.get("line_items", [])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, qty_label in _IA_SHEETS:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Sr No", "SKU Code", "Product Name", qty_label]
        ws.append(headers)
        _style_ia_header(ws, len(headers))
        for idx, li in enumerate(line_items, start=1):
            ws.append([idx, li.get("sku", ""), li.get("name", ""), None])
        ws.freeze_panes = "A2"
        _auto_col_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"inventory_adjustment_{bill.get('bill_number', bill_id)}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _parse_ia_workbook(wb, bill_line_items: list[dict]) -> dict:
    sku_to_item = {li.get("sku"): li for li in bill_line_items if li.get("sku")}

    def _parse_sheet(sheet_name: str):
        rows = []
        unresolved = []
        if sheet_name not in wb.sheetnames:
            return rows, unresolved
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4:
                continue
            _, sku, name, qty = row[0], row[1], row[2], row[3]
            sku = str(sku).strip() if sku is not None else ""
            if not sku or qty in (None, ""):
                continue
            try:
                qty = float(qty)
            except (TypeError, ValueError):
                continue
            if qty == 0:
                continue
            li = sku_to_item.get(sku)
            entry = {"sku": sku, "name": name or (li or {}).get("name", ""), "qty": abs(qty)}
            if li:
                entry["item_id"] = li.get("item_id")
                rows.append(entry)
            else:
                unresolved.append(entry)
        return rows, unresolved

    short_rows, short_unresolved = _parse_sheet("Short Received")
    extra_rows, extra_unresolved = _parse_sheet("Extra Received")
    return {
        "short_received": short_rows,
        "extra_received": extra_rows,
        "unresolved": short_unresolved + extra_unresolved,
    }


@router.post("/{bill_id}/inventory-adjustment/preview")
async def inventory_adjustment_preview(bill_id: str, file: UploadFile = File(...)):
    try:
        bill = _fetch_bill(bill_id)
    except Exception as e:
        logger.error("Failed to fetch bill %s for IA preview: %s", bill_id, e)
        raise HTTPException(status_code=500, detail=str(e))

    try:
        data = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        result = _parse_ia_workbook(wb, bill.get("line_items", []))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse workbook: {e}")

    result["bill_number"] = bill.get("bill_number")
    return result


@router.post("/{bill_id}/inventory-adjustment/confirm")
async def inventory_adjustment_confirm(
    bill_id: str,
    file: UploadFile = File(...),
    description: Optional[str] = Form(None),
    db=Depends(get_database),
):
    try:
        bill = _fetch_bill(bill_id)
    except Exception as e:
        logger.error("Failed to fetch bill %s for IA confirm: %s", bill_id, e)
        raise HTTPException(status_code=500, detail=str(e))

    try:
        data = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        parsed = _parse_ia_workbook(wb, bill.get("line_items", []))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse workbook: {e}")

    # Net short (-) vs extra (+) per item_id, in case the same SKU appears on both sheets.
    net_by_item: dict[str, float] = {}
    name_by_item: dict[str, str] = {}
    for row in parsed["short_received"]:
        net_by_item[row["item_id"]] = net_by_item.get(row["item_id"], 0) - row["qty"]
        name_by_item[row["item_id"]] = row["name"]
    for row in parsed["extra_received"]:
        net_by_item[row["item_id"]] = net_by_item.get(row["item_id"], 0) + row["qty"]
        name_by_item[row["item_id"]] = row["name"]

    adjustment_account_id = _get_inventory_transit_account_id()

    line_items = [
        {
            "item_id": item_id,
            "name": name_by_item.get(item_id, ""),
            "quantity_adjusted": qty,
            "location_id": DEFAULT_WAREHOUSE_ID,
            **({"adjustment_account_id": adjustment_account_id} if adjustment_account_id else {}),
        }
        for item_id, qty in net_by_item.items()
        if qty != 0
    ]
    if not line_items:
        raise HTTPException(status_code=400, detail="No non-zero short/extra quantities found in the uploaded sheet.")

    # Resolve the PO number + order number/name backing this bill, e.g. bill_number "SH26E0558",
    # po_number "PO-PETZOO066", order_name "66" → ref# "SH26E0558 ( PO-PETZOO066 )" and description
    # "Shortage & Excess Received in Order 66" — mirrors the manually-created adjustment format.
    po_number = None
    po_ids = bill.get("purchaseorder_ids") or []
    if po_ids:
        po_doc = db.get_collection(PURCHASE_ORDERS_COLLECTION).find_one(
            {"purchaseorder_id": po_ids[0]}, {"purchaseorder_number": 1}
        )
        po_number = (po_doc or {}).get("purchaseorder_number")

    order_name = None
    if po_number:
        brand_order = db.get_collection(BRAND_ORDERS_COLLECTION).find_one({"purchaseorder_number": po_number}, {"name": 1})
        order_name = (brand_order or {}).get("name")

    bill_number = bill.get("bill_number")
    reference_number = f"{bill_number} ( {po_number} )" if po_number else bill_number
    # User-supplied notes win; otherwise fall back to the auto-generated description.
    final_description = (description or "").strip() or (
        f"Shortage & Excess Received in Order {order_name or po_number or bill_number}"
    )

    def _do_create():
        token = _get_inventory_token()
        headers = {"Authorization": f"Zoho-oauthtoken {token}"}
        payload = {
            "date": datetime.now().strftime("%Y-%m-%d"),
            "reason": "Shipment Inward adjustment",
            "description": final_description,
            "adjustment_type": "quantity",
            "reference_number": reference_number,
            "location_id": DEFAULT_WAREHOUSE_ID,
            "line_items": line_items,
        }
        logger.info("Zoho inventory adjustment payload: %s", payload)
        r = requests.post(
            f"{ZOHO_INVENTORY_BASE}/inventoryadjustments",
            headers=headers,
            json=payload,
            params={"organization_id": ORGANIZATION_ID},
            timeout=30,
        )
        logger.info("Zoho inventory adjustment response %s: %s", r.status_code, r.text)
        r.raise_for_status()
        resp_data = r.json()
        if resp_data.get("code") != 0:
            raise ValueError(f"Zoho error: {resp_data.get('message', 'Unknown error')}")
        return resp_data["inventory_adjustment"]

    try:
        adjustment = _do_create()
    except Exception as e:
        logger.error("Failed to create inventory adjustment for bill %s: %s", bill_id, e)
        raise HTTPException(status_code=500, detail=str(e))

    _upsert_inventory_adjustment(db, adjustment)
    if adjustment.get("inventory_adjustment_id"):
        _link_adjustment_to_bill(db, bill_id, adjustment["inventory_adjustment_id"])

    return {"inventory_adjustment": adjustment}


# ─── Linking existing inventory adjustments to a bill ─────────────────────────


@router.get("/inventory-adjustments/search")
def search_inventory_adjustments(search: Optional[str] = Query(None), limit: int = Query(20, ge=1, le=100), db=Depends(get_database)):
    query: dict = {}
    if search:
        query["$or"] = [
            {"reference_number": {"$regex": search, "$options": "i"}},
            {"reason": {"$regex": search, "$options": "i"}},
            {"inventory_adjustment_id": {"$regex": search, "$options": "i"}},
        ]
    docs = list(
        db.get_collection(INVENTORY_ADJUSTMENTS_COLLECTION)
        .find(query, {"inventory_adjustment_id": 1, "reference_number": 1, "date": 1, "reason": 1, "status": 1})
        .sort([("date", -1), ("_id", -1)])
        .limit(limit)
    )
    return {"options": serialize_mongo_document(docs)}


@router.get("/inventory-adjustments/{inventory_adjustment_id}")
def get_inventory_adjustment_detail(inventory_adjustment_id: str, db=Depends(get_database)):
    doc = db.get_collection(INVENTORY_ADJUSTMENTS_COLLECTION).find_one({"inventory_adjustment_id": inventory_adjustment_id})
    if doc and doc.get("line_items"):
        return {"inventory_adjustment": serialize_mongo_document(doc)}

    # Local doc missing or line-item-less (e.g. only summary fields synced so far) — fetch live.
    try:
        token = _get_inventory_token()
        r = requests.get(
            f"{ZOHO_INVENTORY_BASE}/inventoryadjustments/{inventory_adjustment_id}",
            headers={"Authorization": f"Zoho-oauthtoken {token}"},
            params={"organization_id": ORGANIZATION_ID},
            timeout=30,
        )
        r.raise_for_status()
        data = r.json()
        if data.get("code") != 0:
            raise ValueError(f"Zoho error: {data.get('message', 'Unknown error')}")
        adjustment = data["inventory_adjustment"]
    except Exception as e:
        logger.error("Failed to fetch inventory adjustment %s: %s", inventory_adjustment_id, e)
        raise HTTPException(status_code=500, detail=str(e))

    _upsert_inventory_adjustment(db, adjustment)
    return {"inventory_adjustment": adjustment}


@router.patch("/inventory-adjustments/{inventory_adjustment_id}/description")
def update_inventory_adjustment_description(
    inventory_adjustment_id: str,
    description: str = Body("", embed=True),
    db=Depends(get_database),
):
    """Rewrite an adjustment's description (the notes shown on the adjustment in Zoho)."""

    def _do_update():
        token = _get_inventory_token()
        r = requests.put(
            f"{ZOHO_INVENTORY_BASE}/inventoryadjustments/{inventory_adjustment_id}",
            headers={"Authorization": f"Zoho-oauthtoken {token}"},
            json={"description": description},
            params={"organization_id": ORGANIZATION_ID},
            timeout=30,
        )
        logger.info("Zoho IA description update response %s: %s", r.status_code, r.text)
        if r.status_code >= 400:
            # Zoho refuses edits on some adjustment states — surface its own wording.
            try:
                msg = r.json().get("message")
            except ValueError:
                msg = None
            raise ValueError(f"Zoho error: {msg or r.text}")
        data = r.json()
        if data.get("code") != 0:
            raise ValueError(f"Zoho error: {data.get('message', 'Unknown error')}")
        return data["inventory_adjustment"]

    try:
        adjustment = _do_update()
    except Exception as e:
        logger.error("Failed to update description for adjustment %s: %s", inventory_adjustment_id, e)
        raise HTTPException(status_code=500, detail=str(e))

    _upsert_inventory_adjustment(db, adjustment)
    return {"inventory_adjustment": serialize_mongo_document(adjustment)}


@router.get("/{bill_id}/inventory-adjustments")
def get_linked_inventory_adjustments(bill_id: str, db=Depends(get_database)):
    return {"inventory_adjustments": _linked_adjustments_for_bill(db, bill_id)}


@router.post("/{bill_id}/inventory-adjustment/link")
def link_inventory_adjustment(bill_id: str, inventory_adjustment_id: str = Body(..., embed=True), db=Depends(get_database)):
    if not db.get_collection(INVENTORY_ADJUSTMENTS_COLLECTION).find_one({"inventory_adjustment_id": inventory_adjustment_id}):
        raise HTTPException(status_code=404, detail="Inventory adjustment not found")
    _link_adjustment_to_bill(db, bill_id, inventory_adjustment_id)
    return {"inventory_adjustments": _linked_adjustments_for_bill(db, bill_id)}


@router.delete("/{bill_id}/inventory-adjustment/link/{inventory_adjustment_id}")
def unlink_inventory_adjustment(bill_id: str, inventory_adjustment_id: str, db=Depends(get_database)):
    db.get_collection(BILLS_COLLECTION).update_one(
        {"bill_id": bill_id},
        {"$pull": {"linked_inventory_adjustment_ids": inventory_adjustment_id}},
    )
    return {"inventory_adjustments": _linked_adjustments_for_bill(db, bill_id)}
