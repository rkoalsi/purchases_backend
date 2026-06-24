from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from datetime import datetime
from ..helpers.datetime_utils import utcnow
from ..database import get_database
from pydantic import BaseModel
from typing import Optional
import asyncio
import io
import re
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)
router = APIRouter()

# Managed unit prices live in their own collection, keyed by sku_code (BB code).
# This is the source of truth for the master report "Unit Price" column — it is
# NO LONGER derived from the latest purchase order rate.
COLLECTION = "product_unit_prices"
PRODUCTS_COLLECTION = "products"

TEMPLATE_COLUMNS = ["SKU Code", "Unit Price", "Currency"]
KNOWN_CURRENCIES = {"USD", "INR", "CNY", "RMB", "EUR", "GBP"}


def _clean(raw) -> str:
    """Strip Excel ="..." formula wrappers and surrounding whitespace."""
    if raw is None:
        return ""
    s = str(raw).strip()
    m = re.match(r'^="?([^"]*)"?$', s)
    return (m.group(1) if m else s).strip()


def _safe_float(val) -> Optional[float]:
    if val is None or val == "":
        return None
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _norm_currency(raw) -> str:
    c = _clean(raw).upper()
    if c == "RMB":
        return "CNY"
    return c


# ─── Request models ─────────────────────────────────────────────────────────────

class UnitPriceUpdate(BaseModel):
    unit_price: Optional[float] = None
    currency: Optional[str] = None
    updated_by: Optional[str] = None


# ─── Helpers ────────────────────────────────────────────────────────────────────

def _serialize(doc: dict) -> dict:
    return {
        "sku_code": doc.get("sku_code", ""),
        "unit_price": doc.get("unit_price"),
        "currency": doc.get("currency", "") or "",
        "updated_by": doc.get("updated_by", "") or "",
        "updated_at": doc.get("updated_at").isoformat()
        if isinstance(doc.get("updated_at"), datetime)
        else doc.get("updated_at"),
    }


def _build_template_xlsx(rows: list[dict] | None = None, with_meta: bool = False) -> bytes:
    """Build the upload template. When ``with_meta`` is True an export adds
    Item Name + Brand context columns (ignored by the upload parser, which matches
    columns by header name) so the filled sheet is easy to read."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unit Prices"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    if with_meta:
        columns = ["SKU Code", "Item Name", "Brand", "Purchase Status", "Zoho Status", "Unit Price", "Currency"]
        widths = [22, 40, 18, 16, 14, 16, 12]
    else:
        columns = TEMPLATE_COLUMNS
        widths = [22, 16, 12]

    for i, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]

    if rows:
        for r_idx, row in enumerate(rows, start=2):
            if with_meta:
                ws.cell(row=r_idx, column=1, value=row.get("sku_code", ""))
                ws.cell(row=r_idx, column=2, value=row.get("item_name", ""))
                ws.cell(row=r_idx, column=3, value=row.get("brand", ""))
                ws.cell(row=r_idx, column=4, value=row.get("purchase_status", ""))
                ws.cell(row=r_idx, column=5, value=row.get("zoho_status", ""))
                ws.cell(row=r_idx, column=6, value=row.get("unit_price"))
                ws.cell(row=r_idx, column=7, value=row.get("currency", ""))
            else:
                ws.cell(row=r_idx, column=1, value=row.get("sku_code", ""))
                ws.cell(row=r_idx, column=2, value=row.get("unit_price"))
                ws.cell(row=r_idx, column=3, value=row.get("currency", ""))

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ─── Endpoints ──────────────────────────────────────────────────────────────────

@router.get("/")
async def list_unit_prices(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    search: str = Query(None, description="Search by SKU code, item name, or brand"),
    only_set: bool = Query(False, description="Only return SKUs that have a managed unit price"),
    db=Depends(get_database),
):
    """Paginated list of active products joined with their managed unit price.

    Drives the inline-edit table. Each row is a product (so buyers can set a price
    for any SKU); the managed price is attached when present.
    """

    def _fetch():
        q: dict = {"cf_sku_code": {"$exists": True, "$nin": [None, ""]}}
        if search:
            rx = {"$regex": re.escape(search), "$options": "i"}
            q["$or"] = [
                {"cf_sku_code": rx},
                {"name": rx},
                {"brand": rx},
            ]

        # Pre-load managed prices for matching SKUs when filtering to only_set,
        # otherwise we page over products and attach prices afterwards.
        if only_set:
            priced = list(
                db[COLLECTION].find({}, {"_id": 0})
            )
            priced_skus = {p["sku_code"] for p in priced if p.get("sku_code")}
            q["cf_sku_code"]["$in"] = list(priced_skus)

        total = db[PRODUCTS_COLLECTION].count_documents(q)
        skip = (page - 1) * limit
        cursor = (
            db[PRODUCTS_COLLECTION]
            .find(q, {"_id": 0, "cf_sku_code": 1, "name": 1, "brand": 1, "purchase_status": 1, "status": 1})
            .sort([("name", 1), ("cf_sku_code", 1)])
            .skip(skip)
            .limit(limit)
        )
        products = list(cursor)

        skus = [p.get("cf_sku_code") for p in products if p.get("cf_sku_code")]
        price_map = {
            d["sku_code"]: d
            for d in db[COLLECTION].find({"sku_code": {"$in": skus}}, {"_id": 0})
        }
        return total, products, price_map

    total, products, price_map = await asyncio.to_thread(_fetch)

    rows = []
    for p in products:
        sku = p.get("cf_sku_code", "")
        pr = price_map.get(sku, {})
        rows.append({
            "sku_code": sku,
            "item_name": p.get("name", "") or "",
            "brand": p.get("brand", "") or "",
            "purchase_status": p.get("purchase_status", "") or "",
            "zoho_status": p.get("status", "") or "",
            "unit_price": pr.get("unit_price"),
            "currency": pr.get("currency", "") or "",
            "updated_by": pr.get("updated_by", "") or "",
            "updated_at": pr.get("updated_at").isoformat()
            if isinstance(pr.get("updated_at"), datetime)
            else pr.get("updated_at"),
        })

    return {
        "data": rows,
        "pagination": {
            "page": page,
            "limit": limit,
            "total": total,
            "total_pages": (total + limit - 1) // limit,
        },
    }


@router.put("/{sku_code}")
async def set_unit_price(
    sku_code: str,
    payload: UnitPriceUpdate,
    db=Depends(get_database),
):
    """Upsert the managed unit price + currency for a single SKU.

    Sending unit_price = null (or omitting it) clears the managed price for the SKU.
    """
    sku_code = _clean(sku_code)
    if not sku_code:
        raise HTTPException(status_code=400, detail="SKU code is required")

    if payload.unit_price is None:
        def _clear():
            return db[COLLECTION].delete_one({"sku_code": sku_code})

        await asyncio.to_thread(_clear)
        return {"sku_code": sku_code, "unit_price": None, "currency": ""}

    if payload.unit_price < 0:
        raise HTTPException(status_code=400, detail="Unit price cannot be negative")

    currency = _norm_currency(payload.currency)
    doc = {
        "sku_code": sku_code,
        "unit_price": round(float(payload.unit_price), 4),
        "currency": currency,
        "updated_by": (payload.updated_by or "").strip(),
        "updated_at": utcnow(),
    }

    def _upsert():
        db[COLLECTION].update_one(
            {"sku_code": sku_code}, {"$set": doc}, upsert=True
        )

    await asyncio.to_thread(_upsert)
    return _serialize(doc)


@router.delete("/{sku_code}")
async def clear_unit_price(sku_code: str, db=Depends(get_database)):
    sku_code = _clean(sku_code)

    def _delete():
        return db[COLLECTION].delete_one({"sku_code": sku_code})

    result = await asyncio.to_thread(_delete)
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="No managed unit price for that SKU")
    return {"deleted": result.deleted_count}


@router.get("/template")
async def download_template(
    prefill: bool = Query(
        False,
        description="Export every product with its current managed price filled in (blank where unset)",
    ),
    db=Depends(get_database),
):
    """Blank template (prefill=False) or a filled export of all products with their
    current unit price (prefill=True). The export adds Item Name + Brand columns and
    round-trips cleanly back through /upload."""
    rows = None
    if prefill:
        def _fetch():
            price_map = {
                d["sku_code"]: d
                for d in db[COLLECTION].find({}, {"_id": 0})
                if d.get("sku_code")
            }
            products = db[PRODUCTS_COLLECTION].find(
                {"cf_sku_code": {"$exists": True, "$nin": [None, ""]}},
                {"_id": 0, "cf_sku_code": 1, "name": 1, "brand": 1, "purchase_status": 1, "status": 1},
            ).sort([("name", 1), ("cf_sku_code", 1)])

            out_rows = []
            for p in products:
                sku = p.get("cf_sku_code", "")
                pr = price_map.get(sku, {})
                out_rows.append({
                    "sku_code": sku,
                    "item_name": p.get("name", "") or "",
                    "brand": p.get("brand", "") or "",
                    "purchase_status": p.get("purchase_status", "") or "",
                    "zoho_status": p.get("status", "") or "",
                    "unit_price": pr.get("unit_price"),
                    "currency": pr.get("currency", "") or "",
                })
            return out_rows

        rows = await asyncio.to_thread(_fetch)

    xlsx_bytes = _build_template_xlsx(rows, with_meta=prefill)
    filename = "unit_prices_export.xlsx" if prefill else "unit_prices_template.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/upload")
async def upload_unit_prices(
    file: UploadFile = File(...),
    updated_by: str = Query("", description="Username performing the upload"),
    db=Depends(get_database),
):
    """Bulk set managed unit prices from an Excel file (SKU Code, Unit Price, Currency)."""
    content = await file.read()
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only XLSX files are supported")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read the Excel file")
    ws = wb.active

    headers = [_clean(c.value).lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def _find(*names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return None

    idx_sku = _find("sku code", "sku_code", "sku", "bb code", "bbcode")
    idx_price = _find("unit price", "unit_price", "price", "rate")
    idx_cur = _find("currency", "currency code", "currency_code")

    if idx_sku is None or idx_price is None:
        raise HTTPException(
            status_code=400,
            detail="Template must have at least 'SKU Code' and 'Unit Price' columns.",
        )

    now = utcnow()
    by = (updated_by or "").strip()
    operations = []
    skipped = []
    seen = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        sku = _clean(row[idx_sku]) if idx_sku < len(row) else ""
        if not sku or sku in seen:
            continue
        seen.add(sku)
        price = _safe_float(row[idx_price]) if idx_price < len(row) else None
        currency = _norm_currency(row[idx_cur]) if idx_cur is not None and idx_cur < len(row) else ""

        if price is None:
            # Blank price clears any existing managed value for that SKU.
            operations.append(("clear", sku))
            continue
        if price < 0:
            skipped.append(sku)
            continue
        operations.append(("set", {
            "sku_code": sku,
            "unit_price": round(price, 4),
            "currency": currency,
            "updated_by": by,
            "updated_at": now,
        }))

    if not operations:
        raise HTTPException(status_code=400, detail="No valid data rows found in file")

    def _apply():
        from pymongo import UpdateOne
        # Blank-price rows delete any existing managed value to keep the collection tidy.
        clear_skus = [p for k, p in operations if k == "clear"]
        cleared_count = 0
        if clear_skus:
            cleared_count = db[COLLECTION].delete_many(
                {"sku_code": {"$in": clear_skus}}
            ).deleted_count

        set_ops = [
            UpdateOne({"sku_code": p["sku_code"]}, {"$set": p}, upsert=True)
            for k, p in operations
            if k == "set"
        ]
        updated_count = 0
        if set_ops:
            r = db[COLLECTION].bulk_write(set_ops, ordered=False)
            updated_count = (r.upserted_count or 0) + (r.modified_count or 0)
        return updated_count, cleared_count

    updated, cleared = await asyncio.to_thread(_apply)

    return {
        "updated": updated,
        "cleared": cleared,
        "skipped": skipped,
        "filename": file.filename,
    }
